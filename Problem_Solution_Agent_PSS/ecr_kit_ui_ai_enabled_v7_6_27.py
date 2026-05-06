# ecr_kit_ui.py (Enhanced v7.5.5 – Auto Excel conversion: silently open in Excel, SaveAs .xlsx, then import; OBS copy buttons)
import html
import importlib.util
import json
import re
import sys
import traceback
from pathlib import Path
from typing import Any, Dict, List

try:
    from inventory_demand_cost_query import fetch_inventory_demand_cost
except ImportError:
    fetch_inventory_demand_cost = None

try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    _HAS_MATPLOTLIB = True
except Exception:
    FigureCanvas = None
    Figure = None
    _HAS_MATPLOTLIB = False

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PyQt6.QtCore import Qt, QRect
from PyQt6.QtGui import QAction, QBrush, QColor, QFont, QFontMetrics, QGuiApplication, QPalette, QPen, QTextCursor
from PyQt6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QStyle,
    QStyleOptionHeader,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QToolBar,
    QVBoxLayout,
    QWidget,
)

APP_TITLE = "ECR Kit Assistant"
DATA_FILE = Path(__file__).with_name('ecr_kit_data.json')
TEMPLATE_FILE = Path(__file__).with_name('Obs_parts_template.xlsx')

def _handle_exception(exc_type, exc_value, exc_tb):
    try:
        log_path = Path(__file__).with_name('error_log.txt')
        msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log_path.write_text(msg, encoding='utf-8')
        m = QMessageBox()
        m.setIcon(QMessageBox.Icon.Critical)
        m.setWindowTitle('Unexpected Error')
        m.setText('An unexpected error occurred. A log was written to error_log.txt')
        m.setDetailedText(msg)
        m.setStandardButtons(QMessageBox.StandardButton.Ok)
        m.show()
        app = QApplication.instance()
        if app is not None:
            if not hasattr(app, '_error_dialogs'):
                app._error_dialogs = []
            app._error_dialogs.append(m)
    except Exception:
        print('Unhandled exception:', file=sys.stderr)
        traceback.print_exception(exc_type, exc_value, exc_tb)

sys.excepthook = _handle_exception


class ShiftEnterTextEdit(QTextEdit):
    """Enter moves focus; Shift+Enter inserts a newline."""

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                super().keyPressEvent(event)
            else:
                self.focusNextPrevChild(True)
                event.accept()
            return
        super().keyPressEvent(event)


def _format_solution_for_display(text: str) -> str:
    """
    Post-processes solution text before displaying in solution_txt widget.
    Splits any inline ' - ' / em-dash / en-dash action separators into
    individual bullet lines (UI-side safety net only — no word-wrapping).
    """
    if not text:
        return ""

    INLINE_SEP = re.compile(
        r"(?<=[a-zA-Z0-9.)]) - (?=[A-Z])"  # hyphen with spaces after word/period/paren
        r"|\.?\s*\u2014\s*(?=[A-Z])"        # em-dash (U+2014)
        r"| \u2013 (?=[A-Z])",              # en-dash (U+2013)
        re.UNICODE,
    )

    out: List[str] = []
    for raw in text.splitlines():
        stripped = raw.strip()
        if not stripped:
            out.append("")
            continue

        # Preserve section-header lines unchanged (e.g. "Solution Description:")
        if re.match(r"^[A-Za-z][\w /()&]+:\s*$", stripped):
            out.append(raw)
            continue

        # Strip any existing bullet prefix so we can re-apply it consistently
        prefix_m = re.match(r"^(\s*[-*\u2022]\s*)", raw)
        body = raw[len(prefix_m.group(0)):].strip() if prefix_m else raw.strip()

        # Split on inline action separators — each part becomes its own bullet
        parts = [p.strip() for p in INLINE_SEP.split(body) if p.strip()]
        for part in parts:
            out.append("- " + part)

    # ── Enforce max 15 words per bullet (split longer ones) ──────────────────
    enforced: List[str] = []
    for line in out:
        if not line.startswith("- "):
            enforced.append(line)
            continue
        words = line[2:].split()
        if len(words) <= 15:
            enforced.append(line)
        else:
            for i in range(0, len(words), 15):
                enforced.append("- " + " ".join(words[i:i + 15]))

    # ── Ensure at least 3 bullet points (split longest if needed) ────────────
    bullets = [l for l in enforced if l.startswith("- ")]
    while len(bullets) < 3:
        # Find the longest bullet and split it at its midpoint
        longest_idx = max(
            (i for i, l in enumerate(enforced) if l.startswith("- ")),
            key=lambda i: len(enforced[i].split()),
            default=None,
        )
        if longest_idx is None:
            break
        words = enforced[longest_idx][2:].split()
        if len(words) < 4:          # too short to split meaningfully
            break
        mid = len(words) // 2
        enforced[longest_idx:longest_idx + 1] = [
            "- " + " ".join(words[:mid]),
            "- " + " ".join(words[mid:]),
        ]
        bullets = [l for l in enforced if l.startswith("- ")]

    return "\n".join(l for l in enforced if l)


def _format_pss_for_html_display(text: str) -> str:
    """
    Converts plain PSS text to HTML for rich display in QTextEdit widgets.
    - Strips any markdown bold markers (- ** ... **) from section headers and renders them bold.
    - Makes From: and To: bold wherever they appear inline.
    - Makes from: and to: bold wherever they appear inline.
    - Adds an empty line before 'Benefits of the Proposed Solution:'.
    """
    import html as html_mod

    if not text:
        return ""

    # Insert blank line before Benefits section if not already present
    text = re.sub(
        r"(?<!\n)\n(Benefits of the Proposed Solution\s*:)",
        r"\n\n\1",
        text,
    )

    lines = text.splitlines()
    html_parts: List[str] = []

    for line in lines:
        stripped = line.strip()

        if not stripped:
            html_parts.append("")
            continue

        # Strip markdown bold/bullet artefacts from header lines: "- **Header:**" -> "Header:"
        clean = re.sub(r"^[-*\u2022]?\s*\*{1,2}([^*]+)\*{1,2}\s*$", r"\1", stripped)
        # Also handle "- **Header:** trailing text" patterns
        clean = re.sub(r"^[-*\u2022]?\s*\*{1,2}([^*]+?)\*{1,2}(.*)", r"\1\2", clean)

        escaped = html_mod.escape(clean)

        # Bold section headers (lines ending with ":")
        if re.match(r"^[A-Za-z][A-Za-z0-9_ /()&,#.\-]+:\s*$", clean):
            html_parts.append(f"<b>{escaped}</b>")
        else:
            # Bold **From:** and **To:** markdown markers output by LLM
            escaped = re.sub(r"\*\*(From:)\*\*", r"<b>\1</b>", escaped)
            escaped = re.sub(r"\*\*(To:)\*\*", r"<b>\1</b>", escaped)
            # Also bold plain From: / To: as fallback
            escaped = re.sub(r"(?<![*>])(From:)(?![*<])", r"<b>\1</b>", escaped)
            escaped = re.sub(r"(?<![*>])(To:)(?![*<])", r"<b>\1</b>", escaped)
            html_parts.append(escaped)

    return "<br>".join(html_parts)


class ReadmeTab(QWidget):
    def __init__(self, readme_path: Path):
        super().__init__()
        layout = QVBoxLayout(self)
        title = QLabel("READ ME")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.DemiBold))
        layout.addWidget(title)
        text = QTextEdit(); text.setReadOnly(True)
        content = "--Wait for the Instructions on how to use--."
        try:
            if readme_path.exists():
                content = readme_path.read_text(encoding="utf-8")
        except Exception as e:
            content = f"Error opening README: {e}"
        text.setPlainText(content)
        layout.addWidget(text)

# ---------- Helpers ----------

def get_orphan_color(orphan_level: str):
    lvl = orphan_level.lower()
    if lvl == 'orphan1': return QColor('#C0392B')
    if lvl == 'orphan2': return QColor('#E67E22')
    if lvl.startswith('orphan'): return QColor('#2980B9')
    return None


def _excel_width_to_px(widget: QWidget, chars: int) -> int:
    fm = widget.fontMetrics()
    return fm.horizontalAdvance('0' * max(1, chars)) + 22

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                self.insertPlainText('\n'); return
        super().keyPressEvent(event)


class RotatedColumnsHeader(QHeaderView):
    """Horizontal header that renders selected columns with 90-degree rotated text.
    Optionally paints a group label spanning all rotated columns in the top band."""

    GROUP_BAND = 26  # height in px reserved for the group label at top of header

    def __init__(self, orientation, rotated_columns=None, parent=None,
                 group_label=None, group_columns=None):
        super().__init__(orientation, parent)
        self._rotated_columns = set(rotated_columns or [])
        self._group_label = group_label          # e.g. 'Change Type'
        self._group_columns = list(group_columns or [])  # e.g. [2,3,4,5,6]
        self._group_spans = []  # list[(label, [col_idx,...])]
        self._header_texts = {}  # cache for header text
        self.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

    def set_rotated_columns(self, columns):
        self._rotated_columns = set(columns or [])
        self.viewport().update()

    def set_group_spans(self, group_spans):
        # group_spans: list of tuples -> (label: str, columns: list[int])
        self._group_spans = [
            (str(label), list(cols))
            for label, cols in (group_spans or [])
            if cols
        ]
        self.viewport().update()

    def set_header_texts(self, texts_dict):
        """Cache header text: { col_index: text }"""
        self._header_texts = dict(texts_dict or {})

    def _get_header_text(self, logical_index):
        """Get header text from cache or model."""
        if logical_index in self._header_texts:
            return self._header_texts[logical_index]
        try:
            m = self.model()
            if m:
                text = m.headerData(logical_index, self.orientation(), Qt.ItemDataRole.DisplayRole)
                return "" if text is None else str(text)
        except Exception:
            pass
        return ""

    def paintSection(self, painter, rect, logicalIndex):
        if not rect.isValid():
            return

        has_legacy_group = bool(self._group_label and self._group_columns)
        has_multi_groups = bool(self._group_spans)
        has_group = has_legacy_group or has_multi_groups

        if logicalIndex not in self._rotated_columns:
            super().paintSection(painter, rect, logicalIndex)
            return

        # ---- Split rect: top band for group label, rest for rotated text ----
        band = self.GROUP_BAND if has_group else 0
        indiv_rect = QRect(rect.x(), rect.y() + band, rect.width(), rect.height() - band)

        # Draw background/border for the individual section area.
        option = QStyleOptionHeader()
        self.initStyleOption(option)
        option.rect = indiv_rect
        option.section = logicalIndex
        option.text = ""
        self.style().drawControl(QStyle.ControlElement.CE_HeaderSection, option, painter, self)

        # Draw rotated column text in the lower portion.
        text = self._get_header_text(logicalIndex)
        painter.save()
        painter.setPen(self.palette().color(QPalette.ColorRole.ButtonText))
        painter.translate(indiv_rect.center())
        painter.rotate(-90)
        tr = QRect(-indiv_rect.height() // 2 + 4, -indiv_rect.width() // 2 + 2,
                   indiv_rect.height() - 8, indiv_rect.width() - 4)
        painter.drawText(tr, Qt.AlignmentFlag.AlignCenter | Qt.TextFlag.TextWordWrap, text)
        painter.restore()

        # ---- Draw group labels in top band ----
        if has_multi_groups:
            for label, cols in self._group_spans:
                if logicalIndex != cols[0]:
                    continue
                total_w = sum(self.sectionSize(c) for c in cols)
                group_rect = QRect(rect.x(), rect.y(), total_w, band)
                painter.save()
                painter.fillRect(group_rect, QColor('#B0D8F5'))
                painter.setPen(QColor('#0F2D46'))
                bold_font = painter.font()
                bold_font.setBold(True)
                bold_font.setPointSize(bold_font.pointSize() - 1)
                painter.setFont(bold_font)
                painter.drawText(group_rect, Qt.AlignmentFlag.AlignCenter, label)
                painter.restore()
                break
        elif has_legacy_group and logicalIndex == self._group_columns[0]:
            total_w = sum(self.sectionSize(c) for c in self._group_columns)
            group_rect = QRect(rect.x(), rect.y(), total_w, band)
            painter.save()
            painter.fillRect(group_rect, QColor('#B0D8F5'))
            painter.setPen(QColor('#0F2D46'))
            bold_font = painter.font()
            bold_font.setBold(True)
            bold_font.setPointSize(bold_font.pointSize() - 1)
            painter.setFont(bold_font)
            painter.drawText(group_rect, Qt.AlignmentFlag.AlignCenter, self._group_label)
            painter.restore()

class ECRFrontPageTab(QWidget):
    ROW_HEIGHT = 40  # compact rows to show ~5 checkpoints per view
    # Updated checklist per user request
    CHECKLIST_DEF = [
        ("Project Association", ["Does the ECR include a Project (PCR)?"]),
        ("Safety & Compliance", ["Provide PSER details if any safety incident occurred.", "Is a PCN required and approved by ROW and CE?"]),
        ("Part Release Status", ["Are all BTPs and kits EVAL released?", "If the parent part is in production, is the replacement part (BTP) production released?"]),
        ("Design Analysis", ["Is VA/VB analysis completed (for new designs/parts only)?", "PACE / DASH parts addressed?"]),
        ("Watchlist & Spares", ["Are new parts/designs MLO certified / were parent/previous parts MLO certified?", "Do we have AGS approval for OBS parts (sparable) without replacement?"]),
        ("Config Documents", ["Is CCR available with change matrix details for new options/reference designator updates?"]),
        ("ECR Strategies Identified", ["Provide reason code, effective strategy, priority, and disposition."]),
        ("Multi BU Alignment", ["If scope impacts multiple BUs/products, verify and confirm all affected BUs/products are listed in the project and approved."]),
        ("Interchangeability & Testing", [
            "Are the changes FFF compatible? Provide system details.",
            "Are test results/FDR available for all FFF impacted parts/critical parts?",
        ]),
        ("Costs and Savings", [
            "If the project relates to DCR (cost reduction), provide cost-saving details.",
        ])
    ]
    _QUESTIONS_TOTAL = sum(len(qs) for _, qs in CHECKLIST_DEF)
    assert len(CHECKLIST_DEF) == 10 and _QUESTIONS_TOTAL == 15

    def __init__(self):
        super().__init__()
        top = QVBoxLayout(self)

        scroll = QScrollArea(self); scroll.setWidgetResizable(True); top.addWidget(scroll)
        content = QWidget(); outer = QVBoxLayout(content); outer.setSpacing(10); outer.setContentsMargins(8,8,8,8); scroll.setWidget(content)

        # Header
        row1 = QWidget(); r1 = QHBoxLayout(row1); r1.setContentsMargins(12,0,0,0)
        self.ecr_no = QLineEdit(); self.ecr_no.setPlaceholderText("ECR#")
        self.eco_primer = QLineEdit(); self.eco_primer.setPlaceholderText("ECO Primer Refs#")
        self.ec_category = QLineEdit(); self.ec_category.setPlaceholderText("EC Category")
        self.bu = QLineEdit(); self.bu.setPlaceholderText("BU")
        self.tco = QLineEdit(); self.tco.setPlaceholderText("TCO")
        self.project_no = QLineEdit(); self.project_no.setPlaceholderText("Project#")
        self.product = QLineEdit(); self.product.setPlaceholderText("Product")
        for w in [self.ecr_no, self.eco_primer, self.ec_category, self.bu, self.tco, self.project_no, self.product]:
            w.setFixedHeight(28); r1.addWidget(w)
        row2 = QWidget(); r2 = QHBoxLayout(row2); r2.setContentsMargins(12,0,0,0)
        self.affected_modules = QLineEdit(); self.affected_modules.setPlaceholderText("Affected Module(s)")
        self.place = QLineEdit(); self.place.setPlaceholderText("Place")
        for w in [self.affected_modules, self.place]:
            w.setFixedHeight(28); r2.addWidget(w)
        outer.addWidget(row1); outer.addWidget(row2)

        lbl = QLabel("ECR Creation Checklist"); lbl.setFont(QFont("Segoe UI", 5, QFont.Weight.DemiBold))
        temp_lbl_row = QWidget(); tl = QHBoxLayout(temp_lbl_row); tl.setContentsMargins(12,0,0,0); tl.addWidget(lbl); outer.addWidget(temp_lbl_row)

        # Table
        table_row = QWidget(); tr = QHBoxLayout(table_row); tr.setContentsMargins(12,0,0,0)
        self.table = QTableWidget(len(self.CHECKLIST_DEF), 9)
        self.table.setHorizontalHeaderLabels(["Sl.No","Completion","Category","Check Point","Validation","Number","Action Owner","Comments","Due Date"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.table.setWordWrap(True); self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
            QHeaderView::section { background:#E1F0FF; color:#0F2D46; font-weight:600; border:1px solid #C9E2FF; padding:4px; }
            QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        for row,(cat,qlist) in enumerate(self.CHECKLIST_DEF):
            sl = QTableWidgetItem(str(row+1)); sl.setFlags(sl.flags() & ~Qt.ItemFlag.ItemIsEditable); sl.setTextAlignment(Qt.AlignmentFlag.AlignCenter); self.table.setItem(row,0,sl)
            chk = QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); cont._chk=chk; self.table.setCellWidget(row,1,cont)
            cat_item = QTableWidgetItem(cat); cat_item.setFlags(cat_item.flags() & ~Qt.ItemFlag.ItemIsEditable); self.table.setItem(row,2,cat_item)
            cp_label = QLabel("\n".join(qlist)); cp_label.setTextFormat(Qt.TextFormat.PlainText); cp_label.setWordWrap(True); self.table.setCellWidget(row,3,cp_label)
            combo = QComboBox(); combo.addItems(['','YES','NO','N/A']); self.table.setCellWidget(row,4,combo)
            def _on_change(txt, r=row):
                sel=(txt or '').strip().upper(); w=self.table.cellWidget(r,1)
                if isinstance(w, QWidget) and hasattr(w,'_chk'): w._chk.setChecked(sel in ('YES','N/A'))
            combo.currentTextChanged.connect(_on_change)
            self.table.setItem(row,5,QTableWidgetItem(""))
            self.table.setItem(row,6,QTableWidgetItem(""))
            te=ShiftEnterTextEdit(); te.setPlaceholderText("Add comments… (Shift+Enter for new line)"); te.setFrameShape(QFrame.Shape.NoFrame); te.setStyleSheet("QTextEdit { padding:4px; }"); self.table.setCellWidget(row,7,te)
        header=self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2,QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(2,_excel_width_to_px(self.table,14))
        header.setSectionResizeMode(3,QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5,QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(6,QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(7,QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(5,_excel_width_to_px(self.table,10))
        self.table.setColumnWidth(6,_excel_width_to_px(self.table,14))
        for r in range(self.table.rowCount()): self.table.setRowHeight(r,self.ROW_HEIGHT)
        tr.addWidget(self.table); outer.addWidget(table_row)

        # Progress
        progress_row = QWidget(); pr=QHBoxLayout(progress_row); pr.setContentsMargins(12,0,0,0)
        self.progress_label = QLabel(f"Checklist Progress: 0 / {len(self.CHECKLIST_DEF)}")
        self.progress_bar=QProgressBar(); self.progress_bar.setRange(0,len(self.CHECKLIST_DEF)); self.progress_bar.setFixedHeight(26); self.progress_bar.setTextVisible(True)
        pr.addWidget(self.progress_label); pr.addWidget(self.progress_bar,3); pr.addStretch(1); outer.addWidget(progress_row)

        # Inputs
        action_row = QWidget(); ar=QHBoxLayout(action_row); ar.setContentsMargins(12,0,0,0)
        self.btn_generate=QPushButton("Generate Problem Statement")
        self.title_edit=QLineEdit(); self.title_edit.setPlaceholderText("Title")
        ar.addWidget(self.btn_generate,1); ar.addWidget(self.title_edit,3); ar.addStretch(1); outer.addWidget(action_row)

        self.problem_edit=QTextEdit(); self.problem_edit.setPlaceholderText("Write the problem statement here (max 2000 characters)…"); self.problem_edit.setMinimumHeight(120)
        self.problem_edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.problem_edit.textChanged.connect(lambda: self._limit_text(self.problem_edit,2000))

        self.solution_edit=QTextEdit(); self.solution_edit.setPlaceholderText("Write the proposed solution here (max 2000 characters)…"); self.solution_edit.setMinimumHeight(120)
        self.solution_edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # --- AI Assisted ECR Drafting Panel (Right Side) ---
        ai_row = QWidget(); ai_layout = QHBoxLayout(ai_row); ai_layout.setContentsMargins(12,0,0,0)
        left_splitter = QSplitter(Qt.Orientation.Vertical)
        left_splitter.addWidget(self.problem_edit)
        left_splitter.addWidget(self.solution_edit)
        left_splitter.setChildrenCollapsible(False)
        left_splitter.setStretchFactor(0, 1)
        left_splitter.setStretchFactor(1, 1)
        left_splitter.setSizes([220, 220])

        ai_panel = QWidget(); ai_panel.setFixedWidth(300); ai_v = QVBoxLayout(ai_panel)
        ai_title = QLabel('AI Assisted ECR Drafting'); ai_title.setFont(QFont('Segoe UI',11,QFont.Weight.DemiBold))
        btn_paste = QPushButton('-||-')
        btn_file = QPushButton('Upload PDF / Word / PPT')
        btn_email = QPushButton('Upload Email (.msg / .eml)')
        scope_lbl = QLabel('Include Data From:')
        cb_obs = QCheckBox('OBS Parts'); cb_obs.setChecked(True)
        cb_wu = QCheckBox('Where Used'); cb_wu.setChecked(True)
        cb_chk = QCheckBox('Checklist'); cb_chk.setChecked(True)
        cb_struct = QCheckBox('Structure Sheet'); cb_struct.setChecked(True)
        btn_ai = QPushButton('Generate ECR using AI'); btn_ai.setFixedHeight(34)
        info = QLabel('AI will analyze selected inputs and other tabs to generate Title, Problem and Solution.'); info.setWordWrap(True)

        for w in [ai_title, btn_paste, btn_file, btn_email, scope_lbl, cb_obs, cb_wu, cb_chk, cb_struct, btn_ai, info]: ai_v.addWidget(w)
        ai_v.addStretch(1)

        ai_layout.addWidget(left_splitter,3); ai_layout.addWidget(ai_panel,1); outer.addWidget(ai_row)

        self.solution_edit.textChanged.connect(lambda: self._limit_text(self.solution_edit,2000))

        self.setStyleSheet("""
            QLabel { color: #12324A; }
            QLineEdit, QTextEdit { background:#FFFFFF; border:1px solid #BBD3EA; border-radius:4px; padding:4px; }
            QLineEdit:focus, QTextEdit:focus { border-color:#639AD2; }
            QPushButton { background-color:#3BAFDA; color:#ffffff; border:1px solid #2C9CC8; border-radius:5px; padding:6px 10px; }
            QPushButton:hover { background-color:#35A0C9; }
            QProgressBar { border:1px solid #BBD3EA; border-radius:4px; background:#ECF4FF; text-align:center; color:#12324A; }
            QProgressBar::chunk { background-color:#5CC0FF; }
        """)

        def recalc():
            checked=0
            for r in range(self.table.rowCount()):
                w=self.table.cellWidget(r,1)
                if isinstance(w,QWidget) and hasattr(w,'_chk') and w._chk.isChecked(): checked+=1
            self.progress_label.setText(f"Checklist Progress: {checked} / {self.table.rowCount()}")
            self.progress_bar.setValue(checked)
        for r in range(self.table.rowCount()):
            w=self.table.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.stateChanged.connect(recalc)

    def _limit_text(self, editor: QTextEdit, max_chars: int):
        doc = editor.toPlainText()
        if len(doc)>max_chars:
            cursor=editor.textCursor(); pos=cursor.position()
            editor.blockSignals(True); editor.setPlainText(doc[:max_chars]); cursor.setPosition(min(pos,max_chars)); editor.setTextCursor(cursor); editor.blockSignals(False)

    def to_dict(self)->Dict[str,Any]:
        data={'fields':{
            'ecr_no':self.ecr_no.text(),'eco_primer':self.eco_primer.text(),'ec_category':self.ec_category.text(),
            'bu':self.bu.text(),'tco':self.tco.text(),'project_no':self.project_no.text(),'product':self.product.text(),
            'affected_modules':self.affected_modules.text(),'place':self.place.text(),'title':self.title_edit.text(),
            'problem':self.problem_edit.toPlainText(),'solution':self.solution_edit.toPlainText(),},'checklist':[]}
        t=self.table
        for r in range(t.rowCount()):
            completion=False; w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): completion=w._chk.isChecked()
            cp_widget=t.cellWidget(r,3); cp_text=cp_widget.text() if isinstance(cp_widget,QLabel) else ''
            v_widget=t.cellWidget(r,4); v_choice=v_widget.currentText() if isinstance(v_widget,QComboBox) else ''
            row={'slno':t.item(r,0).text() if t.item(r,0) else str(r+1),'completion':completion,'category':t.item(r,2).text() if t.item(r,2) else '',
                 'check_point_text':cp_text,'validation_choice':v_choice,'action_owner':t.item(r,5).text() if t.item(r,5) else '',
                 'due_date':t.item(r,6).text() if t.item(r,6) else '',
                 'comments': t.cellWidget(r,7).toPlainText() if isinstance(t.cellWidget(r,7), QTextEdit) else (t.item(r,7).text() if t.item(r,7) else '')}
            data['checklist'].append(row)
        return data

    def from_dict(self, data: Dict[str,Any]):
        f=data.get('fields',{})
        self.ecr_no.setText(f.get('ecr_no','')); self.eco_primer.setText(f.get('eco_primer','')); self.ec_category.setText(f.get('ec_category',''))
        self.bu.setText(f.get('bu','')); self.tco.setText(f.get('tco','')); self.project_no.setText(f.get('project_no','')); self.product.setText(f.get('product',''))
        self.affected_modules.setText(f.get('affected_modules','')); self.place.setText(f.get('place','')); self.title_edit.setText(f.get('title',''))
        self.problem_edit.setPlainText(f.get('problem','')); self.solution_edit.setPlainText(f.get('solution',''))
        checklist=data.get('checklist',[]); t=self.table; rows=min(len(checklist), t.rowCount())
        for r in range(rows):
            row=checklist[r]
            w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.setChecked(bool(row.get('completion',False)))
            vc=row.get('validation_choice',None); wv=t.cellWidget(r,4)
            if vc is not None and isinstance(wv,QComboBox):
                idx=wv.findText(vc); wv.setCurrentIndex(idx if idx>=0 else 0)
            def _ensure(col):
                it=t.item(r,col)
                if it is None:
                    it=QTableWidgetItem(""); t.setItem(r,col,it)
                return it
            _ensure(5).setText(row.get('action_owner','') or row.get('actioner_owner',''))
            _ensure(6).setText(row.get('due_date',''))
            w_comments=t.cellWidget(r,7)
            if isinstance(w_comments, QTextEdit): w_comments.setPlainText(row.get('comments',''))
        for r in range(t.rowCount()):
            try: t.setRowHeight(r,self.ROW_HEIGHT)
            except Exception: pass

    def reset(self):
        for w in [self.ecr_no,self.eco_primer,self.ec_category,self.bu,self.tco,self.project_no,self.product,self.affected_modules,self.place,self.title_edit]: w.clear()
        self.problem_edit.clear(); self.solution_edit.clear()
        t=self.table
        for r in range(t.rowCount()):
            w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.setChecked(False)
        for c in (5,6):
            it=t.item(r,c)
            if it: it.setText("")
        w_comments=t.cellWidget(r,7)
        if isinstance(w_comments, QTextEdit): w_comments.clear()
        self.progress_label.setText(f"Checklist Progress: 0 / {self.table.rowCount()}"); self.progress_bar.setValue(0)

class OBSTable(QTableWidget):
    def __init__(self, parent=None, initial_rows: int = 10):
        super().__init__(initial_rows, 4, parent)
        self.setHorizontalHeaderLabels(["Select","OBS Parts","Change","Replacement"])
        self.verticalHeader().setVisible(False)
        try:
            header=self.horizontalHeader(); header.setStretchLastSection(False)
            header.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(1,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(2,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(3,QHeaderView.ResizeMode.Interactive)
        except Exception: pass
        self._init_rows(0,self.rowCount()); self._apply_excel_widths(); self.setAlternatingRowColors(True)

    def _apply_excel_widths(self, chars:int=14):
        px=_excel_width_to_px(self,chars)
        for col in (1,2,3):
            try: self.setColumnWidth(col,px)
            except Exception: pass

    def _make_change_combo(self)->QComboBox:
        combo=QComboBox(); combo.addItems(["Obsolete","Inactivate"]); combo.setCurrentText("Obsolete"); return combo

    def _init_rows(self,start,end):
        for r in range(start,end):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); self.setCellWidget(r,0,cont)
            if not self.item(r,1): self.setItem(r,1,QTableWidgetItem(""))
            combo=self._make_change_combo(); self.setCellWidget(r,2,combo)
            if not self.item(r,3): self.setItem(r,3,QTableWidgetItem(""))

    def keyPressEvent(self,event):
        try:
            if event.matches(event.StandardKey.Copy): self._copy_selection_to_clipboard(); return
            if event.matches(event.StandardKey.Paste): self._paste_from_clipboard(); return
        except Exception: traceback.print_exc()
        super().keyPressEvent(event)

    def _copy_selection_to_clipboard(self):
        sel = self.selectedRanges()
        if not sel:
            return
        r=sel[0]; rows=[]
        for i in range(r.rowCount()):
            cols=[]
            for j in range(r.columnCount()):
                row_i=r.topRow()+i; col_j=r.leftColumn()+j
                if col_j==2:
                    w=self.cellWidget(row_i,2); txt=w.currentText() if isinstance(w,QComboBox) else ''
                else:
                    it=self.item(row_i,col_j); txt=it.text() if it else ''
                cols.append(txt)
            rows.append('\t'.join(cols))
        QGuiApplication.clipboard().setText('\n'.join(rows))

    def _ensure_rows(self,upto_row_inclusive:int):
        if upto_row_inclusive>=self.rowCount():
            old=self.rowCount(); self.setRowCount(upto_row_inclusive+1); self._init_rows(old,self.rowCount())

    def _paste_from_clipboard(self):
        text = QGuiApplication.clipboard().text()
        if not text:
            return
        start_row=self.currentRow(); start_col=self.currentColumn()
        if start_row<0:
            start_row=self._first_empty_row()
            if start_row<0: start_row=self.rowCount()
        lines=[ln for ln in text.splitlines() if ln.strip()]
        for r_offset,line in enumerate(lines):
            parts=[p.strip() for p in line.split('\t')]
            row=start_row+r_offset; self._ensure_rows(row)
            for c_offset,val in enumerate(parts):
                col=start_col+c_offset
                if col==0: continue
                if col==2:
                    w=self.cellWidget(row,2)
                    if isinstance(w,QComboBox):
                        idx=w.findText(val)
                        if idx>=0: w.setCurrentIndex(idx)
                else:
                    self.setItem(row,col,QTableWidgetItem(val))
        self._apply_excel_widths()

    def _first_empty_row(self):
        for r in range(self.rowCount()):
            it=self.item(r,1)
            if it is None or not it.text().strip(): return r
        return -1

    def delete_selected_rows(self):
        rows_to_delete = []
        for r in range(self.rowCount()):
            w = self.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and cb.isChecked():
                    rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            self.removeRow(r)
        if self.rowCount() == 0:
            self.setRowCount(10)
            self._init_rows(0, 10)

class OBSPartsTab(QWidget):
    def toggle_select_all(self):
        any_unchecked = False
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and not cb.isChecked():
                    any_unchecked = True
                    break
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(any_unchecked)


    def select_all_rows(self):
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, 'findChild'):
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(True)

    def __init__(self):
        super().__init__()
        outer=QVBoxLayout(self)

        title_row=QHBoxLayout()
        title=QLabel("Final OBS List"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold)); title.setStyleSheet("color:#C1272D;")
        btn_template=QPushButton("Download Template")
        btn_upload=QPushButton("Upload Template")
        # New copy buttons
        btn_copy_obs=QPushButton("Copy OBS Parts")
        btn_copy_rep=QPushButton("Copy Repl Parts")
        delete_btn=QPushButton("Delete Selected")
        title_row.addWidget(title)
        title_row.addStretch(1)
        title_row.addWidget(btn_template)
        title_row.addWidget(btn_upload)
        title_row.addWidget(btn_copy_obs)
        title_row.addWidget(btn_copy_rep)
        select_all_btn = QPushButton(" Select All")
        select_all_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))
        select_all_btn.setToolTip("Toggle select / unselect all rows")
        title_row.addWidget(select_all_btn)
        title_row.addWidget(delete_btn)
        outer.addLayout(title_row)

        # ── Where Used of OBS Parts row ──────────────────────────────────────
        wu_row = QHBoxLayout()
        wu_lbl = QLabel("WU Level (1–6):")
        wu_lbl.setStyleSheet("font-weight:600; color:#7A1C21;")
        self.wu_level_input = QLineEdit()
        self.wu_level_input.setFixedWidth(55)
        self.wu_level_input.setPlaceholderText("1–6")
        self.wu_level_input.setMaxLength(1)
        self.wu_level_input.setToolTip("Maximum Where Used depth to retrieve (1 to 6)")
        self.btn_where_used = QPushButton("Where Used of OBS Parts")
        self.btn_where_used.setToolTip(
            "Query Databricks for multi-level Where Used data for all OBS parts"
        )
        plant_lbl = QLabel("Plant:")
        plant_lbl.setStyleSheet("font-weight:600; color:#7A1C21;")
        self.plant_combo = QComboBox()
        self.plant_combo.addItems(["4020", "4055", "4060", "4070", "4080", "4090"])
        self.plant_combo.setCurrentText("4070")
        self.plant_combo.setToolTip("Plant code to filter Where Used query")
        self.plant_combo.setFixedWidth(75)
        wu_row.addStretch(1)
        wu_row.addWidget(wu_lbl)
        wu_row.addWidget(self.wu_level_input)
        wu_row.addWidget(plant_lbl)
        wu_row.addWidget(self.plant_combo)
        wu_row.addWidget(self.btn_where_used)
        outer.addLayout(wu_row)

        legend = QLabel(
            "<b>Orphan Legend:</b> "
            "<span style='color:#C0392B; font-weight:600;'>Orphan1</span> | "
            "<span style='color:#E67E22; font-weight:600;'>Orphan2</span> | "
            "<span style='color:#2980B9; font-weight:600;'>Orphan3+</span>"
        )
        legend.setStyleSheet("padding:4px;")
        outer.addWidget(legend)


        self.table=OBSTable(self, initial_rows=1); outer.addWidget(self.table)

        self.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#FFF5F5; gridline-color:#F3C2C2; }
            QHeaderView::section { background:#F8D7DA; color:#7A1C21; font-weight:600; border:1px solid #E3AEB2; padding:4px; }
            QTableWidget::item:selected { background:#F5B5B8; color:#4A0E10; }
            QPushButton { background-color:#C1272D; color:#FFFFFF; border:1px solid #9F1F24; border-radius:5px; padding:6px 10px; }
            QPushButton:hover { background-color:#AD2227; }
            QComboBox { border:1px solid #E3AEB2; border-radius:4px; padding:2px 6px; }
        """)

        delete_btn.clicked.connect(self.table.delete_selected_rows)
        btn_template.clicked.connect(self.download_template)
        btn_upload.clicked.connect(self.upload_from_excel)
        btn_copy_obs.clicked.connect(self.copy_obs_parts)
        btn_copy_rep.clicked.connect(self.copy_replacement_parts)
        select_all_btn.clicked.connect(self.toggle_select_all)
        self.btn_where_used.clicked.connect(self.launch_where_used_import)
        self.where_used_tab = None  # linked by MainWindow after both tabs are created

    def download_template(self):
        try:
            path,_=QFileDialog.getSaveFileName(self,'Save Template',str(TEMPLATE_FILE),'Excel Files (*.xlsx)')
            if not path: return
            import pandas as pd
            df=pd.DataFrame({'OBS Parts':[''],'Change':['Obsolete'],'Replacement':['']})
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='OBS_Template')
            QMessageBox.information(self,'Template Saved',f'Template saved to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self,'Template Error',str(e))

    def upload_from_excel(self):
        try:
            path,_=QFileDialog.getOpenFileName(self,'Open OBS Parts Excel','', 'Excel Files (*.xlsx *.xls)')
            if not path: return
            import pandas as pd
            def as_text(value: Any) -> str:
                if value is None:
                    return ''
                try:
                    if pd.isna(value):
                        return ''
                except Exception:
                    pass
                return str(value).strip()
            if path.lower().endswith('.xls'):
                df=pd.read_excel(path, engine='xlrd', dtype=str, keep_default_na=False)
            else:
                df=pd.read_excel(path, engine='openpyxl', dtype=str, keep_default_na=False)
            cols={str(c).strip().lower(): c for c in df.columns}
            def pick(name):
                for key in cols:
                    if key==name: return cols[key]
                return None
            c_obs=pick('obs parts') or pick('obs part') or pick('part')
            c_change=pick('change')
            c_rep=(pick('replacement') or pick('repl') or pick('replace') or
                   pick('replacement part') or pick('new part') or pick('irplacement'))
            if not c_obs: raise ValueError('Column "OBS Parts" is required in the Excel file.')
            rows=[]
            for _,r in df.iterrows():
                obs=as_text(r.get(c_obs,''))
                if not obs: continue
                change_val=as_text(r.get(c_change,'Obsolete')) if c_change else 'Obsolete'
                if change_val not in ['Obsolete','Inactivate']: change_val='Obsolete'
                rep=as_text(r.get(c_rep,'')) if c_rep else ''
                rows.append((obs,change_val,rep))
            if not rows:
                QMessageBox.information(self,'No Data','No valid rows found in the Excel file.'); return
            t=self.table; t.setRowCount(len(rows)); t._init_rows(0,len(rows))
            for r,(obs,change,rep) in enumerate(rows):
                t.setItem(r,1,QTableWidgetItem(obs))
                w=t.cellWidget(r,2)
                if isinstance(w,QComboBox):
                    idx=w.findText(change); w.setCurrentIndex(idx if idx>=0 else 0)
                rep_item = t.item(r,3)
                if rep_item is None:
                    rep_item = QTableWidgetItem('')
                    t.setItem(r,3,rep_item)
                rep_item.setText(rep)
            t._apply_excel_widths(); QMessageBox.information(self,'Upload Complete',f'Loaded {len(rows)} rows from Excel.')
        except Exception as e:
            QMessageBox.warning(self,'Upload Error',str(e))

    # NEW: Copy helpers
    def _collect_column_values(self, col_index: int) -> List[str]:
        t=self.table
        values=[]
        for r in range(t.rowCount()):
            it=t.item(r,col_index)
            if it:
                val=(it.text() or '').strip()
                if val:
                    values.append(val)
        return values

    def copy_obs_parts(self):
        values=self._collect_column_values(1)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} OBS Part number(s) to clipboard.')

    def copy_replacement_parts(self):
        values=self._collect_column_values(3)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} Replacement part number(s) to clipboard (Image).')

    def to_dict(self)->Dict[str,Any]:
        t=self.table; rows:List[Dict[str,Any]]=[]
        for r in range(t.rowCount()):
            obs=t.item(r,1).text() if t.item(r,1) else ''
            rep=t.item(r,3).text() if t.item(r,3) else ''
            w=t.cellWidget(r,2); change=w.currentText() if isinstance(w,QComboBox) else 'Obsolete'
            if any([obs.strip(), rep.strip()]): rows.append({'obs_part':obs,'change':change,'replacement':rep})
        return {'rows': rows}

    def from_dict(self,data:Dict[str,Any]):
        rows=data.get('rows',[]); t=self.table; needed=max(10,len(rows)); t.setRowCount(needed); t._init_rows(0,needed)
        for r,row in enumerate(rows):
            t.setItem(r,1,QTableWidgetItem(row.get('obs_part','')))
            w=t.cellWidget(r,2)
            if isinstance(w,QComboBox):
                idx=w.findText(row.get('change','Obsolete'))
                if idx>=0: w.setCurrentIndex(idx)
            t.setItem(r,3,QTableWidgetItem(row.get('replacement','')))
        t._apply_excel_widths()

    def reset(self):
        t=self.table; t.setRowCount(10); t._init_rows(0,10)

    def launch_where_used_import(self):
        """Validate WU level + OBS parts, then delegate to the linked Where Used tab."""
        # ── Validate WU Level ──────────────────────────────────────────────────
        raw_level = self.wu_level_input.text().strip()
        if not raw_level:
            QMessageBox.warning(
                self, 'WU Level Required',
                'Please enter a WU Level (1 to 6) before importing.'
            )
            return
        try:
            wu_level = int(raw_level)
            if not (1 <= wu_level <= 6):
                raise ValueError
        except ValueError:
            QMessageBox.warning(
                self, 'Invalid WU Level',
                f'"{raw_level}" is not valid.  Please enter a whole number from 1 to 6.'
            )
            return

        # ── Collect non-empty OBS part numbers ────────────────────────────────
        t = self.table
        obs_parts: List[str] = []
        for r in range(t.rowCount()):
            it = t.item(r, 1)
            val = (it.text() if it else '').strip()
            if val:
                obs_parts.append(val)

        if not obs_parts:
            QMessageBox.warning(
                self, 'No OBS Parts',
                'The OBS Parts column is empty.\n'
                'Please enter at least one part number before importing.'
            )
            return

        # ── Delegate to Where Used tab ─────────────────────────────────────────
        if self.where_used_tab is None:
            QMessageBox.warning(self, 'Not Ready', 'Where Used tab is not available yet.')
            return

        plant = self.plant_combo.currentText().strip()
        self.where_used_tab.import_from_databricks(obs_parts, wu_level, plant)

        # Switch focus to the Where Used tab after a successful import
        try:
            main = self.window()
            if hasattr(main, 'tabs'):
                main.tabs.setCurrentWidget(self.where_used_tab)
        except Exception:
            pass


class WhereUsedTab(QWidget):
    """Import and manage Where Used parent rows with OBS-aware mapping.
    Data cleanup is applied to the in-app table only and source files are unchanged."""
    def __init__(self, obs_provider=None):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)
        title_row = QHBoxLayout()
        title = QLabel("Where Used - Parents"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold))
        btn_import = QPushButton("Import 'Where Used' Parents")
        btn_import.setToolTip("Select Excel/CSV/HTML. Legacy/mismatched files are auto-converted to .xlsx using Excel, then imported. Cleanup + OBS mapping are applied to the in-app view only.")
        title_row.addWidget(title); title_row.addStretch(1); title_row.addWidget(btn_import)
        outer.addLayout(title_row)

    


        # Action buttons row
        btn_row = QHBoxLayout()
        self.btn_import_obs_multi = QPushButton("Import Where Used of OBS Parts (Multiple Level)")
        self.btn_sel_9024 = QPushButton("Select all 9024 Parents")
        self.btn_sel_options = QPushButton("Select all Options/O Class")
        self.btn_sel_esw = QPushButton("Select ESW Parents")
        self.btn_delete_sel = QPushButton("Delete the selected")
        self.btn_move_to_struct = QPushButton("Move selected Items to Structure Sheet")
        self.btn_append_obs = QPushButton("Append the selected to OBS List")
        self.btn_export = QPushButton("Export WhereUsed")
        self.btn_reset = QPushButton("Reset Where_Used")
        for b in [self.btn_import_obs_multi, self.btn_sel_9024, self.btn_sel_options, self.btn_sel_esw, self.btn_delete_sel, self.btn_move_to_struct, self.btn_append_obs, self.btn_export, self.btn_reset]:
            btn_row.addWidget(b)
        outer.addLayout(btn_row)

        self.table=QTableWidget(0,0); self.table.verticalHeader().setVisible(False); self.table.setAlternatingRowColors(True); self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        outer.addWidget(self.table)

        self.setStyleSheet("""
        QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
        QHeaderView::section { background:#E1F0FF; color:#0F2D46; font-weight:600; border:1px solid #C9E2FF; padding:4px; }
        QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        btn_import.clicked.connect(self.import_where_used)
        self.btn_import_obs_multi.clicked.connect(self.import_where_used_of_obs_multi_level)
        self.btn_sel_9024.clicked.connect(self.select_all_9024_parents)
        self.btn_sel_options.clicked.connect(self.select_all_options)
        self.btn_sel_esw.clicked.connect(self.select_esw_parents)
        self.btn_delete_sel.clicked.connect(self.delete_selected_rows)
        self.btn_move_to_struct.clicked.connect(self.move_selected_to_structure)
        self.btn_append_obs.clicked.connect(self.append_selected_to_obs)
        self.btn_export.clicked.connect(self.export_where_used)
        self.btn_reset.clicked.connect(self.reset_where_used)

    # ---------------- Excel conversion helpers ----------------
    def _is_html_like(self, path: str)->bool:
        try:
            with open(path, 'rb') as f:
                head = f.read(2048).lstrip().lower()
                return (head.startswith(b'<!') or head.startswith(b'<html') or b'<table' in head)
        except Exception:
            return False

    def _convert_to_xlsx_via_excel(self, src_path: str)->str|None:
        try:
            from pathlib import Path as _Path
            import win32com.client
            src_abs=str(_Path(src_path).resolve())
            dst_path=str(_Path(src_abs).with_suffix(''))+'_converted.xlsx'
            excel=win32com.client.DispatchEx('Excel.Application'); excel.Visible=False; excel.DisplayAlerts=False
            wb=excel.Workbooks.Open(src_abs); wb.SaveAs(dst_path, FileFormat=51); wb.Close(SaveChanges=False); excel.Quit(); return dst_path
        except Exception:
            try:
                excel.Quit()
            except Exception:
                pass
            return None

    # ---------------- Helpers for OBS, colors, selection ----------------
    def _find_part_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            if str(c).strip().lower()=='part': return i
        return -1
    def _find_parent_col_index(self, cols:list[str])->int:
        keys=['parent','parent part','parent pn','parent number','parent part number']
        low=[str(c).strip().lower() for c in cols]
        for i,name in enumerate(low):
            for k in keys:
                if k==name or k in name: return i
        return -1
    def _find_class_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            s=str(c).strip().lower()
            if any(k in s for k in ['class','type','category']): return i
        return -1
    def _build_obs_map(self)->dict:
        mapping={}
        try:
            if self.obs_provider is None: return mapping
            t=self.obs_provider.table
            for r in range(t.rowCount()):
                obs_item=t.item(r,1); rep_item=t.item(r,3)
                obs=(obs_item.text() if obs_item else '').strip(); rep=(rep_item.text() if rep_item else '')
                if obs: mapping[obs.upper()]=rep
        except Exception: pass
        return mapping
    def _read_xlsx_background_colors(self, xlsx_path:str, target_ncols:int):
        try:
            from openpyxl import load_workbook
            wb=load_workbook(xlsx_path, data_only=True); ws=wb.active
            colors=[]; first_data_row=2; ncols=target_ncols
            for r in range(first_data_row, ws.max_row+1):
                row_colors=[]
                for c in range(1, ncols+1):
                    cell=ws.cell(row=r,column=c); col=None
                    try:
                        fill=cell.fill
                        if fill and getattr(fill,'fill_type',None):
                            start=getattr(fill,'start_color',None); rgb=getattr(start,'rgb',None)
                            if isinstance(rgb,str) and len(rgb) in (6,8):
                                rgb_hex=rgb[-6:]; col=QColor('#'+rgb_hex)
                    except Exception: col=None
                    row_colors.append(col)
                colors.append(row_colors)
            return colors
        except Exception:
            return None
    def _apply_cleanup_rules(self, df, part_idx:int):
        import pandas as pd
        part_raw=df.iloc[:,part_idx].astype(str); part_trim=part_raw.str.strip(); lengths=part_trim.str.len().fillna(0)
        mask_gt=(lengths>10) & part_trim.str.upper().str.startswith('ESW')
        mask_eq=(lengths==10) & (part_trim.str[4]=='-')
        mask=mask_gt | mask_eq
        df_kept=df.loc[mask].copy(); df_kept.reset_index(drop=True, inplace=True)
        return df_kept, mask.reset_index(drop=True)
    def _find_replacement_col(self)->int:
        for i in range(self.table.columnCount()):
            h=self.table.horizontalHeaderItem(i)
            if h and str(h.text()).strip().lower()=='replacement': return i
        return -1
    def _center_replacement_column(self):
        rep_col=self._find_replacement_col()
        if rep_col>=0:
            for r in range(self.table.rowCount()):
                it=self.table.item(r,rep_col)
                if it is None:
                    it=QTableWidgetItem(''); self.table.setItem(r,rep_col,it)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def _populate_table_with_extras(self, df, orig_cols:list[str], part_idx:int, colors_2d=None):
        headers=['Select']
        for i,name in enumerate(orig_cols):
            headers.append(str(name))
            if i==part_idx: headers.append('Replacement')
        self.table.clear(); self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers); self.table.setRowCount(len(df))
        obs_map=self._build_obs_map()
        def map_col(c:int)->int: return 1 + c + (1 if c>part_idx else 0)
        for r in range(len(df)):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); cont._chk=chk; self.table.setCellWidget(r,0,cont)
            for c in range(len(orig_cols)):
                val=df.iloc[r,c]; txt='' if val is None else str(val)
                item=QTableWidgetItem(txt)
                if colors_2d and r<len(colors_2d) and c<len(colors_2d[r]):
                    col=colors_2d[r][c]
                    if isinstance(col,QColor): item.setBackground(col)
                self.table.setItem(r, map_col(c), item)
            part_val=str(df.iloc[r,part_idx]) if df.iloc[r,part_idx] is not None else ''
            part_key=part_val.strip().upper(); replacement=obs_map.get(part_key,'')
            rep_item=QTableWidgetItem(replacement); rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if colors_2d and r<len(colors_2d) and part_idx<len(colors_2d[r]):
                col=colors_2d[r][part_idx]
                if isinstance(col,QColor): rep_item.setBackground(col)
            self.table.setItem(r, 1+part_idx+1, rep_item)
        header=self.table.horizontalHeader()
        for i in range(len(headers)):
            if i==len(headers)-1: header.setSectionResizeMode(i,QHeaderView.ResizeMode.Stretch)
            else: header.setSectionResizeMode(i,QHeaderView.ResizeMode.ResizeToContents)

    # -----------------------------------------------------------
    def import_where_used(self):
        try:
            path,_=QFileDialog.getOpenFileName(self, "Import 'Where Used' Parents", '', "Excel/CSV/HTML (*.xlsx *.xls *.csv *.htm *.html);;All Files (*.*)")
            if not path: return
            import pandas as pd
            df=None; errors=[]; lower=path.lower()
            needs_excel_conversion = lower.endswith('.xls') or self._is_html_like(path)
            converted_path=None
            if needs_excel_conversion:
                converted_path=self._convert_to_xlsx_via_excel(path)
                if converted_path:
                    try: df=pd.read_excel(converted_path, engine='openpyxl')
                    except Exception as e: errors.append(f'Converted .xlsx read failed: {e}')
            if df is None:
                try:
                    if lower.endswith('.xlsx'): df=pd.read_excel(path, engine='openpyxl')
                except Exception as e: errors.append(f'XLSX read failed: {e}')
            if df is None and (lower.endswith('.htm') or lower.endswith('.html') or self._is_html_like(path)):
                try:
                    tables=pd.read_html(path)
                    if tables: df=tables[0]
                except Exception as e: errors.append(f'HTML parse failed: {e}')
            if df is None and lower.endswith('.csv'):
                try: df=pd.read_csv(path)
                except Exception as e: errors.append(f'CSV comma failed: {e}')
                if df is None:
                    try: df=pd.read_csv(path, sep='\t')
                    except Exception as e: errors.append(f'CSV tab failed: {e}')
            if df is None:
                try: df=pd.read_excel(path)
                except Exception as e: errors.append(f'Generic read_excel failed: {e}')
            if df is None:
                msg = "Failed to open the file. Tried Excel (with Excel-based conversion), HTML and CSV paths.\n" + "\n".join(errors[-6:])
                QMessageBox.warning(self,'Import Error', msg); return
            cols=[str(c) if c is not None else '' for c in df.columns]
            part_idx=self._find_part_col_index(cols)
            if part_idx<0:
                QMessageBox.warning(self,'Missing Column', "Couldn't find a 'Part' column (case-insensitive) in the selected file."); return
            # Optional OBS-only prefilter
            if hasattr(self,'_obs_only_filter') and self._obs_only_filter:
                df=self._apply_obs_only_filter(df, part_idx)
            colors_2d=None
            xlsx_source=converted_path if converted_path else (path if path.lower().endswith('.xlsx') else None)
            if xlsx_source: colors_raw=self._read_xlsx_background_colors(xlsx_source, target_ncols=len(cols))
            else: colors_raw=None
            df_kept, mask=self._apply_cleanup_rules(df, part_idx)
            if colors_raw is not None:
                kept_colors=[]; mask_list=mask.tolist()
                for ok,row_colors in zip(mask_list, colors_raw):
                    if ok: kept_colors.append(row_colors)
                colors_2d=kept_colors
            self._populate_table_with_extras(df_kept, list(df_kept.columns), part_idx, colors_2d=colors_2d)
            self._center_replacement_column()
            msg=f"Imported {len(df_kept)} cleaned row(s) from:\n{converted_path or path}"
            if converted_path: msg+="\n(The source file was auto-converted to .xlsx using Excel.)"
            QMessageBox.information(self,'Import Complete', msg)
        except Exception as e:
            QMessageBox.warning(self,'Import Error', str(e))

    # ----------------- Extra actions -----------------
    def import_where_used_of_obs_multi_level(self):
        try:
            self._obs_only_filter=True; self.import_where_used()
        finally:
            if hasattr(self,'_obs_only_filter'): delattr(self,'_obs_only_filter')
    def _apply_obs_only_filter(self, df, part_idx:int):
        obs_map=self._build_obs_map()
        if not obs_map: return df
        keys=set(obs_map.keys()); col=df.iloc[:,part_idx].astype(str).fillna('')
        mask=col.str.strip().str.upper().isin(keys)
        return df.loc[mask].reset_index(drop=True)
    def _selected_row_indices(self):
        rows=[]
        for r in range(self.table.rowCount()):
            w=self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk') and w._chk.isChecked(): rows.append(r)
        return rows
    def _headers(self)->list[str]:
        return [self.table.horizontalHeaderItem(i).text() if self.table.horizontalHeaderItem(i) else '' for i in range(self.table.columnCount())]
    def _map_original_to_table_col(self, orig_idx:int, part_idx:int)->int:
        rep_col=self._find_replacement_col()
        if part_idx>=0 and rep_col>=0 and (orig_idx>(rep_col-2)): return 1+orig_idx+1
        else: return 1+orig_idx
    def select_all_9024_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Select 9024 Parents', "Couldn't find a 'Parent' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=it.text() if it else ''
            self.table.cellWidget(r,0)._chk.setChecked(txt.strip().startswith('9024'))
    def select_all_options(self):
        headers=self._headers(); cidx=self._find_class_col_index(headers[1:])
        if cidx<0: QMessageBox.information(self,'Select Options/O Class', "Couldn't find a 'Class/Type/Category' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(cidx, part_idx)
            it=self.table.item(r, tbl_col); t=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(('OPTION' in t) or ('O CLASS' in t))
    def select_esw_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0:
            pidx=self._find_part_col_index(headers[1:])
            if pidx<0: QMessageBox.information(self,'Select ESW Parents', "Couldn't find 'Parent' or 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(txt.startswith('ESW'))
    def delete_selected_rows(self):
        rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Delete Selected','No rows are selected (checkbox).'); return
        for r in reversed(rows): self.table.removeRow(r)
    def move_selected_to_structure(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'append_rows'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.'); return
            headers=self._headers(); rows_idx=self._selected_row_indices()
            if not rows_idx: QMessageBox.information(self,'Move to Structure Sheet','No rows selected.'); return
            data=[]
            for r in rows_idx:
                row=[(self.table.item(r,c).text() if self.table.item(r,c) else '') for c in range(1,self.table.columnCount())]
                data.append(row)
            target.append_rows(headers[1:], data)
            QMessageBox.information(self,'Move to Structure Sheet', f'Moved {len(data)} row(s) to Structure Sheet.')
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))
    def _append_obs_part(self, part:str):
        try:
            if not self.obs_provider: return
            t=self.obs_provider.table; key=(part or '').strip().upper()
            if not key: return
            existing=set()
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it: existing.add((it.text() or '').strip().upper())
            if key in existing: return
            target=None
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it is None or not (it.text() or '').strip(): target=r; break
            if target is None:
                target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
            t.setItem(target,1,QTableWidgetItem(part))
            w=t.cellWidget(target,2)
            if isinstance(w,QComboBox): idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
        except Exception: pass
    def append_selected_to_obs(self):
        headers=self._headers(); pidx=self._find_part_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Append to OBS', "Couldn't find a 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:]); rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Append to OBS','No rows selected.'); return
        count=0
        for r in rows:
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r,tbl_col); part=it.text() if it else ''
            if part.strip(): self._append_obs_part(part); count+=1
        QMessageBox.information(self,'Append to OBS', f'Appended {count} part(s) to OBS List.')
    def export_where_used(self):
        try:
            dialog = QFileDialog(self, 'Export WhereUsed', 'WhereUsed.xlsx', 'Excel Files (*.xlsx)')
            dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
            dialog.setFileMode(QFileDialog.FileMode.AnyFile)
            dialog.setDefaultSuffix('xlsx')
            if not dialog.exec():
                return
            files = dialog.selectedFiles()
            path = files[0] if files else ''
            if not path: return
            if not path.lower().endswith('.xlsx'):
                path += '.xlsx'
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            wb=Workbook(); ws=wb.active; ws.title='WhereUsed'
            headers=self._headers(); ws.append(headers)
            for c in range(1,len(headers)+1):
                cell=ws.cell(row=1,column=c); cell.font=Font(bold=True)
                if headers[c-1].strip().lower()=='replacement': cell.alignment=Alignment(horizontal='center')
            for r in range(self.table.rowCount()):
                row_vals=[]
                for c in range(self.table.columnCount()):
                    if c==0:
                        w=self.table.cellWidget(r,0); row_vals.append('Yes' if (w and hasattr(w,'_chk') and w._chk.isChecked()) else 'No')
                    else:
                        it=self.table.item(r,c); row_vals.append(it.text() if it else '')
                ws.append(row_vals)
                wu_item = self.table.item(r, 1)
                row_fill_rgb = 'C7DEFA' if (wu_item and wu_item.text().strip() == '0') else None
                for c in range(self.table.columnCount()):
                    it=self.table.item(r,c)
                    if row_fill_rgb:
                        ws.cell(row=r+2,column=c+1).fill=PatternFill('solid', fgColor=row_fill_rgb)
                    elif it and it.background().style() != Qt.BrushStyle.NoBrush:
                        qcol=it.background().color()
                        if qcol.isValid() and qcol.alpha() > 0:
                            rgb=f"{qcol.red():02X}{qcol.green():02X}{qcol.blue():02X}"; ws.cell(row=r+2,column=c+1).fill=PatternFill('solid', fgColor=rgb)
                    if headers[c].strip().lower()=='replacement': ws.cell(row=r+2,column=c+1).alignment=Alignment(horizontal='center')
            wb.save(path); QMessageBox.information(self,'Export Complete', f'Exported {self.table.rowCount()} row(s) to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self,'Export Error', str(e))
    def reset_where_used(self):
        try:
            self.table.clear(); self.table.setRowCount(0); self.table.setColumnCount(0)
            QMessageBox.information(self,'Reset','Where Used view has been reset.')
        except Exception as e:
            QMessageBox.warning(self,'Reset Error', str(e))

    def import_from_databricks(self, obs_parts: List[str], max_level: int, plant: str = "4070"):
        """Import multi-level Where Used from Databricks for the given OBS parts.

        Called by OBSPartsTab.launch_where_used_import() after input validation.
        Inputs are already validated (obs_parts non-empty, max_level 1–6).
        """
        # ── Confirm overwrite of existing data ────────────────────────────────
        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            reply = QMessageBox.question(
                self,
                'Where Used – Existing Data',
                'The Where Used tab already contains data.\n'
                'Delete it and import fresh data from Databricks?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        # ── Import the query module ────────────────────────────────────────────
        try:
            import sys as _sys
            import importlib as _il
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from where_used_query import fetch_where_used, DISPLAY_HEADERS  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(
                self, 'Module Not Found',
                f'where_used_query.py could not be imported:\n{exc}'
            )
            return

        # ── Query Databricks ───────────────────────────────────────────────────
        try:
            records = fetch_where_used(obs_parts, max_level, plant)
        except Exception as exc:
            QMessageBox.warning(self, 'Databricks Query Error', str(exc))
            return

        if not records:
            QMessageBox.information(
                self, 'No Data',
                f'Databricks returned no results for {len(obs_parts)} OBS part(s) '
                f'at max WU level {max_level}.'
            )
            return

        # ── Build OBS map for Replacement column ──────────────────────────────
        obs_map = self._build_obs_map()

        # ── Column layout ──────────────────────────────────────────────────────
        # all_headers: [Select=0, WU Level=1, Part=2, Replacement=3, Rev/Ln=4, ...]
        all_headers = ['Select'] + DISPLAY_HEADERS
        _WU_COL     = 1
        _PART_COL   = 2
        _REPL_COL   = 3
        _DATA_START = 4   # Rev/Ln and onwards

        # Lambdas that extract the display value for each column from col 4 onwards.
        # Order must match DISPLAY_HEADERS[3:] (i.e. after WU Level / Part / Replacement).
        _DB_COL_FUNCS = [
            lambda r: r.get('rev_ln', ''),
            lambda r: r.get('plant', ''),
            lambda r: r.get('description', ''),
            lambda r: r.get('item_status', ''),
            lambda r: r.get('base_qty', ''),
            lambda r: r.get('ext_qty', ''),
            lambda r: r.get('uom', ''),
            lambda r: r.get('eco_number', ''),
            lambda r: r.get('procurement_type', ''),
            lambda r: r.get('effectivity_date', ''),
            lambda r: r.get('user_item_type', ''),
            lambda r: r.get('item_seq', ''),
            lambda r: r.get('kit_code', ''),
            lambda r: r.get('sparable_flag', ''),
            lambda r: r.get('designator', ''),
            lambda r: r.get('option_class', ''),
            lambda r: (r.get('pace_or_dash', '') if 'pace' in r.get('pace_or_dash', '').lower() else ''),
            lambda r: r.get('mlo_class', ''),
        ]

        # ── Populate table ─────────────────────────────────────────────────────
        self.table.clear()
        self.table.setColumnCount(len(all_headers))
        self.table.setHorizontalHeaderLabels(all_headers)
        self.table.setRowCount(len(records))

        for row_idx, record in enumerate(records):
            # Select checkbox
            chk  = QCheckBox()
            cont = QWidget()
            h    = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            cont._chk = chk
            self.table.setCellWidget(row_idx, 0, cont)

            # WU Level – numeric value from Databricks
            wu_val = record.get('wu_level', '')
            self.table.setItem(row_idx, _WU_COL, QTableWidgetItem(wu_val))

            # Part – display with leading spaces (2 spaces per WU level) for hierarchy
            raw_part = record.get('part', '')
            try:
                level_int = int(wu_val)
            except (ValueError, TypeError):
                level_int = 0
            indented_part = ('      ' * level_int) + raw_part
            self.table.setItem(row_idx, _PART_COL, QTableWidgetItem(indented_part))

            # Replacement – auto-fill from OBS map using the raw (unindented) part key
            replacement = obs_map.get(raw_part.strip().upper(), '')
            rep_item = QTableWidgetItem(replacement)
            rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, _REPL_COL, rep_item)

            # Remaining Databricks columns
            for c_off, fn in enumerate(_DB_COL_FUNCS):
                self.table.setItem(row_idx, _DATA_START + c_off,
                                   QTableWidgetItem(fn(record)))

            # Blue background for level-0 rows (the input OBS parts)
            if wu_val == '0':
                _blue = QColor('#C7DEFA')
                cont.setStyleSheet('background-color: #C7DEFA;')
                for _col in range(1, len(all_headers)):
                    _item = self.table.item(row_idx, _col)
                    if _item is not None:
                        _item.setBackground(_blue)

        # ── Resize columns ─────────────────────────────────────────────────────
        hdr = self.table.horizontalHeader()
        for i in range(len(all_headers)):
            if i == len(all_headers) - 1:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

        QMessageBox.information(
            self, 'Import Complete',
            f'Imported {len(records)} row(s) from Databricks.\n'
            f'OBS parts: {len(obs_parts)}  |  Max WU level: {max_level}  |  Plant: {plant}'
        )


class WhereUsedTabV2(WhereUsedTab):
    """UI-customized Where Used variant with revised action buttons and selection rules.
    Extends base import behavior while preserving data handling and table structure."""
    def __init__(self, obs_provider=None):
        super().__init__(obs_provider=obs_provider)
        # Hide legacy buttons/row
        try:
            # Hide old buttons if they exist
            for btn_name in [
                'btn_import_obs_multi','btn_sel_9024','btn_sel_options','btn_sel_esw',
                'btn_delete_sel','btn_move_to_struct','btn_append_obs','btn_export','btn_reset'
            ]:
                btn = getattr(self, btn_name, None)
                if btn is not None:
                    btn.hide()
        except Exception:
            pass
        
        # Build new two-row action panel
        from PyQt6.QtWidgets import QWidget, QGridLayout, QLabel
        panel = QWidget(self)
        grid = QGridLayout(panel)
        grid.setContentsMargins(0,0,0,0)
        grid.setHorizontalSpacing(6)
        grid.setVerticalSpacing(6)

        # Buttons (new)
        self.v2_btn_sel_9024 = QPushButton("Select all 9024 Parents")
        self.v2_btn_sel_opt_struct = QPushButton("Select Options/Class to Create Structure Sheet")
        self.v2_btn_sel_esw = QPushButton("Select ESW Parents")
        self.v2_btn_sel_above_cfg = QPushButton("Select above Config")
        self.v2_btn_move = QPushButton("Move to Structure Sheet")
        self.v2_btn_append = QPushButton("Append to OBS List")
        self.v2_btn_refresh = QPushButton("Refresh")
        self.v2_btn_delete = QPushButton("Delete selected")
        self.v2_btn_reset = QPushButton("Reset Tab")
        self.v2_btn_export = QPushButton("Export WhereUsed")
        # Place Import first, then Export (no custom colors)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # Import button is the last widget; insert Export right after Import
            title_row.insertWidget(title_row.count(), self.v2_btn_export)
        except Exception:
            pass
        # Move Export button to title row (right side)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # export will be inserted after import
        except Exception:
            pass

        # Colors (smooth, subtle gradients)
        def btn_style(bg1, bg2, border, text='#FFFFFF'):
            return f"""
            QPushButton {{
                color:{text}; padding:4px 8px; font-size:11px; border-radius:5px;
                border:1px solid {border};
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg2});
            }}
            QPushButton:hover {{
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg1});
            }}
            """
        
        common_blue = btn_style('#26C6DA','#14A7BE','#0F93AA')
        refresh_green = btn_style('#66BB6A','#4EA85A','#3F8F4A')
        reset_gray = btn_style('#90A4AE','#7C919B','#6A7E87')
        delete_red = btn_style('#EF5350','#D32F2F','#B71C1C')

        # Compact sizing – let buttons fit their text
        for btn in [
            self.v2_btn_sel_9024, self.v2_btn_sel_opt_struct, self.v2_btn_sel_esw,
            self.v2_btn_sel_above_cfg, self.v2_btn_move, self.v2_btn_append,
            self.v2_btn_delete, self.v2_btn_export, self.v2_btn_refresh, self.v2_btn_reset,
        ]:
            btn.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)

        # Fit button widths to text so labels are fully visible.
        def _fit_btn(btn: QPushButton, pad: int = 24, min_h: int = 31):
            fm: QFontMetrics = btn.fontMetrics()
            btn.setFixedWidth(fm.horizontalAdvance(btn.text()) + pad)
            btn.setFixedHeight(min_h)

        self.v2_btn_sel_9024.setStyleSheet(common_blue)
        self.v2_btn_sel_esw.setStyleSheet(common_blue)
        self.v2_btn_sel_above_cfg.setStyleSheet(common_blue)
        self.v2_btn_sel_opt_struct.setStyleSheet(common_blue)
        self.v2_btn_append.setStyleSheet(common_blue)
        self.v2_btn_move.setStyleSheet(common_blue)
        # Keep Export styling similar to Import button (default look).
        self.v2_btn_export.setStyleSheet('')
        self.v2_btn_delete.setStyleSheet(delete_red)
        self.v2_btn_refresh.setStyleSheet(refresh_green)
        self.v2_btn_reset.setStyleSheet(reset_gray)

        for btn in [
            self.v2_btn_sel_9024,
            self.v2_btn_sel_esw,
            self.v2_btn_sel_above_cfg,
            self.v2_btn_sel_opt_struct,
            self.v2_btn_append,
            self.v2_btn_move,
            self.v2_btn_delete,
            self.v2_btn_refresh,
            self.v2_btn_reset,
            self.v2_btn_export,
        ]:
            _fit_btn(btn)

        # Keep title-row Import/Export controls aligned right and sized to text.
        for btn in self.findChildren(QPushButton):
            txt = (btn.text() or '').strip()
            if txt in {"Import 'Where Used' Parents", 'Export WhereUsed'}:
                _fit_btn(btn)

        # Button layout: single compact row, matching app style.
        row1 = [
            self.v2_btn_sel_9024,
            self.v2_btn_sel_esw,
            self.v2_btn_sel_above_cfg,
            self.v2_btn_sel_opt_struct,
            self.v2_btn_move,
            self.v2_btn_append,
        ]
        for c, btn in enumerate(row1):
            grid.addWidget(btn, 0, c)
        grid.setColumnStretch(len(row1), 1)
        grid.addWidget(self.v2_btn_delete, 0, len(row1) + 1)
        grid.addWidget(self.v2_btn_reset, 0, len(row1) + 2)
        grid.addWidget(self.v2_btn_refresh, 0, len(row1) + 3)

        # Insert our panel right after the title row (index 1)
        try:
            outer: QVBoxLayout = self.layout()
            # self.layout() returns the QWidget layout of WhereUsedTab (outer QVBoxLayout)
            outer.insertWidget(1, panel)
        except Exception:
            # If insertion fails, just add at end
            self.layout().addWidget(panel)

        # Wire up actions
        self.v2_btn_sel_9024.clicked.connect(self._v2_select_9024_by_part)
        self.v2_btn_sel_opt_struct.clicked.connect(self._v2_select_options_for_structure_sheet)
        self.v2_btn_sel_esw.clicked.connect(self._v2_select_esw_by_part)
        self.v2_btn_sel_above_cfg.clicked.connect(self._v2_select_above_config_block)
        self.v2_btn_move.clicked.connect(self._v2_move_selected_part_only)
        self.v2_btn_append.clicked.connect(self._v2_append_selected_with_replacement)
        self.v2_btn_refresh.clicked.connect(self._v2_refresh_replacements)
        self.v2_btn_delete.clicked.connect(self.delete_selected_rows)
        self.v2_btn_reset.clicked.connect(self.reset_where_used)
        self.v2_btn_export.clicked.connect(self.export_where_used)

          # Keep refresh/reset in distinct colors; other buttons remain common blue.

        # Also hide the legacy "Import Where Used of OBS Parts (Multiple Level)" if present in title row
        try:
            if hasattr(self, 'btn_import_obs_multi'):
                self.btn_import_obs_multi.hide()
        except Exception:
            pass

    # ---------- Helpers to find current table columns ----------
    def _v2_find_table_col_index(self, header_exact: str) -> int:
        name = (header_exact or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and (h.text() or '').strip().lower() == name:
                return i
        return -1

    def _v2_find_table_col_contains(self, keyword: str) -> int:
        key = (keyword or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and key in (h.text() or '').strip().lower():
                return i
        return -1

    # ---------- New selection logics ----------
    def _v2_select_9024_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select 9024 Parents', "Couldn't find a 'Part' column.")
            return
        wucol = self._v2_find_table_col_contains('wu level')
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
            ok_prefix = part.startswith('9024')
            ok_wu = True
            if wucol >= 0:
                val = (self.table.item(r, wucol).text() if self.table.item(r, wucol) else '').strip()
                try:
                    ok_wu = int(val) != 0
                except Exception:
                    ok_wu = (val != '0')
            chk = ok_prefix and ok_wu
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    def _v2_select_opt_optclass_by_prefix(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select Opt & Opt Class', "Couldn't find a 'Part' column.")
            return
        prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
            '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
            '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
        }
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
            first4 = part[:4]
            chk = first4 in prefixes
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    def _v2_select_esw_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select ESW Parents', "Couldn't find a 'Part' column.")
            return
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip().upper()
            chk = part.startswith('ESW')
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    # ---------- Move / Append / Refresh ----------
    
    def _v2_toggle_rows(self, rows):
        checks = []
        for r in rows:
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk'):
                checks.append(w._chk)
        if not checks:
            return
        select = any(not c.isChecked() for c in checks)
        for c in checks:
            c.setChecked(select)

    def _v2_select_9024_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select 9024 Parents',"Couldn't find a 'Part' column.")
            return
        wucol = self._v2_find_table_col_contains('wu level')
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip()
            ok = part.startswith('9024')
            if ok and wucol >= 0:
                wu = (self.table.item(r,wucol).text() if self.table.item(r,wucol) else '').strip()
                ok = wu != '0'
            if ok:
                rows.append(r)
        self._v2_toggle_rows(rows)

    def _v2_select_opt_optclass_by_prefix(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select Opt & Opt Class',"Couldn't find a 'Part' column.")
            return
        prefixes = {'0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335','0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465','0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'}
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip()
            if part[:4] in prefixes:
                rows.append(r)
        self._v2_toggle_rows(rows)

    def _v2_select_esw_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select ESW Parents',"Couldn't find a 'Part' column.")
            return
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip().upper()
            if part.startswith('ESW'):
                rows.append(r)
        self._v2_toggle_rows(rows)

    def _v2_select_options_for_structure_sheet(self):
        """Select Option/Class rows and their immediate direct children in each WU block.

        Example: if an Option is at WU level 3, select that row plus only level-4
        children under that row. Indentation in Part text is used as a hierarchy
        guard to avoid selecting unrelated rows.

        Some datasets list the linked item one level in the opposite direction;
        if no direct child is found, we fallback to the nearest one-level linked row.
        """
        part_col = self._v2_find_table_col_index('Part')
        wu_col = self._v2_find_table_col_contains('wu level')
        if part_col < 0 or wu_col < 0:
            QMessageBox.information(
                self,
                'Select Options/Class to Create Structure Sheet',
                "Couldn't find required 'Part' and 'WU Level' columns.",
            )
            return

        option_prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
            '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
            '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
        }

        def _row_meta(row_idx: int):
            part_item = self.table.item(row_idx, part_col)
            part_raw = (part_item.text() if part_item else '') or ''
            part_trim = part_raw.lstrip(' \t')
            indent = len(part_raw) - len(part_trim)

            wu_item = self.table.item(row_idx, wu_col)
            wu_raw = (wu_item.text() if wu_item else '') or ''
            try:
                wu_level = int(float(wu_raw.strip()))
            except Exception:
                wu_level = -1

            return part_trim, indent, wu_level

        rows = self.table.rowCount()
        target_rows = set()

        r = 0
        while r < rows:
            _, _, curr_wu = _row_meta(r)
            if curr_wu != 0:
                r += 1
                continue

            block_start = r
            block_end = rows
            for i in range(r + 1, rows):
                _, _, wu_i = _row_meta(i)
                if wu_i == 0:
                    block_end = i
                    break

            for i in range(block_start + 1, block_end):
                part_trim_i, indent_i, wu_i = _row_meta(i)
                if not part_trim_i:
                    continue
                if part_trim_i[:4] not in option_prefixes:
                    continue

                # Always include the Option/Class row itself.
                target_rows.add(i)

                # If Option/Class is at level 1, also include level-0 parent row
                # for this WU block (the block starts at level 0).
                if wu_i == 1:
                    _, indent_0, wu_0 = _row_meta(block_start)
                    if wu_0 == 0 and indent_0 < indent_i:
                        target_rows.add(block_start)

                # Include only immediate direct children one WU level below,
                # inside this option/class subtree.
                child_level = wu_i + 1
                matched_direct_child = False
                for j in range(i + 1, block_end):
                    _, indent_j, wu_j = _row_meta(j)

                    # End of this option/class subtree.
                    if wu_j >= 0 and wu_j <= wu_i and indent_j <= indent_i:
                        break

                    # Primary rule: one level deeper and deeper indentation.
                    if wu_j == child_level and indent_j > indent_i:
                        target_rows.add(j)
                        matched_direct_child = True

                # Fallback: if no direct child exists, select the nearest linked
                # one-level row in the opposite direction (common in some WU layouts).
                if not matched_direct_child:
                    parent_level = wu_i - 1
                    for j in range(i - 1, block_start - 1, -1):
                        _, indent_j, wu_j = _row_meta(j)
                        if wu_j == parent_level and indent_j < indent_i:
                            target_rows.add(j)
                            break

            r = block_end

        self._v2_toggle_rows(sorted(target_rows))
    def _v2_move_selected_part_only(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'table'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.')
                return
            rows_idx = self._selected_row_indices()
            if not rows_idx:
                QMessageBox.information(self,'Move to Structure Sheet','No rows selected.')
                return

            src_headers = self._headers()
            if len(src_headers) <= 1:
                QMessageBox.information(self,'Move to Structure Sheet','Where Used table has no usable columns.')
                return

            # Drop Select column and rename WU Level -> BOM Level.
            out_headers = list(src_headers[1:])
            wu_out_idx = -1
            for i, h in enumerate(out_headers):
                if (h or '').strip().lower() == 'wu level':
                    out_headers[i] = 'BOM Level'
                    wu_out_idx = i
                    break
            if wu_out_idx < 0:
                for i, h in enumerate(out_headers):
                    if 'wu level' in (h or '').strip().lower():
                        out_headers[i] = 'BOM Level'
                        wu_out_idx = i
                        break

            def _norm_header(h: str) -> str:
                return ' '.join((h or '').strip().lower().split())

            # Remove columns requested by user from Structure Sheet output.
            drop_headers = {
                'rev/ln',
                'plant',
                'eco number',
                'effectivity date',
                'user item type',
                'bom source type (debug)',
            }

            # New leading columns.
            # New leading columns.  Column 1 = 'Action' (user-editable change type).
            leading_headers = [
                'Action',
                'Part Description',
                'Seq#',
                'Qty',
                'Kit Code',
                'Ref Designator(RD)',
            ]

            part_out_idx = -1
            for i, h in enumerate(out_headers):
                if _norm_header(h) == 'part':
                    part_out_idx = i
                    break

            change_type_src_idx = -1
            for i, h in enumerate(out_headers):
                if _norm_header(h) == 'change type':
                    change_type_src_idx = i
                    break

            pcol = self._v2_find_table_col_index('Part')
            wu_col = self._v2_find_table_col_contains('wu level')
            if pcol < 0 or wu_col < 0:
                QMessageBox.information(self,'Move to Structure Sheet',"Couldn't find required 'Part' and 'WU Level' columns.")
                return

            option_prefixes = {
                '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
                '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
                '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
            }
            selected_set = set(rows_idx)

            def _src_row_values(row_idx: int):
                return [
                    (self.table.item(row_idx, c).text() if self.table.item(row_idx, c) else '')
                    for c in range(1, self.table.columnCount())
                ]

            def _part_clean(row_idx: int) -> str:
                it = self.table.item(row_idx, pcol)
                return ((it.text() if it else '') or '').lstrip(' \t').strip()

            def _meta(row_idx: int):
                p_raw = (self.table.item(row_idx, pcol).text() if self.table.item(row_idx, pcol) else '') or ''
                p_trim = p_raw.lstrip(' \t')
                indent = len(p_raw) - len(p_trim)
                w_raw = (self.table.item(row_idx, wu_col).text() if self.table.item(row_idx, wu_col) else '') or ''
                try:
                    lvl = int(float(w_raw.strip()))
                except Exception:
                    lvl = -1
                return p_trim.strip(), indent, lvl

            # Precompute WU level-0 blocks.
            rows_total = self.table.rowCount()
            block_bounds = []
            start = 0
            for r in range(rows_total):
                _, _, lvl = _meta(r)
                if lvl == 0:
                    if block_bounds:
                        prev_s, _ = block_bounds[-1]
                        block_bounds[-1] = (prev_s, r)
                    block_bounds.append((r, rows_total))
            if not block_bounds:
                block_bounds = [(0, rows_total)]

            # Build option groups, de-duplicated by option part.
            option_order = []
            option_groups = {}  # key -> {'option_row': list[str], 'children': [list[str]], 'seen_children': set[str]}

            def _find_block_for_row(row_idx: int):
                for bs, be in block_bounds:
                    if bs <= row_idx < be:
                        return bs, be
                return 0, rows_total

            # Candidate options are selected option-prefix rows.
            for i in rows_idx:
                part_i, indent_i, wu_i = _meta(i)
                if not part_i or part_i[:4] not in option_prefixes:
                    continue

                opt_key = part_i.upper()
                if opt_key not in option_groups:
                    option_order.append(opt_key)
                    option_groups[opt_key] = {
                        'option_row': _src_row_values(i),
                        'children': [],
                        'seen_children': set(),
                    }

                bs, be = _find_block_for_row(i)

                # Include selected linked rows inside the option subtree.
                for j in range(i + 1, be):
                    _, indent_j, wu_j = _meta(j)
                    if wu_j >= 0 and wu_j <= wu_i and indent_j <= indent_i:
                        break
                    if j not in selected_set:
                        continue
                    child_part = _part_clean(j).upper()
                    if not child_part or child_part == opt_key:
                        continue
                    if child_part in option_groups[opt_key]['seen_children']:
                        continue
                    option_groups[opt_key]['children'].append(_src_row_values(j))
                    option_groups[opt_key]['seen_children'].add(child_part)

                # Include nearest selected one-level parent as the first linked row.
                parent_candidate = None
                for j in range(i - 1, bs - 1, -1):
                    _, indent_j, wu_j = _meta(j)
                    if wu_j == (wu_i - 1) and indent_j < indent_i:
                        if j in selected_set:
                            parent_candidate = j
                        break

                if parent_candidate is not None:
                    parent_part = _part_clean(parent_candidate).upper()
                    if parent_part and parent_part not in option_groups[opt_key]['seen_children']:
                        option_groups[opt_key]['children'].insert(0, _src_row_values(parent_candidate))
                        option_groups[opt_key]['seen_children'].add(parent_part)

            if not option_order:
                QMessageBox.information(
                    self,
                    'Move to Structure Sheet',
                    'No selected Option/Class rows were found to build Structure Sheet output.',
                )
                return

            # Decide move mode when overlaps exist in current Structure Sheet.
            move_mode = 'replace'
            existing_full_duplicates = []
            overlap_options = []
            overlap_part_count = 0
            if target.table.columnCount() > 0 and target.table.rowCount() > 0:
                t_existing = target.table
                bom_col_ex = -1
                part_col_ex = -1
                for c in range(t_existing.columnCount()):
                    h = t_existing.horizontalHeaderItem(c)
                    if not h:
                        continue
                    nh = _norm_header(h.text())
                    if nh == 'bom level':
                        bom_col_ex = c
                    elif nh == 'part':
                        part_col_ex = c

                if bom_col_ex >= 0 and part_col_ex >= 0:
                    existing_bom_map = {}  # parent_part -> set(child_parts)
                    existing_all_parts = set()
                    current_parent = None
                    for rr in range(t_existing.rowCount()):
                        bom_item = t_existing.item(rr, bom_col_ex)
                        part_item = t_existing.item(rr, part_col_ex)
                        bom_txt = (bom_item.text() if bom_item else '').strip()
                        part_txt = ((part_item.text() if part_item else '') or '').lstrip(' \t').strip().upper()
                        if not part_txt:
                            continue
                        existing_all_parts.add(part_txt)
                        if bom_txt == '0':
                            current_parent = part_txt
                            existing_bom_map.setdefault(current_parent, set())
                        elif bom_txt == '1' and current_parent:
                            existing_bom_map.setdefault(current_parent, set()).add(part_txt)

                    for opt_key in option_order:
                        grp = option_groups.get(opt_key, {})
                        children = grp.get('children', [])
                        child_set = set()
                        for ch in children:
                            if 0 <= part_out_idx < len(ch):
                                pch = ((ch[part_out_idx] or '') if isinstance(ch[part_out_idx], str)
                                       else str(ch[part_out_idx] or '')).lstrip(' \t').strip().upper()
                                if pch and pch != opt_key:
                                    child_set.add(pch)

                        if opt_key in existing_bom_map:
                            overlap_options.append(opt_key)
                            overlap_part_count += len(child_set.intersection(existing_bom_map.get(opt_key, set())))

                        if opt_key in existing_bom_map and child_set and child_set.issubset(existing_bom_map.get(opt_key, set())):
                            existing_full_duplicates.append(opt_key)

                    if overlap_options or overlap_part_count > 0:
                        msg = QMessageBox(self)
                        msg.setIcon(QMessageBox.Icon.Question)
                        msg.setWindowTitle('Move to Structure Sheet')
                        msg.setText(
                            'Some selected Option(s)/Part(s) already exist in Structure Sheet.\n\n'
                            'Choose how to proceed:'
                        )
                        msg.setInformativeText(
                            'Replace Existing Data: overwrite current Structure Sheet with selected move data.\n'
                            'Merge Data: keep existing rows and update Replacement for matching BOM items under each Option.'
                        )
                        btn_replace = msg.addButton('Replace Existing Data', QMessageBox.ButtonRole.AcceptRole)
                        btn_merge = msg.addButton('Merge Data', QMessageBox.ButtonRole.ActionRole)
                        btn_cancel = msg.addButton(QMessageBox.StandardButton.Cancel)
                        msg.exec()
                        clicked = msg.clickedButton()
                        if clicked == btn_cancel:
                            return
                        if clicked == btn_merge:
                            move_mode = 'merge'

            if not option_order:
                if existing_full_duplicates:
                    preview = ', '.join(existing_full_duplicates[:10])
                    more = '' if len(existing_full_duplicates) <= 10 else f' and {len(existing_full_duplicates) - 10} more'
                    QMessageBox.information(
                        self,
                        'Move to Structure Sheet',
                        f'Already exists in Structure Sheet: {preview}{more}.'
                    )
                else:
                    QMessageBox.information(
                        self,
                        'Move to Structure Sheet',
                        'No selected Option/Class rows were found to build Structure Sheet output.',
                    )
                return

            # Compose output rows: Option first (BOM Level 0), linked rows next (BOM Level 1).
            out_rows = []

            def _split_ws(val: str):
                raw = val or ''
                body = raw.lstrip(' \t')
                ws = raw[:len(raw) - len(body)]
                return ws, body

            for opt_key in option_order:
                grp = option_groups[opt_key]
                opt_row = list(grp['option_row'])
                if wu_out_idx >= 0 and wu_out_idx < len(opt_row):
                    opt_row[wu_out_idx] = '0'

                child_rows = [list(ch) for ch in grp['children']]

                # Swap indentation between option and first linked row in Part column.
                if part_out_idx >= 0 and child_rows:
                    opt_part = opt_row[part_out_idx] if part_out_idx < len(opt_row) else ''
                    ch_part = child_rows[0][part_out_idx] if part_out_idx < len(child_rows[0]) else ''
                    opt_ws, opt_body = _split_ws(opt_part)
                    ch_ws, ch_body = _split_ws(ch_part)
                    if part_out_idx < len(opt_row):
                        opt_row[part_out_idx] = ch_ws + opt_body
                    if part_out_idx < len(child_rows[0]):
                        child_rows[0][part_out_idx] = opt_ws + ch_body

                out_rows.append(opt_row)

                for child in child_rows:
                    c_row = list(child)
                    if wu_out_idx >= 0 and wu_out_idx < len(c_row):
                        c_row[wu_out_idx] = '1'
                    out_rows.append(c_row)

            # Build final headers (prepend Select and new columns, remove dropped columns,
            # and avoid duplicate Change Type from source section).
            keep_indices = []
            for i, h in enumerate(out_headers):
                nh = _norm_header(h)
                if nh in drop_headers:
                    continue
                if nh == 'change type':
                    continue
                keep_indices.append(i)
            final_headers = ['Select'] + leading_headers + [out_headers[i] for i in keep_indices]

            # Build final rows with Select column and new leading columns.
            final_rows = []
            for row in out_rows:
                change_type_val = ''
                if 0 <= change_type_src_idx < len(row):
                    change_type_val = row[change_type_src_idx]

                lead = [change_type_val, '', '', '', '', '']
                tail = [row[i] if i < len(row) else '' for i in keep_indices]
                final_rows.append([''] + lead + tail)
            keep_indices = []
            for i, h in enumerate(out_headers):
                nh = _norm_header(h)
                if nh in drop_headers:
                    continue
                if nh == 'change type':
                    continue
                keep_indices.append(i)
            final_headers = ['Select'] + leading_headers + [out_headers[i] for i in keep_indices]

            # Build final rows with Select column and new leading columns.
            final_rows = []
            for row in out_rows:
                change_type_val = ''
                if 0 <= change_type_src_idx < len(row):
                    change_type_val = row[change_type_src_idx]

                lead = [change_type_val, '', '', '', '', '']
                tail = [row[i] if i < len(row) else '' for i in keep_indices]
                final_rows.append([''] + lead + tail)

            # Insert blank 'Ref Designator' column between PACE and MLO Class.
            pace_pos = -1
            mlo_pos = -1
            for ci, hh in enumerate(final_headers):
                nh = _norm_header(hh)
                if 'pace' in nh:
                    pace_pos = ci
                if 'mlo' in nh and 'class' in nh:
                    mlo_pos = ci
            insert_ref_at = -1
            if pace_pos >= 0 and mlo_pos == pace_pos + 1:
                insert_ref_at = mlo_pos
            elif pace_pos >= 0:
                insert_ref_at = pace_pos + 1
            if insert_ref_at >= 0:
                final_headers.insert(insert_ref_at, 'Ref Designator(RD)')
                for fr in final_rows:
                    fr.insert(insert_ref_at, '')

            # Merge mode: keep existing rows and update/append incoming rows.
            merge_added_count = 0
            merge_replacement_updates = 0
            if move_mode == 'merge' and target.table.columnCount() > 0 and target.table.rowCount() > 0:
                t_existing = target.table
                existing_headers = []
                for c in range(t_existing.columnCount()):
                    h = t_existing.horizontalHeaderItem(c)
                    existing_headers.append(h.text() if h else '')

                merged_headers = list(existing_headers)
                existing_header_keys = {_norm_header(h): i for i, h in enumerate(existing_headers)}
                for h in final_headers:
                    if _norm_header(h) not in existing_header_keys:
                        merged_headers.append(h)
                        existing_header_keys[_norm_header(h)] = len(merged_headers) - 1

                bom_col_idx = next((i for i, h in enumerate(merged_headers) if _norm_header(h) == 'bom level'), -1)
                part_col_idx = next((i for i, h in enumerate(merged_headers) if _norm_header(h) == 'part'), -1)
                repl_col_idx = next((i for i, h in enumerate(merged_headers) if _norm_header(h) == 'replacement'), -1)

                if bom_col_idx >= 0 and part_col_idx >= 0:
                    def _align_row(row_vals, source_headers, target_headers):
                        aligned = ['' for _ in target_headers]
                        src_map = {_norm_header(h): i for i, h in enumerate(source_headers)}
                        for target_i, target_h in enumerate(target_headers):
                            source_i = src_map.get(_norm_header(target_h), -1)
                            if 0 <= source_i < len(row_vals):
                                aligned[target_i] = row_vals[source_i]
                        return aligned

                    existing_rows = []
                    for rr in range(t_existing.rowCount()):
                        row_vals = []
                        for cc in range(t_existing.columnCount()):
                            if cc == 0:
                                row_vals.append('')
                            elif cc == 1:
                                w = t_existing.cellWidget(rr, cc)
                                if w and isinstance(w, QComboBox):
                                    row_vals.append((w.currentText() or '').strip())
                                else:
                                    it = t_existing.item(rr, cc)
                                    row_vals.append((it.text() if it else '').strip())
                            else:
                                it = t_existing.item(rr, cc)
                                row_vals.append((it.text() if it else ''))
                        existing_rows.append(_align_row(row_vals, existing_headers, merged_headers))

                    incoming_rows = [_align_row(row, final_headers, merged_headers) for row in final_rows]

                    def _row_key_stream(rows_list):
                        keys = []
                        current_opt = ''
                        for row in rows_list:
                            bom_txt = (row[bom_col_idx] if 0 <= bom_col_idx < len(row) else '').strip()
                            part_txt = ((row[part_col_idx] if 0 <= part_col_idx < len(row) else '') or '').lstrip(' \t').strip().upper()
                            if bom_txt == '0' and part_txt:
                                current_opt = part_txt
                            keys.append((current_opt, bom_txt, part_txt))
                        return keys

                    existing_keys = _row_key_stream(existing_rows)
                    existing_index_by_key = {k: i for i, k in enumerate(existing_keys) if k[2]}
                    incoming_keys = _row_key_stream(incoming_rows)
                    appended_count = 0
                    replacement_updates = 0

                    for inc_i, inc_row in enumerate(incoming_rows):
                        ikey = incoming_keys[inc_i]
                        if not ikey[2]:
                            continue

                        if ikey in existing_index_by_key:
                            ex_i = existing_index_by_key[ikey]
                            if repl_col_idx >= 0 and repl_col_idx < len(inc_row):
                                incoming_rep = (inc_row[repl_col_idx] or '').strip()
                                existing_rep = (existing_rows[ex_i][repl_col_idx] or '').strip()
                                if incoming_rep and existing_rep != incoming_rep:
                                    existing_rows[ex_i][repl_col_idx] = incoming_rep
                                    replacement_updates += 1
                        else:
                            existing_rows.append(list(inc_row))
                            existing_index_by_key[ikey] = len(existing_rows) - 1
                            appended_count += 1

                    final_headers = merged_headers
                    final_rows = existing_rows
                    merge_added_count = appended_count
                    merge_replacement_updates = replacement_updates
                else:
                    QMessageBox.warning(
                        self,
                        'Move to Structure Sheet',
                        'Merge requested, but current Structure Sheet is missing required BOM Level/Part columns. Replacing data instead.'
                    )
                    move_mode = 'replace'

            # Replace Structure Sheet table with rearranged, de-duplicated grouped output.
            t = target.table
            _bulk_guard = hasattr(target, '_struct_item_change_guard')
            if _bulk_guard:
                target._struct_item_change_guard = True
            try:
                t.clear()
                t.setRowCount(0)
                t.setColumnCount(len(final_headers))
                hdr = RotatedColumnsHeader(
                    Qt.Orientation.Horizontal,
                    rotated_columns=range(2, 7),
                    parent=t,
                    group_label='Change Type',
                    group_columns=list(range(2, 7)),
                )
                hdr.setStretchLastSection(True)
                hdr.setSectionsClickable(True)
                t.setHorizontalHeader(hdr)
                t.setHorizontalHeaderLabels(final_headers)
                t.setRowCount(len(final_rows))
                if hasattr(target, '_apply_inserted_header_colors'):
                    target._apply_inserted_header_colors()

                t.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
                t.horizontalHeader().setMinimumHeight(117)  # reduced ~20%
                for vc in range(2, 7):
                    t.horizontalHeader().setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                    t.setColumnWidth(vc, 40)

                bom_col = -1
                part_col = -1
                replacement_col = -1
                for c, h in enumerate(final_headers):
                    if _norm_header(h) == 'bom level':
                        bom_col = c
                    if _norm_header(h) == 'part':
                        part_col = c
                    if _norm_header(h) == 'replacement':
                        replacement_col = c

                vertical_checkbox_cols = {2, 3, 4, 5, 6}  # Description..Reference Designator (shifted by Select + Change Type)
                # Default action is blank until user explicitly selects it.
                change_type_options_l0 = ['', 'Revised', 'Change']
                change_type_options_l1 = ['', 'Repl Item at Same Seq', 'Remove Item', 'Change']

                for r, row in enumerate(final_rows):
                    for c, val in enumerate(row):
                        if c == 0:  # Select column with checkbox
                            cont = QWidget()
                            h_lay = QHBoxLayout(cont)
                            h_lay.setContentsMargins(0, 0, 0, 0)
                            h_lay.addStretch(1)
                            chk = QCheckBox()
                            h_lay.addWidget(chk)
                            h_lay.addStretch(1)
                            cont._chk = chk
                            t.setCellWidget(r, c, cont)
                        elif c == 1:  # Change Type column
                            bom_val = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            rep_val = (str(row[replacement_col]).strip() if (replacement_col >= 0 and replacement_col < len(row)) else '')

                            combo = QComboBox()
                            # BOM Level 0 rows: Revised + Change; non-zero rows: no Revised
                            if bom_val == '0':
                                combo.addItems(change_type_options_l0)
                            else:
                                combo.addItems(change_type_options_l1)

                            combo.setCurrentIndex(0)

                            combo.setStyleSheet('QComboBox { padding: 2px; }')
                            if hasattr(target, '_on_change_type_changed'):
                                combo.currentTextChanged.connect(lambda text, rr=r: target._on_change_type_changed(rr))
                            t.setCellWidget(r, c, combo)
                        elif c in vertical_checkbox_cols:
                            bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            part_here = (str(row[part_col]).lstrip(' \t') if (part_col >= 0 and part_col < len(row)) else '')
                            is_option_part = bool(part_here) and (part_here[:4] in option_prefixes)

                            # BOM 0 rows: only Description(2) and Ref Designator(6) have checkboxes.
                            if bom_here == '0' and c not in {2, 6}:
                                t.setItem(r, c, QTableWidgetItem(''))
                            # Non-option rows: remove Ref Designator checkbox.
                            elif c == 6 and not is_option_part:
                                t.setItem(r, c, QTableWidgetItem(''))
                            else:
                                cont_v = QWidget()
                                hv = QHBoxLayout(cont_v)
                                hv.setContentsMargins(0, 0, 0, 0)
                                hv.addStretch(1)
                                chk_v = QCheckBox()
                                hv.addWidget(chk_v)
                                hv.addStretch(1)
                                cont_v._chk = chk_v
                                if hasattr(target, '_on_selector_checkbox_toggled'):
                                    chk_v.stateChanged.connect(target._on_selector_checkbox_toggled)
                                t.setCellWidget(r, c, cont_v)
                        else:
                            display_val = val
                            if c == part_col:
                                part_txt = ((val or '') if isinstance(val, str) else str(val or '')).lstrip(' \t')
                                bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                                display_val = part_txt if bom_here == '0' else ('      ' + part_txt if part_txt else '')

                            item = QTableWidgetItem(display_val)
                            if c == part_col:
                                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            elif _norm_header(final_headers[c]) == 'description':
                                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            else:
                                item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                            t.setItem(r, c, item)

                    # Keep Part and Description data left-aligned.
                    if part_col >= 0:
                        part_item = t.item(r, part_col)
                        if part_item is not None:
                            part_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    desc_col_final = next(
                        (ci for ci in range(t.columnCount())
                         if _norm_header((t.horizontalHeaderItem(ci) or QTableWidgetItem()).text()) == 'description'),
                        -1
                    )
                    if desc_col_final >= 0:
                        desc_item = t.item(r, desc_col_final)
                        if desc_item is not None:
                            desc_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

                    # Skyblue template for BOM level 0 rows.
                    if bom_col >= 0 and bom_col < len(row) and str(row[bom_col]).strip() == '0':
                        for c in range(t.columnCount()):
                            if c == 0:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                            elif c == 1:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                            elif c in vertical_checkbox_cols:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                                else:
                                    # Plain item (no checkbox for this column on this row)
                                    cell = t.item(r, c)
                                    if cell:
                                        cell.setBackground(QColor('#87CEEB'))
                            else:
                                cell = t.item(r, c)
                                if cell:
                                    cell.setBackground(QColor('#87CEEB'))

                    # Enforce action-first behavior for each row.
                    if hasattr(target, '_on_change_type_changed'):
                        target._on_change_type_changed(r)

                # Final enforcement: keep Part and Description data left-aligned in all rows.
                if part_col >= 0:
                    for rr in range(t.rowCount()):
                        it = t.item(rr, part_col)
                        if it is not None:
                            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                desc_col_final = next(
                    (ci for ci in range(t.columnCount())
                     if _norm_header((t.horizontalHeaderItem(ci) or QTableWidgetItem()).text()) == 'description'),
                    -1
                )
                if desc_col_final >= 0:
                    for rr in range(t.rowCount()):
                        it = t.item(rr, desc_col_final)
                        if it is not None:
                            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

                t.resizeColumnsToContents()
                for vc in range(2, 7):
                    t.setColumnWidth(vc, 40)
                if hasattr(target, '_update_change_type_body_box'):
                    target._update_change_type_body_box()
                if hasattr(target, '_refresh_structure_action_buttons'):
                    target._refresh_structure_action_buttons()
            finally:
                if _bulk_guard:
                    target._struct_item_change_guard = False
            if move_mode == 'merge':
                QMessageBox.information(
                    self,
                    'Move to Structure Sheet',
                    f'Merge completed. Added {merge_added_count} new row(s); updated Replacement in {merge_replacement_updates} existing row(s).',
                )
            else:
                QMessageBox.information(
                    self,
                    'Move to Structure Sheet',
                    f'Created Structure Sheet with {len(option_order)} unique option row(s) and {len(out_rows)} total row(s).',
                )
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))

    def _v2_append_selected_with_replacement(self):
        try:
            if not self.obs_provider:
                QMessageBox.information(self,'Append to OBS', 'OBS Parts tab is not available.')
                return
            t = self.obs_provider.table
            rows = self._selected_row_indices()
            if not rows:
                QMessageBox.information(self,'Append to OBS','No rows selected.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            count = 0
            for r in rows:
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
                rep  = (self.table.item(r, rcol).text() if (rcol>=0 and self.table.item(r, rcol)) else '').strip()
                if not part:
                    continue
                # Check existing
                existing=set()
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it: existing.add((it.text() or '').strip().upper())
                if part.upper() in existing:
                    # If exists, update replacement if empty
                    if rep:
                        for rr in range(t.rowCount()):
                            it=t.item(rr,1)
                            if it and (it.text() or '').strip().upper()==part.upper():
                                t.setItem(rr,3,QTableWidgetItem(rep))
                                break
                    continue
                target=None
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it is None or not (it.text() or '').strip():
                        target=rr; break
                if target is None:
                    target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
                t.setItem(target,1,QTableWidgetItem(part))
                w=t.cellWidget(target,2)
                if isinstance(w,QComboBox):
                    idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
                if rep:
                    t.setItem(target,3,QTableWidgetItem(rep))
                count += 1
            QMessageBox.information(self,'Append to OBS', f'Appended/Updated {count} part(s) to OBS List.')
        except Exception as e:
            QMessageBox.warning(self,'Append Error', str(e))

    def _v2_refresh_replacements(self):
        try:
            obs_map = self._build_obs_map()
            if not obs_map:
                QMessageBox.information(self, 'Refresh', 'No OBS parts found to refresh from.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            if pcol < 0 or rcol < 0:
                QMessageBox.information(self,'Refresh','Missing Part or Replacement column.')
                return
            updated = 0
            for r in range(self.table.rowCount()):
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '')
                key = (part or '').strip().upper()
                rep = obs_map.get(key, '')
                it = self.table.item(r, rcol)
                if it is None:
                    it = QTableWidgetItem('')
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(r, rcol, it)
                old = it.text()
                if rep != old:
                    it.setText(rep)
                    updated += 1
            QMessageBox.information(self,'Refresh', f'Refreshed Replacement column using OBS Parts list. Updated {updated} row(s).')
        except Exception as e:
            QMessageBox.warning(self,'Refresh Error', str(e))



    def _v2_select_above_config_block(self):
        part_col = self._v2_find_table_col_index('Part')
        wu_col = self._v2_find_table_col_contains('wu level')

        if part_col < 0 or wu_col < 0:
            return

        config_prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357',
            '0390','0395','0397','0335','0391','0431','0435','0437',
            '0440','0445','0455','0450','0441','0447','0457',
            '0460','0465','0461','0467','0410','0415','0417',
            '0411','0412','0413','0414','0360','0365','0361','0367'
        }

        rows = self.table.rowCount()
        target_rows = []
        
        # Process each WU Level 0 block
        r = 0
        while r < rows:
            wu_val = (self.table.item(r, wu_col).text()
                      if self.table.item(r, wu_col) else '').strip()

            if wu_val == '0':
                block_start = r
                block_end = rows
                # Find end of this block (next WU Level 0 row)
                for i in range(r + 1, rows):
                    nxt = (self.table.item(i, wu_col).text()
                           if self.table.item(i, wu_col) else '').strip()
                    if nxt == '0':
                        block_end = i
                        break

                # For each config-prefix row in the block, select all rows
                # with greater indentation until indent decreases
                for i in range(block_start + 1, block_end):
                    it = self.table.item(i, part_col)
                    if not it:
                        continue
                    raw = it.text()
                    stripped = raw.lstrip()
                    if not stripped:
                        continue
                    indent = len(raw) - len(stripped)
                    part_prefix = stripped[:4]

                    # If this is a config row, select all deeper rows until indent drops
                    if part_prefix in config_prefixes:
                        for j in range(i + 1, block_end):
                            it_j = self.table.item(j, part_col)
                            if not it_j:
                                continue
                            raw_j = it_j.text()
                            stripped_j = raw_j.lstrip()
                            if not stripped_j:
                                continue
                            indent_j = len(raw_j) - len(stripped_j)
                            
                            # Stop if indentation drops to config level or below
                            if indent_j <= indent:
                                break
                            
                            # Select this row (avoid duplicates)
                            if j not in target_rows:
                                target_rows.append(j)

                r = block_end
            else:
                r += 1

        # Toggle: if any target row is unchecked → select all; else deselect all
        any_unchecked = False
        for i in target_rows:
            w = self.table.cellWidget(i, 0)
            if w and hasattr(w, '_chk') and not w._chk.isChecked():
                any_unchecked = True
                break

        for i in target_rows:
            w = self.table.cellWidget(i, 0)
            if w and hasattr(w, '_chk'):
                w._chk.setChecked(any_unchecked)

class StructureSheetTab(QWidget):
    MAX_LINES = 500
    PART_LEN = 10
    INSERTED_COL_TO_SELECTOR = {
        'new description': 2,
        'new item seq': 3,
        'new qty': 4,
        'new kit code': 5,
        'new ref designator': 6,
        'new reference designator': 6,
        'new ref designator(rd)': 6,
    }

    DBKEY_BY_HEADER = {
        'part': 'part',
        'rev/ln': 'rev_ln',
        'plant': 'plant',
        'part description': 'description',
        'description': 'description',
        'item status': 'item_status',
        'base qty': 'base_qty',
        'ext qty': 'ext_qty',
        'uom': 'uom',
        'eco number': 'eco_number',
        'procurement type': 'procurement_type',
        'effectivity date': 'effectivity_date',
        'user item type': 'user_item_type',
        'item seq': 'item_seq',
        'kit code': 'kit_code',
        'sparable flag': 'sparable_flag',
        'pace': 'pace_or_dash',
        'mlo class': 'mlo_class',
    }

    def __init__(self):
        super().__init__()

        outer = QVBoxLayout(self)
        outer.setContentsMargins(8, 8, 8, 8)
        outer.setSpacing(8)
        title = QLabel("Structure Sheet")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        # ---- Structure Sheet Table ----
        structure_tab = QWidget()
        struct_layout = QVBoxLayout(structure_tab)
        struct_layout.setContentsMargins(0, 0, 0, 0)
        struct_layout.setSpacing(8)

        # Action buttons row
        btn_row = QHBoxLayout()
        self.btn_delete_row = QPushButton("Delete Selected Row(s)")
        self.btn_add_row = QPushButton("Add Row(s)")
        self.btn_build_sheet = QPushButton("Build Structure Sheet")
        self.lbl_build_plant = QLabel("Plant:")
        self.cmb_build_plant = QComboBox()
        self.cmb_build_plant.addItems(['4020', '4055', '4060', '4070', '4080', '4090'])
        self.cmb_build_plant.setCurrentText('4070')
        self.btn_insert_chk_cols = QPushButton("Insert/Remove Checkbox Columns")
        self.btn_update_part_info = QPushButton("Update Part Information")
        self.btn_export = QPushButton("Export Excel")
        self.btn_import = QPushButton("Import Excel")
        self.btn_reset_row = QPushButton("Reset")
        btn_row.addWidget(self.btn_add_row)
        btn_row.addWidget(self.btn_delete_row)
        btn_row.addWidget(self.btn_build_sheet)
        btn_row.addWidget(self.lbl_build_plant)
        btn_row.addWidget(self.cmb_build_plant)
        btn_row.addWidget(self.btn_insert_chk_cols)
        btn_row.addWidget(self.btn_update_part_info)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_export)
        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(self.btn_reset_row)
        struct_layout.addLayout(btn_row)

        # Use QSplitter to position input on left and table on right
        self.structure_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Part input widget for building structure sheet from part numbers (LEFT side)
        self.part_input_container = QWidget()
        self.part_input_container.setMinimumWidth(320)
        self.part_input_container.setMaximumWidth(360)
        part_input_layout = QVBoxLayout(self.part_input_container)
        part_input_layout.setContentsMargins(5, 5, 5, 5)
        part_input_layout.setSpacing(6)
        part_input_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        part_input_label = QLabel("Enter Part Numbers (Max 500, one per line or comma-separated):")
        part_input_label.setWordWrap(True)
        part_input_layout.addWidget(part_input_label)
        self.part_input_text = QTextEdit()
        self.part_input_text.setPlaceholderText("Enter option/part numbers...")
        self.part_input_text.setFixedHeight(300)
        part_input_layout.addWidget(self.part_input_text)
        self.structure_splitter.addWidget(self.part_input_container)
        
        # Table widget (RIGHT side)
        self.table = QTableWidget(0, 0)
        self._struct_item_change_guard = False
        self._struct_action_prompt_guard = False
        self._change_type_body_box = None
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
            QHeaderView::section {
                background:#E1F0FF; font-weight:600;
                border:1px solid #C9E2FF; padding:4px;
            }
            QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        self.structure_splitter.addWidget(self.table)
        self.structure_splitter.setCollapsible(0, True)
        self.structure_splitter.setSizes([330, 900])
        self.structure_splitter.setStretchFactor(0, 0)
        self.structure_splitter.setStretchFactor(1, 1)
        struct_layout.addWidget(self.structure_splitter, 1)

        self.btn_delete_row.clicked.connect(self._delete_selected_struct_rows)
        self.btn_add_row.clicked.connect(self._add_struct_row)
        self.btn_build_sheet.clicked.connect(self._on_build_structure_sheet_clicked)
        self.btn_insert_chk_cols.clicked.connect(self._insert_checkbox_columns)
        self.btn_update_part_info.clicked.connect(self._update_part_information)
        self.btn_export.clicked.connect(self._export_structure_sheet)
        self.btn_import.clicked.connect(self._import_structure_sheet)
        self.btn_reset_row.clicked.connect(self._reset_struct_table)

        # Keep blue body box synced with scroll/resize/content changes.
        self.table.horizontalScrollBar().valueChanged.connect(lambda _v: self._update_change_type_body_box())
        self.table.verticalScrollBar().valueChanged.connect(lambda _v: self._update_change_type_body_box())
        self.table.horizontalHeader().sectionResized.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().rowsInserted.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().rowsRemoved.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().columnsInserted.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().columnsRemoved.connect(lambda *_args: self._update_change_type_body_box())
        self.table.itemChanged.connect(self._on_struct_item_changed)

        # Keep reset color consistent with other reset buttons in the app.
        self.btn_reset_row.setStyleSheet(
            "QPushButton{background:#90A4AE;color:white;padding:6px 12px;border-radius:6px;}"
            "QPushButton:hover{background:#7C919B;}"
        )

        self._refresh_structure_action_buttons()

        outer.addWidget(structure_tab, 1)

    def _refresh_structure_action_buttons(self):
        """Show Build Sheet input only when sheet is empty; reveal others when data exists."""
        has_data = self.table.rowCount() > 0
        gated = (
            self.btn_add_row,
            self.btn_delete_row,
            self.btn_insert_chk_cols,
            self.btn_update_part_info,
            self.btn_export,
            self.btn_import,
            self.btn_reset_row,
        )
        for b in gated:
            b.setVisible(has_data)
            b.setEnabled(has_data)

        # Show part input and Build button only when sheet is empty
        self.part_input_container.setVisible(not has_data)
        self.btn_build_sheet.setVisible(not has_data)
        self.btn_build_sheet.setEnabled(not has_data)
        self.lbl_build_plant.setVisible(not has_data)
        self.cmb_build_plant.setVisible(not has_data)
        self.cmb_build_plant.setEnabled(not has_data)
        if hasattr(self, 'structure_splitter'):
            if has_data:
                self.table.setVisible(True)
                self.structure_splitter.setSizes([0, 2000])
            else:
                self.table.setVisible(False)
                self.structure_splitter.setSizes([2000, 0])

    def _on_build_structure_sheet_clicked(self):
        """Handler for Build Structure Sheet button: parse part numbers and fetch BOM."""
        try:
            text = self.part_input_text.toPlainText().strip()
            if not text:
                QMessageBox.information(self, 'Build Structure Sheet', 'Please enter at least one part number.')
                return

            # Parse part numbers: split by newline and/or comma
            parts = []
            for line in text.split('\n'):
                for item in line.split(','):
                    item = item.strip()
                    if item:
                        parts.append(item)

            # Validate max 500 parts
            if len(parts) > 500:
                QMessageBox.warning(
                    self,
                    'Build Structure Sheet',
                    f'Too many part numbers. Maximum is 500, you entered {len(parts)}.'
                )
                return

            if not parts:
                QMessageBox.information(self, 'Build Structure Sheet', 'No valid part numbers found.')
                return

            # Build the structure sheet
            self._build_structure_sheet_from_parts(parts)
            
            # Clear input and refresh buttons
            self.part_input_text.clear()
            self._refresh_structure_action_buttons()

        except Exception as e:
            QMessageBox.warning(self, 'Build Structure Sheet Error', f'{str(e)}')

    def _validate_parts(self):
        pass

    def get_impacted_parts(self):
        return []

    # ---- helper: populate one new empty row at index `row_idx` ----
    def _init_struct_row(self, row_idx: int):
        # For newly added rows, include Add/Replace/Remove/Change (no Revised).
        change_type_options = ['', 'Repl Item at Same Seq', 'Remove Item', 'Add Item', 'Change']
        part_col = self._struct_part_col()
        for c in range(self.table.columnCount()):
            if c == 0:  # Select checkbox
                cont = QWidget()
                h_lay = QHBoxLayout(cont)
                h_lay.setContentsMargins(0, 0, 0, 0)
                h_lay.addStretch(1)
                chk = QCheckBox()
                h_lay.addWidget(chk)
                h_lay.addStretch(1)
                cont._chk = chk
                self.table.setCellWidget(row_idx, c, cont)
            elif c == 1:  # Change Type dropdown
                combo = QComboBox()
                combo.addItems(change_type_options)
                combo.setCurrentIndex(0)
                combo.setStyleSheet('QComboBox { padding: 2px; }')
                combo.currentTextChanged.connect(lambda text, r=row_idx: self._on_change_type_changed(r))
                self.table.setCellWidget(row_idx, c, combo)
            elif c in {2, 3, 4, 5, 6}:  # Vertical-header checkbox columns
                cont_v = QWidget()
                hv = QHBoxLayout(cont_v)
                hv.setContentsMargins(0, 0, 0, 0)
                hv.addStretch(1)
                chk_v = QCheckBox()
                hv.addWidget(chk_v)
                hv.addStretch(1)
                cont_v._chk = chk_v
                chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                self.table.setCellWidget(row_idx, c, cont_v)
            else:
                item = QTableWidgetItem('')
                h = self.table.horizontalHeaderItem(c)
                if h and self._selector_col_for_inserted_header(h.text()) >= 0:
                    item.setForeground(QBrush(QColor('#FF8C00')))
                    selector_col = self._selector_col_for_inserted_header(h.text())
                    selector_widget = self.table.cellWidget(row_idx, selector_col)
                    is_checked = bool(selector_widget and hasattr(selector_widget, '_chk') and selector_widget._chk.isChecked())
                    self._set_item_editable(item, is_checked)
                if c == part_col:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    item.setData(Qt.ItemDataRole.UserRole, 'manual-added')
                self.table.setItem(row_idx, c, item)

        self._apply_manual_added_row_style(row_idx)
        self._on_change_type_changed(row_idx)

    def _norm_header(self, v: str) -> str:
        return (v or '').strip().lower()

    def _find_col_by_header(self, header: str) -> int:
        key = self._norm_header(header)
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h and self._norm_header(h.text()) == key:
                return c
        return -1

    def _current_change_type(self, row_idx: int) -> str:
        combo = self.table.cellWidget(row_idx, 1)
        if combo and isinstance(combo, QComboBox):
            return (combo.currentText() or '').strip()
        it = self.table.item(row_idx, 1)
        return (it.text() if it else '').strip()

    def _is_action_selected(self, row_idx: int) -> bool:
        return bool(self._current_change_type(row_idx))

    def _show_select_action_first_prompt(self):
        if self._struct_action_prompt_guard:
            return
        self._struct_action_prompt_guard = True
        try:
            QMessageBox.information(self, 'Select Action', 'Please select Action first.')
        finally:
            self._struct_action_prompt_guard = False

    def _on_struct_item_changed(self, item: QTableWidgetItem):
        if self._struct_item_change_guard:
            return
        if item is None:
            return
        row_idx = item.row()
        col_idx = item.column()
        h = self.table.horizontalHeaderItem(col_idx)
        is_inserted_col = bool(h and self._selector_col_for_inserted_header(h.text()) >= 0)
        # Ignore select/action/checkbox helper columns.
        if col_idx in {0, 1, 2, 3, 4, 5, 6}:
            return
        txt = (item.text() or '').strip()
        if is_inserted_col:
            self._sync_inserted_item_visual_state(row_idx, item)
        if not txt:
            return
        if self._is_action_selected(row_idx):
            return

    def _get_bom0_parts(self) -> set[str]:
        bom_col = self._find_col_by_header('BOM Level')
        part_col = self._find_col_by_header('Part')
        parts = set()
        if bom_col < 0 or part_col < 0:
            return parts
        for r in range(self.table.rowCount()):
            b = self.table.item(r, bom_col)
            p = self.table.item(r, part_col)
            if b and p and (b.text() or '').strip() == '0':
                pt = (p.text() or '').strip().upper()
                if pt:
                    parts.add(pt)
        return parts

    def _affected_parent_bom0_for_row(self, row_idx: int) -> str:
        """Return nearest BOM level-0 parent part above the row, if any."""
        bom_col = self._find_col_by_header('BOM Level')
        part_col = self._find_col_by_header('Part')
        if bom_col < 0 or part_col < 0:
            return ''
        for r in range(row_idx, -1, -1):
            b = self.table.item(r, bom_col)
            p = self.table.item(r, part_col)
            btxt = (b.text() if b else '').strip()
            ptxt = ((p.text() if p else '') or '').lstrip(' \t').strip().upper()
            if btxt == '0' and ptxt:
                return ptxt
        return ''

    def _is_entered_part_row(self, row_idx: int) -> bool:
        """Treat orange Part cell rows as user-entered rows for update validation."""
        part_col = self._find_col_by_header('Part')
        if part_col < 0:
            return False
        it = self.table.item(row_idx, part_col)
        if it is None:
            return False
        bg = it.background().color()
        return bg.name().lower() == '#ffe5cc'

    def _looks_like_part(self, part: str) -> bool:
        p = (part or '').strip().upper()
        if not p:
            return False
        if p.startswith('ESW') and len(p) > 10:
            return True
        return len(p) == 10 and p[4] == '-'

    def _set_part_cell_style(self, row_idx: int, editable: bool, orange: bool):
        part_col = self._find_col_by_header('Part')
        if part_col < 0:
            return
        it = self.table.item(row_idx, part_col)
        if it is None:
            it = QTableWidgetItem('')
            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, part_col, it)

        self._set_item_editable(it, editable)
        if orange:
            it.setBackground(QColor('#FFE5CC'))
        else:
            bom_col = self._find_col_by_header('BOM Level')
            is_bom0 = False
            if bom_col >= 0:
                bom_it = self.table.item(row_idx, bom_col)
                is_bom0 = bool(bom_it and (bom_it.text() or '').strip() == '0')
            it.setBackground(QColor('#87CEEB') if is_bom0 else QColor('#FFFFFF'))

    def _ensure_change_type_body_box(self):
        if self._change_type_body_box is not None:
            return
        box = QFrame(self.table.viewport())
        box.setObjectName('changeTypeBodyBox')
        box.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        box.setStyleSheet('QFrame#changeTypeBodyBox { border: 2px solid #1565C0; background: transparent; }')
        box.hide()
        self._change_type_body_box = box

    def _update_change_type_body_box(self):
        self._ensure_change_type_body_box()
        box = self._change_type_body_box
        if box is None:
            return
        if self.table.columnCount() <= 6 or self.table.rowCount() == 0:
            box.hide()
            return

        left_col = 2
        right_col = 6
        if right_col >= self.table.columnCount():
            box.hide()
            return

        left = self.table.columnViewportPosition(left_col)
        right = self.table.columnViewportPosition(right_col) + self.table.columnWidth(right_col) - 1
        top = self.table.rowViewportPosition(0)
        last_row = self.table.rowCount() - 1
        bottom = self.table.rowViewportPosition(last_row) + self.table.rowHeight(last_row) - 1

        if right < left or bottom < top:
            box.hide()
            return

        box.setGeometry(left, top, right - left + 1, bottom - top + 1)
        box.show()
        box.raise_()

    def _populate_row_from_where_used_record(self, row_idx: int, rec: dict,
                                              allowed_headers: set[str] | None = None,
                                              only_blank: bool = False):
        # Determine if this row is a BOM L0 row so new items get sky-blue background
        bom_col = self._find_col_by_header('BOM Level')
        is_bom0 = False
        if bom_col >= 0:
            bom_it = self.table.item(row_idx, bom_col)
            if bom_it and (bom_it.text() or '').strip() == '0':
                is_bom0 = True
        sky = QColor('#87CEEB')

        self._struct_item_change_guard = True
        try:
            for c in range(self.table.columnCount()):
                if c in {2, 3, 4, 5, 6}:  # rotated checkbox columns
                    continue
                h = self.table.horizontalHeaderItem(c)
                if not h:
                    continue
                h_norm = self._norm_header(h.text())
                if allowed_headers is not None and h_norm not in allowed_headers:
                    continue
                key = self.DBKEY_BY_HEADER.get(self._norm_header(h.text()))
                if not key:
                    continue
                value = '' if rec.get(key) is None else str(rec.get(key)).strip()
                if key == 'description' and h_norm in {'part description'}:
                    # Keep checkbox helper column untouched; actual part description
                    # should go to the regular Description column.
                    continue
                item = self.table.item(row_idx, c)
                if item is None:
                    item = QTableWidgetItem('')
                    if is_bom0:
                        item.setBackground(sky)
                    self.table.setItem(row_idx, c, item)

                # Determine alignment for this column
                if h_norm == 'part':
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                elif h_norm in {'description', 'part description'}:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)

                # only_blank: skip cells that already have a value
                if only_blank and (item.text() or '').strip():
                    continue

                # Preserve leading indent for Part column
                if h_norm == 'part':
                    existing = item.text()
                    leading = existing[:len(existing) - len(existing.lstrip(' \t'))]
                    value = leading + value.lstrip(' \t')

                item.setText(value)

            # Optional Ref Designator enrichment if query returns any key variant.
            ref_val = ''
            for k in ('ref_designator', 'reference_designator', 'reference designator'):
                if rec.get(k):
                    ref_val = str(rec.get(k)).strip()
                    break
            if ref_val:
                for c in range(self.table.columnCount()):
                    h = self.table.horizontalHeaderItem(c)
                    if not h:
                        continue
                    if self._norm_header(h.text()) in {'ref designator', 'reference designator',
                                                        'ref designator(rd)'}:
                        it = self.table.item(row_idx, c)
                        if it is None:
                            it = QTableWidgetItem('')
                            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            self.table.setItem(row_idx, c, it)
                        if only_blank and (it.text() or '').strip():
                            continue
                        it.setText(ref_val)
        finally:
            self._struct_item_change_guard = False

    def _validate_and_fill_part_for_row(self, row_idx: int, show_message: bool = True) -> tuple[bool, str]:
        part_col = self._find_col_by_header('Part')
        plant_col = self._find_col_by_header('Plant')
        if part_col < 0:
            return False, 'Part column not found.'

        part_item = self.table.item(row_idx, part_col)
        part = (part_item.text() if part_item else '').strip()
        if not self._looks_like_part(part):
            return False, 'Invalid part format.'

        plant = '4070'
        if plant_col >= 0:
            pit = self.table.item(row_idx, plant_col)
            if pit and (pit.text() or '').strip():
                plant = (pit.text() or '').strip()

        try:
            from where_used_query import fetch_where_used
            records = fetch_where_used([part], 1, plant=plant)
        except Exception as e:
            reason = f'Failed to query where-used for part {part}: {e}'
            if show_message:
                QMessageBox.warning(self, 'Validation Failed', reason)
            return False, reason

        action = self._current_change_type(row_idx)

        parent_l1 = {
            (r.get('part') or '').strip().upper()
            for r in records
            if (r.get('wu_level') or '').strip() == '1'
        }
        affected_parent = self._affected_parent_bom0_for_row(row_idx)

        # Replace/Remove must report to an existing level-0 parent.
        if action in {'Repl Item at Same Seq', 'Remove Item'}:
            if not affected_parent:
                full_reason = f'Part {part} cannot be validated because affected 0th Level Parent Part is not found.'
                if show_message:
                    QMessageBox.warning(self, 'Validation Failed', full_reason)
                return False, full_reason
            if affected_parent not in parent_l1:
                full_reason = f'Part {part} is not a BOM Item of {affected_parent} (0th Level Parent Part).'
                if show_message:
                    QMessageBox.warning(self, 'Validation Failed', full_reason)
                return False, full_reason

        rec_l0 = None
        for r in records:
            if (r.get('wu_level') or '').strip() == '0':
                rec_l0 = r
                break
        if rec_l0:
            if action == 'Add Item':
                self._populate_row_from_where_used_record(
                    row_idx,
                    rec_l0,
                    allowed_headers={
                        'description', 'item status', 'uom', 'procurement type',
                        'pace', 'mlo class', 'ref designator', 'reference designator',
                        'ref designator(rd)', 'part',
                    },
                )
            else:
                # Repl/Remove: only fill blank cells to preserve existing data
                self._populate_row_from_where_used_record(row_idx, rec_l0, only_blank=True)

        self._set_part_cell_style(row_idx, editable=True, orange=False)
        return True, ''

    def _update_part_information(self):
        target_actions = {'Repl Item at Same Seq', 'Remove Item', 'Add Item'}
        part_col = self._find_col_by_header('Part')
        if part_col < 0:
            QMessageBox.warning(self, 'Update Part Information', 'Part column not found.')
            return

        rows_to_check = []
        smbom_rows = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, part_col)
            # SmBOM rows are identified by UserRole == 'smbom' on the Part cell
            if it and it.data(Qt.ItemDataRole.UserRole) == 'smbom':
                part = (it.text() or '').strip()
                if part:
                    smbom_rows.append(r)
                continue
            action = self._current_change_type(r)
            if action not in target_actions:
                continue
            part = (it.text() if it else '').strip()
            # Only process rows where user entered/edited part (orange part cell)
            if part and self._is_entered_part_row(r):
                rows_to_check.append(r)

        if not rows_to_check and not smbom_rows:
            QMessageBox.information(
                self,
                'Update Part Information',
                'No entered part numbers found.'
            )
            return

        ok_count = 0
        failed = []
        for r in rows_to_check:
            ok, reason = self._validate_and_fill_part_for_row(r, show_message=False)
            if ok:
                ok_count += 1
            else:
                failed.append(reason)

        # Process SmBOM rows from bottom to top to keep row indices stable after inserts
        for r in sorted(smbom_rows, reverse=True):
            ok, reason = self._load_smbom_for_row(r)
            if ok:
                ok_count += 1
            else:
                failed.append(reason)

        if failed:
            numbered = [f'{i + 1}. {msg}' for i, msg in enumerate(failed[:8])]
            preview = '\n'.join(numbered)
            more = '' if len(failed) <= 8 else f'\n...and {len(failed) - 8} more failure(s).'
            QMessageBox.warning(
                self,
                'Update Part Information',
                f'Updated {ok_count} row(s). Failed {len(failed)} row(s).\n\n{preview}{more}'
            )
        else:
            QMessageBox.information(
                self,
                'Update Part Information',
                f'Updated part information for {ok_count} row(s).'
            )

    def _selector_col_for_inserted_header(self, header_text: str) -> int:
        return self.INSERTED_COL_TO_SELECTOR.get((header_text or '').strip().lower(), -1)

    def _default_struct_cell_bg(self, row_idx: int) -> QColor:
        if self._is_manual_added_row(row_idx):
            return QColor('#FFF2E0')
        bom_col = self._find_col_by_header('BOM Level')
        if bom_col >= 0:
            bom_it = self.table.item(row_idx, bom_col)
            if bom_it and (bom_it.text() or '').strip() == '0':
                return QColor('#87CEEB')
        return QColor('#FFFFFF')

    def _is_manual_added_row(self, row_idx: int) -> bool:
        part_col = self._struct_part_col()
        if part_col < 0:
            return False
        it = self.table.item(row_idx, part_col)
        return bool(it and it.data(Qt.ItemDataRole.UserRole) == 'manual-added')

    def _apply_manual_added_row_style(self, row_idx: int):
        if not self._is_manual_added_row(row_idx):
            return

        row_bg = '#FFF2E0'
        combo_bg = '#FFF2E0'
        widget_bg = 'QWidget { background-color: #FFF2E0; }'
        combo_style = 'QComboBox { padding: 2px; background-color: #FFF2E0; }'

        for c in range(self.table.columnCount()):
            if c == 0:
                w = self.table.cellWidget(row_idx, c)
                if w:
                    w.setStyleSheet(widget_bg)
            elif c == 1:
                w = self.table.cellWidget(row_idx, c)
                if w and isinstance(w, QComboBox):
                    w.setStyleSheet(combo_style)
            elif c in {2, 3, 4, 5, 6}:
                w = self.table.cellWidget(row_idx, c)
                if w:
                    w.setStyleSheet(widget_bg)
                else:
                    it = self.table.item(row_idx, c)
                    if it is not None:
                        it.setBackground(QColor(row_bg))
            else:
                it = self.table.item(row_idx, c)
                if it is not None and c != self._struct_part_col():
                    it.setBackground(QColor(row_bg))

    def _sync_inserted_item_visual_state(self, row_idx: int, item: QTableWidgetItem):
        if item is None:
            return
        header_item = self.table.horizontalHeaderItem(item.column())
        selector_col = self._selector_col_for_inserted_header(header_item.text() if header_item else '')
        if selector_col < 0:
            return

        item.setForeground(QBrush(QColor('#FF8C00')))
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        w = self.table.cellWidget(row_idx, selector_col)
        is_checked = bool(w and hasattr(w, '_chk') and w._chk.isChecked())
        has_text = bool((item.text() or '').strip())
        if is_checked and not has_text:
            item.setBackground(QColor('#FFE5CC'))
        else:
            item.setBackground(self._default_struct_cell_bg(row_idx))

    def _apply_inserted_header_colors(self):
        base = QBrush(QColor('#0F2D46'))
        orange = QBrush(QColor('#FF8C00'))
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h is None:
                continue
            if self._selector_col_for_inserted_header(h.text()) >= 0:
                h.setForeground(orange)
            else:
                h.setForeground(base)

    def _ensure_change_type_group_header(self):
        if self.table.columnCount() == 0:
            return
        hh = self.table.horizontalHeader()
        if isinstance(hh, RotatedColumnsHeader):
            hh.set_rotated_columns(range(2, 7))
            hh._group_label = 'Change Type'
            hh._group_columns = list(range(2, 7))
            hh.setMinimumHeight(117)
            hh.viewport().update()
            return

        headers = []
        for c in range(self.table.columnCount()):
            hi = self.table.horizontalHeaderItem(c)
            headers.append(hi.text() if hi else f'Col{c}')
        hdr = RotatedColumnsHeader(
            Qt.Orientation.Horizontal,
            rotated_columns=range(2, 7),
            parent=self.table,
            group_label='Change Type',
            group_columns=list(range(2, 7)),
        )
        hdr.setStretchLastSection(True)
        hdr.setSectionsClickable(True)
        self.table.setHorizontalHeader(hdr)
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.horizontalHeader().setMinimumHeight(117)
        self._apply_inserted_header_colors()

    def _set_item_editable(self, item: QTableWidgetItem, editable: bool):
        if item is None:
            return
        flags = item.flags()
        if editable:
            item.setFlags(flags | Qt.ItemFlag.ItemIsEditable)
        else:
            item.setFlags(flags & ~Qt.ItemFlag.ItemIsEditable)

    def _ensure_structure_sheet_template_headers(self):
        """Initialize default Structure Sheet headers when table is empty."""
        if self.table.columnCount() > 0:
            return

        headers = self._structure_sheet_headers()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        self._ensure_change_type_group_header()
        self._restore_rotated_column_widths()

        # Baseline widths/alignment similar to loaded template behavior
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.resizeColumnsToContents()
        self._update_change_type_body_box()

    def _find_header_col(self, header_name: str) -> int:
        key = (header_name or '').strip().lower()
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h and (h.text() or '').strip().lower() == key:
                return c
        return -1

    def _structure_sheet_headers(self) -> list[str]:
        return [
            'Select',
            'Action',
            'Part Description',
            'Seq#',
            'Qty',
            'Kit Code',
            'Ref Designator(RD)',
            'BOM Level',
            'Part',
            'Replacement',
            'Description',
            'Item Status',
            'Base Qty',
            'Ext Qty',
            'UOM',
            'Procurement Type',
            'Item Seq',
            'Kit Code',
            'Sparable flag',
            'Pace',
            'Ref Designator(RD)',
            'MLO Class',
        ]

    def _collect_incomplete_actions(self) -> dict:
        selectors = {
            2: 'New Description',
            3: 'New Item Seq',
            4: 'New Qty',
            5: 'New Kit Code',
            6: 'New Ref Designator',
        }
        checked_total = 0
        missing_col = 0
        blank_value = 0

        for r in range(self.table.rowCount()):
            for sel_col, new_hdr in selectors.items():
                w = self.table.cellWidget(r, sel_col)
                if not (w and hasattr(w, '_chk') and w._chk.isChecked()):
                    continue
                checked_total += 1

                new_col = self._find_header_col(new_hdr)
                if new_col < 0 and new_hdr == 'New Ref Designator':
                    new_col = self._find_header_col('New Reference Designator')
                if new_col < 0:
                    missing_col += 1
                    continue

                it = self.table.item(r, new_col)
                if it is None or not (it.text() or '').strip():
                    blank_value += 1

        return {
            'checked_total': checked_total,
            'missing_col': missing_col,
            'blank_value': blank_value,
            'incomplete_total': missing_col + blank_value,
        }

    def _undo_incomplete_actions(self):
        selectors = {
            2: 'New Description',
            3: 'New Item Seq',
            4: 'New Qty',
            5: 'New Kit Code',
            6: 'New Ref Designator',
        }
        for r in range(self.table.rowCount()):
            for sel_col, new_hdr in selectors.items():
                w = self.table.cellWidget(r, sel_col)
                if not (w and hasattr(w, '_chk') and w._chk.isChecked()):
                    continue

                new_col = self._find_header_col(new_hdr)
                if new_col < 0 and new_hdr == 'New Ref Designator':
                    new_col = self._find_header_col('New Reference Designator')

                is_incomplete = False
                if new_col < 0:
                    is_incomplete = True
                else:
                    it = self.table.item(r, new_col)
                    if it is None or not (it.text() or '').strip():
                        is_incomplete = True

                if is_incomplete:
                    w._chk.setChecked(False)

    def confirm_leave_with_incomplete(self) -> bool:
        s = self._collect_incomplete_actions()
        if s['incomplete_total'] <= 0:
            return True

        msg = (
            'Incomplete actions found in Structure sheet:\n\n'
            f"Checked selectors: {s['checked_total']}\n"
            f"Missing inserted columns: {s['missing_col']}\n"
            f"Blank values in inserted columns: {s['blank_value']}\n\n"
            'Incomplete actions will not be saved.\n'
            'Yes: undo only incomplete actions and continue.\n'
            'No: stay on this tab.'
        )
        ans = QMessageBox.question(
            self,
            'Unsaved Incomplete Actions',
            msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if ans == QMessageBox.StandardButton.Yes:
            self._undo_incomplete_actions()
            return True
        return False

    def _find_row_for_selector_checkbox(self, chk: QCheckBox) -> int:
        for r in range(self.table.rowCount()):
            for c in (2, 3, 4, 5, 6):
                w = self.table.cellWidget(r, c)
                if w and hasattr(w, '_chk') and w._chk is chk:
                    return r
        return -1

    def _update_inserted_editability_for_row(self, row_idx: int):
        if row_idx < 0 or row_idx >= self.table.rowCount():
            return
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if not h:
                continue
            selector_col = self._selector_col_for_inserted_header(h.text())
            if selector_col < 0:
                continue
            item = self.table.item(row_idx, c)
            if item is None:
                item = QTableWidgetItem('')
                item.setForeground(QBrush(QColor('#FF8C00')))
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row_idx, c, item)
            w = self.table.cellWidget(row_idx, selector_col)
            is_checked = bool(w and hasattr(w, '_chk') and w._chk.isChecked())
            if not is_checked and item.text():
                item.setText('')
            self._set_item_editable(item, is_checked)
            self._sync_inserted_item_visual_state(row_idx, item)

    def _update_inserted_editability_all_rows(self):
        for r in range(self.table.rowCount()):
            self._update_inserted_editability_for_row(r)

    def _on_selector_checkbox_toggled(self, _state: int):
        chk = self.sender()
        if not isinstance(chk, QCheckBox):
            return
        row_idx = self._find_row_for_selector_checkbox(chk)
        if row_idx >= 0:
            if chk.isChecked() and not self._is_action_selected(row_idx):
                self._show_select_action_first_prompt()
                chk.blockSignals(True)
                chk.setChecked(False)
                chk.blockSignals(False)
                return
            self._update_inserted_editability_for_row(row_idx)

    def _restore_rotated_column_widths(self):
        self._ensure_change_type_group_header()
        hh = self.table.horizontalHeader()
        for vc in range(2, 7):
            if vc < self.table.columnCount():
                hh.setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                self.table.setColumnWidth(vc, 40)
        self._update_change_type_body_box()

    def _struct_part_col(self) -> int:
        """Return the index of the Part column, or -1 if not found."""
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h and h.text().strip().lower() == 'part':
                return c
        return -1

    def _on_change_type_changed(self, row_idx: int):
        """Handle Change Type dropdown value changes to show/hide checkboxes."""
        combo = self.table.cellWidget(row_idx, 1)
        if not combo or not isinstance(combo, QComboBox):
            return

        change_type = combo.currentText()
        show_checkboxes = bool(change_type) and (change_type not in ['Remove Item', 'Add Item'])

        self._struct_item_change_guard = True
        try:
            for col in {2, 3, 4, 5}:
                w = self.table.cellWidget(row_idx, col)
                if w and hasattr(w, '_chk'):
                    if show_checkboxes:
                        w.show()
                    else:
                        w._chk.setChecked(False)
                        w.hide()

            w_ref = self.table.cellWidget(row_idx, 6)
            if w_ref and hasattr(w_ref, '_chk'):
                if show_checkboxes:
                    w_ref.show()
                else:
                    w_ref._chk.setChecked(False)
                    w_ref.hide()

            action = self._current_change_type(row_idx)
            if action in {'Repl Item at Same Seq', 'Remove Item', 'Add Item', 'Revised', 'Change'}:
                self._set_part_cell_style(row_idx, editable=True, orange=True)
            else:
                self._set_part_cell_style(row_idx, editable=False, orange=False)

            self._update_inserted_editability_for_row(row_idx)
            self._apply_manual_added_row_style(row_idx)
        finally:
            self._struct_item_change_guard = False

    def _delete_selected_struct_rows(self):
        """Delete rows where the Select checkbox (column 0) is checked."""
        rows_to_delete = []
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk') and w._chk.isChecked():
                rows_to_delete.append(r)
        if not rows_to_delete:
            QMessageBox.information(self, 'Delete Rows', 'No rows selected. Check the Select checkbox on the row(s) to delete.')
            return
        for r in reversed(rows_to_delete):
            self.table.removeRow(r)
        self._update_change_type_body_box()
        self._refresh_structure_action_buttons()
        QMessageBox.information(self, 'Delete Rows', f'Deleted {len(rows_to_delete)} row(s).')

    def _add_struct_row(self):
        """Insert a new empty row immediately below the last selected row, or append."""
        insert_after = -1
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk') and w._chk.isChecked():
                insert_after = r

        insert_at = insert_after + 1 if insert_after >= 0 else self.table.rowCount()
        self.table.insertRow(insert_at)
        self._init_struct_row(insert_at)
        self._update_change_type_body_box()
        self._refresh_structure_action_buttons()

    def _build_structure_sheet_from_parts(self, part_list: list):
        """Fetch level-0/1 Implemented BOM and render with the Structure Sheet template."""
        try:
            from implemented_bom_query import fetch_implemented_bom  # type: ignore[import]

            plant = (self.cmb_build_plant.currentText() or '').strip() if hasattr(self, 'cmb_build_plant') else ''
            if not plant:
                plant = getattr(self, '_current_plant', '4070')

            requested_parts = [str(p).strip().upper() for p in part_list if str(p).strip()]
            if not requested_parts:
                QMessageBox.information(self, 'Build Structure Sheet', 'No valid part numbers were provided.')
                return

            records = fetch_implemented_bom(requested_parts, max_level=1, plant=plant, include_level0=True)
            if not records:
                QMessageBox.information(
                    self,
                    'Build Structure Sheet',
                    f'No implemented BOM data found for the provided part numbers in plant {plant}.',
                )
                return

            def _norm_header(h: str) -> str:
                return ' '.join((h or '').strip().lower().split())

            ordered_roots = []
            groups = {}
            for root in requested_parts:
                if root not in groups:
                    ordered_roots.append(root)
                    groups[root] = {'root': None, 'children': [], 'seen': set()}

            for rec in records:
                root = str(rec.get('input_part', rec.get('part', '')) or '').strip().upper()
                if not root:
                    continue
                if root not in groups:
                    ordered_roots.append(root)
                    groups[root] = {'root': None, 'children': [], 'seen': set()}

                level_txt = str(rec.get('bom_level', '-1') or '-1').strip()
                try:
                    lvl = int(float(level_txt))
                except Exception:
                    lvl = -1

                if lvl == 0:
                    groups[root]['root'] = rec
                elif lvl == 1:
                    cpart = str(rec.get('part', '') or '').strip().upper()
                    if cpart and cpart not in groups[root]['seen']:
                        groups[root]['children'].append(rec)
                        groups[root]['seen'].add(cpart)

            final_headers = self._structure_sheet_headers()
            dbkey_by_header = {
                'bom level': 'bom_level',
                'part': 'part',
                'replacement': '',
                'rev/ln': 'rev_ln',
                'plant': 'plant',
                'description': 'description',
                'item status': 'item_status',
                'base qty': 'base_qty',
                'ext qty': 'ext_qty',
                'uom': 'uom',
                'eco number': 'eco_number',
                'procurement type': 'procurement_type',
                'effectivity date': 'effectivity_date',
                'user item type': 'user_item_type',
                'item seq': 'item_seq',
                'kit code': 'kit_code',
                'sparable flag': 'sparable_flag',
                'pace': 'pace_or_dash',
                'mlo class': 'mlo_class',
            }

            out_rows = []
            for root in ordered_roots:
                grp = groups.get(root, {'root': None, 'children': []})
                if grp.get('root') is not None:
                    out_rows.append(grp['root'])
                out_rows.extend(grp.get('children', []))

            if not out_rows:
                self._refresh_structure_action_buttons()
                QMessageBox.information(
                    self,
                    'Build Structure Sheet',
                    'No BOM Level 0/1 rows found for the provided part numbers.',
                )
                return

            final_rows = []
            for rec in out_rows:
                row = []
                for idx, h in enumerate(final_headers):
                    hk = _norm_header(h)
                    if hk in {'select', 'action'} or 2 <= idx <= 6:
                        val = ''
                    else:
                        dbk = dbkey_by_header.get(hk, '')
                        val = '' if not dbk else rec.get(dbk, '')
                    row.append('' if val is None else str(val))
                final_rows.append(row)

            obs_replacements = {}
            obs_tab = getattr(self.window(), 'obs_tab', None)
            if obs_tab and hasattr(obs_tab, 'table'):
                obs_table = obs_tab.table
                obs_part_col = -1
                obs_repl_col = -1
                for c in range(obs_table.columnCount()):
                    h = obs_table.horizontalHeaderItem(c)
                    if h:
                        nh = _norm_header(h.text())
                        if nh == 'part':
                            obs_part_col = c
                        elif nh == 'replacement':
                            obs_repl_col = c
                if obs_part_col >= 0 and obs_repl_col >= 0:
                    for rr in range(obs_table.rowCount()):
                        part_item = obs_table.item(rr, obs_part_col)
                        repl_item = obs_table.item(rr, obs_repl_col)
                        if part_item and repl_item:
                            part_val = (part_item.text() or '').strip().upper()
                            repl_val = (repl_item.text() or '').strip()
                            if part_val and repl_val:
                                obs_replacements[part_val] = repl_val

            if obs_replacements:
                ans = QMessageBox.question(
                    self,
                    'OBS Replacement Parts',
                    'Please confirm if the Replacement Parts to be mapped from OBS List.',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                if ans == QMessageBox.StandardButton.Yes:
                    repl_col_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'replacement'), -1)
                    part_col_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'part'), -1)
                    if repl_col_idx >= 0 and part_col_idx >= 0:
                        for row in final_rows:
                            if part_col_idx < len(row):
                                part_val = (row[part_col_idx] or '').strip().upper()
                                if part_val in obs_replacements:
                                    row[repl_col_idx] = obs_replacements[part_val]

            t = self.table
            _bulk_guard = hasattr(self, '_struct_item_change_guard')
            if _bulk_guard:
                self._struct_item_change_guard = True
            try:
                t.clear()
                t.setRowCount(0)
                t.setColumnCount(len(final_headers))
                hdr = RotatedColumnsHeader(
                    Qt.Orientation.Horizontal,
                    rotated_columns=range(2, 7),
                    parent=t,
                    group_label='Change Type',
                    group_columns=list(range(2, 7)),
                )
                hdr.setStretchLastSection(True)
                hdr.setSectionsClickable(True)
                t.setHorizontalHeader(hdr)
                t.setHorizontalHeaderLabels(final_headers)
                t.setRowCount(len(final_rows))

                t.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
                t.horizontalHeader().setMinimumHeight(117)
                for vc in range(2, 7):
                    t.horizontalHeader().setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                    t.setColumnWidth(vc, 40)

                bom_col = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'bom level'), -1)
                part_col = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'part'), -1)
                vertical_checkbox_cols = {2, 3, 4, 5, 6}
                change_type_options_l0 = ['', 'Revised', 'Change']
                change_type_options_l1 = ['', 'Repl Item at Same Seq', 'Remove Item', 'Change']

                for r, row in enumerate(final_rows):
                    for c, val in enumerate(row):
                        if c == 0:
                            cont = QWidget()
                            h_lay = QHBoxLayout(cont)
                            h_lay.setContentsMargins(0, 0, 0, 0)
                            h_lay.addStretch(1)
                            chk = QCheckBox()
                            h_lay.addWidget(chk)
                            h_lay.addStretch(1)
                            cont._chk = chk
                            t.setCellWidget(r, c, cont)
                        elif c == 1:
                            bom_val = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            combo = QComboBox()
                            combo.addItems(change_type_options_l0 if bom_val == '0' else change_type_options_l1)
                            combo.setCurrentIndex(0)
                            combo.setStyleSheet('QComboBox { padding: 2px; }')
                            combo.currentTextChanged.connect(lambda _text, rr=r: self._on_change_type_changed(rr))
                            t.setCellWidget(r, c, combo)
                        elif c in vertical_checkbox_cols:
                            bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            if bom_here == '0' and c not in {2, 6}:
                                t.setItem(r, c, QTableWidgetItem(''))
                            else:
                                cont_v = QWidget()
                                hv = QHBoxLayout(cont_v)
                                hv.setContentsMargins(0, 0, 0, 0)
                                hv.addStretch(1)
                                chk_v = QCheckBox()
                                hv.addWidget(chk_v)
                                hv.addStretch(1)
                                cont_v._chk = chk_v
                                chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                                t.setCellWidget(r, c, cont_v)
                        else:
                            display_val = val
                            if c == part_col:
                                part_txt = ((val or '') if isinstance(val, str) else str(val or '')).lstrip(' \t')
                                bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                                display_val = part_txt if bom_here == '0' else ('      ' + part_txt if part_txt else '')
                            item = QTableWidgetItem(display_val)
                            if c == part_col or _norm_header(final_headers[c]) == 'description':
                                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            else:
                                item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                            t.setItem(r, c, item)

                    if bom_col >= 0 and bom_col < len(row) and str(row[bom_col]).strip() == '0':
                        for c in range(t.columnCount()):
                            if c == 0:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                            elif c == 1:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                            elif c in vertical_checkbox_cols:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                                else:
                                    it = t.item(r, c)
                                    if it:
                                        it.setBackground(QColor('#87CEEB'))
                            else:
                                it = t.item(r, c)
                                if it:
                                    it.setBackground(QColor('#87CEEB'))

                    self._on_change_type_changed(r)

                t.resizeColumnsToContents()
                for vc in range(2, 7):
                    t.setColumnWidth(vc, 40)
                self._update_change_type_body_box()
                self._refresh_structure_action_buttons()
            finally:
                if _bulk_guard:
                    self._struct_item_change_guard = False

            root_count = sum(1 for r in out_rows if str(r.get('bom_level', '')).strip() == '0')
            child_count = sum(1 for r in out_rows if str(r.get('bom_level', '')).strip() == '1')
            QMessageBox.information(
                self,
                'Build Structure Sheet',
                f'Loaded Implemented BOM successfully. Root rows: {root_count}, Level 1 rows: {child_count}.',
            )
        except Exception as e:
            QMessageBox.warning(self, 'Build Structure Sheet Error', str(e))

    # ---- SmBOM row support ----
    def _add_smbom_row(self):
        """Append a new sky-blue SmBOM row at the bottom with Part cell orange for input."""
        if self.table.columnCount() == 0:
            self._ensure_structure_sheet_template_headers()
        bom_col = self._find_col_by_header('BOM Level')
        if bom_col < 0:
            QMessageBox.warning(self, 'Add SmBOM', 'BOM Level column not found.')
            return
        row_idx = self.table.rowCount()
        self.table.insertRow(row_idx)
        self._init_smbom_row(row_idx)
        self._update_change_type_body_box()
        self._refresh_structure_action_buttons()

    def _init_smbom_row(self, row_idx: int):
        """Set up a BOM L0-style sky-blue SmBOM input row."""
        bom_col = self._find_col_by_header('BOM Level')
        part_col = self._struct_part_col()
        sky = QColor('#87CEEB')
        change_type_options_l0 = ['', 'Revised', 'Change']

        for c in range(self.table.columnCount()):
            if c == 0:  # Select checkbox
                cont = QWidget()
                h_lay = QHBoxLayout(cont)
                h_lay.setContentsMargins(0, 0, 0, 0)
                h_lay.addStretch(1)
                chk = QCheckBox()
                h_lay.addWidget(chk)
                h_lay.addStretch(1)
                cont._chk = chk
                cont.setStyleSheet('QWidget { background-color: #87CEEB; }')
                self.table.setCellWidget(row_idx, c, cont)
            elif c == 1:  # Action dropdown with BOM L0 options
                combo = QComboBox()
                combo.addItems(change_type_options_l0)
                combo.setCurrentIndex(0)
                combo.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                combo.currentTextChanged.connect(lambda text, r=row_idx: self._on_change_type_changed(r))
                self.table.setCellWidget(row_idx, c, combo)
            elif c in {2, 3, 4, 5, 6}:
                # BOM L0 rule: only Description(2) and Ref Designator(6) get checkboxes
                if c in {2, 6}:
                    cont_v = QWidget()
                    hv = QHBoxLayout(cont_v)
                    hv.setContentsMargins(0, 0, 0, 0)
                    hv.addStretch(1)
                    chk_v = QCheckBox()
                    hv.addWidget(chk_v)
                    hv.addStretch(1)
                    cont_v._chk = chk_v
                    cont_v.setStyleSheet('QWidget { background-color: #87CEEB; }')
                    chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                    self.table.setCellWidget(row_idx, c, cont_v)
                else:
                    item = QTableWidgetItem('')
                    item.setBackground(sky)
                    self.table.setItem(row_idx, c, item)
            else:
                item = QTableWidgetItem('')
                if c == part_col:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    item.setBackground(QColor('#FFE5CC'))
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setData(Qt.ItemDataRole.UserRole, 'smbom')
                elif c == bom_col:
                    item.setText('0')
                    item.setBackground(sky)
                else:
                    h = self.table.horizontalHeaderItem(c)
                    h_norm = self._norm_header(h.text()) if h else ''
                    if h_norm == 'description':
                        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                    item.setBackground(sky)
                self.table.setItem(row_idx, c, item)

            self._on_change_type_changed(row_idx)

    def _load_smbom_for_row(self, row_idx: int) -> tuple[bool, str]:
        """Fetch 1-level BOM for an SmBOM row, populate it and insert child rows."""
        part_col = self._struct_part_col()
        plant_col = self._find_col_by_header('Plant')
        bom_col = self._find_col_by_header('BOM Level')

        if part_col < 0:
            return False, 'Part column not found.'
        part_item = self.table.item(row_idx, part_col)
        part = (part_item.text() if part_item else '').strip()
        if not self._looks_like_part(part):
            return False, f'Invalid part format: {part!r}'

        # Prevent duplicate SmBOM option insertion if BOM0 part already exists.
        part_upper = part.upper()
        if bom_col >= 0:
            for r in range(self.table.rowCount()):
                if r == row_idx:
                    continue
                b_it = self.table.item(r, bom_col)
                p_it = self.table.item(r, part_col)
                b_txt = (b_it.text() if b_it else '').strip()
                p_txt = ((p_it.text() if p_it else '') or '').lstrip(' \t').strip().upper()
                if b_txt == '0' and p_txt == part_upper:
                    return False, f'SmBOM Option {part} already exists.'

        plant = '4070'
        if plant_col >= 0:
            pit = self.table.item(row_idx, plant_col)
            if pit and (pit.text() or '').strip():
                plant = (pit.text() or '').strip()

        try:
            from implemented_bom_query import fetch_implemented_bom
            records = fetch_implemented_bom([part], max_level=1, plant=plant, include_level0=True)
        except Exception as e:
            return False, f'Failed to query BOM for {part}: {e}'

        l0_recs = [r for r in records if str(r.get('bom_level', '')).strip() == '0']
        l1_recs = [r for r in records if str(r.get('bom_level', '')).strip() == '1']

        # Populate the L0 (SmBOM) row
        if l0_recs:
            self._populate_row_from_where_used_record(row_idx, l0_recs[0])
        # Clear SmBOM marker so subsequent Update won't reload
        if part_item:
            part_item.setData(Qt.ItemDataRole.UserRole, None)
            part_item.setBackground(QColor('#87CEEB'))  # restore sky-blue after load

        # Insert L1 child rows immediately below the SmBOM row
        sky = QColor('#FFFFFF')
        change_type_options_l1 = ['', 'Repl Item at Same Seq', 'Remove Item', 'Change']
        for i, l1_rec in enumerate(l1_recs):
            insert_at = row_idx + 1 + i
            self.table.insertRow(insert_at)
            self._struct_item_change_guard = True
            try:
                for c in range(self.table.columnCount()):
                    if c == 0:
                        cont = QWidget()
                        h_lay = QHBoxLayout(cont)
                        h_lay.setContentsMargins(0, 0, 0, 0)
                        h_lay.addStretch(1)
                        chk = QCheckBox()
                        h_lay.addWidget(chk)
                        h_lay.addStretch(1)
                        cont._chk = chk
                        self.table.setCellWidget(insert_at, c, cont)
                    elif c == 1:
                        combo = QComboBox()
                        combo.addItems(change_type_options_l1)
                        combo.setCurrentIndex(0)
                        combo.setStyleSheet('QComboBox { padding: 2px; }')
                        combo.currentTextChanged.connect(
                            lambda text, r=insert_at: self._on_change_type_changed(r))
                        self.table.setCellWidget(insert_at, c, combo)
                    elif c in {2, 3, 4, 5, 6}:
                        cont_v = QWidget()
                        hv = QHBoxLayout(cont_v)
                        hv.setContentsMargins(0, 0, 0, 0)
                        hv.addStretch(1)
                        chk_v = QCheckBox()
                        hv.addWidget(chk_v)
                        hv.addStretch(1)
                        cont_v._chk = chk_v
                        chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                        self.table.setCellWidget(insert_at, c, cont_v)
                    else:
                        h = self.table.horizontalHeaderItem(c)
                        h_norm = self._norm_header(h.text()) if h else ''
                        key = self.DBKEY_BY_HEADER.get(h_norm)
                        value = ''
                        if key:
                            value = '' if l1_rec.get(key) is None else str(l1_rec.get(key)).strip()
                        item = QTableWidgetItem(value)
                        if c == part_col:
                            item.setText('      ' + value.lstrip(' \t') if value.strip() else '')
                            item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                        elif c == bom_col:
                            item.setText('1')
                            item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                        elif h_norm == 'description':
                            item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                        else:
                            item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                        self.table.setItem(insert_at, c, item)
            finally:
                self._struct_item_change_guard = False

            self._on_change_type_changed(insert_at)

        return True, ''

    def _insert_checkbox_columns(self):
        """Insert/remove 'New ...' columns based on selected checkboxes in columns 2..6."""
        if self.table.columnCount() == 0 or self.table.rowCount() == 0:
            QMessageBox.information(self, 'Insert Columns', 'No Structure Sheet data available.')
            return

        # Determine which selector columns have at least one checked checkbox.
        selected_flags = {
            'description': False,
            'seq': False,
            'qty': False,
            'kit': False,
            'refdes': False,
        }
        col_key_map = {2: 'description', 3: 'seq', 4: 'qty', 5: 'kit', 6: 'refdes'}
        for r in range(self.table.rowCount()):
            for col, key in col_key_map.items():
                w = self.table.cellWidget(r, col)
                if w and hasattr(w, '_chk') and w._chk.isChecked():
                    selected_flags[key] = True

        def _norm(v: str) -> str:
            return (v or '').strip().lower()

        def _find_header_index(name: str) -> int:
            for c in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                if h and _norm(h.text()) == _norm(name):
                    return c
            return -1

        def _find_last_header(possible: set[str]) -> int:
            for c in range(self.table.columnCount() - 1, -1, -1):
                h = self.table.horizontalHeaderItem(c)
                if h and _norm(h.text()) in possible:
                    return c
            return -1

        def _find_bom_col() -> int:
            for c in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                if h and _norm(h.text()) == 'bom level':
                    return c
            return -1

        orange = QBrush(QColor('#FF8C00'))
        insert_specs = [
            ('description', 'New Description', {'description', 'part description'}),
            ('seq', 'New Item Seq', {'item seq', 'seq#', 'seq'}),
            ('qty', 'New Qty', {'ext qty', 'qty'}),
            ('kit', 'New Kit Code', {'kit code'}),
            ('refdes', 'New Ref Designator', {'reference designator', 'ref designator', 'ref designator(rd)'}),
        ]

        inserted = 0
        removed = 0
        bom_col = _find_bom_col()
        for flag_key, new_header, anchors in insert_specs:
            should_exist = selected_flags.get(flag_key, False)
            existing_idx = _find_header_index(new_header)
            if flag_key == 'refdes' and existing_idx < 0:
                existing_idx = _find_header_index('New Reference Designator')

            if should_exist and existing_idx < 0:
                anchor_col = _find_last_header(anchors)
                insert_at = (anchor_col + 1) if anchor_col >= 0 else self.table.columnCount()

                self.table.insertColumn(insert_at)
                h_item = QTableWidgetItem(new_header)
                h_item.setForeground(orange)
                self.table.setHorizontalHeaderItem(insert_at, h_item)

                for r in range(self.table.rowCount()):
                    item = QTableWidgetItem('')
                    item.setForeground(orange)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    selector_col = self.INSERTED_COL_TO_SELECTOR[new_header.lower()]
                    selector_w = self.table.cellWidget(r, selector_col)
                    is_checked = bool(selector_w and hasattr(selector_w, '_chk') and selector_w._chk.isChecked())
                    self._set_item_editable(item, is_checked)
                    if bom_col >= 0:
                        bom_item = self.table.item(r, bom_col)
                        if bom_item and bom_item.text().strip() == '0':
                            item.setBackground(QColor('#87CEEB'))
                    self.table.setItem(r, insert_at, item)

                inserted += 1

            elif (not should_exist) and existing_idx >= 0:
                self.table.removeColumn(existing_idx)
                removed += 1

        self._update_inserted_editability_all_rows()
        self._apply_inserted_header_colors()
        self._restore_rotated_column_widths()
        self._update_change_type_body_box()

        if inserted == 0 and removed == 0:
            QMessageBox.information(self, 'Insert Columns', 'No column change needed.')
        else:
            QMessageBox.information(self, 'Insert Columns',
                                    f'Inserted {inserted} column(s), removed {removed} column(s).')

    def _reset_struct_table(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self._refresh_structure_action_buttons()

    def append_rows(self, headers:list[str], rows:list[list[str]]):
        if self.table.columnCount()==0:
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
        if self.table.columnCount()!=len(headers):
            norm=[]
            for r in rows:
                r2=(r+['']*self.table.columnCount())[:self.table.columnCount()]
                norm.append(r2)
            rows=norm
        start=self.table.rowCount()
        self.table.setRowCount(start+len(rows))
        for i,row in enumerate(rows):
            for j,val in enumerate(row):
                self.table.setItem(start+i,j,QTableWidgetItem(val))
        self._refresh_structure_action_buttons()

    def _export_structure_sheet(self):
        """Export Structure Sheet to Excel with dropdowns, checkboxes, and formatting preserved."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            if self.table.rowCount() == 0:
                QMessageBox.warning(self, 'Export', 'No data to export.')
                return

            file_path, _ = QFileDialog.getSaveFileName(
                self, 'Save Structure Sheet', '', 'Excel Files (*.xlsx);;All Files (*)'
            )
            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Structure Sheet"

            # Get headers (exclude Select column 0)
            headers = []
            for c in range(1, self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                headers.append(h.text() if h else f'Col{c}')

            # Identify vertical-header columns by name (table cols 2-6, export cols 2-6)
            # Vertical rotation ONLY for the 5 leading columns at fixed positions 2-6
            VERT_EXCEL_COLS = {2, 3, 4, 5, 6}

            # Find Part and BOM Level table columns (table col c == excel col c, Select skipped)
            part_excel_col = -1
            bom_col_table = -1
            desc_excel_col = -1
            for c in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                if h:
                    if h.text().strip().lower() == 'part':
                        part_excel_col = c
                    if 'bom level' in h.text().lower():
                        bom_col_table = c
                    if h.text().strip().lower() == 'description':
                        desc_excel_col = c

            thin = Side(border_style='thin', color='C9E2FF')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_fill = PatternFill(start_color='E1F0FF', end_color='E1F0FF', fill_type='solid')
            grp_fill = PatternFill(start_color='B0D8F5', end_color='B0D8F5', fill_type='solid')
            hdr_font = Font(bold=True, color='0F2D46')

            # Row 1: group label 'Change Type' merged over vertical cols, blank for others
            ws.row_dimensions[1].height = 20
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c_idx, value='')
                cell.fill = hdr_fill
                cell.border = border
            grp_start, grp_end = min(VERT_EXCEL_COLS), max(VERT_EXCEL_COLS)
            ws.merge_cells(start_row=1, start_column=grp_start, end_row=1, end_column=grp_end)
            gc = ws.cell(row=1, column=grp_start, value='Change Type')
            gc.font = hdr_font
            gc.fill = grp_fill
            gc.alignment = Alignment(horizontal='center', vertical='center')

            # Row 2: column headers
            ws.row_dimensions[2].height = 90
            for c_idx, h_text in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=c_idx, value=h_text)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = border
                if c_idx in VERT_EXCEL_COLS:
                    cell.alignment = Alignment(text_rotation=90, horizontal='center',
                                               vertical='bottom', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center',
                                               wrap_text=True)

            col_max_len = {i: len(headers[i - 1]) for i in range(1, len(headers) + 1)}

            # Data validation: one per column (reuse)
            # Action dropdown validation (L0: Revised+Change; others: no Revised)
            dv_l0 = DataValidation(type='list', formula1='"Revised,Change"',
                                   allow_blank=True, showDropDown=False)
            dv_l1 = DataValidation(type='list',
                formula1='"Repl Item at Same Seq,Remove Item,Add Item,Change"',
                allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv_l0)
            ws.add_data_validation(dv_l1)

            sky_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

            for r in range(self.table.rowCount()):
                excel_row = r + 3  # data starts at row 3

                bom_level = ''
                if bom_col_table >= 0:
                    it = self.table.item(r, bom_col_table)
                    if it:
                        bom_level = it.text().strip()

                for c in range(1, self.table.columnCount()):
                    excel_col = c
                    widget = self.table.cellWidget(r, c)
                    item = self.table.item(r, c)

                    if c == 1:  # Action dropdown
                        value = (widget.currentText() if isinstance(widget, QComboBox)
                                 else (item.text() if item else ''))
                        cell = ws.cell(row=excel_row, column=excel_col, value=value)
                        (dv_l0 if bom_level == '0' else dv_l1).add(cell)

                    elif c in VERT_EXCEL_COLS:  # checkbox symbol, no dropdown
                        if widget and hasattr(widget, '_chk'):
                            value = '\u2611' if widget._chk.isChecked() else '\u2610'
                        elif item:
                            value = item.text()
                        else:
                            value = ''
                        cell = ws.cell(row=excel_row, column=excel_col, value=value)

                    else:
                        value = item.text() if item else ''
                        cell = ws.cell(row=excel_row, column=excel_col, value=value)

                    if excel_col == part_excel_col:
                        cell.alignment = Alignment(horizontal='left', vertical='center',
                                                   wrap_text=True)
                    elif excel_col == desc_excel_col:
                        cell.alignment = Alignment(horizontal='left', vertical='center',
                                                   wrap_text=True)
                    elif excel_col in VERT_EXCEL_COLS:
                        cell.alignment = Alignment(horizontal='center', vertical='center',
                                                   wrap_text=False)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center',
                                                   wrap_text=True)

                    if bom_level == '0':
                        cell.fill = sky_fill

                    if excel_col not in VERT_EXCEL_COLS:
                        col_max_len[excel_col] = max(col_max_len.get(excel_col, 0),
                                                     len(str(value) if value else ''))

            # Column widths
            for c_idx in range(1, len(headers) + 1):
                if c_idx in VERT_EXCEL_COLS:
                    ws.column_dimensions[get_column_letter(c_idx)].width = 5
                else:
                    fit_w = min(max(col_max_len.get(c_idx, 8) + 2, 8), 50)
                    ws.column_dimensions[get_column_letter(c_idx)].width = fit_w

            ws.freeze_panes = 'A3'

            save_path = file_path
            while True:
                try:
                    wb.save(save_path)
                    QMessageBox.information(self, 'Export', f'Structure Sheet exported to:\n{save_path}')
                    break
                except PermissionError:
                    QMessageBox.warning(
                        self,
                        'Export Error',
                        'Permission denied while saving the file.\n\n'
                        'Please close the target Excel file (if open) or choose a different folder/file name.'
                    )
                    save_path, _ = QFileDialog.getSaveFileName(
                        self, 'Save Structure Sheet', save_path,
                        'Excel Files (*.xlsx);;All Files (*)'
                    )
                    if not save_path:
                        return

        except Exception as e:
            QMessageBox.warning(self, 'Export Error', f'Error exporting:\n{str(e)}\n\n{traceback.format_exc()}')

    def _import_structure_sheet(self):
        """Import Structure Sheet from Excel, restoring dropdowns, checkboxes, and formatting."""
        try:
            from openpyxl import load_workbook
            
            file_path, _ = QFileDialog.getOpenFileName(
                self, 'Open Structure Sheet', '', 'Excel Files (*.xlsx);;All Files (*)'
            )
            if not file_path:
                return
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Extract headers
            # Row 1 is the group label row; column headers are in row 2.
            headers = []
            for c in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=2, column=c).value
                headers.append(str(cell_value) if cell_value else f'Col{c}')
            
            # Prepend "Select" column to match table structure
            headers = ['Select'] + headers
            
            # Clear and reset table
            self.table.clear()
            self.table.setRowCount(0)
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            
            # Create rotated header for vertical columns (2-6)
            from PyQt6.QtCore import Qt
            hdr = RotatedColumnsHeader(
                Qt.Orientation.Horizontal,
                rotated_columns=range(2, 7),
                parent=self.table,
                group_label='Change Type',
                group_columns=list(range(2, 7)),
            )
            hdr.setStretchLastSection(True)
            hdr.setSectionsClickable(True)
            self.table.setHorizontalHeader(hdr)
            self.table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.horizontalHeader().setMinimumHeight(117)
            self._apply_inserted_header_colors()
            
            # Set column widths for rotated columns
            for vc in range(2, 7):
                self.table.horizontalHeader().setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                self.table.setColumnWidth(vc, 40)
            
            # Find BOM Level column
            bom_col = -1
            for c, h in enumerate(headers):
                if 'bom level' in h.lower():
                    bom_col = c
                    break
            
            # Load data rows
            for r in range(3, ws.max_row + 1):  # data starts at row 3
                row_data = []
                for c in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=r, column=c).value
                    row_data.append(str(cell_value) if cell_value else '')
                
                self.table.insertRow(self.table.rowCount())
                current_row = self.table.rowCount() - 1
                
                # Get BOM level for this row
                bom_level = row_data[bom_col - 1].strip() if (bom_col > 0 and bom_col <= len(row_data)) else ''
                
                # Set up cells
                for c in range(len(headers)):
                    if c == 0:  # Select column with checkbox
                        cont = QWidget()
                        h_lay = QHBoxLayout(cont)
                        h_lay.setContentsMargins(0, 0, 0, 0)
                        h_lay.addStretch(1)
                        chk = QCheckBox()
                        h_lay.addWidget(chk)
                        h_lay.addStretch(1)
                        cont._chk = chk
                        self.table.setCellWidget(current_row, c, cont)
                    
                    elif c == 1:  # Change Type dropdown
                        combo = QComboBox()
                        # L0: Revised + Change only;  non-zero: no Revised
                        if bom_level == '0':
                            combo.addItems(['', 'Revised', 'Change'])
                        else:
                            combo.addItems(['', 'Repl Item at Same Seq', 'Remove Item', 'Add Item', 'Change'])

                        # Set value from Excel
                        if c - 1 < len(row_data):
                            combo_value = row_data[c - 1]
                            idx = combo.findText(combo_value)
                            if idx >= 0:
                                combo.setCurrentIndex(idx)

                        combo.setStyleSheet('QComboBox { padding: 2px; }')
                        combo.currentTextChanged.connect(lambda text, rr=current_row: self._on_change_type_changed(rr))
                        self.table.setCellWidget(current_row, c, combo)
                    
                    elif c in {2, 3, 4, 5, 6}:  # Checkbox columns
                        cont_v = QWidget()
                        hv = QHBoxLayout(cont_v)
                        hv.setContentsMargins(0, 0, 0, 0)
                        hv.addStretch(1)
                        chk_v = QCheckBox()
                        
                        # Set checked state from Excel
                        if c - 1 < len(row_data):
                            cell_val = row_data[c - 1]
                            chk_v.setChecked(cell_val in ['☑', 'True', '1', 'TRUE'])
                        
                        hv.addWidget(chk_v)
                        hv.addStretch(1)
                        cont_v._chk = chk_v
                        chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                        self.table.setCellWidget(current_row, c, cont_v)
                    
                    else:
                        item = QTableWidgetItem(row_data[c - 1] if c - 1 < len(row_data) else '')
                        self.table.setItem(current_row, c, item)
                
                # Apply sky-blue for BOM Level 0
                if bom_level == '0':
                    for c in range(self.table.columnCount()):
                        if c == 0:
                            w = self.table.cellWidget(current_row, c)
                            if w:
                                w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                        elif c == 1:
                            w = self.table.cellWidget(current_row, c)
                            if w:
                                w.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                        elif c in {2, 3, 4, 5, 6}:
                            w = self.table.cellWidget(current_row, c)
                            if w:
                                w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                            else:
                                cell = self.table.item(current_row, c)
                                if cell:
                                    cell.setBackground(QColor('#87CEEB'))
                        else:
                            cell = self.table.item(current_row, c)
                            if cell:
                                cell.setBackground(QColor('#87CEEB'))
                
                # Initialize checkbox visibility based on Change Type
                self._on_change_type_changed(current_row)
            
            QMessageBox.information(self, 'Import', f'Structure Sheet imported successfully from:\n{file_path}')
            self._update_inserted_editability_all_rows()
            self._apply_inserted_header_colors()
            self._refresh_structure_action_buttons()
        
        except Exception as e:
            QMessageBox.warning(self, 'Import Error', f'Error importing: {str(e)}\n{traceback.format_exc()}')


# ================= Inventory Cost Tab (FULL UPDATED) =================

PLANTS = [4020,4055,4060,4070,4080,4090]
METRICS = ['On Order Qty','Onhand Qty','Gross Demand-13','Gross Demand-26','Gross Demand-52','Standard Cost USD']
NO_CDW_CODES = {
'0070','0080','0110','0120','0130','0170','0180','0210','0243','0250','0251','0260','0261','0280','0288','0289','0290',
'0301','0302','0303','0304','0305','0320','0330','0335','0340','0345','0350','0355','0360','0365','0370','0375','0380','0385','0390','0395',
'0401','0402','0403','0404','0405','0410','0415','0420','0425','0430','0435','0440','0445','0450','0455','0460','0465','0470','0475','0480','0485','0490','0495'
}

class InventoryCostTab(QWidget):
    def reset_tab(self):
        self.df = None
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self._update_charts()

    def __init__(self):
        super().__init__()
        v = QVBoxLayout(self)
        h = QHBoxLayout()
        self.btn_import = QPushButton('Import MM360 & Create Inventory & Cost')
        self.btn_import_db = QPushButton('Import from Databricks')
        self.btn_export = QPushButton('Export Excel')
        self.btn_reset = QPushButton('Reset Tab')
        for b in (self.btn_import, self.btn_import_db, self.btn_export, self.btn_reset):
            h.addWidget(b)
        h.addStretch(1)
        v.addLayout(h)

        self.subtabs = QTabWidget()

        self.data_tab = QWidget()
        data_layout = QVBoxLayout(self.data_tab)
        data_layout.setContentsMargins(0, 0, 0, 0)
        self.table = QTableWidget(0,0)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setVisible(True)
        data_layout.addWidget(self.table)

        self.charts_tab = QWidget()
        charts_layout = QVBoxLayout(self.charts_tab)
        charts_layout.setContentsMargins(8, 8, 8, 8)

        self.chart_status = QLabel('Import data to view charts.')
        charts_layout.addWidget(self.chart_status)

        self.cost_canvas = None
        self.demand_canvas = None
        self.cost_fig = None
        self.demand_fig = None
        self.cost_ax = None
        self.demand_ax = None

        if _HAS_MATPLOTLIB:
            self.cost_fig = Figure(figsize=(9, 3.8), tight_layout=True)
            self.cost_canvas = FigureCanvas(self.cost_fig)
            self.cost_ax = self.cost_fig.add_subplot(111)
            charts_layout.addWidget(self.cost_canvas)

            self.demand_fig = Figure(figsize=(9, 3.8), tight_layout=True)
            self.demand_canvas = FigureCanvas(self.demand_fig)
            self.demand_ax = self.demand_fig.add_subplot(111)
            charts_layout.addWidget(self.demand_canvas)
        else:
            charts_layout.addWidget(QLabel('matplotlib is not available in this environment.'))

        self.subtabs.addTab(self.data_tab, 'Table')
        self.subtabs.addTab(self.charts_tab, 'Charts')
        v.addWidget(self.subtabs)
        self.df = None
        self.btn_import.clicked.connect(self.import_mm360)
        self.btn_import_db.clicked.connect(self.import_databricks)
        self.btn_export.clicked.connect(self.export_excel)
        self.btn_reset.clicked.connect(self.reset_tab)

    def _update_charts(self):
        if not _HAS_MATPLOTLIB or self.cost_ax is None or self.demand_ax is None:
            return

        self.cost_ax.clear()
        self.demand_ax.clear()

        if self.df is None or self.df.empty:
            self.chart_status.setText('Import data to view charts.')
            self.cost_ax.set_title('Inventory Cost by Plant')
            self.demand_ax.set_title('Gross Demand-52 by Plant')
            self.cost_canvas.draw_idle()
            self.demand_canvas.draw_idle()
            return

        plants = [str(p) for p in PLANTS]
        cost_values = []
        demand_values = []
        for p in plants:
            # Cost should be based on on-hand inventory only.
            onhand_col = f'{p} Onhand Qty'
            std_cost_col = f'{p} Standard Cost USD'
            dem52_col = f'{p} Gross Demand-52'

            onhand = pd.to_numeric(self.df.get(onhand_col, 0), errors='coerce').fillna(0)
            std_cost = pd.to_numeric(self.df.get(std_cost_col, 0), errors='coerce').fillna(0)
            dem52 = pd.to_numeric(self.df.get(dem52_col, 0), errors='coerce').fillna(0)

            plant_cost = float((onhand * std_cost).sum())
            plant_dem52 = float(dem52.sum())
            cost_values.append(plant_cost)
            demand_values.append(plant_dem52)

        self.cost_ax.bar(plants, cost_values, color='#2E86C1')
        self.cost_ax.set_title('Inventory Cost by Plant')
        self.cost_ax.set_ylabel('USD')
        self.cost_ax.tick_params(axis='x', labelrotation=0)

        self.demand_ax.bar(plants, demand_values, color='#27AE60')
        self.demand_ax.set_title('Gross Demand-52 by Plant')
        self.demand_ax.set_ylabel('Qty')
        self.demand_ax.tick_params(axis='x', labelrotation=0)

        self.chart_status.setText('Charts updated for plant-wise Inventory Cost and Demand-52.')
        self.cost_canvas.draw_idle()
        self.demand_canvas.draw_idle()

    def _build_obs_replacement_map(self) -> Dict[str, str]:
        obs_map: Dict[str, str] = {}
        try:
            main = self.window()
            obs_tab = getattr(main, 'obs_tab', None)
            if not obs_tab or not hasattr(obs_tab, 'table'):
                return obs_map
            t = obs_tab.table
            for r in range(t.rowCount()):
                obs = t.item(r, 1).text().strip() if t.item(r, 1) else ''
                rep = t.item(r, 3).text().strip() if t.item(r, 3) else ''
                if obs:
                    obs_map[obs.upper()] = rep
        except Exception:
            pass
        return obs_map

    def import_mm360(self):
        path,_ = QFileDialog.getOpenFileName(self,'Open MM360','', 'Excel Files (*.xlsx)')
        if not path: return
        src = pd.read_excel(path, sheet_name='Material Analysis', engine='openpyxl')

        def pace(v):
            v=str(v)
            if v.startswith('SGP'): return 'PACE'
            if v.startswith('GDS'): return 'DASH'
            return ''
        src['PACE/DASH'] = src['MRP Profile'].apply(pace)

        obs_map = self._build_obs_replacement_map()
        rows=[]
        for (pn,desc,pdsh), g in src.groupby(['Material Number','Material Description','PACE/DASH']):
            code4=str(pn)[:4]
            prim=sec=''
            if code4 in NO_CDW_CODES:
                prim=sec='No Change Required'
            rep = obs_map.get(str(pn).strip().upper(), '')
            row={
                'Material Number':pn,
                'Material Description':desc,
                'Replacement Part':rep,
                'Primary Disposition':prim,
                'Secondary Disposition':sec,
                'PACE/DASH':pdsh
            }
            tot_on=tot_oh=tot_d13=tot_d26=tot_d52=tot_cost=0.0
            for p in PLANTS:
                gp=g[g['Plant Code']==p]
                oo=float(gp['On Order Quantity'].sum()) if not gp.empty else 0
                oh=float(gp['onhand'].sum()) if not gp.empty else 0
                d13=float(gp['Gross Demand-13'].sum()) if not gp.empty else 0
                d26=float(gp['Gross Demand-26'].sum()) if not gp.empty else 0
                d52=float(gp['Gross Demand-52'].sum()) if not gp.empty else 0
                sc=float(gp['Standard Cost USD'].iloc[0]) if oh > 0 else 0
                row[f'{p} On Order Qty']=oo
                row[f'{p} Onhand Qty']=oh
                row[f'{p} Gross Demand-13']=d13
                row[f'{p} Gross Demand-26']=d26
                row[f'{p} Gross Demand-52']=d52
                row[f'{p} Standard Cost USD']=sc
                tot_on+=oo; tot_oh+=oh
                tot_d13+=d13; tot_d26+=d26; tot_d52+=d52
                tot_cost+=oh*sc
            row['Total On Order Quantity']=tot_on
            row['Total Onhand']=tot_oh
            row['Gross Demand-13']=tot_d13
            row['Gross Demand-26']=tot_d26
            row['Gross Demand-52']=tot_d52
            row['Inventory Cost']=tot_cost
            rows.append(row)
        self.df=pd.DataFrame(rows).sort_values('Inventory Cost',ascending=False)
        self.render()

    def import_databricks(self):
        """Import inventory, demand, and cost data from Databricks for OBS parts."""
        if fetch_inventory_demand_cost is None:
            QMessageBox.warning(self, 'Import Error', 'Failed to import Databricks query module. Check installation.')
            return
        try:
            # Get OBS parts from OBS Tab
            obs_parts = []
            try:
                main = self.window()
                obs_tab = getattr(main, 'obs_tab', None)
                if obs_tab and hasattr(obs_tab, 'table'):
                    t = obs_tab.table
                    for r in range(t.rowCount()):
                        obs = t.item(r, 1)
                        if obs:
                            part = obs.text().strip()
                            if part:
                                obs_parts.append(part)
            except Exception as e:
                QMessageBox.warning(self, 'Error', f'Failed to extract OBS parts: {str(e)}')
                return
            if not obs_parts:
                QMessageBox.information(self, 'No Data', 'No OBS parts found in OBS Tab.')
                return
            # Fetch data from Databricks
            QMessageBox.information(self, 'Loading', f'Fetching data for {len(obs_parts)} OBS part(s) from Databricks...')
            results = fetch_inventory_demand_cost(obs_parts, plants=PLANTS)
            if not results:
                QMessageBox.warning(self, 'No Results', 'No data returned from Databricks.')
                return
            # Build DataFrame similar to import_mm360 structure
            # Define PACE/DASH mapper (same as MM360)
            def get_pace_dash(mrp_profile):
                if mrp_profile is None:
                    return ''
                mrp_str = str(mrp_profile).strip().upper()
                if mrp_str.startswith('SGP'):
                    return 'PACE'
                if mrp_str.startswith('GDS'):
                    return 'DASH'
                return ''
            
            rows = []
            # Normalize Databricks rows for stable key matching (string/int/case safe).
            result_by_part = {}
            obs_map = self._build_obs_replacement_map()
            for rec in results:
                part_key = str(rec.get('part_number', '')).strip().upper()
                if part_key:
                    result_by_part.setdefault(part_key, []).append(rec)

            for pn in obs_parts:
                pn_key = str(pn).strip().upper()
                part_results = result_by_part.get(pn_key, [])
                if not part_results:
                    continue
                # Prefer a row that actually has metadata populated (4020 is often blank).
                first = next(
                    (
                        r for r in part_results
                        if str(r.get('part_description', '')).strip()
                        or str(r.get('mrp_profile', '')).strip()
                        or str(r.get('make_buy', '')).strip()
                    ),
                    part_results[0],
                )
                code4 = str(pn)[:4]
                prim = sec = ''
                if code4 in NO_CDW_CODES:
                    prim = sec = 'No Change Required'
                # Derive PACE/DASH from the first non-empty MRP profile row.
                mrp_row = next(
                    (r for r in part_results if str(r.get('mrp_profile', '')).strip()),
                    first,
                )
                pdsh = get_pace_dash(mrp_row.get('mrp_profile'))
                row = {
                    'Material Number': pn,
                    'Material Description': first.get('part_description', ''),
                    'Replacement Part': obs_map.get(pn_key, ''),
                    'Primary Disposition': prim,
                    'Secondary Disposition': sec,
                    'PACE/DASH': pdsh
                }
                tot_on = tot_oh = tot_d13 = tot_d26 = tot_d52 = tot_cost = 0.0
                plant_map = {
                    str(r.get('plant', '')).strip(): r
                    for r in part_results
                }
                for p in PLANTS:
                    p_key = str(p).strip()
                    plant_data = plant_map.get(p_key)
                    if plant_data:
                        oo = float(plant_data.get('on_order_qty') or 0)
                        oh = float(plant_data.get('on_hand_qty') or 0)
                        d13 = float(plant_data.get('gross_demand_13w') or 0)
                        d26 = float(plant_data.get('gross_demand_26w') or 0)
                        d52 = float(plant_data.get('gross_demand_52w') or 0)
                        raw_sc = float(plant_data.get('standard_cost_usd') or 0)
                        sc = raw_sc if oh > 0 else 0.0
                    else:
                        oo = oh = d13 = d26 = d52 = sc = 0.0
                    row[f'{p} On Order Qty'] = oo
                    row[f'{p} Onhand Qty'] = oh
                    row[f'{p} Gross Demand-13'] = d13
                    row[f'{p} Gross Demand-26'] = d26
                    row[f'{p} Gross Demand-52'] = d52
                    row[f'{p} Standard Cost USD'] = sc
                    tot_on += oo
                    tot_oh += oh
                    tot_d13 += d13
                    tot_d26 += d26
                    tot_d52 += d52
                    tot_cost += oh * sc
                row['Total On Order Quantity'] = tot_on
                row['Total Onhand'] = tot_oh
                row['Gross Demand-13'] = tot_d13
                row['Gross Demand-26'] = tot_d26
                row['Gross Demand-52'] = tot_d52
                row['Inventory Cost'] = tot_cost
                rows.append(row)
            self.df = pd.DataFrame(rows).sort_values('Inventory Cost', ascending=False)
            QMessageBox.information(self, 'Success', f'Loaded {len(rows)} OBS part(s) from Databricks.')
            self.render()
        except Exception as e:
            QMessageBox.warning(self, 'Import Error', f'Error importing from Databricks: {str(e)}')

    def render(self):
        if self.df is None: return
        self.table.setRowCount(len(self.df))
        self.table.setColumnCount(len(self.df.columns))

        headers = self.df.columns.tolist()
        display_headers = []
        rotated_cols = []
        plant_groups = []
        header_texts = {}  # { col_index: text }
        
        for idx, name in enumerate(headers):
            m = re.match(r'^(\d{4})\s+(.+)$', str(name))
            if m:
                plant = m.group(1)
                metric = m.group(2)
                display_headers.append(metric)  # remove repeated plant code from rotated labels
                header_texts[idx] = metric  # cache for header
                rotated_cols.append(idx)
                if not plant_groups or plant_groups[-1][0] != plant:
                    plant_groups.append((plant, [idx]))
                else:
                    plant_groups[-1][1].append(idx)
            else:
                display_headers.append(name)
                header_texts[idx] = name

        # Create and install custom rotated header first.
        hdr = RotatedColumnsHeader(Qt.Orientation.Horizontal, rotated_columns=rotated_cols, parent=self.table)
        hdr.set_header_texts(header_texts)  # pass cached header texts
        hdr.set_group_spans(plant_groups)
        self.table.setHorizontalHeader(hdr)
        self.table.horizontalHeader().setVisible(True)

        # Assign labels after installing the header so section/model state is fresh.
        self.table.setHorizontalHeaderLabels(display_headers)
        
        # Force header to have access to model and update rendering
        hdr.reset()
        hdr.viewport().update()
        self.table.horizontalHeader().setFixedHeight(180)

        # Populate table data
        for r in range(len(self.df)):
            for c,col in enumerate(self.df.columns):
                v=self.df.iloc[r][col]
                if isinstance(v, (int, float)) and v == 0:
                    txt = ''
                elif isinstance(v, (int, float)) and 'Cost' in col:
                    txt = f'${v:,.2f}'
                else:
                    txt = str(v)
                self.table.setItem(r,c,QTableWidgetItem(txt))

        # Resize and apply fixed widths to rotated columns
        self.table.resizeColumnsToContents()
        for c in rotated_cols:
            self.table.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)
            col_name = str(self.df.columns[c])
            if 'Cost' in col_name:
                self.table.setColumnWidth(c, 92)
            else:
                self.table.setColumnWidth(c, 36)

        # Ensure all cost columns are wide enough for currency text.
        for c, col_name in enumerate(self.df.columns):
            if 'Cost' not in str(col_name):
                continue
            self.table.resizeColumnToContents(c)
            self.table.setColumnWidth(c, max(self.table.columnWidth(c) + 12, 110))

        self._update_charts()

    def export_excel(self):
        if self.df is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, 'Export', 'inventory_cost.xlsx', 'Excel Files (*.xlsx)'
        )
        if not path:
            return
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import pandas as pd

        df = self.df.copy()
        if 'Replacement Part' not in df.columns:
            obs_map = self._build_obs_replacement_map()
            df.insert(2, 'Replacement Part', '')
            for i, pn in enumerate(df['Material Number']):
                key = str(pn).strip().upper()
                if key in obs_map:
                    df.at[i, 'Replacement Part'] = obs_map[key]

        with pd.ExcelWriter(path, engine='openpyxl') as w:
            df.to_excel(w, index=False, sheet_name='Output Sheet', startrow=1)
        wb = load_workbook(path)
        ws = wb['Output Sheet']

        header_fill = PatternFill('solid', fgColor='BDD7EE')
        ws.row_dimensions[2].height = 118

        # Format row 2 headers
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(2, c)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            header_name = str(cell.value or '')
            is_plant_metric = bool(re.match(r'^\d{4} ', header_name))
            if not is_plant_metric:
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, text_rotation=90)

        # ---- Group plant metric columns in row 1 using plant code ----
        groups = {}
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(2, c).value or '')
            m = re.match(r'^(\d{4})\s+', val)
            if not m:
                continue
            code = m.group(1)
            ws.cell(1, c).value = code
            groups.setdefault(code, []).append(c)

        # Merge each plant's metric columns in row 1
        for code, cols in groups.items():
            if not code:
                continue
            if len(cols) > 1:
                ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])
            cell = ws.cell(1, cols[0])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # Apply currency formatting ($ with 2 decimals) to all cost columns
        cost_columns = [
            c for c in range(1, ws.max_column + 1)
            if 'Cost' in str(ws.cell(2, c).value or '')
        ]
        for r in range(3, ws.max_row + 1):
            for c in cost_columns:
                cell = ws.cell(r, c)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0.00'

        # Column widths
        for c in range(1, ws.max_column + 1):
            header_name = str(ws.cell(2, c).value or '')
            if re.match(r'^\d{4} ', header_name):
                ws.column_dimensions[get_column_letter(c)].width = 4
                continue
            col_letter = get_column_letter(c)
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, c).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = max(10, max_len + 2)

        # Blank zero values
        for r in range(3, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).value == 0:
                    ws.cell(r, c).value = ''

        wb.save(path)
        QMessageBox.information(self, 'Export', 'Excel exported successfully')

class PlaceholderTab(QWidget):
    def __init__(self, title: str):
        super().__init__()
        l = QVBoxLayout(self)
        l.addWidget(QLabel(f"{title} – UI under development"))



class OrphanOBSSubTab(QWidget):
    def __init__(self):
        super().__init__()
        outer = QVBoxLayout(self)

        btn_row = QHBoxLayout()
        self.btn_import = QPushButton("Import BOM of OBS Parts")
        lbl_bom = QLabel("BOM Level (1-18):")
        self.bom_level_input = QLineEdit()
        self.bom_level_input.setFixedWidth(45)
        self.bom_level_input.setText("6")
        self.bom_level_input.setPlaceholderText("1-18")
        self.bom_level_input.setToolTip("Enter the maximum BOM depth (1 to 18)")

        lbl_plant = QLabel("Plant:")
        self.plant_combo = QComboBox()
        self.plant_combo.addItems(["4020", "4055", "4060", "4070", "4080", "4090"])
        self.plant_combo.setCurrentText("4070")
        self.plant_combo.setFixedWidth(70)
        self.plant_combo.setToolTip("Plant code to filter the BOM query")

        self.btn_select_oem = QPushButton("Select OEM's")
        self.btn_delete = QPushButton("Delete")
        self.btn_reset = QPushButton("Reset")
        self.btn_copy_removed = QPushButton("Copy Removed Child Parts")

        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(lbl_bom)
        btn_row.addWidget(self.bom_level_input)
        btn_row.addWidget(lbl_plant)
        btn_row.addWidget(self.plant_combo)
        btn_row.addWidget(self.btn_select_oem)
        btn_row.addWidget(self.btn_delete)
        btn_row.addWidget(self.btn_reset)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_copy_removed)
        outer.addLayout(btn_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self.import_bom)
        self.btn_select_oem.clicked.connect(self.select_oems)
        self.btn_delete.clicked.connect(self.delete_selected)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.btn_copy_removed.clicked.connect(self.copy_removed_child_parts)

    def import_bom(self):
        level_text = self.bom_level_input.text().strip()
        if not level_text:
            QMessageBox.warning(self, 'BOM Level Required', 'Please enter BOM Level (1 to 18).')
            return
        try:
            max_level = int(level_text)
        except ValueError:
            QMessageBox.warning(self, 'Invalid BOM Level', f'BOM Level must be a number from 1 to 18, got: "{level_text}"')
            return
        if max_level < 1 or max_level > 18:
            QMessageBox.warning(self, 'Invalid BOM Level', f'BOM Level must be between 1 and 18, got: {max_level}')
            return

        plant = self.plant_combo.currentText().strip() or '4070'
        self.import_from_sap(max_level=max_level, plant=plant)

    def _is_bom_level_zero(self, bom_val: str) -> bool:
        txt = (bom_val or '').strip()
        if txt == '':
            return False
        try:
            return float(txt) == 0.0
        except ValueError:
            return txt == '0'

    def _apply_comment_and_row_style(self, row_idx: int, bom_val: str, tc_tbl_idx: int):
        if self._is_bom_level_zero(bom_val):
            for c in range(self.table.columnCount()):
                cell = self.table.item(row_idx, c)
                if cell:
                    cell.setBackground(QColor('#87CEEB'))  # Sky Blue
            return

        it = QTableWidgetItem('Removed BOM Item')
        f = it.font()
        f.setBold(True)
        it.setFont(f)
        it.setForeground(QColor('orange'))
        self.table.setItem(row_idx, tc_tbl_idx, it)

    def _collect_obs_parts(self) -> list[str]:
        main = self.window()
        obs_tab = getattr(main, 'obs_tab', None)
        if obs_tab is None or not hasattr(obs_tab, 'table'):
            return []

        parts = []
        seen = set()
        t = obs_tab.table
        for r in range(t.rowCount()):
            it = t.item(r, 1)  # OBS Parts column
            part = (it.text() if it else '').strip()
            key = part.upper()
            if part and key not in seen:
                seen.add(key)
                parts.append(part)
        return parts

    def import_from_sap(self, max_level: int, plant: str):
        obs_parts = self._collect_obs_parts()
        if not obs_parts:
            QMessageBox.warning(
                self,
                'No OBS Parts',
                'OBS Parts tab is empty. Please add OBS parts before importing from SAP.'
            )
            return

        try:
            import sys as _sys
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from implemented_bom_query import fetch_implemented_bom  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(
                self,
                'Module Not Found',
                f'implemented_bom_query.py could not be imported:\n{exc}'
            )
            return

        try:
            records = fetch_implemented_bom(obs_parts, max_level=max_level, plant=plant, include_level0=True)
        except Exception as exc:
            QMessageBox.warning(self, 'SAP Import Error', str(exc))
            return

        if not records:
            QMessageBox.information(
                self,
                'No Data',
                f'No BOM data found in SAP for {len(obs_parts)} OBS part(s) at plant {plant}.'
            )
            return

        headers = [
            'Select',
            'BOM Level',
            'Part',
            'Tool comments',
            'Rev/Ln',
            'Plant',
            'Description',
            'Item Status',
            'Base Qty',
            'Ext Qty',
            'UOM',
            'ECO Number',
            'Procurement Type',
            'Effectivity Date',
            'User Item Type',
            'Item Seq',
            'Kit Code',
            'Designator',
            'Option Class',
            'BOM Source Type (Debug)',
        ]
        self.table.clear()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(records))

        key_by_header = {
            'BOM Level': 'bom_level',
            'Part': 'part',
            'Tool comments': 'tool_comments',
            'Rev/Ln': 'rev_ln',
            'Plant': 'plant',
            'Description': 'description',
            'Item Status': 'item_status',
            'Base Qty': 'base_qty',
            'Ext Qty': 'ext_qty',
            'UOM': 'uom',
            'ECO Number': 'eco_number',
            'Procurement Type': 'procurement_type',
            'Effectivity Date': 'effectivity_date',
            'User Item Type': 'user_item_type',
            'Item Seq': 'item_seq',
            'Kit Code': 'kit_code',
            'Designator': 'designator',
            'Option Class': 'option_class',
            'BOM Source Type (Debug)': 'bom_source_type',
        }

        bom_tbl_idx = headers.index('BOM Level')
        part_tbl_idx = headers.index('Part')
        tc_tbl_idx = headers.index('Tool comments')

        for r, rec in enumerate(records):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            self.table.setCellWidget(r, 0, cont)

            for c, hdr in enumerate(headers[1:], start=1):
                db_key = key_by_header.get(hdr, '')
                val = rec.get(db_key, '')

                # Fallbacks for legacy query outputs where these aliases are not present.
                if hdr == 'Designator' and (val is None or str(val).strip() == ''):
                    val = rec.get('sparable_flag', '')
                elif hdr == 'BOM Source Type (Debug)' and (val is None or str(val).strip() == ''):
                    val = rec.get('pace_or_dash', '')

                self.table.setItem(r, c, QTableWidgetItem(str(val) if val is not None else ''))

            # Part hierarchy indent using BOM level (same visual style as Where Used).
            bom_item_for_indent = self.table.item(r, bom_tbl_idx)
            part_item = self.table.item(r, part_tbl_idx)
            if part_item is not None:
                raw_part = (part_item.text() or '').strip()
                try:
                    level_int = int(float((bom_item_for_indent.text() if bom_item_for_indent else '0') or '0'))
                except (ValueError, TypeError):
                    level_int = 0
                part_item.setText(('      ' * max(level_int, 0)) + raw_part)

            bom_item = self.table.item(r, bom_tbl_idx)
            bom_val = bom_item.text() if bom_item else ''
            self._apply_comment_and_row_style(r, bom_val, tc_tbl_idx)

        self.table.resizeColumnsToContents()
        QMessageBox.information(
            self,
            'Import Complete',
            f'Imported {len(records)} BOM row(s) from SAP for {len(obs_parts)} OBS part(s) at level {max_level}.'
        )

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def find_column(self, header):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text() == header:
                return i
        return -1

    def select_oems(self):
        part_col = self.find_column('Part')
        bom_col = self.find_column('BOM Level')
        if part_col < 0 or bom_col < 0:
            return

        eligible = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, part_col).text() if self.table.item(r, part_col) else '').lstrip()
            bom = (self.table.item(r, bom_col).text() if self.table.item(r, bom_col) else '').strip()
            prefix = part[:4]
            if bom != '0' and prefix.isdigit() and int(prefix) >= 500:
                eligible.append(r)

        should_select = any(
            not self.table.cellWidget(r, 0).findChild(QCheckBox).isChecked()
            for r in eligible
        )

        for r in eligible:
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            chk.setChecked(should_select)

    def delete_selected(self):
        rows = []
        for r in range(self.table.rowCount()):
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            if chk.isChecked():
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)

    def copy_removed_child_parts(self):
        part_col = self.find_column('Part')
        tc_col = self.find_column('Tool comments')

        if part_col < 0 or tc_col < 0:
            QMessageBox.warning(self, 'Column Missing',
                                'Required columns (Part / Tool comments) not found.')
            return

        unique_parts = set()

        for r in range(self.table.rowCount()):
            tc_item = self.table.item(r, tc_col)
            if not tc_item or not tc_item.text().strip():
                continue

            part_item = self.table.item(r, part_col)
            if not part_item:
                continue

            part = part_item.text().replace(' ', '').strip()
            if part:
                unique_parts.add(part)

        if not unique_parts:
            QMessageBox.information(self, 'No Data',
                                    'No removed child parts found to copy.')
            return

        from PyQt6.QtGui import QGuiApplication
        result = '\n'.join(sorted(unique_parts))
        QGuiApplication.clipboard().setText(result)

        QMessageBox.information(self, 'Copied',
                                f'Copied {len(unique_parts)} unique removed child part(s) to clipboard.')


class WURemovedBOMItemsTab(QWidget):

    def append_orphans_to_obs_parts(self, remove_from_orphan_table=False):
        if not self.obs_provider:
            return
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphan Child')
        if pcol < 0 or ocol < 0:
            return
        orphan_map = {}
        orphan_rows = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, pcol)
            oit = self.table.item(r, ocol)
            if pit and oit and oit.text().lower().startswith('orphan'):
                orphan_map[pit.text().strip().upper()] = oit.text().strip()
                orphan_rows.append(r)
        if not orphan_map:
            return
        t = self.obs_provider.table
        existing = {t.item(r,1).text().strip().upper() for r in range(t.rowCount()) if t.item(r,1) and t.item(r,1).text().strip()}
        for part, lvl in orphan_map.items():
            if part in existing:
                continue
            target = next((r for r in range(t.rowCount()) if not t.item(r,1) or not t.item(r,1).text().strip()), None)
            if target is None:
                target = t.rowCount(); t.setRowCount(target+1); t._init_rows(target,target+1)
            t.setItem(target,1,QTableWidgetItem(part))
            cb = QComboBox(); cb.addItem(lvl); cb.setCurrentText(lvl)
            col = get_orphan_color(lvl)
            if col: cb.setStyleSheet(f"QComboBox {{ color:{col.name()}; font-weight:bold; }}")
            t.setCellWidget(target,2,cb)
            t.setItem(target,3,QTableWidgetItem('No Replacement'))
        if remove_from_orphan_table:
            for r in reversed(orphan_rows): self.table.removeRow(r)

    def __init__(self, obs_provider, imp_bom_provider=None):
        super().__init__()
        self.obs_provider = obs_provider
        self.imp_bom_provider = imp_bom_provider
        outer = QVBoxLayout(self)

        # ── Button row ────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        wu_lbl = QLabel("WU Level (1-6):")
        self.wu_level_input = QLineEdit()
        self.wu_level_input.setFixedWidth(45)
        self.wu_level_input.setText("1")
        self.wu_level_input.setPlaceholderText("1-6")
        self.wu_level_input.setToolTip("Enter the maximum Where Used depth (1 to 6)")

        plant_lbl = QLabel("Plant:")
        self.plant_combo = QComboBox()
        self.plant_combo.addItems(["4020", "4055", "4060", "4070", "4080", "4090"])
        self.plant_combo.setCurrentText("4070")
        self.plant_combo.setFixedWidth(70)
        self.plant_combo.setToolTip("Plant code to filter the WU query")

        self.btn_import = QPushButton("Import WU of Removed BOM items")
        self.btn_sap_import_wu_removed = QPushButton("SAP Import WU of Removed BOM Item")
        self.btn_analyze = QPushButton("Perform Orphan Analysis")
        self.btn_reset = QPushButton("Reset")
        self.btn_remove_esw = QPushButton("Remove ESW Parents")
        self.btn_remove_9024 = QPushButton("Remove 9024")

        btn_row.addWidget(wu_lbl)
        btn_row.addWidget(self.wu_level_input)
        btn_row.addWidget(plant_lbl)
        btn_row.addWidget(self.plant_combo)
        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(self.btn_sap_import_wu_removed)
        btn_row.addWidget(self.btn_analyze)
        btn_row.addWidget(self.btn_reset)
        btn_row.addWidget(self.btn_remove_esw)
        btn_row.addWidget(self.btn_remove_9024)
        btn_row.addStretch(1)
        outer.addLayout(btn_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self.import_from_databricks_for_removed_items)
        self.btn_sap_import_wu_removed.clicked.connect(self.import_removed_bom_wu_from_sap)
        self.btn_analyze.clicked.connect(self.perform_orphan_analysis)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.btn_remove_esw.clicked.connect(lambda: self.remove_by_prefix('ESW'))
        self.btn_remove_9024.clicked.connect(lambda: self.remove_by_prefix('9024'))

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def _find_col(self, name):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text().strip().lower() == name.lower():
                return i
        return -1

    def _import_wu_by_tool_comment(self, required_comment: str, flow_name: str):
        """Generic WU import by matching Imp BOM Tool comments value."""
        from PyQt6.QtCore import QThread, pyqtSignal, QObject
        from PyQt6.QtGui import QColor
        from PyQt6.QtWidgets import QProgressDialog

        wu_text = self.wu_level_input.text().strip()
        if not wu_text:
            QMessageBox.warning(self, 'WU Level Required',
                                f'Please enter a WU Level (1 to 6) before {flow_name}.')
            return
        try:
            max_level = int(wu_text)
        except ValueError:
            QMessageBox.warning(self, 'Invalid WU Level',
                                f'WU Level must be a number between 1 and 6, got: "{wu_text}"')
            return
        if max_level < 1 or max_level > 6:
            QMessageBox.warning(self, 'Invalid WU Level',
                                f'WU Level must be between 1 and 6, got: {max_level}')
            return

        selected_plant = self.plant_combo.currentText().strip() or '4070'

        if self.imp_bom_provider is None:
            QMessageBox.warning(self, 'Imp BOM Not Available',
                                'The Imp BOM tab is not connected. Cannot read removed parts.')
            return

        imp_table = self.imp_bom_provider.table
        part_col = self.imp_bom_provider.find_column('Part')
        tc_col = self.imp_bom_provider.find_column('Tool comments')

        if part_col < 0 or tc_col < 0:
            QMessageBox.warning(self, 'Imp BOM Missing Columns',
                                'Imp BOM does not contain both a Part and a Tool comments column.\n'
                                'Please import a BOM file first.')
            return

        seen = set()
        target_parts = []
        for r in range(imp_table.rowCount()):
            tc_item = imp_table.item(r, tc_col)
            if not tc_item:
                continue
            comment = tc_item.text().strip().lower()
            if comment != required_comment:
                continue

            part_item = imp_table.item(r, part_col)
            if not part_item:
                continue
            part = part_item.text().strip()
            key = part.upper()
            if part and key not in seen:
                seen.add(key)
                target_parts.append(part)

        if not target_parts:
            QMessageBox.information(
                self,
                'No Matching Parts',
                f'No parts found in Imp BOM where Tool comments = {required_comment.title()}.',
            )
            return

        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            reply = QMessageBox.question(
                self, 'Existing Data',
                'This tab already contains data.\nDelete it and import fresh WU data?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        try:
            import sys as _sys
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from where_used_query import DISPLAY_HEADERS  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(self, 'Module Not Found',
                                f'where_used_query.py could not be imported:\n{exc}')
            return

        class _Worker(QObject):
            finished = pyqtSignal(list)
            error = pyqtSignal(str)

            def __init__(self, parts, level, plant, child_only):
                super().__init__()
                self._parts = parts
                self._level = level
                self._plant = plant
                self._child_only = child_only

            def run(self):
                try:
                    if self._child_only:
                        from where_used_query import fetch_where_used_level1_fast as _fwu_fast
                        result = _fwu_fast(self._parts, self._plant)
                    else:
                        from where_used_query import fetch_where_used_parents_only as _fwu
                        result = _fwu(self._parts, self._level, self._plant)
                    self.finished.emit(result)
                except Exception as exc:
                    self.error.emit(str(exc))

        progress = QProgressDialog(
            f'Querying Databricks for {len(target_parts)} part(s) where Tool comments = {required_comment.title()}\n'
            f'at WU level {max_level} and plant {selected_plant}...',
            None,
            0,
            0,
            self,
        )
        progress.setWindowTitle(flow_name)
        progress.setMinimumDuration(0)
        progress.setModal(True)
        progress.show()

        thread = QThread(self)
        worker = _Worker(target_parts, max_level, selected_plant, False)
        worker.moveToThread(thread)

        _display_headers = DISPLAY_HEADERS
        _obs_map = self._build_obs_change_map()

        def _on_error(msg):
            progress.close()
            thread.quit()
            QMessageBox.warning(self, 'Databricks Query Error', msg)

        def _on_finished(records):
            progress.close()
            thread.quit()

            if not records:
                QMessageBox.information(
                    self,
                    'No Data',
                    f'Databricks returned no records for {len(target_parts)} part(s) at WU level {max_level}.',
                )
                return

            _WU_COL = 1
            _PART_COL = 2
            _ORPHAN_COL = 3
            _REPL_COL = 4
            _DATA_START = 5

            wu_headers = list(_display_headers)
            wu_headers.insert(2, 'Orphan Child')
            all_headers = ['Select'] + wu_headers

            _DB_COL_FUNCS = [
                lambda r: r.get('rev_ln', ''),
                lambda r: r.get('plant', ''),
                lambda r: r.get('description', ''),
                lambda r: r.get('item_status', ''),
                lambda r: r.get('base_qty', ''),
                lambda r: r.get('ext_qty', ''),
                lambda r: r.get('uom', ''),
                lambda r: r.get('eco_number', ''),
                lambda r: r.get('procurement_type', ''),
                lambda r: r.get('effectivity_date', ''),
                lambda r: r.get('user_item_type', ''),
                lambda r: r.get('item_seq', ''),
                lambda r: r.get('kit_code', ''),
                lambda r: r.get('sparable_flag', ''),
                lambda r: r.get('designator', ''),
                lambda r: r.get('option_class', ''),
                lambda r: (r.get('pace_or_dash', '') if 'pace' in r.get('pace_or_dash', '').lower() else ''),
                lambda r: r.get('mlo_class', ''),
            ]

            self.table.setUpdatesEnabled(False)
            self.table.clear()
            self.table.setColumnCount(len(all_headers))
            self.table.setHorizontalHeaderLabels(all_headers)
            self.table.setRowCount(len(records))

            for row_idx, record in enumerate(records):
                chk = QCheckBox()
                cont = QWidget()
                h = QHBoxLayout(cont)
                h.setContentsMargins(0, 0, 0, 0)
                h.setAlignment(Qt.AlignmentFlag.AlignCenter)
                h.addWidget(chk)
                cont._chk = chk
                self.table.setCellWidget(row_idx, 0, cont)

                wu_val = str(record.get('wu_level', ''))
                self.table.setItem(row_idx, _WU_COL, QTableWidgetItem(wu_val))

                raw_part = record.get('part', '')
                try:
                    level_int = int(wu_val)
                except (ValueError, TypeError):
                    level_int = 0
                self.table.setItem(row_idx, _PART_COL,
                                   QTableWidgetItem(('      ' * level_int) + raw_part))

                self.table.setItem(row_idx, _ORPHAN_COL, QTableWidgetItem(''))

                replacement = _obs_map.get(raw_part.strip().upper(), '')
                rep_item = QTableWidgetItem(replacement)
                rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row_idx, _REPL_COL, rep_item)

                for c_off, fn in enumerate(_DB_COL_FUNCS):
                    self.table.setItem(row_idx, _DATA_START + c_off,
                                       QTableWidgetItem(fn(record)))

                if wu_val == '0':
                    _blue = QColor('#C7DEFA')
                    cont.setStyleSheet('background-color: #C7DEFA;')
                    for _col in range(1, len(all_headers)):
                        _item = self.table.item(row_idx, _col)
                        if _item is not None:
                            _item.setBackground(_blue)

            self.table.setUpdatesEnabled(True)

            hdr = self.table.horizontalHeader()
            # ResizeToContents is very expensive on huge result sets.
            if len(records) > 4000:
                for i in range(len(all_headers)):
                    if i == len(all_headers) - 1:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
                    else:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
                # Apply practical defaults for the first few key columns.
                if len(all_headers) > 0:
                    self.table.setColumnWidth(0, 65)   # Select
                if len(all_headers) > 1:
                    self.table.setColumnWidth(1, 85)   # WU Level
                if len(all_headers) > 2:
                    self.table.setColumnWidth(2, 260)  # Part
            else:
                for i in range(len(all_headers)):
                    if i == len(all_headers) - 1:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
                    else:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

            QMessageBox.information(
                self,
                'Import Complete',
                f'Imported {len(records)} row(s) from Databricks.\n'
                f'Parts queried from Imp BOM comments {required_comment.title()}: {len(target_parts)}\n'
                f'WU Level: {max_level} | Plant: {selected_plant}',
            )

        worker.finished.connect(_on_finished)
        worker.error.connect(_on_error)
        thread.started.connect(worker.run)
        thread.start()

    def import_from_databricks_for_removed_items(self):
        """Legacy flow: imports WU for rows marked as Removed Child Part in Imp BOM."""
        self._import_wu_by_tool_comment(
            required_comment='removed child part',
            flow_name='Import WU of Removed BOM items',
        )

    def import_removed_bom_wu_from_sap(self):
        """SAP flow: imports WU for rows marked as Removed BOM Item in Imp BOM.
        
        Asks user to select scope:
        - "Orphan Child alone": Add "Orphan Child" column, force WU level to 1
        - "Orphan Child and Parent": Add both columns, use selected WU levels
        """
        from PyQt6.QtCore import QThread, pyqtSignal, QObject
        from PyQt6.QtGui import QColor
        from PyQt6.QtWidgets import QProgressDialog

        # Validate WU Level input
        wu_text = self.wu_level_input.text().strip()
        if not wu_text:
            QMessageBox.warning(self, 'WU Level Required',
                                'Please enter a WU Level (1 to 6) before importing.')
            return
        try:
            selected_wu_level = int(wu_text)
        except ValueError:
            QMessageBox.warning(self, 'Invalid WU Level',
                                f'WU Level must be a number between 1 and 6, got: "{wu_text}"')
            return
        if selected_wu_level < 1 or selected_wu_level > 6:
            QMessageBox.warning(self, 'Invalid WU Level',
                                f'WU Level must be between 1 and 6, got: {selected_wu_level}')
            return

        selected_plant = self.plant_combo.currentText().strip() or '4070'

        # Check if Imp BOM tab is available
        if self.imp_bom_provider is None:
            QMessageBox.warning(self, 'Imp BOM Not Available',
                                'The Imp BOM tab is not connected. Cannot read removed parts.')
            return

        # Extract parts from Imp BOM where Tool Comments indicates removed BOM items.
        imp_table = self.imp_bom_provider.table
        part_col = self.imp_bom_provider.find_column('Part')
        tc_col = self.imp_bom_provider.find_column('Tool comments')

        if part_col < 0 or tc_col < 0:
            QMessageBox.warning(self, 'Imp BOM Missing Columns',
                                'Imp BOM does not contain both a Part and a Tool comments column.\n'
                                'Please import a BOM file first.')
            return

        valid_comments = {'removed bom item', 'removed bom items'}
        seen = set()
        target_parts = []
        for r in range(imp_table.rowCount()):
            tc_item = imp_table.item(r, tc_col)
            if not tc_item:
                continue
            comment = tc_item.text().strip().lower()
            if comment not in valid_comments:
                continue

            part_item = imp_table.item(r, part_col)
            if not part_item:
                continue
            part = part_item.text().strip()
            key = part.upper()
            if part and key not in seen:
                seen.add(key)
                target_parts.append(part)

        if not target_parts:
            QMessageBox.information(
                self,
                'No Matching Parts',
                'No parts found in Imp BOM where Tool comments is "Removed BOM Item" or "Removed BOM Items".',
            )
            return

        # Ask user to select scope with explicit button labels (avoids Yes/No confusion).
        scope_box = QMessageBox(self)
        scope_box.setIcon(QMessageBox.Icon.Question)
        scope_box.setWindowTitle('Select Orphan Analysis Scope')
        scope_box.setText(
            'Choose the scope for orphan identification:\n\n'
            f'• Orphan Child alone: Use your selected WU Level ({selected_wu_level})\n'
            '  (tracks Orphan Child only)\n\n'
            f'• Orphan Child and Parent: Use your selected WU Level ({selected_wu_level})\n'
            '  (tracks both columns, may take longer)'
        )
        btn_child_only = scope_box.addButton('Orphan Child alone', QMessageBox.ButtonRole.AcceptRole)
        btn_child_parent = scope_box.addButton('Orphan Child and Parent', QMessageBox.ButtonRole.ActionRole)
        scope_box.addButton(QMessageBox.StandardButton.Cancel)
        scope_box.setDefaultButton(btn_child_only)
        scope_box.exec()

        clicked = scope_box.clickedButton()
        if clicked is None or clicked not in (btn_child_only, btn_child_parent):
            return

        is_child_only = (clicked == btn_child_only)
        
        # Set WU level based on scope
        wu_level_for_query = selected_wu_level

        # Guardrail for large runs that are expected to take longer.
        if len(target_parts) >= 200 and wu_level_for_query >= 2:
            proceed = QMessageBox.question(
                self,
                'Large Query Warning',
                f'This run has {len(target_parts)} parts at WU level {wu_level_for_query}.\n'
                'It may take several minutes.\n\n'
                'For faster results, use WU level 1 when possible.\n\n'
                'Do you want to continue with the current scope?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if proceed != QMessageBox.StandardButton.Yes:
                return

        # Confirm clear existing data
        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            reply = QMessageBox.question(
                self, 'Existing Data',
                'This tab already contains data.\nDelete it and import fresh WU data?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        try:
            import sys as _sys
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from where_used_query import DISPLAY_HEADERS  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(self, 'Module Not Found',
                                f'where_used_query.py could not be imported:\n{exc}')
            return

        class _Worker(QObject):
            finished = pyqtSignal(list)
            error = pyqtSignal(str)

            def __init__(self, parts, level, plant, child_only):
                super().__init__()
                self._parts = parts
                self._level = level
                self._plant = plant
                self._child_only = child_only

            def run(self):
                try:
                    if self._child_only and self._level == 1:
                        from where_used_query import fetch_where_used_level1_fast as _fwu_fast
                        result = _fwu_fast(self._parts, self._plant)
                    else:
                        from where_used_query import fetch_where_used_parents_only as _fwu
                        result = _fwu(self._parts, self._level, self._plant)
                    self.finished.emit(result)
                except Exception as exc:
                    self.error.emit(str(exc))

        scope_label = "Orphan Child only" if is_child_only else "Orphan Child and Parent"
        import time as _time
        _start_ts = _time.perf_counter()
        progress = QProgressDialog(
            f'Querying Databricks for {len(target_parts)} part(s) (Scope: {scope_label})\n'
            f'WU level {wu_level_for_query} and plant {selected_plant}...',
            None,
            0,
            0,
            self,
        )
        progress.setWindowTitle('SAP Import WU of Removed BOM Item')
        progress.setMinimumDuration(0)
        progress.setModal(True)
        progress.show()

        thread = QThread(self)
        worker = _Worker(target_parts, wu_level_for_query, selected_plant, is_child_only)
        worker.moveToThread(thread)

        _display_headers = DISPLAY_HEADERS
        _obs_map = self._build_obs_change_map()
        _is_child_only = is_child_only

        def _on_error(msg):
            progress.close()
            thread.quit()
            elapsed = _time.perf_counter() - _start_ts
            QMessageBox.warning(self, 'Databricks Query Error',
                                f'{msg}\n\nElapsed: {elapsed:.1f} sec')

        def _on_finished(records):
            progress.close()
            thread.quit()

            if not records:
                QMessageBox.information(
                    self,
                    'No Data',
                    f'Databricks returned no records for {len(target_parts)} part(s) at WU level {wu_level_for_query}.',
                )
                return

            _WU_COL = 1
            _PART_COL = 2
            _ORPHAN_CHILD_COL = 3
            _ORPHAN_PARENT_COL = 4 if not _is_child_only else -1
            _REPL_COL = 5 if not _is_child_only else 4
            _DATA_START = 6 if not _is_child_only else 5

            # Build headers based on scope
            wu_headers = list(_display_headers)
            wu_headers.insert(2, 'Orphan Child')
            if not _is_child_only:
                wu_headers.insert(3, 'Orphan Parent')
            all_headers = ['Select'] + wu_headers

            _DB_COL_FUNCS = [
                lambda r: r.get('rev_ln', ''),
                lambda r: r.get('plant', ''),
                lambda r: r.get('description', ''),
                lambda r: r.get('item_status', ''),
                lambda r: r.get('base_qty', ''),
                lambda r: r.get('ext_qty', ''),
                lambda r: r.get('uom', ''),
                lambda r: r.get('eco_number', ''),
                lambda r: r.get('procurement_type', ''),
                lambda r: r.get('effectivity_date', ''),
                lambda r: r.get('user_item_type', ''),
                lambda r: r.get('item_seq', ''),
                lambda r: r.get('kit_code', ''),
                lambda r: r.get('sparable_flag', ''),
                lambda r: r.get('designator', ''),
                lambda r: r.get('option_class', ''),
                lambda r: (r.get('pace_or_dash', '') if 'pace' in r.get('pace_or_dash', '').lower() else ''),
                lambda r: r.get('mlo_class', ''),
            ]

            self.table.setUpdatesEnabled(False)
            self.table.clear()
            self.table.setColumnCount(len(all_headers))
            self.table.setHorizontalHeaderLabels(all_headers)
            self.table.setRowCount(len(records))

            for row_idx, record in enumerate(records):
                chk = QCheckBox()
                cont = QWidget()
                h = QHBoxLayout(cont)
                h.setContentsMargins(0, 0, 0, 0)
                h.setAlignment(Qt.AlignmentFlag.AlignCenter)
                h.addWidget(chk)
                cont._chk = chk
                self.table.setCellWidget(row_idx, 0, cont)

                wu_val = str(record.get('wu_level', ''))
                self.table.setItem(row_idx, _WU_COL, QTableWidgetItem(wu_val))

                raw_part = record.get('part', '')
                try:
                    level_int = int(wu_val)
                except (ValueError, TypeError):
                    level_int = 0
                self.table.setItem(row_idx, _PART_COL,
                                   QTableWidgetItem(('      ' * level_int) + raw_part))

                # Set Orphan Child column
                self.table.setItem(row_idx, _ORPHAN_CHILD_COL, QTableWidgetItem(''))

                # Set Orphan Parent column if not child-only scope
                if not _is_child_only:
                    self.table.setItem(row_idx, _ORPHAN_PARENT_COL, QTableWidgetItem(''))

                replacement = _obs_map.get(raw_part.strip().upper(), '')
                rep_item = QTableWidgetItem(replacement)
                rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row_idx, _REPL_COL, rep_item)

                for c_off, fn in enumerate(_DB_COL_FUNCS):
                    self.table.setItem(row_idx, _DATA_START + c_off,
                                       QTableWidgetItem(fn(record)))

                if wu_val == '0':
                    _blue = QColor('#C7DEFA')
                    cont.setStyleSheet('background-color: #C7DEFA;')
                    for _col in range(1, len(all_headers)):
                        _item = self.table.item(row_idx, _col)
                        if _item is not None:
                            _item.setBackground(_blue)

            self.table.setUpdatesEnabled(True)

            hdr = self.table.horizontalHeader()
            # ResizeToContents is very expensive on huge result sets.
            if len(records) > 4000:
                for i in range(len(all_headers)):
                    if i == len(all_headers) - 1:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
                    else:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
                if len(all_headers) > 0:
                    self.table.setColumnWidth(0, 65)   # Select
                if len(all_headers) > 1:
                    self.table.setColumnWidth(1, 85)   # WU Level
                if len(all_headers) > 2:
                    self.table.setColumnWidth(2, 260)  # Part
            else:
                for i in range(len(all_headers)):
                    if i == len(all_headers) - 1:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
                    else:
                        hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

            scope_msg = (
                f"Orphan Child only (WU Level {wu_level_for_query})"
                if _is_child_only
                else f"Orphan Child and Parent (WU Level {wu_level_for_query})"
            )
            elapsed = _time.perf_counter() - _start_ts
            QMessageBox.information(
                self,
                'Import Complete',
                f'Imported {len(records)} row(s) from Databricks.\n'
                f'Parts queried from Imp BOM: {len(target_parts)}\n'
                f'Scope: {scope_msg}\n'
                f'Plant: {selected_plant}\n'
                f'Elapsed: {elapsed:.1f} sec',
            )

        worker.finished.connect(_on_finished)
        worker.error.connect(_on_error)
        thread.started.connect(worker.run)
        thread.start()

    def _build_obs_change_map(self):
        mapping = {}
        t = self.obs_provider.table
        for r in range(t.rowCount()):
            obs_it = t.item(r, 1)
            chg_w = t.cellWidget(r, 2)
            key = obs_it.text().strip().upper() if obs_it else ''
            if key:
                mapping[key] = chg_w.currentText() if chg_w else ''
        return mapping

    def perform_orphan_analysis(self):
        # Find required columns
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphan Child')
        wucol = self._find_col('WU Level')

        if pcol < 0 or ocol < 0 or wucol < 0:
            QMessageBox.information(self, 'Orphan Analysis', 'Required columns not found.')
            return

        row_count = self.table.rowCount()

        def is_zero(v):
            return str(v).strip() in ('0', '0.0')

        def is_one(v):
            return str(v).strip() in ('1', '1.0')

        ignore_prefixes = ('0243', '0299', '0289', '0290')

        # Step 0: OBS → Obsolete / Inactivate mapping (existing behavior)
        obs_map = self._build_obs_change_map()
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if not it:
                continue
            key = it.text().lstrip().upper()
            if key in obs_map:
                self.table.setItem(r, ocol, QTableWidgetItem(obs_map[key]))

        # Step 1: Orphan1
        orphan_parts = set()
        r = 0
        while r < row_count:
            wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
            if is_zero(wu):
                child_row = r
                part_it = self.table.item(child_row, pcol)
                orphan_it = self.table.item(child_row, ocol)
                part = part_it.text().strip() if part_it else ''

                if orphan_it and orphan_it.text().strip():
                    r += 1
                    continue

                block_end = row_count
                for i in range(r + 1, row_count):
                    nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if is_zero(nxt):
                        block_end = i
                        break

                parents = []
                for i in range(child_row + 1, block_end):
                    wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if not is_one(wu_p):
                        continue
                    pit = self.table.item(i, pcol)
                    pval = pit.text().lstrip() if pit else ''
                    if pval.startswith(ignore_prefixes):
                        continue
                    parents.append(i)

                mark = False
                if not parents:
                    mark = True
                else:
                    all_bad = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        status = oit.text().strip().lower() if oit else ''
                        if status not in ('obsolete', 'inactivate'):
                            all_bad = False
                            break
                    if all_bad:
                        mark = True

                if mark and part:
                    self.table.setItem(child_row, ocol, QTableWidgetItem('Orphan1'))
                    orphan_parts.add(part.upper())

                r = block_end
            else:
                r += 1

        # Propagate Orphan1
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if it and it.text().strip().upper() in orphan_parts:
                oit = self.table.item(r, ocol)
                if not oit or not oit.text().strip():
                    self.table.setItem(r, ocol, QTableWidgetItem('Orphan1'))

        # Step 2+: Orphan2, Orphan3...
        current_level = 2
        while True:
            new_found = False
            new_parts = set()
            r = 0
            while r < row_count:
                wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
                orphan_it = self.table.item(r, ocol)
                if is_zero(wu) and (not orphan_it or not orphan_it.text().strip()):
                    part_it = self.table.item(r, pcol)
                    part = part_it.text().strip() if part_it else ''

                    block_end = row_count
                    for i in range(r + 1, row_count):
                        nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if is_zero(nxt):
                            block_end = i
                            break

                    parents = []
                    for i in range(r + 1, block_end):
                        wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if not is_one(wu_p):
                            continue
                        pit = self.table.item(i, pcol)
                        pval = pit.text().lstrip() if pit else ''
                        if pval.startswith(ignore_prefixes):
                            continue
                        parents.append(i)

                    parents_filled = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        if not oit or not oit.text().strip():
                            parents_filled = False
                            break

                    if part and (not parents or parents_filled):
                        self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))
                        new_parts.add(part.upper())
                        new_found = True
                r += 1

            for r in range(row_count):
                it = self.table.item(r, pcol)
                oit = self.table.item(r, ocol)
                if it and it.text().strip().upper() in new_parts and (not oit or not oit.text().strip()):
                    self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))

            if not new_found:
                break
            current_level += 1

            self.append_orphans_to_obs_parts()
        from PyQt6.QtGui import QColor
        for r in range(self.table.rowCount()):
            wu = self.table.item(r, wucol)
            if wu and wu.text().strip() == '0':
                for c in range(self.table.columnCount()):
                    it = self.table.item(r, c)
                    if it:
                        it.setBackground(QColor('#87CEEB'))


    def remove_by_prefix(self, prefix):
        pcol = self._find_col('Part')
        if pcol < 0:
            return
        rows = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, pcol)
            if it and it.text().lstrip().startswith(prefix):
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)


class  OBSAllPartsWithoutReplacementTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)

        subtabs = QTabWidget()
        # Equal width secondary tabs
        subtabs.setStyleSheet("""
        QTabBar::tab {
            min-width: 220px;
            padding: 6px 12px;
            text-align: center;
        }
        """)

        self.imp_bom_tab = OrphanOBSSubTab()
        self.wu_removed_tab = WURemovedBOMItemsTab(self.obs_provider, imp_bom_provider=self.imp_bom_tab)
        subtabs.addTab(self.imp_bom_tab, "Imp BOM")
        subtabs.addTab(self.wu_removed_tab, "WU of Removed BOM Items")

        outer.addWidget(subtabs)

class _OrphanAnalysisTab_OLD(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        outer = QVBoxLayout(self)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        self.subtabs = QTabWidget()
        self.subtabs.setStyleSheet("""
        QTabBar::tab:selected {
            background-color: #87CEEB;
            color: #0F2D46;
            font-weight: 600;
        }
        QTabBar::tab {
            background-color: #EAF6FD;
        }
        """)
        self.subtabs.addTab(
            OBSAllPartsWithoutReplacementTab(obs_provider),
            "OBS all Parts without Replacements"
        )
        self.subtabs.addTab(
            PlaceholderTab("OBS with or without Replacement"),
            "OBS with or without Replacement"
        )

        outer.addWidget(self.subtabs)




# === EC Creation Inputs Form (Embedded) ===

EC_CATEGORY_DESC = {
    "A1": "SMBoM Options as revised items and having CDW",
    "A2": "SMBoM Options as revised items, and No CDW",
    "B1": "No SMBoM Options as revised items and having CDW",
    "B2": "No SMBoM Options as revised items, No CDW, and revised item status at Eval and/or moving to Eval or adding already released parts to Proto buckets",
    "B3": "No SMBoM Option as revised items, No CDW, and revised Item status at Production and/or moving to Production",
}


def header(text):
    lbl = QLabel(text)
    lbl.setFont(QFont("Segoe UI", 12, QFont.Weight.DemiBold))
    return lbl


def highlight_label(text):
    lbl = QLabel(text)
    lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
    lbl.setStyleSheet("color:#1F4E79")
    return lbl


class LimitedTextEdit(QTextEdit):
    def __init__(self, limit=2000):
        super().__init__()
        self.limit = limit
        self.textChanged.connect(self._limit)

    def _limit(self):
        text = self.toPlainText()
        if len(text) > self.limit:
            self.blockSignals(True)
            self.setPlainText(text[:self.limit])
            self.moveCursor(QTextCursor.MoveOperation.End)
            self.blockSignals(False)

    def _update_counter(self):
        if not self.counter_label:
            return
        count = len(self.toPlainText())
        self.counter_label.setText(f"{count} / {self.limit} characters")
        if count >= self.limit:
            self.counter_label.setStyleSheet("color:red;font-size:10px")
        else:
            self.counter_label.setStyleSheet("color:gray;font-size:10px")

class ECCreationInputsFormTab(QWidget):
    def __init__(self):
        super().__init__()
        self.email_file_path = None  # Stores the path of the browsed email file
        root = QVBoxLayout(self)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        root.addWidget(scroll)
        container = QWidget(); scroll.setWidget(container)
        outer = QVBoxLayout(container)

        # ---- Section A ----
        outer.addWidget(header("Section A: EC Category Form"))
        secA = QFrame(); a = QVBoxLayout(secA)

        scope_row = QHBoxLayout(); scope_grp = QButtonGroup(self)
        for scope in ["Up Revision", "Status Roll", "Production Release", "OBS / Inactivate", "Product Release"]:
            rb = QRadioButton(scope)
            scope_grp.addButton(rb); rb.toggled.connect(self.start_flow)
            scope_row.addWidget(rb)
        a.addLayout(scope_row)

        self.flow_area = QVBoxLayout(); a.addLayout(self.flow_area)
        self.ec_result_lbl = QLabel(""); self.ec_result_lbl.setWordWrap(True)
        self.ec_result_lbl.setStyleSheet("color:green;font-weight:600;font-size:14px")
        # Hidden – category is shown in the 'EC Category based on Selected Criteria' field below

        # ── AI Proposed EC Category row ───────────────────────────────────────
        ai_ec_row = QHBoxLayout()
        ai_ec_row.setSpacing(8)
        ai_ec_prop_lbl = QLabel("EC Category based on Selected Criteria:")
        ai_ec_prop_lbl.setStyleSheet("color:#1F4E79; font-weight:600;")
        ai_ec_row.addWidget(ai_ec_prop_lbl)

        self.ai_ec_category_edit = QLineEdit()
        self.ai_ec_category_edit.setReadOnly(True)
        self.ai_ec_category_edit.setPlaceholderText("")
        self.ai_ec_category_edit.setFixedHeight(26)
        self.ai_ec_category_edit.setMinimumWidth(60)
        self.ai_ec_category_edit.setMaximumWidth(120)
        self.ai_ec_category_edit.setStyleSheet(
            "background:#E8F5E9; border:1px solid #81C784; border-radius:4px; "
            "color:#1B5E20; font-weight:bold; padding:2px 6px;"
        )
        ai_ec_row.addWidget(self.ai_ec_category_edit)

        self.ai_ec_justification_btn = QPushButton("\u24d8")
        self.ai_ec_justification_btn.setFixedSize(22, 22)
        self.ai_ec_justification_btn.setStyleSheet(
            "QPushButton { background:#1F4E79; color:white; border-radius:11px; "
            "font-weight:bold; font-size:12px; border:none; }"
            "QPushButton:hover { background:#2E75B6; }"
        )
        self.ai_ec_justification_btn.setToolTip("Run Problem Summary to see EC Category justification")
        self.ai_ec_justification_btn.setCursor(Qt.CursorShape.WhatsThisCursor)
        self.ai_ec_justification_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        ai_ec_row.addWidget(self.ai_ec_justification_btn)
        ai_ec_row.addStretch(1)
        a.addLayout(ai_ec_row)
        # ───────────────────────────────────────────────────────────────────────

        self.ec_divider = QFrame(); self.ec_divider.setFrameShape(QFrame.Shape.HLine)
        self.ec_divider.setVisible(False); a.addWidget(self.ec_divider)
        outer.addWidget(secA)
        outer.addSpacing(12)

# ---- Section B ----
        self.secB_header = header("Section B: ECR Change Assessment")
        self.secB_header.setVisible(False)
        outer.addWidget(self.secB_header)

        self.secB = QFrame()
        self.secB.setVisible(False)
        outer.addWidget(self.secB)

        # === Main horizontal layout for Section B ===
        section_b_layout = QHBoxLayout(self.secB)
        section_b_left = QVBoxLayout()
        section_b_right = QVBoxLayout()

    #    setContentsMargins(left, top, right, bottom)
        BOX_WIDTH = 240          # width of the box
        BOX_HEIGHT = 300         # height of the box
        BOX_OFFSET_FROM_RIGHT = 500   # 👈 INCREASE → moves box LEFT


        section_b_layout.addLayout(section_b_left, 3)   # Left: questions
        section_b_layout.addLayout(section_b_right, 1)  # Right: PN box


        # Reference fields
        self.ref_boxes = {}
        self.ref_radios = {}
        self.ref_labels = {}

        def yes_no_with_box(key, label_text):
                row = QHBoxLayout()

                lbl = QLabel(label_text)
                row.addWidget(lbl)

                grp = QButtonGroup(self)
                rb_y = QRadioButton("Yes")
                rb_n = QRadioButton("No")
                grp.addButton(rb_y)
                grp.addButton(rb_n)

                row.addWidget(rb_y)
                row.addWidget(rb_n)

                txt = QTextEdit()
                txt.setFixedHeight(30)
                txt.setMaximumWidth(500)
                txt.setStyleSheet("background:#FFF2CC; border:none")
                txt.setVisible(False)

                row.addWidget(txt)
                row.addStretch(1)

                rb_y.toggled.connect(lambda c: txt.setVisible(c))
                rb_n.toggled.connect(lambda c: txt.setVisible(False))

                self.ref_boxes[key] = txt
                self.ref_radios[key] = (rb_y, rb_n)
                self.ref_labels[key] = lbl

                section_b_left.addLayout(row)


        yes_no_with_box("PCR_PCN", "1.  Does this Project include Product Change Request (PCR)")
        yes_no_with_box("PSN", "             Is there any Associated Product Safety Note (PSN)")
        yes_no_with_box("PCR", "2.  Associated Project created for PCR")
        yes_no_with_box("SPS", "3.  Associated Open SPSs")
        yes_no_with_box("ESW", "4.  Associated ESWs")
        yes_no_with_box("REF_ECR", "5.  Reference ECR Number(s)")


        # PCR → PSN dependency
        pcr_yes, pcr_no = self.ref_radios["PCR_PCN"]
        psn_yes, psn_no = self.ref_radios["PSN"]
        psn_txt = self.ref_boxes["PSN"]
        psn_lbl = self.ref_labels["PSN"]

        def disable_psn():
                # Temporarily disable auto-exclusive to allow both radios to be unchecked
                grp_psn = psn_yes.group()
                if grp_psn:
                    grp_psn.setExclusive(False)
                
                # Block signals and uncheck both
                psn_yes.blockSignals(True)
                psn_no.blockSignals(True)
                psn_yes.setChecked(False)
                psn_no.setChecked(False)
                psn_yes.blockSignals(False)
                psn_no.blockSignals(False)
                
                # Re-enable auto-exclusive
                if grp_psn:
                    grp_psn.setExclusive(True)
                
                # Disable and hide
                psn_yes.setEnabled(False)
                psn_no.setEnabled(False)
                psn_lbl.setEnabled(False)
                psn_txt.setPlainText("")
                psn_txt.setVisible(False)

        def enable_psn():
                psn_yes.setEnabled(True)
                psn_no.setEnabled(True)
                psn_lbl.setEnabled(True)

        pcr_no.toggled.connect(lambda c: disable_psn() if c else None)
        pcr_yes.toggled.connect(lambda c: enable_psn() if c else None)


        # Reference Email Attachments
        rowm = QHBoxLayout()
        rowm.addWidget(QLabel("6.  Reference e-mail/attachments?"))

        grp_m = QButtonGroup(self)
        rb_my = QRadioButton("Yes")
        rb_mn = QRadioButton("No")
        grp_m.addButton(rb_my)
        grp_m.addButton(rb_mn)
        self.ref_attachment_yes_radio = rb_my
        self.ref_attachment_no_radio = rb_mn

        rowm.addWidget(rb_my)
        rowm.addWidget(rb_mn)

        browse = QPushButton("Browse Attachment")
        browse.setFixedHeight(24)
        browse.setVisible(False)

        rowm.addWidget(browse)
        rowm.addStretch(1)
        section_b_left.addLayout(rowm)

        rb_my.toggled.connect(self._on_reference_attachments_toggled)
        browse.clicked.connect(self._on_browse_email_clicked)
        self._browse_email_btn = browse  # keep reference to update button label


        # Impact caused by
        row = QHBoxLayout()
        row.addWidget(QLabel("7.  Impact caused by:"))

        grp_sc = QButtonGroup(self)
        rb_sup = QRadioButton("Supplier")
        rb_cust = QRadioButton("Customer")
        grp_sc.addButton(rb_sup)
        grp_sc.addButton(rb_cust)
        self.impact_supplier_radio = rb_sup
        self.impact_customer_radio = rb_cust

        row.addWidget(rb_sup)
        row.addWidget(rb_cust)
        row.addStretch(1)
        section_b_left.addLayout(row)

        sub_row = QHBoxLayout()
        sub_row.setContentsMargins(20, 0, 0, 0)

        grp_c = QButtonGroup(self)
        rb_int = QRadioButton("Internal")
        rb_ext = QRadioButton("External")
        grp_c.addButton(rb_int)
        grp_c.addButton(rb_ext)
        self.impact_internal_radio = rb_int
        self.impact_external_radio = rb_ext

        sub_row.addWidget(rb_int)
        sub_row.addWidget(rb_ext)
        sub_row.addStretch(1)
        section_b_left.addLayout(sub_row)

        def clear_customer_scope_selection():
            # Temporarily disable exclusivity so both radios can be fully unchecked.
            grp_c.setExclusive(False)
            rb_int.setChecked(False)
            rb_ext.setChecked(False)
            grp_c.setExclusive(True)

        rb_sup.toggled.connect(
            lambda c: (
                clear_customer_scope_selection(),
                rb_int.setEnabled(False),
                rb_ext.setEnabled(False)
            ) if c else None
        )

        rb_cust.toggled.connect(
            lambda c: (
                rb_int.setEnabled(c),
                rb_ext.setEnabled(c)
            )
        )


        yes_no_with_box("QN", "8. Is there any QN")

        # Reason Code
        rc_row = QHBoxLayout()
        rc_row.setSpacing(8)
        rc_row.addWidget(QLabel("9.  ECR Reason Code:"))
        self.reason_cb = QComboBox()
        self.reason_cb.addItems([
            "Beyond Spec Request","Cap Code Management","CES","Cost Reduction",
            "Design Correction","Document Correction","Manufacturing Improvement",
            "Obsolescence","Option Reduction and Product End of Life","Order BOM Change",
            "Product Improvement","Product Release","Safety Event"
        ])
        rc_row.addWidget(self.reason_cb)

        # Proposed Reason Code – read-only label next to the dropdown
        ai_prop_lbl = QLabel("Proposed Reason Code:")
        ai_prop_lbl.setStyleSheet("color:#1F4E79; font-weight:600;")
        rc_row.addWidget(ai_prop_lbl)

        self.ai_reason_code_edit = QLineEdit()
        self.ai_reason_code_edit.setReadOnly(True)
        self.ai_reason_code_edit.setPlaceholderText("")
        self.ai_reason_code_edit.setFixedHeight(26)
        self.ai_reason_code_edit.setMinimumWidth(180)
        self.ai_reason_code_edit.setStyleSheet(
            "background:#E8F5E9; border:1px solid #81C784; border-radius:4px; "
            "color:#1B5E20; font-weight:bold; padding:2px 6px;"
        )
        rc_row.addWidget(self.ai_reason_code_edit)

        # ⓘ info button – opens Scenario & Examples dialog for the proposed Reason Code
        self.ai_justification_btn = QPushButton("\u24d8")
        self.ai_justification_btn.setFixedSize(22, 22)
        self.ai_justification_btn.setStyleSheet(
            "QPushButton { background:#1F4E79; color:white; border-radius:11px; "
            "font-weight:bold; font-size:12px; border:none; }"
            "QPushButton:hover { background:#2E75B6; }"
        )
        self.ai_justification_btn.setToolTip("Click to see Scenario and Examples for the proposed Reason Code")
        self.ai_justification_btn.setCursor(Qt.CursorShape.WhatsThisCursor)
        self.ai_justification_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.ai_justification_btn.clicked.connect(self._show_rc_info_dialog)
        rc_row.addWidget(self.ai_justification_btn)

        rc_row.addStretch(1)
        section_b_left.addLayout(rc_row)


        scope_lbl = QLabel("Initial Scope Part Numbers")
        scope_lbl.setStyleSheet("font-weight: bold;")

        self.scope_parts_txt = QTextEdit()
        self.scope_parts_txt.setPlaceholderText(
                "Enter Part Number (up to 500 lines)"
        )

        # --- Size tuning ---
        self.scope_parts_txt.setFixedWidth(235)      # ~10% narrower
        self.scope_parts_txt.setMaximumHeight(280)   # slightly taller

        # --- Scrollbars ---
        self.scope_parts_txt.setVerticalScrollBarPolicy(
                Qt.ScrollBarPolicy.ScrollBarAsNeeded
        )
        self.scope_parts_txt.setHorizontalScrollBarPolicy(
                Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )

        # --- Styling (box + scrollbar) ---
        self.scope_parts_txt.setStyleSheet("""
                QTextEdit {
                        background: #F8F8F8;
                        border: 1px solid #999;
                }
                QScrollBar:vertical {
                        background: #E6E6E6;
                        width: 10px;
                        margin: 0px;
                }
                QScrollBar::handle:vertical {
                        background: #A6A6A6;
                        min-height: 20px;
                        border-radius: 4px;
                }
                QScrollBar::handle:vertical:hover {
                        background: #8C8C8C;
                }
                QScrollBar::add-line:vertical,
                QScrollBar::sub-line:vertical {
                        height: 0px;
                }
        """)

        
        section_b_right.setContentsMargins(0, 0, 300, 0)   # 👈 right margin pushes box left
        section_b_right.addWidget(scope_lbl)
        section_b_right.addWidget(self.scope_parts_txt)
        section_b_right.addSpacing(10)

        include_tabs_lbl = QLabel("Include Data From Tabs")
        include_tabs_lbl.setStyleSheet("font-weight: bold;")
        section_b_right.addWidget(include_tabs_lbl)

        self.cb_where_used = QCheckBox("Where Used")
        self.cb_obs_parts = QCheckBox("OBS Parts")
        self.cb_structure_sheet = QCheckBox("Structure Sheet")

        section_b_right.addWidget(self.cb_where_used)
        section_b_right.addWidget(self.cb_obs_parts)
        section_b_right.addWidget(self.cb_structure_sheet)

        self.selected_tab_payload = {}

        self.cb_where_used.stateChanged.connect(self.refresh_selected_tab_payload)
        self.cb_obs_parts.stateChanged.connect(self.refresh_selected_tab_payload)
        self.cb_structure_sheet.stateChanged.connect(self.refresh_selected_tab_payload)

        section_b_right.addStretch(1)
       
    # Short Title
        title_row = QHBoxLayout()
        title_row.addWidget(highlight_label("Short Title"))
        title_row.addStretch(1)
        section_b_left.addLayout(title_row)

        self.short_title_edit = QLineEdit()
        self.short_title_edit.setPlaceholderText("Enter short title")
        self.short_title_edit.setFixedHeight(30)
        self.short_title_edit.setMaximumWidth(900)
        section_b_left.addWidget(self.short_title_edit)

        section_b_left.addSpacing(6)

# Problem Summary
        ps_row = QHBoxLayout()
        ps_row.addWidget(
                highlight_label("Problem Summary from PCR, PCN, SPS and ESW")
        )
        self.btn_ps = QPushButton("Problem Summary")
        self.btn_ps.setFixedSize(180, 26)
        ps_row.addWidget(self.btn_ps)
        ps_row.addStretch(1)

        section_b_left.addLayout(ps_row)

        self.problem_txt = LimitedTextEdit(2000)
        self.problem_txt.setMinimumHeight(260)
        self.problem_txt.setMaximumWidth(900)
        section_b_left.addWidget(self.problem_txt)

        self.btn_ps.clicked.connect(self.on_problem_summary_clicked)

        section_b_left.addSpacing(6)




        # Solution
        sol_row = QHBoxLayout()
        sol_row.addWidget(highlight_label("Proposed Solution"))
        self.btn_sol = QPushButton("Regenerate Proposed Solution")
        self.btn_sol.setFixedSize(220, 26)
        self.btn_sol.setToolTip(
            "Clears the existing solution and regenerates a new solution using AI."
        )
        sol_row.addWidget(self.btn_sol)
        sol_row.addStretch(1)

        section_b_left.addLayout(sol_row)

        self.solution_txt = LimitedTextEdit(2000)
        self.solution_txt.setMinimumHeight(260)
        self.solution_txt.setMaximumWidth(900)

        section_b_left.addWidget(self.solution_txt)

        self.btn_sol.clicked.connect(self.on_regenerate_solution_clicked)

        outer.addWidget(self.secB)
        outer.addStretch(1)


    def refresh_selected_tab_payload(self):
        main = self.window()
        payload = {}

        if self.cb_where_used.isChecked():
            whereused_tab = getattr(main, "whereused_tab", None)
            if whereused_tab and hasattr(whereused_tab, "table"):
                payload["Where Used"] = self._table_to_rows(whereused_tab.table)

        if self.cb_obs_parts.isChecked():
            obs_tab = getattr(main, "obs_tab", None)
            if obs_tab and hasattr(obs_tab, "table"):
                payload["OBS Parts"] = self._table_to_rows(obs_tab.table)

        if self.cb_structure_sheet.isChecked():
            structure_tab = getattr(main, "structure_tab", None)
            if structure_tab:
                structure_data = {}

                impacted_text = getattr(structure_tab, "impacted_text", None)
                if impacted_text:
                    structure_data["Impacted Options/Parts"] = [
                        line.strip()
                        for line in impacted_text.toPlainText().splitlines()
                        if line.strip()
                    ]

                structure_table = getattr(structure_tab, "table", None)
                if structure_table:
                    structure_data["Structure Sheet"] = self._table_to_rows(structure_table)

                payload["Structure Sheet"] = structure_data

        self.selected_tab_payload = payload

    def get_selected_tab_payload(self):
        self.refresh_selected_tab_payload()
        return self.selected_tab_payload

    def _table_to_rows(self, table):
        if table is None or table.columnCount() == 0:
            return []

        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text().strip() if h and h.text() else f"Column {c + 1}")

        rows = []
        for r in range(table.rowCount()):
            row_data = {}
            has_value = False

            for c, header in enumerate(headers):
                value = ""
                cell_widget = table.cellWidget(r, c)

                if isinstance(cell_widget, QComboBox):
                    value = cell_widget.currentText().strip()
                elif isinstance(cell_widget, QTextEdit):
                    value = cell_widget.toPlainText().strip()
                elif isinstance(cell_widget, QWidget):
                    chk = cell_widget.findChild(QCheckBox)
                    if chk:
                        value = "Yes" if chk.isChecked() else "No"
                    else:
                        item = table.item(r, c)
                        value = item.text().strip() if item and item.text() else ""
                else:
                    item = table.item(r, c)
                    value = item.text().strip() if item and item.text() else ""

                if value:
                    has_value = True
                row_data[header] = value

            if has_value:
                rows.append(row_data)

        return rows
# ...existing code...

    def _selected_radio_text(self, key: str) -> str:
        pair = self.ref_radios.get(key)
        if not pair:
            return ""
        rb_y, rb_n = pair
        if rb_y.isChecked():
            return "Yes"
        if rb_n.isChecked():
            return "No"
        return ""

# ...existing code...
    def _collect_pss_payload(self) -> Dict[str, Any]:
        self.refresh_selected_tab_payload()

        scope_parts = [
            ln.strip()
            for ln in self.scope_parts_txt.toPlainText().splitlines()
            if ln.strip()
        ]

        reference_inputs = {}
        for key, txt in self.ref_boxes.items():
            reference_inputs[key] = {
                "answer": self._selected_radio_text(key),   # "Yes" / "No" / ""
                "text": txt.toPlainText().strip(),
            }

        return {
            "short_title": self.short_title_edit.text().strip(),
            "reason_code": self.reason_cb.currentText().strip(),
            "scope_parts": scope_parts,
            "reference_inputs": reference_inputs,
            "selected_tab_payload": self.selected_tab_payload,
            "current_problem_text": self.problem_txt.toPlainText().strip(),
            "proposed_solution_text": self.solution_txt.toPlainText().strip(),
            "solution_regeneration_requested": bool(
                getattr(self, "_solution_regen_requested", False)
            ),
            "solution_regeneration_reason": (
                "Existing solution is incomplete or incorrect. "
                "Regenerate a complete, engineer-quality solution."
                if getattr(self, "_solution_regen_requested", False)
                else ""
            ),
            "include_tabs_flags": {
                "where_used": self.cb_where_used.isChecked(),
                "obs_parts": self.cb_obs_parts.isChecked(),
                "structure_sheet": self.cb_structure_sheet.isChecked(),
            },
        }
# ...existing code...

    def _load_ai_pss_module(self):
        mod_path = Path(__file__).with_name("AI_Assisted_PSS.py")
        if not mod_path.exists():
            raise FileNotFoundError(f"AI_Assisted_PSS.py not found: {mod_path}")

        spec = importlib.util.spec_from_file_location("AI_Assisted_PSS", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load AI_Assisted_PSS module spec")

        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    # ...existing code...

    def _normalize_pss_output(self, out: Any) -> Dict[str, str]:
        if isinstance(out, dict):
            title = str(out.get("title") or "").strip()
            problem = str(
                out.get("problem_statement")
                or out.get("problem_summary")
                or out.get("problem")
                or ""
            ).strip()
            solution = str(
                out.get("solution_statement")
                or out.get("solution")
                or ""
            ).strip()
            return {"title": title, "problem": problem, "solution": solution}

        text = str(out or "").strip()
        return {"title": "", "problem": text, "solution": ""}

    def _run_ai_pss_full(self, payload: Dict[str, Any]) -> Dict[str, str]:
        module = self._load_ai_pss_module()

        fn_full = getattr(module, "generate_full_pss", None)
        if callable(fn_full):
            try:
                out = fn_full(payload)
            except TypeError:
                out = fn_full()
            return self._normalize_pss_output(out)

        for fn_name in ("generate_problem_summary", "generate_pss", "run_problem_summary"):
            fn = getattr(module, fn_name, None)
            if callable(fn):
                try:
                    out = fn(payload)
                except TypeError:
                    out = fn()
                return self._normalize_pss_output(out)

        raise AttributeError(
            "No supported entry point in AI_Assisted_PSS.py. "
            "Expected generate_full_pss(payload) or generate_problem_summary(payload)."
        )

# ...existing code...

# ...existing code...

    def _run_ai_pss(self, payload: Dict[str, Any]) -> str:
        out = self._run_ai_pss_full(payload)
        return out.get("problem", "")

# ...existing code...

        for fn_name in ("generate_problem_summary", "generate_pss", "run_problem_summary"):
            fn = getattr(module, fn_name, None)
            if callable(fn):
                try:
                    out = fn(payload)
                except TypeError:
                    out = fn()

                if isinstance(out, dict):
                    out = out.get("problem_summary") or out.get("problem") or ""
                return str(out or "").strip()

        for cls_name in ("AIAssistedPSS", "ProblemSolutionAgent", "PSSAgent"):
            cls = getattr(module, cls_name, None)
            if cls is None:
                continue
            obj = cls()
            for m_name in ("generate_problem_summary", "generate_pss", "run"):
                method = getattr(obj, m_name, None)
                if callable(method):
                    try:
                        out = method(payload)
                    except TypeError:
                        out = method()

                    if isinstance(out, dict):
                        out = out.get("problem_summary") or out.get("problem") or ""
                    return str(out or "").strip()

        raise AttributeError(
            "No supported entry point in AI_Assisted_PSS.py. "
            "Expected generate_problem_summary(payload) or equivalent."
        )

# ...existing code...

    def _on_reference_attachments_toggled(self, checked: bool):
        """Show/hide browse button and clear selected attachment when toggled off."""
        if hasattr(self, "_browse_email_btn"):
            self._browse_email_btn.setVisible(checked)

        if not checked:
            self.email_file_path = None
            if hasattr(self, "_browse_email_btn"):
                self._browse_email_btn.setText("Browse Attachment")
                self._browse_email_btn.setToolTip("")

    def _on_browse_email_clicked(self):
        """Open file dialog, store selected attachment path, update button label."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Reference Attachment",
            "",
            "Supported Files (*.msg *.eml *.txt *.pdf *.ppt *.pptx *.doc *.docx)"
        )
        if file_path:
            self.email_file_path = file_path
            file_name = Path(file_path).name
            self._browse_email_btn.setText(f"📎 {file_name}")
            self._browse_email_btn.setToolTip(file_path)
            QMessageBox.information(
                self,
                "Attachment Loaded",
                f"Reference attachment loaded:\n{file_name}\n\nClick 'Problem Summary' to process it.",
            )

    def _validate_problem_summary_inputs(self) -> str:
        """Validate ECR change assessment inputs before generating Problem Summary."""
        missing_answers = []
        missing_values = []
        qn_missing_answer = None
        qn_missing_value = None
        pcr_pair = self.ref_radios.get("PCR_PCN")
        pcr_yes_selected = bool(pcr_pair and pcr_pair[0].isChecked())
        field_names = {
            "PCR_PCN": "PCR",
            "PSN": "PSN",
            "PCR": "PCR Project",
            "SPS": "SPS",
            "ESW": "ESW",
            "REF_ECR": "Reference ECR",
            "QN": "QN",
        }

        for key, (rb_yes, rb_no) in self.ref_radios.items():
            # PSN is auto-derived from PCR records when Question 1 (PCR_PCN) is Yes.
            # Do not require manual PSN Yes/No or value in that flow.
            if key == "PSN" and pcr_yes_selected:
                continue

            # Skip questions that are currently disabled by dependency logic (for example PSN).
            if not rb_yes.isEnabled() and not rb_no.isEnabled():
                continue

            if not rb_yes.isChecked() and not rb_no.isChecked():
                label = self.ref_labels.get(key)
                q_text = label.text() if label else key
                if key == "QN":
                    qn_missing_answer = q_text
                else:
                    missing_answers.append(q_text)
                continue

            if rb_yes.isChecked():
                value_widget = self.ref_boxes.get(key)
                value_text = value_widget.toPlainText().strip() if value_widget else ""
                if not value_text:
                    if key == "QN":
                        qn_missing_value = field_names.get(key, key)
                    else:
                        missing_values.append(field_names.get(key, key))

        attachment_yes = getattr(self, "ref_attachment_yes_radio", None)
        attachment_no = getattr(self, "ref_attachment_no_radio", None)
        if attachment_yes and attachment_no:
            if not attachment_yes.isChecked() and not attachment_no.isChecked():
                missing_answers.append(
                    "6. Reference e-mail/attachments"
                )
            elif attachment_yes.isChecked() and not getattr(self, "email_file_path", None):
                missing_answers.append(
                    "6. Reference e-mail/attachments: please attach relevant attachments"
                )

        impact_supplier = getattr(self, "impact_supplier_radio", None)
        impact_customer = getattr(self, "impact_customer_radio", None)
        impact_internal = getattr(self, "impact_internal_radio", None)
        impact_external = getattr(self, "impact_external_radio", None)

        if impact_supplier and impact_customer:
            if not impact_supplier.isChecked() and not impact_customer.isChecked():
                missing_answers.append("7. Impact caused by: select Supplier or Customer")
            elif impact_customer.isChecked() and impact_internal and impact_external:
                if not impact_internal.isChecked() and not impact_external.isChecked():
                    missing_answers.append(
                        "7. Impact caused by Customer: select Internal or External"
                    )

        if qn_missing_answer:
            missing_answers.append(qn_missing_answer)

        if qn_missing_value:
            missing_values.append(qn_missing_value)

        if missing_answers:
            return (
                "Please address Assessment question(s) with Yes/No selection:\n\n"
                + "\n".join(f"{q}" for q in missing_answers)
            )

        if missing_values:
            return (
                "Please provide details in the corresponding field for these 'Yes' selections:\n\n"
                + "\n".join(f"{name}: details/number is missing" for name in missing_values)
            )

        return ""

    # ------------------------------------------------------------------
    # PCR-driven Problem Summary helpers (Question 1 – Databricks path)
    # ------------------------------------------------------------------

    def _load_pcr_query_module(self):
        """Load pcr_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "pcr_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"pcr_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("pcr_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load pcr_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_project_query_module(self):
        """Load project_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "project_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"project_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("project_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load project_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_sps_query_module(self):
        """Load sps_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "sps_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"sps_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("sps_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load sps_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_esw_query_module(self):
        """Load esw_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "esw_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"esw_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("esw_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load esw_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_ecr_query_module(self):
        """Load ecr_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "ecr_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"ecr_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("ecr_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load ecr_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _fetch_q2_project_context(self) -> Dict[str, Any]:
        """Fetch Question 2 (Associated Project created for PCR) records if enabled.

        Returns empty structure when Q2 is not selected as Yes.
        """
        pair = self.ref_radios.get("PCR")
        if not pair:
            self._project_fetch_result = {"skipped": [], "not_found": []}
            return {"project_records": [], "skipped_projects": [], "not_found_projects": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._project_fetch_result = {"skipped": [], "not_found": []}
            return {"project_records": [], "skipped_projects": [], "not_found_projects": []}

        project_text = self.ref_boxes["PCR"].toPlainText().strip()
        if not project_text:
            raise ValueError(
                "Please enter at least one Project number in Question 2 "
                "(Associated Project created for PCR)."
            )

        mod = self._load_project_query_module()
        result = mod.fetch_project_records(project_text)

        project_records = result.get("valid", [])
        skipped_projects = result.get("skipped", [])
        not_found_projects = result.get("not_found", [])

        self._project_fetch_result = {
            "skipped": skipped_projects,
            "not_found": not_found_projects,
        }

        return {
            "project_records": project_records,
            "skipped_projects": skipped_projects,
            "not_found_projects": not_found_projects,
        }

    def _fetch_q3_sps_context(self) -> Dict[str, Any]:
        """Fetch Question 3 (Associated Open SPSs) records if enabled.

        Returns empty structure when Q3 is not selected as Yes.
        """
        pair = self.ref_radios.get("SPS")
        if not pair:
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            return {"sps_records": [], "skipped_sps": [], "not_found_sps": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            return {"sps_records": [], "skipped_sps": [], "not_found_sps": []}

        sps_text = self.ref_boxes["SPS"].toPlainText().strip()
        if not sps_text:
            raise ValueError(
                "Please enter at least one SPS number in Question 3 "
                "(Associated Open SPSs)."
            )

        mod = self._load_sps_query_module()
        result = mod.fetch_sps_records(sps_text)

        sps_records = result.get("valid", [])
        skipped_sps = result.get("skipped", [])
        not_found_sps = result.get("not_found", [])

        self._sps_fetch_result = {
            "skipped": skipped_sps,
            "not_found": not_found_sps,
        }

        return {
            "sps_records": sps_records,
            "skipped_sps": skipped_sps,
            "not_found_sps": not_found_sps,
        }

    def _fetch_q4_esw_context(self) -> Dict[str, Any]:
        """Fetch Question 4 (Associated ESWs) records if enabled.

        Returns empty structure when Q4 is not selected as Yes.
        """
        pair = self.ref_radios.get("ESW")
        if not pair:
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            return {"esw_records": [], "skipped_esw": [], "not_found_esw": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            return {"esw_records": [], "skipped_esw": [], "not_found_esw": []}

        esw_text = self.ref_boxes["ESW"].toPlainText().strip()
        if not esw_text:
            raise ValueError(
                "Please enter at least one ESW number in Question 4 "
                "(Associated ESWs)."
            )

        mod = self._load_esw_query_module()
        result = mod.fetch_esw_records(esw_text)

        esw_records = result.get("valid", [])
        # Keep an alias so downstream code that expects ec_number can still read ESW IDs.
        for rec in esw_records:
            if rec.get("ec_number") is None and rec.get("esw_number") is not None:
                rec["ec_number"] = rec.get("esw_number")
        skipped_esw = result.get("skipped", [])
        not_found_esw = result.get("not_found", [])

        self._esw_fetch_result = {
            "skipped": skipped_esw,
            "not_found": not_found_esw,
        }

        return {
            "esw_records": esw_records,
            "skipped_esw": skipped_esw,
            "not_found_esw": not_found_esw,
        }

    def _fetch_q5_ecr_context(self) -> Dict[str, Any]:
        """Fetch Question 5 (Reference ECR Numbers) records if enabled.

        Connects to Databricks via ODBC, reads problem and solution for each
        ECR number entered, stores results in a transient instance attribute
        (_ecr_fetch_result) and returns them for inclusion in the AI payload.

        Returns empty structure when Q5 is not selected as Yes.
        """
        pair = self.ref_radios.get("REF_ECR")
        if not pair:
            self._ecr_fetch_result = {"skipped": [], "not_found": []}
            return {"ecr_records": [], "skipped_ecr": [], "not_found_ecr": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._ecr_fetch_result = {"skipped": [], "not_found": []}
            return {"ecr_records": [], "skipped_ecr": [], "not_found_ecr": []}

        ecr_text = self.ref_boxes["REF_ECR"].toPlainText().strip()
        if not ecr_text:
            raise ValueError(
                "Please enter at least one ECR number in Question 5 "
                "(Reference ECR Number(s))."
            )

        mod = self._load_ecr_query_module()
        result = mod.fetch_ecr_records(ecr_text)

        ecr_records = result.get("valid", [])
        skipped_ecr = result.get("skipped", [])
        not_found_ecr = result.get("not_found", [])

        # Store transient fetch metadata for surface warnings after generation.
        self._ecr_fetch_result = {
            "skipped": skipped_ecr,
            "not_found": not_found_ecr,
        }

        return {
            "ecr_records": ecr_records,
            "skipped_ecr": skipped_ecr,
            "not_found_ecr": not_found_ecr,
        }

    def _append_project_context(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Enrich payload with Q2 project, Q3 SPS, Q4 ESW and Q5 ECR context when applicable."""
        enriched = dict(payload or {})
        project_ctx = self._fetch_q2_project_context()
        sps_ctx = self._fetch_q3_sps_context()
        esw_ctx = self._fetch_q4_esw_context()
        ecr_ctx = self._fetch_q5_ecr_context()
        enriched.update(project_ctx)
        enriched.update(sps_ctx)
        enriched.update(esw_ctx)
        enriched.update(ecr_ctx)
        return enriched

    def _extract_numeric_ids(self, text: str) -> List[int]:
        """Extract numeric IDs from free-form text preserving first-seen order."""
        raw = re.findall(r"\d+", text or "")
        if not raw:
            return []
        return list(dict.fromkeys(int(x) for x in raw))

    def _fetch_existing_ec_numbers(self, ids: List[int]) -> List[int]:
        """Return IDs that exist in tbl_projectx_ec.ec_number."""
        if not ids:
            return []

        try:
            import pyodbc
        except Exception as exc:
            raise RuntimeError(f"pyodbc import failed: {exc}") from exc

        # Databricks Hive SQL does not support positional ? parameters; inline validated ints.
        placeholders = ", ".join(str(n) for n in ids)
        sql = f"""
        SELECT DISTINCT ec.ec_number
        FROM prd.rd_core.tbl_projectx_ec ec
        WHERE ec.ec_number IN ({placeholders})
        """

        conn = pyodbc.connect("DSN=Spark-PRD", autocommit=True)
        try:
            cursor = conn.cursor()
            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
            finally:
                cursor.close()
        finally:
            conn.close()

        found = []
        for row in rows:
            if row and row[0] is not None:
                try:
                    found.append(int(row[0]))
                except Exception:
                    continue
        return list(dict.fromkeys(found))

    def _fetch_ec_records(self, ids: List[int]) -> List[Dict[str, Any]]:
        """Return EC records with problem/solution/status for the provided IDs."""
        if not ids:
            return []

        try:
            import pyodbc
        except Exception as exc:
            raise RuntimeError(f"pyodbc import failed: {exc}") from exc

        # Databricks Hive SQL does not support positional ? parameters; inline validated ints.
        placeholders = ", ".join(str(n) for n in ids)
        sql = f"""
        SELECT
            ec.ec_number,
            ec.ec_problem,
            ec.ec_solution,
            st.ec_status
        FROM prd.rd_core.tbl_projectx_ec ec
        LEFT JOIN prd.rd_core.tbl_projectx_ec_status st
            ON ec.ec_status_id = st.ec_status_id
        WHERE ec.ec_number IN ({placeholders})
        ORDER BY ec.ec_number
        """

        conn = pyodbc.connect("DSN=Spark-PRD", autocommit=True)
        try:
            cursor = conn.cursor()
            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
            finally:
                cursor.close()
        finally:
            conn.close()

        records: List[Dict[str, Any]] = []
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                ec_number = int(row[0])
            except Exception:
                continue
            records.append({
                "ec_number": ec_number,
                "problem": str(row[1]).strip() if row[1] is not None else "",
                "solution": str(row[2]).strip() if row[2] is not None else "",
                "status": str(row[3]).strip() if row[3] is not None else "",
            })
        return records

    def _fetch_selected_ec_reference_context(self) -> Dict[str, Any]:
        """Fetch detailed EC context for yes-selected reference fields."""
        mapping = {
            "SPS": "sps_records",
            "ESW": "esw_records",
            "REF_ECR": "reference_ecr_records",
            "QN": "qn_records",
        }
        context: Dict[str, Any] = {value: [] for value in mapping.values()}

        for key, payload_key in mapping.items():
            pair = self.ref_radios.get(key)
            if not pair:
                continue
            rb_yes, _rb_no = pair
            if not rb_yes.isChecked():
                continue

            text_widget = self.ref_boxes.get(key)
            field_text = text_widget.toPlainText().strip() if text_widget else ""
            ids = self._extract_numeric_ids(field_text)
            if not ids:
                continue

            context[payload_key] = self._fetch_ec_records(ids)

        return context

    def _collect_not_found_verification_items(
        self,
        pcr_not_found: List[int],
        project_not_found: List[int],
        sps_not_found: List[int],
        esw_not_found: List[int],
        ecr_not_found: List[int] = None,
    ) -> List[str]:
        """Collect all user-facing verification items for missing/unverifiable IDs."""
        issues: List[str] = []

        if pcr_not_found:
            issues.append(
                "PCR ID(s) not found: "
                + ", ".join(str(x) for x in pcr_not_found)
            )

        if project_not_found:
            issues.append(
                "Project ID(s) not found: "
                + ", ".join(str(x) for x in project_not_found)
            )

        if sps_not_found:
            issues.append(
                "SPS ID(s) not found: "
                + ", ".join(str(x) for x in sps_not_found)
            )

        if esw_not_found:
            issues.append(
                "ESW ID(s) not found: "
                + ", ".join(str(x) for x in esw_not_found)
            )

        if ecr_not_found:
            issues.append(
                "Reference ECR number(s) not found: "
                + ", ".join(str(x) for x in ecr_not_found)
            )

        return issues

    def _validate_q1_pcr(self) -> str:
        """Validate Question 1 (PCR_PCN) only – used by the PCR-driven summary path."""
        pair = self.ref_radios.get("PCR_PCN")
        if not pair:
            return "PCR question widgets not found."
        rb_yes, rb_no = pair
        if not rb_yes.isChecked() and not rb_no.isChecked():
            return (
                "Please answer Question 1: "
                "Does this Project include a Product Change Request (PCR)?"
            )
        if rb_yes.isChecked():
            pcr_text = (self.ref_boxes.get("PCR_PCN") or {}).toPlainText().strip() \
                if hasattr(self.ref_boxes.get("PCR_PCN"), "toPlainText") else ""
            if not pcr_text:
                return (
                    "Please enter at least one PCR number in the PCR field "
                    "before generating a Problem Summary."
                )
        return ""

    def _collect_pcr_payload(self) -> dict:
        """Fetch PCR/PSN records from Databricks and build an AI-ready payload.

        The PCR records and derived PSN state are stored only in runtime
        memory – no files are written.
        """
        pcr_text = self.ref_boxes["PCR_PCN"].toPlainText().strip()

        mod = self._load_pcr_query_module()
        result = mod.fetch_pcr_records(pcr_text)

        valid_records  = result.get("valid", [])
        skipped_records = result.get("skipped", [])
        not_found      = result.get("not_found", [])

        # Derive PSN Yes/No from the psnnumber field on each valid PCR row.
        psn_numbers = [
            str(rec["psnnumber"]).strip()
            for rec in valid_records
            if rec.get("psnnumber")
        ]
        derived_psn = (
            {"answer": "Yes", "numbers": psn_numbers}
            if psn_numbers
            else {"answer": "No", "numbers": []}
        )

        # Keep fetch metadata in a transient instance attribute so the
        # button handler can surface skipped/not-found IDs after generation.
        self._pcr_fetch_result = {
            "skipped":   skipped_records,
            "not_found": not_found,
        }

        scope_parts = [
            ln.strip()
            for ln in self.scope_parts_txt.toPlainText().splitlines()
            if ln.strip()
        ]

        reference_inputs = {}
        for key, txt in self.ref_boxes.items():
            reference_inputs[key] = {
                "answer": self._selected_radio_text(key),
                "text": txt.toPlainText().strip(),
            }

        return {
            "short_title":           self.short_title_edit.text().strip(),
            "reason_code":           self.reason_cb.currentText().strip(),
            "scope_parts":           scope_parts,
            "pcr_records":           valid_records,
            "derived_psn":           derived_psn,
            "skipped_pcrs":          skipped_records,
            "current_problem_text":  self.problem_txt.toPlainText().strip(),
            "proposed_solution_text": self.solution_txt.toPlainText().strip(),
            "solution_regeneration_requested": bool(
                getattr(self, "_solution_regen_requested", False)
            ),
            "solution_regeneration_reason": (
                "Existing solution is incomplete or incorrect. "
                "Regenerate a complete, engineer-quality solution."
                if getattr(self, "_solution_regen_requested", False)
                else ""
            ),
            "selected_tab_payload":  self.get_selected_tab_payload(),
            "reference_inputs": reference_inputs,
        }

    def on_regenerate_solution_clicked(self):
        """Clear existing solution and trigger AI regeneration of proposed solution."""
        try:
            # User clicked regenerate because current output is incomplete/incorrect.
            previous_title = self.short_title_edit.text()
            previous_problem_html = self.problem_txt.toHtml()
            previous_problem_text = self.problem_txt.toPlainText().strip()

            self.solution_txt.clear()
            self._solution_regen_requested = True
            self.on_problem_summary_clicked()

            # Keep this action solution-focused: preserve existing title/problem content.
            self.short_title_edit.setText(previous_title)
            if previous_problem_text:
                self.problem_txt.setHtml(previous_problem_html)
            else:
                self.problem_txt.clear()
        finally:
            self._solution_regen_requested = False

    # ------------------------------------------------------------------
    # Reason Code classification – called after Problem Summary is built
    # ------------------------------------------------------------------
    _REASON_CODE_DATA: Dict[str, Dict] = {
        "Cost Reduction": {
            "justification": (
                "Design change intended to reduce cost. Cost savings must be quantified "
                "in ECR PSS, showing values before and after ECO."
            ),
            "scenario": (
                "4.7\tCost Reduction \u2013 Cost reduction of part or assembly\n"
                "Scenarios\n"
                "This includes design changes in order to reduce cost of material and/or assembly cycle time.\n"
                "The EC engineer shall state quantitatively on the ECO what the existing cost was, along with the estimated savings after ECO implementation.\n"
                "Examples:\n"
                "\u2022\tA base plate is redesigned to change it to a casting in order to save overall material cost of the module.\n"
                "\u2022\tA machined part was redesigned with looser tolerance and less number of drilled holes as a part of cost reduction efforts. The drawing was released with a new PN to reflect the design changes.\n"
                "\u2022\tThe material on a part was changed from Stainless Steel to Aluminum without sacrificing design intent and functionality. The new Aluminum design is less expensive than the Stainless-Steel version."
            ),
            "keywords": [
                "cost reduction", "cost saving", "reduce cost", "dcr", "savings",
                "cost reduc", "cost improv", "price reduc",
            ],
        },
        "Product Release": {
            "justification": (
                "Used for initial release of new products, part status promotion from "
                "Concept or Eval to Prod, or first-pass FAI changes only."
            ),
            "scenario": (
                "4.1\tProduct Release \u2013 Release of New Products (Customer Release) / and Components\n"
                "Scenarios\n"
                "Used for initial release of new products/components. This includes NSR/CES releases and item status change from evaluation to production. It also includes Concept to Eval item status changes.\n"
                "\u2022\tOn changes from Concept to Evaluation\n"
                "\u2022\tAs a result, from managed builds during new product introduction\n"
                "\u2022\tFor the first pass FAI(s)\n"
                "\tChanges that result from subsequent FAIs on the same part number shall be counted as Corrective.\n"
                "\u2022\tDrawing/document corrections only while product-released part numbers shall be considered \u201cCorrective.\u201d\n"
                
            ),
            "keywords": [
                "product release", "new product", "initial release", "fai",
                "eval to prod", "concept to prod", "first release",
                "part status promotion", "new design release",
            ],
        },
        "Product Improvement": {
            "justification": (
                "Change improves functionality, quality, or reliability. Quantitative "
                "performance data must be provided in the ECR (PSS or attachments)."
            ),
            "scenario": (
                "4.2\tProduct Improvement - Improvement(Redesign or replacement New Design) to the Quality/Reliability of the Tool\n"
                "Scenarios\n"
                "This applies to design changes made to parts and assemblies as an improvement, released after July 1, 2003, that currently meet ERAMS+ and C-RAMS performance requirements in Quality (QPPM) and reliability (MTBF, MTBI). This applies to an increase in functionality or performance.\n"
                "It does apply to changes that are required to address Customer specific issues.\n"
                "It does not apply to correcting problems in the design or Drawing.\n"
                "The Design Engineering Authority (DEA) shall state quantitatively on the ECO what the product performance was, (i.e., MTBF, MTTR, MTBI, etc.) and what the product performance is after ECO implementation.\n"
                "Example:\n"
                "\u2022\tA new lift assembly is redesigned to make it more reliable. The current lift assembly already meets ERAMS+ MTBF."
            ),
            "keywords": [
                "product improvement", "improve function", "improve quality",
                "improve reliab", "performance improv", "enhance", "enhancement",
                "increased performance", "better performance",
            ],
        },
        "Manufacturing Improvement": {
            "justification": (
                "Change supports manufacturing efficiency through BOM restructuring, "
                "OMS or test procedure updates, second source additions, "
                "policy/procedure updates, or sparing improvements."
            ),
            "scenario": (
                "4.3\tManufacturing Improvement \u2013 Manufacturing Improvement/Maintenance\n"
                "Scenarios\n"
                "ECOs written to improve/maintain product producibility or manufacturing methods including:\n"
                "\u2022\tPolicy/Procedure creation/updates (all)\n"
                "\u2022\tChanging BOM structure to support work center or supplier changes (no corrections)\n"
                "\u2022\tOMS creation/updates\n"
                "\u2022\tCreating/updating test procedures, fixtures, tools etc.\n"
                "\u2022\tUpdating the OEM PDF to add second sources, update manufacturer\u2019s part numbers, etc. alternate part (no obsolete)\n"
                "\u2022\tIn the process of adding an approved second source supplier for a part released after July 1, 2003, if a SPS, SNF or SCAR is generated that requires an ECO to correct a Design or a document, as defined in the Design Correction or the Document Correction category, that ECO shall be classified as \u201cCorrective\u201d and not included in this category.\n"
                "\u2022\tAdd/updates to the critical parts lists.\n"
                "The code also applies to BOM restructuring due to outsourcing.\n"
                "Examples:\n"
                "\u2022\tThe chamber work center is reconfiguring the feeder lines and incorporating previously released options into it. The ECO restructures to configurable BOM to support this.\n"
                "\u2022\tThe spares group requests making a component part in a PDF part sparable to make the field fix simpler and less expensive. The ECO restructures the kit to accomplish this.\n"
                "\u2022\tAn ECO is written to release OMSs on the integration line.\n"
                "\u2022\tAn excess and zero demand report identifies parts in stock from two years ago that have no usage. The ECO deletes and obsoletes the parts."
            ),
            "keywords": [
                "manufacturing improvement", "manufacturing efficiency", "bom restructur",
                "oms update", "test procedure", "second source", "sparing", "spare",
                "manufacturing process", "assembly process",
            ],
        },
        "Design Correction": {
            "justification": (
                "Change corrects design issues impacting form, fit, or function, "
                "including field failures, safety compliance issues, or dimensional "
                "non-conformance."
            ),
            "scenario": (
                "4.4\tDesign Correction \u2013 Correct Design to meet Specs.\n"
                "Scenarios\n"
                "Changes made to the functionality, reliability, or quality of a product to bring it up to the specification that it was originally specified to meet including the requirements set forth by C-RAMS and ERAMS+. This includes changes to meet the products originally established safety standards.\n"
                "The change is a core engineering change that affects form, fit and function. It does not include document changes that do not affect the basic design. It does not apply to design changes made to reduce material cost. It does not apply to changes that are required to address Customer specific issues.\n"
                "Examples:\n"
                "\u2022\tA chamber was released to Production status with a specification that it could process 20 wafers / hour, however, customers reported that they could only process 10 wafers / hour. Redesign was done to enable the product to meet the original specification and an ECO was written to document the changes.\n"
                "\u2022\tAn SPS was received from a module supplier to correct a dimensional error that resulted in two mating parts not fitting together. The ECO modifies a drawing to fix the problem.\n"
                "\u2022\tAn ECO is fixing a label previously released to make the product S2 compliant.\n"
                "\u2022\tA part/assembly has poor performance in the field, high failure rate, low life one, this shall be considered as design correction not product improvement."
            ),
            "keywords": [
                "design correction", "form fit function", "fff", "field failure",
                "dimensional", "non-conformance", "nonconformance",
                "design issue", "design defect", "fit issue", "fit problem",
                "function issue", "function problem",
            ],
        },
        "Document Correction": {
            "justification": (
                "Change corrects errors in drawings, OMSs, procedures, or policy "
                "documents without impacting product form, fit, or function."
            ),
            "scenario": (
                "4.5\tDocument Correction \u2013 Correction of Document/Drawing/Model/etc.\n"
                "Scenarios\n"
                "Changes to correct document or drawing errors; this includes drawing dimension, Revision, Tolerance, Ballon Callouts, OMSs, Procedures and Policy type documents. It does not apply to drawing changes that will result in how the part will be manufactured. It does not apply to changes that result from engineering, managed and controlled builds or resulting from the initial FAI.\n"
                "Examples:\n"
                "\u2022\tAn SPS was written on volume released product to remove a feature that was double dimensioned. An ECO is written to correct and up rev the drawing.\n"
                "\u2022\tA controlled build was completed in manufacturing during a transition program. Seventy OMS/document errors were identified from the build.\n"
                "\u2022\tA reference designator was found that incorrectly identifies an option on chamber position D. An ECO is written to fix it."
            ),
            "keywords": [
                "document correction", "drawing error", "document error",
                "procedure error", "oms error", "typographical", "documentation error",
                "policy document", "drawing update", "document update",
            ],
        },
        "Obsolescence": {
            "justification": (
                "Supplier or sub-supplier part / OEM is becoming obsolete, End of Life(EOL) "
                "Includes a DEA-approved replacement. Not intended for Kit or Assembly BTP parts."
            ),
            "scenario": (
                "4.6\tObsolescence \u2013 Changes due to supplier part obsolescence or preference\n"
                "Scenarios\n"
                "Changes due to supplier or sub-supplier part obsolescence for parts/components replacement if authorized by DEA.\n"
                "Examples:\n"
                "\u2022\tA cable supplier submits an SPS because the connector is no longer available. The drawing is updated to reflect another supplier\u2019s PN.\n"
                "\u2022\tThe sensor for a cable is no longer available from the manufacturer. The engineering group selects a new sensor and releases a new PN with the new sensor and inactivates the old sensor cable.\n"
                "\u2022\tA Design Engineer wants a fastener purchased from a specific vendor. The documentation needs to be updated to reflect the preferred vendor."
            ),
            "keywords": [
                "obsolescence", "obsolete", "end of life", "eol", "discontinu",
                "oem replacement", "supplier discontinu", "obs part",
                "last time buy", "ltb", "no longer available",
            ],
        },
        "Compliance Product Specification": {
            "justification": (
                "Change addresses non-compliance with customer requirements. "
                "A unique PCR is mandatory."
            ),
            "scenario": (
                "4.9\tCompliance Product Specification\n"
                "Scenarios\n"
                "This is used when we have to update items that are not in Compliance with Customer Requirements. When selected, a unique PCR is required (CORE coded to require the PCR)."
            ),
            "keywords": [
                "compliance", "non-compliance", "noncompliance",
                "customer requirement", "customer spec", "regulatory",
                "specification violation", "does not meet spec",
            ],
        },
        "Safety Event": {
            "justification": (
                "Change impacts product safety and may require PSER reporting, "
                "Core Plus Project (PCR), or compliance-related support."
            ),
            "scenario": (
                "4.10\tSafety Event\n"
                "Scenarios\n"
                "An ECO shall be classified a Safety Event if any of the following impact criteria is met:\n"
                "\u2022\tPSER:  Mandatory capturing internal or external of corporation product Safety Event.\n"
                "\u2022\tPCR:  Mandatory unique Core Plus Project required in order to track all required changes and overall implementation of the Safety Event.\n"
                "\u2022\tSUPPORT:\n"
                "o\tSupplier or Manufacturing: Build, Testing, Shipment while adhering to corporation manufacturing practices & Compliance Product Specification\n"
                "o\tCustomer: Installation, Maintenance, Improvement while adhering to corporation manufacturing practices.  Production usage while adhering with Compliance Product Specification."
            ),
            "keywords": [
                "safety", "pser", "safety event", "safety incident",
                "hazard", "safety compliance", "safety issue", "safety concern",
                "injury", "fire", "electric shock",
            ],
        },
        "Order BOM Change \u2013 CRN Support": {
            "justification": (
                "Change requests updates to System BOMs and is directly linked "
                "to a valid CRN request."
            ),
            "scenario": (
                "4.11\tOrder BOM Change \u2013 CRN Support\n"
                "Scenarios\n"
                "This \u201cReason Code\u201d is not actually a reason code, but rather the way to support the CRN ECR Workflow rather than going through on of the ECR Workflows.  This is only to be selected for CRN ECRs requesting updates to System BOMs."
            ),
            "keywords": [
                "crn", "order bom", "system bom", "crn support",
                "configuration request number",
            ],
        },
        "Cap Code Management": {
            "justification": (
                "Change manages or modifies CAP Code options, with validation of "
                "CAP Code logic and configuration."
            ),
            "scenario": (
                "4.12\tCap Code Management\n"
                "Scenarios\n"
                "Use for ECRs managing/changing CAP Code Options"
            ),
            "keywords": [
                "cap code", "capcode", "cap code management",
                "configuration option cap",
            ],
        },
        "Option Reduction and Product End of Life": {
            "justification": (
                "Change supports structure tree trimming, movement of options to "
                "Unmaintained product structures, or obsolescence of orphaned components."
            ),
            "scenario": (
                "4.14\tOption Reduction and Product End of Life\n"
                "Scenarios\n"
                "Used for changes related to trimming of structure trees to meet corporate option optimization targets.  Options in these changes could be moved to Unmaintained product structure, removed from a specific product structure (in cases where it is still used elsewhere), or obsoleted.\n"
                "It can also be used, or the obsoleting of components orphaned by the unmaintained option process."
            ),
            "keywords": [
                "option reduction", "product end of life", "peol",
                "structure tree", "unmaintained", "orphan",
                "orphaned component", "option trim", "option elim",
            ],
        },
    }

    # ------------------------------------------------------------------
    # EC Category classification – called after Problem Summary is built
    # ------------------------------------------------------------------
    _EC_CATEGORY_DATA: Dict[str, Dict] = {
        "A1": {
            "description": EC_CATEGORY_DESC["A1"],
            "justification": (
                "Revised items include SMBoM options (e.g. kit or assembly options) "
                "and the change involves a CDW (Configuration Design Work order). "
                "OBS Kit or Pieces are present."
            ),
            "keywords": [
                "smbom", "option", "cdw", "obs kit", "kit option",
                "assembly option", "configuration design", "cdw required",
            ],
        },
        "A2": {
            "description": EC_CATEGORY_DESC["A2"],
            "justification": (
                "Revised items include SMBoM options but there is no CDW involved. "
                "Typical for product release with SmBOM impact or structure changes "
                "without a configuration design work order."
            ),
            "keywords": [
                "smbom", "option", "no cdw", "product release smbom",
                "new option", "structure option", "bom option",
            ],
        },
        "B1": {
            "description": EC_CATEGORY_DESC["B1"],
            "justification": (
                "No SMBoM options among revised items but a CDW is required. "
                "Typical for OBS / Inactivate changes where parts are removed "
                "and a configuration design work order is needed."
            ),
            "keywords": [
                "obsolete", "obsolescence", "inactivate", "obs", "cdw",
                "no smbom option", "obs part", "replace obsolete", "end of life",
            ],
        },
        "B2": {
            "description": EC_CATEGORY_DESC["B2"],
            "justification": (
                "No SMBoM options, no CDW, and the revised item(s) are at Eval "
                "status or moving to Eval / being added to Proto buckets."
            ),
            "keywords": [
                "eval", "evaluation", "eval release", "proto", "prototype",
                "eval status", "moving to eval", "added to proto",
                "no cdw", "no smbom",
            ],
        },
        "B3": {
            "description": EC_CATEGORY_DESC["B3"],
            "justification": (
                "No SMBoM options, no CDW, and the revised item(s) are at Production "
                "status or moving to Production. Typical for standard production "
                "revisions, design corrections, or document corrections."
            ),
            "keywords": [
                "production", "prod release", "production release",
                "moving to production", "production status",
                "design correction", "document correction", "revision",
                "production revision",
            ],
        },
    }

    def _classify_ec_category(self, problem: str, solution: str) -> tuple:
        """Analyse problem and solution text and return (ec_category_code, justification).

        Uses keyword-scoring heuristic over EC_CATEGORY_DATA.
        Falls back to 'B3' when no keywords match (most common category).
        """
        combined = (problem + " " + solution).lower()
        best_code = ""
        best_score = -1

        for code, info in self._EC_CATEGORY_DATA.items():
            score = sum(1 for kw in info["keywords"] if kw in combined)
            if score > best_score:
                best_score = score
                best_code = code

        if best_score == 0 or not best_code:
            best_code = "B3"

        info = self._EC_CATEGORY_DATA[best_code]
        tooltip = f"{info['description']}\n\n{info['justification']}"
        return best_code, tooltip

    # ------------------------------------------------------------------

    def _extract_reason_evidence(self, problem: str, solution: str, reason_code: str, seed_texts: List[str]) -> List[str]:
        """Collect short source snippets from problem/solution for highlight display."""
        source_text = (problem or "") + "\n" + (solution or "")
        snippets: List[str] = []

        def _add_snippet(text: str):
            t = (text or "").strip()
            if not t:
                return
            if t.lower() in {s.lower() for s in snippets}:
                return
            snippets.append(t)

        # Keep LLM-provided evidence first when available.
        for s in seed_texts or []:
            _add_snippet(s)

        # Heuristic fallback: pull lines that contain rule keywords for selected code.
        for kw in self._REASON_CODE_DATA.get(reason_code, {}).get("keywords", []):
            kw = (kw or "").strip().lower()
            if not kw:
                continue
            for line in source_text.splitlines():
                line_clean = line.strip()
                if len(line_clean) < 8:
                    continue
                if kw in line_clean.lower():
                    _add_snippet(line_clean[:220])
                    break
            if len(snippets) >= 3:
                break

        return snippets[:3]

    def _highlight_evidence_html(self, text: str, evidence_snippets: List[str]) -> str:
        """Return HTML with matched evidence snippets bold-highlighted."""
        base = text or ""
        needles = [s.strip() for s in (evidence_snippets or []) if s and s.strip()]
        if not needles:
            return html.escape(base).replace("\n", "<br>")

        # Prioritize longer phrases so smaller phrases do not split larger matches.
        needles = sorted(set(needles), key=len, reverse=True)
        pattern = re.compile("|".join(re.escape(n) for n in needles), re.IGNORECASE)

        out: List[str] = []
        last = 0
        for m in pattern.finditer(base):
            out.append(html.escape(base[last:m.start()]))
            out.append(
                "<b style='background:#FFF59D; color:#111; padding:0 1px;'>"
                + html.escape(m.group(0))
                + "</b>"
            )
            last = m.end()
        out.append(html.escape(base[last:]))
        return "".join(out).replace("\n", "<br>")

    def _pick_scenario_reason_lines(
        self,
        scenario_text: str,
        reason_code: str,
        problem: str,
        solution: str,
        evidence_snippets: List[str],
    ) -> List[str]:
        """Pick existing scenario lines that best explain why the reason code was selected."""
        lines = [ln.strip() for ln in (scenario_text or "").splitlines() if ln.strip()]
        if not lines:
            return []

        combined = ((problem or "") + " " + (solution or "")).lower()
        kw_all = [str(k).strip().lower() for k in self._REASON_CODE_DATA.get(reason_code, {}).get("keywords", []) if str(k).strip()]
        kw_active = [k for k in kw_all if k in combined]
        keywords_to_use = kw_active or kw_all

        picked: List[str] = []
        picked_norm = set()

        for raw in lines:
            line = raw.strip()
            line_l = line.lower()
            if any(k in line_l for k in keywords_to_use):
                n = re.sub(r"\s+", " ", line_l)
                if n not in picked_norm:
                    picked.append(line)
                    picked_norm.add(n)

        # Optional second signal from evidence snippets.
        if len(picked) < 3:
            ev_words = set()
            for ev in evidence_snippets or []:
                for tok in re.findall(r"[A-Za-z]{4,}", ev.lower()):
                    ev_words.add(tok)
            if ev_words:
                for raw in lines:
                    if len(picked) >= 3:
                        break
                    line = raw.strip()
                    line_l = line.lower()
                    if len(line_l) < 16:
                        continue
                    if any(w in line_l for w in ev_words):
                        n = re.sub(r"\s+", " ", line_l)
                        if n not in picked_norm:
                            picked.append(line)
                            picked_norm.add(n)

        # Fallback: choose key explanatory lines already present in scenario.
        if not picked:
            for raw in lines:
                line = raw.strip()
                line_l = line.lower()
                if line_l in {"scenarios", "examples:"}:
                    continue
                if line.startswith("4.") or line.startswith("•"):
                    continue
                if len(line) >= 30:
                    picked.append(line)
                if len(picked) >= 2:
                    break

        return picked[:3]

    def _render_scenario_with_reason_lines_html(self, scenario_text: str, reason_lines: List[str]) -> str:
        """Render scenario text as HTML and bold-highlight selected existing lines."""
        chosen = {
            re.sub(r"\s+", " ", (ln or "").strip().lower())
            for ln in (reason_lines or [])
            if (ln or "").strip()
        }

        html_lines: List[str] = []
        for raw in (scenario_text or "").splitlines():
            esc = html.escape(raw)
            norm = re.sub(r"\s+", " ", raw.strip().lower())
            if norm and norm in chosen:
                html_lines.append(
                    "<b style='background:#FFF59D; color:#111; padding:0 1px;'>"
                    + esc
                    + "</b>"
                )
            else:
                html_lines.append(esc)

        return "<br>".join(html_lines)

    def _classify_reason_code(self, problem: str, solution: str) -> tuple:
        """You are a Senior Manufacturing Engineer, analyse problem and solution text and return (reason_code, scenario).

        Attempts LLM-based classification first via AI_Assisted_PSS.classify_reason_code_with_llm,
        using a comprehensive ECM expert prompt with all 14 reason codes and governing rules.
        Falls back to keyword-scoring heuristic on any LLM error.
        """
        # ── Attempt LLM-based classification ────────────────────────────────
        try:
            module = self._load_ai_pss_module()
            classify_fn = getattr(module, "classify_reason_code_with_llm", None)
            if callable(classify_fn):
                try:
                    llm_result = classify_fn(problem, solution, return_evidence=True)
                except TypeError:
                    llm_result = classify_fn(problem, solution)

                llm_code = ""
                llm_just = ""
                llm_evidence: List[str] = []
                if isinstance(llm_result, tuple):
                    if len(llm_result) >= 2:
                        llm_code, llm_just = llm_result[0], llm_result[1]
                    if len(llm_result) >= 3 and isinstance(llm_result[2], list):
                        llm_evidence = llm_result[2]

                if llm_code:
                    info = self._REASON_CODE_DATA.get(llm_code, {})
                    justification = llm_just or info.get("justification", llm_just)
                    scenario = info.get("scenario", justification)
                    evidence = self._extract_reason_evidence(problem, solution, llm_code, llm_evidence)
                    return llm_code, justification, scenario, evidence
        except Exception:
            pass

        # ── Keyword-scoring heuristic fallback ──────────────────────────────
        combined = (problem + " " + solution).lower()
        best_code = ""
        best_score = -1

        for code, info in self._REASON_CODE_DATA.items():
            score = sum(1 for kw in info["keywords"] if kw in combined)
            if score > best_score:
                best_score = score
                best_code = code

        if best_score == 0 or not best_code:
            best_code = "Product Improvement"

        justification = self._REASON_CODE_DATA[best_code]["justification"]
        scenario = self._REASON_CODE_DATA[best_code].get("scenario", justification)
        evidence = self._extract_reason_evidence(problem, solution, best_code, [])
        return best_code, justification, scenario, evidence

    # ------------------------------------------------------------------

    def _show_rc_info_dialog(self):
        """Show Scenario and Examples for the currently proposed Reason Code."""
        scenario_text = getattr(self, "_current_rc_scenario", "")
        reason_lines = getattr(self, "_current_rc_reason_lines", [])
        rc_code = self.ai_reason_code_edit.text().strip() if hasattr(self, "ai_reason_code_edit") else ""
        if not scenario_text:
            QMessageBox.information(
                self,
                "Reason Code Info",
                "Run Problem Summary first to see the Scenario and Examples for the proposed Reason Code.",
            )
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Reason Code: {rc_code} – Scenario & Examples")
        dlg.setMinimumWidth(640)
        dlg.setMinimumHeight(420)
        layout = QVBoxLayout(dlg)
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)

        scenario_html = self._render_scenario_with_reason_lines_html(scenario_text, reason_lines)
        reason_lines_html = "<br>".join(f"- {html.escape(x)}" for x in reason_lines) if reason_lines else "- No explicit scenario lines matched"

        detail_html = (
            "<div style='font-size:12px; line-height:1.45;'>"
            "<h3 style='margin:0 0 8px 0; color:#1F4E79;'>Scenario and Examples</h3>"
            "<div style='margin:0 0 8px 0; color:#444;'><i>Highlighted lines are the specific scenario lines used to justify the proposed Reason Code.</i></div>"
            f"<div>{scenario_html or 'No scenario details available.'}</div>"
            "<h3 style='margin:14px 0 8px 0; color:#1F4E79;'>Reason-Driving Scenario Lines</h3>"
            f"<div>{reason_lines_html}</div>"
            "</div>"
        )
        text_edit.setHtml(detail_html)
        text_edit.setStyleSheet(
            "font-size: 12px; background: #FAFAFA; border: 1px solid #CCCCCC; padding: 6px;"
        )
        layout.addWidget(text_edit)
        close_btn = QPushButton("Close")
        close_btn.setFixedWidth(90)
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)
        dlg.exec()

    # ------------------------------------------------------------------

    def on_problem_summary_clicked(self):
        try:
            # Reset transient fetch state per click to avoid stale cross-field notices.
            self._project_fetch_result = {"skipped": [], "not_found": []}
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            self._ecr_fetch_result = {"skipped": [], "not_found": []}

            # Always validate the full assessment first so unanswered Yes/No
            # selections are prompted regardless of PCR branch.
            validation_error = self._validate_problem_summary_inputs()
            if validation_error:
                QMessageBox.warning(self, "Validation Required", validation_error)
                return

            pcr_yes = self.ref_radios.get("PCR_PCN", (None, None))[0]
            use_pcr_path = pcr_yes is not None and pcr_yes.isChecked()

            if use_pcr_path:
                # ---- Question 1 – PCR-driven Databricks path ----
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                payload = self._append_project_context(self._collect_pcr_payload())

                # Auto-fill PSN radio button and text field from derived_psn (PCR path only).
                # This must happen immediately after PCR fetch so users are not asked to enter PSN manually.
                derived_psn = payload.get("derived_psn") or {}
                psn_pair = self.ref_radios.get("PSN")
                psn_box = self.ref_boxes.get("PSN")
                if psn_pair:
                    psn_yes_rb, psn_no_rb = psn_pair
                    if derived_psn.get("answer") == "Yes":
                        psn_yes_rb.setChecked(True)
                        if psn_box:
                            psn_box.setPlainText(", ".join(derived_psn.get("numbers") or []))
                            psn_box.setVisible(True)
                    else:
                        psn_no_rb.setChecked(True)
                        if psn_box:
                            psn_box.setPlainText("")
                            psn_box.setVisible(False)

                fetch_meta = getattr(self, "_pcr_fetch_result", {})
                skipped    = fetch_meta.get("skipped", [])
                not_found  = fetch_meta.get("not_found", [])
                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])
                ecr_meta = getattr(self, "_ecr_fetch_result", {})
                skipped_ecr = ecr_meta.get("skipped", [])
                not_found_ecr = ecr_meta.get("not_found", [])

                if not payload.get("pcr_records"):
                    msgs = []
                    if skipped:
                        msgs.append(
                            "The following PCR(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  PCR {s['pcr_id']} – {s['status']}" for s in skipped
                            )
                        )
                    if not_found:
                        msgs.append(
                            "The following PCR ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found)
                        )
                    raise ValueError(
                        "No eligible PCR records found to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("PCR", (None, None))[0] and self.ref_radios["PCR"][0].isChecked() and not payload.get("project_records"):
                    msgs = []
                    if skipped_projects:
                        msgs.append(
                            "The following Project(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  Project {s['project_id']} - {s['status']}" for s in skipped_projects
                            )
                        )
                    if not_found_projects:
                        msgs.append(
                            "The following Project ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_projects)
                        )
                    raise ValueError(
                        "No eligible Project records found for Question 2 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("SPS", (None, None))[0] and self.ref_radios["SPS"][0].isChecked() and not payload.get("sps_records"):
                    msgs = []
                    if skipped_sps:
                        msgs.append(
                            "The following SPS(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                            )
                        )
                    if not_found_sps:
                        msgs.append(
                            "The following SPS ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_sps)
                        )
                    raise ValueError(
                        "No eligible SPS records found for Question 3 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("ESW", (None, None))[0] and self.ref_radios["ESW"][0].isChecked() and not payload.get("esw_records"):
                    msgs = []
                    if skipped_esw:
                        msgs.append(
                            "The following ESW(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                            )
                        )
                    if not_found_esw:
                        msgs.append(
                            "The following ESW ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_esw)
                        )
                    raise ValueError(
                        "No eligible ESW records found for Question 4 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("REF_ECR", (None, None))[0] and self.ref_radios["REF_ECR"][0].isChecked() and not payload.get("ecr_records"):
                    msgs = []
                    if skipped_ecr:
                        msgs.append(
                            "The following Reference ECR(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ECR {s['ecr_number']} - {s['status']}" for s in skipped_ecr
                            )
                        )
                    if not_found_ecr:
                        msgs.append(
                            "The following Reference ECR number(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_ecr)
                        )
                    raise ValueError(
                        "No eligible Reference ECR records found for Question 5 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                verify_items = self._collect_not_found_verification_items(
                    pcr_not_found=not_found,
                    project_not_found=not_found_projects,
                    sps_not_found=not_found_sps,
                    esw_not_found=not_found_esw,
                    ecr_not_found=not_found_ecr,
                )
                if verify_items:
                    QMessageBox.warning(
                        self,
                        "Reference Verification Required",
                        "Some entered values were not found. Please verify and correct before generating Problem Summary:\n\n"
                        + "\n".join(f"- {x}" for x in verify_items),
                    )
                    return

                out = self._run_ai_pss_full(payload)

            else:
                # ---- Existing all-questions path ----
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                payload = self._append_project_context(self._collect_pss_payload())

                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])
                ecr_meta = getattr(self, "_ecr_fetch_result", {})
                skipped_ecr = ecr_meta.get("skipped", [])
                not_found_ecr = ecr_meta.get("not_found", [])

                if self.ref_radios.get("SPS", (None, None))[0] and self.ref_radios["SPS"][0].isChecked() and not payload.get("sps_records"):
                    msgs = []
                    if skipped_sps:
                        msgs.append(
                            "The following SPS(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                            )
                        )
                    if not_found_sps:
                        msgs.append(
                            "The following SPS ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_sps)
                        )
                    raise ValueError(
                        "No eligible SPS records found for Question 3 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("ESW", (None, None))[0] and self.ref_radios["ESW"][0].isChecked() and not payload.get("esw_records"):
                    msgs = []
                    if skipped_esw:
                        msgs.append(
                            "The following ESW(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                            )
                        )
                    if not_found_esw:
                        msgs.append(
                            "The following ESW ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_esw)
                        )
                    raise ValueError(
                        "No eligible ESW records found for Question 4 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("REF_ECR", (None, None))[0] and self.ref_radios["REF_ECR"][0].isChecked() and not payload.get("ecr_records"):
                    msgs = []
                    if skipped_ecr:
                        msgs.append(
                            "The following Reference ECR(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ECR {s['ecr_number']} - {s['status']}" for s in skipped_ecr
                            )
                        )
                    if not_found_ecr:
                        msgs.append(
                            "The following Reference ECR number(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_ecr)
                        )
                    raise ValueError(
                        "No eligible Reference ECR records found for Question 5 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                verify_items = self._collect_not_found_verification_items(
                    pcr_not_found=[],
                    project_not_found=not_found_projects,
                    sps_not_found=not_found_sps,
                    esw_not_found=not_found_esw,
                    ecr_not_found=not_found_ecr,
                )
                if verify_items:
                    QMessageBox.warning(
                        self,
                        "Reference Verification Required",
                        "Some entered values were not found. Please verify and correct before generating Problem Summary:\n\n"
                        + "\n".join(f"- {x}" for x in verify_items),
                    )
                    return

                if getattr(self, "email_file_path", None):
                    module = self._load_ai_pss_module()
                    email_content = module.read_email_file(self.email_file_path)
                    if str(email_content).startswith("Error reading"):
                        raise ValueError(f"Could not read email file:\n{email_content}")
                    user_problem = self.problem_txt.toPlainText().strip()
                    out = module.correlate_email_with_problem(
                        email_content,
                        user_problem,
                        payload
                    )
                    if out.get("error"):
                        raise ValueError(out["error"])
                else:
                    out = self._run_ai_pss_full(payload)

                fetch_meta = {}
                skipped    = []
                not_found  = []
                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])

            # ---- Common output handling ----
            title    = (out.get("title") or "").strip()
            problem  = (out.get("problem") or out.get("problem_statement") or "").strip()
            solution = (out.get("solution") or out.get("solution_statement") or "").strip()

            if not any([title, problem, solution]):
                raise ValueError("AI_Assisted_PSS returned empty output.")

            if title:
                self.short_title_edit.setText(title)
            if problem:
                self.problem_txt.setHtml(_format_pss_for_html_display(problem[:2000]))
            if solution:
                formatted_sol = _format_solution_for_display(solution)[:2000]
                self.solution_txt.setHtml(_format_pss_for_html_display(formatted_sol))

            # Surface skipped / not-found PCRs as a non-blocking info message.
            if use_pcr_path and (skipped or not_found):
                info_parts = []
                if skipped:
                    info_parts.append(
                        "Skipped PCR(s) due to inactive status:\n"
                        + "\n".join(
                            f"  PCR {s['pcr_id']} – {s['status']}" for s in skipped
                        )
                    )
                if not_found:
                    info_parts.append(
                        "PCR ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found)
                    )
                QMessageBox.information(
                    self,
                    "PCR Lookup Notice",
                    "Summary generated from eligible PCR(s).\n\n"
                    + "\n\n".join(info_parts),
                )

            if (payload.get("project_records") or skipped_projects or not_found_projects):
                info_parts = []
                if skipped_projects:
                    info_parts.append(
                        "Skipped Project(s) due to inactive status:\n"
                        + "\n".join(
                            f"  Project {s['project_id']} - {s['status']}" for s in skipped_projects
                        )
                    )
                if not_found_projects:
                    info_parts.append(
                        "Project ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_projects)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "Project Lookup Notice",
                        "Summary generated from eligible Project record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            if (payload.get("sps_records") or skipped_sps or not_found_sps):
                info_parts = []
                if skipped_sps:
                    info_parts.append(
                        "Skipped SPS(s) due to inactive status:\n"
                        + "\n".join(
                            f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                        )
                    )
                if not_found_sps:
                    info_parts.append(
                        "SPS ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_sps)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "SPS Lookup Notice",
                        "Summary generated from eligible SPS record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            if (payload.get("esw_records") or skipped_esw or not_found_esw):
                info_parts = []
                if skipped_esw:
                    info_parts.append(
                        "Skipped ESW(s) due to inactive status:\n"
                        + "\n".join(
                            f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                        )
                    )
                if not_found_esw:
                    info_parts.append(
                        "ESW ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_esw)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "ESW Lookup Notice",
                        "Summary generated from eligible ESW record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            # ── Propose Reason Code based on generated problem + solution ────
            _prob = self.problem_txt.toPlainText()
            _sol  = self.solution_txt.toPlainText()
            try:
                rc_code, rc_just, rc_scenario, rc_evidence = self._classify_reason_code(_prob, _sol)
                rc_reason_lines = self._pick_scenario_reason_lines(
                    rc_scenario,
                    rc_code,
                    _prob,
                    _sol,
                    rc_evidence,
                )
                self.ai_reason_code_edit.setText(rc_code)
                self._current_rc_scenario = rc_scenario
                self._current_rc_evidence = rc_evidence
                self._current_rc_reason_lines = rc_reason_lines
                self.ai_justification_btn.setToolTip(
                    "Click \u24d8 to see Scenario and highlighted reason-driving lines for the proposed Reason Code"
                )
            except Exception:
                pass
            # ─────────────────────────────────────────────────────────────────

        except Exception as e:
            QMessageBox.warning(self, "Problem Summary Error", str(e))
        finally:
            QApplication.restoreOverrideCursor()


    # ---- Flow Logic ----
    def clear_flow(self):
        while self.flow_area.count():
            w = self.flow_area.takeAt(0).widget()
            if w: w.deleteLater()
        self.ec_result_lbl.setText(""); self.ec_divider.setVisible(False)

    def add_q(self,q,opts,cb):
        row = QHBoxLayout(); row.addWidget(QLabel(q))
        grp = QButtonGroup(self)
        for o in opts:
            rb = QRadioButton(o); grp.addButton(rb)
            rb.toggled.connect(lambda c,v=o: c and cb(v))
            row.addWidget(rb)
        row.addStretch(1)
        w = QWidget(); w.setLayout(row); self.flow_area.addWidget(w)

    def start_flow(self):
        if not self.sender().isChecked(): return
        self.clear_flow(); s=self.sender().text()
        if s=="Production Release": return self.finish("B3")
        if s=="OBS / Inactivate": return self.finish("B1")
        if s=="Up Revision": return self.add_q("SmBOM Impacted?",["Yes","No"],self.up1)
        if s=="Status Roll": return self.add_q("Transition Type?",["EVAL → Production","Concept → BOM List"],self.status1)
        if s=="Product Release": return self.add_q("SmBOM Impacted?",["Yes","No"],self.prod1)

    def up1(self,v): self.trim(1); self.finish("A1" if v=="Yes" else "A2") if v in ["Yes","No"] and False else (self.add_q("OBS Kit / Pieces?",["Yes","No"],lambda x:self.finish("A1" if x=="Yes" else "A2")) if v=="Yes" else self.add_q("Part Status?",["EVAL","PROD"],lambda x:self.finish("B2" if x=="EVAL" else "B3")))
    def status1(self,v): self.trim(1); self.finish("B3") if v.startswith("EVAL") else self.add_q("OBS Kit / Pieces?",["Yes","No"],lambda x:self.finish("A1" if x=="Yes" else "A2"))
    def prod1(self,v): self.trim(1); self.finish("A2") if v=="Yes" else self.add_q("Part Status?",["EVAL","PROD"],lambda x:self.finish("B2" if x=="EVAL" else "B3"))

    def trim(self,k):
        while self.flow_area.count()>k:
            w=self.flow_area.takeAt(self.flow_area.count()-1).widget()
            if w: w.deleteLater()

    def finish(self,c):
        self.ec_result_lbl.setText(f"EC Category: {c} \u2013 {EC_CATEGORY_DESC[c]}")
        self.ec_divider.setVisible(True); self.secB_header.setVisible(True); self.secB.setVisible(True)
        # Populate the AI Proposed EC Category widgets from the user's flow selection
        try:
            self.ai_ec_category_edit.setText(c)
            info = self._EC_CATEGORY_DATA.get(c, {})
            desc = info.get('description', EC_CATEGORY_DESC.get(c, ''))
            just = info.get('justification', '')
            self.ai_ec_justification_btn.setToolTip(
                f"<b>EC Category {c}:</b><br>{desc}<br><br>{just}"
            )
        except Exception:
            pass



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1280, 860)
        self.setStyleSheet("""
            QScrollBar:vertical {
                background: #EAF1F8;
                width: 12px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #BDD0E3;
                min-height: 28px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover {
                background: #AFC6DD;
            }
            QScrollBar:horizontal {
                background: #EAF1F8;
                height: 12px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #BDD0E3;
                min-width: 28px;
                border-radius: 6px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #AFC6DD;
            }
            QScrollBar::add-line, QScrollBar::sub-line,
            QScrollBar::add-page, QScrollBar::sub-page {
                background: transparent;
                border: none;
            }
        """)
        self.tabs = QTabWidget()
        
        self.tabs.setStyleSheet("""
            QTabBar::tab { background: #EAF2FB; color: #1F3B57; padding: 8px 14px; border: 1px solid #D5E3F6; border-bottom: none; border-top-left-radius:6px; border-top-right-radius:6px; }
            QTabBar::tab:selected { background: #FFFFFF; color: #0F2D46; font-weight: 600; }
            QTabWidget::pane { border: 1px solid #D5E3F6; top: -1px; }
        """)

        readme_path = Path(__file__).with_name('README.txt')
        self.readme_tab = ReadmeTab(readme_path)
        self.front_tab = ECRFrontPageTab()
        self.obs_tab = OBSPartsTab()
        self.structure_tab = None  # to be created later
        self.whereused_tab = WhereUsedTabV2(obs_provider=self.obs_tab)
        self.obs_tab.where_used_tab = self.whereused_tab  # enables OBS → Where Used import

        self.tabs.addTab(self.readme_tab, "READ ME")
        self.tabs.addTab(ECCreationInputsFormTab(), "EC Creation Inputs Form"); self.tabs.setDocumentMode(True); self.tabs.setMovable(True)
        self.tabs.addTab(self.front_tab, "ECR Front Page")
        self.tabs.addTab(self.obs_tab, "OBS Parts")
        self.tabs.addTab(self.whereused_tab, "Where Used")
        self.tabs.addTab(OrphanAnalysisTab(self.obs_tab), "Orphan Analysis")
        self.structure_tab = StructureSheetTab(); self.tabs.addTab(self.structure_tab, "Structure sheet")
        self.inventory_cost_tab=InventoryCostTab(); self.tabs.addTab(self.inventory_cost_tab,'Inventory_Cost')
        self.tabs.addTab(PlaceholderTab("Safety"), "Safety")
        self.tabs.addTab(PlaceholderTab("CE!"), "CE!")
        self.tabs.addTab(PlaceholderTab("CDW"), "CDW")
        self.tabs.addTab(PlaceholderTab("Report"), "Report")
        self.tabs.addTab(PlaceholderTab("User Notes"), "User Notes")
        self.setCentralWidget(self.tabs)
        self._tab_switch_guard = False
        self._prev_tab_index = self.tabs.currentIndex()
        self.tabs.currentChanged.connect(self._on_main_tab_changed)
        def update_tab_colors():
            for i in range(self.tabs.count()):
                txt = self.tabs.tabText(i)
                # heuristic: Front Page checklist
                if 'ECR Front Page' in txt:
                    done = self.front_tab.progress_bar.value()==self.front_tab.progress_bar.maximum()
                    if done:
                        self.tabs.setTabText(i, '✓ ' + txt.strip('✓ '))
                        self.tabs.tabBar().setTabBackgroundColor(i, QColor('#E8F5E9'))
                    elif self.front_tab.progress_bar.value()>0:
                        self.tabs.setTabText(i, '✓ ' + txt.strip('✓ '))
  #                      self.tabs.tabBar().setTabBackgroundColor(i, QColor('#E3F2FD'))
            
        try:
            self.front_tab.progress_bar.valueChanged.connect(update_tab_colors)
        except Exception: pass


        tb=QToolBar("File"); self.addToolBar(tb)
        find_edit = QLineEdit(); find_edit.setPlaceholderText('Find part / text'); find_edit.setFixedWidth(220); tb.addWidget(find_edit)
        def _do_find():
            text = find_edit.text().strip().lower()
            if not text:
                return
            w = self.tabs.currentWidget()
            from PyQt6.QtWidgets import QTableWidget, QTextEdit
            found = False
            for t in w.findChildren(QTableWidget):
                for r in range(t.rowCount()):
                    for c in range(t.columnCount()):
                        it = t.item(r,c)
                        if it and text in it.text().lower():
                            it.setBackground(QColor('#FFF59D'))
                            t.setCurrentCell(r,c)
                            found = True
                            break
                    if found:
                        break
                if found:
                    break
            if not found:
                for te in w.findChildren(QTextEdit):
                    if text in te.toPlainText().lower():
                        te.find(text)
                        found = True
                        break
            if not found:
                QMessageBox.information(self, 'Find', 'Not Found')
        find_edit.returnPressed.connect(_do_find)
        act_save=QAction("Save", self); act_save.triggered.connect(self.save_data); tb.addAction(act_save)
        
        spacer=QWidget(); spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred); tb.addWidget(spacer)

        reset_btn=QPushButton("Reset App")
        reset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reset_btn.setStyleSheet("""
            QPushButton { color:#FFFFFF; padding:6px 14px; border-radius:6px; border:1px solid #0D5EA6;
            background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2994FF, stop:1 #0A67C2); }
            QPushButton:hover { background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2FA0FF, stop:1 #0D6ED0); }
            QPushButton:pressed { padding-top:7px; padding-bottom:5px; }
        """)
        shadow=QGraphicsDropShadowEffect(self); shadow.setBlurRadius(12); shadow.setXOffset(0); shadow.setYOffset(2); shadow.setColor(QColor(0,0,0,80))
        reset_btn.setGraphicsEffect(shadow); reset_btn.clicked.connect(self.reset_app); tb.addWidget(reset_btn)

        self.load_data_if_exists()

    def _on_main_tab_changed(self, new_index: int):
        if self._tab_switch_guard:
            return
        prev = getattr(self, '_prev_tab_index', -1)
        if prev < 0 or prev == new_index:
            self._prev_tab_index = new_index
            return

        prev_widget = self.tabs.widget(prev)
        if isinstance(prev_widget, StructureSheetTab):
            if not prev_widget.confirm_leave_with_incomplete():
                self._tab_switch_guard = True
                try:
                    self.tabs.setCurrentIndex(prev)
                finally:
                    self._tab_switch_guard = False
                return

        self._prev_tab_index = new_index

    def aggregate_data(self)->Dict[str,Any]:
        return {'front_page': self.front_tab.to_dict(), 'obs_parts': self.obs_tab.to_dict()}

    def apply_data(self, data: Dict[str,Any]):
        if not data: return
        if 'front_page' in data: self.front_tab.from_dict(data['front_page'])
        if 'obs_parts' in data: self.obs_tab.from_dict(data['obs_parts'])

    def save_data(self):
        data=self.aggregate_data()
        try:
            DATA_FILE.write_text(json.dumps(data, indent=2), encoding='utf-8')
            QMessageBox.information(self, "Saved", f"Data saved to {DATA_FILE.name}")
        except Exception as e:
            QMessageBox.warning(self, "Save Failed", str(e))

    def load_data_if_exists(self):
        try:
            if DATA_FILE.exists():
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
        except Exception as e:
            QMessageBox.warning(self, "Load Failed", str(e))

    def load_data_dialog(self):
        if DATA_FILE.exists():
            try:
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
                QMessageBox.information(self, "Loaded", f"Data loaded from {DATA_FILE.name}")
            except Exception as e:
                QMessageBox.warning(self, "Load Failed", str(e))
        else:
            QMessageBox.information(self, "No Save Found", "No saved data file found yet. Click Save to create one.")

    def reset_app(self):
        ans=QMessageBox.question(self, "Reset App", "This will clear all fields, reset tables and remove any saved data. Do you want to continue?",
                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if ans==QMessageBox.StandardButton.Yes:
            try:
                self.front_tab.reset(); self.obs_tab.reset()
                if DATA_FILE.exists(): DATA_FILE.unlink()
                QMessageBox.information(self, "Reset Complete", "Application data has been reset.")
            except Exception as e:
                QMessageBox.warning(self, "Reset Failed", str(e))

class OrphanAnalysisTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        outer = QVBoxLayout(self)

        header = QWidget(self)
        h = QHBoxLayout(header)
        h.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))

        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        rb_without = QRadioButton("Without Replacement")
        rb_with = QRadioButton("With / Without Replacement")
        rb_without.setChecked(True)

        grp = QButtonGroup(self)
        grp.addButton(rb_without, 0)
        grp.addButton(rb_with, 1)

        h.addWidget(title)
        h.addSpacing(12)
        h.addWidget(rb_without)
        h.addWidget(rb_with)
        h.addStretch(1)
        outer.addWidget(header)

        self.tabs = QTabWidget()
        tab1 = OBSAllPartsWithoutReplacementTab(obs_provider)
        tab2 = PlaceholderTab("OBS with or without Replacement")

        self.tabs.addTab(tab1, "OBS all Parts without Replacements")
        self.tabs.addTab(tab2, "OBS with or without Replacement")
        outer.addWidget(self.tabs)

        def apply(idx):
            self.tabs.setCurrentIndex(idx)
            self.tabs.setTabEnabled(0, idx == 0)
            self.tabs.setTabEnabled(1, idx == 1)

        rb_without.toggled.connect(lambda v: v and apply(0))
        rb_with.toggled.connect(lambda v: v and apply(1))
        apply(0)



def run():
    app=QApplication(sys.argv); app.setStyle("Fusion")
    pal=QPalette(); pal.setColor(QPalette.ColorRole.Window, QColor(247,250,253)); pal.setColor(QPalette.ColorRole.Base, QColor(255,255,255))
    pal.setColor(QPalette.ColorRole.AlternateBase, QColor(241,246,252)); pal.setColor(QPalette.ColorRole.Text, QColor(28,41,56)); app.setPalette(pal)
    win=MainWindow(); win.show(); sys.exit(app.exec())

if __name__=='__main__': run()

