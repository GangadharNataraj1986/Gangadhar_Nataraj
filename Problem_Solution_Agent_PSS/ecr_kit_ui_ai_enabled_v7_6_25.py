# ecr_kit_ui.py (Enhanced v7.5.5 – Auto Excel conversion: silently open in Excel, SaveAs .xlsx, then import; OBS copy buttons)
import sys
import json
import re
from pathlib import Path
from typing import Any, Dict, List
from PyQt6.QtWidgets import (
    QStyle,
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QTextEdit, QTableWidget, QTableWidgetItem, QCheckBox,
    QPushButton, QProgressBar, QSizePolicy, QScrollArea, QToolBar,
    QMessageBox, QComboBox, QFileDialog, QHeaderView, QGraphicsDropShadowEffect, QFrame, QSplitter,
 QGridLayout
)
from PyQt6.QtGui import QPalette, QColor, QFont, QGuiApplication, QAction, QTextCursor
from PyQt6.QtCore import Qt
APP_TITLE = "ECR Kit Assistant"
DATA_FILE = Path(__file__).with_name('ecr_kit_data.json')
TEMPLATE_FILE = Path(__file__).with_name('Obs_parts_template.xlsx')
import traceback
import importlib.util

def _handle_exception(exc_type, exc_value, exc_tb):
    try:
        log_path = Path(__file__).with_name('error_log.txt')
        msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log_path.write_text(msg, encoding='utf-8')
        m = QMessageBox()
        m.setIcon(QMessageBox.Icon.Critical)
        m.setWindowTitle('Unexpected Error')
        m.setText('An unexpected error occurred. A log was written to error_log.txt')
        m.setDetailedText(msg)
        m.setStandardButtons(QMessageBox.StandardButton.Ok)
        m.show()
        app = QApplication.instance()
        if app is not None:
            if not hasattr(app, '_error_dialogs'):
                app._error_dialogs = []
            app._error_dialogs.append(m)
    except Exception:
        print('Unhandled exception:', file=sys.stderr)
        traceback.print_exception(exc_type, exc_value, exc_tb)

sys.excepthook = _handle_exception

class ReadmeTab(QWidget):
    def __init__(self, readme_path: Path):
        super().__init__()
        layout = QVBoxLayout(self)
        title = QLabel("READ ME")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.DemiBold))
        layout.addWidget(title)
        text = QTextEdit(); text.setReadOnly(True)
        content = "--Wait for the Instructions on how to use--."
        try:
            if readme_path.exists():
                content = readme_path.read_text(encoding="utf-8")
        except Exception as e:
            content = f"Error opening README: {e}"
        text.setPlainText(content)
        layout.addWidget(text)

# ---------- Helpers ----------

from PyQt6.QtGui import QColor

def get_orphan_color(orphan_level: str):
    lvl = orphan_level.lower()
    if lvl == 'orphan1': return QColor('#C0392B')
    if lvl == 'orphan2': return QColor('#E67E22')
    if lvl.startswith('orphan'): return QColor('#2980B9')
    return None


def _excel_width_to_px(widget: QWidget, chars: int) -> int:
    fm = widget.fontMetrics()
    return fm.horizontalAdvance('0' * max(1, chars)) + 22

class ShiftEnterTextEdit(QTextEdit):
    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                self.insertPlainText('\n'); return
        super().keyPressEvent(event)

class ECRFrontPageTab(QWidget):
    ROW_HEIGHT = 40  # compact rows to show ~5 checkpoints per view
    # Updated checklist per user request
    CHECKLIST_DEF = [
        ("Project Association", ["Does the ECR include a Project (PCR)?"]),
        ("Safety & Compliance", ["Provide PSER details if any safety incident occurred.", "Is a PCN required and approved by ROW and CE?"]),
        ("Part Release Status", ["Are all BTPs and kits EVAL released?", "If the parent part is in production, is the replacement part (BTP) production released?"]),
        ("Design Analysis", ["Is VA/VB analysis completed (for new designs/parts only)?", "PACE / DASH parts addressed?"]),
        ("Watchlist & Spares", ["Are new parts/designs MLO certified / were parent/previous parts MLO certified?", "Do we have AGS approval for OBS parts (sparable) without replacement?"]),
        ("Config Documents", ["Is CCR available with change matrix details for new options/reference designator updates?"]),
        ("ECR Strategies Identified", ["Provide reason code, effective strategy, priority, and disposition."]),
        ("Multi BU Alignment", ["If scope impacts multiple BUs/products, verify and confirm all affected BUs/products are listed in the project and approved."]),
        ("Interchangeability & Testing", [
            "Are the changes FFF compatible? Provide system details.",
            "Are test results/FDR available for all FFF impacted parts/critical parts?",
        ]),
        ("Costs and Savings", [
            "If the project relates to DCR (cost reduction), provide cost-saving details.",
        ])
    ]
    _QUESTIONS_TOTAL = sum(len(qs) for _, qs in CHECKLIST_DEF)
    assert len(CHECKLIST_DEF) == 10 and _QUESTIONS_TOTAL == 15

    def __init__(self):
        super().__init__()
        top = QVBoxLayout(self)

        scroll = QScrollArea(self); scroll.setWidgetResizable(True); top.addWidget(scroll)
        content = QWidget(); outer = QVBoxLayout(content); outer.setSpacing(10); outer.setContentsMargins(8,8,8,8); scroll.setWidget(content)

        # Header
        row1 = QWidget(); r1 = QHBoxLayout(row1); r1.setContentsMargins(12,0,0,0)
        self.ecr_no = QLineEdit(); self.ecr_no.setPlaceholderText("ECR#")
        self.eco_primer = QLineEdit(); self.eco_primer.setPlaceholderText("ECO Primer Refs#")
        self.ec_category = QLineEdit(); self.ec_category.setPlaceholderText("EC Category")
        self.bu = QLineEdit(); self.bu.setPlaceholderText("BU")
        self.tco = QLineEdit(); self.tco.setPlaceholderText("TCO")
        self.project_no = QLineEdit(); self.project_no.setPlaceholderText("Project#")
        self.product = QLineEdit(); self.product.setPlaceholderText("Product")
        for w in [self.ecr_no, self.eco_primer, self.ec_category, self.bu, self.tco, self.project_no, self.product]:
            w.setFixedHeight(28); r1.addWidget(w)
        row2 = QWidget(); r2 = QHBoxLayout(row2); r2.setContentsMargins(12,0,0,0)
        self.affected_modules = QLineEdit(); self.affected_modules.setPlaceholderText("Affected Module(s)")
        self.place = QLineEdit(); self.place.setPlaceholderText("Place")
        for w in [self.affected_modules, self.place]:
            w.setFixedHeight(28); r2.addWidget(w)
        outer.addWidget(row1); outer.addWidget(row2)

        lbl = QLabel("ECR Creation Checklist"); lbl.setFont(QFont("Segoe UI", 5, QFont.Weight.DemiBold))
        temp_lbl_row = QWidget(); tl = QHBoxLayout(temp_lbl_row); tl.setContentsMargins(12,0,0,0); tl.addWidget(lbl); outer.addWidget(temp_lbl_row)

        # Table
        table_row = QWidget(); tr = QHBoxLayout(table_row); tr.setContentsMargins(12,0,0,0)
        self.table = QTableWidget(len(self.CHECKLIST_DEF), 9)
        self.table.setHorizontalHeaderLabels(["Sl.No","Completion","Category","Check Point","Validation","Number","Action Owner","Comments","Due Date"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.table.setWordWrap(True); self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
            QHeaderView::section { background:#E1F0FF; color:#0F2D46; font-weight:600; border:1px solid #C9E2FF; padding:4px; }
            QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        for row,(cat,qlist) in enumerate(self.CHECKLIST_DEF):
            sl = QTableWidgetItem(str(row+1)); sl.setFlags(sl.flags() & ~Qt.ItemFlag.ItemIsEditable); sl.setTextAlignment(Qt.AlignmentFlag.AlignCenter); self.table.setItem(row,0,sl)
            chk = QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); cont._chk=chk; self.table.setCellWidget(row,1,cont)
            cat_item = QTableWidgetItem(cat); cat_item.setFlags(cat_item.flags() & ~Qt.ItemFlag.ItemIsEditable); self.table.setItem(row,2,cat_item)
            cp_label = QLabel("\n".join(qlist)); cp_label.setTextFormat(Qt.TextFormat.PlainText); cp_label.setWordWrap(True); self.table.setCellWidget(row,3,cp_label)
            combo = QComboBox(); combo.addItems(['','YES','NO','N/A']); self.table.setCellWidget(row,4,combo)
            def _on_change(txt, r=row):
                sel=(txt or '').strip().upper(); w=self.table.cellWidget(r,1)
                if isinstance(w, QWidget) and hasattr(w,'_chk'): w._chk.setChecked(sel in ('YES','N/A'))
            combo.currentTextChanged.connect(_on_change)
            self.table.setItem(row,5,QTableWidgetItem(""))
            self.table.setItem(row,6,QTableWidgetItem(""))
            te=ShiftEnterTextEdit(); te.setPlaceholderText("Add comments… (Shift+Enter for new line)"); te.setFrameShape(QFrame.Shape.NoFrame); te.setStyleSheet("QTextEdit { padding:4px; }"); self.table.setCellWidget(row,7,te)
        header=self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2,QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(2,_excel_width_to_px(self.table,14))
        header.setSectionResizeMode(3,QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5,QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(6,QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(7,QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(5,_excel_width_to_px(self.table,10))
        self.table.setColumnWidth(6,_excel_width_to_px(self.table,14))
        for r in range(self.table.rowCount()): self.table.setRowHeight(r,self.ROW_HEIGHT)
        tr.addWidget(self.table); outer.addWidget(table_row)

        # Progress
        progress_row = QWidget(); pr=QHBoxLayout(progress_row); pr.setContentsMargins(12,0,0,0)
        self.progress_label = QLabel(f"Checklist Progress: 0 / {len(self.CHECKLIST_DEF)}")
        self.progress_bar=QProgressBar(); self.progress_bar.setRange(0,len(self.CHECKLIST_DEF)); self.progress_bar.setFixedHeight(26); self.progress_bar.setTextVisible(True)
        pr.addWidget(self.progress_label); pr.addWidget(self.progress_bar,3); pr.addStretch(1); outer.addWidget(progress_row)

        # Inputs
        action_row = QWidget(); ar=QHBoxLayout(action_row); ar.setContentsMargins(12,0,0,0)
        self.btn_generate=QPushButton("Generate Problem Statement")
        self.title_edit=QLineEdit(); self.title_edit.setPlaceholderText("Title (max 75 chars)"); self.title_edit.setMaxLength(75)
        ar.addWidget(self.btn_generate,1); ar.addWidget(self.title_edit,3); ar.addStretch(1); outer.addWidget(action_row)

        self.problem_edit=QTextEdit(); self.problem_edit.setPlaceholderText("Write the problem statement here (max 2000 characters)…"); self.problem_edit.setMinimumHeight(120)
        self.problem_edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.problem_edit.textChanged.connect(lambda: self._limit_text(self.problem_edit,2000))

        self.solution_edit=QTextEdit(); self.solution_edit.setPlaceholderText("Write the proposed solution here (max 2000 characters)…"); self.solution_edit.setMinimumHeight(120)
        self.solution_edit.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # --- AI Assisted ECR Drafting Panel (Right Side) ---
        ai_row = QWidget(); ai_layout = QHBoxLayout(ai_row); ai_layout.setContentsMargins(12,0,0,0)
        left_splitter = QSplitter(Qt.Orientation.Vertical)
        left_splitter.addWidget(self.problem_edit)
        left_splitter.addWidget(self.solution_edit)
        left_splitter.setChildrenCollapsible(False)
        left_splitter.setStretchFactor(0, 1)
        left_splitter.setStretchFactor(1, 1)
        left_splitter.setSizes([220, 220])

        ai_panel = QWidget(); ai_panel.setFixedWidth(300); ai_v = QVBoxLayout(ai_panel)
        ai_title = QLabel('AI Assisted ECR Drafting'); ai_title.setFont(QFont('Segoe UI',11,QFont.Weight.DemiBold))
        btn_paste = QPushButton('-||-')
        btn_file = QPushButton('Upload PDF / Word / PPT')
        btn_email = QPushButton('Upload Email (.msg / .eml)')
        scope_lbl = QLabel('Include Data From:')
        cb_obs = QCheckBox('OBS Parts'); cb_obs.setChecked(True)
        cb_wu = QCheckBox('Where Used'); cb_wu.setChecked(True)
        cb_chk = QCheckBox('Checklist'); cb_chk.setChecked(True)
        cb_struct = QCheckBox('Structure Sheet'); cb_struct.setChecked(True)
        btn_ai = QPushButton('Generate ECR using AI'); btn_ai.setFixedHeight(34)
        info = QLabel('AI will analyze selected inputs and other tabs to generate Title, Problem and Solution.'); info.setWordWrap(True)

        for w in [ai_title, btn_paste, btn_file, btn_email, scope_lbl, cb_obs, cb_wu, cb_chk, cb_struct, btn_ai, info]: ai_v.addWidget(w)
        ai_v.addStretch(1)

        ai_layout.addWidget(left_splitter,3); ai_layout.addWidget(ai_panel,1); outer.addWidget(ai_row)

        self.solution_edit.textChanged.connect(lambda: self._limit_text(self.solution_edit,2000))

        self.setStyleSheet("""
            QLabel { color: #12324A; }
            QLineEdit, QTextEdit { background:#FFFFFF; border:1px solid #BBD3EA; border-radius:4px; padding:4px; }
            QLineEdit:focus, QTextEdit:focus { border-color:#639AD2; }
            QPushButton { background-color:#3BAFDA; color:#ffffff; border:1px solid #2C9CC8; border-radius:5px; padding:6px 10px; }
            QPushButton:hover { background-color:#35A0C9; }
            QProgressBar { border:1px solid #BBD3EA; border-radius:4px; background:#ECF4FF; text-align:center; color:#12324A; }
            QProgressBar::chunk { background-color:#5CC0FF; }
        """)

        def recalc():
            checked=0
            for r in range(self.table.rowCount()):
                w=self.table.cellWidget(r,1)
                if isinstance(w,QWidget) and hasattr(w,'_chk') and w._chk.isChecked(): checked+=1
            self.progress_label.setText(f"Checklist Progress: {checked} / {self.table.rowCount()}")
            self.progress_bar.setValue(checked)
        for r in range(self.table.rowCount()):
            w=self.table.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.stateChanged.connect(recalc)

    def _limit_text(self, editor: QTextEdit, max_chars: int):
        doc = editor.toPlainText()
        if len(doc)>max_chars:
            cursor=editor.textCursor(); pos=cursor.position()
            editor.blockSignals(True); editor.setPlainText(doc[:max_chars]); cursor.setPosition(min(pos,max_chars)); editor.setTextCursor(cursor); editor.blockSignals(False)

    def to_dict(self)->Dict[str,Any]:
        data={'fields':{
            'ecr_no':self.ecr_no.text(),'eco_primer':self.eco_primer.text(),'ec_category':self.ec_category.text(),
            'bu':self.bu.text(),'tco':self.tco.text(),'project_no':self.project_no.text(),'product':self.product.text(),
            'affected_modules':self.affected_modules.text(),'place':self.place.text(),'title':self.title_edit.text(),
            'problem':self.problem_edit.toPlainText(),'solution':self.solution_edit.toPlainText(),},'checklist':[]}
        t=self.table
        for r in range(t.rowCount()):
            completion=False; w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): completion=w._chk.isChecked()
            cp_widget=t.cellWidget(r,3); cp_text=cp_widget.text() if isinstance(cp_widget,QLabel) else ''
            v_widget=t.cellWidget(r,4); v_choice=v_widget.currentText() if isinstance(v_widget,QComboBox) else ''
            row={'slno':t.item(r,0).text() if t.item(r,0) else str(r+1),'completion':completion,'category':t.item(r,2).text() if t.item(r,2) else '',
                 'check_point_text':cp_text,'validation_choice':v_choice,'action_owner':t.item(r,5).text() if t.item(r,5) else '',
                 'due_date':t.item(r,6).text() if t.item(r,6) else '',
                 'comments': t.cellWidget(r,7).toPlainText() if isinstance(t.cellWidget(r,7), QTextEdit) else (t.item(r,7).text() if t.item(r,7) else '')}
            data['checklist'].append(row)
        return data

    def from_dict(self, data: Dict[str,Any]):
        f=data.get('fields',{})
        self.ecr_no.setText(f.get('ecr_no','')); self.eco_primer.setText(f.get('eco_primer','')); self.ec_category.setText(f.get('ec_category',''))
        self.bu.setText(f.get('bu','')); self.tco.setText(f.get('tco','')); self.project_no.setText(f.get('project_no','')); self.product.setText(f.get('product',''))
        self.affected_modules.setText(f.get('affected_modules','')); self.place.setText(f.get('place','')); self.title_edit.setText(f.get('title',''))
        self.problem_edit.setPlainText(f.get('problem','')); self.solution_edit.setPlainText(f.get('solution',''))
        checklist=data.get('checklist',[]); t=self.table; rows=min(len(checklist), t.rowCount())
        for r in range(rows):
            row=checklist[r]
            w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.setChecked(bool(row.get('completion',False)))
            vc=row.get('validation_choice',None); wv=t.cellWidget(r,4)
            if vc is not None and isinstance(wv,QComboBox):
                idx=wv.findText(vc); wv.setCurrentIndex(idx if idx>=0 else 0)
            def _ensure(col):
                it=t.item(r,col)
                if it is None:
                    it=QTableWidgetItem(""); t.setItem(r,col,it)
                return it
            _ensure(5).setText(row.get('action_owner','') or row.get('actioner_owner',''))
            _ensure(6).setText(row.get('due_date',''))
            w_comments=t.cellWidget(r,7)
            if isinstance(w_comments, QTextEdit): w_comments.setPlainText(row.get('comments',''))
        for r in range(t.rowCount()):
            try: t.setRowHeight(r,self.ROW_HEIGHT)
            except Exception: pass

    def reset(self):
        for w in [self.ecr_no,self.eco_primer,self.ec_category,self.bu,self.tco,self.project_no,self.product,self.affected_modules,self.place,self.title_edit]: w.clear()
        self.problem_edit.clear(); self.solution_edit.clear()
        t=self.table
        for r in range(t.rowCount()):
            w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.setChecked(False)
        for c in (5,6):
            it=t.item(r,c)
            if it: it.setText("")
        w_comments=t.cellWidget(r,7)
        if isinstance(w_comments, QTextEdit): w_comments.clear()
        self.progress_label.setText(f"Checklist Progress: 0 / {self.table.rowCount()}"); self.progress_bar.setValue(0)

class OBSTable(QTableWidget):
    def __init__(self, parent=None, initial_rows: int = 10):
        super().__init__(initial_rows, 4, parent)
        self.setHorizontalHeaderLabels(["Select","OBS Parts","Change","Replacement"])
        self.verticalHeader().setVisible(False)
        try:
            header=self.horizontalHeader(); header.setStretchLastSection(False)
            header.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(1,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(2,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(3,QHeaderView.ResizeMode.Interactive)
        except Exception: pass
        self._init_rows(0,self.rowCount()); self._apply_excel_widths(); self.setAlternatingRowColors(True)

    def _apply_excel_widths(self, chars:int=14):
        px=_excel_width_to_px(self,chars)
        for col in (1,2,3):
            try: self.setColumnWidth(col,px)
            except Exception: pass

    def _make_change_combo(self)->QComboBox:
        combo=QComboBox(); combo.addItems(["Obsolete","Inactivate"]); combo.setCurrentText("Obsolete"); return combo

    def _init_rows(self,start,end):
        for r in range(start,end):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); self.setCellWidget(r,0,cont)
            if not self.item(r,1): self.setItem(r,1,QTableWidgetItem(""))
            combo=self._make_change_combo(); self.setCellWidget(r,2,combo)
            if not self.item(r,3): self.setItem(r,3,QTableWidgetItem(""))

    def keyPressEvent(self,event):
        try:
            if event.matches(event.StandardKey.Copy): self._copy_selection_to_clipboard(); return
            if event.matches(event.StandardKey.Paste): self._paste_from_clipboard(); return
        except Exception: traceback.print_exc()
        super().keyPressEvent(event)

    def _copy_selection_to_clipboard(self):
        sel = self.selectedRanges()
        if not sel:
            return
        r=sel[0]; rows=[]
        for i in range(r.rowCount()):
            cols=[]
            for j in range(r.columnCount()):
                row_i=r.topRow()+i; col_j=r.leftColumn()+j
                if col_j==2:
                    w=self.cellWidget(row_i,2); txt=w.currentText() if isinstance(w,QComboBox) else ''
                else:
                    it=self.item(row_i,col_j); txt=it.text() if it else ''
                cols.append(txt)
            rows.append('\t'.join(cols))
        QGuiApplication.clipboard().setText('\n'.join(rows))

    def _ensure_rows(self,upto_row_inclusive:int):
        if upto_row_inclusive>=self.rowCount():
            old=self.rowCount(); self.setRowCount(upto_row_inclusive+1); self._init_rows(old,self.rowCount())

    def _paste_from_clipboard(self):
        text = QGuiApplication.clipboard().text()
        if not text:
            return
        start_row=self.currentRow(); start_col=self.currentColumn()
        if start_row<0:
            start_row=self._first_empty_row()
            if start_row<0: start_row=self.rowCount()
        lines=[ln for ln in text.splitlines() if ln.strip()]
        for r_offset,line in enumerate(lines):
            parts=[p.strip() for p in line.split('\t')]
            row=start_row+r_offset; self._ensure_rows(row)
            for c_offset,val in enumerate(parts):
                col=start_col+c_offset
                if col==0: continue
                if col==2:
                    w=self.cellWidget(row,2)
                    if isinstance(w,QComboBox):
                        idx=w.findText(val)
                        if idx>=0: w.setCurrentIndex(idx)
                else:
                    self.setItem(row,col,QTableWidgetItem(val))
        self._apply_excel_widths()

    def _first_empty_row(self):
        for r in range(self.rowCount()):
            it=self.item(r,1)
            if it is None or not it.text().strip(): return r
        return -1

    def delete_selected_rows(self):
        rows_to_delete = []
        for r in range(self.rowCount()):
            w = self.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and cb.isChecked():
                    rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            self.removeRow(r)
        if self.rowCount() == 0:
            self.setRowCount(10)
            self._init_rows(0, 10)

class OBSPartsTab(QWidget):
    def toggle_select_all(self):
        any_unchecked = False
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and not cb.isChecked():
                    any_unchecked = True
                    break
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(any_unchecked)


    def select_all_rows(self):
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, 'findChild'):
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(True)

    def __init__(self):
        super().__init__()
        outer=QVBoxLayout(self)

        title_row=QHBoxLayout()
        title=QLabel("Final OBS List"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold)); title.setStyleSheet("color:#C1272D;")
        btn_template=QPushButton("Download Template")
        btn_upload=QPushButton("Upload Template")
        # New copy buttons
        btn_copy_obs=QPushButton("Copy OBS Parts")
        btn_copy_rep=QPushButton("Copy Repl Parts")
        delete_btn=QPushButton("Delete Selected")
        title_row.addWidget(title)
        title_row.addStretch(1)
        title_row.addWidget(btn_template)
        title_row.addWidget(btn_upload)
        title_row.addWidget(btn_copy_obs)
        title_row.addWidget(btn_copy_rep)
        select_all_btn = QPushButton(" Select All")
        select_all_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))
        select_all_btn.setToolTip("Toggle select / unselect all rows")
        title_row.addWidget(select_all_btn)
        title_row.addWidget(delete_btn)
        outer.addLayout(title_row)

        # ── Where Used of OBS Parts row ──────────────────────────────────────
        wu_row = QHBoxLayout()
        wu_lbl = QLabel("WU Level (1–6):")
        wu_lbl.setStyleSheet("font-weight:600; color:#7A1C21;")
        self.wu_level_input = QLineEdit()
        self.wu_level_input.setFixedWidth(55)
        self.wu_level_input.setPlaceholderText("1–6")
        self.wu_level_input.setMaxLength(1)
        self.wu_level_input.setToolTip("Maximum Where Used depth to retrieve (1 to 6)")
        self.btn_where_used = QPushButton("Where Used of OBS Parts")
        self.btn_where_used.setToolTip(
            "Query Databricks for multi-level Where Used data for all OBS parts"
        )
        plant_lbl = QLabel("Plant:")
        plant_lbl.setStyleSheet("font-weight:600; color:#7A1C21;")
        self.plant_combo = QComboBox()
        self.plant_combo.addItems(["4020", "4055", "4060", "4070", "4080", "4090"])
        self.plant_combo.setCurrentText("4070")
        self.plant_combo.setToolTip("Plant code to filter Where Used query")
        self.plant_combo.setFixedWidth(75)
        wu_row.addStretch(1)
        wu_row.addWidget(wu_lbl)
        wu_row.addWidget(self.wu_level_input)
        wu_row.addWidget(plant_lbl)
        wu_row.addWidget(self.plant_combo)
        wu_row.addWidget(self.btn_where_used)
        outer.addLayout(wu_row)

        legend = QLabel(
            "<b>Orphan Legend:</b> "
            "<span style='color:#C0392B; font-weight:600;'>Orphan1</span> | "
            "<span style='color:#E67E22; font-weight:600;'>Orphan2</span> | "
            "<span style='color:#2980B9; font-weight:600;'>Orphan3+</span>"
        )
        legend.setStyleSheet("padding:4px;")
        outer.addWidget(legend)


        self.table=OBSTable(self, initial_rows=1); outer.addWidget(self.table)

        self.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#FFF5F5; gridline-color:#F3C2C2; }
            QHeaderView::section { background:#F8D7DA; color:#7A1C21; font-weight:600; border:1px solid #E3AEB2; padding:4px; }
            QTableWidget::item:selected { background:#F5B5B8; color:#4A0E10; }
            QPushButton { background-color:#C1272D; color:#FFFFFF; border:1px solid #9F1F24; border-radius:5px; padding:6px 10px; }
            QPushButton:hover { background-color:#AD2227; }
            QComboBox { border:1px solid #E3AEB2; border-radius:4px; padding:2px 6px; }
        """)

        delete_btn.clicked.connect(self.table.delete_selected_rows)
        btn_template.clicked.connect(self.download_template)
        btn_upload.clicked.connect(self.upload_from_excel)
        btn_copy_obs.clicked.connect(self.copy_obs_parts)
        btn_copy_rep.clicked.connect(self.copy_replacement_parts)
        select_all_btn.clicked.connect(self.toggle_select_all)
        self.btn_where_used.clicked.connect(self.launch_where_used_import)
        self.where_used_tab = None  # linked by MainWindow after both tabs are created

    def download_template(self):
        try:
            path,_=QFileDialog.getSaveFileName(self,'Save Template',str(TEMPLATE_FILE),'Excel Files (*.xlsx)')
            if not path: return
            import pandas as pd
            df=pd.DataFrame({'OBS Parts':[''],'Change':['Obsolete'],'Replacement':['']})
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='OBS_Template')
            QMessageBox.information(self,'Template Saved',f'Template saved to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self,'Template Error',str(e))

    def upload_from_excel(self):
        try:
            path,_=QFileDialog.getOpenFileName(self,'Open OBS Parts Excel','', 'Excel Files (*.xlsx *.xls)')
            if not path: return
            import pandas as pd
            def as_text(value: Any) -> str:
                if value is None:
                    return ''
                try:
                    if pd.isna(value):
                        return ''
                except Exception:
                    pass
                return str(value).strip()
            if path.lower().endswith('.xls'):
                df=pd.read_excel(path, engine='xlrd', dtype=str, keep_default_na=False)
            else:
                df=pd.read_excel(path, engine='openpyxl', dtype=str, keep_default_na=False)
            cols={str(c).strip().lower(): c for c in df.columns}
            def pick(name):
                for key in cols:
                    if key==name: return cols[key]
                return None
            c_obs=pick('obs parts') or pick('obs part') or pick('part')
            c_change=pick('change')
            c_rep=(pick('replacement') or pick('repl') or pick('replace') or
                   pick('replacement part') or pick('new part') or pick('irplacement'))
            if not c_obs: raise ValueError('Column "OBS Parts" is required in the Excel file.')
            rows=[]
            for _,r in df.iterrows():
                obs=as_text(r.get(c_obs,''))
                if not obs: continue
                change_val=as_text(r.get(c_change,'Obsolete')) if c_change else 'Obsolete'
                if change_val not in ['Obsolete','Inactivate']: change_val='Obsolete'
                rep=as_text(r.get(c_rep,'')) if c_rep else ''
                rows.append((obs,change_val,rep))
            if not rows:
                QMessageBox.information(self,'No Data','No valid rows found in the Excel file.'); return
            t=self.table; t.setRowCount(len(rows)); t._init_rows(0,len(rows))
            for r,(obs,change,rep) in enumerate(rows):
                t.setItem(r,1,QTableWidgetItem(obs))
                w=t.cellWidget(r,2)
                if isinstance(w,QComboBox):
                    idx=w.findText(change); w.setCurrentIndex(idx if idx>=0 else 0)
                rep_item = t.item(r,3)
                if rep_item is None:
                    rep_item = QTableWidgetItem('')
                    t.setItem(r,3,rep_item)
                rep_item.setText(rep)
            t._apply_excel_widths(); QMessageBox.information(self,'Upload Complete',f'Loaded {len(rows)} rows from Excel.')
        except Exception as e:
            QMessageBox.warning(self,'Upload Error',str(e))

    # NEW: Copy helpers
    def _collect_column_values(self, col_index: int) -> List[str]:
        t=self.table
        values=[]
        for r in range(t.rowCount()):
            it=t.item(r,col_index)
            if it:
                val=(it.text() or '').strip()
                if val:
                    values.append(val)
        return values

    def copy_obs_parts(self):
        values=self._collect_column_values(1)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} OBS Part number(s) to clipboard.')

    def copy_replacement_parts(self):
        values=self._collect_column_values(3)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} Replacement part number(s) to clipboard (Image).')

    def to_dict(self)->Dict[str,Any]:
        t=self.table; rows:List[Dict[str,Any]]=[]
        for r in range(t.rowCount()):
            obs=t.item(r,1).text() if t.item(r,1) else ''
            rep=t.item(r,3).text() if t.item(r,3) else ''
            w=t.cellWidget(r,2); change=w.currentText() if isinstance(w,QComboBox) else 'Obsolete'
            if any([obs.strip(), rep.strip()]): rows.append({'obs_part':obs,'change':change,'replacement':rep})
        return {'rows': rows}

    def from_dict(self,data:Dict[str,Any]):
        rows=data.get('rows',[]); t=self.table; needed=max(10,len(rows)); t.setRowCount(needed); t._init_rows(0,needed)
        for r,row in enumerate(rows):
            t.setItem(r,1,QTableWidgetItem(row.get('obs_part','')))
            w=t.cellWidget(r,2)
            if isinstance(w,QComboBox):
                idx=w.findText(row.get('change','Obsolete'))
                if idx>=0: w.setCurrentIndex(idx)
            t.setItem(r,3,QTableWidgetItem(row.get('replacement','')))
        t._apply_excel_widths()

    def reset(self):
        t=self.table; t.setRowCount(10); t._init_rows(0,10)

    def launch_where_used_import(self):
        """Validate WU level + OBS parts, then delegate to the linked Where Used tab."""
        # ── Validate WU Level ──────────────────────────────────────────────────
        raw_level = self.wu_level_input.text().strip()
        if not raw_level:
            QMessageBox.warning(
                self, 'WU Level Required',
                'Please enter a WU Level (1 to 6) before importing.'
            )
            return
        try:
            wu_level = int(raw_level)
            if not (1 <= wu_level <= 6):
                raise ValueError
        except ValueError:
            QMessageBox.warning(
                self, 'Invalid WU Level',
                f'"{raw_level}" is not valid.  Please enter a whole number from 1 to 6.'
            )
            return

        # ── Collect non-empty OBS part numbers ────────────────────────────────
        t = self.table
        obs_parts: List[str] = []
        for r in range(t.rowCount()):
            it = t.item(r, 1)
            val = (it.text() if it else '').strip()
            if val:
                obs_parts.append(val)

        if not obs_parts:
            QMessageBox.warning(
                self, 'No OBS Parts',
                'The OBS Parts column is empty.\n'
                'Please enter at least one part number before importing.'
            )
            return

        # ── Delegate to Where Used tab ─────────────────────────────────────────
        if self.where_used_tab is None:
            QMessageBox.warning(self, 'Not Ready', 'Where Used tab is not available yet.')
            return

        plant = self.plant_combo.currentText().strip()
        self.where_used_tab.import_from_databricks(obs_parts, wu_level, plant)

        # Switch focus to the Where Used tab after a successful import
        try:
            main = self.window()
            if hasattr(main, 'tabs'):
                main.tabs.setCurrentWidget(self.where_used_tab)
        except Exception:
            pass


class WhereUsedTab(QWidget):
    """Import 'Where Used' parents with cleanup and OBS mapping.
    - Cleans rows (backend-trim of Part) then filters per rules; display keeps original Part text
    - Preserves Excel cell background colors when available (.xlsx)
    - Adds a leading Select column (with checkboxes)
    - Inserts a Replacement column right after the Part column and auto-fills using OBS Parts tab
    - Provides utility actions: select subsets, delete, move to Structure sheet, append to OBS, export, reset
    - Does NOT modify source files on disk
    """
    def __init__(self, obs_provider=None):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)
        title_row = QHBoxLayout()
        title = QLabel("Where Used - Parents"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold))
        btn_import = QPushButton("Import 'Where Used' Parents")
        btn_import.setToolTip("Select Excel/CSV/HTML. Legacy/mismatched files are auto-converted to .xlsx using Excel, then imported. Cleanup + OBS mapping are applied to the in-app view only.")
        title_row.addWidget(title); title_row.addStretch(1); title_row.addWidget(btn_import)
        outer.addLayout(title_row)

    


        # Action buttons row
        btn_row = QHBoxLayout()
        self.btn_import_obs_multi = QPushButton("Import Where Used of OBS Parts (Multiple Level)")
        self.btn_sel_9024 = QPushButton("Select all 9024 Parents")
        self.btn_sel_options = QPushButton("Select all Options/O Class")
        self.btn_sel_esw = QPushButton("Select ESW Parents")
        self.btn_delete_sel = QPushButton("Delete the selected")
        self.btn_move_to_struct = QPushButton("Move selected Items to Structure Sheet")
        self.btn_append_obs = QPushButton("Append the selected to OBS List")
        self.btn_export = QPushButton("Export WhereUsed")
        self.btn_reset = QPushButton("Reset Where_Used")
        for b in [self.btn_import_obs_multi, self.btn_sel_9024, self.btn_sel_options, self.btn_sel_esw, self.btn_delete_sel, self.btn_move_to_struct, self.btn_append_obs, self.btn_export, self.btn_reset]:
            btn_row.addWidget(b)
        outer.addLayout(btn_row)

        self.table=QTableWidget(0,0); self.table.verticalHeader().setVisible(False); self.table.setAlternatingRowColors(True); self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        outer.addWidget(self.table)

        self.setStyleSheet("""
        QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
        QHeaderView::section { background:#E1F0FF; color:#0F2D46; font-weight:600; border:1px solid #C9E2FF; padding:4px; }
        QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        btn_import.clicked.connect(self.import_where_used)
        self.btn_import_obs_multi.clicked.connect(self.import_where_used_of_obs_multi_level)
        self.btn_sel_9024.clicked.connect(self.select_all_9024_parents)
        self.btn_sel_options.clicked.connect(self.select_all_options)
        self.btn_sel_esw.clicked.connect(self.select_esw_parents)
        self.btn_delete_sel.clicked.connect(self.delete_selected_rows)
        self.btn_move_to_struct.clicked.connect(self.move_selected_to_structure)
        self.btn_append_obs.clicked.connect(self.append_selected_to_obs)
        self.btn_export.clicked.connect(self.export_where_used)
        self.btn_reset.clicked.connect(self.reset_where_used)

    # ---------------- Excel conversion helpers ----------------
    def _is_html_like(self, path: str)->bool:
        try:
            with open(path, 'rb') as f:
                head = f.read(2048).lstrip().lower()
                return (head.startswith(b'<!') or head.startswith(b'<html') or b'<table' in head)
        except Exception:
            return False

    def _convert_to_xlsx_via_excel(self, src_path: str)->str|None:
        try:
            from pathlib import Path as _Path
            import win32com.client
            src_abs=str(_Path(src_path).resolve())
            dst_path=str(_Path(src_abs).with_suffix(''))+'_converted.xlsx'
            excel=win32com.client.DispatchEx('Excel.Application'); excel.Visible=False; excel.DisplayAlerts=False
            wb=excel.Workbooks.Open(src_abs); wb.SaveAs(dst_path, FileFormat=51); wb.Close(SaveChanges=False); excel.Quit(); return dst_path
        except Exception:
            try:
                excel.Quit()
            except Exception:
                pass
            return None

    # ---------------- Helpers for OBS, colors, selection ----------------
    def _find_part_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            if str(c).strip().lower()=='part': return i
        return -1
    def _find_parent_col_index(self, cols:list[str])->int:
        keys=['parent','parent part','parent pn','parent number','parent part number']
        low=[str(c).strip().lower() for c in cols]
        for i,name in enumerate(low):
            for k in keys:
                if k==name or k in name: return i
        return -1
    def _find_class_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            s=str(c).strip().lower()
            if any(k in s for k in ['class','type','category']): return i
        return -1
    def _build_obs_map(self)->dict:
        mapping={}
        try:
            if self.obs_provider is None: return mapping
            t=self.obs_provider.table
            for r in range(t.rowCount()):
                obs_item=t.item(r,1); rep_item=t.item(r,3)
                obs=(obs_item.text() if obs_item else '').strip(); rep=(rep_item.text() if rep_item else '')
                if obs: mapping[obs.upper()]=rep
        except Exception: pass
        return mapping
    def _read_xlsx_background_colors(self, xlsx_path:str, target_ncols:int):
        try:
            from openpyxl import load_workbook
            wb=load_workbook(xlsx_path, data_only=True); ws=wb.active
            colors=[]; first_data_row=2; ncols=target_ncols
            for r in range(first_data_row, ws.max_row+1):
                row_colors=[]
                for c in range(1, ncols+1):
                    cell=ws.cell(row=r,column=c); col=None
                    try:
                        fill=cell.fill
                        if fill and getattr(fill,'fill_type',None):
                            start=getattr(fill,'start_color',None); rgb=getattr(start,'rgb',None)
                            if isinstance(rgb,str) and len(rgb) in (6,8):
                                rgb_hex=rgb[-6:]; col=QColor('#'+rgb_hex)
                    except Exception: col=None
                    row_colors.append(col)
                colors.append(row_colors)
            return colors
        except Exception:
            return None
    def _apply_cleanup_rules(self, df, part_idx:int):
        import pandas as pd
        part_raw=df.iloc[:,part_idx].astype(str); part_trim=part_raw.str.strip(); lengths=part_trim.str.len().fillna(0)
        mask_gt=(lengths>10) & part_trim.str.upper().str.startswith('ESW')
        mask_eq=(lengths==10) & (part_trim.str[4]=='-')
        mask=mask_gt | mask_eq
        df_kept=df.loc[mask].copy(); df_kept.reset_index(drop=True, inplace=True)
        return df_kept, mask.reset_index(drop=True)
    def _find_replacement_col(self)->int:
        for i in range(self.table.columnCount()):
            h=self.table.horizontalHeaderItem(i)
            if h and str(h.text()).strip().lower()=='replacement': return i
        return -1
    def _center_replacement_column(self):
        rep_col=self._find_replacement_col()
        if rep_col>=0:
            for r in range(self.table.rowCount()):
                it=self.table.item(r,rep_col)
                if it is None:
                    it=QTableWidgetItem(''); self.table.setItem(r,rep_col,it)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def _populate_table_with_extras(self, df, orig_cols:list[str], part_idx:int, colors_2d=None):
        headers=['Select']
        for i,name in enumerate(orig_cols):
            headers.append(str(name))
            if i==part_idx: headers.append('Replacement')
        self.table.clear(); self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers); self.table.setRowCount(len(df))
        obs_map=self._build_obs_map()
        def map_col(c:int)->int: return 1 + c + (1 if c>part_idx else 0)
        for r in range(len(df)):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); cont._chk=chk; self.table.setCellWidget(r,0,cont)
            for c in range(len(orig_cols)):
                val=df.iloc[r,c]; txt='' if val is None else str(val)
                item=QTableWidgetItem(txt)
                if colors_2d and r<len(colors_2d) and c<len(colors_2d[r]):
                    col=colors_2d[r][c]
                    if isinstance(col,QColor): item.setBackground(col)
                self.table.setItem(r, map_col(c), item)
            part_val=str(df.iloc[r,part_idx]) if df.iloc[r,part_idx] is not None else ''
            part_key=part_val.strip().upper(); replacement=obs_map.get(part_key,'')
            rep_item=QTableWidgetItem(replacement); rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if colors_2d and r<len(colors_2d) and part_idx<len(colors_2d[r]):
                col=colors_2d[r][part_idx]
                if isinstance(col,QColor): rep_item.setBackground(col)
            self.table.setItem(r, 1+part_idx+1, rep_item)
        header=self.table.horizontalHeader()
        for i in range(len(headers)):
            if i==len(headers)-1: header.setSectionResizeMode(i,QHeaderView.ResizeMode.Stretch)
            else: header.setSectionResizeMode(i,QHeaderView.ResizeMode.ResizeToContents)

    # -----------------------------------------------------------
    def import_where_used(self):
        try:
            path,_=QFileDialog.getOpenFileName(self, "Import 'Where Used' Parents", '', "Excel/CSV/HTML (*.xlsx *.xls *.csv *.htm *.html);;All Files (*.*)")
            if not path: return
            import pandas as pd
            df=None; errors=[]; lower=path.lower()
            needs_excel_conversion = lower.endswith('.xls') or self._is_html_like(path)
            converted_path=None
            if needs_excel_conversion:
                converted_path=self._convert_to_xlsx_via_excel(path)
                if converted_path:
                    try: df=pd.read_excel(converted_path, engine='openpyxl')
                    except Exception as e: errors.append(f'Converted .xlsx read failed: {e}')
            if df is None:
                try:
                    if lower.endswith('.xlsx'): df=pd.read_excel(path, engine='openpyxl')
                except Exception as e: errors.append(f'XLSX read failed: {e}')
            if df is None and (lower.endswith('.htm') or lower.endswith('.html') or self._is_html_like(path)):
                try:
                    tables=pd.read_html(path)
                    if tables: df=tables[0]
                except Exception as e: errors.append(f'HTML parse failed: {e}')
            if df is None and lower.endswith('.csv'):
                try: df=pd.read_csv(path)
                except Exception as e: errors.append(f'CSV comma failed: {e}')
                if df is None:
                    try: df=pd.read_csv(path, sep='\t')
                    except Exception as e: errors.append(f'CSV tab failed: {e}')
            if df is None:
                try: df=pd.read_excel(path)
                except Exception as e: errors.append(f'Generic read_excel failed: {e}')
            if df is None:
                msg = "Failed to open the file. Tried Excel (with Excel-based conversion), HTML and CSV paths.\n" + "\n".join(errors[-6:])
                QMessageBox.warning(self,'Import Error', msg); return
            cols=[str(c) if c is not None else '' for c in df.columns]
            part_idx=self._find_part_col_index(cols)
            if part_idx<0:
                QMessageBox.warning(self,'Missing Column', "Couldn't find a 'Part' column (case-insensitive) in the selected file."); return
            # Optional OBS-only prefilter
            if hasattr(self,'_obs_only_filter') and self._obs_only_filter:
                df=self._apply_obs_only_filter(df, part_idx)
            colors_2d=None
            xlsx_source=converted_path if converted_path else (path if path.lower().endswith('.xlsx') else None)
            if xlsx_source: colors_raw=self._read_xlsx_background_colors(xlsx_source, target_ncols=len(cols))
            else: colors_raw=None
            df_kept, mask=self._apply_cleanup_rules(df, part_idx)
            if colors_raw is not None:
                kept_colors=[]; mask_list=mask.tolist()
                for ok,row_colors in zip(mask_list, colors_raw):
                    if ok: kept_colors.append(row_colors)
                colors_2d=kept_colors
            self._populate_table_with_extras(df_kept, list(df_kept.columns), part_idx, colors_2d=colors_2d)
            self._center_replacement_column()
            msg=f"Imported {len(df_kept)} cleaned row(s) from:\n{converted_path or path}"
            if converted_path: msg+="\n(The source file was auto-converted to .xlsx using Excel.)"
            QMessageBox.information(self,'Import Complete', msg)
        except Exception as e:
            QMessageBox.warning(self,'Import Error', str(e))

    # ----------------- Extra actions -----------------
    def import_where_used_of_obs_multi_level(self):
        try:
            self._obs_only_filter=True; self.import_where_used()
        finally:
            if hasattr(self,'_obs_only_filter'): delattr(self,'_obs_only_filter')
    def _apply_obs_only_filter(self, df, part_idx:int):
        obs_map=self._build_obs_map()
        if not obs_map: return df
        keys=set(obs_map.keys()); col=df.iloc[:,part_idx].astype(str).fillna('')
        mask=col.str.strip().str.upper().isin(keys)
        return df.loc[mask].reset_index(drop=True)
    def _selected_row_indices(self):
        rows=[]
        for r in range(self.table.rowCount()):
            w=self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk') and w._chk.isChecked(): rows.append(r)
        return rows
    def _headers(self)->list[str]:
        return [self.table.horizontalHeaderItem(i).text() if self.table.horizontalHeaderItem(i) else '' for i in range(self.table.columnCount())]
    def _map_original_to_table_col(self, orig_idx:int, part_idx:int)->int:
        rep_col=self._find_replacement_col()
        if part_idx>=0 and rep_col>=0 and (orig_idx>(rep_col-2)): return 1+orig_idx+1
        else: return 1+orig_idx
    def select_all_9024_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Select 9024 Parents', "Couldn't find a 'Parent' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=it.text() if it else ''
            self.table.cellWidget(r,0)._chk.setChecked(txt.strip().startswith('9024'))
    def select_all_options(self):
        headers=self._headers(); cidx=self._find_class_col_index(headers[1:])
        if cidx<0: QMessageBox.information(self,'Select Options/O Class', "Couldn't find a 'Class/Type/Category' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(cidx, part_idx)
            it=self.table.item(r, tbl_col); t=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(('OPTION' in t) or ('O CLASS' in t))
    def select_esw_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0:
            pidx=self._find_part_col_index(headers[1:])
            if pidx<0: QMessageBox.information(self,'Select ESW Parents', "Couldn't find 'Parent' or 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(txt.startswith('ESW'))
    def delete_selected_rows(self):
        rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Delete Selected','No rows are selected (checkbox).'); return
        for r in reversed(rows): self.table.removeRow(r)
    def move_selected_to_structure(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'append_rows'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.'); return
            headers=self._headers(); rows_idx=self._selected_row_indices()
            if not rows_idx: QMessageBox.information(self,'Move to Structure Sheet','No rows selected.'); return
            data=[]
            for r in rows_idx:
                row=[(self.table.item(r,c).text() if self.table.item(r,c) else '') for c in range(1,self.table.columnCount())]
                data.append(row)
            target.append_rows(headers[1:], data)
            QMessageBox.information(self,'Move to Structure Sheet', f'Moved {len(data)} row(s) to Structure Sheet.')
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))
    def _append_obs_part(self, part:str):
        try:
            if not self.obs_provider: return
            t=self.obs_provider.table; key=(part or '').strip().upper()
            if not key: return
            existing=set()
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it: existing.add((it.text() or '').strip().upper())
            if key in existing: return
            target=None
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it is None or not (it.text() or '').strip(): target=r; break
            if target is None:
                target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
            t.setItem(target,1,QTableWidgetItem(part))
            w=t.cellWidget(target,2)
            if isinstance(w,QComboBox): idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
        except Exception: pass
    def append_selected_to_obs(self):
        headers=self._headers(); pidx=self._find_part_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Append to OBS', "Couldn't find a 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:]); rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Append to OBS','No rows selected.'); return
        count=0
        for r in rows:
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r,tbl_col); part=it.text() if it else ''
            if part.strip(): self._append_obs_part(part); count+=1
        QMessageBox.information(self,'Append to OBS', f'Appended {count} part(s) to OBS List.')
    def export_where_used(self):
        try:
            dialog = QFileDialog(self, 'Export WhereUsed', 'WhereUsed.xlsx', 'Excel Files (*.xlsx)')
            dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
            dialog.setFileMode(QFileDialog.FileMode.AnyFile)
            dialog.setDefaultSuffix('xlsx')
            if not dialog.exec():
                return
            files = dialog.selectedFiles()
            path = files[0] if files else ''
            if not path: return
            if not path.lower().endswith('.xlsx'):
                path += '.xlsx'
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            wb=Workbook(); ws=wb.active; ws.title='WhereUsed'
            headers=self._headers(); ws.append(headers)
            for c in range(1,len(headers)+1):
                cell=ws.cell(row=1,column=c); cell.font=Font(bold=True)
                if headers[c-1].strip().lower()=='replacement': cell.alignment=Alignment(horizontal='center')
            for r in range(self.table.rowCount()):
                row_vals=[]
                for c in range(self.table.columnCount()):
                    if c==0:
                        w=self.table.cellWidget(r,0); row_vals.append('Yes' if (w and hasattr(w,'_chk') and w._chk.isChecked()) else 'No')
                    else:
                        it=self.table.item(r,c); row_vals.append(it.text() if it else '')
                ws.append(row_vals)
                wu_item = self.table.item(r, 1)
                row_fill_rgb = 'C7DEFA' if (wu_item and wu_item.text().strip() == '0') else None
                for c in range(self.table.columnCount()):
                    it=self.table.item(r,c)
                    if row_fill_rgb:
                        ws.cell(row=r+2,column=c+1).fill=PatternFill('solid', fgColor=row_fill_rgb)
                    elif it and it.background().style() != Qt.BrushStyle.NoBrush:
                        qcol=it.background().color()
                        if qcol.isValid() and qcol.alpha() > 0:
                            rgb=f"{qcol.red():02X}{qcol.green():02X}{qcol.blue():02X}"; ws.cell(row=r+2,column=c+1).fill=PatternFill('solid', fgColor=rgb)
                    if headers[c].strip().lower()=='replacement': ws.cell(row=r+2,column=c+1).alignment=Alignment(horizontal='center')
            wb.save(path); QMessageBox.information(self,'Export Complete', f'Exported {self.table.rowCount()} row(s) to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self,'Export Error', str(e))
    def reset_where_used(self):
        try:
            self.table.clear(); self.table.setRowCount(0); self.table.setColumnCount(0)
            QMessageBox.information(self,'Reset','Where Used view has been reset.')
        except Exception as e:
            QMessageBox.warning(self,'Reset Error', str(e))

    def import_from_databricks(self, obs_parts: List[str], max_level: int, plant: str = "4070"):
        """Import multi-level Where Used from Databricks for the given OBS parts.

        Called by OBSPartsTab.launch_where_used_import() after input validation.
        Inputs are already validated (obs_parts non-empty, max_level 1–6).
        """
        # ── Confirm overwrite of existing data ────────────────────────────────
        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            reply = QMessageBox.question(
                self,
                'Where Used – Existing Data',
                'The Where Used tab already contains data.\n'
                'Delete it and import fresh data from Databricks?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        # ── Import the query module ────────────────────────────────────────────
        try:
            import sys as _sys
            import importlib as _il
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from where_used_query import fetch_where_used, DISPLAY_HEADERS  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(
                self, 'Module Not Found',
                f'where_used_query.py could not be imported:\n{exc}'
            )
            return

        # ── Query Databricks ───────────────────────────────────────────────────
        try:
            records = fetch_where_used(obs_parts, max_level, plant)
        except Exception as exc:
            QMessageBox.warning(self, 'Databricks Query Error', str(exc))
            return

        if not records:
            QMessageBox.information(
                self, 'No Data',
                f'Databricks returned no results for {len(obs_parts)} OBS part(s) '
                f'at max WU level {max_level}.'
            )
            return

        # ── Build OBS map for Replacement column ──────────────────────────────
        obs_map = self._build_obs_map()

        # ── Column layout ──────────────────────────────────────────────────────
        # all_headers: [Select=0, WU Level=1, Part=2, Replacement=3, Rev/Ln=4, ...]
        all_headers = ['Select'] + DISPLAY_HEADERS
        _WU_COL     = 1
        _PART_COL   = 2
        _REPL_COL   = 3
        _DATA_START = 4   # Rev/Ln and onwards

        # Lambdas that extract the display value for each column from col 4 onwards.
        # Order must match DISPLAY_HEADERS[3:] (i.e. after WU Level / Part / Replacement).
        _DB_COL_FUNCS = [
            lambda r: r.get('rev_ln', ''),
            lambda r: r.get('plant', ''),
            lambda r: r.get('description', ''),
            lambda r: r.get('item_status', ''),
            lambda r: r.get('base_qty', ''),
            lambda r: r.get('ext_qty', ''),
            lambda r: r.get('uom', ''),
            lambda r: r.get('eco_number', ''),
            lambda r: r.get('procurement_type', ''),
            lambda r: r.get('effectivity_date', ''),
            lambda r: r.get('user_item_type', ''),
            lambda r: r.get('item_seq', ''),
            lambda r: r.get('kit_code', ''),
            lambda r: r.get('sparable_flag', ''),
            lambda r: (r.get('pace_or_dash', '') if 'pace' in r.get('pace_or_dash', '').lower() else ''),
            lambda r: r.get('mlo_class', ''),
        ]

        # ── Populate table ─────────────────────────────────────────────────────
        self.table.clear()
        self.table.setColumnCount(len(all_headers))
        self.table.setHorizontalHeaderLabels(all_headers)
        self.table.setRowCount(len(records))

        for row_idx, record in enumerate(records):
            # Select checkbox
            chk  = QCheckBox()
            cont = QWidget()
            h    = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            cont._chk = chk
            self.table.setCellWidget(row_idx, 0, cont)

            # WU Level – numeric value from Databricks
            wu_val = record.get('wu_level', '')
            self.table.setItem(row_idx, _WU_COL, QTableWidgetItem(wu_val))

            # Part – display with leading spaces (2 spaces per WU level) for hierarchy
            raw_part = record.get('part', '')
            try:
                level_int = int(wu_val)
            except (ValueError, TypeError):
                level_int = 0
            indented_part = ('      ' * level_int) + raw_part
            self.table.setItem(row_idx, _PART_COL, QTableWidgetItem(indented_part))

            # Replacement – auto-fill from OBS map using the raw (unindented) part key
            replacement = obs_map.get(raw_part.strip().upper(), '')
            rep_item = QTableWidgetItem(replacement)
            rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, _REPL_COL, rep_item)

            # Remaining Databricks columns
            for c_off, fn in enumerate(_DB_COL_FUNCS):
                self.table.setItem(row_idx, _DATA_START + c_off,
                                   QTableWidgetItem(fn(record)))

            # Blue background for level-0 rows (the input OBS parts)
            if wu_val == '0':
                _blue = QColor('#C7DEFA')
                cont.setStyleSheet('background-color: #C7DEFA;')
                for _col in range(1, len(all_headers)):
                    _item = self.table.item(row_idx, _col)
                    if _item is not None:
                        _item.setBackground(_blue)

        # ── Resize columns ─────────────────────────────────────────────────────
        hdr = self.table.horizontalHeader()
        for i in range(len(all_headers)):
            if i == len(all_headers) - 1:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

        QMessageBox.information(
            self, 'Import Complete',
            f'Imported {len(records)} row(s) from Databricks.\n'
            f'OBS parts: {len(obs_parts)}  |  Max WU level: {max_level}  |  Plant: {plant}'
        )


class WhereUsedTabV2(WhereUsedTab):
    """Customized Where Used tab per latest requirements.
    - Remove the "Import Where Used of OBS Parts (Multiple Level)" button (hidden)
    - New two-row action panel with colored buttons
    - Selections based on Part column (with WU Level filter for 9024)
    - Select Opt & Opt Class using first 4-digit prefixes list
    - Select ESW parents using Part prefix 'ESW'
    - Move to Structure Sheet: only Part number
    - Append to OBS List: OBS Part + Replacement (if available)
    - Refresh: recompute Replacement column from current OBS Parts list
    - Delete selected, Reset Tab labels updated
    """
    def __init__(self, obs_provider=None):
        super().__init__(obs_provider=obs_provider)
        # Hide legacy buttons/row
        try:
            # Hide old buttons if they exist
            for btn_name in [
                'btn_import_obs_multi','btn_sel_9024','btn_sel_options','btn_sel_esw',
                'btn_delete_sel','btn_move_to_struct','btn_append_obs','btn_export','btn_reset'
            ]:
                btn = getattr(self, btn_name, None)
                if btn is not None:
                    btn.hide()
        except Exception:
            pass
        
        # Build new two-row action panel
        from PyQt6.QtWidgets import QWidget, QGridLayout, QLabel
        panel = QWidget(self)
        grid = QGridLayout(panel)
        grid.setContentsMargins(0,0,0,0)
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(8)

        # Buttons (new)
        self.v2_btn_sel_9024 = QPushButton("Select all 9024 Parents")
        self.v2_btn_sel_opt = QPushButton("Select all Opt & Opt Class")
        self.v2_btn_sel_esw = QPushButton("Select ESW Parents")
        self.v2_btn_sel_above_cfg = QPushButton("Select above Config")
        self.v2_btn_move = QPushButton("Move to Structure Sheet")
        self.v2_btn_append = QPushButton("Append to OBS List")
        self.v2_btn_refresh = QPushButton("Refresh")
        self.v2_btn_delete = QPushButton("Delete selected")
        self.v2_btn_reset = QPushButton("Reset Tab")
        self.v2_btn_export = QPushButton("Export WhereUsed")
        # Place Import first, then Export (no custom colors)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # Import button is the last widget; insert Export right after Import
            title_row.insertWidget(title_row.count(), self.v2_btn_export)
        except Exception:
            pass
        # Move Export button to title row (right side)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # export will be inserted after import
        except Exception:
            pass

        # Colors (smooth, subtle gradients)
        def btn_style(bg1, bg2, border, text='#FFFFFF'):
            return f"""
            QPushButton {{
                color:{text}; padding:4px 8px; font-size:11px; border-radius:5px;
                border:1px solid {border};
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg2});
            }}
            QPushButton:hover {{
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg1});
            }}
            """
        
        self.v2_btn_sel_9024.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))    # cyan
        self.v2_btn_sel_opt.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_sel_esw.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_move.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_append.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_delete.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_export.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_reset.setStyleSheet(btn_style('#90A4AE','#7C919B','#6A7E87'))       # gray
        self.v2_btn_refresh.setStyleSheet(btn_style('#66BB6A','#4EA85A','#3F8F4A'))     # fresh green

        # Arrange in 2 rows
        row1 = [self.v2_btn_sel_9024, self.v2_btn_sel_opt, self.v2_btn_sel_esw, self.v2_btn_sel_above_cfg, self.v2_btn_refresh]
        row2 = [self.v2_btn_move, self.v2_btn_append, self.v2_btn_delete, self.v2_btn_reset]
        for c,btn in enumerate(row1):
            grid.addWidget(btn, 0, c)
        for c,btn in enumerate(row2):
            grid.addWidget(btn, 1, c)

        # Insert our panel right after the title row (index 1)
        try:
            outer: QVBoxLayout = self.layout()
            # self.layout() returns the QWidget layout of WhereUsedTab (outer QVBoxLayout)
            outer.insertWidget(1, panel)
        except Exception:
            # If insertion fails, just add at end
            self.layout().addWidget(panel)

        # Wire up actions
        self.v2_btn_sel_9024.clicked.connect(self._v2_select_9024_by_part)
        self.v2_btn_sel_opt.clicked.connect(self._v2_select_opt_optclass_by_prefix)
        self.v2_btn_sel_esw.clicked.connect(self._v2_select_esw_by_part)
        self.v2_btn_sel_above_cfg.clicked.connect(self._v2_select_above_config_block)
        self.v2_btn_move.clicked.connect(self._v2_move_selected_part_only)
        self.v2_btn_append.clicked.connect(self._v2_append_selected_with_replacement)
        self.v2_btn_refresh.clicked.connect(self._v2_refresh_replacements)
        self.v2_btn_delete.clicked.connect(self.delete_selected_rows)
        self.v2_btn_reset.clicked.connect(self.reset_where_used)
        self.v2_btn_export.clicked.connect(self.export_where_used)

        # Unified single-color style for Where Used buttons
        for b in (self.v2_btn_sel_9024, self.v2_btn_sel_opt, self.v2_btn_sel_esw,
                  self.v2_btn_sel_above_cfg, self.v2_btn_move, self.v2_btn_append,
                  self.v2_btn_refresh, self.v2_btn_delete, self.v2_btn_export, self.v2_btn_reset):
            b.setStyleSheet("QPushButton{background:#2F80ED;color:white;padding:6px 12px;border-radius:6px;}QPushButton:hover{background:#2567C7;}")

        # Also hide the legacy "Import Where Used of OBS Parts (Multiple Level)" if present in title row
        try:
            if hasattr(self, 'btn_import_obs_multi'):
                self.btn_import_obs_multi.hide()
        except Exception:
            pass

    # ---------- Helpers to find current table columns ----------
    def _v2_find_table_col_index(self, header_exact: str) -> int:
        name = (header_exact or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and (h.text() or '').strip().lower() == name:
                return i
        return -1

    def _v2_find_table_col_contains(self, keyword: str) -> int:
        key = (keyword or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and key in (h.text() or '').strip().lower():
                return i
        return -1

    # ---------- New selection logics ----------
    def _v2_select_9024_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select 9024 Parents', "Couldn't find a 'Part' column.")
            return
        wucol = self._v2_find_table_col_contains('wu level')
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
            ok_prefix = part.startswith('9024')
            ok_wu = True
            if wucol >= 0:
                val = (self.table.item(r, wucol).text() if self.table.item(r, wucol) else '').strip()
                try:
                    ok_wu = int(val) != 0
                except Exception:
                    ok_wu = (val != '0')
            chk = ok_prefix and ok_wu
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    def _v2_select_opt_optclass_by_prefix(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select Opt & Opt Class', "Couldn't find a 'Part' column.")
            return
        prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
            '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
            '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
        }
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
            first4 = part[:4]
            chk = first4 in prefixes
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    def _v2_select_esw_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select ESW Parents', "Couldn't find a 'Part' column.")
            return
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip().upper()
            chk = part.startswith('ESW')
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    # ---------- Move / Append / Refresh ----------
    
    def _v2_toggle_rows(self, rows):
        checks = []
        for r in rows:
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk'):
                checks.append(w._chk)
        if not checks:
            return
        select = any(not c.isChecked() for c in checks)
        for c in checks:
            c.setChecked(select)

    def _v2_select_9024_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select 9024 Parents',"Couldn't find a 'Part' column.")
            return
        wucol = self._v2_find_table_col_contains('wu level')
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip()
            ok = part.startswith('9024')
            if ok and wucol >= 0:
                wu = (self.table.item(r,wucol).text() if self.table.item(r,wucol) else '').strip()
                ok = wu != '0'
            if ok:
                rows.append(r)
        self._v2_toggle_rows(rows)

    def _v2_select_opt_optclass_by_prefix(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select Opt & Opt Class',"Couldn't find a 'Part' column.")
            return
        prefixes = {'0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335','0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465','0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'}
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip()
            if part[:4] in prefixes:
                rows.append(r)
        self._v2_toggle_rows(rows)

    def _v2_select_esw_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select ESW Parents',"Couldn't find a 'Part' column.")
            return
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip().upper()
            if part.startswith('ESW'):
                rows.append(r)
        self._v2_toggle_rows(rows)
    def _v2_move_selected_part_only(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'append_rows'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.')
                return
            rows_idx = self._selected_row_indices()
            if not rows_idx:
                QMessageBox.information(self,'Move to Structure Sheet','No rows selected.')
                return
            pcol = self._v2_find_table_col_index('Part')
            if pcol < 0:
                QMessageBox.information(self,'Move to Structure Sheet',"Couldn't find a 'Part' column.")
                return
            data = []
            for r in rows_idx:
                val = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '')
                data.append([val])
            target.append_rows(['Part'], data)
            QMessageBox.information(self,'Move to Structure Sheet', f'Moved {len(data)} row(s) to Structure Sheet (Part only).')
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))

    def _v2_append_selected_with_replacement(self):
        try:
            if not self.obs_provider:
                QMessageBox.information(self,'Append to OBS', 'OBS Parts tab is not available.')
                return
            t = self.obs_provider.table
            rows = self._selected_row_indices()
            if not rows:
                QMessageBox.information(self,'Append to OBS','No rows selected.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            count = 0
            for r in rows:
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
                rep  = (self.table.item(r, rcol).text() if (rcol>=0 and self.table.item(r, rcol)) else '').strip()
                if not part:
                    continue
                # Check existing
                existing=set()
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it: existing.add((it.text() or '').strip().upper())
                if part.upper() in existing:
                    # If exists, update replacement if empty
                    if rep:
                        for rr in range(t.rowCount()):
                            it=t.item(rr,1)
                            if it and (it.text() or '').strip().upper()==part.upper():
                                t.setItem(rr,3,QTableWidgetItem(rep))
                                break
                    continue
                target=None
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it is None or not (it.text() or '').strip():
                        target=rr; break
                if target is None:
                    target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
                t.setItem(target,1,QTableWidgetItem(part))
                w=t.cellWidget(target,2)
                if isinstance(w,QComboBox):
                    idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
                if rep:
                    t.setItem(target,3,QTableWidgetItem(rep))
                count += 1
            QMessageBox.information(self,'Append to OBS', f'Appended/Updated {count} part(s) to OBS List.')
        except Exception as e:
            QMessageBox.warning(self,'Append Error', str(e))

    def _v2_refresh_replacements(self):
        try:
            obs_map = self._build_obs_map()
            if not obs_map:
                QMessageBox.information(self, 'Refresh', 'No OBS parts found to refresh from.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            if pcol < 0 or rcol < 0:
                QMessageBox.information(self,'Refresh','Missing Part or Replacement column.')
                return
            updated = 0
            for r in range(self.table.rowCount()):
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '')
                key = (part or '').strip().upper()
                rep = obs_map.get(key, '')
                it = self.table.item(r, rcol)
                if it is None:
                    it = QTableWidgetItem('')
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(r, rcol, it)
                old = it.text()
                if rep != old:
                    it.setText(rep)
                    updated += 1
            QMessageBox.information(self,'Refresh', f'Refreshed Replacement column using OBS Parts list. Updated {updated} row(s).')
        except Exception as e:
            QMessageBox.warning(self,'Refresh Error', str(e))



    def _v2_select_above_config_block(self):
        part_col = self._v2_find_table_col_index('Part')
        wu_col = self._v2_find_table_col_contains('wu level')

        if part_col < 0 or wu_col < 0:
            return

        config_prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357',
            '0390','0395','0397','0335','0391','0431','0435','0437',
            '0440','0445','0455','0450','0441','0447','0457',
            '0460','0465','0461','0467','0410','0415','0417',
            '0411','0412','0413','0414','0360','0365','0361','0367'
        }

        rows = self.table.rowCount()
        target_rows = []
        
        # Process each WU Level 0 block
        r = 0
        while r < rows:
            wu_val = (self.table.item(r, wu_col).text()
                      if self.table.item(r, wu_col) else '').strip()

            if wu_val == '0':
                block_start = r
                block_end = rows
                # Find end of this block (next WU Level 0 row)
                for i in range(r + 1, rows):
                    nxt = (self.table.item(i, wu_col).text()
                           if self.table.item(i, wu_col) else '').strip()
                    if nxt == '0':
                        block_end = i
                        break

                # For each config-prefix row in the block, select all rows
                # with greater indentation until indent decreases
                for i in range(block_start + 1, block_end):
                    it = self.table.item(i, part_col)
                    if not it:
                        continue
                    raw = it.text()
                    stripped = raw.lstrip()
                    if not stripped:
                        continue
                    indent = len(raw) - len(stripped)
                    part_prefix = stripped[:4]

                    # If this is a config row, select all deeper rows until indent drops
                    if part_prefix in config_prefixes:
                        for j in range(i + 1, block_end):
                            it_j = self.table.item(j, part_col)
                            if not it_j:
                                continue
                            raw_j = it_j.text()
                            stripped_j = raw_j.lstrip()
                            if not stripped_j:
                                continue
                            indent_j = len(raw_j) - len(stripped_j)
                            
                            # Stop if indentation drops to config level or below
                            if indent_j <= indent:
                                break
                            
                            # Select this row (avoid duplicates)
                            if j not in target_rows:
                                target_rows.append(j)

                r = block_end
            else:
                r += 1

        # Toggle: if any target row is unchecked → select all; else deselect all
        any_unchecked = False
        for i in target_rows:
            w = self.table.cellWidget(i, 0)
            if w and hasattr(w, '_chk') and not w._chk.isChecked():
                any_unchecked = True
                break

        for i in target_rows:
            w = self.table.cellWidget(i, 0)
            if w and hasattr(w, '_chk'):
                w._chk.setChecked(any_unchecked)
class StructureSheetTab(QWidget):
    def __init__(self):
        super().__init__()
        outer = QVBoxLayout(self)

        title = QLabel('Structure Sheet')
        title.setFont(QFont('Segoe UI', 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        self.sub_tabs = QTabWidget()
        outer.addWidget(self.sub_tabs)

        # FIRST: Impacted Options/Parts
        self.sub_tabs.addTab(self._build_impacted_tab(), 'Impacted Options/Parts')

        # SECOND: Structure Sheet
        self.sub_tabs.addTab(self._build_structure_tab(), 'Structure Sheet')

    def _build_structure_tab(self):
        w = QWidget()
        l = QVBoxLayout(w)

        btn = QPushButton('Import Implemented BOM')
        btn.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        btn.clicked.connect(self._import_bom)
        l.addWidget(btn)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        l.addWidget(self.table)
        return w

    def _import_bom(self):
        path, _ = QFileDialog.getOpenFileName(self, 'Import Implemented BOM', '', 'Excel Files (*.xlsx *.xls)')
        if not path:
            return
        try:
            import pandas as pd
            df = pd.read_excel(path, engine='openpyxl')
            self.table.clear()
            self.table.setRowCount(len(df.index))
            self.table.setColumnCount(len(df.columns))
            self.table.setHorizontalHeaderLabels([str(c) for c in df.columns])
            for r in range(len(df.index)):
                for c in range(len(df.columns)):
                    val = '' if pd.isna(df.iat[r, c]) else str(df.iat[r, c])
                    self.table.setItem(r, c, QTableWidgetItem(val))
        except Exception as e:
            QMessageBox.critical(self, 'Import Failed', str(e))

    def _build_impacted_tab(self):
        w = QWidget()
        l = QVBoxLayout(w)

        hdr = QLabel('Impacted Options/Parts')
        hdr.setFont(QFont('Segoe UI', 12, QFont.Weight.DemiBold))
        l.addWidget(hdr)

        class _Text(QTextEdit):
            LIMIT = 5000
            MAX_COLS = 10
            def __init__(self):
                super().__init__()
                self._alerted = False
                self.textChanged.connect(self._validate)

            def _wrap_lines(self, text):
                wrapped = []
                for line in text.splitlines() or ['']:
                    while len(line) > self.MAX_COLS:
                        wrapped.append(line[:self.MAX_COLS])
                        line = line[self.MAX_COLS:]
                    wrapped.append(line)
                return ''.join(wrapped)

            def _validate(self):
                text = self.toPlainText()

                # Auto-wrap lines to 10 chars
                wrapped = self._wrap_lines(text)
                if wrapped != text:
                    self.blockSignals(True)
                    self.setPlainText(wrapped)
                    self.moveCursor(QTextCursor.MoveOperation.End)
                    self.blockSignals(False)
                    text = wrapped

                # Remove duplicate lines
                lines = text.splitlines()
                seen=set(); uniq=[]
                for ln in lines:
                    k=ln.strip()
                    if k and k not in seen:
                        seen.add(k); uniq.append(ln)
                new=''.join(uniq)
                if new!=text:
                    self.blockSignals(True)
                    self.setPlainText(new)
                    self.moveCursor(QTextCursor.MoveOperation.End)
                    self.blockSignals(False)
                    text=new

                # Character limit enforcement
                if len(text)>=self.LIMIT:
                    if not self._alerted:
                        QMessageBox.warning(self,'Character limit','Maximum 5000 characters allowed.')
                        self._alerted=True
                    if len(text)>self.LIMIT:
                        self.blockSignals(True)
                        self.setPlainText(text[:self.LIMIT])
                        self.moveCursor(QTextCursor.MoveOperation.End)
                        self.blockSignals(False)
                else:
                    self._alerted=False

        box=_Text()
        fm=box.fontMetrics()
        box.setFixedWidth(fm.horizontalAdvance('0')*9+20)
        box.setFixedHeight(fm.lineSpacing()*500)
        box.setPlaceholderText('Max 10 characters per line. Auto-wrap enabled. 5000 chars total.')
        l.addWidget(box); l.addStretch(1)
        return w

    def append_rows(self, headers:list[str], rows:list[list[str]]):
        if self.table.columnCount()==0:
            self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers)
        if self.table.columnCount()!=len(headers):
            norm=[]
            for r in rows:
                r2=(r+['']*self.table.columnCount())[:self.table.columnCount()]; norm.append(r2)
            rows=norm
        start=self.table.rowCount(); self.table.setRowCount(start+len(rows))
        for i,row in enumerate(rows):
            for j,val in enumerate(row): self.table.setItem(start+i,j,QTableWidgetItem(val))
class StructureSheetTab(QWidget):
    MAX_LINES = 500
    PART_LEN = 10

    def __init__(self):
        super().__init__()

        outer = QVBoxLayout(self)
        title = QLabel("Structure Sheet")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        self.sub_tabs = QTabWidget()
        outer.addWidget(self.sub_tabs)

        # ---- Impacted Options / Parts ----
        impacted_tab = QWidget()
        imp_layout = QVBoxLayout(impacted_tab)

        imp_label = QLabel(
            "Impacted Options/Parts"
            f"(Max {self.MAX_LINES} Parts)"
        )
        imp_label.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
        imp_layout.addWidget(imp_label)

        self.impacted_text = QTextEdit()
        self.impacted_text.setPlaceholderText(
            "Paste or enter part numbers here..."
            "One part number per line"
        )
        self.impacted_text.setFixedHeight(350)
        self.impacted_text.textChanged.connect(self._validate_parts)
        imp_layout.addWidget(self.impacted_text)

        self.validation_label = QLabel("")
        self.validation_label.setStyleSheet("color:#C0392B")
        imp_layout.addWidget(self.validation_label)

        self.sub_tabs.addTab(impacted_tab, "Impacted Options/Parts")

        # ---- Structure Sheet Table ----
        structure_tab = QWidget()
        struct_layout = QVBoxLayout(structure_tab)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        struct_layout.addWidget(self.table)

        self.sub_tabs.addTab(structure_tab, "Structure Sheet")

    def _validate_parts(self):
        text = self.impacted_text.toPlainText()
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        errors = []

        if len(lines) > self.MAX_LINES:
            errors.append(f"Too many lines: {len(lines)} (Max {self.MAX_LINES})")

        bad = [l for l in lines if len(l) != self.PART_LEN]
        if bad:
            errors.append(f"{len(bad)} entries not {self.PART_LEN} characters")

        if errors:
            self.validation_label.setText(" | ".join(errors))
            self.validation_label.setStyleSheet("color:#C0392B")
        else:
            self.validation_label.setText(f"✅ {len(lines)} valid part numbers")
            self.validation_label.setStyleSheet("color:#2E7D32")

    def get_impacted_parts(self):
        lines = [l.strip() for l in self.impacted_text.toPlainText().splitlines() if l.strip()]
        return [l for l in lines if len(l) == self.PART_LEN]

    def append_rows(self, headers:list[str], rows:list[list[str]]):
        if self.table.columnCount()==0:
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
        if self.table.columnCount()!=len(headers):
            norm=[]
            for r in rows:
                r2=(r+['']*self.table.columnCount())[:self.table.columnCount()]
                norm.append(r2)
            rows=norm
        start=self.table.rowCount()
        self.table.setRowCount(start+len(rows))
        for i,row in enumerate(rows):
            for j,val in enumerate(row):
                self.table.setItem(start+i,j,QTableWidgetItem(val))


# ================= Inventory Cost Tab (FULL UPDATED) =================
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

PLANTS = [4020,4055,4060,4070,4080,4090]
METRICS = ['On Order Qty','Onhand Qty','Gross Demand-13','Gross Demand-26','Gross Demand-52','Standard Cost USD']
NO_CDW_CODES = {
'0070','0080','0110','0120','0130','0170','0180','0210','0243','0250','0251','0260','0261','0280','0288','0289','0290',
'0301','0302','0303','0304','0305','0320','0330','0335','0340','0345','0350','0355','0360','0365','0370','0375','0380','0385','0390','0395',
'0401','0402','0403','0404','0405','0410','0415','0420','0425','0430','0435','0440','0445','0450','0455','0460','0465','0470','0475','0480','0485','0490','0495'
}

class InventoryCostTab(QWidget):
    def reset_tab(self):
        self.df = None
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def __init__(self):
        super().__init__()
        v = QVBoxLayout(self)
        h = QHBoxLayout()
        self.btn_import = QPushButton('Import MM360 & Create Inventory & Cost')
        self.btn_export = QPushButton('Export Excel')
        self.btn_reset = QPushButton('Reset Tab')
        for b in (self.btn_import, self.btn_export, self.btn_reset):
            h.addWidget(b)
        h.addStretch(1)
        v.addLayout(h)
        self.table = QTableWidget(0,0)
        self.table.setAlternatingRowColors(True)
        v.addWidget(self.table)
        self.df = None
        self.btn_import.clicked.connect(self.import_mm360)
        self.btn_export.clicked.connect(self.export_excel)
        self.btn_reset.clicked.connect(self.reset_tab)

    def import_mm360(self):
        path,_ = QFileDialog.getOpenFileName(self,'Open MM360','', 'Excel Files (*.xlsx)')
        if not path: return
        src = pd.read_excel(path, sheet_name='Material Analysis', engine='openpyxl')

        def pace(v):
            v=str(v)
            if v.startswith('SGP'): return 'PACE'
            if v.startswith('GDS'): return 'DASH'
            return ''
        src['PACE/DASH'] = src['MRP Profile'].apply(pace)

        rows=[]
        for (pn,desc,pdsh), g in src.groupby(['Material Number','Material Description','PACE/DASH']):
            code4=str(pn)[:4]
            prim=sec=''
            if code4 in NO_CDW_CODES:
                prim=sec='No Change required'
            row={
                'Material Number':pn,
                'Material Description':desc,
                'Primary Disposition':prim,
                'Secondary Disposition':sec,
                'PACE/DASH':pdsh
            }
            tot_on=tot_oh=tot_d13=tot_d26=tot_d52=tot_cost=0.0
            for p in PLANTS:
                gp=g[g['Plant Code']==p]
                oo=float(gp['On Order Quantity'].sum()) if not gp.empty else 0
                oh=float(gp['onhand'].sum()) if not gp.empty else 0
                d13=float(gp['Gross Demand-13'].sum()) if not gp.empty else 0
                d26=float(gp['Gross Demand-26'].sum()) if not gp.empty else 0
                d52=float(gp['Gross Demand-52'].sum()) if not gp.empty else 0
                sc=float(gp['Standard Cost USD'].iloc[0]) if (oo or oh) else 0
                row[f'{p} On Order Qty']=oo
                row[f'{p} Onhand Qty']=oh
                row[f'{p} Gross Demand-13']=d13
                row[f'{p} Gross Demand-26']=d26
                row[f'{p} Gross Demand-52']=d52
                row[f'{p} Standard Cost USD']=sc
                tot_on+=oo; tot_oh+=oh
                tot_d13+=d13; tot_d26+=d26; tot_d52+=d52
                tot_cost+=(oo+oh)*sc
            row['Total On Order Quantity']=tot_on
            row['Total Onhand']=tot_oh
            row['Gross Demand-13']=tot_d13
            row['Gross Demand-26']=tot_d26
            row['Gross Demand-52']=tot_d52
            row['Inventory Cost']=tot_cost
            rows.append(row)
        self.df=pd.DataFrame(rows).sort_values('Inventory Cost',ascending=False)
        self.render()

    def render(self):
        if self.df is None: return
        self.table.setRowCount(len(self.df))
        self.table.setColumnCount(len(self.df.columns))
        self.table.setHorizontalHeaderLabels(self.df.columns.tolist())
        for r in range(len(self.df)):
            for c,col in enumerate(self.df.columns):
                v=self.df.iloc[r][col]
                txt='' if (isinstance(v,(int,float)) and v==0) else str(v)
                self.table.setItem(r,c,QTableWidgetItem(txt))
        self.table.resizeColumnsToContents()

    def export_excel(self):
        if self.df is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, 'Export', 'inventory_cost.xlsx', 'Excel Files (*.xlsx)'
        )
        if not path:
            return
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import pandas as pd

        # Insert Replacement column between B & C
        df = self.df.copy()
        df.insert(2, 'Replacement', '')

        # Build OBS -> Replacement map
        obs_map = {}
        try:
            main = self.window()
            obs_tab = getattr(main, 'obs_tab', None)
            if obs_tab:
                t = obs_tab.table
                for r in range(t.rowCount()):
                    obs = t.item(r,1).text().strip() if t.item(r,1) else ''
                    rep = t.item(r,3).text().strip() if t.item(r,3) else ''
                    if obs:
                        obs_map[obs.upper()] = rep
        except Exception:
            pass

        # Fill Replacement column by matching Material Number
        for i, pn in enumerate(df['Material Number']):
            key = str(pn).strip().upper()
            if key in obs_map:
                df.at[i, 'Replacement'] = obs_map[key]

        with pd.ExcelWriter(path, engine='openpyxl') as w:
            df.to_excel(w, index=False, sheet_name='Output Sheet', startrow=1)
        wb = load_workbook(path)
        ws = wb['Output Sheet']

        header_fill = PatternFill('solid', fgColor='BDD7EE')
        ws.row_dimensions[2].height = 118

        # Format row 2 headers
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(2, c)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            if c <= 5:
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, text_rotation=90)

        # ---- Extract first 4 digits from row 2 (G onward) and write to row 1 ----
        groups = {}
        for c in range(7, ws.max_column + 1):
            val = str(ws.cell(2, c).value or '')
            code = val[:4]
            ws.cell(1, c).value = code
            groups.setdefault(code, []).append(c)

        # Merge same codes in row 1
        for code, cols in groups.items():
            if not code:
                continue
            if len(cols) > 1:
                ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])
            cell = ws.cell(1, cols[0])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # Column widths
        for c in range(6, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(c)].width = 4
        for c in range(1,6):
            col_letter = get_column_letter(c)
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, c).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = max(10, max_len + 2)

        # Blank zero values
        for r in range(3, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).value == 0:
                    ws.cell(r, c).value = ''

        wb.save(path)
        QMessageBox.information(self, 'Export', 'Excel exported successfully')
def reset_tab(self):
        self.df = None
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

class PlaceholderTab(QWidget):
    def __init__(self, title: str):
        super().__init__()
        l = QVBoxLayout(self)
        l.addWidget(QLabel(f"{title} – UI under development"))



class OrphanOBSSubTab(QWidget):
    def __init__(self):
        super().__init__()
        outer = QVBoxLayout(self)

        btn_row = QHBoxLayout()
        self.btn_import = QPushButton("Import Imp BOM of OBS Parts (.xlsx)")
        self.btn_select_oem = QPushButton("Select OEM's")
        self.btn_delete = QPushButton("Delete")
        self.btn_reset = QPushButton("Reset")
        self.btn_copy_removed = QPushButton("Copy Removed Child Parts")

        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(self.btn_select_oem)
        btn_row.addWidget(self.btn_delete)
        btn_row.addWidget(self.btn_reset)
        btn_row.addWidget(self.btn_copy_removed)
        btn_row.addStretch(1)
        outer.addLayout(btn_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self.import_excel)
        self.btn_select_oem.clicked.connect(self.select_oems)
        self.btn_delete.clicked.connect(self.delete_selected)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.btn_copy_removed.clicked.connect(self.copy_removed_child_parts)

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import Imp BOM", '', 'Excel Files (*.xlsx)')
        if not path:
            return
        import pandas as pd
        from PyQt6.QtGui import QColor

        df = pd.read_excel(path, engine='openpyxl')
        cols = list(df.columns)
        if 'Part' not in cols or 'BOM Level' not in cols:
            QMessageBox.warning(self, 'Missing Column', 'Required columns not found')
            return

        part_idx = cols.index('Part')
        cols2 = cols[:part_idx+1] + ['Tool comments'] + cols[part_idx+1:]

        headers = ['Select'] + cols2
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(df))

        bom_df_idx = cols.index('BOM Level')
        tc_tbl_idx = headers.index('Tool comments')

        for r in range(len(df)):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            self.table.setCellWidget(r, 0, cont)

            for c, col in enumerate(cols):
                val = '' if pd.isna(df.iloc[r, c]) else str(df.iloc[r, c])
                tbl_c = 1 + c if c <= part_idx else 1 + c + 1
                self.table.setItem(r, tbl_c, QTableWidgetItem(val))

            bom_val = str(df.iloc[r, bom_df_idx]).strip()
            if bom_val != '0':
                it = QTableWidgetItem('Removed child Part')
                f = it.font(); f.setBold(True); it.setFont(f)
                it.setForeground(QColor('orange'))
                self.table.setItem(r, tc_tbl_idx, it)
            else:
                # BOM Level 0 -> BLUE row
                for c in range(self.table.columnCount()):
                    cell = self.table.item(r, c)
                    if cell:
                        cell.setBackground(QColor('#87CEEB'))  # Sky Blue

        self.table.resizeColumnsToContents()

    def find_column(self, header):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text() == header:
                return i
        return -1

    def select_oems(self):
        part_col = self.find_column('Part')
        bom_col = self.find_column('BOM Level')
        if part_col < 0 or bom_col < 0:
            return

        eligible = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, part_col).text() if self.table.item(r, part_col) else '').lstrip()
            bom = (self.table.item(r, bom_col).text() if self.table.item(r, bom_col) else '').strip()
            prefix = part[:4]
            if bom != '0' and prefix.isdigit() and int(prefix) >= 500:
                eligible.append(r)

        should_select = any(
            not self.table.cellWidget(r, 0).findChild(QCheckBox).isChecked()
            for r in eligible
        )

        for r in eligible:
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            chk.setChecked(should_select)

    def delete_selected(self):
        rows = []
        for r in range(self.table.rowCount()):
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            if chk.isChecked():
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)

    def copy_removed_child_parts(self):
        part_col = self.find_column('Part')
        tc_col = self.find_column('Tool comments')

        if part_col < 0 or tc_col < 0:
            QMessageBox.warning(self, 'Column Missing',
                                'Required columns (Part / Tool comments) not found.')
            return

        unique_parts = set()

        for r in range(self.table.rowCount()):
            tc_item = self.table.item(r, tc_col)
            if not tc_item or not tc_item.text().strip():
                continue

            part_item = self.table.item(r, part_col)
            if not part_item:
                continue

            part = part_item.text().replace(' ', '').strip()
            if part:
                unique_parts.add(part)

        if not unique_parts:
            QMessageBox.information(self, 'No Data',
                                    'No removed child parts found to copy.')
            return

        from PyQt6.QtGui import QGuiApplication
        result = '\n'.join(sorted(unique_parts))
        QGuiApplication.clipboard().setText(result)

        QMessageBox.information(self, 'Copied',
                                f'Copied {len(unique_parts)} unique removed child part(s) to clipboard.')


class WURemovedBOMItemsTab(QWidget):

    def append_orphans_to_obs_parts(self, remove_from_orphan_table=False):
        if not self.obs_provider:
            return
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphan Child')
        if pcol < 0 or ocol < 0:
            return
        orphan_map = {}
        orphan_rows = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, pcol)
            oit = self.table.item(r, ocol)
            if pit and oit and oit.text().lower().startswith('orphan'):
                orphan_map[pit.text().strip().upper()] = oit.text().strip()
                orphan_rows.append(r)
        if not orphan_map:
            return
        t = self.obs_provider.table
        existing = {t.item(r,1).text().strip().upper() for r in range(t.rowCount()) if t.item(r,1) and t.item(r,1).text().strip()}
        for part, lvl in orphan_map.items():
            if part in existing:
                continue
            target = next((r for r in range(t.rowCount()) if not t.item(r,1) or not t.item(r,1).text().strip()), None)
            if target is None:
                target = t.rowCount(); t.setRowCount(target+1); t._init_rows(target,target+1)
            t.setItem(target,1,QTableWidgetItem(part))
            cb = QComboBox(); cb.addItem(lvl); cb.setCurrentText(lvl)
            col = get_orphan_color(lvl)
            if col: cb.setStyleSheet(f"QComboBox {{ color:{col.name()}; font-weight:bold; }}")
            t.setCellWidget(target,2,cb)
            t.setItem(target,3,QTableWidgetItem('No Replacement'))
        if remove_from_orphan_table:
            for r in reversed(orphan_rows): self.table.removeRow(r)

    def __init__(self, obs_provider):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)

        btn_row = QHBoxLayout()
        self.btn_import = QPushButton("Import WU of Removed BOM items")
        self.btn_analyze = QPushButton("Perform Orphan Analysis")
        self.btn_reset = QPushButton("Reset")
        self.btn_remove_esw = QPushButton("Remove ESW Parents")
        self.btn_remove_9024 = QPushButton("Remove 9024")

        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(self.btn_analyze)
        btn_row.addWidget(self.btn_reset)
        btn_row.addWidget(self.btn_remove_esw)
        btn_row.addWidget(self.btn_remove_9024)
        btn_row.addStretch(1)
        outer.addLayout(btn_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self.import_excel)
        self.btn_analyze.clicked.connect(self.perform_orphan_analysis)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.btn_remove_esw.clicked.connect(lambda: self.remove_by_prefix('ESW'))
        self.btn_remove_9024.clicked.connect(lambda: self.remove_by_prefix('9024'))

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def _find_col(self, name):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text().strip().lower() == name.lower():
                return i
        return -1

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import WU of Removed BOM items", '', "Excel Files (*.xlsx)")
        if not path:
            return

        import pandas as pd
        from PyQt6.QtGui import QColor

        df = pd.read_excel(path, engine='openpyxl')
        if 'Part' not in df.columns:
            QMessageBox.warning(self, 'Missing Column', 'Part column not found')
            return

        cleaned = []
        for _, row in df.iterrows():
            original_part = '' if pd.isna(row['Part']) else str(row['Part'])
            p = original_part.lstrip()
            lp = len(p)

            keep = False
            if lp > 10 and p.upper().startswith('ESW'):
                keep = True
            elif lp == 10 and p[4] == '-':
                keep = True

            if keep:
                cleaned.append(row)

        if not cleaned:
            QMessageBox.information(self, 'No Valid Data', 'All rows removed by cleanup rules.')
            return

        clean_df = pd.DataFrame(cleaned)
        clean_df.insert(2, 'Orphan Child', '')

        headers = ['Select'] + list(clean_df.columns)
        self.table.setRowCount(len(clean_df))
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        for r in range(len(clean_df)):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            self.table.setCellWidget(r, 0, cont)

            wu_zero = False
            if 'WU Level' in clean_df.columns:
                wu_zero = str(clean_df.iloc[r]['WU Level']).strip() == '0'

            for c, col in enumerate(clean_df.columns):
                val = '' if pd.isna(clean_df.iloc[r, c]) else str(clean_df.iloc[r, c])
                item = QTableWidgetItem(val)
                if wu_zero:
                    item.setBackground(QColor('#87CEEB'))
                self.table.setItem(r, c + 1, item)

        self.table.resizeColumnsToContents()

    def _build_obs_change_map(self):
        mapping = {}
        t = self.obs_provider.table
        for r in range(t.rowCount()):
            obs_it = t.item(r, 1)
            chg_w = t.cellWidget(r, 2)
            key = obs_it.text().strip().upper() if obs_it else ''
            if key:
                mapping[key] = chg_w.currentText() if chg_w else ''
        return mapping

    def perform_orphan_analysis(self):
        # Find required columns
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphan Child')
        wucol = self._find_col('WU Level')

        if pcol < 0 or ocol < 0 or wucol < 0:
            QMessageBox.information(self, 'Orphan Analysis', 'Required columns not found.')
            return

        row_count = self.table.rowCount()

        def is_zero(v):
            return str(v).strip() in ('0', '0.0')

        def is_one(v):
            return str(v).strip() in ('1', '1.0')

        ignore_prefixes = ('0243', '0299', '0289', '0290')

        # Step 0: OBS → Obsolete / Inactivate mapping (existing behavior)
        obs_map = self._build_obs_change_map()
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if not it:
                continue
            key = it.text().lstrip().upper()
            if key in obs_map:
                self.table.setItem(r, ocol, QTableWidgetItem(obs_map[key]))

        # Step 1: Orphan1
        orphan_parts = set()
        r = 0
        while r < row_count:
            wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
            if is_zero(wu):
                child_row = r
                part_it = self.table.item(child_row, pcol)
                orphan_it = self.table.item(child_row, ocol)
                part = part_it.text().strip() if part_it else ''

                if orphan_it and orphan_it.text().strip():
                    r += 1
                    continue

                block_end = row_count
                for i in range(r + 1, row_count):
                    nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if is_zero(nxt):
                        block_end = i
                        break

                parents = []
                for i in range(child_row + 1, block_end):
                    wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if not is_one(wu_p):
                        continue
                    pit = self.table.item(i, pcol)
                    pval = pit.text().lstrip() if pit else ''
                    if pval.startswith(ignore_prefixes):
                        continue
                    parents.append(i)

                mark = False
                if not parents:
                    mark = True
                else:
                    all_bad = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        status = oit.text().strip().lower() if oit else ''
                        if status not in ('obsolete', 'inactivate'):
                            all_bad = False
                            break
                    if all_bad:
                        mark = True

                if mark and part:
                    self.table.setItem(child_row, ocol, QTableWidgetItem('Orphan1'))
                    orphan_parts.add(part.upper())

                r = block_end
            else:
                r += 1

        # Propagate Orphan1
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if it and it.text().strip().upper() in orphan_parts:
                oit = self.table.item(r, ocol)
                if not oit or not oit.text().strip():
                    self.table.setItem(r, ocol, QTableWidgetItem('Orphan1'))

        # Step 2+: Orphan2, Orphan3...
        current_level = 2
        while True:
            new_found = False
            new_parts = set()
            r = 0
            while r < row_count:
                wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
                orphan_it = self.table.item(r, ocol)
                if is_zero(wu) and (not orphan_it or not orphan_it.text().strip()):
                    part_it = self.table.item(r, pcol)
                    part = part_it.text().strip() if part_it else ''

                    block_end = row_count
                    for i in range(r + 1, row_count):
                        nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if is_zero(nxt):
                            block_end = i
                            break

                    parents = []
                    for i in range(r + 1, block_end):
                        wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if not is_one(wu_p):
                            continue
                        pit = self.table.item(i, pcol)
                        pval = pit.text().lstrip() if pit else ''
                        if pval.startswith(ignore_prefixes):
                            continue
                        parents.append(i)

                    parents_filled = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        if not oit or not oit.text().strip():
                            parents_filled = False
                            break

                    if part and (not parents or parents_filled):
                        self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))
                        new_parts.add(part.upper())
                        new_found = True
                r += 1

            for r in range(row_count):
                it = self.table.item(r, pcol)
                oit = self.table.item(r, ocol)
                if it and it.text().strip().upper() in new_parts and (not oit or not oit.text().strip()):
                    self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))

            if not new_found:
                break
            current_level += 1

            self.append_orphans_to_obs_parts()
        from PyQt6.QtGui import QColor
        for r in range(self.table.rowCount()):
            wu = self.table.item(r, wucol)
            if wu and wu.text().strip() == '0':
                for c in range(self.table.columnCount()):
                    it = self.table.item(r, c)
                    if it:
                        it.setBackground(QColor('#87CEEB'))


    def remove_by_prefix(self, prefix):
        pcol = self._find_col('Part')
        if pcol < 0:
            return
        rows = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, pcol)
            if it and it.text().lstrip().startswith(prefix):
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)


class  OBSAllPartsWithoutReplacementTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)

        subtabs = QTabWidget()
        # Equal width secondary tabs
        subtabs.setStyleSheet("""
        QTabBar::tab {
            min-width: 220px;
            padding: 6px 12px;
            text-align: center;
        }
        """)

        subtabs.addTab(OrphanOBSSubTab(), "Imp BOM")
        subtabs.addTab(WURemovedBOMItemsTab(self.obs_provider), "WU of Removed BOM Items")

        outer.addWidget(subtabs)

class _OrphanAnalysisTab_OLD(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        outer = QVBoxLayout(self)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        self.subtabs = QTabWidget()
        self.subtabs.setStyleSheet("""
        QTabBar::tab:selected {
            background-color: #87CEEB;
            color: #0F2D46;
            font-weight: 600;
        }
        QTabBar::tab {
            background-color: #EAF6FD;
        }
        """)
        self.subtabs.addTab(
            OBSAllPartsWithoutReplacementTab(obs_provider),
            "OBS all Parts without Replacements"
        )
        self.subtabs.addTab(
            PlaceholderTab("OBS with or without Replacement"),
            "OBS with or without Replacement"
        )

        outer.addWidget(self.subtabs)




# === EC Creation Inputs Form (Embedded) ===
from PyQt6.QtWidgets import QRadioButton
import sys
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget,
    QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QFrame,
    QButtonGroup, QTextEdit, QScrollArea, QPushButton, QFileDialog
)
from PyQt6.QtGui import QFont

APP_TITLE = "ECR Kit Assistant"

EC_CATEGORY_DESC = {
    "A1": "SMBoM Options as revised items and having CDW",
    "A2": "SMBoM Options as revised items, and No CDW",
    "B1": "No SMBoM Options as revised items and having CDW",
    "B2": "No SMBoM Options as revised items, No CDW, and revised item status at Eval and/or moving to Eval or adding already released parts to Proto buckets",
    "B3": "No SMBoM Option as revised items, No CDW, and revised Item status at Production and/or moving to Production",
}


def header(text):
    lbl = QLabel(text)
    lbl.setFont(QFont("Segoe UI", 12, QFont.Weight.DemiBold))
    return lbl


def highlight_label(text):
    lbl = QLabel(text)
    lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
    lbl.setStyleSheet("color:#1F4E79")
    return lbl


class LimitedTextEdit(QTextEdit):
    def __init__(self, limit=2000):
        super().__init__()
        self.limit = limit
        self.textChanged.connect(self._limit)

    def _limit(self):
        text = self.toPlainText()
        if len(text) > self.limit:
            self.blockSignals(True)
            self.setPlainText(text[:self.limit])
            self.moveCursor(QTextCursor.MoveOperation.End)
            self.blockSignals(False)

    def _update_counter(self):
        if not self.counter_label:
            return
        count = len(self.toPlainText())
        self.counter_label.setText(f"{count} / {self.limit} characters")
        if count >= self.limit:
            self.counter_label.setStyleSheet("color:red;font-size:10px")
        else:
            self.counter_label.setStyleSheet("color:gray;font-size:10px")

class ECCreationInputsFormTab(QWidget):
    def __init__(self):
        super().__init__()
        self.email_file_path = None  # Stores the path of the browsed email file
        root = QVBoxLayout(self)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        root.addWidget(scroll)
        container = QWidget(); scroll.setWidget(container)
        outer = QVBoxLayout(container)

        # ---- Section A ----
        outer.addWidget(header("Section A: EC Category Form"))
        secA = QFrame(); a = QVBoxLayout(secA)

        scope_row = QHBoxLayout(); scope_grp = QButtonGroup(self)
        for scope in ["Up Revision", "Status Roll", "Production Release", "OBS / Inactivate", "Product Release"]:
            rb = QRadioButton(scope)
            scope_grp.addButton(rb); rb.toggled.connect(self.start_flow)
            scope_row.addWidget(rb)
        a.addLayout(scope_row)

        self.flow_area = QVBoxLayout(); a.addLayout(self.flow_area)
        self.ec_result_lbl = QLabel(""); self.ec_result_lbl.setWordWrap(True)
        self.ec_result_lbl.setStyleSheet("color:green;font-weight:600;font-size:14px")
        a.addWidget(self.ec_result_lbl)

        self.ec_divider = QFrame(); self.ec_divider.setFrameShape(QFrame.Shape.HLine)
        self.ec_divider.setVisible(False); a.addWidget(self.ec_divider)
        outer.addWidget(secA)

        outer.addSpacing(12)

# ---- Section B ----
        self.secB_header = header("Section B: ECR Change Assessment")
        self.secB_header.setVisible(False)
        outer.addWidget(self.secB_header)

        self.secB = QFrame()
        self.secB.setVisible(False)
        outer.addWidget(self.secB)

        # === Main horizontal layout for Section B ===
        section_b_layout = QHBoxLayout(self.secB)
        section_b_left = QVBoxLayout()
        section_b_right = QVBoxLayout()

    #    setContentsMargins(left, top, right, bottom)
        BOX_WIDTH = 240          # width of the box
        BOX_HEIGHT = 300         # height of the box
        BOX_OFFSET_FROM_RIGHT = 500   # 👈 INCREASE → moves box LEFT


        section_b_layout.addLayout(section_b_left, 3)   # Left: questions
        section_b_layout.addLayout(section_b_right, 1)  # Right: PN box


        # Reference fields
        self.ref_boxes = {}
        self.ref_radios = {}
        self.ref_labels = {}

        def yes_no_with_box(key, label_text):
                row = QHBoxLayout()

                lbl = QLabel(label_text)
                row.addWidget(lbl)

                grp = QButtonGroup(self)
                rb_y = QRadioButton("Yes")
                rb_n = QRadioButton("No")
                grp.addButton(rb_y)
                grp.addButton(rb_n)

                row.addWidget(rb_y)
                row.addWidget(rb_n)

                txt = QTextEdit()
                txt.setFixedHeight(30)
                txt.setMaximumWidth(500)
                txt.setStyleSheet("background:#FFF2CC; border:none")
                txt.setVisible(False)

                row.addWidget(txt)
                row.addStretch(1)

                rb_y.toggled.connect(lambda c: txt.setVisible(c))
                rb_n.toggled.connect(lambda c: txt.setVisible(False))

                self.ref_boxes[key] = txt
                self.ref_radios[key] = (rb_y, rb_n)
                self.ref_labels[key] = lbl

                section_b_left.addLayout(row)


        yes_no_with_box("PCR_PCN", "1.  Does this Project include Product Change Request (PCR)")
        yes_no_with_box("PSN", "             Is there any Associated Product Safety Note (PSN)")
        yes_no_with_box("PCR", "2.  Associated Project created for PCR")
        yes_no_with_box("SPS", "3.  Associated Open SPSs")
        yes_no_with_box("ESW", "4.  Associated ESWs")
        yes_no_with_box("REF_ECR", "5.  Reference ECR Number(s)")


        # PCR → PSN dependency
        pcr_yes, pcr_no = self.ref_radios["PCR_PCN"]
        psn_yes, psn_no = self.ref_radios["PSN"]
        psn_txt = self.ref_boxes["PSN"]
        psn_lbl = self.ref_labels["PSN"]

        def disable_psn():
                # Temporarily disable auto-exclusive to allow both radios to be unchecked
                grp_psn = psn_yes.group()
                if grp_psn:
                    grp_psn.setExclusive(False)
                
                # Block signals and uncheck both
                psn_yes.blockSignals(True)
                psn_no.blockSignals(True)
                psn_yes.setChecked(False)
                psn_no.setChecked(False)
                psn_yes.blockSignals(False)
                psn_no.blockSignals(False)
                
                # Re-enable auto-exclusive
                if grp_psn:
                    grp_psn.setExclusive(True)
                
                # Disable and hide
                psn_yes.setEnabled(False)
                psn_no.setEnabled(False)
                psn_lbl.setEnabled(False)
                psn_txt.setPlainText("")
                psn_txt.setVisible(False)

        def enable_psn():
                psn_yes.setEnabled(True)
                psn_no.setEnabled(True)
                psn_lbl.setEnabled(True)

        pcr_no.toggled.connect(lambda c: disable_psn() if c else None)
        pcr_yes.toggled.connect(lambda c: enable_psn() if c else None)


        # Reference Email Attachments
        rowm = QHBoxLayout()
        rowm.addWidget(QLabel("6.  Reference e-mail/attachments?"))

        grp_m = QButtonGroup(self)
        rb_my = QRadioButton("Yes")
        rb_mn = QRadioButton("No")
        grp_m.addButton(rb_my)
        grp_m.addButton(rb_mn)
        self.ref_attachment_yes_radio = rb_my
        self.ref_attachment_no_radio = rb_mn

        rowm.addWidget(rb_my)
        rowm.addWidget(rb_mn)

        browse = QPushButton("Browse Attachment")
        browse.setFixedHeight(24)
        browse.setVisible(False)

        rowm.addWidget(browse)
        rowm.addStretch(1)
        section_b_left.addLayout(rowm)

        rb_my.toggled.connect(self._on_reference_attachments_toggled)
        browse.clicked.connect(self._on_browse_email_clicked)
        self._browse_email_btn = browse  # keep reference to update button label


        # Impact caused by
        row = QHBoxLayout()
        row.addWidget(QLabel("7.  Impact caused by:"))

        grp_sc = QButtonGroup(self)
        rb_sup = QRadioButton("Supplier")
        rb_cust = QRadioButton("Customer")
        grp_sc.addButton(rb_sup)
        grp_sc.addButton(rb_cust)
        self.impact_supplier_radio = rb_sup
        self.impact_customer_radio = rb_cust

        row.addWidget(rb_sup)
        row.addWidget(rb_cust)
        row.addStretch(1)
        section_b_left.addLayout(row)

        sub_row = QHBoxLayout()
        sub_row.setContentsMargins(20, 0, 0, 0)

        grp_c = QButtonGroup(self)
        rb_int = QRadioButton("Internal")
        rb_ext = QRadioButton("External")
        grp_c.addButton(rb_int)
        grp_c.addButton(rb_ext)
        self.impact_internal_radio = rb_int
        self.impact_external_radio = rb_ext

        sub_row.addWidget(rb_int)
        sub_row.addWidget(rb_ext)
        sub_row.addStretch(1)
        section_b_left.addLayout(sub_row)

        def clear_customer_scope_selection():
            # Temporarily disable exclusivity so both radios can be fully unchecked.
            grp_c.setExclusive(False)
            rb_int.setChecked(False)
            rb_ext.setChecked(False)
            grp_c.setExclusive(True)

        rb_sup.toggled.connect(
            lambda c: (
                clear_customer_scope_selection(),
                rb_int.setEnabled(False),
                rb_ext.setEnabled(False)
            ) if c else None
        )

        rb_cust.toggled.connect(
            lambda c: (
                rb_int.setEnabled(c),
                rb_ext.setEnabled(c)
            )
        )


        yes_no_with_box("QN", "8. Is there any QN")

        # Reason Code
        rc_row = QHBoxLayout()
        rc_row.addWidget(QLabel("9.  ECR Reason Code:"))
        self.reason_cb = QComboBox()
        self.reason_cb.addItems([
            "Beyond Spec Request","Cap Code Management","CES","Cost Reduction",
            "Design Correction","Document Correction","Manufacturing Improvement",
            "Obsolescence","Option Reduction and Product End of Life","Order BOM Change",
            "Product Improvement","Product Release","Safety Event"
        ])
        rc_row.addWidget(self.reason_cb)
        rc_row.addStretch(1)
        section_b_left.addLayout(rc_row)


        scope_lbl = QLabel("Initial Scope Part Numbers")
        scope_lbl.setStyleSheet("font-weight: bold;")

        self.scope_parts_txt = QTextEdit()
        self.scope_parts_txt.setPlaceholderText(
                "Enter Part Number (up to 500 lines)"
        )

        # --- Size tuning ---
        self.scope_parts_txt.setFixedWidth(235)      # ~10% narrower
        self.scope_parts_txt.setMaximumHeight(280)   # slightly taller

        # --- Scrollbars ---
        self.scope_parts_txt.setVerticalScrollBarPolicy(
                Qt.ScrollBarPolicy.ScrollBarAsNeeded
        )
        self.scope_parts_txt.setHorizontalScrollBarPolicy(
                Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )

        # --- Styling (box + scrollbar) ---
        self.scope_parts_txt.setStyleSheet("""
                QTextEdit {
                        background: #F8F8F8;
                        border: 1px solid #999;
                }
                QScrollBar:vertical {
                        background: #E6E6E6;
                        width: 10px;
                        margin: 0px;
                }
                QScrollBar::handle:vertical {
                        background: #A6A6A6;
                        min-height: 20px;
                        border-radius: 4px;
                }
                QScrollBar::handle:vertical:hover {
                        background: #8C8C8C;
                }
                QScrollBar::add-line:vertical,
                QScrollBar::sub-line:vertical {
                        height: 0px;
                }
        """)

        
        section_b_right.setContentsMargins(0, 0, 300, 0)   # 👈 right margin pushes box left
        section_b_right.addWidget(scope_lbl)
        section_b_right.addWidget(self.scope_parts_txt)
        section_b_right.addSpacing(10)

        include_tabs_lbl = QLabel("Include Data From Tabs")
        include_tabs_lbl.setStyleSheet("font-weight: bold;")
        section_b_right.addWidget(include_tabs_lbl)

        self.cb_where_used = QCheckBox("Where Used")
        self.cb_obs_parts = QCheckBox("OBS Parts")
        self.cb_structure_sheet = QCheckBox("Structure Sheet")

        section_b_right.addWidget(self.cb_where_used)
        section_b_right.addWidget(self.cb_obs_parts)
        section_b_right.addWidget(self.cb_structure_sheet)

        self.selected_tab_payload = {}

        self.cb_where_used.stateChanged.connect(self.refresh_selected_tab_payload)
        self.cb_obs_parts.stateChanged.connect(self.refresh_selected_tab_payload)
        self.cb_structure_sheet.stateChanged.connect(self.refresh_selected_tab_payload)

        section_b_right.addStretch(1)
       
    # Short Title
        title_row = QHBoxLayout()
        title_row.addWidget(highlight_label("Short Title"))
        title_row.addStretch(1)
        section_b_left.addLayout(title_row)

        self.short_title_edit = QLineEdit()
        self.short_title_edit.setPlaceholderText("Enter short title (max 75 characters)")
        self.short_title_edit.setMaxLength(75)
        self.short_title_edit.setFixedHeight(30)
        self.short_title_edit.setMaximumWidth(900)
        section_b_left.addWidget(self.short_title_edit)

        section_b_left.addSpacing(6)

# Problem Summary
        ps_row = QHBoxLayout()
        ps_row.addWidget(
                highlight_label("Problem Summary from PCR, PCN, SPS and ESW")
        )
        self.btn_ps = QPushButton("Problem Summary")
        self.btn_ps.setFixedSize(180, 26)
        ps_row.addWidget(self.btn_ps)
        ps_row.addStretch(1)

        section_b_left.addLayout(ps_row)

        self.problem_txt = LimitedTextEdit(2000)
        self.problem_txt.setMinimumHeight(140)
        self.problem_txt.setMaximumWidth(900)
        section_b_left.addWidget(self.problem_txt)

        self.btn_ps.clicked.connect(self.on_problem_summary_clicked)

        section_b_left.addSpacing(6)

        # Solution
        sol_row = QHBoxLayout()
        sol_row.addWidget(highlight_label("Proposed Solution"))
        btn_sol = QPushButton("Generate Proposed Solution")
        btn_sol.setFixedSize(180, 26)
        sol_row.addWidget(btn_sol)
        sol_row.addStretch(1)

        section_b_left.addLayout(sol_row)

        self.solution_txt = LimitedTextEdit(2000)
        self.solution_txt.setMinimumHeight(140)
        self.solution_txt.setMaximumWidth(900)
        section_b_left.addWidget(self.solution_txt)

        btn_sol.clicked.connect(
                lambda: self.solution_txt.setPlainText("Will be enabled in Future")
        )

        outer.addWidget(self.secB)
        outer.addStretch(1)


    def refresh_selected_tab_payload(self):
        main = self.window()
        payload = {}

        if self.cb_where_used.isChecked():
            whereused_tab = getattr(main, "whereused_tab", None)
            if whereused_tab and hasattr(whereused_tab, "table"):
                payload["Where Used"] = self._table_to_rows(whereused_tab.table)

        if self.cb_obs_parts.isChecked():
            obs_tab = getattr(main, "obs_tab", None)
            if obs_tab and hasattr(obs_tab, "table"):
                payload["OBS Parts"] = self._table_to_rows(obs_tab.table)

        if self.cb_structure_sheet.isChecked():
            structure_tab = getattr(main, "structure_tab", None)
            if structure_tab:
                structure_data = {}

                impacted_text = getattr(structure_tab, "impacted_text", None)
                if impacted_text:
                    structure_data["Impacted Options/Parts"] = [
                        line.strip()
                        for line in impacted_text.toPlainText().splitlines()
                        if line.strip()
                    ]

                structure_table = getattr(structure_tab, "table", None)
                if structure_table:
                    structure_data["Structure Sheet"] = self._table_to_rows(structure_table)

                payload["Structure Sheet"] = structure_data

        self.selected_tab_payload = payload

    def get_selected_tab_payload(self):
        self.refresh_selected_tab_payload()
        return self.selected_tab_payload

    def _table_to_rows(self, table):
        if table is None or table.columnCount() == 0:
            return []

        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text().strip() if h and h.text() else f"Column {c + 1}")

        rows = []
        for r in range(table.rowCount()):
            row_data = {}
            has_value = False

            for c, header in enumerate(headers):
                value = ""
                cell_widget = table.cellWidget(r, c)

                if isinstance(cell_widget, QComboBox):
                    value = cell_widget.currentText().strip()
                elif isinstance(cell_widget, QTextEdit):
                    value = cell_widget.toPlainText().strip()
                elif isinstance(cell_widget, QWidget):
                    chk = cell_widget.findChild(QCheckBox)
                    if chk:
                        value = "Yes" if chk.isChecked() else "No"
                    else:
                        item = table.item(r, c)
                        value = item.text().strip() if item and item.text() else ""
                else:
                    item = table.item(r, c)
                    value = item.text().strip() if item and item.text() else ""

                if value:
                    has_value = True
                row_data[header] = value

            if has_value:
                rows.append(row_data)

        return rows
# ...existing code...

    def _selected_radio_text(self, key: str) -> str:
        pair = self.ref_radios.get(key)
        if not pair:
            return ""
        rb_y, rb_n = pair
        if rb_y.isChecked():
            return "Yes"
        if rb_n.isChecked():
            return "No"
        return ""

# ...existing code...
    def _collect_pss_payload(self) -> Dict[str, Any]:
        self.refresh_selected_tab_payload()

        scope_parts = [
            ln.strip()
            for ln in self.scope_parts_txt.toPlainText().splitlines()
            if ln.strip()
        ]

        reference_inputs = {}
        for key, txt in self.ref_boxes.items():
            reference_inputs[key] = {
                "answer": self._selected_radio_text(key),   # "Yes" / "No" / ""
                "text": txt.toPlainText().strip(),
            }

        return {
            "short_title": self.short_title_edit.text().strip(),
            "reason_code": self.reason_cb.currentText().strip(),
            "scope_parts": scope_parts,
            "reference_inputs": reference_inputs,
            "selected_tab_payload": self.selected_tab_payload,
            "current_problem_text": self.problem_txt.toPlainText().strip(),
            "proposed_solution_text": self.solution_txt.toPlainText().strip(),
            "include_tabs_flags": {
                "where_used": self.cb_where_used.isChecked(),
                "obs_parts": self.cb_obs_parts.isChecked(),
                "structure_sheet": self.cb_structure_sheet.isChecked(),
            },
        }
# ...existing code...

    def _load_ai_pss_module(self):
        mod_path = Path(__file__).with_name("AI_Assisted_PSS.py")
        if not mod_path.exists():
            raise FileNotFoundError(f"AI_Assisted_PSS.py not found: {mod_path}")

        spec = importlib.util.spec_from_file_location("AI_Assisted_PSS", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load AI_Assisted_PSS module spec")

        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    # ...existing code...

    def _normalize_pss_output(self, out: Any) -> Dict[str, str]:
        if isinstance(out, dict):
            title = str(out.get("title") or "").strip()
            problem = str(
                out.get("problem_statement")
                or out.get("problem_summary")
                or out.get("problem")
                or ""
            ).strip()
            solution = str(
                out.get("solution_statement")
                or out.get("solution")
                or ""
            ).strip()
            return {"title": title, "problem": problem, "solution": solution}

        text = str(out or "").strip()
        return {"title": "", "problem": text, "solution": ""}

    def _run_ai_pss_full(self, payload: Dict[str, Any]) -> Dict[str, str]:
        module = self._load_ai_pss_module()

        fn_full = getattr(module, "generate_full_pss", None)
        if callable(fn_full):
            try:
                out = fn_full(payload)
            except TypeError:
                out = fn_full()
            return self._normalize_pss_output(out)

        for fn_name in ("generate_problem_summary", "generate_pss", "run_problem_summary"):
            fn = getattr(module, fn_name, None)
            if callable(fn):
                try:
                    out = fn(payload)
                except TypeError:
                    out = fn()
                return self._normalize_pss_output(out)

        raise AttributeError(
            "No supported entry point in AI_Assisted_PSS.py. "
            "Expected generate_full_pss(payload) or generate_problem_summary(payload)."
        )

# ...existing code...

# ...existing code...

    def _run_ai_pss(self, payload: Dict[str, Any]) -> str:
        out = self._run_ai_pss_full(payload)
        return out.get("problem", "")

# ...existing code...

        for fn_name in ("generate_problem_summary", "generate_pss", "run_problem_summary"):
            fn = getattr(module, fn_name, None)
            if callable(fn):
                try:
                    out = fn(payload)
                except TypeError:
                    out = fn()

                if isinstance(out, dict):
                    out = out.get("problem_summary") or out.get("problem") or ""
                return str(out or "").strip()

        for cls_name in ("AIAssistedPSS", "ProblemSolutionAgent", "PSSAgent"):
            cls = getattr(module, cls_name, None)
            if cls is None:
                continue
            obj = cls()
            for m_name in ("generate_problem_summary", "generate_pss", "run"):
                method = getattr(obj, m_name, None)
                if callable(method):
                    try:
                        out = method(payload)
                    except TypeError:
                        out = method()

                    if isinstance(out, dict):
                        out = out.get("problem_summary") or out.get("problem") or ""
                    return str(out or "").strip()

        raise AttributeError(
            "No supported entry point in AI_Assisted_PSS.py. "
            "Expected generate_problem_summary(payload) or equivalent."
        )

# ...existing code...

    def _on_reference_attachments_toggled(self, checked: bool):
        """Show/hide browse button and clear selected attachment when toggled off."""
        if hasattr(self, "_browse_email_btn"):
            self._browse_email_btn.setVisible(checked)

        if not checked:
            self.email_file_path = None
            if hasattr(self, "_browse_email_btn"):
                self._browse_email_btn.setText("Browse Attachment")
                self._browse_email_btn.setToolTip("")

    def _on_browse_email_clicked(self):
        """Open file dialog, store selected attachment path, update button label."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Reference Attachment",
            "",
            "Supported Files (*.msg *.eml *.txt *.pdf *.ppt *.pptx *.doc *.docx)"
        )
        if file_path:
            self.email_file_path = file_path
            file_name = Path(file_path).name
            self._browse_email_btn.setText(f"📎 {file_name}")
            self._browse_email_btn.setToolTip(file_path)
            QMessageBox.information(
                self,
                "Attachment Loaded",
                f"Reference attachment loaded:\n{file_name}\n\nClick 'Problem Summary' to process it.",
            )

    def _validate_problem_summary_inputs(self) -> str:
        """Validate ECR change assessment inputs before generating Problem Summary."""
        missing_answers = []
        missing_values = []
        qn_missing_answer = None
        qn_missing_value = None
        pcr_pair = self.ref_radios.get("PCR_PCN")
        pcr_yes_selected = bool(pcr_pair and pcr_pair[0].isChecked())
        field_names = {
            "PCR_PCN": "PCR",
            "PSN": "PSN",
            "PCR": "PCR Project",
            "SPS": "SPS",
            "ESW": "ESW",
            "REF_ECR": "Reference ECR",
            "QN": "QN",
        }

        for key, (rb_yes, rb_no) in self.ref_radios.items():
            # PSN is auto-derived from PCR records when Question 1 (PCR_PCN) is Yes.
            # Do not require manual PSN Yes/No or value in that flow.
            if key == "PSN" and pcr_yes_selected:
                continue

            # Skip questions that are currently disabled by dependency logic (for example PSN).
            if not rb_yes.isEnabled() and not rb_no.isEnabled():
                continue

            if not rb_yes.isChecked() and not rb_no.isChecked():
                label = self.ref_labels.get(key)
                q_text = label.text() if label else key
                if key == "QN":
                    qn_missing_answer = q_text
                else:
                    missing_answers.append(q_text)
                continue

            if rb_yes.isChecked():
                value_widget = self.ref_boxes.get(key)
                value_text = value_widget.toPlainText().strip() if value_widget else ""
                if not value_text:
                    if key == "QN":
                        qn_missing_value = field_names.get(key, key)
                    else:
                        missing_values.append(field_names.get(key, key))

        attachment_yes = getattr(self, "ref_attachment_yes_radio", None)
        attachment_no = getattr(self, "ref_attachment_no_radio", None)
        if attachment_yes and attachment_no:
            if not attachment_yes.isChecked() and not attachment_no.isChecked():
                missing_answers.append(
                    "6. Reference e-mail/attachments"
                )
            elif attachment_yes.isChecked() and not getattr(self, "email_file_path", None):
                missing_answers.append(
                    "6. Reference e-mail/attachments: please attach relevant attachments"
                )

        impact_supplier = getattr(self, "impact_supplier_radio", None)
        impact_customer = getattr(self, "impact_customer_radio", None)
        impact_internal = getattr(self, "impact_internal_radio", None)
        impact_external = getattr(self, "impact_external_radio", None)

        if impact_supplier and impact_customer:
            if not impact_supplier.isChecked() and not impact_customer.isChecked():
                missing_answers.append("7. Impact caused by: select Supplier or Customer")
            elif impact_customer.isChecked() and impact_internal and impact_external:
                if not impact_internal.isChecked() and not impact_external.isChecked():
                    missing_answers.append(
                        "7. Impact caused by Customer: select Internal or External"
                    )

        if qn_missing_answer:
            missing_answers.append(qn_missing_answer)

        if qn_missing_value:
            missing_values.append(qn_missing_value)

        if missing_answers:
            return (
                "Please address Assessment question(s) with Yes/No selection:\n\n"
                + "\n".join(f"{q}" for q in missing_answers)
            )

        if missing_values:
            return (
                "Please provide details in the corresponding field for these 'Yes' selections:\n\n"
                + "\n".join(f"{name}: details/number is missing" for name in missing_values)
            )

        return ""

    # ------------------------------------------------------------------
    # PCR-driven Problem Summary helpers (Question 1 – Databricks path)
    # ------------------------------------------------------------------

    def _load_pcr_query_module(self):
        """Load pcr_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "pcr_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"pcr_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("pcr_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load pcr_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_project_query_module(self):
        """Load project_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "project_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"project_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("project_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load project_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_sps_query_module(self):
        """Load sps_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "sps_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"sps_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("sps_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load sps_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_esw_query_module(self):
        """Load esw_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "esw_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"esw_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("esw_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load esw_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _fetch_q2_project_context(self) -> Dict[str, Any]:
        """Fetch Question 2 (Associated Project created for PCR) records if enabled.

        Returns empty structure when Q2 is not selected as Yes.
        """
        pair = self.ref_radios.get("PCR")
        if not pair:
            self._project_fetch_result = {"skipped": [], "not_found": []}
            return {"project_records": [], "skipped_projects": [], "not_found_projects": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._project_fetch_result = {"skipped": [], "not_found": []}
            return {"project_records": [], "skipped_projects": [], "not_found_projects": []}

        project_text = self.ref_boxes["PCR"].toPlainText().strip()
        if not project_text:
            raise ValueError(
                "Please enter at least one Project number in Question 2 "
                "(Associated Project created for PCR)."
            )

        mod = self._load_project_query_module()
        result = mod.fetch_project_records(project_text)

        project_records = result.get("valid", [])
        skipped_projects = result.get("skipped", [])
        not_found_projects = result.get("not_found", [])

        self._project_fetch_result = {
            "skipped": skipped_projects,
            "not_found": not_found_projects,
        }

        return {
            "project_records": project_records,
            "skipped_projects": skipped_projects,
            "not_found_projects": not_found_projects,
        }

    def _fetch_q3_sps_context(self) -> Dict[str, Any]:
        """Fetch Question 3 (Associated Open SPSs) records if enabled.

        Returns empty structure when Q3 is not selected as Yes.
        """
        pair = self.ref_radios.get("SPS")
        if not pair:
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            return {"sps_records": [], "skipped_sps": [], "not_found_sps": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            return {"sps_records": [], "skipped_sps": [], "not_found_sps": []}

        sps_text = self.ref_boxes["SPS"].toPlainText().strip()
        if not sps_text:
            raise ValueError(
                "Please enter at least one SPS number in Question 3 "
                "(Associated Open SPSs)."
            )

        mod = self._load_sps_query_module()
        result = mod.fetch_sps_records(sps_text)

        sps_records = result.get("valid", [])
        skipped_sps = result.get("skipped", [])
        not_found_sps = result.get("not_found", [])

        self._sps_fetch_result = {
            "skipped": skipped_sps,
            "not_found": not_found_sps,
        }

        return {
            "sps_records": sps_records,
            "skipped_sps": skipped_sps,
            "not_found_sps": not_found_sps,
        }

    def _fetch_q4_esw_context(self) -> Dict[str, Any]:
        """Fetch Question 4 (Associated ESWs) records if enabled.

        Returns empty structure when Q4 is not selected as Yes.
        """
        pair = self.ref_radios.get("ESW")
        if not pair:
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            return {"esw_records": [], "skipped_esw": [], "not_found_esw": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            return {"esw_records": [], "skipped_esw": [], "not_found_esw": []}

        esw_text = self.ref_boxes["ESW"].toPlainText().strip()
        if not esw_text:
            raise ValueError(
                "Please enter at least one ESW number in Question 4 "
                "(Associated ESWs)."
            )

        mod = self._load_esw_query_module()
        result = mod.fetch_esw_records(esw_text)

        esw_records = result.get("valid", [])
        # Keep an alias so downstream code that expects ec_number can still read ESW IDs.
        for rec in esw_records:
            if rec.get("ec_number") is None and rec.get("esw_number") is not None:
                rec["ec_number"] = rec.get("esw_number")
        skipped_esw = result.get("skipped", [])
        not_found_esw = result.get("not_found", [])

        self._esw_fetch_result = {
            "skipped": skipped_esw,
            "not_found": not_found_esw,
        }

        return {
            "esw_records": esw_records,
            "skipped_esw": skipped_esw,
            "not_found_esw": not_found_esw,
        }

    def _append_project_context(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Enrich payload with Q2 project, Q3 SPS and Q4 ESW context when applicable."""
        enriched = dict(payload or {})
        project_ctx = self._fetch_q2_project_context()
        sps_ctx = self._fetch_q3_sps_context()
        esw_ctx = self._fetch_q4_esw_context()
        enriched.update(project_ctx)
        enriched.update(sps_ctx)
        enriched.update(esw_ctx)
        return enriched

    def _extract_numeric_ids(self, text: str) -> List[int]:
        """Extract numeric IDs from free-form text preserving first-seen order."""
        raw = re.findall(r"\d+", text or "")
        if not raw:
            return []
        return list(dict.fromkeys(int(x) for x in raw))

    def _fetch_existing_ec_numbers(self, ids: List[int]) -> List[int]:
        """Return IDs that exist in tbl_projectx_ec.ec_number."""
        if not ids:
            return []

        try:
            import pyodbc
        except Exception as exc:
            raise RuntimeError(f"pyodbc import failed: {exc}") from exc

        # Databricks Hive SQL does not support positional ? parameters; inline validated ints.
        placeholders = ", ".join(str(n) for n in ids)
        sql = f"""
        SELECT DISTINCT ec.ec_number
        FROM prd.rd_core.tbl_projectx_ec ec
        WHERE ec.ec_number IN ({placeholders})
        """

        conn = pyodbc.connect("DSN=Spark-PRD", autocommit=True)
        try:
            cursor = conn.cursor()
            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
            finally:
                cursor.close()
        finally:
            conn.close()

        found = []
        for row in rows:
            if row and row[0] is not None:
                try:
                    found.append(int(row[0]))
                except Exception:
                    continue
        return list(dict.fromkeys(found))

    def _fetch_ec_records(self, ids: List[int]) -> List[Dict[str, Any]]:
        """Return EC records with problem/solution/status for the provided IDs."""
        if not ids:
            return []

        try:
            import pyodbc
        except Exception as exc:
            raise RuntimeError(f"pyodbc import failed: {exc}") from exc

        # Databricks Hive SQL does not support positional ? parameters; inline validated ints.
        placeholders = ", ".join(str(n) for n in ids)
        sql = f"""
        SELECT
            ec.ec_number,
            ec.ec_problem,
            ec.ec_solution,
            st.ec_status
        FROM prd.rd_core.tbl_projectx_ec ec
        LEFT JOIN prd.rd_core.tbl_projectx_ec_status st
            ON ec.ec_status_id = st.ec_status_id
        WHERE ec.ec_number IN ({placeholders})
        ORDER BY ec.ec_number
        """

        conn = pyodbc.connect("DSN=Spark-PRD", autocommit=True)
        try:
            cursor = conn.cursor()
            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
            finally:
                cursor.close()
        finally:
            conn.close()

        records: List[Dict[str, Any]] = []
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                ec_number = int(row[0])
            except Exception:
                continue
            records.append({
                "ec_number": ec_number,
                "problem": str(row[1]).strip() if row[1] is not None else "",
                "solution": str(row[2]).strip() if row[2] is not None else "",
                "status": str(row[3]).strip() if row[3] is not None else "",
            })
        return records

    def _fetch_selected_ec_reference_context(self) -> Dict[str, Any]:
        """Fetch detailed EC context for yes-selected reference fields."""
        mapping = {
            "SPS": "sps_records",
            "ESW": "esw_records",
            "REF_ECR": "reference_ecr_records",
            "QN": "qn_records",
        }
        context: Dict[str, Any] = {value: [] for value in mapping.values()}

        for key, payload_key in mapping.items():
            pair = self.ref_radios.get(key)
            if not pair:
                continue
            rb_yes, _rb_no = pair
            if not rb_yes.isChecked():
                continue

            text_widget = self.ref_boxes.get(key)
            field_text = text_widget.toPlainText().strip() if text_widget else ""
            ids = self._extract_numeric_ids(field_text)
            if not ids:
                continue

            context[payload_key] = self._fetch_ec_records(ids)

        return context

    def _collect_not_found_verification_items(
        self,
        pcr_not_found: List[int],
        project_not_found: List[int],
        sps_not_found: List[int],
        esw_not_found: List[int],
    ) -> List[str]:
        """Collect all user-facing verification items for missing/unverifiable IDs."""
        issues: List[str] = []

        if pcr_not_found:
            issues.append(
                "PCR ID(s) not found: "
                + ", ".join(str(x) for x in pcr_not_found)
            )

        if project_not_found:
            issues.append(
                "Project ID(s) not found: "
                + ", ".join(str(x) for x in project_not_found)
            )

        if sps_not_found:
            issues.append(
                "SPS ID(s) not found: "
                + ", ".join(str(x) for x in sps_not_found)
            )

        if esw_not_found:
            issues.append(
                "ESW ID(s) not found: "
                + ", ".join(str(x) for x in esw_not_found)
            )

        return issues

    def _validate_q1_pcr(self) -> str:
        """Validate Question 1 (PCR_PCN) only – used by the PCR-driven summary path."""
        pair = self.ref_radios.get("PCR_PCN")
        if not pair:
            return "PCR question widgets not found."
        rb_yes, rb_no = pair
        if not rb_yes.isChecked() and not rb_no.isChecked():
            return (
                "Please answer Question 1: "
                "Does this Project include a Product Change Request (PCR)?"
            )
        if rb_yes.isChecked():
            pcr_text = (self.ref_boxes.get("PCR_PCN") or {}).toPlainText().strip() \
                if hasattr(self.ref_boxes.get("PCR_PCN"), "toPlainText") else ""
            if not pcr_text:
                return (
                    "Please enter at least one PCR number in the PCR field "
                    "before generating a Problem Summary."
                )
        return ""

    def _collect_pcr_payload(self) -> dict:
        """Fetch PCR/PSN records from Databricks and build an AI-ready payload.

        The PCR records and derived PSN state are stored only in runtime
        memory – no files are written.
        """
        pcr_text = self.ref_boxes["PCR_PCN"].toPlainText().strip()

        mod = self._load_pcr_query_module()
        result = mod.fetch_pcr_records(pcr_text)

        valid_records  = result.get("valid", [])
        skipped_records = result.get("skipped", [])
        not_found      = result.get("not_found", [])

        # Derive PSN Yes/No from the psnnumber field on each valid PCR row.
        psn_numbers = [
            str(rec["psnnumber"]).strip()
            for rec in valid_records
            if rec.get("psnnumber")
        ]
        derived_psn = (
            {"answer": "Yes", "numbers": psn_numbers}
            if psn_numbers
            else {"answer": "No", "numbers": []}
        )

        # Keep fetch metadata in a transient instance attribute so the
        # button handler can surface skipped/not-found IDs after generation.
        self._pcr_fetch_result = {
            "skipped":   skipped_records,
            "not_found": not_found,
        }

        scope_parts = [
            ln.strip()
            for ln in self.scope_parts_txt.toPlainText().splitlines()
            if ln.strip()
        ]

        reference_inputs = {}
        for key, txt in self.ref_boxes.items():
            reference_inputs[key] = {
                "answer": self._selected_radio_text(key),
                "text": txt.toPlainText().strip(),
            }

        return {
            "short_title":           self.short_title_edit.text().strip(),
            "reason_code":           self.reason_cb.currentText().strip(),
            "scope_parts":           scope_parts,
            "pcr_records":           valid_records,
            "derived_psn":           derived_psn,
            "skipped_pcrs":          skipped_records,
            "current_problem_text":  self.problem_txt.toPlainText().strip(),
            "proposed_solution_text": self.solution_txt.toPlainText().strip(),
            "selected_tab_payload":  self.get_selected_tab_payload(),
            "reference_inputs": reference_inputs,
        }

    def on_problem_summary_clicked(self):
        try:
            # Reset transient fetch state per click to avoid stale cross-field notices.
            self._project_fetch_result = {"skipped": [], "not_found": []}
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            self._esw_fetch_result = {"skipped": [], "not_found": []}

            # Always validate the full assessment first so unanswered Yes/No
            # selections are prompted regardless of PCR branch.
            validation_error = self._validate_problem_summary_inputs()
            if validation_error:
                QMessageBox.warning(self, "Validation Required", validation_error)
                return

            pcr_yes = self.ref_radios.get("PCR_PCN", (None, None))[0]
            use_pcr_path = pcr_yes is not None and pcr_yes.isChecked()

            if use_pcr_path:
                # ---- Question 1 – PCR-driven Databricks path ----
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                payload = self._append_project_context(self._collect_pcr_payload())

                # Auto-fill PSN radio button and text field from derived_psn (PCR path only).
                # This must happen immediately after PCR fetch so users are not asked to enter PSN manually.
                derived_psn = payload.get("derived_psn") or {}
                psn_pair = self.ref_radios.get("PSN")
                psn_box = self.ref_boxes.get("PSN")
                if psn_pair:
                    psn_yes_rb, psn_no_rb = psn_pair
                    if derived_psn.get("answer") == "Yes":
                        psn_yes_rb.setChecked(True)
                        if psn_box:
                            psn_box.setPlainText(", ".join(derived_psn.get("numbers") or []))
                            psn_box.setVisible(True)
                    else:
                        psn_no_rb.setChecked(True)
                        if psn_box:
                            psn_box.setPlainText("")
                            psn_box.setVisible(False)

                fetch_meta = getattr(self, "_pcr_fetch_result", {})
                skipped    = fetch_meta.get("skipped", [])
                not_found  = fetch_meta.get("not_found", [])
                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])

                if not payload.get("pcr_records"):
                    msgs = []
                    if skipped:
                        msgs.append(
                            "The following PCR(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  PCR {s['pcr_id']} – {s['status']}" for s in skipped
                            )
                        )
                    if not_found:
                        msgs.append(
                            "The following PCR ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found)
                        )
                    raise ValueError(
                        "No eligible PCR records found to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("PCR", (None, None))[0] and self.ref_radios["PCR"][0].isChecked() and not payload.get("project_records"):
                    msgs = []
                    if skipped_projects:
                        msgs.append(
                            "The following Project(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  Project {s['project_id']} - {s['status']}" for s in skipped_projects
                            )
                        )
                    if not_found_projects:
                        msgs.append(
                            "The following Project ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_projects)
                        )
                    raise ValueError(
                        "No eligible Project records found for Question 2 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("SPS", (None, None))[0] and self.ref_radios["SPS"][0].isChecked() and not payload.get("sps_records"):
                    msgs = []
                    if skipped_sps:
                        msgs.append(
                            "The following SPS(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                            )
                        )
                    if not_found_sps:
                        msgs.append(
                            "The following SPS ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_sps)
                        )
                    raise ValueError(
                        "No eligible SPS records found for Question 3 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("ESW", (None, None))[0] and self.ref_radios["ESW"][0].isChecked() and not payload.get("esw_records"):
                    msgs = []
                    if skipped_esw:
                        msgs.append(
                            "The following ESW(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                            )
                        )
                    if not_found_esw:
                        msgs.append(
                            "The following ESW ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_esw)
                        )
                    raise ValueError(
                        "No eligible ESW records found for Question 4 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                verify_items = self._collect_not_found_verification_items(
                    pcr_not_found=not_found,
                    project_not_found=not_found_projects,
                    sps_not_found=not_found_sps,
                    esw_not_found=not_found_esw,
                )
                if verify_items:
                    QMessageBox.warning(
                        self,
                        "Reference Verification Required",
                        "Some entered values were not found. Please verify and correct before generating Problem Summary:\n\n"
                        + "\n".join(f"- {x}" for x in verify_items),
                    )
                    return

                out = self._run_ai_pss_full(payload)

            else:
                # ---- Existing all-questions path ----
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                payload = self._append_project_context(self._collect_pss_payload())

                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])

                if self.ref_radios.get("SPS", (None, None))[0] and self.ref_radios["SPS"][0].isChecked() and not payload.get("sps_records"):
                    msgs = []
                    if skipped_sps:
                        msgs.append(
                            "The following SPS(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                            )
                        )
                    if not_found_sps:
                        msgs.append(
                            "The following SPS ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_sps)
                        )
                    raise ValueError(
                        "No eligible SPS records found for Question 3 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("ESW", (None, None))[0] and self.ref_radios["ESW"][0].isChecked() and not payload.get("esw_records"):
                    msgs = []
                    if skipped_esw:
                        msgs.append(
                            "The following ESW(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                            )
                        )
                    if not_found_esw:
                        msgs.append(
                            "The following ESW ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_esw)
                        )
                    raise ValueError(
                        "No eligible ESW records found for Question 4 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                verify_items = self._collect_not_found_verification_items(
                    pcr_not_found=[],
                    project_not_found=not_found_projects,
                    sps_not_found=not_found_sps,
                    esw_not_found=not_found_esw,
                )
                if verify_items:
                    QMessageBox.warning(
                        self,
                        "Reference Verification Required",
                        "Some entered values were not found. Please verify and correct before generating Problem Summary:\n\n"
                        + "\n".join(f"- {x}" for x in verify_items),
                    )
                    return

                if getattr(self, "email_file_path", None):
                    module = self._load_ai_pss_module()
                    email_content = module.read_email_file(self.email_file_path)
                    if str(email_content).startswith("Error reading"):
                        raise ValueError(f"Could not read email file:\n{email_content}")
                    user_problem = self.problem_txt.toPlainText().strip()
                    out = module.correlate_email_with_problem(
                        email_content,
                        user_problem,
                        payload
                    )
                    if out.get("error"):
                        raise ValueError(out["error"])
                else:
                    out = self._run_ai_pss_full(payload)

                fetch_meta = {}
                skipped    = []
                not_found  = []
                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])

            # ---- Common output handling ----
            title    = (out.get("title") or "").strip()
            problem  = (out.get("problem") or out.get("problem_statement") or "").strip()
            solution = (out.get("solution") or out.get("solution_statement") or "").strip()

            if not any([title, problem, solution]):
                raise ValueError("AI_Assisted_PSS returned empty output.")

            if title:
                self.short_title_edit.setText(title[:75])
            if problem:
                self.problem_txt.setPlainText(problem[:2000])
            if solution:
                self.solution_txt.setPlainText(solution[:2000])

            # Surface skipped / not-found PCRs as a non-blocking info message.
            if use_pcr_path and (skipped or not_found):
                info_parts = []
                if skipped:
                    info_parts.append(
                        "Skipped PCR(s) due to inactive status:\n"
                        + "\n".join(
                            f"  PCR {s['pcr_id']} – {s['status']}" for s in skipped
                        )
                    )
                if not_found:
                    info_parts.append(
                        "PCR ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found)
                    )
                QMessageBox.information(
                    self,
                    "PCR Lookup Notice",
                    "Summary generated from eligible PCR(s).\n\n"
                    + "\n\n".join(info_parts),
                )

            if (payload.get("project_records") or skipped_projects or not_found_projects):
                info_parts = []
                if skipped_projects:
                    info_parts.append(
                        "Skipped Project(s) due to inactive status:\n"
                        + "\n".join(
                            f"  Project {s['project_id']} - {s['status']}" for s in skipped_projects
                        )
                    )
                if not_found_projects:
                    info_parts.append(
                        "Project ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_projects)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "Project Lookup Notice",
                        "Summary generated from eligible Project record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            if (payload.get("sps_records") or skipped_sps or not_found_sps):
                info_parts = []
                if skipped_sps:
                    info_parts.append(
                        "Skipped SPS(s) due to inactive status:\n"
                        + "\n".join(
                            f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                        )
                    )
                if not_found_sps:
                    info_parts.append(
                        "SPS ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_sps)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "SPS Lookup Notice",
                        "Summary generated from eligible SPS record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            if (payload.get("esw_records") or skipped_esw or not_found_esw):
                info_parts = []
                if skipped_esw:
                    info_parts.append(
                        "Skipped ESW(s) due to inactive status:\n"
                        + "\n".join(
                            f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                        )
                    )
                if not_found_esw:
                    info_parts.append(
                        "ESW ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_esw)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "ESW Lookup Notice",
                        "Summary generated from eligible ESW record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

        except Exception as e:
            QMessageBox.warning(self, "Problem Summary Error", str(e))
        finally:
            QApplication.restoreOverrideCursor()


    # ---- Flow Logic ----
    def clear_flow(self):
        while self.flow_area.count():
            w = self.flow_area.takeAt(0).widget()
            if w: w.deleteLater()
        self.ec_result_lbl.setText(""); self.ec_divider.setVisible(False)

    def add_q(self,q,opts,cb):
        row = QHBoxLayout(); row.addWidget(QLabel(q))
        grp = QButtonGroup(self)
        for o in opts:
            rb = QRadioButton(o); grp.addButton(rb)
            rb.toggled.connect(lambda c,v=o: c and cb(v))
            row.addWidget(rb)
        row.addStretch(1)
        w = QWidget(); w.setLayout(row); self.flow_area.addWidget(w)

    def start_flow(self):
        if not self.sender().isChecked(): return
        self.clear_flow(); s=self.sender().text()
        if s=="Production Release": return self.finish("B3")
        if s=="OBS / Inactivate": return self.finish("B1")
        if s=="Up Revision": return self.add_q("SmBOM Impacted?",["Yes","No"],self.up1)
        if s=="Status Roll": return self.add_q("Transition Type?",["EVAL → Production","Concept → BOM List"],self.status1)
        if s=="Product Release": return self.add_q("SmBOM Impacted?",["Yes","No"],self.prod1)

    def up1(self,v): self.trim(1); self.finish("A1" if v=="Yes" else "A2") if v in ["Yes","No"] and False else (self.add_q("OBS Kit / Pieces?",["Yes","No"],lambda x:self.finish("A1" if x=="Yes" else "A2")) if v=="Yes" else self.add_q("Part Status?",["EVAL","PROD"],lambda x:self.finish("B2" if x=="EVAL" else "B3")))
    def status1(self,v): self.trim(1); self.finish("B3") if v.startswith("EVAL") else self.add_q("OBS Kit / Pieces?",["Yes","No"],lambda x:self.finish("A1" if x=="Yes" else "A2"))
    def prod1(self,v): self.trim(1); self.finish("A2") if v=="Yes" else self.add_q("Part Status?",["EVAL","PROD"],lambda x:self.finish("B2" if x=="EVAL" else "B3"))

    def trim(self,k):
        while self.flow_area.count()>k:
            w=self.flow_area.takeAt(self.flow_area.count()-1).widget()
            if w: w.deleteLater()

    def finish(self,c):
        self.ec_result_lbl.setText(f"EC Category: {c} – {EC_CATEGORY_DESC[c]}")
        self.ec_divider.setVisible(True); self.secB_header.setVisible(True); self.secB.setVisible(True)



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1280, 860)
        self.tabs = QTabWidget()
        
        self.tabs.setStyleSheet("""
            QTabBar::tab { background: #EAF2FB; color: #1F3B57; padding: 8px 14px; border: 1px solid #D5E3F6; border-bottom: none; border-top-left-radius:6px; border-top-right-radius:6px; }
            QTabBar::tab:selected { background: #FFFFFF; color: #0F2D46; font-weight: 600; }
            QTabWidget::pane { border: 1px solid #D5E3F6; top: -1px; }
        """)

        readme_path = Path(__file__).with_name('README.txt')
        self.readme_tab = ReadmeTab(readme_path)
        self.front_tab = ECRFrontPageTab()
        self.obs_tab = OBSPartsTab()
        self.structure_tab = None  # to be created later
        self.whereused_tab = WhereUsedTabV2(obs_provider=self.obs_tab)
        self.obs_tab.where_used_tab = self.whereused_tab  # enables OBS → Where Used import

        self.tabs.addTab(self.readme_tab, "READ ME")
        self.tabs.addTab(ECCreationInputsFormTab(), "EC Creation Inputs Form"); self.tabs.setDocumentMode(True); self.tabs.setMovable(True)
        self.tabs.addTab(self.front_tab, "ECR Front Page")
        self.tabs.addTab(self.obs_tab, "OBS Parts")
        self.tabs.addTab(self.whereused_tab, "Where Used")
        self.tabs.addTab(OrphanAnalysisTab(self.obs_tab), "Orphan Analysis")
        self.structure_tab = StructureSheetTab(); self.tabs.addTab(self.structure_tab, "Structure sheet")
        self.inventory_cost_tab=InventoryCostTab(); self.tabs.addTab(self.inventory_cost_tab,'Inventory_Cost')
        self.tabs.addTab(PlaceholderTab("Safety"), "Safety")
        self.tabs.addTab(PlaceholderTab("CE!"), "CE!")
        self.tabs.addTab(PlaceholderTab("CDW"), "CDW")
        self.tabs.addTab(PlaceholderTab("Report"), "Report")
        self.tabs.addTab(PlaceholderTab("User Notes"), "User Notes")
        self.setCentralWidget(self.tabs)
        def update_tab_colors():
            for i in range(self.tabs.count()):
                txt = self.tabs.tabText(i)
                # heuristic: Front Page checklist
                if 'ECR Front Page' in txt:
                    done = self.front_tab.progress_bar.value()==self.front_tab.progress_bar.maximum()
                    if done:
                        self.tabs.setTabText(i, '✓ ' + txt.strip('✓ '))
                        self.tabs.tabBar().setTabBackgroundColor(i, QColor('#E8F5E9'))
                    elif self.front_tab.progress_bar.value()>0:
                        self.tabs.setTabText(i, '✓ ' + txt.strip('✓ '))
  #                      self.tabs.tabBar().setTabBackgroundColor(i, QColor('#E3F2FD'))
            
        try:
            self.front_tab.progress_bar.valueChanged.connect(update_tab_colors)
        except Exception: pass


        tb=QToolBar("File"); self.addToolBar(tb)
        find_edit = QLineEdit(); find_edit.setPlaceholderText('Find part / text'); find_edit.setFixedWidth(220); tb.addWidget(find_edit)
        def _do_find():
            text = find_edit.text().strip().lower()
            if not text:
                return
            w = self.tabs.currentWidget()
            from PyQt6.QtWidgets import QTableWidget, QTextEdit
            found = False
            for t in w.findChildren(QTableWidget):
                for r in range(t.rowCount()):
                    for c in range(t.columnCount()):
                        it = t.item(r,c)
                        if it and text in it.text().lower():
                            it.setBackground(QColor('#FFF59D'))
                            t.setCurrentCell(r,c)
                            found = True
                            break
                    if found:
                        break
                if found:
                    break
            if not found:
                for te in w.findChildren(QTextEdit):
                    if text in te.toPlainText().lower():
                        te.find(text)
                        found = True
                        break
            if not found:
                QMessageBox.information(self, 'Find', 'Not Found')
        find_edit.returnPressed.connect(_do_find)
        act_save=QAction("Save", self); act_save.triggered.connect(self.save_data); tb.addAction(act_save)
        
        spacer=QWidget(); spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred); tb.addWidget(spacer)

        reset_btn=QPushButton("Reset App")
        reset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reset_btn.setStyleSheet("""
            QPushButton { color:#FFFFFF; padding:6px 14px; border-radius:6px; border:1px solid #0D5EA6;
            background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2994FF, stop:1 #0A67C2); }
            QPushButton:hover { background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2FA0FF, stop:1 #0D6ED0); }
            QPushButton:pressed { padding-top:7px; padding-bottom:5px; }
        """)
        shadow=QGraphicsDropShadowEffect(self); shadow.setBlurRadius(12); shadow.setXOffset(0); shadow.setYOffset(2); shadow.setColor(QColor(0,0,0,80))
        reset_btn.setGraphicsEffect(shadow); reset_btn.clicked.connect(self.reset_app); tb.addWidget(reset_btn)

        self.load_data_if_exists()

    def aggregate_data(self)->Dict[str,Any]:
        return {'front_page': self.front_tab.to_dict(), 'obs_parts': self.obs_tab.to_dict()}

    def apply_data(self, data: Dict[str,Any]):
        if not data: return
        if 'front_page' in data: self.front_tab.from_dict(data['front_page'])
        if 'obs_parts' in data: self.obs_tab.from_dict(data['obs_parts'])

    def save_data(self):
        data=self.aggregate_data()
        try:
            DATA_FILE.write_text(json.dumps(data, indent=2), encoding='utf-8')
            QMessageBox.information(self, "Saved", f"Data saved to {DATA_FILE.name}")
        except Exception as e:
            QMessageBox.warning(self, "Save Failed", str(e))

    def load_data_if_exists(self):
        try:
            if DATA_FILE.exists():
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
        except Exception as e:
            QMessageBox.warning(self, "Load Failed", str(e))

    def load_data_dialog(self):
        if DATA_FILE.exists():
            try:
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
                QMessageBox.information(self, "Loaded", f"Data loaded from {DATA_FILE.name}")
            except Exception as e:
                QMessageBox.warning(self, "Load Failed", str(e))
        else:
            QMessageBox.information(self, "No Save Found", "No saved data file found yet. Click Save to create one.")

    def reset_app(self):
        ans=QMessageBox.question(self, "Reset App", "This will clear all fields, reset tables and remove any saved data. Do you want to continue?",
                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if ans==QMessageBox.StandardButton.Yes:
            try:
                self.front_tab.reset(); self.obs_tab.reset()
                if DATA_FILE.exists(): DATA_FILE.unlink()
                QMessageBox.information(self, "Reset Complete", "Application data has been reset.")
            except Exception as e:
                QMessageBox.warning(self, "Reset Failed", str(e))

class OrphanAnalysisTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        outer = QVBoxLayout(self)

        header = QWidget(self)
        h = QHBoxLayout(header)
        h.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))

        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        rb_without = QRadioButton("Without Replacement")
        rb_with = QRadioButton("With / Without Replacement")
        rb_without.setChecked(True)

        grp = QButtonGroup(self)
        grp.addButton(rb_without, 0)
        grp.addButton(rb_with, 1)

        h.addWidget(title)
        h.addSpacing(12)
        h.addWidget(rb_without)
        h.addWidget(rb_with)
        h.addStretch(1)
        outer.addWidget(header)

        self.tabs = QTabWidget()
        tab1 = OBSAllPartsWithoutReplacementTab(obs_provider)
        tab2 = PlaceholderTab("OBS with or without Replacement")

        self.tabs.addTab(tab1, "OBS all Parts without Replacements")
        self.tabs.addTab(tab2, "OBS with or without Replacement")
        outer.addWidget(self.tabs)

        def apply(idx):
            self.tabs.setCurrentIndex(idx)
            self.tabs.setTabEnabled(0, idx == 0)
            self.tabs.setTabEnabled(1, idx == 1)

        rb_without.toggled.connect(lambda v: v and apply(0))
        rb_with.toggled.connect(lambda v: v and apply(1))
        apply(0)



def run():
    app=QApplication(sys.argv); app.setStyle("Fusion")
    pal=QPalette(); pal.setColor(QPalette.ColorRole.Window, QColor(247,250,253)); pal.setColor(QPalette.ColorRole.Base, QColor(255,255,255))
    pal.setColor(QPalette.ColorRole.AlternateBase, QColor(241,246,252)); pal.setColor(QPalette.ColorRole.Text, QColor(28,41,56)); app.setPalette(pal)
    win=MainWindow(); win.show(); sys.exit(app.exec())

if __name__=='__main__': run()

