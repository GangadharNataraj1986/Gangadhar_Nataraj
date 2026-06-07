import threading
import pyodbc
def fetch_distinct_kit_codes(plant='4070'):
    """Fetch distinct kit codes from Databricks BOM table."""
    DSN = "Spark-PRD"
    thread_local = threading.local()
    def get_conn():
        if getattr(thread_local, 'conn', None) is None:
            thread_local.conn = pyodbc.connect(f"DSN={DSN}", autocommit=True)
        return thread_local.conn
    sql = f"""
        SELECT DISTINCT sortstring AS kit_code
        FROM prd.pd_mm.factbomlvl1
        WHERE plantcd = '{plant}' AND sortstring IS NOT NULL AND sortstring <> ''
        ORDER BY kit_code
    """
    conn = get_conn()
    try:
        cursor = conn.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        return [row[0] for row in rows if row[0]]
    except Exception as e:
        print(f"Error fetching kit codes: {e}")
        return []
# ecr_kit_ui.py (Enhanced v7.5.5 – Auto Excel conversion: silently open in Excel, SaveAs .xlsx, then import; OBS copy buttons)
import html
import importlib.util
import json
import re
import sys
import traceback
from pathlib import Path
from typing import Any, Dict, List

try:
    from inventory_demand_cost_query import (
        fetch_inventory_demand_cost,
        fetch_kit_code_descriptions,
        fetch_open_purchase_order_details,
    )
except ImportError:
    fetch_inventory_demand_cost = None
    fetch_kit_code_descriptions = None
    fetch_open_purchase_order_details = None

try:
    from report_recommendations import REPORT_COLUMNS, build_supply_chain_report
except ImportError:
    REPORT_COLUMNS = []
    build_supply_chain_report = None

try:
    from Watch_List import WatchListTab
except ImportError:
    WatchListTab = None

try:
    from PSS_Change_Summary import (
        export_change_summary,
        generate_change_summary_sentences,
    )
except ImportError:
    export_change_summary = None
    generate_change_summary_sentences = None

try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    _HAS_MATPLOTLIB = True
except Exception:
    FigureCanvas = None
    Figure = None
    _HAS_MATPLOTLIB = False

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PyQt6.QtCore import Qt, QRect, QObject, QEvent, pyqtSignal, QTimer
from PyQt6.QtGui import QAction, QBrush, QColor, QFont, QFontMetrics, QGuiApplication, QPalette, QPen, QTextCursor
from PyQt6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QStyle,
    QStyleOptionHeader,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QToolBar,
    QVBoxLayout,
    QWidget,
)

APP_TITLE = "ECR Kit Assistant"
DATA_FILE = Path(__file__).with_name('ecr_kit_data.json')
TEMPLATE_FILE = Path(__file__).with_name('Obs_parts_template.xlsx')

def _handle_exception(exc_type, exc_value, exc_tb):
    try:
        log_path = Path(__file__).with_name('error_log.txt')
        msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log_path.write_text(msg, encoding='utf-8')
        m = QMessageBox()
        m.setIcon(QMessageBox.Icon.Critical)
        m.setWindowTitle('Unexpected Error')
        m.setText('An unexpected error occurred. A log was written to error_log.txt')
        m.setDetailedText(msg)
        m.setStandardButtons(QMessageBox.StandardButton.Ok)
        m.show()
        app = QApplication.instance()
        if app is not None:
            if not hasattr(app, '_error_dialogs'):
                app._error_dialogs = []
            app._error_dialogs.append(m)
    except Exception:
        print('Unhandled exception:', file=sys.stderr)
        traceback.print_exception(exc_type, exc_value, exc_tb)

sys.excepthook = _handle_exception


class ShiftEnterTextEdit(QTextEdit):
    """Enter moves focus; Shift+Enter inserts a newline."""

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                super().keyPressEvent(event)
            else:
                self.focusNextPrevChild(True)
                event.accept()
            return
        super().keyPressEvent(event)


def _format_solution_for_display(text: str) -> str:
    """
    Post-processes solution text before displaying in solution_txt widget.
    Splits any inline ' - ' / em-dash / en-dash action separators into
    individual bullet lines (UI-side safety net only).
    """
    if not text:
        return ""

    INLINE_SEP = re.compile(
        r"(?<=[a-zA-Z0-9.)]) - (?=[A-Z])"  # hyphen with spaces after word/period/paren
        r"|\.?\s*\u2014\s*(?=[A-Z])"        # em-dash (U+2014)
        r"| \u2013 (?=[A-Z])",              # en-dash (U+2013)
        re.UNICODE,
    )

    out: List[str] = []
    for raw in text.splitlines():
        stripped = raw.strip()
        if not stripped:
            out.append("")
            continue

        # Preserve section-header lines unchanged (e.g. "Solution Description:")
        if re.match(r"^[A-Za-z][\w /()&]+:\s*$", stripped):
            out.append(raw)
            continue

        # Strip any existing bullet prefix so we can re-apply it consistently
        prefix_m = re.match(r"^(\s*[-*\u2022]\s*)", raw)
        body = raw[len(prefix_m.group(0)):].strip() if prefix_m else raw.strip()

        # Split on inline action separators — each part becomes its own bullet
        parts = [p.strip() for p in INLINE_SEP.split(body) if p.strip()]
        for part in parts:
            out.append("- " + part)

    return "\n".join(out)


def _format_pss_for_html_display(text: str) -> str:
    """
    Converts plain PSS text to HTML for rich display in QTextEdit widgets.
    - Strips any markdown bold markers (- ** ... **) from section headers and renders them bold.
    - Makes From: and To: bold wherever they appear inline.
    - Makes from: and to: bold wherever they appear inline.
    - Adds an empty line before 'Benefits of the Proposed Solution:'.
    """
    import html as html_mod

    if not text:
        return ""

    # Insert blank line before Benefits section if not already present
    text = re.sub(
        r"(?<!\n)\n(Benefits of the Proposed Solution\s*:)",
        r"\n\n\1",
        text,
    )

    lines = text.splitlines()
    html_parts: List[str] = []
    always_bold_headers = {
        "change drivers",
        "esw turn around",
        "problem description",
        "impact details",
        "issue part number(s)",
    }

    def _canon_header_name(s: str) -> str:
        # Canonical form for robust header matching.
        s = re.sub(r"^[-*\u2022]\s*", "", s.strip())
        s = re.sub(r"\*{1,2}", "", s)
        s = re.sub(r"\s*:\s*$", "", s)
        s = re.sub(r"\s+", " ", s)
        return s.strip().lower()

    fixed_headers_pattern = re.compile(
        r"(?i)(?<![*>])(Change\s*Drivers\s*:?|ESW\s*Turn\s*Around\s*:?|Problem\s*Description\s*:|Impact\s*Details\s*:|Issue\s*Part\s*Number\(s\)\s*:)(?![*<])"
    )

    for line in lines:
        stripped = line.strip()

        if not stripped:
            html_parts.append("")
            continue

        # Strip markdown bold/bullet artefacts from header lines: "- **Header:**" -> "Header:"
        clean = re.sub(r"^[-*\u2022]?\s*\*{1,2}([^*]+)\*{1,2}\s*$", r"\1", stripped)
        # Also handle "- **Header:** trailing text" patterns
        clean = re.sub(r"^[-*\u2022]?\s*\*{1,2}([^*]+?)\*{1,2}(.*)", r"\1\2", clean)

        escaped = html_mod.escape(clean)
        canon_header = _canon_header_name(clean)

        # Bold section headers (lines ending with ":")
        if canon_header in always_bold_headers:
            label = clean.strip()
            if not label.endswith(":"):
                label = label + ":"
            html_parts.append(f"<b>{html_mod.escape(label)}</b>")
        elif re.match(r"^[A-Za-z][A-Za-z0-9_ /()&,#.\-]+:\s*$", clean):
            html_parts.append(f"<b>{escaped}</b>")
        else:
            # Always bold key PSS headers even when they appear inline.
            escaped = fixed_headers_pattern.sub(r"<b>\1</b>", escaped)
            # Bold markdown and plain From/To tokens (case-insensitive, tolerant of spacing).
            escaped = re.sub(r"\*\*(from\s*:)\*\*", r"<b>\1</b>", escaped, flags=re.IGNORECASE)
            escaped = re.sub(r"\*\*(to\s*:)\*\*", r"<b>\1</b>", escaped, flags=re.IGNORECASE)
            escaped = re.sub(r"(?<![*>])(from\s*:)(?![*<])", r"<b>\1</b>", escaped, flags=re.IGNORECASE)
            escaped = re.sub(r"(?<![*>])(to\s*:)(?![*<])", r"<b>\1</b>", escaped, flags=re.IGNORECASE)
            html_parts.append(escaped)

    return "<br>".join(html_parts)


class ReadmeTab(QWidget):
    def __init__(self, readme_path: Path):
        super().__init__()
        layout = QVBoxLayout(self)
        title = QLabel("READ ME")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.DemiBold))
        layout.addWidget(title)
        text = QTextEdit(); text.setReadOnly(True)
        content = "--Wait for the Instructions on how to use--."
        try:
            if readme_path.exists():
                content = readme_path.read_text(encoding="utf-8")
        except Exception as e:
            content = f"Error opening README: {e}"
        text.setPlainText(content)
        layout.addWidget(text)

# ---------- Helpers ----------

def get_orphan_color(orphan_level: str):
    lvl = orphan_level.lower()
    if lvl == 'orphan1': return QColor('#C0392B')
    if lvl == 'orphan2': return QColor('#E67E22')
    if lvl.startswith('orphan'): return QColor('#2980B9')
    return None


def _excel_width_to_px(widget: QWidget, chars: int) -> int:
    fm = widget.fontMetrics()
    return fm.horizontalAdvance('0' * max(1, chars)) + 22

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                self.insertPlainText('\n'); return
        super().keyPressEvent(event)


class RotatedColumnsHeader(QHeaderView):
    """Horizontal header that renders selected columns with 90-degree rotated text.
    Optionally paints a group label spanning all rotated columns in the top band."""

    GROUP_BAND = 26  # height in px reserved for the group label at top of header

    def __init__(self, orientation, rotated_columns=None, parent=None,
                 group_label=None, group_columns=None):
        super().__init__(orientation, parent)
        self._rotated_columns = set(rotated_columns or [])
        self._group_label = group_label          # e.g. 'Change Type'
        self._group_columns = list(group_columns or [])  # e.g. [2,3,4,5,6]
        self._group_spans = []  # list[(label, [col_idx,...])]
        self._header_texts = {}  # cache for header text
        self.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

    def set_rotated_columns(self, columns):
        self._rotated_columns = set(columns or [])
        self.viewport().update()

    def set_group_spans(self, group_spans):
        # group_spans: list of tuples -> (label: str, columns: list[int])
        self._group_spans = [
            (str(label), list(cols))
            for label, cols in (group_spans or [])
            if cols
        ]
        self.viewport().update()

    def set_header_texts(self, texts_dict):
        """Cache header text: { col_index: text }"""
        self._header_texts = dict(texts_dict or {})

    def _get_header_text(self, logical_index):
        """Get header text from cache or model. For demand columns, blank if no inventory."""
        # Try to blank demand columns if no inventory (On Hand and On Order == 0)
        text = self._header_texts.get(logical_index, None)
        if text is not None:
            # Check if this is a demand column (e.g., '4020 Gross Demand-13')
            m = re.match(r'^(\d{4}) Gross Demand-(\d+)', str(text))
            if m and hasattr(self, 'parent_table') and hasattr(self, 'parent_row'):
                plant = m.group(1)
                onhand_col = f'{plant} Onhand Qty'
                onorder_col = f'{plant} On Order Qty'
                df = getattr(self, 'parent_df', None)
                row = getattr(self, 'parent_row', None)
                if df is not None and row is not None:
                    try:
                        onhand = df.iloc[row][onhand_col] if onhand_col in df.columns else 0
                        onorder = df.iloc[row][onorder_col] if onorder_col in df.columns else 0
                        if (onhand == 0 or pd.isna(onhand)) and (onorder == 0 or pd.isna(onorder)):
                            return ''
                    except Exception:
                        pass
            return str(text)
        try:
            m = self.model()
            if m:
                text = m.headerData(logical_index, self.orientation(), Qt.ItemDataRole.DisplayRole)
                return "" if text is None else str(text)
        except Exception:
            pass
        return ""

    def paintSection(self, painter, rect, logicalIndex):
        if not rect.isValid():
            return

        has_legacy_group = bool(self._group_label and self._group_columns)
        has_multi_groups = bool(self._group_spans)
        has_group = has_legacy_group or has_multi_groups

        if logicalIndex not in self._rotated_columns:
            super().paintSection(painter, rect, logicalIndex)
            return

        # ---- Split rect: top band for group label, rest for rotated text ----
        band = self.GROUP_BAND if has_group else 0
        indiv_rect = QRect(rect.x(), rect.y() + band, rect.width(), rect.height() - band)

        # Draw background/border for the individual section area.
        option = QStyleOptionHeader()
        self.initStyleOption(option)
        option.rect = indiv_rect
        option.section = logicalIndex
        option.text = ""
        self.style().drawControl(QStyle.ControlElement.CE_HeaderSection, option, painter, self)

        # Draw rotated column text in the lower portion.
        text = self._get_header_text(logicalIndex)
        painter.save()
        painter.setPen(self.palette().color(QPalette.ColorRole.ButtonText))
        painter.translate(indiv_rect.center())
        painter.rotate(-90)
        tr = QRect(-indiv_rect.height() // 2 + 4, -indiv_rect.width() // 2 + 2,
                   indiv_rect.height() - 8, indiv_rect.width() - 4)
        painter.drawText(tr, Qt.AlignmentFlag.AlignCenter | Qt.TextFlag.TextWordWrap, text)
        painter.restore()

        # ---- Draw group labels in top band ----
        if has_multi_groups:
            for label, cols in self._group_spans:
                if logicalIndex != cols[0]:
                    continue
                total_w = sum(self.sectionSize(c) for c in cols)
                group_rect = QRect(rect.x(), rect.y(), total_w, band)
                painter.save()
                painter.fillRect(group_rect, QColor('#B0D8F5'))
                painter.setPen(QPen(QColor('#1F6FB2'), 2))
                painter.drawRect(group_rect.adjusted(0, 0, -1, -1))
                painter.setPen(QColor('#0F2D46'))
                bold_font = painter.font()
                bold_font.setBold(True)
                bold_font.setPointSize(bold_font.pointSize() - 1)
                painter.setFont(bold_font)
                painter.drawText(group_rect, Qt.AlignmentFlag.AlignCenter, label)
                painter.restore()
                break
        elif has_legacy_group and logicalIndex == self._group_columns[0]:
            total_w = sum(self.sectionSize(c) for c in self._group_columns)
            group_rect = QRect(rect.x(), rect.y(), total_w, band)
            painter.save()
            painter.fillRect(group_rect, QColor('#B0D8F5'))
            painter.setPen(QPen(QColor('#1F6FB2'), 2))
            painter.drawRect(group_rect.adjusted(0, 0, -1, -1))
            painter.setPen(QColor('#0F2D46'))
            bold_font = painter.font()
            bold_font.setBold(True)
            bold_font.setPointSize(bold_font.pointSize() - 1)
            painter.setFont(bold_font)
            painter.drawText(group_rect, Qt.AlignmentFlag.AlignCenter, self._group_label)
            painter.restore()


class OBSTable(QTableWidget):
    def __init__(self, parent=None, initial_rows: int = 10):
        super().__init__(initial_rows, 4, parent)
        self.setHorizontalHeaderLabels(["Select","OBS Parts","Change","Replacement"])
        self.verticalHeader().setVisible(False)
        try:
            header=self.horizontalHeader(); header.setStretchLastSection(False)
            header.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(1,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(2,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(3,QHeaderView.ResizeMode.Interactive)
        except Exception: pass
        self._init_rows(0,self.rowCount()); self._apply_excel_widths(); self.setAlternatingRowColors(True)

    def _apply_excel_widths(self, chars:int=14):
        px=_excel_width_to_px(self,chars)
        for col in (1,2,3):
            try: self.setColumnWidth(col,px)
            except Exception: pass

    def _make_change_combo(self)->QComboBox:
        combo=QComboBox(); combo.addItems(["", "Obsolete","Inactivate"]); combo.setCurrentText("Obsolete"); return combo

    def _init_rows(self,start,end):
        for r in range(start,end):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); self.setCellWidget(r,0,cont)
            if not self.item(r,1): self.setItem(r,1,QTableWidgetItem(""))
            combo=self._make_change_combo(); self.setCellWidget(r,2,combo)
            if not self.item(r,3): self.setItem(r,3,QTableWidgetItem(""))

    def keyPressEvent(self,event):
        try:
            if event.matches(event.StandardKey.Copy): self._copy_selection_to_clipboard(); return
            if event.matches(event.StandardKey.Paste): self._paste_from_clipboard(); return
        except Exception: traceback.print_exc()
        super().keyPressEvent(event)

    def _copy_selection_to_clipboard(self):
        sel = self.selectedRanges()
        if not sel:
            return
        r=sel[0]; rows=[]
        for i in range(r.rowCount()):
            cols=[]
            for j in range(r.columnCount()):
                row_i=r.topRow()+i; col_j=r.leftColumn()+j
                if col_j==2:
                    w=self.cellWidget(row_i,2); txt=w.currentText() if isinstance(w,QComboBox) else ''
                else:
                    it=self.item(row_i,col_j); txt=it.text() if it else ''
                cols.append(txt)
            rows.append('\t'.join(cols))
        QGuiApplication.clipboard().setText('\n'.join(rows))

    def _ensure_rows(self,upto_row_inclusive:int):
        if upto_row_inclusive>=self.rowCount():
            old=self.rowCount(); self.setRowCount(upto_row_inclusive+1); self._init_rows(old,self.rowCount())

    def _paste_from_clipboard(self):
        text = QGuiApplication.clipboard().text()
        if not text:
            return
        start_row=self.currentRow(); start_col=self.currentColumn()
        if start_row<0:
            start_row=self._first_empty_row()
            if start_row<0: start_row=self.rowCount()
        lines=[ln for ln in text.splitlines() if ln.strip()]
        for r_offset,line in enumerate(lines):
            parts=[p.strip() for p in line.split('\t')]
            row=start_row+r_offset; self._ensure_rows(row)
            for c_offset,val in enumerate(parts):
                col=start_col+c_offset
                if col==0: continue
                if col==2:
                    w=self.cellWidget(row,2)
                    if isinstance(w,QComboBox):
                        idx=w.findText(val)
                        if idx>=0: w.setCurrentIndex(idx)
                else:
                    self.setItem(row,col,QTableWidgetItem(val))
        self._apply_excel_widths()

    def _first_empty_row(self):
        for r in range(self.rowCount()):
            it=self.item(r,1)
            if it is None or not it.text().strip(): return r
        return -1

    def delete_selected_rows(self):
        rows_to_delete = []
        for r in range(self.rowCount()):
            w = self.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and cb.isChecked():
                    rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            self.removeRow(r)
        if self.rowCount() == 0:
            self.setRowCount(10)
            self._init_rows(0, 10)

class OBSPartsTab(QWidget):
    def toggle_select_all(self):
        any_unchecked = False
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and not cb.isChecked():
                    any_unchecked = True
                    break
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(any_unchecked)


    def select_all_rows(self):
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, 'findChild'):
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(True)

    def _selected_obs_rows_and_parts(self):
        rows_to_delete = []
        deleted_parts = []
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if not w:
                continue
            cb = w.findChild(QCheckBox)
            if not (cb and cb.isChecked()):
                continue
            rows_to_delete.append(r)
            pit = self.table.item(r, 1)
            part = (pit.text() if pit else '').strip().upper()
            if part:
                deleted_parts.append(part)
        return rows_to_delete, deleted_parts

    def _has_orphan_hierarchy_tab(self):
        if self.orphan_hierarchy_tab is None:
            return False
        return self.subtabs.indexOf(self.orphan_hierarchy_tab) >= 0

    def _nonempty_obs_rows(self):
        rows = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, 1)
            part = (pit.text() if pit else '').strip()
            if part:
                rows.append(r)
        return rows

    def _is_full_obs_deletion(self, rows_to_delete):
        data_rows = self._nonempty_obs_rows()
        if not data_rows:
            return False
        selected = set(rows_to_delete or [])
        return all(r in selected for r in data_rows)

    def _reset_orphan_hierarchy_state(self):
        if self.orphan_hierarchy_table is not None:
            self.orphan_hierarchy_table.setRowCount(0)
        if self.orphan_hierarchy_tab is not None:
            idx = self.subtabs.indexOf(self.orphan_hierarchy_tab)
            if idx >= 0:
                self.subtabs.removeTab(idx)
        # Reset to initial state (same as before first hierarchy build).
        self.orphan_hierarchy_tab = None
        self.orphan_hierarchy_table = None

    def _hier_parse_level(self, table: QTableWidget, row_idx: int):
        it = table.item(row_idx, 0)
        txt = (it.text() if it else '').strip()
        if not txt:
            return None
        try:
            return int(txt)
        except Exception:
            return None

    def _hier_part_key(self, table: QTableWidget, row_idx: int) -> str:
        it = table.item(row_idx, 1)
        return (it.text() if it else '').strip().upper()

    def _hier_subtree_end(self, table: QTableWidget, anchor_row: int) -> int:
        base_lvl = self._hier_parse_level(table, anchor_row)
        if base_lvl is None:
            return anchor_row
        end = anchor_row
        for rr in range(anchor_row + 1, table.rowCount()):
            lvl = self._hier_parse_level(table, rr)
            if lvl is None:
                break
            if lvl <= base_lvl:
                break
            end = rr
        return end

    def _remove_parts_from_orphan_hierarchy(self, deleted_parts):
        ht = self.orphan_hierarchy_table
        if ht is None or not deleted_parts:
            return set()

        targets = {p.strip().upper() for p in deleted_parts if p and p.strip()}
        if not targets:
            return set()

        dependent_parts = set()
        removed_any = True
        while removed_any:
            removed_any = False
            rr = ht.rowCount() - 1
            while rr >= 0:
                lvl = self._hier_parse_level(ht, rr)
                if lvl is None:
                    rr -= 1
                    continue

                root_part = self._hier_part_key(ht, rr)
                if root_part not in targets:
                    rr -= 1
                    continue

                end = self._hier_subtree_end(ht, rr)
                base_lvl = lvl
                for i in range(rr, end + 1):
                    ilvl = self._hier_parse_level(ht, i)
                    if ilvl is None:
                        continue
                    pkey = self._hier_part_key(ht, i)
                    if not pkey:
                        continue
                    if ilvl > base_lvl:
                        dependent_parts.add(pkey)
                    # Ensure duplicates are removed across all blocks in subsequent scans.
                    targets.add(pkey)

                for _ in range(end - rr + 1):
                    ht.removeRow(rr)

                removed_any = True
                rr -= 1

        return dependent_parts

    def _remove_orphan_children_from_final_obs(self, candidate_parts):
        if not candidate_parts:
            return
        part_set = {p.strip().upper() for p in candidate_parts if p and p.strip()}
        if not part_set:
            return

        identified_col = -1
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h and h.text().strip().lower() == 'identified orphans':
                identified_col = c
                break
        if identified_col < 0:
            return

        rows_to_delete = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, 1)
            part = (pit.text() if pit else '').strip().upper()
            if part not in part_set:
                continue
            iit = self.table.item(r, identified_col)
            identified = (iit.text() if iit else '').strip().lower()
            if re.match(r'^orphan\s*\d+$', identified):
                rows_to_delete.append(r)

        for r in reversed(rows_to_delete):
            self.table.removeRow(r)

    def _cleanup_invalid_hierarchy_blocks(self):
        ht = self.orphan_hierarchy_table
        if ht is None:
            return

        blocks = []
        start = -1
        for rr in range(ht.rowCount()):
            lvl = self._hier_parse_level(ht, rr)
            if lvl is None:
                part = self._hier_part_key(ht, rr).lower()
                if part.startswith('block '):
                    if start >= 0:
                        blocks.append((start, rr - 1))
                    start = rr
        if start >= 0:
            blocks.append((start, ht.rowCount() - 1))

        to_remove = []
        for bstart, bend in blocks:
            has_orphan_child = False
            has_data_row = False
            for rr in range(bstart + 1, bend + 1):
                lvl = self._hier_parse_level(ht, rr)
                if lvl is None:
                    continue
                has_data_row = True
                iit = ht.item(rr, 2)
                identified = (iit.text() if iit else '').strip().lower()
                if re.match(r'^orphan\s*\d+$', identified):
                    has_orphan_child = True
                    break
            if (not has_data_row) or (not has_orphan_child):
                to_remove.append((bstart, bend))

        for bstart, bend in reversed(to_remove):
            for _ in range(bend - bstart + 1):
                ht.removeRow(bstart)

        # Hide tab if hierarchy has no data rows left.
        has_data = False
        for rr in range(ht.rowCount()):
            if self._hier_parse_level(ht, rr) is not None:
                has_data = True
                break
        if not has_data and self.orphan_hierarchy_tab is not None:
            idx = self.subtabs.indexOf(self.orphan_hierarchy_tab)
            if idx >= 0:
                self.subtabs.removeTab(idx)

    def delete_selected_rows_with_sync(self):
        rows_to_delete, deleted_parts = self._selected_obs_rows_and_parts()
        if not rows_to_delete:
            return

        has_hierarchy_tab = self._has_orphan_hierarchy_tab()
        full_delete = has_hierarchy_tab and self._is_full_obs_deletion(rows_to_delete)

        if not has_hierarchy_tab:
            for r in reversed(rows_to_delete):
                self.table.removeRow(r)
            return

        if full_delete:
            msg = (
                'Warning: You are deleting all entries from Final OBS List.\n\n'
                'This will completely clear the Orphan Hierarchy, including all blocks and relationships.\n\n'
                'Do you want to continue?'
            )
        else:
            msg = (
                'Warning: Deleting this part will remove all its dependent child items.\n\n'
                'This action will:\n'
                '- Update all related blocks in Orphan Hierarchy\n'
                '- Remove occurrences across ALL blocks\n'
                '- Delete associated child parts from Final OBS Parts (only if classified as Orphans)\n'
                '- Remove invalid or empty hierarchy blocks\n\n'
                'Do you want to continue?'
            )

        reply = QMessageBox.question(
            self,
            'Delete Confirmation',
            msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        for r in reversed(rows_to_delete):
            self.table.removeRow(r)

        if full_delete:
            # Clean slate requested by user after full delete.
            self.table.setRowCount(0)
            self._reset_orphan_hierarchy_state()
            return

        dependent_parts = self._remove_parts_from_orphan_hierarchy(deleted_parts)
        self._remove_orphan_children_from_final_obs(dependent_parts)
        self._cleanup_invalid_hierarchy_blocks()

        # If no OBS parts remain after dependent cleanup, reset hierarchy fully.
        if not self._nonempty_obs_rows():
            self.table.setRowCount(0)
            self._reset_orphan_hierarchy_state()

    def __init__(self):
        super().__init__()
        outer=QVBoxLayout(self)

        title_row=QHBoxLayout()
        title=QLabel("Final OBS List"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold)); title.setStyleSheet("color:#C1272D;")
        btn_template=QPushButton("Download Template")
        btn_upload=QPushButton("Upload OBS List")
        btn_export_obs=QPushButton("Export OBS List")
        btn_export_obs.setToolTip("Export the OBS table (including Proposed Replacement column) to an Excel file")
        # New copy buttons
        btn_copy_obs=QPushButton("Copy OBS Parts")
        btn_copy_rep=QPushButton("Copy Repl Parts")
        delete_btn=QPushButton("Delete Selected")
        title_row.addWidget(title)
        title_row.addStretch(1)
        title_row.addWidget(btn_template)
        title_row.addWidget(btn_upload)
        title_row.addWidget(btn_export_obs)
        title_row.addWidget(btn_copy_obs)
        title_row.addWidget(btn_copy_rep)
        select_all_btn = QPushButton(" Select All")
        select_all_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))
        select_all_btn.setToolTip("Toggle select / unselect all rows")
        btn_approve_replacement = QPushButton("Approve replacement for selected parts")
        btn_approve_replacement.setToolTip(
            "Copy Proposed Replacement values into the Replacement column for checked rows.\n"
            "If no rows are checked, prompts to accept all proposed replacements."
        )
        title_row.addWidget(select_all_btn)
        title_row.addWidget(btn_approve_replacement)
        title_row.addWidget(delete_btn)
        outer.addLayout(title_row)

        # (WU controls moved to Where Used tab)

        legend = QLabel(
            "<b>Orphan Legend:</b> "
            "<span style='color:#C0392B; font-weight:600;'>Orphan1</span> | "
            "<span style='color:#E67E22; font-weight:600;'>Orphan2</span> | "
            "<span style='color:#2980B9; font-weight:600;'>Orphan3+</span>"
        )
        legend.setStyleSheet("padding:4px;")
        outer.addWidget(legend)

        self.proposed_replacement_note = QLabel(
            "Note: The replacement is determined by comparing the BOM items of the "
            "<b>OBS</b> and <b>Replacement</b> parent parts, and replacements are proposed "
            "by matching item descriptions of BOM Items."
        )
        self.proposed_replacement_note.setWordWrap(True)
        self.proposed_replacement_note.setVisible(False)
        self.proposed_replacement_note.setStyleSheet(
            "color:#C1272D; padding:2px 0 6px 0;"
        )
        outer.addWidget(self.proposed_replacement_note)


        self.subtabs = QTabWidget(self)
        self.final_obs_tab = QWidget(self)
        final_layout = QVBoxLayout(self.final_obs_tab)
        final_layout.setContentsMargins(0, 0, 0, 0)
        final_layout.setSpacing(0)

        self.table=OBSTable(self, initial_rows=1)
        final_layout.addWidget(self.table)
        self.subtabs.addTab(self.final_obs_tab, "Final OBS List")

        self.orphan_hierarchy_tab = None
        self.orphan_hierarchy_table = None

        outer.addWidget(self.subtabs)

        self.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#FFF5F5; gridline-color:#F3C2C2; }
            QHeaderView::section { background:#F8D7DA; color:#7A1C21; font-weight:600; border:1px solid #E3AEB2; padding:4px; }
            QTableWidget::item:selected { background:#F5B5B8; color:#4A0E10; }
            QPushButton { background-color:#C1272D; color:#FFFFFF; border:1px solid #9F1F24; border-radius:5px; padding:6px 10px; }
            QPushButton:hover { background-color:#AD2227; }
            QComboBox { border:1px solid #E3AEB2; border-radius:4px; padding:2px 6px; }
        """)

        delete_btn.clicked.connect(self.delete_selected_rows_with_sync)
        btn_template.clicked.connect(self.download_template)
        btn_upload.clicked.connect(self.upload_from_excel)
        btn_copy_obs.clicked.connect(self.copy_obs_parts)
        btn_copy_rep.clicked.connect(self.copy_replacement_parts)
        select_all_btn.clicked.connect(self.toggle_select_all)
        btn_approve_replacement.clicked.connect(self.approve_proposed_replacements)
        btn_export_obs.clicked.connect(self.export_obs_list)
        # (WU controls moved to Where Used tab)
        self.where_used_tab = None  # linked by MainWindow after both tabs are created

    def _ensure_orphan_hierarchy_tab(self):
        if self.orphan_hierarchy_tab is not None and self.orphan_hierarchy_table is not None:
            return
        self.orphan_hierarchy_tab = QWidget(self)
        lay = QVBoxLayout(self.orphan_hierarchy_tab)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self.orphan_hierarchy_table = QTableWidget(0, 5, self.orphan_hierarchy_tab)
        self.orphan_hierarchy_table.setHorizontalHeaderLabels([
            'BOM Level',
            'Part Hierarchy',
            'Identified Orphans',
            'Change',
            'Replacement',
        ])
        self.orphan_hierarchy_table.verticalHeader().setVisible(False)
        self.orphan_hierarchy_table.setAlternatingRowColors(True)
        self.orphan_hierarchy_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.orphan_hierarchy_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.orphan_hierarchy_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        try:
            hh = self.orphan_hierarchy_table.horizontalHeader()
            hh.setStretchLastSection(False)
            hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
            hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
            hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
            hh.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
            hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
            self.orphan_hierarchy_table.setColumnWidth(0, _excel_width_to_px(self.orphan_hierarchy_table, 10))
            self.orphan_hierarchy_table.setColumnWidth(1, _excel_width_to_px(self.orphan_hierarchy_table, 34))
            self.orphan_hierarchy_table.setColumnWidth(2, _excel_width_to_px(self.orphan_hierarchy_table, 18))
            self.orphan_hierarchy_table.setColumnWidth(3, _excel_width_to_px(self.orphan_hierarchy_table, 14))
            self.orphan_hierarchy_table.setColumnWidth(4, _excel_width_to_px(self.orphan_hierarchy_table, 18))
        except Exception:
            pass

        lay.addWidget(self.orphan_hierarchy_table)
        self.subtabs.addTab(self.orphan_hierarchy_tab, 'Orphan Hierarchy')

    def update_orphan_hierarchy_view(self, orphan_blocks: List[Dict[str, Any]], wu_replacement_map=None):
        """Append-only hierarchy renderer using flat Block groups.

        - Structure comes from WU orphan blocks only.
        - Final OBS List is used for orphan labels/change/replacement lookups.
        - Existing hierarchy rows are preserved; only new relations are appended.
        """
        t = self.table

        rows = []
        for r in range(t.rowCount()):
            pit = t.item(r, 1)
            part = (pit.text() if pit else '').strip()
            if not part:
                continue

            change_val = ''
            cw = t.cellWidget(r, 2)
            if isinstance(cw, QComboBox):
                change_val = (cw.currentText() or '').strip()
            else:
                cit = t.item(r, 2)
                change_val = (cit.text() if cit else '').strip()

            rep_it = t.item(r, 3)
            rep_val = (rep_it.text() if rep_it else '').strip()

            identified = ''
            for c in range(t.columnCount()):
                h = t.horizontalHeaderItem(c)
                if h and h.text().strip().lower() == 'identified orphans':
                    iit = t.item(r, c)
                    identified = (iit.text() if iit else '').strip()
                    break

            rows.append({
                'idx': r,
                'part': part,
                'key': part.upper(),
                'identified': identified,
                'change': change_val,
                'replacement': rep_val,
            })

        if not rows:
            # Final OBS List is empty: reset and hide Orphan Hierarchy tab.
            if self.orphan_hierarchy_table is not None:
                self.orphan_hierarchy_table.setRowCount(0)
            if self.orphan_hierarchy_tab is not None:
                idx = self.subtabs.indexOf(self.orphan_hierarchy_tab)
                if idx >= 0:
                    self.subtabs.removeTab(idx)
            return

        self._ensure_orphan_hierarchy_tab()
        ht = self.orphan_hierarchy_table
        if ht is None:
            return

        if not orphan_blocks:
            ht.setRowCount(0)
            return

        by_key: Dict[str, dict] = {}
        label_by_key: Dict[str, str] = {}
        for row in rows:
            if row['key'] not in by_key:
                by_key[row['key']] = row
            lbl = (row.get('identified') or '').strip()
            if lbl:
                label_by_key[row['key']] = lbl

        wu_replacement_map = wu_replacement_map or {}

        def _row_for_part(part_key: str) -> dict:
            row = by_key.get(part_key)
            if row:
                rep = (row.get('replacement') or '').strip()
                if not rep:
                    rep = (wu_replacement_map.get(part_key, '') or '').strip()
                out = dict(row)
                out['replacement'] = rep
                return out
            return {
                'part': part_key,
                'key': part_key,
                'identified': '',
                'change': '',
                'replacement': (wu_replacement_map.get(part_key, '') or '').strip(),
            }

        def _parse_level(row_idx: int):
            it = ht.item(row_idx, 0)
            txt = (it.text() if it else '').strip()
            if not txt:
                return None
            try:
                return int(txt)
            except ValueError:
                return None

        def _part_at(row_idx: int) -> str:
            it = ht.item(row_idx, 1)
            return (it.text() if it else '').strip()

        def _part_key_at(row_idx: int) -> str:
            return _part_at(row_idx).strip().upper()

        def _write_hierarchy_row(row_idx: int, level: int, part_key: str, is_header: bool = False):
            if is_header:
                ht.setItem(row_idx, 0, QTableWidgetItem(''))
                head = QTableWidgetItem(part_key)
                _f = head.font(); _f.setBold(True); head.setFont(_f)
                ht.setItem(row_idx, 1, head)
                ht.setItem(row_idx, 2, QTableWidgetItem(''))
                ht.setItem(row_idx, 3, QTableWidgetItem(''))
                ht.setItem(row_idx, 4, QTableWidgetItem(''))
                return

            row = _row_for_part(part_key)
            indent = '    ' * max(0, level)
            ht.setItem(row_idx, 0, QTableWidgetItem(str(max(0, level))))
            ht.setItem(row_idx, 1, QTableWidgetItem(f"{indent}{row.get('part', part_key)}"))
            ht.setItem(row_idx, 2, QTableWidgetItem(row.get('identified', '')))
            ht.setItem(row_idx, 3, QTableWidgetItem(row.get('change', '')))
            ht.setItem(row_idx, 4, QTableWidgetItem(row.get('replacement', '')))

        def _find_last_row_for_part(part_key: str):
            key = (part_key or '').strip().upper()
            found = -1
            for rr in range(ht.rowCount()):
                if _part_key_at(rr) == key:
                    found = rr
            return found

        def _subtree_end(anchor_row: int):
            base_lvl = _parse_level(anchor_row)
            if base_lvl is None:
                return anchor_row
            end = anchor_row
            for rr in range(anchor_row + 1, ht.rowCount()):
                lvl = _parse_level(rr)
                if lvl is None:
                    break
                if lvl <= base_lvl:
                    break
                end = rr
            return end

        def _parent_child_pairs_existing():
            pairs = set()
            stack: List[tuple[int, str]] = []
            for rr in range(ht.rowCount()):
                lvl = _parse_level(rr)
                if lvl is None:
                    stack.clear()
                    continue
                part_key = _part_key_at(rr)
                if not part_key:
                    continue
                while stack and stack[-1][0] >= lvl:
                    stack.pop()
                if stack:
                    pairs.add((stack[-1][1], part_key))
                stack.append((lvl, part_key))
            return pairs

        def _next_block_label() -> str:
            max_block = 0
            for rr in range(ht.rowCount()):
                if _parse_level(rr) is not None:
                    continue
                txt = _part_at(rr).strip()
                if not txt.lower().startswith('block '):
                    continue
                try:
                    num = int(txt.split(' ', 1)[1].strip())
                    if num > max_block:
                        max_block = num
                except Exception:
                    continue
            return f'Block {max_block + 1}'

        # Remove scope section rows if present from previous runs.
        for rr in range(ht.rowCount() - 1, -1, -1):
            if _parse_level(rr) is not None:
                continue
            txt = _part_at(rr).strip().lower()
            if txt in ('=== original scope ===', '=== orphan parents scope (extended) ==='):
                ht.removeRow(rr)

        # Refresh existing hierarchy row values so Replacement gets filled after each run.
        for rr in range(ht.rowCount()):
            lvl = _parse_level(rr)
            if lvl is None:
                continue
            part_key = _part_key_at(rr)
            if not part_key:
                continue
            row = _row_for_part(part_key)
            ht.setItem(rr, 2, QTableWidgetItem(row.get('identified', '')))
            ht.setItem(rr, 3, QTableWidgetItem(row.get('change', '')))
            ht.setItem(rr, 4, QTableWidgetItem(row.get('replacement', '')))

        existing_edges = _parent_child_pairs_existing()

        def _is_orphan_parent_part(part_key: str) -> bool:
            lbl = (label_by_key.get(part_key, '') or '').strip().lower()
            return lbl.startswith('orphan')

        def _append_group(header_label: str, parents: List[str], child_key: str):
            insert_at = ht.rowCount()
            ht.insertRow(insert_at)
            _write_hierarchy_row(insert_at, 0, header_label, is_header=True)

            current = insert_at + 1
            for p in parents:
                ht.insertRow(current)
                _write_hierarchy_row(current, 0, p)
                current += 1

            ht.insertRow(current)
            _write_hierarchy_row(current, 1, child_key)

        for block in orphan_blocks:
            parents = [(p or '').strip().upper() for p in (block.get('parents', []) or [])]
            parents = [p for p in parents if p]
            uniq_parents = []
            for p in parents:
                if p not in uniq_parents:
                    uniq_parents.append(p)
            parents = uniq_parents

            child_key = (block.get('child') or '').strip().upper()
            if not child_key:
                continue

            if any((p, child_key) in existing_edges for p in parents):
                continue

            orphan_parent_roots = [p for p in parents if _is_orphan_parent_part(p)]

            if orphan_parent_roots:
                for op in orphan_parent_roots:
                    if (op, child_key) in existing_edges:
                        continue
                    anchor = _find_last_row_for_part(op)
                    if anchor >= 0:
                        lvl = _parse_level(anchor)
                        if lvl is None:
                            lvl = 0
                        at = _subtree_end(anchor) + 1
                        ht.insertRow(at)
                        _write_hierarchy_row(at, lvl + 1, child_key)
                    else:
                        _append_group(_next_block_label(), [op], child_key)
                    existing_edges.add((op, child_key))
            else:
                _append_group(_next_block_label(), parents, child_key)
                for p in parents:
                    existing_edges.add((p, child_key))

        if self.subtabs.indexOf(self.orphan_hierarchy_tab) >= 0:
            self.subtabs.setCurrentWidget(self.orphan_hierarchy_tab)

    def update_proposed_replacement_note_visibility(self, visible=None):
        if visible is None:
            visible = any(
                (self.table.horizontalHeaderItem(c)
                 and self.table.horizontalHeaderItem(c).text().strip().lower() == 'proposed replacement')
                for c in range(self.table.columnCount())
            )
        self.proposed_replacement_note.setVisible(bool(visible))

    def download_template(self):
        try:
            path,_=QFileDialog.getSaveFileName(self,'Save Template',str(TEMPLATE_FILE),'Excel Files (*.xlsx)')
            if not path: return
            import pandas as pd
            df=pd.DataFrame({'OBS Parts':[''],'Change':['Obsolete'],'Replacement':['']})
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='OBS_Template')
            QMessageBox.information(self,'Template Saved',f'Template saved to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self,'Template Error',str(e))

    def upload_from_excel(self):
        try:
            path,_=QFileDialog.getOpenFileName(self,'Open OBS Parts Excel','', 'Excel Files (*.xlsx *.xls)')
            if not path: return
            import pandas as pd
            def as_text(value: Any) -> str:
                if value is None:
                    return ''
                try:
                    if pd.isna(value):
                        return ''
                except Exception:
                    pass
                return str(value).strip()
            if path.lower().endswith('.xls'):
                df=pd.read_excel(path, engine='xlrd', dtype=str, keep_default_na=False)
            else:
                df=pd.read_excel(path, engine='openpyxl', dtype=str, keep_default_na=False)
            cols={str(c).strip().lower(): c for c in df.columns}
            def pick(name):
                for key in cols:
                    if key==name: return cols[key]
                return None
            c_obs=pick('obs parts') or pick('obs part') or pick('part')
            c_change=pick('change')
            c_rep=(pick('replacement') or pick('repl') or pick('replace') or
                   pick('replacement part') or pick('new part') or pick('irplacement'))
            if not c_obs: raise ValueError('Column "OBS Parts" is required in the Excel file.')
            rows=[]
            for _,r in df.iterrows():
                obs=as_text(r.get(c_obs,''))
                if not obs: continue
                change_val=as_text(r.get(c_change,'Obsolete')) if c_change else 'Obsolete'
                if change_val not in ['', 'Obsolete','Inactivate']: change_val='Obsolete'
                rep=as_text(r.get(c_rep,'')) if c_rep else ''
                rows.append((obs,change_val,rep))
            if not rows:
                QMessageBox.information(self,'No Data','No valid rows found in the Excel file.'); return
            t=self.table; t.setRowCount(len(rows)); t._init_rows(0,len(rows))
            for r,(obs,change,rep) in enumerate(rows):
                t.setItem(r,1,QTableWidgetItem(obs))
                w=t.cellWidget(r,2)
                if isinstance(w,QComboBox):
                    idx=w.findText(change); w.setCurrentIndex(idx if idx>=0 else 0)
                rep_item = t.item(r,3)
                if rep_item is None:
                    rep_item = QTableWidgetItem('')
                    t.setItem(r,3,rep_item)
                rep_item.setText(rep)
            t._apply_excel_widths(); QMessageBox.information(self,'Upload Complete',f'Loaded {len(rows)} rows from Excel.')
        except Exception as e:
            QMessageBox.warning(self,'Upload Error',str(e))

    def export_obs_list(self):
        """Export the OBS table as-is (including Proposed Replacement if present) to Excel."""
        try:
            import pandas as pd
            t = self.table
            col_count = t.columnCount()
            row_count = t.rowCount()

            # Collect headers (skip the Select checkbox column at index 0)
            headers = []
            col_indices = []
            for c in range(col_count):
                if c == 0:
                    continue
                h = t.horizontalHeaderItem(c)
                headers.append(h.text().strip() if h else f'Col{c}')
                col_indices.append(c)

            # Collect data rows — skip rows with no OBS part value
            data_rows = []
            obs_col_local = next(
                (i for i, h in enumerate(headers) if h.strip().lower() == 'obs parts'), -1
            )
            for r in range(row_count):
                row_data = []
                for c in col_indices:
                    if c == 2:  # Change combo
                        w = t.cellWidget(r, c)
                        val = w.currentText() if isinstance(w, QComboBox) else ''
                    else:
                        it = t.item(r, c)
                        val = it.text().strip() if it else ''
                    row_data.append(val)
                # Skip entirely empty rows
                if obs_col_local >= 0 and not row_data[obs_col_local].strip():
                    continue
                data_rows.append(row_data)

            if not data_rows:
                QMessageBox.information(self, 'Export OBS List', 'No data to export.')
                return

            df = pd.DataFrame(data_rows, columns=headers)
            path, _ = QFileDialog.getSaveFileName(
                self, 'Export OBS List', 'OBS_List.xlsx', 'Excel Files (*.xlsx)'
            )
            if not path:
                return
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='OBS List')
            QMessageBox.information(self, 'Export Complete', f'OBS List exported to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self, 'Export Error', str(e))

    # NEW: Copy helpers
    def _collect_column_values(self, col_index: int) -> List[str]:
        t=self.table
        values=[]
        for r in range(t.rowCount()):
            it=t.item(r,col_index)
            if it:
                val=(it.text() or '').strip()
                if val:
                    values.append(val)
        return values

    def copy_obs_parts(self):
        values=self._collect_column_values(1)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} OBS Part number(s) to clipboard.')

    def copy_replacement_parts(self):
        values=self._collect_column_values(3)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} Replacement part number(s) to clipboard (Image).')

    def to_dict(self)->Dict[str,Any]:
        t=self.table; rows:List[Dict[str,Any]]=[]
        for r in range(t.rowCount()):
            obs=t.item(r,1).text() if t.item(r,1) else ''
            rep=t.item(r,3).text() if t.item(r,3) else ''
            w=t.cellWidget(r,2); change=w.currentText() if isinstance(w,QComboBox) else 'Obsolete'
            if any([obs.strip(), rep.strip()]): rows.append({'obs_part':obs,'change':change,'replacement':rep})
        return {'rows': rows}

    def from_dict(self,data:Dict[str,Any]):
        rows=data.get('rows',[]); t=self.table; needed=max(10,len(rows)); t.setRowCount(needed); t._init_rows(0,needed)
        for r,row in enumerate(rows):
            t.setItem(r,1,QTableWidgetItem(row.get('obs_part','')))
            w=t.cellWidget(r,2)
            if isinstance(w,QComboBox):
                idx=w.findText(row.get('change','Obsolete'))
                if idx>=0: w.setCurrentIndex(idx)
            t.setItem(r,3,QTableWidgetItem(row.get('replacement','')))
        t._apply_excel_widths()

    def reset(self):
        t=self.table; t.setRowCount(10); t._init_rows(0,10)

    def approve_proposed_replacements(self):
        """Copy Proposed Replacement → Replacement for selected rows (or all if none selected)."""
        t = self.table

        # Locate required columns
        prop_col = -1
        repl_col = -1
        obs_col = -1
        for c in range(t.columnCount()):
            h = t.horizontalHeaderItem(c)
            if h:
                txt = h.text().strip().lower()
                if txt == 'proposed replacement':
                    prop_col = c
                elif txt == 'replacement':
                    repl_col = c
                elif txt == 'obs parts':
                    obs_col = c

        if prop_col < 0:
            QMessageBox.information(self, 'No Proposed Replacements',
                'No "Proposed Replacement" column found.\n'
                'Run the orphan analysis to generate proposed replacements first.')
            return
        if repl_col < 0:
            QMessageBox.warning(self, 'Error', 'Could not find "Replacement" column.')
            return

        # Determine which rows the user has checked
        selected_rows = []
        for r in range(t.rowCount()):
            w = t.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and cb.isChecked():
                    selected_rows.append(r)

        if selected_rows:
            rows_to_process = selected_rows
        else:
            reply = QMessageBox.question(
                self, 'Confirm Approve All',
                'Accept all proposed replacements as Valid Replacements for the Identified Orphans.',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            rows_to_process = list(range(t.rowCount()))

        # Apply proposed → replacement and build a map for Where Used update
        accepted_map: Dict[str, str] = {}  # upper(part_number) -> replacement_value
        for r in rows_to_process:
            prop_item = t.item(r, prop_col)
            prop_val = prop_item.text().strip() if prop_item else ''
            if not prop_val:
                continue
            repl_item = t.item(r, repl_col)
            if repl_item is None:
                repl_item = QTableWidgetItem('')
                t.setItem(r, repl_col, repl_item)
            repl_item.setText(prop_val)
            if obs_col >= 0:
                obs_item = t.item(r, obs_col)
                part_num = obs_item.text().strip() if obs_item else ''
                if part_num:
                    accepted_map[part_num.upper()] = prop_val

        if not accepted_map:
            QMessageBox.information(self, 'Nothing to Accept',
                'No non-empty proposed replacement values found for the selected rows.')
            return

        # Check whether ALL rows that have a proposed value were processed
        rows_to_process_set = set(rows_to_process)
        all_accepted = all(
            r in rows_to_process_set
            for r in range(t.rowCount())
            if (t.item(r, prop_col) and t.item(r, prop_col).text().strip())
        )

        if all_accepted:
            t.removeColumn(prop_col)
            self.update_proposed_replacement_note_visibility(False)

        # Update Where Used tab if available
        if self.where_used_tab is not None and accepted_map:
            wu_table = self.where_used_tab.table
            if (wu_table is not None
                    and wu_table.rowCount() > 0
                    and wu_table.columnCount() > 0):
                wu_part_col = -1
                wu_repl_col = -1
                for c in range(wu_table.columnCount()):
                    h = wu_table.horizontalHeaderItem(c)
                    if h:
                        txt = h.text().strip().lower()
                        if txt in ('part', 'part number', 'material number'):
                            wu_part_col = c
                        elif txt == 'replacement':
                            wu_repl_col = c
                if wu_part_col >= 0 and wu_repl_col >= 0:
                    for r in range(wu_table.rowCount()):
                        part_item = wu_table.item(r, wu_part_col)
                        part_val = part_item.text().strip().upper() if part_item else ''
                        if part_val in accepted_map:
                            repl_item = wu_table.item(r, wu_repl_col)
                            if repl_item is None:
                                repl_item = QTableWidgetItem('')
                                wu_table.setItem(r, wu_repl_col, repl_item)
                            repl_item.setText(accepted_map[part_val])

        QMessageBox.information(
            self, 'Update Complete',
            'Replacement update completed. Please proceed to the next level of Orphan Analysis.'
        )

    def launch_where_used_import(self):
        """Validate WU level + OBS parts, then delegate to the linked Where Used tab."""
        # ── Validate WU Level ──────────────────────────────────────────────────
        raw_level = self.wu_level_input.text().strip()
        if not raw_level:
            QMessageBox.warning(
                self, 'WU Level Required',
                'Please enter a WU Level (1 to 6) before importing.'
            )
            return
        try:
            wu_level = int(raw_level)
            if not (1 <= wu_level <= 6):
                raise ValueError
        except ValueError:
            QMessageBox.warning(
                self, 'Invalid WU Level',
                f'"{raw_level}" is not valid.  Please enter a whole number from 1 to 6.'
            )
            return

        # ── Collect non-empty OBS part numbers ────────────────────────────────
        t = self.table
        obs_parts: List[str] = []
        for r in range(t.rowCount()):
            it = t.item(r, 1)
            val = (it.text() if it else '').strip()
            if val:
                obs_parts.append(val)

        if not obs_parts:
            QMessageBox.warning(
                self, 'No OBS Parts',
                'The OBS Parts column is empty.\n'
                'Please enter at least one part number before importing.'
            )
            return

        # ── Delegate to Where Used tab ─────────────────────────────────────────
        if self.where_used_tab is None:
            QMessageBox.warning(self, 'Not Ready', 'Where Used tab is not available yet.')
            return

        plant = self.plant_combo.currentText().strip()
        self.where_used_tab.import_from_databricks(obs_parts, wu_level, plant)

        # Switch focus to the Where Used tab after a successful import
        try:
            main = self.window()
            if hasattr(main, 'tabs'):
                main.tabs.setCurrentWidget(self.where_used_tab)
        except Exception:
            pass


class WhereUsedTab(QWidget):
    """Import and manage Where Used parent rows with OBS-aware mapping.
    Data cleanup is applied to the in-app table only and source files are unchanged."""
    def __init__(self, obs_provider=None):
        super().__init__()
        self.table = QTableWidget(0,0)
        print('DEBUG: WhereUsedTab self.table initialized:', self.table)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)
        title_row = QHBoxLayout()
        title = QLabel("Where Used - Parents"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold))
        btn_import = QPushButton("Import 'Where Used' Parents")
        btn_import.setToolTip("Select Excel/CSV/HTML. Legacy/mismatched files are auto-converted to .xlsx using Excel, then imported. Cleanup + OBS mapping are applied to the in-app view only.")
        title_row.addWidget(title); title_row.addStretch(1); title_row.addWidget(btn_import)
        outer.addLayout(title_row)

        # --- Where Used of OBS Parts controls (moved from OBS tab) ---

        # --- Controls row: WU Level, Plant, Where Used of OBS Parts button, and radio selections ---
        controls_row = QHBoxLayout()
        wu_lbl = QLabel("WU Level (1–6):")
        wu_lbl.setStyleSheet("font-weight:600; color:#7A1C21;")
        self.wu_level_input = QLineEdit()
        self.wu_level_input.setFixedWidth(55)
        self.wu_level_input.setPlaceholderText("1–6")
        self.wu_level_input.setMaxLength(1)
        self.wu_level_input.setToolTip("Maximum Where Used depth to retrieve (1 to 6)")
        plant_lbl = QLabel("Plant:")
        plant_lbl.setStyleSheet("font-weight:600; color:#7A1C21;")
        self.plant_combo = QComboBox()
        self.plant_combo.addItems(["4020", "4055", "4060", "4070", "4080", "4090"])
        self.plant_combo.setCurrentText("4070")
        self.plant_combo.setToolTip("Plant code to filter Where Used query")
        self.plant_combo.setFixedWidth(75)
        controls_row.addWidget(wu_lbl)
        controls_row.addWidget(self.wu_level_input)
        controls_row.addWidget(plant_lbl)
        controls_row.addWidget(self.plant_combo)

        # --- Where Used of OBS Parts button (in place of select buttons) ---
        self.btn_where_used = QPushButton("Where Used of OBS Parts")
        self.btn_where_used.setToolTip("Query Databricks for multi-level Where Used data for all OBS parts")
        self.btn_where_used.setFixedWidth(180)
        self.btn_where_used.setFixedHeight(28)
        self.btn_where_used.setStyleSheet("font-size:12px; padding:2px 8px;")
        self.btn_where_used.clicked.connect(self._handle_where_used_obs_parts)
        controls_row.addSpacing(16)
        controls_row.addWidget(self.btn_where_used)

        # --- Radio selections next to Plant selection with indentation ---
        controls_row.addSpacing(32)
        def add_radio_group_inline(layout, label_text, group_attr, retain_attr, remove_attr):
            label = QLabel(f"<b>{label_text}</b>")
            label.setStyleSheet("font-size:13px; margin-right:6px;")
            group = QButtonGroup(self)
            retain = QRadioButton("Retain")
            remove = QRadioButton("Remove")
            remove.setChecked(True)
            group.addButton(retain)
            group.addButton(remove)
            layout.addWidget(label)
            layout.addWidget(retain)
            layout.addWidget(remove)
            setattr(self, group_attr, group)
            setattr(self, retain_attr, retain)
            setattr(self, remove_attr, remove)

        add_radio_group_inline(controls_row, "9024 Parents:", 'radio_9024_group', 'radio_9024_retain', 'radio_9024_remove')
        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.VLine)
        sep1.setFrameShadow(QFrame.Shadow.Sunken)
        sep1.setLineWidth(1)
        controls_row.addWidget(sep1)
        add_radio_group_inline(controls_row, "ESW Parents:", 'radio_esw_group', 'radio_esw_retain', 'radio_esw_remove')
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.VLine)
        sep2.setFrameShadow(QFrame.Shadow.Sunken)
        sep2.setLineWidth(1)
        controls_row.addWidget(sep2)
        add_radio_group_inline(controls_row, "SmBOM above Config:", 'radio_above_cfg_group', 'radio_above_cfg_retain', 'radio_above_cfg_remove')
        controls_row.addStretch(1)
        outer.addLayout(controls_row)

        outer.addWidget(self.table)

        self.setStyleSheet("""
        QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
        QHeaderView::section { background:#E1F0FF; color:#0F2D46; font-weight:600; border:1px solid #C9E2FF; padding:4px; }
        QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        btn_import.clicked.connect(self.import_where_used)
    def _handle_where_used_obs_parts(self):
        # Validate WU Level
        raw_level = self.wu_level_input.text().strip()
        if not raw_level:
            QMessageBox.warning(self, 'WU Level Required', 'Please enter a WU Level (1 to 6) before importing.')
            return
        try:
            wu_level = int(raw_level)
            if not (1 <= wu_level <= 6):
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, 'Invalid WU Level', f'"{raw_level}" is not valid.  Please enter a whole number from 1 to 6.')
            return
        # Get OBS parts from OBS tab
        obs_tab = getattr(self.window(), 'obs_tab', None)
        obs_parts = []
        if obs_tab and hasattr(obs_tab, 'table'):
            t = obs_tab.table
            for r in range(t.rowCount()):
                it = t.item(r, 1)
                val = (it.text() if it else '').strip()
                if val:
                    obs_parts.append(val)
        if not obs_parts:
            QMessageBox.warning(self, 'No OBS Parts', 'The OBS Parts column is empty. Please enter at least one part number before importing.')
            return
        plant = self.plant_combo.currentText().strip()
        # Read radio button selections
        retain_9024 = self.radio_9024_retain.isChecked()
        retain_esw = self.radio_esw_retain.isChecked()
        retain_above_cfg = self.radio_above_cfg_retain.isChecked()
        # Pass these selections to import_from_databricks
        self.import_from_databricks(obs_parts, wu_level, plant, retain_9024, retain_esw, retain_above_cfg)

    # (Removed misplaced constructor-only UI wiring from this method)

    # ---------------- Excel conversion helpers ----------------
    def _is_html_like(self, path: str)->bool:
        try:
            with open(path, 'rb') as f:
                head = f.read(2048).lstrip().lower()
                return (head.startswith(b'<!') or head.startswith(b'<html') or b'<table' in head)
        except Exception:
            return False

    def _convert_to_xlsx_via_excel(self, src_path: str)->str|None:
        try:
            from pathlib import Path as _Path
            import win32com.client
            src_abs=str(_Path(src_path).resolve())
            dst_path=str(_Path(src_abs).with_suffix(''))+'_converted.xlsx'
            excel=win32com.client.DispatchEx('Excel.Application'); excel.Visible=False; excel.DisplayAlerts=False
            wb=excel.Workbooks.Open(src_abs); wb.SaveAs(dst_path, FileFormat=51); wb.Close(SaveChanges=False); excel.Quit(); return dst_path
        except Exception:
            try:
                excel.Quit()
            except Exception:
                pass
            return None

    # ---------------- Helpers for OBS, colors, selection ----------------
    def _find_part_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            if str(c).strip().lower()=='part': return i
        return -1
    def _find_parent_col_index(self, cols:list[str])->int:
        keys=['parent','parent part','parent pn','parent number','parent part number']
        low=[str(c).strip().lower() for c in cols]
        for i,name in enumerate(low):
            for k in keys:
                if k==name or k in name: return i
        return -1
    def _find_class_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            s=str(c).strip().lower()
            if any(k in s for k in ['class','type','category']): return i
        return -1
    def _build_obs_map(self)->dict:
        mapping={}
        try:
            if self.obs_provider is None: return mapping
            t=self.obs_provider.table
            for r in range(t.rowCount()):
                obs_item=t.item(r,1); rep_item=t.item(r,3)
                obs=(obs_item.text() if obs_item else '').strip(); rep=(rep_item.text() if rep_item else '')
                if obs: mapping[obs.upper()]=rep
        except Exception: pass
        return mapping
    def _read_xlsx_background_colors(self, xlsx_path:str, target_ncols:int):
        try:
            from openpyxl import load_workbook
            wb=load_workbook(xlsx_path, data_only=True); ws=wb.active
            colors=[]; first_data_row=2; ncols=target_ncols
            for r in range(first_data_row, ws.max_row+1):
                row_colors=[]
                for c in range(1, ncols+1):
                    cell=ws.cell(row=r,column=c); col=None
                    try:
                        fill=cell.fill
                        if fill and getattr(fill,'fill_type',None):
                            start=getattr(fill,'start_color',None); rgb=getattr(start,'rgb',None)
                            if isinstance(rgb,str) and len(rgb) in (6,8):
                                rgb_hex=rgb[-6:]; col=QColor('#'+rgb_hex)
                    except Exception: col=None
                    row_colors.append(col)
                colors.append(row_colors)
            return colors
        except Exception:
            return None
    def _apply_cleanup_rules(self, df, part_idx:int):
        import pandas as pd
        part_raw=df.iloc[:,part_idx].astype(str); part_trim=part_raw.str.strip(); lengths=part_trim.str.len().fillna(0)
        mask_gt=(lengths>10) & part_trim.str.upper().str.startswith('ESW')
        mask_eq=(lengths==10) & (part_trim.str[4]=='-')
        mask=mask_gt | mask_eq
        df_kept=df.loc[mask].copy(); df_kept.reset_index(drop=True, inplace=True)
        return df_kept, mask.reset_index(drop=True)
    def _find_replacement_col(self)->int:
        for i in range(self.table.columnCount()):
            h=self.table.horizontalHeaderItem(i)
            if h and str(h.text()).strip().lower()=='replacement': return i
        return -1
    def _center_replacement_column(self):
        rep_col=self._find_replacement_col()
        if rep_col>=0:
            for r in range(self.table.rowCount()):
                it=self.table.item(r,rep_col)
                if it is None:
                    it=QTableWidgetItem(''); self.table.setItem(r,rep_col,it)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def _populate_table_with_extras(self, df, orig_cols:list[str], part_idx:int, colors_2d=None):
        headers=['Select']
        for i,name in enumerate(orig_cols):
            headers.append(str(name))
            if i==part_idx: headers.append('Replacement')
        self.table.clear(); self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers); self.table.setRowCount(len(df))
        obs_map=self._build_obs_map()
        def map_col(c:int)->int: return 1 + c + (1 if c>part_idx else 0)
        for r in range(len(df)):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); cont._chk=chk; self.table.setCellWidget(r,0,cont)
            for c in range(len(orig_cols)):
                val=df.iloc[r,c]; txt='' if val is None else str(val)
                item=QTableWidgetItem(txt)
                if colors_2d and r<len(colors_2d) and c<len(colors_2d[r]):
                    col=colors_2d[r][c]
                    if isinstance(col,QColor): item.setBackground(col)
                self.table.setItem(r, map_col(c), item)
            part_val=str(df.iloc[r,part_idx]) if df.iloc[r,part_idx] is not None else ''
            part_key=part_val.strip().upper(); replacement=obs_map.get(part_key,'')
            rep_item=QTableWidgetItem(replacement); rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if colors_2d and r<len(colors_2d) and part_idx<len(colors_2d[r]):
                col=colors_2d[r][part_idx]
                if isinstance(col,QColor): rep_item.setBackground(col)
            self.table.setItem(r, 1+part_idx+1, rep_item)
        header=self.table.horizontalHeader()
        for i in range(len(headers)):
            if i==len(headers)-1: header.setSectionResizeMode(i,QHeaderView.ResizeMode.Stretch)
            else: header.setSectionResizeMode(i,QHeaderView.ResizeMode.ResizeToContents)

    # -----------------------------------------------------------
    def import_where_used(self):
        try:
            path,_=QFileDialog.getOpenFileName(self, "Import 'Where Used' Parents", '', "Excel/CSV/HTML (*.xlsx *.xls *.csv *.htm *.html);;All Files (*.*)")
            if not path: return
            import pandas as pd
            df=None; errors=[]; lower=path.lower()
            needs_excel_conversion = lower.endswith('.xls') or self._is_html_like(path)
            converted_path=None
            if needs_excel_conversion:
                converted_path=self._convert_to_xlsx_via_excel(path)
                if converted_path:
                    try: df=pd.read_excel(converted_path, engine='openpyxl')
                    except Exception as e: errors.append(f'Converted .xlsx read failed: {e}')
            if df is None:
                try:
                    if lower.endswith('.xlsx'): df=pd.read_excel(path, engine='openpyxl')
                except Exception as e: errors.append(f'XLSX read failed: {e}')
            if df is None and (lower.endswith('.htm') or lower.endswith('.html') or self._is_html_like(path)):
                try:
                    tables=pd.read_html(path)
                    if tables: df=tables[0]
                except Exception as e: errors.append(f'HTML parse failed: {e}')
            if df is None and lower.endswith('.csv'):
                try: df=pd.read_csv(path)
                except Exception as e: errors.append(f'CSV comma failed: {e}')
                if df is None:
                    try: df=pd.read_csv(path, sep='\t')
                    except Exception as e: errors.append(f'CSV tab failed: {e}')
            if df is None:
                try: df=pd.read_excel(path)
                except Exception as e: errors.append(f'Generic read_excel failed: {e}')
            if df is None:
                msg = "Failed to open the file. Tried Excel (with Excel-based conversion), HTML and CSV paths.\n" + "\n".join(errors[-6:])
                QMessageBox.warning(self,'Import Error', msg); return
            cols=[str(c) if c is not None else '' for c in df.columns]
            part_idx=self._find_part_col_index(cols)
            if part_idx<0:
                QMessageBox.warning(self,'Missing Column', "Couldn't find a 'Part' column (case-insensitive) in the selected file."); return
            # Optional OBS-only prefilter
            if hasattr(self,'_obs_only_filter') and self._obs_only_filter:
                df=self._apply_obs_only_filter(df, part_idx)
            colors_2d=None
            xlsx_source=converted_path if converted_path else (path if path.lower().endswith('.xlsx') else None)
            if xlsx_source: colors_raw=self._read_xlsx_background_colors(xlsx_source, target_ncols=len(cols))
            else: colors_raw=None
            df_kept, mask=self._apply_cleanup_rules(df, part_idx)
            if colors_raw is not None:
                kept_colors=[]; mask_list=mask.tolist()
                for ok,row_colors in zip(mask_list, colors_raw):
                    if ok: kept_colors.append(row_colors)
                colors_2d=kept_colors
            self._populate_table_with_extras(df_kept, list(df_kept.columns), part_idx, colors_2d=colors_2d)
            self._center_replacement_column()
            msg=f"Imported {len(df_kept)} cleaned row(s) from:\n{converted_path or path}"
            if converted_path: msg+="\n(The source file was auto-converted to .xlsx using Excel.)"
            QMessageBox.information(self,'Import Complete', msg)
        except Exception as e:
            QMessageBox.warning(self,'Import Error', str(e))

    # ----------------- Extra actions -----------------
    def import_where_used_of_obs_multi_level(self):
        try:
            self._obs_only_filter=True; self.import_where_used()
        finally:
            if hasattr(self,'_obs_only_filter'): delattr(self,'_obs_only_filter')
    def _apply_obs_only_filter(self, df, part_idx:int):
        obs_map=self._build_obs_map()
        if not obs_map: return df
        keys=set(obs_map.keys()); col=df.iloc[:,part_idx].astype(str).fillna('')
        mask=col.str.strip().str.upper().isin(keys)
        return df.loc[mask].reset_index(drop=True)
    def _selected_row_indices(self):
        rows=[]
        for r in range(self.table.rowCount()):
            w=self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk') and w._chk.isChecked(): rows.append(r)
        return rows
    def _headers(self)->list[str]:
        return [self.table.horizontalHeaderItem(i).text() if self.table.horizontalHeaderItem(i) else '' for i in range(self.table.columnCount())]
    def _map_original_to_table_col(self, orig_idx:int, part_idx:int)->int:
        rep_col=self._find_replacement_col()
        if part_idx>=0 and rep_col>=0 and (orig_idx>(rep_col-2)): return 1+orig_idx+1
        else: return 1+orig_idx
    def select_all_9024_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Select 9024 Parents', "Couldn't find a 'Parent' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=it.text() if it else ''
            self.table.cellWidget(r,0)._chk.setChecked(txt.strip().startswith('9024'))
    def select_all_options(self):
        headers=self._headers(); cidx=self._find_class_col_index(headers[1:])
        if cidx<0: QMessageBox.information(self,'Select Options/O Class', "Couldn't find a 'Class/Type/Category' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(cidx, part_idx)
            it=self.table.item(r, tbl_col); t=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(('OPTION' in t) or ('O CLASS' in t))
    def select_esw_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0:
            pidx=self._find_part_col_index(headers[1:])
            if pidx<0: QMessageBox.information(self,'Select ESW Parents', "Couldn't find 'Parent' or 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(txt.startswith('ESW'))
    def delete_selected_rows(self):
        rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Delete Selected','No rows are selected (checkbox).'); return
        for r in reversed(rows): self.table.removeRow(r)
    def move_selected_to_structure(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'append_rows'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.'); return
            headers=self._headers(); rows_idx=self._selected_row_indices()
            if not rows_idx: QMessageBox.information(self,'Move to Structure Sheet','No rows selected.'); return
            data=[]
            for r in rows_idx:
                row=[(self.table.item(r,c).text() if self.table.item(r,c) else '') for c in range(1,self.table.columnCount())]
                data.append(row)
            target.append_rows(headers[1:], data)
            QMessageBox.information(self,'Move to Structure Sheet', f'Moved {len(data)} row(s) to Structure Sheet.')
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))
    def _append_obs_part(self, part:str):
        try:
            if not self.obs_provider: return
            t=self.obs_provider.table; key=(part or '').strip().upper()
            if not key: return
            existing=set()
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it: existing.add((it.text() or '').strip().upper())
            if key in existing: return
            target=None
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it is None or not (it.text() or '').strip(): target=r; break
            if target is None:
                target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
            t.setItem(target,1,QTableWidgetItem(part))
            w=t.cellWidget(target,2)
            if isinstance(w,QComboBox): idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
        except Exception: pass
    def append_selected_to_obs(self):
        headers=self._headers(); pidx=self._find_part_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Append to OBS', "Couldn't find a 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:]); rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Append to OBS','No rows selected.'); return
        count=0
        for r in rows:
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r,tbl_col); part=it.text() if it else ''
            if part.strip(): self._append_obs_part(part); count+=1
        QMessageBox.information(self,'Append to OBS', f'Appended {count} part(s) to OBS List.')
    def export_where_used(self):
        try:
            dialog = QFileDialog(self, 'Export WhereUsed', 'WhereUsed.xlsx', 'Excel Files (*.xlsx)')
            dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
            dialog.setFileMode(QFileDialog.FileMode.AnyFile)
            dialog.setDefaultSuffix('xlsx')
            if not dialog.exec():
                return
            files = dialog.selectedFiles()
            path = files[0] if files else ''
            if not path: return
            if not path.lower().endswith('.xlsx'):
                path += '.xlsx'
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            wb=Workbook(); ws=wb.active; ws.title='WhereUsed'
            headers=self._headers(); ws.append(headers)
            for c in range(1,len(headers)+1):
                cell=ws.cell(row=1,column=c); cell.font=Font(bold=True)
                if headers[c-1].strip().lower()=='replacement': cell.alignment=Alignment(horizontal='center')
            for r in range(self.table.rowCount()):
                row_vals=[]
                for c in range(self.table.columnCount()):
                    if c==0:
                        w=self.table.cellWidget(r,0); row_vals.append('Yes' if (w and hasattr(w,'_chk') and w._chk.isChecked()) else 'No')
                    else:
                        it=self.table.item(r,c); row_vals.append(it.text() if it else '')
                ws.append(row_vals)
                wu_item = self.table.item(r, 1)
                row_fill_rgb = 'C7DEFA' if (wu_item and wu_item.text().strip() == '0') else None
                for c in range(self.table.columnCount()):
                    it=self.table.item(r,c)
                    if row_fill_rgb:
                        ws.cell(row=r+2,column=c+1).fill=PatternFill('solid', fgColor=row_fill_rgb)
                    elif it and it.background().style() != Qt.BrushStyle.NoBrush:
                        qcol=it.background().color()
                        if qcol.isValid() and qcol.alpha() > 0:
                            rgb=f"{qcol.red():02X}{qcol.green():02X}{qcol.blue():02X}"; ws.cell(row=r+2,column=c+1).fill=PatternFill('solid', fgColor=rgb)
                    if headers[c].strip().lower()=='replacement': ws.cell(row=r+2,column=c+1).alignment=Alignment(horizontal='center')

            # Add RawData sheet with formatted raw_databricks_records

            try:
                from Problem_Solution_Agent_PSS.where_used_query import DB_KEYS
            except ImportError:
                DB_KEYS = [
                    "wu_level", "part", "rev_ln", "plant", "description", "item_status", "base_qty", "ext_qty", "uom", "eco_number", "procurement_type", "effectivity_date", "user_item_type", "item_seq", "kit_code", "sparable_flag", "designator", "option_class", "pace_or_dash", "mlo_class", "input_part"
                ]
            raw_data = getattr(self, 'raw_databricks_records', None)
            if raw_data and isinstance(raw_data, list) and len(raw_data) > 0:
                ws_raw = wb.create_sheet('RawData')
                ws_raw.append(DB_KEYS)
                for rec in raw_data:
                    row = []
                    for k in DB_KEYS:
                        val = rec.get(k, '')
                        if k == 'part':
                            try:
                                level = int(rec.get('wu_level', 0))
                            except (ValueError, TypeError):
                                level = 0
                            val = ('      ' * level) + str(val)
                        row.append(val)
                    ws_raw.append(row)

            wb.save(path)
            wb.save(path)
            QMessageBox.information(self,'Export Complete', f'Exported {self.table.rowCount()} row(s) to:\n{path}\nRawData sheet included.')
        except Exception as e:
            QMessageBox.warning(self,'Export Error', str(e))
    def reset_where_used(self):
        try:
            self.table.clear(); self.table.setRowCount(0); self.table.setColumnCount(0)
            QMessageBox.information(self,'Reset','Where Used view has been reset.')
        except Exception as e:
            QMessageBox.warning(self,'Reset Error', str(e))

    def import_from_databricks(self, obs_parts: List[str], max_level: int, plant: str = "4070", retain_9024=False, retain_esw=False, retain_above_cfg=False):
        """Import multi-level Where Used from Databricks for the given OBS parts.

        Called by OBSPartsTab.launch_where_used_import() after input validation.
        Inputs are already validated (obs_parts non-empty, max_level 1–6).
        """
        # ── Confirm overwrite of existing data ────────────────────────────────
        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            reply = QMessageBox.question(
                self,
                'Where Used – Existing Data',
                'The Where Used tab already contains data.\n'
                'Delete it and import fresh data from Databricks?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        # ── Import the query module ────────────────────────────────────────────
        try:
            import sys as _sys
            import importlib as _il
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from where_used_query import fetch_where_used, DISPLAY_HEADERS  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(
                self, 'Module Not Found',
                f'where_used_query.py could not be imported:\n{exc}'
            )
            return

        # ── Query Databricks ───────────────────────────────────────────────────
        try:
            records = fetch_where_used(
                obs_parts,
                max_level,
                plant,
                retain_9024=retain_9024,
                retain_esw=retain_esw,
                retain_above_cfg=retain_above_cfg,
            )
            self.raw_databricks_records = records.copy() if isinstance(records, list) else None
        except Exception as exc:
            QMessageBox.warning(self, 'Databricks Query Error', str(exc))
            return

        if not records:
            QMessageBox.information(
                self, 'No Data',
                f'Databricks returned no results for {len(obs_parts)} OBS part(s) '
                f'at max WU level {max_level}.'
            )
            return

        # ── Build OBS map for Replacement column ──────────────────────────────
        obs_map = self._build_obs_map()

        # ── Column layout ──────────────────────────────────────────────────────
        # all_headers: [Select=0, WU Level=1, Part=2, Replacement=3, Rev/Ln=4, ...]
        all_headers = ['Select'] + DISPLAY_HEADERS
        _WU_COL     = 1
        _PART_COL   = 2
        _REPL_COL   = 3
        _DATA_START = 4   # Rev/Ln and onwards

        # Lambdas that extract the display value for each column from col 4 onwards.
        # Order must match DISPLAY_HEADERS[3:] (i.e. after WU Level / Part / Replacement).
        _DB_COL_FUNCS = [
            lambda r: r.get('rev_ln', ''),
            lambda r: r.get('plant', ''),
            lambda r: r.get('description', ''),
            lambda r: r.get('item_status', ''),
            lambda r: r.get('base_qty', ''),
            lambda r: r.get('ext_qty', ''),
            lambda r: r.get('uom', ''),
            lambda r: r.get('eco_number', ''),
            lambda r: r.get('procurement_type', ''),
            lambda r: r.get('effectivity_date', ''),
            lambda r: r.get('user_item_type', ''),
            lambda r: r.get('item_seq', ''),
            lambda r: r.get('kit_code', ''),
            lambda r: r.get('sparable_flag', ''),
            lambda r: r.get('designator', ''),
            lambda r: r.get('option_class', ''),
            lambda r: (r.get('pace_or_dash', '') if 'pace' in r.get('pace_or_dash', '').lower() else ''),
            lambda r: r.get('mlo_class', ''),
        ]

        # ── Populate table ─────────────────────────────────────────────────────
        self.table.clear()
        self.table.setColumnCount(len(all_headers))
        self.table.setHorizontalHeaderLabels(all_headers)
        self.table.setRowCount(len(records))

        for row_idx, record in enumerate(records):
            # Select checkbox
            chk  = QCheckBox()
            cont = QWidget()
            h    = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            cont._chk = chk
            self.table.setCellWidget(row_idx, 0, cont)

            # WU Level – numeric value from Databricks
            wu_val = record.get('wu_level', '')
            self.table.setItem(row_idx, _WU_COL, QTableWidgetItem(wu_val))

            # Part – display with leading spaces (2 spaces per WU level) for hierarchy
            raw_part = record.get('part', '')
            try:
                level_int = int(wu_val)
            except (ValueError, TypeError):
                level_int = 0
            indented_part = ('      ' * level_int) + raw_part
            self.table.setItem(row_idx, _PART_COL, QTableWidgetItem(indented_part))

            # Replacement – auto-fill from OBS map using the raw (unindented) part key
            replacement = obs_map.get(raw_part.strip().upper(), '')
            rep_item = QTableWidgetItem(replacement)
            rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, _REPL_COL, rep_item)

            # Remaining Databricks columns
            for c_off, fn in enumerate(_DB_COL_FUNCS):
                self.table.setItem(row_idx, _DATA_START + c_off,
                                   QTableWidgetItem(fn(record)))

            # Blue background for level-0 rows (the input OBS parts)
            if wu_val == '0':
                _blue = QColor('#C7DEFA')
                cont.setStyleSheet('background-color: #C7DEFA;')
                for _col in range(1, len(all_headers)):
                    _item = self.table.item(row_idx, _col)
                    if _item is not None:
                        _item.setBackground(_blue)

        # ── Resize columns ─────────────────────────────────────────────────────
        hdr = self.table.horizontalHeader()
        for i in range(len(all_headers)):
            if i == len(all_headers) - 1:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

        QMessageBox.information(
            self, 'Import Complete',
            f'Imported {len(records)} row(s) from Databricks.\n'
            f'OBS parts: {len(obs_parts)}  |  Max WU level: {max_level}  |  Plant: {plant}'
        )
        try:
            main = self.window()
            report_tab = getattr(main, 'report_tab', None)
            if report_tab is not None and hasattr(report_tab, 'refresh_report'):
                report_tab.refresh_report()
        except Exception:
            pass


class WhereUsedTabV2(WhereUsedTab):
    """UI-customized Where Used variant with revised action buttons and selection rules.
    Extends base import behavior while preserving data handling and table structure."""
    def __init__(self, obs_provider=None):
        super().__init__(obs_provider=obs_provider)
        # Hide legacy buttons/row
        try:
            # Hide old buttons if they exist
            for btn_name in [
                'btn_import_obs_multi','btn_sel_9024','btn_sel_options','btn_sel_esw',
                'btn_delete_sel','btn_move_to_struct','btn_append_obs','btn_export','btn_reset'
            ]:
                btn = getattr(self, btn_name, None)
                if btn is not None:
                    btn.hide()
        except Exception:
            pass
        
        # Build new two-row action panel
        from PyQt6.QtWidgets import QWidget, QGridLayout, QLabel
        panel = QWidget(self)
        grid = QGridLayout(panel)
        grid.setContentsMargins(0,0,0,0)
        grid.setHorizontalSpacing(6)
        grid.setVerticalSpacing(6)

        # Buttons (new)
        self.v2_btn_sel_opt_struct = QPushButton("Select Options/Class to Create Structure Sheet")
        self.v2_btn_move = QPushButton("Move to Structure Sheet")
        self.v2_btn_sel_parents_cfg = QPushButton("Select All Parents till Config")
        self.v2_btn_append = QPushButton("Append to OBS List")
        self.v2_btn_refresh = QPushButton("Refresh")
        self.v2_btn_delete = QPushButton("Delete selected")
        self.v2_btn_reset = QPushButton("Reset Tab")
        self.v2_btn_export = QPushButton("Export WhereUsed")
        # Place Import first, then Export (no custom colors)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # Import button is the last widget; insert Export right after Import
            title_row.insertWidget(title_row.count(), self.v2_btn_export)
        except Exception:
            pass
        # Move Export button to title row (right side)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # export will be inserted after import
        except Exception:
            pass

        # Colors (smooth, subtle gradients)
        def btn_style(bg1, bg2, border, text='#FFFFFF'):
            return f"""
            QPushButton {{
                color:{text}; padding:4px 8px; font-size:11px; border-radius:5px;
                border:1px solid {border};
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg2});
            }}
            QPushButton:hover {{
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg1});
            }}
            """
        
        common_blue = btn_style('#26C6DA','#14A7BE','#0F93AA')
        light_blue = btn_style('#8ED7EA','#6CC6DF','#54B6D3', text='#0F2D46')
        refresh_green = btn_style('#66BB6A','#4EA85A','#3F8F4A')
        reset_gray = btn_style('#90A4AE','#7C919B','#6A7E87')
        delete_red = btn_style('#EF5350','#D32F2F','#B71C1C')

        # Compact sizing – let buttons fit their text
        for btn in [
            self.v2_btn_sel_opt_struct, self.v2_btn_move, self.v2_btn_sel_parents_cfg, self.v2_btn_append,
            self.v2_btn_delete, self.v2_btn_export, self.v2_btn_refresh, self.v2_btn_reset,
        ]:
            btn.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)

        # Fit button widths to text so labels are fully visible.
        def _fit_btn(btn: QPushButton, pad: int = 24, min_h: int = 31):
            fm: QFontMetrics = btn.fontMetrics()
            btn.setFixedWidth(fm.horizontalAdvance(btn.text()) + pad)
            btn.setFixedHeight(min_h)

        self.v2_btn_sel_opt_struct.setStyleSheet(common_blue)
        self.v2_btn_sel_parents_cfg.setStyleSheet(light_blue)
        self.v2_btn_append.setStyleSheet(light_blue)
        self.v2_btn_move.setStyleSheet(common_blue)
        # Keep Export styling similar to Import button (default look).
        self.v2_btn_export.setStyleSheet('')
        self.v2_btn_delete.setStyleSheet(delete_red)
        self.v2_btn_refresh.setStyleSheet(refresh_green)
        self.v2_btn_reset.setStyleSheet(reset_gray)

        for btn in [
            self.v2_btn_sel_opt_struct,
            self.v2_btn_sel_parents_cfg,
            self.v2_btn_append,
            self.v2_btn_move,
            self.v2_btn_delete,
            self.v2_btn_refresh,
            self.v2_btn_reset,
            self.v2_btn_export,
        ]:
            _fit_btn(btn)

        # Keep title-row Import/Export controls aligned right and sized to text.
        for btn in self.findChildren(QPushButton):
            txt = (btn.text() or '').strip()
            if txt in {"Import 'Where Used' Parents", 'Export WhereUsed'}:
                _fit_btn(btn)

        # Button layout: single compact row, matching app style.
        row1 = [
            self.v2_btn_sel_opt_struct,
            self.v2_btn_move,
            self.v2_btn_sel_parents_cfg,
            self.v2_btn_append,
        ]
        for c, btn in enumerate(row1):
            grid.addWidget(btn, 0, c)
        grid.setColumnStretch(len(row1), 1)
        grid.addWidget(self.v2_btn_delete, 0, len(row1) + 1)
        grid.addWidget(self.v2_btn_reset, 0, len(row1) + 2)
        grid.addWidget(self.v2_btn_refresh, 0, len(row1) + 3)

        # Insert our panel right after the title row (index 1)
        try:
            outer: QVBoxLayout = self.layout()
            # self.layout() returns the QWidget layout of WhereUsedTab (outer QVBoxLayout)
            outer.insertWidget(1, panel)
        except Exception:
            # If insertion fails, just add at end
            self.layout().addWidget(panel)

        # Wire up actions
        self.v2_btn_sel_opt_struct.clicked.connect(self._v2_select_options_for_structure_sheet)
        self.v2_btn_move.clicked.connect(self._v2_move_selected_part_only)
        self.v2_btn_sel_parents_cfg.clicked.connect(self._v2_select_all_parents_till_config)
        self.v2_btn_append.clicked.connect(self._v2_append_selected_with_replacement)
        self.v2_btn_refresh.clicked.connect(self._v2_refresh_replacements)
        self.v2_btn_delete.clicked.connect(self.delete_selected_rows)
        self.v2_btn_reset.clicked.connect(self.reset_where_used)
        self.v2_btn_export.clicked.connect(self.export_where_used)

          # Keep refresh/reset in distinct colors; other buttons remain common blue.

        # Also hide the legacy "Import Where Used of OBS Parts (Multiple Level)" if present in title row
        try:
            if hasattr(self, 'btn_import_obs_multi'):
                self.btn_import_obs_multi.hide()
        except Exception:
            pass

    # ---------- Helpers to find current table columns ----------
    def _v2_find_table_col_index(self, header_exact: str) -> int:
        name = (header_exact or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and (h.text() or '').strip().lower() == name:
                return i
        return -1

    def _v2_find_table_col_contains(self, keyword: str) -> int:
        key = (keyword or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and key in (h.text() or '').strip().lower():
                return i
        return -1

    # ---------- Move / Append / Refresh ----------
    
    def _v2_toggle_rows(self, rows):
        checks = []
        for r in rows:
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk'):
                checks.append(w._chk)
        if not checks:
            return
        select = any(not c.isChecked() for c in checks)
        for c in checks:
            c.setChecked(select)

    def _v2_set_row_checked(self, row_idx: int, checked: bool):
        w = self.table.cellWidget(row_idx, 0)
        if w and hasattr(w, '_chk'):
            w._chk.setChecked(checked)

    def _v2_select_all_parents_till_config(self):
        """Select parent rows (WU Level != 0) per WU block up to Config.

        Kits/Assemblies are always selected. Options/Option Class selection is
        user-controlled via prompt + checkboxes.
        """
        config_prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357',
            '0390','0395','0397','0335','0391','0431','0435','0437',
            '0440','0445','0455','0450','0441','0447','0457',
            '0460','0465','0461','0467','0410','0415','0417',
            '0411','0412','0413','0414','0360','0365','0361','0367'
        }
        listing_prefixes = {'0243', '0288', '0289', '0290'}

        part_col = self._v2_find_table_col_index('Part')
        wu_col = self._v2_find_table_col_contains('wu level')
        option_class_col = self._v2_find_table_col_contains('option class')
        mlo_class_col = self._v2_find_table_col_contains('mlo class')
        if part_col < 0 or wu_col < 0:
            QMessageBox.information(
                self,
                'Select All Parents till Config',
                "Couldn't find required 'Part' and 'WU Level' columns.",
            )
            return

        def _meta(row_idx: int):
            p_raw = (self.table.item(row_idx, part_col).text() if self.table.item(row_idx, part_col) else '') or ''
            p_trim = p_raw.lstrip(' \t').strip()
            w_raw = (self.table.item(row_idx, wu_col).text() if self.table.item(row_idx, wu_col) else '') or ''
            try:
                lvl = int(float(w_raw.strip()))
            except Exception:
                lvl = -1
            oc_txt = ''
            mlo_txt = ''
            if option_class_col >= 0:
                oc_item = self.table.item(row_idx, option_class_col)
                oc_txt = (oc_item.text() if oc_item else '') or ''
            if mlo_class_col >= 0:
                mlo_item = self.table.item(row_idx, mlo_class_col)
                mlo_txt = (mlo_item.text() if mlo_item else '') or ''
            return p_trim, lvl, oc_txt, mlo_txt

        def _option_kind(option_class_text: str, mlo_class_text: str):
            oc = (option_class_text or '').strip().upper()
            mlo = (mlo_class_text or '').strip().upper()

            # MLO Class like "Class 1/2/3" should be treated as Options only.
            if re.search(r'\bCLASS\s*[123]\b', mlo):
                return 'options'

            if ('OPTION CLASS' in oc) or ('O CLASS' in oc):
                return 'option_class'
            if 'OPTION' in oc:
                return 'options'
            return 'unknown'

        rows_total = self.table.rowCount()
        kits_rows = set()
        options_rows = set()
        option_class_rows = set()
        optional_unknown_rows = set()

        r = 0
        while r < rows_total:
            _, lvl_r, _, _ = _meta(r)
            if lvl_r != 0:
                r += 1
                continue

            block_start = r
            block_end = rows_total
            for i in range(r + 1, rows_total):
                _, lvl_i, _, _ = _meta(i)
                if lvl_i == 0:
                    block_end = i
                    break

            seen_config = False
            for i in range(block_start + 1, block_end):
                part_i, lvl_i, oc_i, mlo_i = _meta(i)
                if lvl_i <= 0:
                    continue

                # Listing commodity rows are always excluded from selection.
                if part_i[:4] in listing_prefixes:
                    continue

                is_config_or_option = (part_i[:4] in config_prefixes) or (_option_kind(oc_i, mlo_i) != 'unknown')
                if is_config_or_option:
                    seen_config = True

                if not seen_config:
                    kits_rows.add(i)
                    continue

                kind = _option_kind(oc_i, mlo_i)
                if kind == 'options':
                    options_rows.add(i)
                elif kind == 'option_class':
                    option_class_rows.add(i)
                else:
                    optional_unknown_rows.add(i)

            r = block_end

        if not (kits_rows or options_rows or option_class_rows or optional_unknown_rows):
            QMessageBox.information(
                self,
                'Select All Parents till Config',
                'No parent rows (WU Level != 0) were found to select.',
            )
            return

        ans = QMessageBox.question(
            self,
            'Select All Parents till Config',
            'Do you want to obsolete the Options and Option Class along with Kits/Assemblies?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        include_options = False
        include_option_class = False

        if ans == QMessageBox.StandardButton.Yes:
            dlg = QDialog(self)
            dlg.setWindowTitle('Include in Obsoletion')
            lay = QVBoxLayout(dlg)

            msg = QLabel('Select the items to include in obsoletion.')
            msg.setWordWrap(True)
            lay.addWidget(msg)

            chk_opt = QCheckBox('Options')
            chk_opt_cls = QCheckBox('Option Class')
            chk_opt.setEnabled(True)
            chk_opt_cls.setEnabled(True)
            lay.addWidget(chk_opt)
            lay.addWidget(chk_opt_cls)

            btn_row = QHBoxLayout()
            btn_row.addStretch(1)
            btn_ok = QPushButton('Apply')
            btn_cancel = QPushButton('Cancel')
            btn_row.addWidget(btn_ok)
            btn_row.addWidget(btn_cancel)
            lay.addLayout(btn_row)

            btn_ok.clicked.connect(dlg.accept)
            btn_cancel.clicked.connect(dlg.reject)

            if dlg.exec() != QDialog.DialogCode.Accepted:
                return

            include_options = chk_opt.isChecked()
            include_option_class = chk_opt_cls.isChecked()

        selected_rows = set(kits_rows)
        if include_options:
            selected_rows.update(options_rows)
        if include_option_class:
            selected_rows.update(option_class_rows)
        if include_options and include_option_class:
            selected_rows.update(optional_unknown_rows)

        for rr in range(rows_total):
            self._v2_set_row_checked(rr, False)
        for rr in sorted(selected_rows):
            self._v2_set_row_checked(rr, True)

        QMessageBox.information(
            self,
            'Selection Applied',
            f'Selected {len(selected_rows)} row(s). Kits/Assemblies are always included.',
        )

    def _v2_select_options_for_structure_sheet(self):
        """Select Option/Class rows and their immediate direct children in each WU block.

        Example: if an Option is at WU level 3, select that row plus only level-4
        children under that row. Indentation in Part text is used as a hierarchy
        guard to avoid selecting unrelated rows.

        Some datasets list the linked item one level in the opposite direction;
        if no direct child is found, we fallback to the nearest one-level linked row.
        """
        part_col = self._v2_find_table_col_index('Part')
        wu_col = self._v2_find_table_col_contains('wu level')
        rep_col = self._find_replacement_col()
        if part_col < 0 or wu_col < 0:
            QMessageBox.information(
                self,
                'Select Options/Class to Create Structure Sheet',
                "Couldn't find required 'Part' and 'WU Level' columns.",
            )
            return

        option_prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
            '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
            '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
        }

        def _row_meta(row_idx: int):
            part_item = self.table.item(row_idx, part_col)
            part_raw = (part_item.text() if part_item else '') or ''
            part_trim = part_raw.lstrip(' \t')
            indent = len(part_raw) - len(part_trim)

            wu_item = self.table.item(row_idx, wu_col)
            wu_raw = (wu_item.text() if wu_item else '') or ''
            try:
                wu_level = int(float(wu_raw.strip()))
            except Exception:
                wu_level = -1

            return part_trim, indent, wu_level

        rows = self.table.rowCount()
        target_rows = set()
        blocked_rows = set()

        r = 0
        while r < rows:
            _, _, curr_wu = _row_meta(r)
            if curr_wu != 0:
                r += 1
                continue

            block_start = r
            block_end = rows
            for i in range(r + 1, rows):
                _, _, wu_i = _row_meta(i)
                if wu_i == 0:
                    block_end = i
                    break

            for i in range(block_start + 1, block_end):
                if i in blocked_rows:
                    continue

                part_trim_i, indent_i, wu_i = _row_meta(i)
                if not part_trim_i:
                    continue
                if part_trim_i[:4] not in option_prefixes:
                    continue

                rep_txt = ''
                if rep_col >= 0:
                    rep_item = self.table.item(i, rep_col)
                    rep_txt = (rep_item.text() if rep_item else '').strip()

                # If an option/class row already has a replacement and is below WU level 0,
                # exclude that option and everything reporting to it.
                if wu_i != 0 and rep_txt:
                    blocked_rows.add(i)
                    for j in range(i + 1, block_end):
                        _, indent_j, wu_j = _row_meta(j)
                        if wu_j >= 0 and wu_j <= wu_i and indent_j <= indent_i:
                            break
                        blocked_rows.add(j)
                    continue

                # Always include the Option/Class row itself.
                target_rows.add(i)

                # If Option/Class is at level 1, also include level-0 parent row
                # for this WU block (the block starts at level 0).
                if wu_i == 1:
                    _, indent_0, wu_0 = _row_meta(block_start)
                    if wu_0 == 0 and indent_0 < indent_i:
                        target_rows.add(block_start)

                # Include only immediate direct children one WU level below,
                # inside this option/class subtree.
                child_level = wu_i + 1
                matched_direct_child = False
                for j in range(i + 1, block_end):
                    _, indent_j, wu_j = _row_meta(j)

                    # End of this option/class subtree.
                    if wu_j >= 0 and wu_j <= wu_i and indent_j <= indent_i:
                        break

                    # Primary rule: one level deeper and deeper indentation.
                    if wu_j == child_level and indent_j > indent_i:
                        target_rows.add(j)
                        matched_direct_child = True

                # Fallback: if no direct child exists, select the nearest linked
                # one-level row in the opposite direction (common in some WU layouts).
                if not matched_direct_child:
                    parent_level = wu_i - 1
                    for j in range(i - 1, block_start - 1, -1):
                        _, indent_j, wu_j = _row_meta(j)
                        if wu_j == parent_level and indent_j < indent_i:
                            target_rows.add(j)
                            break

            r = block_end

        self._v2_toggle_rows(sorted(target_rows))
    def _v2_move_selected_part_only(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'table'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.')
                return
            rows_idx = self._selected_row_indices()
            if not rows_idx:
                QMessageBox.information(self,'Move to Structure Sheet','No rows selected.')
                return

            src_headers = self._headers()
            if len(src_headers) <= 1:
                QMessageBox.information(self,'Move to Structure Sheet','Where Used table has no usable columns.')
                return

            # Drop Select column and rename WU Level -> BOM Level.
            out_headers = list(src_headers[1:])
            wu_out_idx = -1
            for i, h in enumerate(out_headers):
                if (h or '').strip().lower() == 'wu level':
                    out_headers[i] = 'BOM Level'
                    wu_out_idx = i
                    break
            if wu_out_idx < 0:
                for i, h in enumerate(out_headers):
                    if 'wu level' in (h or '').strip().lower():
                        out_headers[i] = 'BOM Level'
                        wu_out_idx = i
                        break

            def _norm_header(h: str) -> str:
                return ' '.join((h or '').strip().lower().split())

            # Remove columns requested by user from Structure Sheet output.
            drop_headers = {
                'rev/ln',
                'plant',
                'eco number',
                'effectivity date',
                'user item type',
                'bom source type (debug)',
            }

            # New leading columns.
            # New leading columns.  Column 1 = 'Action' (user-editable change type).
            leading_headers = [
                'Action',
                'Part Description',
                'Seq#',
                'Qty',
                'Kit Code',
                'Ref Designator(RD)',
            ]

            part_out_idx = -1
            for i, h in enumerate(out_headers):
                if _norm_header(h) == 'part':
                    part_out_idx = i
                    break

            change_type_src_idx = -1
            for i, h in enumerate(out_headers):
                if _norm_header(h) == 'change type':
                    change_type_src_idx = i
                    break

            pcol = self._v2_find_table_col_index('Part')
            wu_col = self._v2_find_table_col_contains('wu level')
            if pcol < 0 or wu_col < 0:
                QMessageBox.information(self,'Move to Structure Sheet',"Couldn't find required 'Part' and 'WU Level' columns.")
                return

            option_prefixes = {
                '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
                '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
                '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
            }
            selected_set = set(rows_idx)

            def _src_row_values(row_idx: int):
                return [
                    (self.table.item(row_idx, c).text() if self.table.item(row_idx, c) else '')
                    for c in range(1, self.table.columnCount())
                ]

            def _part_clean(row_idx: int) -> str:
                it = self.table.item(row_idx, pcol)
                return ((it.text() if it else '') or '').lstrip(' \t').strip()

            def _meta(row_idx: int):
                p_raw = (self.table.item(row_idx, pcol).text() if self.table.item(row_idx, pcol) else '') or ''
                p_trim = p_raw.lstrip(' \t')
                indent = len(p_raw) - len(p_trim)
                w_raw = (self.table.item(row_idx, wu_col).text() if self.table.item(row_idx, wu_col) else '') or ''
                try:
                    lvl = int(float(w_raw.strip()))
                except Exception:
                    lvl = -1
                return p_trim.strip(), indent, lvl

            # Precompute WU level-0 blocks.
            rows_total = self.table.rowCount()
            block_bounds = []
            start = 0
            for r in range(rows_total):
                _, _, lvl = _meta(r)
                if lvl == 0:
                    if block_bounds:
                        prev_s, _ = block_bounds[-1]
                        block_bounds[-1] = (prev_s, r)
                    block_bounds.append((r, rows_total))
            if not block_bounds:
                block_bounds = [(0, rows_total)]

            # Build option groups, de-duplicated by option part.
            option_order = []
            option_groups = {}  # key -> {'option_row': list[str], 'children': [list[str]], 'seen_children': set[str]}

            def _find_block_for_row(row_idx: int):
                for bs, be in block_bounds:
                    if bs <= row_idx < be:
                        return bs, be
                return 0, rows_total

            # Candidate options are selected option-prefix rows.
            for i in rows_idx:
                part_i, indent_i, wu_i = _meta(i)
                if not part_i or part_i[:4] not in option_prefixes:
                    continue

                opt_key = part_i.upper()
                if opt_key not in option_groups:
                    option_order.append(opt_key)
                    option_groups[opt_key] = {
                        'option_row': _src_row_values(i),
                        'children': [],
                        'seen_children': set(),
                    }

                bs, be = _find_block_for_row(i)

                # Include selected linked rows inside the option subtree.
                for j in range(i + 1, be):
                    _, indent_j, wu_j = _meta(j)
                    if wu_j >= 0 and wu_j <= wu_i and indent_j <= indent_i:
                        break
                    if j not in selected_set:
                        continue
                    child_part = _part_clean(j).upper()
                    if not child_part or child_part == opt_key:
                        continue
                    if child_part in option_groups[opt_key]['seen_children']:
                        continue
                    option_groups[opt_key]['children'].append(_src_row_values(j))
                    option_groups[opt_key]['seen_children'].add(child_part)

                # Include nearest selected one-level parent as the first linked row.
                parent_candidate = None
                for j in range(i - 1, bs - 1, -1):
                    _, indent_j, wu_j = _meta(j)
                    if wu_j == (wu_i - 1) and indent_j < indent_i:
                        if j in selected_set:
                            parent_candidate = j
                        break

                if parent_candidate is not None:
                    parent_part = _part_clean(parent_candidate).upper()
                    if parent_part and parent_part not in option_groups[opt_key]['seen_children']:
                        option_groups[opt_key]['children'].insert(0, _src_row_values(parent_candidate))
                        option_groups[opt_key]['seen_children'].add(parent_part)

            if not option_order:
                QMessageBox.information(
                    self,
                    'Move to Structure Sheet',
                    'No selected Option/Class rows were found to build Structure Sheet output.',
                )
                return

            # Decide move mode when overlaps exist in current Structure Sheet.
            move_mode = 'replace'
            existing_full_duplicates = []
            overlap_options = []
            overlap_part_count = 0
            if target.table.columnCount() > 0 and target.table.rowCount() > 0:
                t_existing = target.table
                bom_col_ex = -1
                part_col_ex = -1
                for c in range(t_existing.columnCount()):
                    h = t_existing.horizontalHeaderItem(c)
                    if not h:
                        continue
                    nh = _norm_header(h.text())
                    if nh == 'bom level':
                        bom_col_ex = c
                    elif nh == 'part':
                        part_col_ex = c

                if bom_col_ex >= 0 and part_col_ex >= 0:
                    existing_bom_map = {}  # parent_part -> set(child_parts)
                    existing_all_parts = set()
                    current_parent = None
                    for rr in range(t_existing.rowCount()):
                        bom_item = t_existing.item(rr, bom_col_ex)
                        part_item = t_existing.item(rr, part_col_ex)
                        bom_txt = (bom_item.text() if bom_item else '').strip()
                        part_txt = ((part_item.text() if part_item else '') or '').lstrip(' \t').strip().upper()
                        if not part_txt:
                            continue
                        existing_all_parts.add(part_txt)
                        if bom_txt == '0':
                            current_parent = part_txt
                            existing_bom_map.setdefault(current_parent, set())
                        elif bom_txt == '1' and current_parent:
                            existing_bom_map.setdefault(current_parent, set()).add(part_txt)

                    for opt_key in option_order:
                        grp = option_groups.get(opt_key, {})
                        children = grp.get('children', [])
                        child_set = set()
                        for ch in children:
                            if 0 <= part_out_idx < len(ch):
                                pch = ((ch[part_out_idx] or '') if isinstance(ch[part_out_idx], str)
                                       else str(ch[part_out_idx] or '')).lstrip(' \t').strip().upper()
                                if pch and pch != opt_key:
                                    child_set.add(pch)

                        if opt_key in existing_bom_map:
                            overlap_options.append(opt_key)
                            overlap_part_count += len(child_set.intersection(existing_bom_map.get(opt_key, set())))

                        if opt_key in existing_bom_map and child_set and child_set.issubset(existing_bom_map.get(opt_key, set())):
                            existing_full_duplicates.append(opt_key)

                    if overlap_options or overlap_part_count > 0:
                        msg = QMessageBox(self)
                        msg.setIcon(QMessageBox.Icon.Question)
                        msg.setWindowTitle('Move to Structure Sheet')
                        msg.setText(
                            'Some selected Option(s)/Part(s) already exist in Structure Sheet.\n\n'
                            'Choose how to proceed:'
                        )
                        msg.setInformativeText(
                            'Replace Existing Data: overwrite current Structure Sheet with selected move data.\n'
                            'Merge Data: keep existing rows and update Replacement for matching BOM items under each Option.'
                        )
                        btn_replace = msg.addButton('Replace Existing Data', QMessageBox.ButtonRole.AcceptRole)
                        btn_merge = msg.addButton('Merge Data', QMessageBox.ButtonRole.ActionRole)
                        btn_cancel = msg.addButton(QMessageBox.StandardButton.Cancel)
                        msg.exec()
                        clicked = msg.clickedButton()
                        if clicked == btn_cancel:
                            return
                        if clicked == btn_merge:
                            move_mode = 'merge'

            if not option_order:
                if existing_full_duplicates:
                    preview = ', '.join(existing_full_duplicates[:10])
                    more = '' if len(existing_full_duplicates) <= 10 else f' and {len(existing_full_duplicates) - 10} more'
                    QMessageBox.information(
                        self,
                        'Move to Structure Sheet',
                        f'Already exists in Structure Sheet: {preview}{more}.'
                    )
                else:
                    QMessageBox.information(
                        self,
                        'Move to Structure Sheet',
                        'No selected Option/Class rows were found to build Structure Sheet output.',
                    )
                return

            # Compose output rows by reversing WU hierarchy into BOM hierarchy:
            # highest WU level becomes BOM 0, then descend (0 -> deepest BOM level).
            out_rows = []

            def _row_wu_level(vals: list[str]) -> int:
                if wu_out_idx < 0 or wu_out_idx >= len(vals):
                    return -1
                try:
                    return int(float((vals[wu_out_idx] or '').strip()))
                except Exception:
                    return -1

            for opt_key in option_order:
                grp = option_groups[opt_key]
                group_rows = [list(grp['option_row'])] + [list(ch) for ch in grp['children']]

                # Sort by WU desc so parent chain appears above child chain in Structure Sheet.
                indexed = list(enumerate(group_rows))
                indexed.sort(key=lambda ir: (-_row_wu_level(ir[1]), ir[0]))
                ordered_rows = [ir[1] for ir in indexed]

                valid_wu = [_row_wu_level(r) for r in ordered_rows if _row_wu_level(r) >= 0]
                max_wu = max(valid_wu) if valid_wu else 0

                for seq, src_row in enumerate(ordered_rows):
                    mapped_row = list(src_row)
                    wu_here = _row_wu_level(src_row)

                    # Reverse WU->BOM (WU max => BOM 0, WU 0 => BOM max).
                    bom_level = (max_wu - wu_here) if wu_here >= 0 else seq
                    if wu_out_idx >= 0 and wu_out_idx < len(mapped_row):
                        mapped_row[wu_out_idx] = str(max(0, bom_level))

                    out_rows.append(mapped_row)

            # Build final headers (prepend Select and new columns, remove dropped columns,
            # and avoid duplicate Change Type from source section).
            keep_indices = []
            for i, h in enumerate(out_headers):
                nh = _norm_header(h)
                if nh in drop_headers:
                    continue
                if nh == 'change type':
                    continue
                keep_indices.append(i)
            final_headers = ['Select'] + leading_headers + [out_headers[i] for i in keep_indices]

            # Build final rows with Select column and new leading columns.
            final_rows = []
            for row in out_rows:
                change_type_val = ''
                if 0 <= change_type_src_idx < len(row):
                    change_type_val = row[change_type_src_idx]

                lead = [change_type_val, '', '', '', '', '']
                tail = [row[i] if i < len(row) else '' for i in keep_indices]
                final_rows.append([''] + lead + tail)
            keep_indices = []
            for i, h in enumerate(out_headers):
                nh = _norm_header(h)
                if nh in drop_headers:
                    continue
                if nh == 'change type':
                    continue
                keep_indices.append(i)
            final_headers = ['Select'] + leading_headers + [out_headers[i] for i in keep_indices]

            # Build final rows with Select column and new leading columns.
            final_rows = []
            for row in out_rows:
                change_type_val = ''
                if 0 <= change_type_src_idx < len(row):
                    change_type_val = row[change_type_src_idx]

                lead = [change_type_val, '', '', '', '', '']
                tail = [row[i] if i < len(row) else '' for i in keep_indices]
                final_rows.append([''] + lead + tail)

            # Insert blank 'Ref Designator' column between PACE and MLO Class.
            pace_pos = -1
            mlo_pos = -1
            for ci, hh in enumerate(final_headers):
                nh = _norm_header(hh)
                if 'pace' in nh:
                    pace_pos = ci
                if 'mlo' in nh and 'class' in nh:
                    mlo_pos = ci
            insert_ref_at = -1
            if pace_pos >= 0 and mlo_pos == pace_pos + 1:
                insert_ref_at = mlo_pos
            elif pace_pos >= 0:
                insert_ref_at = pace_pos + 1
            if insert_ref_at >= 0:
                final_headers.insert(insert_ref_at, 'Ref Designator(RD)')
                for fr in final_rows:
                    fr.insert(insert_ref_at, '')

            # Merge mode: keep existing rows and update/append incoming rows.
            merge_added_count = 0
            merge_replacement_updates = 0
            if move_mode == 'merge' and target.table.columnCount() > 0 and target.table.rowCount() > 0:
                t_existing = target.table
                existing_headers = []
                for c in range(t_existing.columnCount()):
                    h = t_existing.horizontalHeaderItem(c)
                    existing_headers.append(h.text() if h else '')

                merged_headers = list(existing_headers)
                existing_header_keys = {_norm_header(h): i for i, h in enumerate(existing_headers)}
                for h in final_headers:
                    if _norm_header(h) not in existing_header_keys:
                        merged_headers.append(h)
                        existing_header_keys[_norm_header(h)] = len(merged_headers) - 1

                bom_col_idx = next((i for i, h in enumerate(merged_headers) if _norm_header(h) == 'bom level'), -1)
                part_col_idx = next((i for i, h in enumerate(merged_headers) if _norm_header(h) == 'part'), -1)
                repl_col_idx = next((i for i, h in enumerate(merged_headers) if _norm_header(h) == 'replacement'), -1)

                if bom_col_idx >= 0 and part_col_idx >= 0:
                    def _align_row(row_vals, source_headers, target_headers):
                        aligned = ['' for _ in target_headers]
                        src_map = {_norm_header(h): i for i, h in enumerate(source_headers)}
                        for target_i, target_h in enumerate(target_headers):
                            source_i = src_map.get(_norm_header(target_h), -1)
                            if 0 <= source_i < len(row_vals):
                                aligned[target_i] = row_vals[source_i]
                        return aligned

                    existing_rows = []
                    for rr in range(t_existing.rowCount()):
                        row_vals = []
                        for cc in range(t_existing.columnCount()):
                            if cc == 0:
                                row_vals.append('')
                            elif cc == 1:
                                w = t_existing.cellWidget(rr, cc)
                                if w and isinstance(w, QComboBox):
                                    row_vals.append((w.currentText() or '').strip())
                                else:
                                    it = t_existing.item(rr, cc)
                                    row_vals.append((it.text() if it else '').strip())
                            else:
                                it = t_existing.item(rr, cc)
                                row_vals.append((it.text() if it else ''))
                        existing_rows.append(_align_row(row_vals, existing_headers, merged_headers))

                    incoming_rows = [_align_row(row, final_headers, merged_headers) for row in final_rows]

                    def _row_key_stream(rows_list):
                        keys = []
                        current_opt = ''
                        for row in rows_list:
                            bom_txt = (row[bom_col_idx] if 0 <= bom_col_idx < len(row) else '').strip()
                            part_txt = ((row[part_col_idx] if 0 <= part_col_idx < len(row) else '') or '').lstrip(' \t').strip().upper()
                            if bom_txt == '0' and part_txt:
                                current_opt = part_txt
                            keys.append((current_opt, bom_txt, part_txt))
                        return keys

                    existing_keys = _row_key_stream(existing_rows)
                    existing_index_by_key = {k: i for i, k in enumerate(existing_keys) if k[2]}
                    incoming_keys = _row_key_stream(incoming_rows)
                    appended_count = 0
                    replacement_updates = 0

                    for inc_i, inc_row in enumerate(incoming_rows):
                        ikey = incoming_keys[inc_i]
                        if not ikey[2]:
                            continue

                        if ikey in existing_index_by_key:
                            ex_i = existing_index_by_key[ikey]
                            if repl_col_idx >= 0 and repl_col_idx < len(inc_row):
                                incoming_rep = (inc_row[repl_col_idx] or '').strip()
                                existing_rep = (existing_rows[ex_i][repl_col_idx] or '').strip()
                                if incoming_rep and existing_rep != incoming_rep:
                                    existing_rows[ex_i][repl_col_idx] = incoming_rep
                                    replacement_updates += 1
                        else:
                            existing_rows.append(list(inc_row))
                            existing_index_by_key[ikey] = len(existing_rows) - 1
                            appended_count += 1

                    final_headers = merged_headers
                    final_rows = existing_rows
                    merge_added_count = appended_count
                    merge_replacement_updates = replacement_updates
                else:
                    QMessageBox.warning(
                        self,
                        'Move to Structure Sheet',
                        'Merge requested, but current Structure Sheet is missing required BOM Level/Part columns. Replacing data instead.'
                    )
                    move_mode = 'replace'

            # Item Seq correction:
            # - BOM Level 0 rows must be blank.
            # - BOM Level >= 1 rows should use Implemented BOM item_seq by parent->child edge.
            try:
                bom_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'bom level'), -1)
                part_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'part'), -1)
                item_seq_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'item seq'), -1)
                plant_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'plant'), -1)

                if bom_idx >= 0 and part_idx >= 0 and item_seq_idx >= 0 and final_rows:
                    parent_roots = []
                    seen_roots = set()
                    max_bom = 1
                    plant_val = ''

                    for row in final_rows:
                        bom_txt = (row[bom_idx] if bom_idx < len(row) else '') or ''
                        part_txt = ((row[part_idx] if part_idx < len(row) else '') or '').lstrip(' \t').strip().upper()

                        if not plant_val and plant_idx >= 0 and plant_idx < len(row):
                            plant_val = str(row[plant_idx] or '').strip()

                        try:
                            bom_int = int(float(str(bom_txt).strip()))
                        except Exception:
                            bom_int = -1

                        if bom_int > max_bom:
                            max_bom = bom_int

                        if bom_int == 0 and part_txt and part_txt not in seen_roots:
                            parent_roots.append(part_txt)
                            seen_roots.add(part_txt)

                    if parent_roots:
                        try:
                            from implemented_bom_query import fetch_implemented_bom  # type: ignore[import]

                            if not plant_val:
                                plant_val = (self.plant_combo.currentText() or '').strip() if hasattr(self, 'plant_combo') else ''
                            if not plant_val:
                                plant_val = '4070'

                            impl_rows = fetch_implemented_bom(
                                parent_roots,
                                max_level=max(1, min(18, max_bom)),
                                plant=plant_val,
                                include_level0=False,
                            )
                        except Exception:
                            impl_rows = []

                        seq_by_edge = {}
                        for rec in impl_rows:
                            parent_part = str(rec.get('parent_part', '') or '').strip().upper()
                            child_part = str(rec.get('part', '') or '').strip().upper()
                            seq_val = str(rec.get('item_seq', '') or '').strip()
                            if not parent_part or not child_part or not seq_val:
                                continue
                            key = (parent_part, child_part)
                            if key not in seq_by_edge:
                                seq_by_edge[key] = seq_val

                        current_part_at_level = {}
                        for row in final_rows:
                            bom_txt = (row[bom_idx] if bom_idx < len(row) else '') or ''
                            part_txt = ((row[part_idx] if part_idx < len(row) else '') or '').lstrip(' \t').strip().upper()

                            try:
                                bom_int = int(float(str(bom_txt).strip()))
                            except Exception:
                                bom_int = -1

                            if bom_int == 0:
                                if item_seq_idx < len(row):
                                    row[item_seq_idx] = ''
                            elif bom_int > 0:
                                parent_part = ''
                                for lv in range(bom_int - 1, -1, -1):
                                    cand = current_part_at_level.get(lv, '')
                                    if cand:
                                        parent_part = cand
                                        break

                                mapped_seq = seq_by_edge.get((parent_part, part_txt), '') if parent_part and part_txt else ''
                                if mapped_seq and item_seq_idx < len(row):
                                    row[item_seq_idx] = mapped_seq

                            if part_txt and bom_int >= 0:
                                current_part_at_level[bom_int] = part_txt
                                prune_levels = [lv for lv in current_part_at_level.keys() if lv > bom_int]
                                for lv in prune_levels:
                                    current_part_at_level.pop(lv, None)
            except Exception:
                pass

            # Remove duplicate parent blocks in Structure output.
            # Block rule: starts at BOM Level 0 and ends right before next BOM Level 0.
            # Merge rule: same parent part => keep first parent row, keep unique child parts only.
            try:
                bom_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'bom level'), -1)
                part_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'part'), -1)

                if bom_idx >= 0 and part_idx >= 0 and final_rows:
                    blocks = []
                    orphan_rows = []
                    current_block = None

                    for row in final_rows:
                        bom_txt = (row[bom_idx] if bom_idx < len(row) else '') or ''
                        part_txt = ((row[part_idx] if part_idx < len(row) else '') or '').lstrip(' \t').strip()
                        try:
                            bom_int = int(float(str(bom_txt).strip()))
                        except Exception:
                            bom_int = -1

                        if bom_int == 0:
                            if current_block is not None:
                                blocks.append(current_block)
                            current_block = {
                                'parent_row': list(row),
                                'parent_key': part_txt.upper(),
                                'children': [],
                            }
                        else:
                            if current_block is None:
                                orphan_rows.append(list(row))
                            else:
                                current_block['children'].append(list(row))

                    if current_block is not None:
                        blocks.append(current_block)

                    if blocks:
                        merged_order = []
                        merged = {}

                        for blk in blocks:
                            pkey = blk.get('parent_key', '') or ''
                            if not pkey:
                                # Parent row exists but part is blank; preserve as standalone.
                                orphan_rows.append(list(blk.get('parent_row', [])))
                                orphan_rows.extend([list(r) for r in blk.get('children', [])])
                                continue

                            if pkey not in merged:
                                merged_order.append(pkey)
                                merged[pkey] = {
                                    'parent_row': list(blk.get('parent_row', [])),
                                    'children': [],
                                    'seen_child_parts': set(),
                                }

                            group = merged[pkey]
                            for crow in blk.get('children', []):
                                cpart = ((crow[part_idx] if part_idx < len(crow) else '') or '').lstrip(' \t').strip().upper()
                                if not cpart or cpart in group['seen_child_parts']:
                                    continue
                                group['children'].append(list(crow))
                                group['seen_child_parts'].add(cpart)

                        dedup_rows = []
                        for pkey in merged_order:
                            group = merged[pkey]
                            dedup_rows.append(group['parent_row'])
                            dedup_rows.extend(group['children'])

                        if orphan_rows:
                            dedup_rows.extend(orphan_rows)

                        final_rows = dedup_rows
            except Exception:
                pass

            # Replace Structure Sheet table with rearranged, de-duplicated grouped output.
            t = target.table
            _bulk_guard = hasattr(target, '_struct_item_change_guard')
            if _bulk_guard:
                target._struct_item_change_guard = True
            try:
                t.clear()
                t.setRowCount(0)
                t.setColumnCount(len(final_headers))
                hdr = RotatedColumnsHeader(
                    Qt.Orientation.Horizontal,
                    rotated_columns=range(2, 7),
                    parent=t,
                    group_label='Change Type',
                    group_columns=list(range(2, 7)),
                )
                hdr.setStretchLastSection(True)
                hdr.setSectionsClickable(True)
                t.setHorizontalHeader(hdr)
                t.setHorizontalHeaderLabels(final_headers)
                t.setRowCount(len(final_rows))
                if hasattr(target, '_apply_inserted_header_colors'):
                    target._apply_inserted_header_colors()

                t.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
                t.horizontalHeader().setMinimumHeight(117)  # reduced ~20%
                for vc in range(2, 7):
                    t.horizontalHeader().setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                    t.setColumnWidth(vc, 40)

                bom_col = -1
                part_col = -1
                replacement_col = -1
                for c, h in enumerate(final_headers):
                    if _norm_header(h) == 'bom level':
                        bom_col = c
                    if _norm_header(h) == 'part':
                        part_col = c
                    if _norm_header(h) == 'replacement':
                        replacement_col = c

                vertical_checkbox_cols = {2, 3, 4, 5, 6}  # Description..Reference Designator (shifted by Select + Change Type)
                # Default action is blank until user explicitly selects it.
                change_type_options_l0 = ['', 'Revised', 'Change']
                change_type_options_l1 = ['', 'Repl Item at Same Seq', 'Remove Item', 'Change']

                for r, row in enumerate(final_rows):
                    for c, val in enumerate(row):
                        if c == 0:  # Select column with checkbox
                            cont = QWidget()
                            h_lay = QHBoxLayout(cont)
                            h_lay.setContentsMargins(0, 0, 0, 0)
                            h_lay.addStretch(1)
                            chk = QCheckBox()
                            h_lay.addWidget(chk)
                            h_lay.addStretch(1)
                            cont._chk = chk
                            t.setCellWidget(r, c, cont)
                        elif c == 1:  # Change Type column
                            bom_val = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            rep_val = (str(row[replacement_col]).strip() if (replacement_col >= 0 and replacement_col < len(row)) else '')
                            part_val = (str(row[part_col]).lstrip(' \t') if (part_col >= 0 and part_col < len(row)) else '')
                            is_option_part = bool(part_val) and (part_val[:4] in option_prefixes)

                            combo = QComboBox()
                            # BOM Level 0 rows: Revised + Change; non-zero rows: no Revised
                            if bom_val == '0':
                                combo.addItems(change_type_options_l0)
                                # If Option part, set default to 'Revised'
                                if is_option_part:
                                    idx = combo.findText('Revised')
                                    if idx >= 0:
                                        combo.setCurrentIndex(idx)
                                    else:
                                        combo.setCurrentIndex(0)
                                else:
                                    combo.setCurrentIndex(0)
                            else:
                                combo.addItems(change_type_options_l1)
                                combo.setCurrentIndex(0)

                            combo.setStyleSheet('QComboBox { padding: 2px; }')
                            if hasattr(target, '_bind_change_type_combo'):
                                target._bind_change_type_combo(combo, r)
                            elif hasattr(target, '_on_change_type_changed'):
                                combo.currentTextChanged.connect(lambda text, rr=r: target._on_change_type_changed(rr))
                            t.setCellWidget(r, c, combo)
                        elif c in vertical_checkbox_cols:
                            bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            part_here = (str(row[part_col]).lstrip(' \t') if (part_col >= 0 and part_col < len(row)) else '')
                            is_option_part = bool(part_here) and (part_here[:4] in option_prefixes)

                            # BOM 0 rows: only Description(2) and Ref Designator(6) have checkboxes.
                            if bom_here == '0' and c not in {2, 6}:
                                t.setItem(r, c, QTableWidgetItem(''))
                            # Non-option rows: remove Ref Designator checkbox.
                            elif c == 6 and not is_option_part:
                                t.setItem(r, c, QTableWidgetItem(''))
                            else:
                                cont_v = QWidget()
                                hv = QHBoxLayout(cont_v)
                                hv.setContentsMargins(0, 0, 0, 0)
                                hv.addStretch(1)
                                chk_v = QCheckBox()
                                hv.addWidget(chk_v)
                                hv.addStretch(1)
                                cont_v._chk = chk_v
                                if hasattr(target, '_on_selector_checkbox_toggled'):
                                    chk_v.stateChanged.connect(target._on_selector_checkbox_toggled)
                                t.setCellWidget(r, c, cont_v)
                        else:
                            display_val = val
                            if c == part_col:
                                part_txt = ((val or '') if isinstance(val, str) else str(val or '')).lstrip(' \t')
                                bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                                try:
                                    bom_int = max(0, int(float(bom_here)))
                                except Exception:
                                    bom_int = 0
                                display_val = ('      ' * bom_int + part_txt) if part_txt else ''

                            item = QTableWidgetItem(display_val)
                            if c == part_col:
                                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            elif _norm_header(final_headers[c]) == 'description':
                                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            else:
                                item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                            t.setItem(r, c, item)

                    # Keep Part and Description data left-aligned.
                    if part_col >= 0:
                        part_item = t.item(r, part_col)
                        if part_item is not None:
                            part_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    desc_col_final = next(
                        (ci for ci in range(t.columnCount())
                         if _norm_header((t.horizontalHeaderItem(ci) or QTableWidgetItem()).text()) == 'description'),
                        -1
                    )
                    if desc_col_final >= 0:
                        desc_item = t.item(r, desc_col_final)
                        if desc_item is not None:
                            desc_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

                    # Skyblue template for BOM level 0 rows.
                    if bom_col >= 0 and bom_col < len(row) and str(row[bom_col]).strip() == '0':
                        for c in range(t.columnCount()):
                            if c == 0:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                            elif c == 1:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                            elif c in vertical_checkbox_cols:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                                else:
                                    # Plain item (no checkbox for this column on this row)
                                    cell = t.item(r, c)
                                    if cell:
                                        cell.setBackground(QColor('#87CEEB'))
                            else:
                                cell = t.item(r, c)
                                if cell:
                                    cell.setBackground(QColor('#87CEEB'))

                    # Enforce action-first behavior for each row.
                    if hasattr(target, '_on_change_type_changed'):
                        target._on_change_type_changed(r)

                # Final enforcement: keep Part and Description data left-aligned in all rows.
                if part_col >= 0:
                    for rr in range(t.rowCount()):
                        it = t.item(rr, part_col)
                        if it is not None:
                            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                desc_col_final = next(
                    (ci for ci in range(t.columnCount())
                     if _norm_header((t.horizontalHeaderItem(ci) or QTableWidgetItem()).text()) == 'description'),
                    -1
                )
                if desc_col_final >= 0:
                    for rr in range(t.rowCount()):
                        it = t.item(rr, desc_col_final)
                        if it is not None:
                            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

                t.resizeColumnsToContents()
                for vc in range(2, 7):
                    t.setColumnWidth(vc, 40)
                if hasattr(target, '_update_change_type_body_box'):
                    target._update_change_type_body_box()
                if hasattr(target, '_refresh_structure_action_buttons'):
                    target._refresh_structure_action_buttons()
            finally:
                if _bulk_guard:
                    target._struct_item_change_guard = False
            if move_mode == 'merge':
                QMessageBox.information(
                    self,
                    'Move to Structure Sheet',
                    f'Merge completed. Added {merge_added_count} new row(s); updated Replacement in {merge_replacement_updates} existing row(s).',
                )
            else:
                QMessageBox.information(
                    self,
                    'Move to Structure Sheet',
                    f'Created Structure Sheet with {len(option_order)} unique option row(s) and {len(out_rows)} total row(s).',
                )
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))

    def _v2_append_selected_with_replacement(self):
        try:
            if not self.obs_provider:
                QMessageBox.information(self,'Append to OBS', 'OBS Parts tab is not available.')
                return
            t = self.obs_provider.table
            rows = self._selected_row_indices()
            if not rows:
                QMessageBox.information(self,'Append to OBS','No rows selected.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            count = 0
            for r in rows:
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
                rep  = (self.table.item(r, rcol).text() if (rcol>=0 and self.table.item(r, rcol)) else '').strip()
                if not part:
                    continue
                # Check existing
                existing=set()
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it: existing.add((it.text() or '').strip().upper())
                if part.upper() in existing:
                    # If exists, update replacement if empty
                    if rep:
                        for rr in range(t.rowCount()):
                            it=t.item(rr,1)
                            if it and (it.text() or '').strip().upper()==part.upper():
                                t.setItem(rr,3,QTableWidgetItem(rep))
                                break
                    continue
                target=None
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it is None or not (it.text() or '').strip():
                        target=rr; break
                if target is None:
                    target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
                t.setItem(target,1,QTableWidgetItem(part))
                w=t.cellWidget(target,2)
                if isinstance(w,QComboBox):
                    idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
                if rep:
                    t.setItem(target,3,QTableWidgetItem(rep))
                count += 1
            QMessageBox.information(self,'Append to OBS', f'Appended/Updated {count} part(s) to OBS List.')
        except Exception as e:
            QMessageBox.warning(self,'Append Error', str(e))

    def _v2_refresh_replacements(self):
        try:
            obs_map = self._build_obs_map()
            if not obs_map:
                QMessageBox.information(self, 'Refresh', 'No OBS parts found to refresh from.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            if pcol < 0 or rcol < 0:
                QMessageBox.information(self,'Refresh','Missing Part or Replacement column.')
                return
            updated = 0
            for r in range(self.table.rowCount()):
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '')
                key = (part or '').strip().upper()
                rep = obs_map.get(key, '')
                it = self.table.item(r, rcol)
                if it is None:
                    it = QTableWidgetItem('')
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(r, rcol, it)
                old = it.text()
                if rep != old:
                    it.setText(rep)
                    updated += 1
            QMessageBox.information(self,'Refresh', f'Refreshed Replacement column using OBS Parts list. Updated {updated} row(s).')
        except Exception as e:
            QMessageBox.warning(self,'Refresh Error', str(e))
class StructureSheetTab(QWidget):
    MAX_LINES = 500
    PART_LEN = 10
    INSERTED_COL_TO_SELECTOR = {
        'new description': 2,
        'new item seq': 3,
        'new qty': 4,
        'new kit code': 5,
        'new ref designator': 6,
        'new reference designator': 6,
        'new ref designator(rd)': 6,
    }

    DBKEY_BY_HEADER = {
        'part': 'part',
        'rev/ln': 'rev_ln',
        'plant': 'plant',
        'part description': 'description',
        'description': 'description',
        'item status': 'item_status',
        'base qty': 'base_qty',
        'ext qty': 'ext_qty',
        'uom': 'uom',
        'eco number': 'eco_number',
        'procurement type': 'procurement_type',
        'effectivity date': 'effectivity_date',
        'user item type': 'user_item_type',
        'item seq': 'item_seq',
        'kit code': 'kit_code',
        'sparable flag': 'sparable_flag',
        'pace': 'pace_or_dash',
        'mlo class': 'mlo_class',
    }

    def __init__(self):
        super().__init__()

        outer = QVBoxLayout(self)
        outer.setContentsMargins(8, 8, 8, 8)
        outer.setSpacing(8)
        title = QLabel("Structure Sheet")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        # ---- Structure Sheet Table ----
        structure_tab = QWidget()
        struct_layout = QVBoxLayout(structure_tab)
        struct_layout.setContentsMargins(0, 0, 0, 0)
        struct_layout.setSpacing(8)

        # Action buttons row
        btn_row = QHBoxLayout()
        self.btn_delete_row = QPushButton("Delete Selected Row(s)")
        self.btn_add_row = QPushButton("Add Row(s)")
        self.btn_build_sheet = QPushButton("Build Structure Sheet")
        self.lbl_build_plant = QLabel("Plant:")
        self.cmb_build_plant = QComboBox()
        self.cmb_build_plant.addItems(['4020', '4055', '4060', '4070', '4080', '4090'])
        self.cmb_build_plant.setCurrentText('4070')
        self.btn_insert_chk_cols = QPushButton("Insert/Remove Checkbox Columns")
        self.btn_update_part_info = QPushButton("Update Part Information")
        self.btn_export = QPushButton("Export Excel")
        self.btn_import = QPushButton("Import Excel")
        self.btn_reset_row = QPushButton("Reset")
        btn_row.addWidget(self.btn_add_row)
        btn_row.addWidget(self.btn_delete_row)
        btn_row.addWidget(self.btn_build_sheet)
        btn_row.addWidget(self.lbl_build_plant)
        btn_row.addWidget(self.cmb_build_plant)
        btn_row.addWidget(self.btn_insert_chk_cols)
        btn_row.addWidget(self.btn_update_part_info)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_export)
        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(self.btn_reset_row)
        struct_layout.addLayout(btn_row)

        # Use QSplitter to position input on left and table on right
        self.structure_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Part input widget for building structure sheet from part numbers (LEFT side)
        self.part_input_container = QWidget()
        self.part_input_container.setMinimumWidth(320)
        self.part_input_container.setMaximumWidth(360)
        part_input_layout = QVBoxLayout(self.part_input_container)
        part_input_layout.setContentsMargins(5, 5, 5, 5)
        part_input_layout.setSpacing(6)
        part_input_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        part_input_label = QLabel("Enter Part Numbers (Max 500, one per line or comma-separated):")
        part_input_label.setWordWrap(True)
        part_input_layout.addWidget(part_input_label)
        self.part_input_text = QTextEdit()
        self.part_input_text.setPlaceholderText("Enter option/part numbers...")
        self.part_input_text.setFixedHeight(300)
        part_input_layout.addWidget(self.part_input_text)
        self.structure_splitter.addWidget(self.part_input_container)
        
        # Table widget (RIGHT side)
        self.table = QTableWidget(0, 0)
        self._struct_item_change_guard = False
        self._struct_action_prompt_guard = False
        self._suppress_struct_action_prompt = False
        self._change_type_body_box = None
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
            QHeaderView::section {
                background:#E1F0FF; font-weight:600;
                border:1px solid #C9E2FF; padding:4px;
            }
            QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        self.structure_splitter.addWidget(self.table)
        self.structure_splitter.setCollapsible(0, True)
        self.structure_splitter.setSizes([330, 900])
        self.structure_splitter.setStretchFactor(0, 0)
        self.structure_splitter.setStretchFactor(1, 1)
        struct_layout.addWidget(self.structure_splitter, 1)

        self.btn_delete_row.clicked.connect(self._delete_selected_struct_rows)
        self.btn_add_row.clicked.connect(self._add_struct_row)
        self.btn_build_sheet.clicked.connect(self._on_build_structure_sheet_clicked)
        self.btn_insert_chk_cols.clicked.connect(self._insert_checkbox_columns)
        self.btn_update_part_info.clicked.connect(self._update_part_information)
        self.btn_export.clicked.connect(self._export_structure_sheet)
        self.btn_import.clicked.connect(self._import_structure_sheet)
        self.btn_reset_row.clicked.connect(self._reset_struct_table)

        # Keep blue body box synced with scroll/resize/content changes.
        self.table.horizontalScrollBar().valueChanged.connect(lambda _v: self._update_change_type_body_box())
        self.table.verticalScrollBar().valueChanged.connect(lambda _v: self._update_change_type_body_box())
        self.table.horizontalHeader().sectionResized.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().rowsInserted.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().rowsRemoved.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().columnsInserted.connect(lambda *_args: self._update_change_type_body_box())
        self.table.model().columnsRemoved.connect(lambda *_args: self._update_change_type_body_box())
        self.table.itemChanged.connect(self._on_struct_item_changed)

        # Keep reset color consistent with other reset buttons in the app.
        self.btn_reset_row.setStyleSheet(
            "QPushButton{background:#90A4AE;color:white;padding:6px 12px;border-radius:6px;}"
            "QPushButton:hover{background:#7C919B;}"
        )

        self._refresh_structure_action_buttons()

        outer.addWidget(structure_tab, 1)

    def _refresh_structure_action_buttons(self):
        """Show Build Sheet input only when sheet is empty; reveal others when data exists."""
        has_data = self.table.rowCount() > 0
        gated = (
            self.btn_add_row,
            self.btn_delete_row,
            self.btn_insert_chk_cols,
            self.btn_update_part_info,
            self.btn_export,
            self.btn_import,
            self.btn_reset_row,
        )
        for b in gated:
            b.setVisible(has_data)
            b.setEnabled(has_data)

        # Show part input and Build button only when sheet is empty
        self.part_input_container.setVisible(not has_data)
        self.btn_build_sheet.setVisible(not has_data)
        self.btn_build_sheet.setEnabled(not has_data)
        self.lbl_build_plant.setVisible(not has_data)
        self.cmb_build_plant.setVisible(not has_data)
        self.cmb_build_plant.setEnabled(not has_data)
        if hasattr(self, 'structure_splitter'):
            if has_data:
                self.table.setVisible(True)
                self.structure_splitter.setSizes([0, 2000])
            else:
                self.table.setVisible(False)
                self.structure_splitter.setSizes([2000, 0])

    def showEvent(self, event):
        super().showEvent(event)
        self._reapply_action_visibility_all_rows()

    def _on_build_structure_sheet_clicked(self):
        """Handler for Build Structure Sheet button: parse part numbers and fetch BOM."""
        try:
            text = self.part_input_text.toPlainText().strip()
            if not text:
                QMessageBox.information(self, 'Build Structure Sheet', 'Please enter at least one part number.')
                return

            # Parse part numbers: split by newline and/or comma
            parts = []
            for line in text.split('\n'):
                for item in line.split(','):
                    item = item.strip()
                    if item:
                        parts.append(item)

            # Validate max 500 parts
            if len(parts) > 500:
                QMessageBox.warning(
                    self,
                    'Build Structure Sheet',
                    f'Too many part numbers. Maximum is 500, you entered {len(parts)}.'
                )
                return

            if not parts:
                QMessageBox.information(self, 'Build Structure Sheet', 'No valid part numbers found.')
                return

            # Build the structure sheet
            self._build_structure_sheet_from_parts(parts)
            
            # Clear input and refresh buttons
            self.part_input_text.clear()
            self._refresh_structure_action_buttons()

        except Exception as e:
            QMessageBox.warning(self, 'Build Structure Sheet Error', f'{str(e)}')

    def _validate_parts(self):
        pass

    def get_impacted_parts(self):
        return []

    # ---- helper: populate one new empty row at index `row_idx` ----
    def _init_struct_row(self, row_idx: int):
        # For newly added rows, include Add/Replace/Remove/Change (no Revised).
        change_type_options = ['', 'Repl Item at Same Seq', 'Remove Item', 'Add Item', 'Change']
        part_col = self._struct_part_col()
        for c in range(self.table.columnCount()):
            if c == 0:  # Select checkbox
                cont = QWidget()
                h_lay = QHBoxLayout(cont)
                h_lay.setContentsMargins(0, 0, 0, 0)
                h_lay.addStretch(1)
                chk = QCheckBox()
                h_lay.addWidget(chk)
                h_lay.addStretch(1)
                cont._chk = chk
                self.table.setCellWidget(row_idx, c, cont)
            elif c == 1:  # Change Type dropdown
                combo = QComboBox()
                combo.addItems(change_type_options)
                combo.setCurrentIndex(0)
                combo.setStyleSheet('QComboBox { padding: 2px; }')
                self._bind_change_type_combo(combo, row_idx)
                self.table.setCellWidget(row_idx, c, combo)
            elif c in {2, 3, 4, 5, 6}:  # Vertical-header checkbox columns
                cont_v = QWidget()
                hv = QHBoxLayout(cont_v)
                hv.setContentsMargins(0, 0, 0, 0)
                hv.addStretch(1)
                chk_v = QCheckBox()
                hv.addWidget(chk_v)
                hv.addStretch(1)
                cont_v._chk = chk_v
                chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                self.table.setCellWidget(row_idx, c, cont_v)
            else:
                item = QTableWidgetItem('')
                h = self.table.horizontalHeaderItem(c)
                if h and self._selector_col_for_inserted_header(h.text()) >= 0:
                    item.setForeground(QBrush(QColor('#FF8C00')))
                    selector_col = self._selector_col_for_inserted_header(h.text())
                    selector_widget = self.table.cellWidget(row_idx, selector_col)
                    is_checked = bool(selector_widget and hasattr(selector_widget, '_chk') and selector_widget._chk.isChecked())
                    self._set_item_editable(item, is_checked)
                if c == part_col:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    item.setData(Qt.ItemDataRole.UserRole, 'manual-added')
                self.table.setItem(row_idx, c, item)

        self._apply_manual_added_row_style(row_idx)
        self._on_change_type_changed(row_idx)

    def _norm_header(self, v: str) -> str:
        return (v or '').strip().lower()

    def _find_col_by_header(self, header: str) -> int:
        key = self._norm_header(header)
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h and self._norm_header(h.text()) == key:
                return c
        return -1

    def _current_change_type(self, row_idx: int) -> str:
        combo = self.table.cellWidget(row_idx, 1)
        if combo and isinstance(combo, QComboBox):
            return (combo.currentText() or '').strip()
        it = self.table.item(row_idx, 1)
        return (it.text() if it else '').strip()

    def _find_row_for_action_combo(self, combo: QComboBox) -> int:
        """Find the row index of an action combo widget in column 1."""
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 1)
            if w is combo:
                return r
        return -1

    def _on_action_combo_changed(self, _text: str):
        """Wrapper for action combo change that finds the row dynamically."""
        combo = self.sender()
        if not isinstance(combo, QComboBox):
            return
        row_idx = self._find_row_for_action_combo(combo)
        if row_idx >= 0:
            self._on_change_type_changed(row_idx)

    def _bind_change_type_combo(self, combo: QComboBox, row_idx: int):
        if combo is None:
            return
        combo.setProperty('_last_action', (combo.currentText() or '').strip())
        combo.currentTextChanged.connect(self._on_action_combo_changed)

    def _clear_row_action_changes(self, row_idx: int):
        if row_idx < 0 or row_idx >= self.table.rowCount():
            return

        for col in {2, 3, 4, 5, 6}:
            w = self.table.cellWidget(row_idx, col)
            if w and hasattr(w, '_chk'):
                w._chk.blockSignals(True)
                w._chk.setChecked(False)
                w._chk.blockSignals(False)

        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if not h:
                continue
            if self._selector_col_for_inserted_header(h.text()) < 0:
                continue
            item = self.table.item(row_idx, c)
            if item is not None and (item.text() or ''):
                item.setText('')

        self._update_inserted_editability_for_row(row_idx)

    def _row_has_changes(self, row_idx: int) -> bool:
        """Check if a row has any actual changes (checkboxes checked or data in inserted columns)."""
        if row_idx < 0 or row_idx >= self.table.rowCount():
            return False
        for col in {2, 3, 4, 5, 6}:
            w = self.table.cellWidget(row_idx, col)
            if w and hasattr(w, '_chk') and w._chk.isChecked():
                return True
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if not h:
                continue
            if self._selector_col_for_inserted_header(h.text()) >= 0:
                item = self.table.item(row_idx, c)
                if item is not None and (item.text() or '').strip():
                    return True
        return False

    def _get_bom_level_for_row(self, row_idx: int) -> str:
        """Get BOM level for a row."""
        bom_col = self._find_col_by_header('BOM Level')
        if bom_col < 0:
            return ''
        item = self.table.item(row_idx, bom_col)
        return (item.text() or '').strip() if item else ''

    def _is_option_part_in_row(self, row_idx: int) -> bool:
        """Check if the part in a row is an option part."""
        option_prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
            '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
            '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
        }
        part_col = self._struct_part_col()
        if part_col < 0:
            return False
        item = self.table.item(row_idx, part_col)
        part = (item.text() if item else '').strip().lstrip(' \t')
        return bool(part) and (part[:4] in option_prefixes)

    def _row_has_kit_code_value(self, row_idx: int) -> bool:
        """True when at least one data 'Kit Code' cell in the row has a value."""
        def _cell_text_from_widget(widget) -> str:
            if widget is None:
                return ''
            if isinstance(widget, QComboBox):
                return (widget.currentText() or '').strip()
            if isinstance(widget, QLineEdit):
                return (widget.text() or '').strip()
            if isinstance(widget, QLabel):
                return (widget.text() or '').strip()
            if hasattr(widget, 'toPlainText'):
                try:
                    return (widget.toPlainText() or '').strip()
                except Exception:
                    return ''
            if hasattr(widget, 'text'):
                try:
                    return (widget.text() or '').strip()
                except Exception:
                    return ''
            return ''

        for c in range(self.table.columnCount()):
            # Ignore selector/helper columns and check only real data columns.
            if c in {0, 1, 2, 3, 4, 5, 6}:
                continue
            h = self.table.horizontalHeaderItem(c)
            if not h or self._norm_header(h.text()) != 'kit code':
                continue

            w = self.table.cellWidget(row_idx, c)
            if _cell_text_from_widget(w):
                return True

            item = self.table.item(row_idx, c)
            if item and (item.text() or '').strip():
                return True
        return False

    def _apply_action_checkbox_visibility(self, row_idx: int, action: str, reset_checks: bool = True):
        """Apply checkbox visibility rules based on action and BOM level."""
        bom_level = self._get_bom_level_for_row(row_idx)
        is_option = self._is_option_part_in_row(row_idx)
        has_kit_code = self._row_has_kit_code_value(row_idx)
        visible_cols = set()
        if reset_checks:
            for col in {2, 3, 4, 5, 6}:
                w = self.table.cellWidget(row_idx, col)
                if w and hasattr(w, '_chk'):
                    w._chk.blockSignals(True)
                    w._chk.setChecked(False)
                    w._chk.blockSignals(False)
        if bom_level == '0':
            if action == 'Revised':
                visible_cols = {2, 6}
            elif action == 'Change':
                visible_cols = {2, 6}
        else:
            if action == 'Repl Item at Same Seq':
                visible_cols = {4, 5}
                if is_option:
                    visible_cols.add(6)
            elif action == 'Change':
                # Explicitly remove 'Part Description' (col 2) for BOM Level != 0 and Action == 'Change'
                visible_cols = {3, 4, 5}
                if is_option:
                    visible_cols.add(6)

        # Kit Code selector rule:
        # - If row has a Kit Code value and an action is selected, keep it visible
        #   irrespective of BOM level.
        # - Keep existing Remove-item exception for BOM Level != 0.
        if action and has_kit_code and not (bom_level != '0' and action == 'Remove Item'):
            visible_cols.add(5)
        else:
            visible_cols.discard(5)

        for col in {2, 3, 4, 5, 6}:
            if col in visible_cols:
                self._ensure_selector_checkbox_widget(row_idx, col)
            w = self.table.cellWidget(row_idx, col)
            if w and hasattr(w, '_chk'):
                if col in visible_cols:
                    w._chk.setEnabled(True)
                    w.show()
                else:
                    w._chk.setEnabled(False)
                    w.hide()

    def _reapply_action_visibility_all_rows(self):
        """Reapply row action-based selector visibility (useful after tab switches)."""
        if self.table.rowCount() <= 0:
            return
        for r in range(self.table.rowCount()):
            # Preserve user-entered selector choices while refreshing UI after tab switches.
            self._apply_action_checkbox_visibility(r, self._current_change_type(r), reset_checks=False)
            self._update_inserted_editability_for_row(r)

    def _confirm_struct_action_change(self) -> bool:
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setWindowTitle('Confirm Action Change')
        msg.setText('If the Action is changed, all the changes made under the previously selected Action will not be saved.')
        msg.setInformativeText('Continue or Cancel?')
        btn_continue = msg.addButton('Continue', QMessageBox.ButtonRole.AcceptRole)
        btn_cancel = msg.addButton('Cancel', QMessageBox.ButtonRole.RejectRole)
        msg.setDefaultButton(btn_cancel)
        msg.exec()
        return msg.clickedButton() is btn_continue

    def _is_action_selected(self, row_idx: int) -> bool:
        return bool(self._current_change_type(row_idx))

    def _show_select_action_first_prompt(self):
        if self._struct_action_prompt_guard:
            return
        self._struct_action_prompt_guard = True
        try:
            QMessageBox.information(self, 'Select Action', 'Please select Action first.')
        finally:
            self._struct_action_prompt_guard = False

    def _on_struct_item_changed(self, item: QTableWidgetItem):
        if self._struct_item_change_guard:
            return
        if item is None:
            return
        row_idx = item.row()
        col_idx = item.column()
        h = self.table.horizontalHeaderItem(col_idx)
        is_inserted_col = bool(h and self._selector_col_for_inserted_header(h.text()) >= 0)
        # Ignore select/action/checkbox helper columns.
        if col_idx in {0, 1, 2, 3, 4, 5, 6}:
            return
        txt = (item.text() or '').strip()
        if is_inserted_col:
            self._sync_inserted_item_visual_state(row_idx, item)
        if not txt:
            return
        if self._is_action_selected(row_idx):
            return

    def _get_bom0_parts(self) -> set[str]:
        bom_col = self._find_col_by_header('BOM Level')
        part_col = self._find_col_by_header('Part')
        parts = set()
        if bom_col < 0 or part_col < 0:
            return parts
        for r in range(self.table.rowCount()):
            b = self.table.item(r, bom_col)
            p = self.table.item(r, part_col)
            if b and p and (b.text() or '').strip() == '0':
                pt = (p.text() or '').strip().upper()
                if pt:
                    parts.add(pt)
        return parts

    def _affected_parent_bom0_for_row(self, row_idx: int) -> str:
        """Return nearest BOM level-0 parent part above the row, if any."""
        bom_col = self._find_col_by_header('BOM Level')
        part_col = self._find_col_by_header('Part')
        if bom_col < 0 or part_col < 0:
            return ''
        for r in range(row_idx, -1, -1):
            b = self.table.item(r, bom_col)
            p = self.table.item(r, part_col)
            btxt = (b.text() if b else '').strip()
            ptxt = ((p.text() if p else '') or '').lstrip(' \t').strip().upper()
            if btxt == '0' and ptxt:
                return ptxt
        return ''

    def _is_entered_part_row(self, row_idx: int) -> bool:
        """Treat orange Part cell rows as user-entered rows for update validation."""
        part_col = self._find_col_by_header('Part')
        if part_col < 0:
            return False
        it = self.table.item(row_idx, part_col)
        if it is None:
            return False
        bg = it.background().color()
        return bg.name().lower() == '#ffe5cc'

    def _looks_like_part(self, part: str) -> bool:
        p = (part or '').strip().upper()
        if not p:
            return False
        if p.startswith('ESW') and len(p) > 10:
            return True
        return len(p) == 10 and p[4] == '-'

    def _set_part_cell_style(self, row_idx: int, editable: bool, orange: bool):
        part_col = self._find_col_by_header('Part')
        if part_col < 0:
            return
        it = self.table.item(row_idx, part_col)
        if it is None:
            it = QTableWidgetItem('')
            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row_idx, part_col, it)

        self._set_item_editable(it, editable)
        if orange:
            it.setBackground(QColor('#FFE5CC'))
        else:
            bom_col = self._find_col_by_header('BOM Level')
            is_bom0 = False
            if bom_col >= 0:
                bom_it = self.table.item(row_idx, bom_col)
                is_bom0 = bool(bom_it and (bom_it.text() or '').strip() == '0')
            it.setBackground(QColor('#87CEEB') if is_bom0 else QColor('#FFFFFF'))

    def _ensure_change_type_body_box(self):
        if self._change_type_body_box is not None:
            return
        box = QFrame(self.table.viewport())
        box.setObjectName('changeTypeBodyBox')
        box.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        box.setStyleSheet('QFrame#changeTypeBodyBox { border: 2px solid #1565C0; background: transparent; }')
        box.hide()
        self._change_type_body_box = box

    def _update_change_type_body_box(self):
        self._ensure_change_type_body_box()
        box = self._change_type_body_box
        if box is None:
            return
        if self.table.columnCount() <= 6 or self.table.rowCount() == 0:
            box.hide()
            return

        left_col = 2
        right_col = 6
        if right_col >= self.table.columnCount():
            box.hide()
            return

        left = self.table.columnViewportPosition(left_col)
        right = self.table.columnViewportPosition(right_col) + self.table.columnWidth(right_col) - 1
        top = self.table.rowViewportPosition(0)
        last_row = self.table.rowCount() - 1
        bottom = self.table.rowViewportPosition(last_row) + self.table.rowHeight(last_row) - 1

        if right < left or bottom < top:
            box.hide()
            return

        box.setGeometry(left, top, right - left + 1, bottom - top + 1)
        box.show()
        box.raise_()

    def _populate_row_from_where_used_record(self, row_idx: int, rec: dict,
                                              allowed_headers: set[str] | None = None,
                                              only_blank: bool = False):
        # Determine if this row is a BOM L0 row so new items get sky-blue background
        bom_col = self._find_col_by_header('BOM Level')
        is_bom0 = False
        if bom_col >= 0:
            bom_it = self.table.item(row_idx, bom_col)
            if bom_it and (bom_it.text() or '').strip() == '0':
                is_bom0 = True
        sky = QColor('#87CEEB')

        self._struct_item_change_guard = True
        try:
            for c in range(self.table.columnCount()):
                if c in {2, 3, 4, 5, 6}:  # rotated checkbox columns
                    continue
                h = self.table.horizontalHeaderItem(c)
                if not h:
                    continue
                h_norm = self._norm_header(h.text())
                if allowed_headers is not None and h_norm not in allowed_headers:
                    continue
                key = self.DBKEY_BY_HEADER.get(self._norm_header(h.text()))
                if not key:
                    continue
                value = '' if rec.get(key) is None else str(rec.get(key)).strip()
                if key == 'description' and h_norm in {'part description'}:
                    # Keep checkbox helper column untouched; actual part description
                    # should go to the regular Description column.
                    continue
                item = self.table.item(row_idx, c)
                if item is None:
                    item = QTableWidgetItem('')
                    if is_bom0:
                        item.setBackground(sky)
                    self.table.setItem(row_idx, c, item)

                # Determine alignment for this column
                if h_norm == 'part':
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                elif h_norm in {'description', 'part description'}:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)

                # only_blank: skip cells that already have a value
                if only_blank and (item.text() or '').strip():
                    continue

                # Preserve leading indent for Part column
                if h_norm == 'part':
                    existing = item.text()
                    leading = existing[:len(existing) - len(existing.lstrip(' \t'))]
                    value = leading + value.lstrip(' \t')

                item.setText(value)

            # Optional Ref Designator enrichment if query returns any key variant.
            ref_val = ''
            for k in ('ref_designator', 'reference_designator', 'reference designator'):
                if rec.get(k):
                    ref_val = str(rec.get(k)).strip()
                    break
            if ref_val:
                for c in range(self.table.columnCount()):
                    h = self.table.horizontalHeaderItem(c)
                    if not h:
                        continue
                    if self._norm_header(h.text()) in {'ref designator', 'reference designator',
                                                        'ref designator(rd)'}:
                        it = self.table.item(row_idx, c)
                        if it is None:
                            it = QTableWidgetItem('')
                            it.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            self.table.setItem(row_idx, c, it)
                        if only_blank and (it.text() or '').strip():
                            continue
                        it.setText(ref_val)
        finally:
            self._struct_item_change_guard = False

    def _validate_and_fill_part_for_row(self, row_idx: int, show_message: bool = True) -> tuple[bool, str]:
        part_col = self._find_col_by_header('Part')
        plant_col = self._find_col_by_header('Plant')
        if part_col < 0:
            return False, 'Part column not found.'

        part_item = self.table.item(row_idx, part_col)
        part = (part_item.text() if part_item else '').strip()
        if not self._looks_like_part(part):
            return False, 'Invalid part format.'

        plant = '4070'
        if plant_col >= 0:
            pit = self.table.item(row_idx, plant_col)
            if pit and (pit.text() or '').strip():
                plant = (pit.text() or '').strip()

        try:
            from where_used_query import fetch_where_used
            records = fetch_where_used([part], 1, plant=plant)
        except Exception as e:
            reason = f'Failed to query where-used for part {part}: {e}'
            if show_message:
                QMessageBox.warning(self, 'Validation Failed', reason)
            return False, reason

        action = self._current_change_type(row_idx)

        parent_l1 = {
            (r.get('part') or '').strip().upper()
            for r in records
            if (r.get('wu_level') or '').strip() == '1'
        }
        affected_parent = self._affected_parent_bom0_for_row(row_idx)

        # Replace/Remove must report to an existing level-0 parent.
        if action in {'Repl Item at Same Seq', 'Remove Item'}:
            if not affected_parent:
                full_reason = f'Part {part} cannot be validated because affected 0th Level Parent Part is not found.'
                if show_message:
                    QMessageBox.warning(self, 'Validation Failed', full_reason)
                return False, full_reason
            if affected_parent not in parent_l1:
                full_reason = f'Part {part} is not a BOM Item of {affected_parent} (0th Level Parent Part).'
                if show_message:
                    QMessageBox.warning(self, 'Validation Failed', full_reason)
                return False, full_reason

        rec_l0 = None
        for r in records:
            if (r.get('wu_level') or '').strip() == '0':
                rec_l0 = r
                break
        if rec_l0:
            if action == 'Add Item':
                self._populate_row_from_where_used_record(
                    row_idx,
                    rec_l0,
                    allowed_headers={
                        'description', 'item status', 'uom', 'procurement type',
                        'pace', 'mlo class', 'ref designator', 'reference designator',
                        'ref designator(rd)', 'part',
                    },
                )
            else:
                # Repl/Remove: only fill blank cells to preserve existing data
                self._populate_row_from_where_used_record(row_idx, rec_l0, only_blank=True)

        self._set_part_cell_style(row_idx, editable=True, orange=False)
        return True, ''

    def _update_part_information(self):
        target_actions = {'Repl Item at Same Seq', 'Remove Item', 'Add Item'}
        part_col = self._find_col_by_header('Part')
        if part_col < 0:
            QMessageBox.warning(self, 'Update Part Information', 'Part column not found.')
            return

        rows_to_check = []
        smbom_rows = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, part_col)
            # SmBOM rows are identified by UserRole == 'smbom' on the Part cell
            if it and it.data(Qt.ItemDataRole.UserRole) == 'smbom':
                part = (it.text() or '').strip()
                if part:
                    smbom_rows.append(r)
                continue
            action = self._current_change_type(r)
            if action not in target_actions:
                continue
            part = (it.text() if it else '').strip()
            # Only process rows where user entered/edited part (orange part cell)
            if part and self._is_entered_part_row(r):
                rows_to_check.append(r)

        if not rows_to_check and not smbom_rows:
            QMessageBox.information(
                self,
                'Update Part Information',
                'No entered part numbers found.'
            )
            return

        ok_count = 0
        failed = []
        for r in rows_to_check:
            ok, reason = self._validate_and_fill_part_for_row(r, show_message=False)
            if ok:
                ok_count += 1
            else:
                failed.append(reason)

        # Process SmBOM rows from bottom to top to keep row indices stable after inserts
        for r in sorted(smbom_rows, reverse=True):
            ok, reason = self._load_smbom_for_row(r)
            if ok:
                ok_count += 1
            else:
                failed.append(reason)

        if failed:
            numbered = [f'{i + 1}. {msg}' for i, msg in enumerate(failed[:8])]
            preview = '\n'.join(numbered)
            more = '' if len(failed) <= 8 else f'\n...and {len(failed) - 8} more failure(s).'
            QMessageBox.warning(
                self,
                'Update Part Information',
                f'Updated {ok_count} row(s). Failed {len(failed)} row(s).\n\n{preview}{more}'
            )
        else:
            QMessageBox.information(
                self,
                'Update Part Information',
                f'Updated part information for {ok_count} row(s).'
            )

    def _selector_col_for_inserted_header(self, header_text: str) -> int:
        return self.INSERTED_COL_TO_SELECTOR.get((header_text or '').strip().lower(), -1)

    def _default_struct_cell_bg(self, row_idx: int) -> QColor:
        if self._is_manual_added_row(row_idx):
            return QColor('#FFF2E0')
        bom_col = self._find_col_by_header('BOM Level')
        if bom_col >= 0:
            bom_it = self.table.item(row_idx, bom_col)
            if bom_it and (bom_it.text() or '').strip() == '0':
                return QColor('#87CEEB')
        return QColor('#FFFFFF')

    def _is_manual_added_row(self, row_idx: int) -> bool:
        part_col = self._struct_part_col()
        if part_col < 0:
            return False
        it = self.table.item(row_idx, part_col)
        return bool(it and it.data(Qt.ItemDataRole.UserRole) == 'manual-added')

    def _apply_manual_added_row_style(self, row_idx: int):
        if not self._is_manual_added_row(row_idx):
            return

        row_bg = '#FFF2E0'
        combo_bg = '#FFF2E0'
        widget_bg = 'QWidget { background-color: #FFF2E0; }'
        combo_style = 'QComboBox { padding: 2px; background-color: #FFF2E0; }'

        for c in range(self.table.columnCount()):
            if c == 0:
                w = self.table.cellWidget(row_idx, c)
                if w:
                    w.setStyleSheet(widget_bg)
            elif c == 1:
                w = self.table.cellWidget(row_idx, c)
                if w and isinstance(w, QComboBox):
                    w.setStyleSheet(combo_style)
            elif c in {2, 3, 4, 5, 6}:
                w = self.table.cellWidget(row_idx, c)
                if w:
                    w.setStyleSheet(widget_bg)
                else:
                    it = self.table.item(row_idx, c)
                    if it is not None:
                        it.setBackground(QColor(row_bg))
            else:
                it = self.table.item(row_idx, c)
                if it is not None and c != self._struct_part_col():
                    it.setBackground(QColor(row_bg))

    def _sync_inserted_item_visual_state(self, row_idx: int, item: QTableWidgetItem):
        if item is None:
            return
        header_item = self.table.horizontalHeaderItem(item.column())
        selector_col = self._selector_col_for_inserted_header(header_item.text() if header_item else '')
        if selector_col < 0:
            return

        item.setForeground(QBrush(QColor('#FF8C00')))
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        w = self.table.cellWidget(row_idx, selector_col)
        is_checked = bool(w and hasattr(w, '_chk') and w._chk.isChecked())
        has_text = bool((item.text() or '').strip())
        if is_checked and not has_text:
            item.setBackground(QColor('#FFE5CC'))
        else:
            item.setBackground(self._default_struct_cell_bg(row_idx))

    def _apply_inserted_header_colors(self):
        base = QBrush(QColor('#0F2D46'))
        orange = QBrush(QColor('#FF8C00'))
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h is None:
                continue
            if self._selector_col_for_inserted_header(h.text()) >= 0:
                h.setForeground(orange)
            else:
                h.setForeground(base)

    def _ensure_change_type_group_header(self):
        if self.table.columnCount() == 0:
            return
        hh = self.table.horizontalHeader()
        if isinstance(hh, RotatedColumnsHeader):
            hh.set_rotated_columns(range(2, 7))
            hh._group_label = 'Change Type'
            hh._group_columns = list(range(2, 7))
            hh.setMinimumHeight(117)
            hh.viewport().update()
            return

        headers = []
        for c in range(self.table.columnCount()):
            hi = self.table.horizontalHeaderItem(c)
            headers.append(hi.text() if hi else f'Col{c}')
        hdr = RotatedColumnsHeader(
            Qt.Orientation.Horizontal,
            rotated_columns=range(2, 7),
            parent=self.table,
            group_label='Change Type',
            group_columns=list(range(2, 7)),
        )
        hdr.setStretchLastSection(True)
        hdr.setSectionsClickable(True)
        self.table.setHorizontalHeader(hdr)
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.horizontalHeader().setMinimumHeight(117)
        self._apply_inserted_header_colors()

    def _set_item_editable(self, item: QTableWidgetItem, editable: bool):
        if item is None:
            return
        flags = item.flags()
        if editable:
            item.setFlags(flags | Qt.ItemFlag.ItemIsEditable)
        else:
            item.setFlags(flags & ~Qt.ItemFlag.ItemIsEditable)

    def _ensure_structure_sheet_template_headers(self):
        """Initialize default Structure Sheet headers when table is empty."""
        if self.table.columnCount() > 0:
            return

        headers = self._structure_sheet_headers()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        self._ensure_change_type_group_header()
        self._restore_rotated_column_widths()

        # Baseline widths/alignment similar to loaded template behavior
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.resizeColumnsToContents()
        self._update_change_type_body_box()

    def _find_header_col(self, header_name: str) -> int:
        key = (header_name or '').strip().lower()
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h and (h.text() or '').strip().lower() == key:
                return c
        return -1

    def _structure_sheet_headers(self) -> list[str]:
        return [
            'Select',
            'Action',
            'Part Description',
            'Seq#',
            'Qty',
            'Kit Code',
            'Ref Designator(RD)',
            'BOM Level',
            'Part',
            'Replacement',
            'Description',
            'Item Status',
            'Base Qty',
            'Ext Qty',
            'UOM',
            'Procurement Type',
            'Item Seq',
            'Kit Code',
            'Sparable flag',
            'Pace',
            'Ref Designator(RD)',
            'MLO Class',
        ]

    def _collect_incomplete_actions(self) -> dict:
        selectors = {
            2: 'New Description',
            3: 'New Item Seq',
            4: 'New Qty',
            5: 'New Kit Code',
            6: 'New Ref Designator',
        }
        checked_total = 0
        missing_col = 0
        blank_value = 0

        for r in range(self.table.rowCount()):
            for sel_col, new_hdr in selectors.items():
                w = self.table.cellWidget(r, sel_col)
                if not (w and hasattr(w, '_chk') and w._chk.isChecked()):
                    continue
                checked_total += 1

                new_col = self._find_header_col(new_hdr)
                if new_col < 0 and new_hdr == 'New Ref Designator':
                    new_col = self._find_header_col('New Reference Designator')
                if new_col < 0:
                    missing_col += 1
                    continue

                it = self.table.item(r, new_col)
                if it is None or not (it.text() or '').strip():
                    blank_value += 1

        return {
            'checked_total': checked_total,
            'missing_col': missing_col,
            'blank_value': blank_value,
            'incomplete_total': missing_col + blank_value,
        }

    def _undo_incomplete_actions(self):
        selectors = {
            2: 'New Description',
            3: 'New Item Seq',
            4: 'New Qty',
            5: 'New Kit Code',
            6: 'New Ref Designator',
        }
        for r in range(self.table.rowCount()):
            for sel_col, new_hdr in selectors.items():
                w = self.table.cellWidget(r, sel_col)
                if not (w and hasattr(w, '_chk') and w._chk.isChecked()):
                    continue

                new_col = self._find_header_col(new_hdr)
                if new_col < 0 and new_hdr == 'New Ref Designator':
                    new_col = self._find_header_col('New Reference Designator')

                is_incomplete = False
                if new_col < 0:
                    is_incomplete = True
                else:
                    it = self.table.item(r, new_col)
                    if it is None or not (it.text() or '').strip():
                        is_incomplete = True

                if is_incomplete:
                    w._chk.setChecked(False)

    def confirm_leave_with_incomplete(self) -> bool:
        s = self._collect_incomplete_actions()
        if s['incomplete_total'] <= 0:
            return True

        msg = (
            'Incomplete actions found in Structure sheet:\n\n'
            f"Checked selectors: {s['checked_total']}\n"
            f"Missing inserted columns: {s['missing_col']}\n"
            f"Blank values in inserted columns: {s['blank_value']}\n\n"
            'Incomplete actions will not be saved.\n'
            'Yes: undo only incomplete actions and continue.\n'
            'No: stay on this tab.'
        )
        ans = QMessageBox.question(
            self,
            'Unsaved Incomplete Actions',
            msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if ans == QMessageBox.StandardButton.Yes:
            self._undo_incomplete_actions()
            return True
        return False

    def _find_row_for_action_combo(self, combo: QComboBox) -> int:
        """Find the row index of an action combo widget in column 1."""
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 1)
            if w is combo:
                return r
        return -1

    def _find_row_for_selector_checkbox(self, chk: QCheckBox) -> int:
        for r in range(self.table.rowCount()):
            for c in (2, 3, 4, 5, 6):
                w = self.table.cellWidget(r, c)
                if w and hasattr(w, '_chk') and w._chk is chk:
                    return r
        return -1

    def _ensure_selector_checkbox_widget(self, row_idx: int, col_idx: int):
        """Create selector checkbox widget if the cell currently has only a plain item."""
        if row_idx < 0 or row_idx >= self.table.rowCount() or col_idx not in {2, 3, 4, 5, 6}:
            return
        existing = self.table.cellWidget(row_idx, col_idx)
        if existing and hasattr(existing, '_chk'):
            return

        cont_v = QWidget()
        hv = QHBoxLayout(cont_v)
        hv.setContentsMargins(0, 0, 0, 0)
        hv.addStretch(1)
        chk_v = QCheckBox()
        hv.addWidget(chk_v)
        hv.addStretch(1)
        cont_v._chk = chk_v
        chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)

        # Set background color to match the row style
        bg = self._default_struct_cell_bg(row_idx)
        cont_v.setStyleSheet(f"QWidget {{ background-color: {bg.name()}; }}")

        self.table.setCellWidget(row_idx, col_idx, cont_v)

    def _update_inserted_editability_for_row(self, row_idx: int):
        if row_idx < 0 or row_idx >= self.table.rowCount():
            return
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if not h:
                continue
            selector_col = self._selector_col_for_inserted_header(h.text())
            if selector_col < 0:
                continue
            item = self.table.item(row_idx, c)
            if item is None:
                item = QTableWidgetItem('')
                item.setForeground(QBrush(QColor('#FF8C00')))
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row_idx, c, item)
            w = self.table.cellWidget(row_idx, selector_col)
            is_checked = bool(w and hasattr(w, '_chk') and w._chk.isChecked())
            if not is_checked and item.text():
                item.setText('')
            self._set_item_editable(item, is_checked)
            self._sync_inserted_item_visual_state(row_idx, item)

    def _update_inserted_editability_all_rows(self):
        for r in range(self.table.rowCount()):
            self._update_inserted_editability_for_row(r)

    def _on_selector_checkbox_toggled(self, _state: int):
        chk = self.sender()
        if not isinstance(chk, QCheckBox):
            return
        row_idx = self._find_row_for_selector_checkbox(chk)
        if row_idx >= 0:
            if chk.isChecked() and not self._is_action_selected(row_idx):
                self._show_select_action_first_prompt()
                chk.blockSignals(True)
                chk.setChecked(False)
                chk.blockSignals(False)
                return
            self._update_inserted_editability_for_row(row_idx)

    def _restore_rotated_column_widths(self):
        self._ensure_change_type_group_header()
        hh = self.table.horizontalHeader()
        for vc in range(2, 7):
            if vc < self.table.columnCount():
                hh.setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                self.table.setColumnWidth(vc, 40)
        self._update_change_type_body_box()

    def _struct_part_col(self) -> int:
        """Return the index of the Part column, or -1 if not found."""
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            if h and h.text().strip().lower() == 'part':
                return c
        return -1

    def _on_change_type_changed(self, row_idx: int):
        """Handle Change Type dropdown value changes to show/hide checkboxes."""
        combo = self.table.cellWidget(row_idx, 1)
        if not combo or not isinstance(combo, QComboBox):
            return

        previous_action = (combo.property('_last_action') or '').strip()
        change_type = (combo.currentText() or '').strip()

        if (
            not self._suppress_struct_action_prompt
            and previous_action
            and previous_action != change_type
            and self._row_has_changes(row_idx)
        ):
            if not self._confirm_struct_action_change():
                self._suppress_struct_action_prompt = True
                try:
                    combo.blockSignals(True)
                    combo.setCurrentText(previous_action)
                finally:
                    combo.blockSignals(False)
                    self._suppress_struct_action_prompt = False
                return
            self._clear_row_action_changes(row_idx)

        self._struct_item_change_guard = True
        try:
            self._apply_action_checkbox_visibility(row_idx, change_type)

            action = self._current_change_type(row_idx)
            if action in {'Repl Item at Same Seq', 'Remove Item', 'Add Item', 'Revised', 'Change'}:
                self._set_part_cell_style(row_idx, editable=True, orange=True)
            else:
                self._set_part_cell_style(row_idx, editable=False, orange=False)

            self._update_inserted_editability_for_row(row_idx)
            self._apply_manual_added_row_style(row_idx)
            combo.setProperty('_last_action', change_type)
        finally:
            self._struct_item_change_guard = False

    def _delete_selected_struct_rows(self):
        """Delete rows where the Select checkbox (column 0) is checked."""
        rows_to_delete = []
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk') and w._chk.isChecked():
                rows_to_delete.append(r)
        if not rows_to_delete:
            QMessageBox.information(self, 'Delete Rows', 'No rows selected. Check the Select checkbox on the row(s) to delete.')
            return
        for r in reversed(rows_to_delete):
            self.table.removeRow(r)
        self._update_change_type_body_box()
        self._refresh_structure_action_buttons()
        QMessageBox.information(self, 'Delete Rows', f'Deleted {len(rows_to_delete)} row(s).')

    def _add_struct_row(self):
        """Insert a new empty row immediately below the last selected row, or append."""
        insert_after = -1
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk') and w._chk.isChecked():
                insert_after = r

        insert_at = insert_after + 1 if insert_after >= 0 else self.table.rowCount()
        self.table.insertRow(insert_at)
        self._init_struct_row(insert_at)
        self._update_change_type_body_box()
        self._refresh_structure_action_buttons()

    def _build_structure_sheet_from_parts(self, part_list: list):
        """Fetch level-0/1 Implemented BOM and render with the Structure Sheet template."""
        try:
            from implemented_bom_query import fetch_implemented_bom  # type: ignore[import]

            plant = (self.cmb_build_plant.currentText() or '').strip() if hasattr(self, 'cmb_build_plant') else ''
            if not plant:
                plant = getattr(self, '_current_plant', '4070')

            requested_parts = [str(p).strip().upper() for p in part_list if str(p).strip()]
            if not requested_parts:
                QMessageBox.information(self, 'Build Structure Sheet', 'No valid part numbers were provided.')
                return

            records = fetch_implemented_bom(requested_parts, max_level=1, plant=plant, include_level0=True)
            if not records:
                QMessageBox.information(
                    self,
                    'Build Structure Sheet',
                    f'No implemented BOM data found for the provided part numbers in plant {plant}.',
                )
                return

            def _norm_header(h: str) -> str:
                return ' '.join((h or '').strip().lower().split())

            ordered_roots = []
            groups = {}
            for root in requested_parts:
                if root not in groups:
                    ordered_roots.append(root)
                    groups[root] = {'root': None, 'children': [], 'seen': set()}

            for rec in records:
                root = str(rec.get('input_part', rec.get('part', '')) or '').strip().upper()
                if not root:
                    continue
                if root not in groups:
                    ordered_roots.append(root)
                    groups[root] = {'root': None, 'children': [], 'seen': set()}

                level_txt = str(rec.get('bom_level', '-1') or '-1').strip()
                try:
                    lvl = int(float(level_txt))
                except Exception:
                    lvl = -1

                if lvl == 0:
                    groups[root]['root'] = rec
                elif lvl == 1:
                    cpart = str(rec.get('part', '') or '').strip().upper()
                    if cpart and cpart not in groups[root]['seen']:
                        groups[root]['children'].append(rec)
                        groups[root]['seen'].add(cpart)

            final_headers = self._structure_sheet_headers()
            dbkey_by_header = {
                'bom level': 'bom_level',
                'part': 'part',
                'replacement': '',
                'rev/ln': 'rev_ln',
                'plant': 'plant',
                'description': 'description',
                'item status': 'item_status',
                'base qty': 'base_qty',
                'ext qty': 'ext_qty',
                'uom': 'uom',
                'eco number': 'eco_number',
                'procurement type': 'procurement_type',
                'effectivity date': 'effectivity_date',
                'user item type': 'user_item_type',
                'item seq': 'item_seq',
                'kit code': 'kit_code',
                'sparable flag': 'sparable_flag',
                'pace': 'pace_or_dash',
                'mlo class': 'mlo_class',
            }

            out_rows = []
            for root in ordered_roots:
                grp = groups.get(root, {'root': None, 'children': []})
                if grp.get('root') is not None:
                    out_rows.append(grp['root'])
                out_rows.extend(grp.get('children', []))

            if not out_rows:
                self._refresh_structure_action_buttons()
                QMessageBox.information(
                    self,
                    'Build Structure Sheet',
                    'No BOM Level 0/1 rows found for the provided part numbers.',
                )
                return

            final_rows = []
            for rec in out_rows:
                row = []
                for idx, h in enumerate(final_headers):
                    hk = _norm_header(h)
                    if hk in {'select', 'action'} or 2 <= idx <= 6:
                        val = ''
                    else:
                        dbk = dbkey_by_header.get(hk, '')
                        val = '' if not dbk else rec.get(dbk, '')
                    row.append('' if val is None else str(val))
                final_rows.append(row)

            obs_replacements = {}
            obs_tab = getattr(self.window(), 'obs_tab', None)
            if obs_tab and hasattr(obs_tab, 'table'):
                obs_table = obs_tab.table
                obs_part_col = -1
                obs_repl_col = -1
                for c in range(obs_table.columnCount()):
                    h = obs_table.horizontalHeaderItem(c)
                    if h:
                        nh = _norm_header(h.text())
                        if nh == 'part':
                            obs_part_col = c
                        elif nh == 'replacement':
                            obs_repl_col = c
                if obs_part_col >= 0 and obs_repl_col >= 0:
                    for rr in range(obs_table.rowCount()):
                        part_item = obs_table.item(rr, obs_part_col)
                        repl_item = obs_table.item(rr, obs_repl_col)
                        if part_item and repl_item:
                            part_val = (part_item.text() or '').strip().upper()
                            repl_val = (repl_item.text() or '').strip()
                            if part_val and repl_val:
                                obs_replacements[part_val] = repl_val

            if obs_replacements:
                ans = QMessageBox.question(
                    self,
                    'OBS Replacement Parts',
                    'Please confirm if the Replacement Parts to be mapped from OBS List.',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                if ans == QMessageBox.StandardButton.Yes:
                    repl_col_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'replacement'), -1)
                    part_col_idx = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'part'), -1)
                    if repl_col_idx >= 0 and part_col_idx >= 0:
                        for row in final_rows:
                            if part_col_idx < len(row):
                                part_val = (row[part_col_idx] or '').strip().upper()
                                if part_val in obs_replacements:
                                    row[repl_col_idx] = obs_replacements[part_val]

            t = self.table
            _bulk_guard = hasattr(self, '_struct_item_change_guard')
            if _bulk_guard:
                self._struct_item_change_guard = True
            try:
                t.clear()
                t.setRowCount(0)
                t.setColumnCount(len(final_headers))
                hdr = RotatedColumnsHeader(
                    Qt.Orientation.Horizontal,
                    rotated_columns=range(2, 7),
                    parent=t,
                    group_label='Change Type',
                    group_columns=list(range(2, 7)),
                )
                hdr.setStretchLastSection(True)
                hdr.setSectionsClickable(True)
                t.setHorizontalHeader(hdr)
                t.setHorizontalHeaderLabels(final_headers)
                t.setRowCount(len(final_rows))

                t.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
                t.horizontalHeader().setMinimumHeight(117)
                for vc in range(2, 7):
                    t.horizontalHeader().setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                    t.setColumnWidth(vc, 40)

                bom_col = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'bom level'), -1)
                part_col = next((i for i, h in enumerate(final_headers) if _norm_header(h) == 'part'), -1)
                vertical_checkbox_cols = {2, 3, 4, 5, 6}
                change_type_options_l0 = ['', 'Revised', 'Change']
                change_type_options_l1 = ['', 'Repl Item at Same Seq', 'Remove Item', 'Change']

                for r, row in enumerate(final_rows):
                    for c, val in enumerate(row):
                        if c == 0:
                            cont = QWidget()
                            h_lay = QHBoxLayout(cont)
                            h_lay.setContentsMargins(0, 0, 0, 0)
                            h_lay.addStretch(1)
                            chk = QCheckBox()
                            h_lay.addWidget(chk)
                            h_lay.addStretch(1)
                            cont._chk = chk
                            t.setCellWidget(r, c, cont)
                        elif c == 1:
                            bom_val = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            part_val = (str(row[part_col]).lstrip(' \t') if (part_col >= 0 and part_col < len(row)) else '')
                            option_prefixes = {
                                '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
                                '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
                                '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
                            }
                            is_option_part = bool(part_val) and (part_val[:4] in option_prefixes)
                            combo = QComboBox()
                            if bom_val == '0':
                                combo.addItems(change_type_options_l0)
                                if is_option_part:
                                    idx = combo.findText('Revised')
                                    if idx >= 0:
                                        combo.setCurrentIndex(idx)
                                    else:
                                        combo.setCurrentIndex(0)
                                else:
                                    combo.setCurrentIndex(0)
                            else:
                                combo.addItems(change_type_options_l1)
                                combo.setCurrentIndex(0)
                            combo.setStyleSheet('QComboBox { padding: 2px; }')
                            self._bind_change_type_combo(combo, r)
                            t.setCellWidget(r, c, combo)
                        elif c in vertical_checkbox_cols:
                            bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                            if bom_here == '0' and c not in {2, 6}:
                                t.setItem(r, c, QTableWidgetItem(''))
                            else:
                                cont_v = QWidget()
                                hv = QHBoxLayout(cont_v)
                                hv.setContentsMargins(0, 0, 0, 0)
                                hv.addStretch(1)
                                chk_v = QCheckBox()
                                hv.addWidget(chk_v)
                                hv.addStretch(1)
                                cont_v._chk = chk_v
                                chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                                t.setCellWidget(r, c, cont_v)
                        else:
                            display_val = val
                            if c == part_col:
                                part_txt = ((val or '') if isinstance(val, str) else str(val or '')).lstrip(' \t')
                                bom_here = (str(row[bom_col]).strip() if (bom_col >= 0 and bom_col < len(row)) else '')
                                display_val = part_txt if bom_here == '0' else ('      ' + part_txt if part_txt else '')
                            item = QTableWidgetItem(display_val)
                            if c == part_col or _norm_header(final_headers[c]) == 'description':
                                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                            else:
                                item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                            t.setItem(r, c, item)

                    if bom_col >= 0 and bom_col < len(row) and str(row[bom_col]).strip() == '0':
                        for c in range(t.columnCount()):
                            if c == 0:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                            elif c == 1:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                            elif c in vertical_checkbox_cols:
                                w = t.cellWidget(r, c)
                                if w:
                                    w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                                else:
                                    it = t.item(r, c)
                                    if it:
                                        it.setBackground(QColor('#87CEEB'))
                            else:
                                it = t.item(r, c)
                                if it:
                                    it.setBackground(QColor('#87CEEB'))

                    self._on_change_type_changed(r)

                t.resizeColumnsToContents()
                for vc in range(2, 7):
                    t.setColumnWidth(vc, 40)
                self._update_change_type_body_box()
                self._refresh_structure_action_buttons()
            finally:
                if _bulk_guard:
                    self._struct_item_change_guard = False

            root_count = sum(1 for r in out_rows if str(r.get('bom_level', '')).strip() == '0')
            child_count = sum(1 for r in out_rows if str(r.get('bom_level', '')).strip() == '1')
            QMessageBox.information(
                self,
                'Build Structure Sheet',
                f'Loaded Implemented BOM successfully. Root rows: {root_count}, Level 1 rows: {child_count}.',
            )
        except Exception as e:
            QMessageBox.warning(self, 'Build Structure Sheet Error', str(e))

    # ---- SmBOM row support ----
    def _add_smbom_row(self):
        """Append a new sky-blue SmBOM row at the bottom with Part cell orange for input."""
        if self.table.columnCount() == 0:
            self._ensure_structure_sheet_template_headers()
        bom_col = self._find_col_by_header('BOM Level')
        if bom_col < 0:
            QMessageBox.warning(self, 'Add SmBOM', 'BOM Level column not found.')
            return
        row_idx = self.table.rowCount()
        self.table.insertRow(row_idx)
        self._init_smbom_row(row_idx)
        self._update_change_type_body_box()
        self._refresh_structure_action_buttons()

    def _init_smbom_row(self, row_idx: int):
        """Set up a BOM L0-style sky-blue SmBOM input row."""
        bom_col = self._find_col_by_header('BOM Level')
        part_col = self._struct_part_col()
        sky = QColor('#87CEEB')
        change_type_options_l0 = ['', 'Revised', 'Change']

        for c in range(self.table.columnCount()):
            if c == 0:  # Select checkbox
                cont = QWidget()
                h_lay = QHBoxLayout(cont)
                h_lay.setContentsMargins(0, 0, 0, 0)
                h_lay.addStretch(1)
                chk = QCheckBox()
                h_lay.addWidget(chk)
                h_lay.addStretch(1)
                cont._chk = chk
                cont.setStyleSheet('QWidget { background-color: #87CEEB; }')
                self.table.setCellWidget(row_idx, c, cont)
            elif c == 1:  # Action dropdown with BOM L0 options
                combo = QComboBox()
                combo.addItems(change_type_options_l0)
                combo.setCurrentIndex(0)
                combo.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                self._bind_change_type_combo(combo, row_idx)
                self.table.setCellWidget(row_idx, c, combo)
            elif c in {2, 3, 4, 5, 6}:
                # BOM L0 rule: only Description(2) and Ref Designator(6) get checkboxes
                if c in {2, 6}:
                    cont_v = QWidget()
                    hv = QHBoxLayout(cont_v)
                    hv.setContentsMargins(0, 0, 0, 0)
                    hv.addStretch(1)
                    chk_v = QCheckBox()
                    hv.addWidget(chk_v)
                    hv.addStretch(1)
                    cont_v._chk = chk_v
                    cont_v.setStyleSheet('QWidget { background-color: #87CEEB; }')
                    chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                    self.table.setCellWidget(row_idx, c, cont_v)
                else:
                    item = QTableWidgetItem('')
                    item.setBackground(sky)
                    self.table.setItem(row_idx, c, item)
            else:
                item = QTableWidgetItem('')
                if c == part_col:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    item.setBackground(QColor('#FFE5CC'))
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setData(Qt.ItemDataRole.UserRole, 'smbom')
                elif c == bom_col:
                    item.setText('0')
                    item.setBackground(sky)
                else:
                    h = self.table.horizontalHeaderItem(c)
                    h_norm = self._norm_header(h.text()) if h else ''
                    if h_norm == 'description':
                        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                    item.setBackground(sky)
                self.table.setItem(row_idx, c, item)

            self._on_change_type_changed(row_idx)

    def _load_smbom_for_row(self, row_idx: int) -> tuple[bool, str]:
        """Fetch 1-level BOM for an SmBOM row, populate it and insert child rows."""
        part_col = self._struct_part_col()
        plant_col = self._find_col_by_header('Plant')
        bom_col = self._find_col_by_header('BOM Level')

        if part_col < 0:
            return False, 'Part column not found.'
        part_item = self.table.item(row_idx, part_col)
        part = (part_item.text() if part_item else '').strip()
        if not self._looks_like_part(part):
            return False, f'Invalid part format: {part!r}'

        # Prevent duplicate SmBOM option insertion if BOM0 part already exists.
        part_upper = part.upper()
        if bom_col >= 0:
            for r in range(self.table.rowCount()):
                if r == row_idx:
                    continue
                b_it = self.table.item(r, bom_col)
                p_it = self.table.item(r, part_col)
                b_txt = (b_it.text() if b_it else '').strip()
                p_txt = ((p_it.text() if p_it else '') or '').lstrip(' \t').strip().upper()
                if b_txt == '0' and p_txt == part_upper:
                    return False, f'SmBOM Option {part} already exists.'

        plant = '4070'
        if plant_col >= 0:
            pit = self.table.item(row_idx, plant_col)
            if pit and (pit.text() or '').strip():
                plant = (pit.text() or '').strip()

        try:
            from implemented_bom_query import fetch_implemented_bom
            records = fetch_implemented_bom([part], max_level=1, plant=plant, include_level0=True)
        except Exception as e:
            return False, f'Failed to query BOM for {part}: {e}'

        l0_recs = [r for r in records if str(r.get('bom_level', '')).strip() == '0']
        l1_recs = [r for r in records if str(r.get('bom_level', '')).strip() == '1']

        # Populate the L0 (SmBOM) row
        if l0_recs:
            self._populate_row_from_where_used_record(row_idx, l0_recs[0])
        # Clear SmBOM marker so subsequent Update won't reload
        if part_item:
            part_item.setData(Qt.ItemDataRole.UserRole, None)
            part_item.setBackground(QColor('#87CEEB'))  # restore sky-blue after load

        # Insert L1 child rows immediately below the SmBOM row
        sky = QColor('#FFFFFF')
        change_type_options_l1 = ['', 'Repl Item at Same Seq', 'Remove Item', 'Change']
        for i, l1_rec in enumerate(l1_recs):
            insert_at = row_idx + 1 + i
            self.table.insertRow(insert_at)
            self._struct_item_change_guard = True
            try:
                for c in range(self.table.columnCount()):
                    if c == 0:
                        cont = QWidget()
                        h_lay = QHBoxLayout(cont)
                        h_lay.setContentsMargins(0, 0, 0, 0)
                        h_lay.addStretch(1)
                        chk = QCheckBox()
                        h_lay.addWidget(chk)
                        h_lay.addStretch(1)
                        cont._chk = chk
                        self.table.setCellWidget(insert_at, c, cont)
                    elif c == 1:
                        combo = QComboBox()
                        combo.addItems(change_type_options_l1)
                        combo.setCurrentIndex(0)
                        combo.setStyleSheet('QComboBox { padding: 2px; }')
                        self._bind_change_type_combo(combo, insert_at)
                        self.table.setCellWidget(insert_at, c, combo)
                    elif c in {2, 3, 4, 5, 6}:
                        cont_v = QWidget()
                        hv = QHBoxLayout(cont_v)
                        hv.setContentsMargins(0, 0, 0, 0)
                        hv.addStretch(1)
                        chk_v = QCheckBox()
                        hv.addWidget(chk_v)
                        hv.addStretch(1)
                        cont_v._chk = chk_v
                        chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                        self.table.setCellWidget(insert_at, c, cont_v)
                    else:
                        h = self.table.horizontalHeaderItem(c)
                        h_norm = self._norm_header(h.text()) if h else ''
                        key = self.DBKEY_BY_HEADER.get(h_norm)
                        value = ''
                        if key:
                            value = '' if l1_rec.get(key) is None else str(l1_rec.get(key)).strip()
                        item = QTableWidgetItem(value)
                        if c == part_col:
                            item.setText('      ' + value.lstrip(' \t') if value.strip() else '')
                            item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                        elif c == bom_col:
                            item.setText('1')
                            item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                        elif h_norm == 'description':
                            item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                        else:
                            item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
                        self.table.setItem(insert_at, c, item)
            finally:
                self._struct_item_change_guard = False

            self._on_change_type_changed(insert_at)

        return True, ''

    def _insert_checkbox_columns(self):
        """Insert/remove 'New ...' columns based on selected checkboxes in columns 2..6."""
        if self.table.columnCount() == 0 or self.table.rowCount() == 0:
            QMessageBox.information(self, 'Insert Columns', 'No Structure Sheet data available.')
            return

        # Determine which selector columns have at least one checked checkbox.
        selected_flags = {
            'description': False,
            'seq': False,
            'qty': False,
            'kit': False,
            'refdes': False,
        }
        col_key_map = {2: 'description', 3: 'seq', 4: 'qty', 5: 'kit', 6: 'refdes'}
        for r in range(self.table.rowCount()):
            for col, key in col_key_map.items():
                w = self.table.cellWidget(r, col)
                if w and hasattr(w, '_chk') and w._chk.isChecked():
                    selected_flags[key] = True

        def _norm(v: str) -> str:
            return (v or '').strip().lower()

        def _find_header_index(name: str) -> int:
            for c in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                if h and _norm(h.text()) == _norm(name):
                    return c
            return -1

        def _find_last_header(possible: set[str]) -> int:
            for c in range(self.table.columnCount() - 1, -1, -1):
                h = self.table.horizontalHeaderItem(c)
                if h and _norm(h.text()) in possible:
                    return c
            return -1

        def _find_bom_col() -> int:
            for c in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                if h and _norm(h.text()) == 'bom level':
                    return c
            return -1

        orange = QBrush(QColor('#FF8C00'))
        insert_specs = [
            ('description', 'New Description', {'description', 'part description'}),
            ('seq', 'New Item Seq', {'item seq', 'seq#', 'seq'}),
            ('qty', 'New Qty', {'ext qty', 'qty'}),
            ('kit', 'New Kit Code', {'kit code'}),
            ('refdes', 'New Ref Designator', {'reference designator', 'ref designator', 'ref designator(rd)'}),
        ]

        inserted = 0
        removed = 0
        bom_col = _find_bom_col()
        for flag_key, new_header, anchors in insert_specs:
            should_exist = selected_flags.get(flag_key, False)
            existing_idx = _find_header_index(new_header)
            if flag_key == 'refdes' and existing_idx < 0:
                existing_idx = _find_header_index('New Reference Designator')

            if should_exist and existing_idx < 0:
                anchor_col = _find_last_header(anchors)
                insert_at = (anchor_col + 1) if anchor_col >= 0 else self.table.columnCount()

                self.table.insertColumn(insert_at)
                h_item = QTableWidgetItem(new_header)
                h_item.setForeground(orange)
                self.table.setHorizontalHeaderItem(insert_at, h_item)


                # Special handling for New Kit Code: use dropdown
                if new_header.lower() == 'new kit code':
                    plant_val = (self.cmb_build_plant.currentText() or '').strip() if hasattr(self, 'cmb_build_plant') else '4070'
                    try:
                        kit_codes = fetch_distinct_kit_codes(plant_val)
                        if not kit_codes:
                            print(f"WARNING: No kit codes found for plant '{plant_val}'. Dropdown will be empty.")
                    except Exception as e:
                        print(f"ERROR fetching kit codes for plant '{plant_val}': {e}")
                        kit_codes = []
                    
                    for r in range(self.table.rowCount()):
                        combo = QComboBox()
                        combo.addItem("")
                        if kit_codes:
                            combo.addItems(kit_codes)
                        combo.setEditable(True)
                        combo.setStyleSheet('QComboBox { padding: 2px; }')
                        selector_col = self.INSERTED_COL_TO_SELECTOR[new_header.lower()]
                        selector_w = self.table.cellWidget(r, selector_col)
                        is_checked = bool(selector_w and hasattr(selector_w, '_chk') and selector_w._chk.isChecked())
                        combo.setEnabled(is_checked)
                        if bom_col >= 0:
                            bom_item = self.table.item(r, bom_col)
                            if bom_item and bom_item.text().strip() == '0':
                                combo.setStyleSheet('QComboBox { background-color: #87CEEB; }')
                        self.table.setCellWidget(r, insert_at, combo)
                else:
                    for r in range(self.table.rowCount()):
                        item = QTableWidgetItem('')
                        item.setForeground(orange)
                        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                        selector_col = self.INSERTED_COL_TO_SELECTOR[new_header.lower()]
                        selector_w = self.table.cellWidget(r, selector_col)
                        is_checked = bool(selector_w and hasattr(selector_w, '_chk') and selector_w._chk.isChecked())
                        self._set_item_editable(item, is_checked)
                        if bom_col >= 0:
                            bom_item = self.table.item(r, bom_col)
                            if bom_item and bom_item.text().strip() == '0':
                                item.setBackground(QColor('#87CEEB'))
                        self.table.setItem(r, insert_at, item)

                inserted += 1

            elif (not should_exist) and existing_idx >= 0:
                self.table.removeColumn(existing_idx)
                removed += 1

        self._update_inserted_editability_all_rows()
        self._apply_inserted_header_colors()
        self._restore_rotated_column_widths()
        self._update_change_type_body_box()

        if inserted == 0 and removed == 0:
            QMessageBox.information(self, 'Insert Columns', 'No column change needed.')
        else:
            QMessageBox.information(self, 'Insert Columns',
                                    f'Inserted {inserted} column(s), removed {removed} column(s).')

    def _reset_struct_table(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self._refresh_structure_action_buttons()

    def append_rows(self, headers:list[str], rows:list[list[str]]):
        if self.table.columnCount()==0:
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
        if self.table.columnCount()!=len(headers):
            norm=[]
            for r in rows:
                r2=(r+['']*self.table.columnCount())[:self.table.columnCount()]
                norm.append(r2)
            rows=norm
        start=self.table.rowCount()
        self.table.setRowCount(start+len(rows))
        for i,row in enumerate(rows):
            for j,val in enumerate(row):
                self.table.setItem(start+i,j,QTableWidgetItem(val))
        self._refresh_structure_action_buttons()

    def _export_structure_sheet(self):
        """Export Structure Sheet to Excel with dropdowns, checkboxes, and formatting preserved."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            if self.table.rowCount() == 0:
                QMessageBox.warning(self, 'Export', 'No data to export.')
                return

            file_path, _ = QFileDialog.getSaveFileName(
                self, 'Save Structure Sheet', '', 'Excel Files (*.xlsx);;All Files (*)'
            )
            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Structure Sheet"

            # Get headers (exclude Select column 0)
            headers = []
            for c in range(1, self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                headers.append(h.text() if h else f'Col{c}')

            # Identify vertical-header columns by name (table cols 2-6, export cols 2-6)
            # Vertical rotation ONLY for the 5 leading columns at fixed positions 2-6
            VERT_EXCEL_COLS = {2, 3, 4, 5, 6}

            # Find Part and BOM Level table columns (table col c == excel col c, Select skipped)
            part_excel_col = -1
            bom_col_table = -1
            desc_excel_col = -1
            for c in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                if h:
                    if h.text().strip().lower() == 'part':
                        part_excel_col = c
                    if 'bom level' in h.text().lower():
                        bom_col_table = c
                    if h.text().strip().lower() == 'description':
                        desc_excel_col = c

            thin = Side(border_style='thin', color='C9E2FF')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_fill = PatternFill(start_color='E1F0FF', end_color='E1F0FF', fill_type='solid')
            grp_fill = PatternFill(start_color='B0D8F5', end_color='B0D8F5', fill_type='solid')
            hdr_font = Font(bold=True, color='0F2D46')

            # Row 1: group label 'Change Type' merged over vertical cols, blank for others
            ws.row_dimensions[1].height = 20
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c_idx, value='')
                cell.fill = hdr_fill
                cell.border = border
            grp_start, grp_end = min(VERT_EXCEL_COLS), max(VERT_EXCEL_COLS)
            ws.merge_cells(start_row=1, start_column=grp_start, end_row=1, end_column=grp_end)
            gc = ws.cell(row=1, column=grp_start, value='Change Type')
            gc.font = hdr_font
            gc.fill = grp_fill
            gc.alignment = Alignment(horizontal='center', vertical='center')

            # Row 2: column headers
            ws.row_dimensions[2].height = 90
            for c_idx, h_text in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=c_idx, value=h_text)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = border
                if c_idx in VERT_EXCEL_COLS:
                    cell.alignment = Alignment(text_rotation=90, horizontal='center',
                                               vertical='bottom', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center',
                                               wrap_text=True)

            col_max_len = {i: len(headers[i - 1]) for i in range(1, len(headers) + 1)}

            # Data validation: one per column (reuse)
            # Action dropdown validation (L0: Revised+Change; others: no Revised)
            dv_l0 = DataValidation(type='list', formula1='"Revised,Change"',
                                   allow_blank=True, showDropDown=False)
            dv_l1 = DataValidation(type='list',
                formula1='"Repl Item at Same Seq,Remove Item,Add Item,Change"',
                allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv_l0)
            ws.add_data_validation(dv_l1)

            sky_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

            for r in range(self.table.rowCount()):
                excel_row = r + 3  # data starts at row 3

                bom_level = ''
                if bom_col_table >= 0:
                    it = self.table.item(r, bom_col_table)
                    if it:
                        bom_level = it.text().strip()

                for c in range(1, self.table.columnCount()):
                    excel_col = c
                    widget = self.table.cellWidget(r, c)
                    item = self.table.item(r, c)

                    if c == 1:  # Action dropdown
                        value = (widget.currentText() if isinstance(widget, QComboBox)
                                 else (item.text() if item else ''))
                        cell = ws.cell(row=excel_row, column=excel_col, value=value)
                        (dv_l0 if bom_level == '0' else dv_l1).add(cell)

                    elif c in VERT_EXCEL_COLS:  # checkbox symbol, no dropdown
                        if widget and hasattr(widget, '_chk'):
                            value = '\u2611' if widget._chk.isChecked() else '\u2610'
                        elif item:
                            value = item.text()
                        else:
                            value = ''
                        cell = ws.cell(row=excel_row, column=excel_col, value=value)

                    else:
                        value = item.text() if item else ''
                        cell = ws.cell(row=excel_row, column=excel_col, value=value)

                    if excel_col == part_excel_col:
                        cell.alignment = Alignment(horizontal='left', vertical='center',
                                                   wrap_text=True)
                    elif excel_col == desc_excel_col:
                        cell.alignment = Alignment(horizontal='left', vertical='center',
                                                   wrap_text=True)
                    elif excel_col in VERT_EXCEL_COLS:
                        cell.alignment = Alignment(horizontal='center', vertical='center',
                                                   wrap_text=False)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center',
                                                   wrap_text=True)

                    if bom_level == '0':
                        cell.fill = sky_fill

                    if excel_col not in VERT_EXCEL_COLS:
                        col_max_len[excel_col] = max(col_max_len.get(excel_col, 0),
                                                     len(str(value) if value else ''))

            # Column widths
            for c_idx in range(1, len(headers) + 1):
                if c_idx in VERT_EXCEL_COLS:
                    ws.column_dimensions[get_column_letter(c_idx)].width = 5
                else:
                    fit_w = min(max(col_max_len.get(c_idx, 8) + 2, 8), 50)
                    ws.column_dimensions[get_column_letter(c_idx)].width = fit_w

            ws.freeze_panes = 'A3'

            save_path = file_path
            while True:
                try:
                    wb.save(save_path)
                    QMessageBox.information(self, 'Export', f'Structure Sheet exported to:\n{save_path}')
                    break
                except PermissionError:
                    QMessageBox.warning(
                        self,
                        'Export Error',
                        'Permission denied while saving the file.\n\n'
                        'Please close the target Excel file (if open) or choose a different folder/file name.'
                    )
                    save_path, _ = QFileDialog.getSaveFileName(
                        self, 'Save Structure Sheet', save_path,
                        'Excel Files (*.xlsx);;All Files (*)'
                    )
                    if not save_path:
                        return

        except Exception as e:
            QMessageBox.warning(self, 'Export Error', f'Error exporting:\n{str(e)}\n\n{traceback.format_exc()}')

    def _import_structure_sheet(self):
        """Import Structure Sheet from Excel, restoring dropdowns, checkboxes, and formatting."""
        try:
            from openpyxl import load_workbook
            
            file_path, _ = QFileDialog.getOpenFileName(
                self, 'Open Structure Sheet', '', 'Excel Files (*.xlsx);;All Files (*)'
            )
            if not file_path:
                return
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Extract headers
            # Row 1 is the group label row; column headers are in row 2.
            headers = []
            for c in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=2, column=c).value
                headers.append(str(cell_value) if cell_value else f'Col{c}')
            
            # Prepend "Select" column to match table structure
            headers = ['Select'] + headers
            
            # Clear and reset table
            self.table.clear()
            self.table.setRowCount(0)
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            
            # Create rotated header for vertical columns (2-6)
            from PyQt6.QtCore import Qt
            hdr = RotatedColumnsHeader(
                Qt.Orientation.Horizontal,
                rotated_columns=range(2, 7),
                parent=self.table,
                group_label='Change Type',
                group_columns=list(range(2, 7)),
            )
            hdr.setStretchLastSection(True)
            hdr.setSectionsClickable(True)
            self.table.setHorizontalHeader(hdr)
            self.table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.horizontalHeader().setMinimumHeight(117)
            self._apply_inserted_header_colors()
            
            # Set column widths for rotated columns
            for vc in range(2, 7):
                self.table.horizontalHeader().setSectionResizeMode(vc, QHeaderView.ResizeMode.Fixed)
                self.table.setColumnWidth(vc, 40)
            
            # Find BOM Level column
            bom_col = -1
            for c, h in enumerate(headers):
                if 'bom level' in h.lower():
                    bom_col = c
                    break
            
            # Load data rows
            for r in range(3, ws.max_row + 1):  # data starts at row 3
                row_data = []
                for c in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=r, column=c).value
                    row_data.append(str(cell_value) if cell_value else '')
                
                self.table.insertRow(self.table.rowCount())
                current_row = self.table.rowCount() - 1
                
                # Get BOM level for this row
                bom_level = row_data[bom_col - 1].strip() if (bom_col > 0 and bom_col <= len(row_data)) else ''
                
                # Set up cells
                for c in range(len(headers)):
                    if c == 0:  # Select column with checkbox
                        cont = QWidget()
                        h_lay = QHBoxLayout(cont)
                        h_lay.setContentsMargins(0, 0, 0, 0)
                        h_lay.addStretch(1)
                        chk = QCheckBox()
                        h_lay.addWidget(chk)
                        h_lay.addStretch(1)
                        cont._chk = chk
                        self.table.setCellWidget(current_row, c, cont)
                    
                    elif c == 1:  # Change Type dropdown
                        combo = QComboBox()
                        # L0: Revised + Change only;  non-zero: no Revised
                        if bom_level == '0':
                            combo.addItems(['', 'Revised', 'Change'])
                        else:
                            combo.addItems(['', 'Repl Item at Same Seq', 'Remove Item', 'Add Item', 'Change'])

                        # Set value from Excel
                        if c - 1 < len(row_data):
                            combo_value = row_data[c - 1]
                            idx = combo.findText(combo_value)
                            if idx >= 0:
                                combo.setCurrentIndex(idx)

                        combo.setStyleSheet('QComboBox { padding: 2px; }')
                        self._bind_change_type_combo(combo, current_row)
                        self.table.setCellWidget(current_row, c, combo)
                    
                    elif c in {2, 3, 4, 5, 6}:  # Checkbox columns
                        cont_v = QWidget()
                        hv = QHBoxLayout(cont_v)
                        hv.setContentsMargins(0, 0, 0, 0)
                        hv.addStretch(1)
                        chk_v = QCheckBox()
                        
                        # Set checked state from Excel
                        if c - 1 < len(row_data):
                            cell_val = row_data[c - 1]
                            chk_v.setChecked(cell_val in ['☑', 'True', '1', 'TRUE'])
                        
                        hv.addWidget(chk_v)
                        hv.addStretch(1)
                        cont_v._chk = chk_v
                        chk_v.stateChanged.connect(self._on_selector_checkbox_toggled)
                        self.table.setCellWidget(current_row, c, cont_v)
                    
                    else:
                        item = QTableWidgetItem(row_data[c - 1] if c - 1 < len(row_data) else '')
                        self.table.setItem(current_row, c, item)
                
                # Apply sky-blue for BOM Level 0
                if bom_level == '0':
                    for c in range(self.table.columnCount()):
                        if c == 0:
                            w = self.table.cellWidget(current_row, c)
                            if w:
                                w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                        elif c == 1:
                            w = self.table.cellWidget(current_row, c)
                            if w:
                                w.setStyleSheet('QComboBox { padding: 2px; background-color: #87CEEB; }')
                        elif c in {2, 3, 4, 5, 6}:
                            w = self.table.cellWidget(current_row, c)
                            if w:
                                w.setStyleSheet('QWidget { background-color: #87CEEB; }')
                            else:
                                cell = self.table.item(current_row, c)
                                if cell:
                                    cell.setBackground(QColor('#87CEEB'))
                        else:
                            cell = self.table.item(current_row, c)
                            if cell:
                                cell.setBackground(QColor('#87CEEB'))
                
                # Initialize checkbox visibility based on Change Type
                self._on_change_type_changed(current_row)
            
            QMessageBox.information(self, 'Import', f'Structure Sheet imported successfully from:\n{file_path}')
            self._update_inserted_editability_all_rows()
            self._apply_inserted_header_colors()
            self._refresh_structure_action_buttons()
        
        except Exception as e:
            QMessageBox.warning(self, 'Import Error', f'Error importing: {str(e)}\n{traceback.format_exc()}')




# ================= Inventory Cost Tab (FULL UPDATED) =================

PLANTS = [4020,4055,4060,4070,4080,4090]
METRICS = ['On Order Qty','Onhand Qty','Gross Demand-13','Gross Demand-26','Gross Demand-52','Standard Cost USD']
NO_CDW_CODES = {
'0070','0080','0110','0120','0130','0170','0180','0210','0243','0250','0251','0260','0261','0280','0288','0289','0290',
'0301','0302','0303','0304','0305','0320','0330','0335','0340','0345','0350','0355','0360','0365','0370','0375','0380','0385','0390','0395',
'0401','0402','0403','0404','0405','0410','0415','0420','0425','0430','0435','0440','0445','0450','0455','0460','0465','0470','0475','0480','0485','0490','0495'
}

class InventoryCostTab(QWidget):
    def reset_tab(self):
        self.df = None
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        self._update_charts()
        self._refresh_report_tab(reset=True)

    def __init__(self):
        super().__init__()
        v = QVBoxLayout(self)
        h = QHBoxLayout()
        self.btn_import_db = QPushButton('Import from Databricks')
        self.btn_export = QPushButton('Export Excel')
        self.btn_reset = QPushButton('Reset Tab')
        for b in (self.btn_import_db, self.btn_export, self.btn_reset):
            h.addWidget(b)
        h.addSpacing(16)
        h.addWidget(QLabel('Cholesterol View:'))
        self.cholesterol_filter = QComboBox()
        self.cholesterol_filter.addItems(['Both', 'Good Cholesterol', 'Bad Cholesterol'])
        self.cholesterol_filter.setCurrentText('Both')
        h.addWidget(self.cholesterol_filter)
        h.addStretch(1)
        v.addLayout(h)

        self.subtabs = QTabWidget()

        self.data_tab = QWidget()
        data_layout = QVBoxLayout(self.data_tab)
        data_layout.setContentsMargins(0, 0, 0, 0)
        self.table = QTableWidget(0,0)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setVisible(True)
        data_layout.addWidget(self.table)

        self.charts_tab = QWidget()
        charts_layout = QVBoxLayout(self.charts_tab)
        charts_layout.setContentsMargins(8, 8, 8, 8)

        self.chart_status = QLabel('Import data to view charts.')
        charts_layout.addWidget(self.chart_status)

        self.cost_canvas = None
        self.demand_canvas = None
        self.cost_fig = None
        self.demand_fig = None
        self.cost_ax = None
        self.demand_ax = None

        if _HAS_MATPLOTLIB:
            self.cost_fig = Figure(figsize=(9, 3.8), tight_layout=True)
            self.cost_canvas = FigureCanvas(self.cost_fig)
            self.cost_ax = self.cost_fig.add_subplot(111)
            charts_layout.addWidget(self.cost_canvas)

            self.demand_fig = Figure(figsize=(9, 3.8), tight_layout=True)
            self.demand_canvas = FigureCanvas(self.demand_fig)
            self.demand_ax = self.demand_fig.add_subplot(111)
            charts_layout.addWidget(self.demand_canvas)
        else:
            charts_layout.addWidget(QLabel('matplotlib is not available in this environment.'))

        self.subtabs.addTab(self.data_tab, 'Table')
        self.subtabs.addTab(self.charts_tab, 'Charts')
        v.addWidget(self.subtabs)
        self.df = None
        self.btn_import_db.clicked.connect(self.import_databricks)
        self.btn_export.clicked.connect(self.export_excel)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.subtabs.currentChanged.connect(self._on_subtab_changed)
        self.cholesterol_filter.currentTextChanged.connect(self._on_cholesterol_filter_changed)

    def _on_cholesterol_filter_changed(self, _value: str):
        if self.df is not None:
            self.render()

    def _compute_cholesterol_category(self, row: pd.Series) -> str:
        try:
            total_inv = float(row.get('Total Onhand', 0) or 0) + float(row.get('Total On Order Quantity', 0) or 0)
            total_dem = (
                float(row.get('Gross Demand-13', 0) or 0)
                + float(row.get('Gross Demand-26', 0) or 0)
                + float(row.get('Gross Demand-52', 0) or 0)
            )
            if total_inv <= 0:
                return ''
            return 'Good Cholesterol' if (total_inv - total_dem) < 0 else 'Bad Cholesterol'
        except Exception:
            return ''

    def _filtered_inventory_df(self) -> pd.DataFrame:
        if self.df is None:
            return pd.DataFrame()
        out = self.df.copy()
        if 'Cholesterol' not in out.columns:
            out['Cholesterol'] = out.apply(self._compute_cholesterol_category, axis=1)
        mode = self.cholesterol_filter.currentText().strip()
        if mode == 'Good Cholesterol':
            out = out[out['Cholesterol'] == 'Good Cholesterol']
        elif mode == 'Bad Cholesterol':
            out = out[out['Cholesterol'] == 'Bad Cholesterol']
        return out

    def _refresh_report_tab(self, reset: bool = False):
        try:
            main = self.window()
            report_tab = getattr(main, 'report_tab', None)
            if report_tab is None:
                return
            if reset and hasattr(report_tab, 'reset_report'):
                report_tab.reset_report()
            elif hasattr(report_tab, 'refresh_report'):
                report_tab.refresh_report()
        except Exception:
            pass

    def _on_subtab_changed(self, idx: int):
        if self.subtabs.tabText(idx) == 'Charts':
            self._update_charts()

    def _update_charts(self):
        if not _HAS_MATPLOTLIB or self.cost_ax is None or self.demand_ax is None:
            return
        try:
            self._render_charts()
        except Exception:
            pass

    def _render_charts(self):
        self.cost_ax.clear()
        self.demand_ax.clear()

        chart_df = self._filtered_inventory_df()
        if chart_df.empty:
            self.chart_status.setText('Import data to view charts.')
            self.cost_ax.set_title('Inventory Cost by Plant')
            self.demand_ax.set_title('Gross Demand-52 by Plant')
            self.cost_canvas.draw()
            self.demand_canvas.draw()
            return

        plants = [str(p) for p in PLANTS]
        cost_values = []
        demand_values = []
        for p in plants:
            onhand_col = f'{p} Onhand Qty'
            onorder_col = f'{p} On Order Qty'
            std_cost_col = f'{p} Standard Cost USD'
            dem52_col = f'{p} Gross Demand-52'

            onhand = pd.to_numeric(chart_df[onhand_col], errors='coerce').fillna(0) if onhand_col in chart_df.columns else pd.Series([0.0] * len(chart_df))
            onorder = pd.to_numeric(chart_df[onorder_col], errors='coerce').fillna(0) if onorder_col in chart_df.columns else pd.Series([0.0] * len(chart_df))
            std_cost = pd.to_numeric(chart_df[std_cost_col], errors='coerce').fillna(0) if std_cost_col in chart_df.columns else pd.Series([0.0] * len(chart_df))
            dem52 = pd.to_numeric(chart_df[dem52_col], errors='coerce').fillna(0) if dem52_col in chart_df.columns else pd.Series([0.0] * len(chart_df))

            # Cost = (on-hand + on-order) * standard cost
            plant_cost = float(((onhand + onorder) * std_cost).sum())
            plant_dem52 = float(dem52.sum())
            cost_values.append(plant_cost)
            demand_values.append(plant_dem52)

        self.cost_ax.bar(plants, cost_values, color='#2E86C1')
        self.cost_ax.set_title('Inventory Cost by Plant')
        self.cost_ax.set_ylabel('USD')
        self.cost_ax.tick_params(axis='x', labelrotation=0)

        self.demand_ax.bar(plants, demand_values, color='#27AE60')
        self.demand_ax.set_title('Gross Demand-52 by Plant')
        self.demand_ax.set_ylabel('Qty')
        self.demand_ax.tick_params(axis='x', labelrotation=0)

        self.chart_status.setText('Charts updated for plant-wise Inventory Cost and Demand-52.')
        self.cost_canvas.draw()
        self.demand_canvas.draw()

    def _build_obs_replacement_map(self) -> Dict[str, str]:
        obs_map: Dict[str, str] = {}
        try:
            main = self.window()
            obs_tab = getattr(main, 'obs_tab', None)
            if not obs_tab or not hasattr(obs_tab, 'table'):
                return obs_map
            t = obs_tab.table
            for r in range(t.rowCount()):
                obs = t.item(r, 1).text().strip() if t.item(r, 1) else ''
                rep = t.item(r, 3).text().strip() if t.item(r, 3) else ''
                if obs:
                    obs_map[obs.upper()] = rep
        except Exception:
            pass
        return obs_map

    def import_databricks(self):
        """Import inventory, demand, and cost data from Databricks for OBS parts."""
        if fetch_inventory_demand_cost is None:
            QMessageBox.warning(self, 'Import Error', 'Failed to import Databricks query module. Check installation.')
            return
        try:
            # Get OBS parts from OBS Tab
            obs_parts = []
            try:
                main = self.window()
                obs_tab = getattr(main, 'obs_tab', None)
                if obs_tab and hasattr(obs_tab, 'table'):
                    t = obs_tab.table
                    for r in range(t.rowCount()):
                        obs = t.item(r, 1)
                        if obs:
                            part = obs.text().strip()
                            if part:
                                obs_parts.append(part)
            except Exception as e:
                QMessageBox.warning(self, 'Error', f'Failed to extract OBS parts: {str(e)}')
                return
            if not obs_parts:
                QMessageBox.information(self, 'No Data', 'No OBS parts found in OBS Tab.')
                return
            # Fetch data from Databricks
            QMessageBox.information(self, 'Loading', f'Fetching data for {len(obs_parts)} OBS part(s) from Databricks...')
            results = fetch_inventory_demand_cost(obs_parts, plants=PLANTS)
            if not results:
                QMessageBox.warning(self, 'No Results', 'No data returned from Databricks.')
                return
            # Build DataFrame similar to import_mm360 structure
            # Define PACE/DASH mapper (same as MM360)
            def get_pace_dash(mrp_profile):
                if mrp_profile is None:
                    return ''
                mrp_str = str(mrp_profile).strip().upper()
                if mrp_str.startswith('SGP'):
                    return 'PACE'
                if mrp_str.startswith('GDS'):
                    return 'DASH'
                return ''
            
            rows = []
            # Normalize Databricks rows for stable key matching (string/int/case safe).
            result_by_part = {}
            obs_map = self._build_obs_replacement_map()
            for rec in results:
                part_key = str(rec.get('part_number', '')).strip().upper()
                if part_key:
                    result_by_part.setdefault(part_key, []).append(rec)

            for pn in obs_parts:
                pn_key = str(pn).strip().upper()
                part_results = result_by_part.get(pn_key, [])
                if not part_results:
                    continue
                # Prefer a row that actually has metadata populated (4020 is often blank).
                first = next(
                    (
                        r for r in part_results
                        if str(r.get('part_description', '')).strip()
                        or str(r.get('mrp_profile', '')).strip()
                        or str(r.get('make_buy', '')).strip()
                    ),
                    part_results[0],
                )
                code4 = str(pn)[:4]
                prim = sec = ''
                if code4 in NO_CDW_CODES:
                    prim = sec = 'No Change Required'
                # Derive PACE/DASH from the first non-empty MRP profile row.
                mrp_row = next(
                    (r for r in part_results if str(r.get('mrp_profile', '')).strip()),
                    first,
                )
                pdsh = get_pace_dash(mrp_row.get('mrp_profile'))
                row = {
                    'Material Number': pn,
                    'Material Description': first.get('part_description', ''),
                    'Replacement Part': obs_map.get(pn_key, ''),
                    'Primary Disposition': prim,
                    'Secondary Disposition': sec,
                    'PACE/DASH': pdsh
                }
                tot_on = tot_oh = tot_d13 = tot_d26 = tot_d52 = tot_cost = 0.0
                plant_map = {
                    str(r.get('plant', '')).strip(): r
                    for r in part_results
                }
                for p in PLANTS:
                    p_key = str(p).strip()
                    plant_data = plant_map.get(p_key)
                    if plant_data:
                        oo = float(plant_data.get('on_order_qty') or 0)
                        oh = float(plant_data.get('on_hand_qty') or 0)
                        d13 = float(plant_data.get('gross_demand_13w') or 0)
                        d26 = float(plant_data.get('gross_demand_26w') or 0)
                        d52 = float(plant_data.get('gross_demand_52w') or 0)
                        ags_oo = float(plant_data.get('ags_on_order_qty') or 0)
                        ags_oh = float(plant_data.get('ags_on_hand_qty') or 0)
                        ags_d52 = float(plant_data.get('ags_gross_demand_52w') or 0)
                        raw_sc = float(plant_data.get('standard_cost_usd') or 0)
                        sc = raw_sc if (oh > 0 or oo > 0) else 0.0
                    else:
                        oo = oh = d13 = d26 = d52 = ags_oo = ags_oh = ags_d52 = sc = 0.0
                    row[f'{p} On Order Qty'] = oo
                    row[f'{p} Onhand Qty'] = oh
                    row[f'{p} Gross Demand-13'] = d13
                    row[f'{p} Gross Demand-26'] = d26
                    row[f'{p} Gross Demand-52'] = d52
                    row[f'{p} AGS On-Order'] = ags_oo
                    row[f'{p} AGS On-Hand'] = ags_oh
                    row[f'{p} AGS 6M Gross Demand'] = ags_d52
                    row[f'{p} Standard Cost USD'] = sc
                    # Only sum non-AGS columns for totals
                    if not ("AGS" in f'{p} On Order Qty' or "AGS" in f'{p} Onhand Qty' or "AGS" in f'{p} Gross Demand-13' or "AGS" in f'{p} Gross Demand-26' or "AGS" in f'{p} Gross Demand-52'):
                        tot_on += oo
                        tot_oh += oh
                        tot_d13 += d13
                        tot_d26 += d26
                        tot_d52 += d52
                        tot_cost += (oh + oo) * sc
                row['Total On Order Quantity'] = tot_on
                row['Total Onhand'] = tot_oh
                row['Gross Demand-13'] = tot_d13
                row['Gross Demand-26'] = tot_d26
                row['Gross Demand-52'] = tot_d52
                row['Inventory Cost'] = tot_cost
                rows.append(row)
            self.df = pd.DataFrame(rows).sort_values('Inventory Cost', ascending=False)
            self.render()
            self._refresh_report_tab()
            try:
                main = self.window()
                watch_tab = getattr(main, 'watch_list_tab', None)
                if watch_tab is not None and hasattr(watch_tab, 'refresh_from_sources'):
                    watch_tab.refresh_from_sources(show_message=False)
            except Exception:
                pass
            QMessageBox.information(self, 'Success', f'Loaded {len(rows)} OBS part(s) from Databricks.')
        except Exception as e:
            QMessageBox.warning(self, 'Import Error', f'Error importing from Databricks: {str(e)}')

    def render(self):
        if self.df is None: return
        source_df = self._filtered_inventory_df()
        if source_df.empty:
            self.table.clear()
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.chart_status.setText('No rows match the selected Cholesterol filter.')
            self._update_charts()
            return

        headers = source_df.columns.tolist()
        summary_cols = [
            "Total On Order Quantity",
            "Total Onhand",
            "Gross Demand-13",
            "Gross Demand-26",
            "Gross Demand-52",
            "Inventory Cost"
        ]
        plant_non_ags_cols = []
        ags_cols = []
        base_cols = []

        # Show AGS columns only for plants where AGS inventory exists.
        ags_visible_plants = set()
        for h in headers:
            m = re.match(r'^(\d{4})\s+AGS (On-Order|On-Hand)$', str(h))
            if not m:
                continue
            plant = m.group(1)
            series = pd.to_numeric(source_df[h], errors='coerce').fillna(0)
            if series.ne(0).any():
                ags_visible_plants.add(plant)

        for h in headers:
            m = re.match(r'^(\d{4})\s+(.+)$', str(h))
            if m:
                plant = m.group(1)
                metric = m.group(2)
                if metric.startswith("AGS"):
                    if plant in ags_visible_plants:
                        ags_cols.append(h)
                else:
                    plant_non_ags_cols.append(h)
            elif str(h).strip() not in summary_cols:
                base_cols.append(h)
        total_cols = [c for c in summary_cols if c in headers]
        ordered_headers = base_cols + plant_non_ags_cols + ags_cols + total_cols
        # Only keep the explicitly ordered columns — no leftover columns after Total.
        view_df = source_df.loc[:, ordered_headers].copy()

        # --- DEMAND MAPPING: compute new columns BEFORE setting table column count ---
        for _dc in ["Excess Inventory", "Excess Cost", "Cholesterol"]:
            view_df[_dc] = None
        for _ri in range(len(view_df)):
            _row = view_df.iloc[_ri]
            try:
                _tot_inv = float(_row.get("Total Onhand", 0) or 0) + float(_row.get("Total On Order Quantity", 0) or 0)
                _tot_gd = (float(_row.get("Gross Demand-13", 0) or 0)
                           + float(_row.get("Gross Demand-26", 0) or 0)
                           + float(_row.get("Gross Demand-52", 0) or 0))
                _excess_inv = _tot_inv - _tot_gd
                # Blank out rows with no inventory (nothing to analyse)
                if _tot_inv == 0:
                    view_df.iat[_ri, view_df.columns.get_loc("Excess Inventory")] = ''
                    view_df.iat[_ri, view_df.columns.get_loc("Excess Cost")] = ''
                    view_df.iat[_ri, view_df.columns.get_loc("Cholesterol")] = ''
                elif _excess_inv < 0:
                    # Demand exceeds inventory — no excess, leave blank
                    view_df.iat[_ri, view_df.columns.get_loc("Excess Inventory")] = ''
                    view_df.iat[_ri, view_df.columns.get_loc("Excess Cost")] = ''
                    # Demand > Inventory → Good Cholesterol (inventory will be consumed)
                    view_df.iat[_ri, view_df.columns.get_loc("Cholesterol")] = "Good Cholesterol"
                else:
                    # Inventory exceeds demand — Bad Cholesterol
                    view_df.iat[_ri, view_df.columns.get_loc("Excess Inventory")] = _excess_inv
                    _cost = float(_row.get("Inventory Cost", 0) or 0)
                    _exc_cost = (_cost / _tot_inv) * _excess_inv if _tot_inv > 0 else 0
                    view_df.iat[_ri, view_df.columns.get_loc("Excess Cost")] = _exc_cost if _exc_cost >= 0 else ''
                    view_df.iat[_ri, view_df.columns.get_loc("Cholesterol")] = "Bad Cholesterol"
            except Exception:
                view_df.iat[_ri, view_df.columns.get_loc("Excess Inventory")] = ''
                view_df.iat[_ri, view_df.columns.get_loc("Excess Cost")] = ''
                view_df.iat[_ri, view_df.columns.get_loc("Cholesterol")] = ''

        self.table.setRowCount(len(view_df))
        self.table.setColumnCount(len(view_df.columns))

        headers = view_df.columns.tolist()
        display_headers = []
        rotated_cols = []
        plant_groups = []
        header_texts = {}  # { col_index: text }

        # Group plant columns, AGS columns, and Total columns for rotated header.
        summary_indices = []
        ags_indices = []
        plant_groups = []
        for idx, name in enumerate(headers):
            m = re.match(r'^(\d{4})\s+(.+)$', str(name))
            if m:
                plant = m.group(1)
                metric = m.group(2)
                display_label = f"{metric} ({plant})"
                display_headers.append(display_label)
                header_texts[idx] = display_label
                rotated_cols.append(idx)
                if metric.startswith("AGS"):
                    ags_indices.append(idx)
                else:
                    if not plant_groups or plant_groups[-1][0] != plant:
                        plant_groups.append((plant, [idx]))
                    else:
                        plant_groups[-1][1].append(idx)
            else:
                display_headers.append(name)
                header_texts[idx] = name
                if name.strip() in summary_cols:
                    summary_indices.append(idx)
        # Remove AGS columns from plant groups if present
        for i, (gname, idxs) in enumerate(plant_groups):
            plant_groups[i] = (gname, [ix for ix in idxs if ix not in ags_indices])
        plant_groups = [(g, idxs) for g, idxs in plant_groups if idxs]
        # Ensure summary columns are rotated
        rotated_cols.extend(i for i in summary_indices if i not in rotated_cols)
        # Insert AGS group before Total group
        if ags_indices:
            plant_groups.append(("AGS", ags_indices))
        # Add a 'Total' group for summary columns
        if summary_indices:
            plant_groups.append(("Total", summary_indices))
        # Add Demand Mapping group (columns already computed and in view_df)
        _dm_indices = [idx for idx, name in enumerate(view_df.columns)
                       if name in ("Excess Inventory", "Excess Cost", "Cholesterol")]
        if _dm_indices:
            plant_groups.append(("Demand Mapping", _dm_indices))


        # Create and install custom rotated header first.
        hdr = RotatedColumnsHeader(Qt.Orientation.Horizontal, rotated_columns=rotated_cols, parent=self.table)
        hdr.set_header_texts(header_texts)
        hdr.set_group_spans(plant_groups)
        self.table.setHorizontalHeader(hdr)
        self.table.horizontalHeader().setVisible(True)

        # Assign labels after installing the header so section/model state is fresh.
        self.table.setHorizontalHeaderLabels(display_headers)

        # Make all 'Total' group headers bold in the UI
        try:
            total_group_indices = []
            for group_name, indices in plant_groups:
                if group_name == "Total":
                    total_group_indices = indices
                    break
            for idx in total_group_indices:
                item = self.table.horizontalHeaderItem(idx)
                if item:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
        except Exception:
            pass

        # Force header to have access to model and update rendering
        hdr.reset()
        hdr.viewport().update()
        self.table.horizontalHeader().setFixedHeight(180)

        # Build per-column group index for background coloring.
        col_to_group_idx = {}  # col_index -> group_palette_index
        for gi, (gname, idxs) in enumerate(plant_groups):
            for ci in idxs:
                col_to_group_idx[ci] = gi
        GROUP_COLORS = [
            QColor('#D6EAF8'), QColor('#D5F5E3'), QColor('#FDEBD0'),
            QColor('#E8DAEF'), QColor('#D6EAF8'), QColor('#D5F5E3'),
            QColor('#FDEBD0'), QColor('#E8DAEF'),
        ]
        AGS_COLOR = QColor('#BDE3F7')
        TOTAL_COLOR = QColor('#D9E1F2')

        for r in range(len(view_df)):
            row_data = view_df.iloc[r]
            for c, col in enumerate(view_df.columns):
                v = row_data[col]
                # Determine numeric value safely.
                try:
                    num_v = float(v)
                except (TypeError, ValueError):
                    num_v = None

                # --- Resolve display text ---
                txt = str(v) if (num_v is None or num_v != 0) else ''

                # AGS 6M Gross Demand: blank if AGS On-Order and On-Hand are both 0.
                if 'AGS 6M Gross Demand' in str(col):
                    ma = re.match(r'^(\d{4}) AGS 6M Gross Demand$', str(col))
                    plant = ma.group(1) if ma else ''
                    try:
                        ags_oh = float(row_data.get(f'{plant} AGS On-Hand', 0) or 0)
                    except (TypeError, ValueError):
                        ags_oh = 0
                    try:
                        ags_oo = float(row_data.get(f'{plant} AGS On-Order', 0) or 0)
                    except (TypeError, ValueError):
                        ags_oo = 0
                    txt = '' if (ags_oh == 0 and ags_oo == 0) else (str(v) if (num_v is None or num_v != 0) else '0')

                # Total-group summary demand: blank if no total inventory.
                elif col in ('Gross Demand-13', 'Gross Demand-26', 'Gross Demand-52'):
                    try:
                        tot_oo = float(row_data.get('Total On Order Quantity', 0) or 0)
                        tot_oh = float(row_data.get('Total Onhand', 0) or 0)
                    except (TypeError, ValueError):
                        tot_oo = tot_oh = 0
                    if tot_oo == 0 and tot_oh == 0:
                        txt = ''
                    elif num_v == 0:
                        txt = '0'

                # Per-plant Gross Demand: blank if plant has no inventory.
                elif 'Gross Demand' in str(col) and re.match(r'^\d{4} Gross Demand-', str(col)):
                    mp = re.match(r'^(\d{4}) Gross Demand-', str(col))
                    if mp:
                        plant = mp.group(1)
                        try:
                            oh = float(row_data.get(f'{plant} Onhand Qty', 0) or 0)
                            oo = float(row_data.get(f'{plant} On Order Qty', 0) or 0)
                        except (TypeError, ValueError):
                            oh = oo = 0
                        if oh == 0 and oo == 0:
                            txt = ''
                        elif num_v == 0:
                            txt = '0'

                # Cost formatting.
                elif num_v is not None and 'Cost' in str(col) and num_v != 0:
                    txt = f'${num_v:,.2f}'

                item = QTableWidgetItem(txt)

                # --- Background coloring per group ---
                if c in col_to_group_idx:
                    gi = col_to_group_idx[c]
                    gname = plant_groups[gi][0]
                    if gname == 'AGS':
                        item.setBackground(QBrush(AGS_COLOR))
                    elif gname == 'Total':
                        item.setBackground(QBrush(TOTAL_COLOR))
                    elif gname == 'Demand Mapping':
                        if col == 'Cholesterol':
                            if txt == 'Bad Cholesterol':
                                item.setBackground(QBrush(QColor(220, 50, 50)))
                                item.setForeground(QBrush(QColor(255, 255, 255)))
                            elif txt == 'Good Cholesterol':
                                item.setBackground(QBrush(QColor(50, 180, 50)))
                                item.setForeground(QBrush(QColor(255, 255, 255)))
                            else:
                                item.setBackground(QBrush(QColor('#FFF2CC')))
                        else:
                            item.setBackground(QBrush(QColor('#FFF2CC')))
                    else:
                        item.setBackground(QBrush(GROUP_COLORS[gi % len(GROUP_COLORS)]))

                self.table.setItem(r, c, item)


        # Resize and apply fixed widths to rotated columns
        self.table.resizeColumnsToContents()
        for c in rotated_cols:
            self.table.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)
            col_name = str(view_df.columns[c])
            if col_name.strip() in ("Total Gross Demand-13", "Total Gross Demand-26", "Total Gross Demand-52"):
                self.table.setColumnWidth(c, 36)
            elif col_name.strip() == "Inventory Cost":
                # Increase Inventory Cost column width by 10% (additional, now 21% over base)
                base_width = 50
                self.table.setColumnWidth(c, int(base_width * 1.21))
            elif 'Cost' in col_name:
                self.table.setColumnWidth(c, 50)
            else:
                self.table.setColumnWidth(c, 30)

        # Reduce width and wrap text for 'Total On Order Quantity' header
        try:
            idx = view_df.columns.get_loc('Total On Order Quantity')
            # Match width to 'Total Onhand' column if present, else use 44
            try:
                idx_onhand = view_df.columns.get_loc('Total Onhand')
                width = self.table.columnWidth(idx_onhand)
            except Exception:
                width = 44
            self.table.setColumnWidth(idx, width)
            item = self.table.horizontalHeaderItem(idx)
            if item:
                item.setText('Total\nOn Order\nQuantity')
        except Exception:
            pass

        # Ensure only non-rotated cost columns are widened for readability.
        for c, col_name in enumerate(view_df.columns):
            if 'Cost' not in str(col_name):
                continue
            if c in rotated_cols:
                continue
            self.table.resizeColumnToContents(c)
            self.table.setColumnWidth(c, max(self.table.columnWidth(c) + 5, 44))  # 110 * 0.4 (-60%)

        self._update_charts()

    def export_excel(self):
        if self.df is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, 'Export', 'inventory_cost.xlsx', 'Excel Files (*.xlsx)'
        )
        if not path:
            return
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import pandas as pd

        df = self.df.copy()
        summary_headers = [
            "Total On Order Quantity",
            "Total Onhand",
            "Gross Demand-13",
            "Gross Demand-26",
            "Gross Demand-52",
            "Inventory Cost"
        ]

        # Reorder export columns: base -> plant(non-AGS) -> AGS -> Total
        export_headers = df.columns.tolist()
        export_plant_non_ags = []
        export_ags = []
        export_base = []
        for h in export_headers:
            m = re.match(r'^(\d{4})\s+(.+)$', str(h))
            if m:
                metric = m.group(2)
                if metric.startswith("AGS"):
                    export_ags.append(h)
                else:
                    export_plant_non_ags.append(h)
            elif str(h).strip() not in summary_headers:
                export_base.append(h)
        export_total = [c for c in summary_headers if c in export_headers]
        ordered_export_headers = export_base + export_plant_non_ags + export_ags + export_total
        ordered_export_set = set(ordered_export_headers)
        ordered_export_headers.extend([h for h in export_headers if h not in ordered_export_set])
        df = df.loc[:, ordered_export_headers]

        # Blank AGS demand when AGS inventory is zero.
        for col in list(df.columns):
            m = re.match(r'^(\d{4})\s+AGS 6M Gross Demand$', str(col))
            if not m:
                continue
            plant = m.group(1)
            oh_col = f"{plant} AGS On-Hand"
            oo_col = f"{plant} AGS On-Order"
            if oh_col in df.columns and oo_col in df.columns:
                mask = (
                    pd.to_numeric(df[oh_col], errors='coerce').fillna(0).eq(0)
                    & pd.to_numeric(df[oo_col], errors='coerce').fillna(0).eq(0)
                )
                df.loc[mask, col] = ''
        if 'Replacement Part' not in df.columns:
            obs_map = self._build_obs_replacement_map()
            df.insert(2, 'Replacement Part', '')
            for i, pn in enumerate(df['Material Number']):
                key = str(pn).strip().upper()
                if key in obs_map:
                    df.at[i, 'Replacement Part'] = obs_map[key]

        with pd.ExcelWriter(path, engine='openpyxl') as w:
            df.to_excel(w, index=False, sheet_name='Output Sheet', startrow=1)
        wb = load_workbook(path)
        ws = wb['Output Sheet']

        header_fill = PatternFill('solid', fgColor='BDD7EE')
        ws.row_dimensions[2].height = 118

        # Format row 2 headers
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(2, c)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            header_name = str(cell.value or '')
            is_plant_metric = bool(re.match(r'^\d{4} ', header_name))
            if not is_plant_metric:
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True, text_rotation=90)


        # ---- Group plant metric columns in row 1 using plant code ----
        groups = {}
        total_group_cols = []
        ags_group_cols = []
        demand_mapping_cols = []
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(2, c).value or '')
            m = re.match(r'^(\d{4})\s+', val)
            if m:
                code = m.group(1)
                metric = val.split(' ', 1)[1] if ' ' in val else ''
                if metric.startswith('AGS'):
                    ws.cell(1, c).value = 'AGS'
                    ags_group_cols.append(c)
                else:
                    ws.cell(1, c).value = code
                    groups.setdefault(code, []).append(c)
            elif val.strip() in summary_headers:
                total_group_cols.append(c)
            elif val.strip() in ["Excess Inventory", "Excess Cost", "Cholesterol"]:
                demand_mapping_cols.append(c)

        # Merge each plant's metric columns in row 1
        for code, cols in groups.items():
            if not code:
                continue
            if len(cols) > 1:
                ws.merge_cells(start_row=1, start_column=cols[0], end_row=1, end_column=cols[-1])
            cell = ws.cell(1, cols[0])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # Merge summary columns under 'Total' in row 1
        if ags_group_cols:
            ws.merge_cells(start_row=1, start_column=ags_group_cols[0], end_row=1, end_column=ags_group_cols[-1])
            cell = ws.cell(1, ags_group_cols[0])
            cell.value = 'AGS'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # Merge summary columns under 'Total' in row 1
        if total_group_cols:
            ws.merge_cells(start_row=1, start_column=total_group_cols[0], end_row=1, end_column=total_group_cols[-1])
            cell = ws.cell(1, total_group_cols[0])
            cell.value = 'Total'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            # Make all 'Total' group headers bold in Excel (row 2)
            for c in total_group_cols:
                cell2 = ws.cell(2, c)
                cell2.font = Font(bold=True)

        # Merge Demand Mapping columns under 'Demand Mapping' in row 1
        if demand_mapping_cols:
            if len(demand_mapping_cols) > 1:
                ws.merge_cells(start_row=1, start_column=demand_mapping_cols[0], end_row=1, end_column=demand_mapping_cols[-1])
            cell = ws.cell(1, demand_mapping_cols[0])
            cell.value = 'Demand Mapping'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # Apply currency formatting ($ with 2 decimals) to all cost columns
        cost_columns = [
            c for c in range(1, ws.max_column + 1)
            if 'Cost' in str(ws.cell(2, c).value or '')
        ]
        for r in range(3, ws.max_row + 1):
            for c in cost_columns:
                cell = ws.cell(r, c)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0.00'

        # Column widths
        for c in range(1, ws.max_column + 1):
            header_name = str(ws.cell(2, c).value or '')
            col_letter = get_column_letter(c)
            if re.match(r'^\d{4} ', header_name):
                ws.column_dimensions[col_letter].width = 4
                continue
            # Increase Inventory Cost column width by another 10% (21% total over base)
            if header_name.strip() == "Inventory Cost":
                max_len = 0
                for r in range(1, ws.max_row + 1):
                    v = ws.cell(r, c).value
                    if v:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = int(max(10, max_len + 2) * 1.21)
                continue
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, c).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = max(10, max_len + 2)

        # Blank zero values
        for r in range(3, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).value == 0:
                    ws.cell(r, c).value = ''

        wb.save(path)
        QMessageBox.information(self, 'Export', 'Excel exported successfully')


class ReportTab(QWidget):
    def __init__(self):
        super().__init__()
        outer = QVBoxLayout(self)

        actions = QHBoxLayout()
        self.btn_refresh = QPushButton('Refresh Report')
        self.btn_reset = QPushButton('Reset Report')
        self.status_label = QLabel('Generate Inventory_Cost data to prepare the report.')
        self.status_label.setWordWrap(True)
        actions.addWidget(self.btn_refresh)
        actions.addWidget(self.btn_reset)
        actions.addWidget(self.status_label, 1)
        outer.addLayout(actions)

        self.report_text = QTextEdit()
        self.report_text.setReadOnly(True)
        self.report_text.setPlaceholderText('Generate Inventory_Cost data to prepare the report narrative.')
        outer.addWidget(self.report_text)

        self.df = pd.DataFrame(columns=REPORT_COLUMNS)
        self.btn_refresh.clicked.connect(lambda: self.refresh_report(show_messages=True))
        self.btn_reset.clicked.connect(self.reset_report)

    def reset_report(self):
        self.df = pd.DataFrame(columns=REPORT_COLUMNS)
        self.report_text.clear()
        self.status_label.setText('Generate Inventory_Cost data to prepare the report.')

    def refresh_report(self, show_messages: bool = False):
        if build_supply_chain_report is None:
            msg = 'Report helper module is unavailable.'
            self.status_label.setText(msg)
            if show_messages:
                QMessageBox.warning(self, 'Report', msg)
            return

        main = self.window()
        inventory_tab = getattr(main, 'inventory_cost_tab', None)
        inventory_df = getattr(inventory_tab, 'df', None)
        if inventory_df is None or inventory_df.empty:
            self.reset_report()
            return

        obs_parts = [
            str(value).strip().upper()
            for value in inventory_df.get('Material Number', pd.Series(dtype=str)).tolist()
            if str(value).strip()
        ]
        whereused_tab = getattr(main, 'whereused_tab', None)
        where_used_records = getattr(whereused_tab, 'raw_databricks_records', None) or []
        warnings = []

        po_details = []
        if fetch_open_purchase_order_details is not None:
            try:
                po_details = fetch_open_purchase_order_details(obs_parts, plants=PLANTS)
            except Exception as exc:
                warnings.append(f'PO details unavailable: {exc}')
        else:
            warnings.append('PO details unavailable.')

        kit_descriptions = {}
        kit_codes = sorted(
            {
                str(record.get('kit_code', '')).strip().upper()
                for record in where_used_records
                if str(record.get('kit_code', '')).strip()
            }
        )
        if kit_codes and fetch_kit_code_descriptions is not None:
            try:
                where_used_plant = next(
                    (
                        str(record.get('plant', '')).strip()
                        for record in where_used_records
                        if str(record.get('plant', '')).strip()
                    ),
                    '',
                )
                kit_descriptions = fetch_kit_code_descriptions(kit_codes, plant=where_used_plant or None)
            except Exception as exc:
                warnings.append(f'Kit descriptions unavailable: {exc}')
        elif not where_used_records:
            warnings.append('Where Used data not loaded; outsourced recommendations may be incomplete.')

        try:
            self.df = build_supply_chain_report(
                inventory_df=inventory_df,
                where_used_records=where_used_records,
                po_details=po_details,
                kit_descriptions=kit_descriptions,
                plants=PLANTS,
            )
        except Exception as exc:
            self.status_label.setText(f'Report generation failed: {exc}')
            if show_messages:
                QMessageBox.warning(self, 'Report', f'Failed to build report:\n{exc}')
            return

        self.render()
        if self.df.empty:
            status_text = 'No actionable report rows were generated from the current Inventory_Cost data.'
        else:
            action_count = self.df['Recommended Action'].nunique() if 'Recommended Action' in self.df.columns else 0
            status_text = f'Generated {len(self.df)} report row(s) across {action_count} action type(s).'
        if warnings:
            status_text = status_text + ' ' + ' '.join(warnings)
        self.status_label.setText(status_text)
        if show_messages and warnings:
            QMessageBox.information(self, 'Report', status_text)

    def render(self):
        lines: list[str] = []
        sl_col_width = 12

        def _sl_line(sl_no: int | str, text: str) -> str:
            prefix = f"{sl_no}." if isinstance(sl_no, int) else str(sl_no)
            return f"{prefix:<{sl_col_width}}{text}"

        sl_no = 1
        if self.df is not None and not self.df.empty:
            # Group by Part only, ignore Plant.
            grouped = self.df.sort_values(['Part', 'Recommended Action']).groupby('Part', dropna=False)
            for part, group in grouped:
                part = str(part or '').strip()
                if not part:
                    continue

                first_row = group.iloc[0]
                total_inv = self._fmt_num(first_row.get('Total Inventory', ''))
                total_demand = self._fmt_num(first_row.get('Total Demand', ''))
                total_excess = self._fmt_num(first_row.get('Excess Quantity', ''))
                summary = f"{part} / {total_inv} - {total_demand} = {total_excess}"

                rec_lines: list[str] = []
                for _, row in group.iterrows():
                    action = str(row.get('Recommended Action', '') or '').strip()
                    if action == 'Cancel PO':
                        po_number = str(row.get('PO Number', '') or '').strip()
                        supplier = str(row.get('Supplier Name', '') or '').strip()
                        qty = self._fmt_num(row.get('PO Qty', ''))
                        delivery = self._fmt_delivery(row.get('Delivery Date', ''))
                        rec_lines.append(f"{po_number} / {supplier} / {qty} / {delivery} - Map the Excess Qty, with PO Cancellation")
                    elif action == 'Sell back to CM':
                        exc_qty = self._fmt_num(row.get('Excess Quantity', ''))
                        kit_code = str(row.get('Kit Code', '') or '').strip()
                        rec_lines.append(f"{exc_qty} -> Sell Back Opportunity({kit_code})")

                if not rec_lines:
                    rec_lines.append('Provide Disposition Plan for this Excess Inventory.')

                lines.append(_sl_line(sl_no, f"{summary}: {rec_lines[0]}"))
                for extra in rec_lines[1:]:
                    lines.append(_sl_line('', extra))
                sl_no += 1
        else:
            lines.append(_sl_line(sl_no, 'No actionable recommendations are available for the current selection.'))
            sl_no += 1

        main = self.window()
        watch_tab = getattr(main, 'watch_list_tab', None)
        watch_entries = self._collect_watch_list_entries(watch_tab)
        if watch_entries:
            lines.append('')
            lines.append(_sl_line(sl_no, 'Add Below Parts to Watch List'))
            sl_no += 1
            for part, reason in watch_entries:
                if reason:
                    lines.append(_sl_line('', f"{part} - {reason}"))
                else:
                    lines.append(_sl_line('', part))

        header = f"{'Sl. No,':<{sl_col_width}}Recommendation"
        body_text = '\n'.join(lines).rstrip()
        html_out = (
            "<pre style=\"font-family:'Consolas','Courier New',monospace;font-size:12pt;margin:0;\">"
            f"<b>{html.escape(header)}</b>\n"
            f"{html.escape(body_text)}"
            "</pre>"
        )
        self.report_text.setHtml(html_out)

    def _collect_watch_list_entries(self, watch_tab: Any) -> list[tuple[str, str]]:
        """Read visible Watch_List rows (table first, dataframe fallback) for report output."""
        if watch_tab is None:
            return []

        reason_headers = {'reason to add in watch list', 'watch list rule'}
        part_headers = {'material number', 'part', 'part number'}

        def _norm(v: Any) -> str:
            return str(v or '').strip()

        entries: list[tuple[str, str]] = []
        seen_parts: set[str] = set()

        # 1) Prefer reading from the Watch_List table so report reflects exactly what user sees.
        table = getattr(watch_tab, 'table', None)
        if isinstance(table, QTableWidget) and table.columnCount() > 0:
            col_names = [_norm(table.horizontalHeaderItem(c).text() if table.horizontalHeaderItem(c) else '') for c in range(table.columnCount())]
            part_idx = next((i for i, name in enumerate(col_names) if name.lower() in part_headers), None)
            reason_idx = next((i for i, name in enumerate(col_names) if name.lower() in reason_headers), None)

            if part_idx is not None:
                for r in range(table.rowCount()):
                    part_item = table.item(r, part_idx)
                    part = _norm(part_item.text() if part_item else '')
                    if not part:
                        continue
                    part_key = part.upper()
                    if part_key in seen_parts:
                        continue
                    seen_parts.add(part_key)

                    reason = ''
                    if reason_idx is not None:
                        reason_item = table.item(r, reason_idx)
                        reason = _norm(reason_item.text() if reason_item else '')
                    entries.append((part, reason))

        # 2) Fallback to dataframe if table didn't provide entries.
        if entries:
            return entries

        watch_df = getattr(watch_tab, 'df', None)
        if isinstance(watch_df, pd.DataFrame) and not watch_df.empty:
            reason_col = 'Reason to add in Watch List'
            part_col = 'Material Number'
            if reason_col not in watch_df.columns and 'Watch List Rule' in watch_df.columns:
                reason_col = 'Watch List Rule'
            if part_col not in watch_df.columns:
                for candidate in ('Part', 'Part Number'):
                    if candidate in watch_df.columns:
                        part_col = candidate
                        break

            for _, row in watch_df.iterrows():
                part = _norm(row.get(part_col, ''))
                if not part:
                    continue
                part_key = part.upper()
                if part_key in seen_parts:
                    continue
                seen_parts.add(part_key)
                reason = _norm(row.get(reason_col, ''))
                entries.append((part, reason))

        return entries

    def _sentence_from_row(self, row: pd.Series, include_supplier: bool = True) -> str:
        part = str(row.get('Part', '') or '').strip()
        plant = str(row.get('Plant', '') or '').strip()
        action = str(row.get('Recommended Action', '') or '').strip()
        excess = self._fmt_num(row.get('Excess Quantity', ''))
        po_number = str(row.get('PO Number', '') or '').strip()
        po_qty = self._fmt_num(row.get('PO Qty', ''))
        supplier = str(row.get('Supplier Name', '') or '').strip()
        delivery = self._fmt_delivery(row.get('Delivery Date', ''))
        destination = str(row.get('Destination Plant', '') or '').strip()
        transfer_qty = self._fmt_num(row.get('Transfer Quantity', ''))
        kit_code = str(row.get('Kit Code', '') or '').strip()
        kit_desc = str(row.get('Kit Description', '') or '').strip()

        supplier_text = f' Supplier {supplier}.' if (include_supplier and supplier) else ''
        if action == 'Cancel PO':
            sent = (
                f'Part {part} in plant {plant} has excess {excess}; cancel PO {po_number} for quantity {po_qty} with delivery date {delivery}.'
            )
            if supplier_text:
                sent = sent.rstrip('.') + '.' + supplier_text
            return sent
        if action == 'Move inventory to AGS':
            return (
                f'Part {part} in plant {plant} has excess {excess}; move {transfer_qty} inventory to AGS and align with demand. '
                f'Target delivery date reference is {delivery}.'
            )
        if action == 'Move inventory':
            return (
                f'Part {part} in plant {plant} has excess {excess}; move {transfer_qty} inventory to plant {destination} to balance demand. '
                f'Target delivery date reference is {delivery}.'
            )
        if action == 'Sell back to CM':
            return (
                f'Part {part} in plant {plant} has excess {excess}; kit code {kit_code} ({kit_desc}) is outsourced, so sell back inventory to CM. '
                f'Delivery date reference is {delivery}.'
            )
        details = str(row.get('Action Details', '') or '').strip()
        base = f'Part {part} in plant {plant} has excess {excess}; recommended action is {action}.'
        if details:
            base += f' {details}'
        if delivery:
            base += f' Delivery date reference is {delivery}.'
        return base

    @staticmethod
    def _fmt_num(value: Any) -> str:
        try:
            if value is None or value == '' or pd.isna(value):
                return '0'
            return f'{float(value):,.3f}'
        except Exception:
            text = str(value or '').strip()
            return text if text else '0'

    @staticmethod
    def _fmt_delivery(value: Any) -> str:
        if value is None or value == '' or (isinstance(value, float) and pd.isna(value)):
            return 'N/A'
        try:
            dt = pd.to_datetime(value, errors='coerce')
            if pd.isna(dt):
                text = str(value).strip()
                return text if text else 'N/A'
            return dt.strftime('%Y-%m-%d')
        except Exception:
            text = str(value).strip()
            return text if text else 'N/A'

class PlaceholderTab(QWidget):
    def __init__(self, title: str):
        super().__init__()
        l = QVBoxLayout(self)
        l.addWidget(QLabel(f"{title} – UI under development"))



class OrphanOBSSubTab(QWidget):
    def __init__(self):
        super().__init__()
        self.fixed_bom_level = None
        self.suppress_import_tool_comments = False
        outer = QVBoxLayout(self)


        btn_row = QHBoxLayout()
        # Move BOM Level and Plant to the left
        lbl_bom = QLabel("BOM Level (1-18):")
        self.bom_level_input = QLineEdit()
        self.bom_level_input.setFixedWidth(45)
        self.bom_level_input.setText("6")
        self.bom_level_input.setPlaceholderText("1-18")
        self.bom_level_input.setToolTip("Enter the maximum BOM depth (1 to 18)")
        lbl_plant = QLabel("Plant:")
        self.plant_combo = QComboBox()
        self.plant_combo.addItems(["4020", "4055", "4060", "4070", "4080", "4090"])
        self.plant_combo.setCurrentText("4070")
        self.plant_combo.setFixedWidth(70)
        self.plant_combo.setToolTip("Plant code to filter the BOM query")

        btn_row.addWidget(lbl_bom)
        btn_row.addWidget(self.bom_level_input)
        btn_row.addWidget(lbl_plant)
        btn_row.addWidget(self.plant_combo)

        self.btn_import = QPushButton("Import BOM of OBS Parts")
        btn_row.addWidget(self.btn_import)

        # --- OEM Retain/Remove radio buttons ---
        oem_label = QLabel("<b>OEM:</b>")
        oem_label.setStyleSheet("font-size:13px; margin-right:2px; margin-left:8px;")
        self.oem_radio_group = QButtonGroup(self)
        self.oem_radio_retain = QRadioButton("Retain")
        self.oem_radio_remove = QRadioButton("Remove")
        self.oem_radio_group.addButton(self.oem_radio_retain)
        self.oem_radio_group.addButton(self.oem_radio_remove)
        self.oem_radio_remove.setChecked(True)
        # Reduce spacing between radio buttons
        # OEM radio group tightly spaced
        oem_radio_container = QWidget()
        oem_radio_layout = QHBoxLayout(oem_radio_container)
        oem_radio_layout.setContentsMargins(0, 0, 0, 0)
        oem_radio_layout.setSpacing(2)
        oem_radio_layout.addWidget(oem_label)
        oem_radio_layout.addWidget(self.oem_radio_retain)
        oem_radio_layout.addWidget(self.oem_radio_remove)
        btn_row.addWidget(oem_radio_container)

        btn_row.addStretch(1)
        self.btn_reset = QPushButton("Reset")
        self.btn_copy_removed = QPushButton("Copy Removed BOM Items")
        btn_row.addWidget(self.btn_copy_removed)
        btn_row.addWidget(self.btn_reset)
        outer.addLayout(btn_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self._import_bom_with_oem_logic)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.btn_copy_removed.clicked.connect(self.copy_removed_child_parts)

    def set_fixed_bom_level(self, level: int, tooltip: str | None = None):
        self.fixed_bom_level = level
        self.bom_level_input.setText(str(level))
        self.bom_level_input.setReadOnly(True)
        self.bom_level_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.bom_level_input.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.bom_level_input.setStyleSheet("background-color:#F3F4F6; color:#555555;")
        if tooltip:
            self.bom_level_input.setToolTip(tooltip)

    def _get_requested_bom_level(self):
        if self.fixed_bom_level is not None:
            self.bom_level_input.setText(str(self.fixed_bom_level))
            return self.fixed_bom_level

        level_text = self.bom_level_input.text().strip()
        if not level_text:
            QMessageBox.warning(self, 'BOM Level Required', 'Please enter BOM Level (1 to 18).')
            return None
        try:
            max_level = int(level_text)
        except ValueError:
            QMessageBox.warning(self, 'Invalid BOM Level', f'BOM Level must be a number from 1 to 18, got: "{level_text}"')
            return None
        if max_level < 1 or max_level > 18:
            QMessageBox.warning(self, 'Invalid BOM Level', f'BOM Level must be between 1 and 18, got: {max_level}')
            return None
        return max_level

    def _import_bom_repl_parts(self):
        # Ensure all listed OBS parts have an explicit replacement entry before import.
        main = self.window()
        obs_tab = getattr(main, 'obs_tab', None)
        missing_replacements = []
        if obs_tab and hasattr(obs_tab, 'table'):
            t = obs_tab.table
            for r in range(t.rowCount()):
                obs_it = t.item(r, 1)
                rep_it = t.item(r, 3)
                obs_val = (obs_it.text() if obs_it else '').strip()
                rep_val = (rep_it.text() if rep_it else '').strip()
                if obs_val and not rep_val:
                    missing_replacements.append(obs_val)
        if missing_replacements:
            preview = ', '.join(missing_replacements[:10])
            suffix = '' if len(missing_replacements) <= 10 else f' ... (+{len(missing_replacements) - 10} more)'
            QMessageBox.warning(
                self,
                'Update Replacements Required',
                'Please update the Replacement column for OBS Parts before using Import BOM of REPL Parts.\n\n'
                f'Missing replacements for: {preview}{suffix}'
            )
            return

        # Collect Replacement parts from the OBS Parts tab (column 3)
        main = self.window()
        obs_tab = getattr(main, 'obs_tab', None)
        repl_parts = []
        if obs_tab and hasattr(obs_tab, 'table'):
            t = obs_tab.table
            for r in range(t.rowCount()):
                it = t.item(r, 3)
                val = (it.text() if it else '').strip()
                if val and not re.search(r'no\s*replacement', val, re.IGNORECASE):
                    repl_parts.append(val)
        if not repl_parts:
            QMessageBox.warning(self, 'No Replacement Parts', 'Replacement column do not have any part numbers listed.')
            return
        max_level = self._get_requested_bom_level()
        if max_level is None:
            return
        plant = self.plant_combo.currentText().strip() or '4070'
        self.import_bom_for_repl_parts(repl_parts, max_level, plant)
        self._remove_oem_rows_if_needed()

    def import_bom_for_repl_parts(self, repl_parts, max_level, plant):
        if not repl_parts:
            QMessageBox.warning(self, 'No Replacement Parts', 'No replacement parts provided.')
            return
        try:
            import sys as _sys
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from implemented_bom_query import fetch_implemented_bom  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(
                self,
                'Module Not Found',
                f'implemented_bom_query.py could not be imported:\n{exc}'
            )
            return

        try:
            records = fetch_implemented_bom(repl_parts, max_level=max_level, plant=plant, include_level0=True)
        except Exception as exc:
            QMessageBox.warning(self, 'SAP Import Error', str(exc))
            return

        if not records:
            QMessageBox.information(
                self,
                'No Data',
                f'No BOM data found in SAP for {len(repl_parts)} replacement part(s) at plant {plant}.'
            )
            return

        headers = [
            'Select',
            'BOM Level',
            'Part',
            'Tool comments',
            'Rev/Ln',
            'Plant',
            'Description',
            'Item Status',
            'Base Qty',
            'Ext Qty',
            'UOM',
            'ECO Number',
            'Procurement Type',
            'Effectivity Date',
            'User Item Type',
            'Item Seq',
            'Kit Code',
            'Designator',
            'Option Class',
            'BOM Source Type (Debug)',
        ]
        self.table.clear()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(records))

        key_by_header = {
            'BOM Level': 'bom_level',
            'Part': 'part',
            'Tool comments': 'tool_comments',
            'Rev/Ln': 'rev_ln',
            'Plant': 'plant',
            'Description': 'description',
            'Item Status': 'item_status',
            'Base Qty': 'base_qty',
            'Ext Qty': 'ext_qty',
            'UOM': 'uom',
            'ECO Number': 'eco_number',
            'Procurement Type': 'procurement_type',
            'Effectivity Date': 'effectivity_date',
            'User Item Type': 'user_item_type',
            'Item Seq': 'item_seq',
            'Kit Code': 'kit_code',
            'Designator': 'designator',
            'Option Class': 'option_class',
            'BOM Source Type (Debug)': 'bom_source_type',
        }

        bom_tbl_idx = headers.index('BOM Level')
        part_tbl_idx = headers.index('Part')
        tc_tbl_idx = headers.index('Tool comments')

        for r, rec in enumerate(records):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            self.table.setCellWidget(r, 0, cont)

            for c, hdr in enumerate(headers[1:], start=1):
                db_key = key_by_header.get(hdr, '')
                val = rec.get(db_key, '')

                if hdr == 'Tool comments' and self.suppress_import_tool_comments:
                    val = ''

                # Fallbacks for legacy query outputs where these aliases are not present.
                if hdr == 'Designator' and (val is None or str(val).strip() == ''):
                    val = rec.get('sparable_flag', '')
                elif hdr == 'BOM Source Type (Debug)' and (val is None or str(val).strip() == ''):
                    val = rec.get('pace_or_dash', '')

                self.table.setItem(r, c, QTableWidgetItem(str(val) if val is not None else ''))

            # Part hierarchy indent using BOM level (same visual style as Where Used).
            bom_item_for_indent = self.table.item(r, bom_tbl_idx)
            part_item = self.table.item(r, part_tbl_idx)
            if part_item is not None:
                raw_part = (part_item.text() or '').strip()
                try:
                    level_int = int(float((bom_item_for_indent.text() if bom_item_for_indent else '0') or '0'))
                except (ValueError, TypeError):
                    level_int = 0
                part_item.setText(('      ' * max(level_int, 0)) + raw_part)

            bom_item = self.table.item(r, bom_tbl_idx)
            bom_val = bom_item.text() if bom_item else ''
            self._apply_comment_and_row_style(
                r,
                bom_val,
                tc_tbl_idx,
                suppress_tool_comments=self.suppress_import_tool_comments,
            )

        self.table.resizeColumnsToContents()
        QMessageBox.information(
            self,
            'Import Complete',
            f'Imported {len(records)} BOM row(s) from SAP for {len(repl_parts)} replacement part(s) at level {max_level}.'
        )

    def _remove_oem_rows_if_needed(self):
        if not self.oem_radio_remove.isChecked():
            return

        part_col = self.find_column('Part')
        bom_col = self.find_column('BOM Level')
        if part_col < 0 or bom_col < 0:
            return

        rows_to_remove = []
        for r in range(self.table.rowCount()):
            part_item = self.table.item(r, part_col)
            bom_item = self.table.item(r, bom_col)
            if not part_item or not bom_item:
                continue

            part = part_item.text().strip()
            bom_level = bom_item.text().strip()
            try:
                is_not_zero = float(bom_level) != 0.0
            except Exception:
                is_not_zero = False

            part_clean = part.replace(' ', '').lstrip()
            prefix = part_clean[:4]
            is_oem = prefix.isdigit() and int(prefix) >= 500
            if is_oem and is_not_zero:
                rows_to_remove.append(r)

        for r in reversed(rows_to_remove):
            self.table.removeRow(r)

    def _import_bom_with_oem_logic(self):
        # First populate the table via import
        self.import_bom()
        self._remove_oem_rows_if_needed()

    def import_bom(self):
        max_level = self._get_requested_bom_level()
        if max_level is None:
            return

        plant = self.plant_combo.currentText().strip() or '4070'
        self.import_from_sap(max_level=max_level, plant=plant)

    def _is_bom_level_zero(self, bom_val: str) -> bool:
        txt = (bom_val or '').strip()
        if txt == '':
            return False
        try:
            return float(txt) == 0.0
        except ValueError:
            return txt == '0'

    def _apply_comment_and_row_style(self, row_idx: int, bom_val: str, tc_tbl_idx: int, suppress_tool_comments: bool = False):
        if self._is_bom_level_zero(bom_val):
            for c in range(self.table.columnCount()):
                cell = self.table.item(row_idx, c)
                if cell:
                    cell.setBackground(QColor('#87CEEB'))  # Sky Blue
            return

        if suppress_tool_comments:
            self.table.setItem(row_idx, tc_tbl_idx, QTableWidgetItem(''))
            return

        it = QTableWidgetItem('Removed BOM Item')
        f = it.font()
        f.setBold(True)
        it.setFont(f)
        it.setForeground(QColor('orange'))
        self.table.setItem(row_idx, tc_tbl_idx, it)

    def _collect_obs_parts(self) -> list[str]:
        main = self.window()
        obs_tab = getattr(main, 'obs_tab', None)
        if obs_tab is None or not hasattr(obs_tab, 'table'):
            return []

        parts = []
        seen = set()
        t = obs_tab.table
        for r in range(t.rowCount()):
            it = t.item(r, 1)  # OBS Parts column
            part = (it.text() if it else '').strip()
            key = part.upper()
            if part and key not in seen:
                seen.add(key)
                parts.append(part)
        return parts

    def import_from_sap(self, max_level: int, plant: str):
        obs_parts = self._collect_obs_parts()
        if not obs_parts:
            QMessageBox.warning(
                self,
                'No OBS Parts',
                'OBS Parts tab is empty. Please add OBS parts before importing from SAP.'
            )
            return

        try:
            import sys as _sys
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from implemented_bom_query import fetch_implemented_bom  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(
                self,
                'Module Not Found',
                f'implemented_bom_query.py could not be imported:\n{exc}'
            )
            return

        try:
            records = fetch_implemented_bom(obs_parts, max_level=max_level, plant=plant, include_level0=True)
        except Exception as exc:
            QMessageBox.warning(self, 'SAP Import Error', str(exc))
            return

        if not records:
            QMessageBox.information(
                self,
                'No Data',
                f'No BOM data found in SAP for {len(obs_parts)} OBS part(s) at plant {plant}.'
            )
            return

        headers = [
            'Select',
            'BOM Level',
            'Part',
            'Tool comments',
            'Rev/Ln',
            'Plant',
            'Description',
            'Item Status',
            'Base Qty',
            'Ext Qty',
            'UOM',
            'ECO Number',
            'Procurement Type',
            'Effectivity Date',
            'User Item Type',
            'Item Seq',
            'Kit Code',
            'Designator',
            'Option Class',
            'BOM Source Type (Debug)',
        ]
        self.table.clear()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(records))

        key_by_header = {
            'BOM Level': 'bom_level',
            'Part': 'part',
            'Tool comments': 'tool_comments',
            'Rev/Ln': 'rev_ln',
            'Plant': 'plant',
            'Description': 'description',
            'Item Status': 'item_status',
            'Base Qty': 'base_qty',
            'Ext Qty': 'ext_qty',
            'UOM': 'uom',
            'ECO Number': 'eco_number',
            'Procurement Type': 'procurement_type',
            'Effectivity Date': 'effectivity_date',
            'User Item Type': 'user_item_type',
            'Item Seq': 'item_seq',
            'Kit Code': 'kit_code',
            'Designator': 'designator',
            'Option Class': 'option_class',
            'BOM Source Type (Debug)': 'bom_source_type',
        }

        bom_tbl_idx = headers.index('BOM Level')
        part_tbl_idx = headers.index('Part')
        tc_tbl_idx = headers.index('Tool comments')

        for r, rec in enumerate(records):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            self.table.setCellWidget(r, 0, cont)

            for c, hdr in enumerate(headers[1:], start=1):
                db_key = key_by_header.get(hdr, '')
                val = rec.get(db_key, '')

                if hdr == 'Tool comments' and self.suppress_import_tool_comments:
                    val = ''

                # Fallbacks for legacy query outputs where these aliases are not present.
                if hdr == 'Designator' and (val is None or str(val).strip() == ''):
                    val = rec.get('sparable_flag', '')
                elif hdr == 'BOM Source Type (Debug)' and (val is None or str(val).strip() == ''):
                    val = rec.get('pace_or_dash', '')

                self.table.setItem(r, c, QTableWidgetItem(str(val) if val is not None else ''))

            # Part hierarchy indent using BOM level (same visual style as Where Used).
            bom_item_for_indent = self.table.item(r, bom_tbl_idx)
            part_item = self.table.item(r, part_tbl_idx)
            if part_item is not None:
                raw_part = (part_item.text() or '').strip()
                try:
                    level_int = int(float((bom_item_for_indent.text() if bom_item_for_indent else '0') or '0'))
                except (ValueError, TypeError):
                    level_int = 0
                part_item.setText(('      ' * max(level_int, 0)) + raw_part)

            bom_item = self.table.item(r, bom_tbl_idx)
            bom_val = bom_item.text() if bom_item else ''
            self._apply_comment_and_row_style(
                r,
                bom_val,
                tc_tbl_idx,
                suppress_tool_comments=self.suppress_import_tool_comments,
            )

        self.table.resizeColumnsToContents()
        QMessageBox.information(
            self,
            'Import Complete',
            f'Imported {len(records)} BOM row(s) from SAP for {len(obs_parts)} OBS part(s) at level {max_level}.'
        )

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def find_column(self, header):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text() == header:
                return i
        return -1

    def select_oems(self):
        part_col = self.find_column('Part')
        bom_col = self.find_column('BOM Level')
        if part_col < 0 or bom_col < 0:
            return

        eligible = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, part_col).text() if self.table.item(r, part_col) else '').lstrip()
            bom = (self.table.item(r, bom_col).text() if self.table.item(r, bom_col) else '').strip()
            prefix = part[:4]
            if bom != '0' and prefix.isdigit() and int(prefix) >= 500:
                eligible.append(r)

        should_select = any(
            not self.table.cellWidget(r, 0).findChild(QCheckBox).isChecked()
            for r in eligible
        )

        for r in eligible:
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            chk.setChecked(should_select)

    def delete_selected(self):
        rows = []
        for r in range(self.table.rowCount()):
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            if chk.isChecked():
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)

    def copy_removed_child_parts(self):
        part_col = self.find_column('Part')
        tc_col = self.find_column('Tool comments')

        if part_col < 0 or tc_col < 0:
            QMessageBox.warning(self, 'Column Missing',
                                'Required columns (Part / Tool comments) not found.')
            return

        unique_parts = set()

        for r in range(self.table.rowCount()):
            tc_item = self.table.item(r, tc_col)
            if not tc_item or not tc_item.text().strip():
                continue

            part_item = self.table.item(r, part_col)
            if not part_item:
                continue

            part = part_item.text().replace(' ', '').strip()
            if part:
                unique_parts.add(part)

        if not unique_parts:
            QMessageBox.information(self, 'No Data',
                                    'No removed child parts found to copy.')
            return

        from PyQt6.QtGui import QGuiApplication
        result = '\n'.join(sorted(unique_parts))
        QGuiApplication.clipboard().setText(result)

        QMessageBox.information(self, 'Copied',
                                f'Copied {len(unique_parts)} unique removed child part(s) to clipboard.')


class WURemovedBOMItemsTab(QWidget):

    def _find_obs_col(self, name: str) -> int:
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return -1
        t = self.obs_provider.table
        for i in range(t.columnCount()):
            h = t.horizontalHeaderItem(i)
            if h and h.text().strip().lower() == name.strip().lower():
                return i
        return -1

    def _ensure_obs_proposed_replacement_column(self) -> int:
        """Ensure OBS table has 'Proposed Replacement' immediately after 'Replacement'."""
        if not self.compare_with_replacement:
            return -1
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return -1

        t = self.obs_provider.table
        existing_idx = self._find_obs_col('Proposed Replacement')
        replacement_idx = self._find_obs_col('Replacement')

        # If already present and already placed after Replacement, reuse it.
        if existing_idx >= 0 and replacement_idx >= 0 and existing_idx == (replacement_idx + 1):
            if hasattr(self.obs_provider, 'update_proposed_replacement_note_visibility'):
                self.obs_provider.update_proposed_replacement_note_visibility(True)
            return existing_idx

        insert_at = replacement_idx + 1 if replacement_idx >= 0 else t.columnCount()
        if existing_idx >= 0 and existing_idx != insert_at:
            t.removeColumn(existing_idx)
            if existing_idx < insert_at:
                insert_at -= 1
            existing_idx = -1

        if existing_idx < 0:
            t.insertColumn(insert_at)
            t.setHorizontalHeaderItem(insert_at, QTableWidgetItem('Proposed Replacement'))
            for r in range(t.rowCount()):
                if t.item(r, insert_at) is None:
                    t.setItem(r, insert_at, QTableWidgetItem(''))

        # Apply a wider width for better visibility (20 chars equivalent)
        try:
            px = _excel_width_to_px(t, 20)
            t.setColumnWidth(insert_at, px)
        except Exception:
            pass
        if hasattr(self.obs_provider, 'update_proposed_replacement_note_visibility'):
            self.obs_provider.update_proposed_replacement_note_visibility(True)
        return insert_at

    @staticmethod
    def _norm_desc(value: str) -> str:
        return re.sub(r'\s+', ' ', (value or '').strip()).upper()

    def _collect_provider_blocks(self, provider) -> Dict[str, List[Dict[str, str]]]:
        """
        Returns BOM block rows keyed by root part.
        Each child row has: part, desc, desc_norm.
        """
        if provider is None:
            return {}
        part_col = provider.find_column('Part')
        bom_col = provider.find_column('BOM Level')
        desc_col = provider.find_column('Description')
        if part_col < 0 or bom_col < 0 or desc_col < 0:
            return {}

        table = provider.table
        blocks: Dict[str, List[Dict[str, str]]] = {}
        current_root = ''
        for r in range(table.rowCount()):
            part_item = table.item(r, part_col)
            bom_item = table.item(r, bom_col)
            desc_item = table.item(r, desc_col)

            part = self._normalize_bom_part(part_item.text() if part_item else '')
            if not part:
                continue
            level = self._parse_bom_level_value(bom_item.text() if bom_item else '')
            if level is None:
                continue
            desc = (desc_item.text() if desc_item else '').strip()

            if level == 0:
                current_root = part.upper()
                blocks.setdefault(current_root, [])
                continue

            if not current_root:
                continue

            blocks.setdefault(current_root, []).append({
                'part': part.upper(),
                'desc': desc,
                'desc_norm': self._norm_desc(desc),
            })

        return blocks

    def _collect_obs_parts_set(self) -> set[str]:
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return set()
        t = self.obs_provider.table
        part_col = self._find_obs_col('OBS Parts')
        if part_col < 0:
            part_col = 1
        out = set()
        for r in range(t.rowCount()):
            it = t.item(r, part_col)
            val = (it.text() if it else '').strip().upper()
            if val:
                out.add(val)
        return out

    def _build_proposed_replacement_map_for_orphans(self, orphan_parts: set[str]) -> Dict[str, str]:
        """
        Build Proposed Replacement values per orphan part using BOM block + Description match:
        OBS block child Description -> matching REPL block child part(s), excluding OBS-listed parts.
        """
        if not self.compare_with_replacement:
            return {}
        if not orphan_parts:
            return {}

        obs_map = self._build_obs_replacement_map()
        obs_blocks = self._collect_provider_blocks(self.imp_bom_provider)
        repl_blocks = self._collect_provider_blocks(self.replacement_bom_provider)
        if not obs_blocks or not repl_blocks:
            return {}

        obs_parts_set = self._collect_obs_parts_set()
        proposed: Dict[str, List[str]] = {}

        for obs_root, obs_children in obs_blocks.items():
            repl_root = (obs_map.get(obs_root) or '').strip().upper()
            if not repl_root or re.search(r'no\s*replacement', repl_root, re.IGNORECASE):
                continue
            repl_children = repl_blocks.get(repl_root)
            if not repl_children:
                continue

            repl_by_desc: Dict[str, List[str]] = {}
            for rec in repl_children:
                desc_norm = rec.get('desc_norm', '')
                part = rec.get('part', '').strip().upper()
                if not desc_norm or not part:
                    continue
                if part in obs_parts_set:
                    continue
                bucket = repl_by_desc.setdefault(desc_norm, [])
                if part not in bucket:
                    bucket.append(part)

            for rec in obs_children:
                orphan_part = rec.get('part', '').strip().upper()
                if orphan_part not in orphan_parts:
                    continue
                desc_norm = rec.get('desc_norm', '')
                if not desc_norm:
                    continue
                matches = repl_by_desc.get(desc_norm, [])
                if not matches:
                    continue
                out = proposed.setdefault(orphan_part, [])
                for m in matches:
                    if m not in out:
                        out.append(m)

        return {k: ', '.join(v) for k, v in proposed.items()}

    def _is_obs_orphan_row(self, row: int, change_col: int) -> bool:
        """Return True when the OBS table row represents an orphan (identified by the
        'Identified Orphans' column, falling back to the Change column for rows written
        before that column existed)."""
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return False
        t = self.obs_provider.table
        # Primary check: dedicated 'Identified Orphans' column
        id_orphan_col = self._find_obs_col('Identified Orphans')
        if id_orphan_col >= 0:
            it = t.item(row, id_orphan_col)
            return bool(it and (it.text() or '').strip().lower().startswith('orphan'))
        # Legacy fallback: Change column starts with 'orphan'
        w = t.cellWidget(row, change_col)
        if isinstance(w, QComboBox):
            return (w.currentText() or '').strip().lower().startswith('orphan')
        it = t.item(row, change_col)
        return bool(it and (it.text() or '').strip().lower().startswith('orphan'))

    def _update_obs_proposed_replacement_rows(self, proposed_by_part: Dict[str, str], orphan_parts: set[str]):
        if not proposed_by_part and not orphan_parts:
            return
        proposed_col = self._ensure_obs_proposed_replacement_column()
        if proposed_col < 0:
            return
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return

        t = self.obs_provider.table
        part_col = self._find_obs_col('OBS Parts')
        if part_col < 0:
            part_col = 1
        change_col = self._find_obs_col('Change')
        if change_col < 0:
            change_col = 2

        for r in range(t.rowCount()):
            pit = t.item(r, part_col)
            part_key = (pit.text() if pit else '').strip().upper()
            if not part_key or part_key not in orphan_parts:
                continue
            if not self._is_obs_orphan_row(r, change_col):
                continue

            value = proposed_by_part.get(part_key, '')
            it = t.item(r, proposed_col)
            if it is None:
                it = QTableWidgetItem('')
                t.setItem(r, proposed_col, it)
            if (it.text() or '') != value:
                it.setText(value)

    def _ensure_obs_identified_orphans_column(self) -> int:
        """Ensure OBS table has 'Identified Orphans' column immediately after 'Replacement'
        (or after 'Proposed Replacement' when that column is also present)."""
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return -1
        t = self.obs_provider.table
        existing_idx = self._find_obs_col('Identified Orphans')
        # Determine desired position: after 'Proposed Replacement' if it exists, else after 'Replacement'.
        anchor_col = self._find_obs_col('Proposed Replacement')
        if anchor_col < 0:
            anchor_col = self._find_obs_col('Replacement')
        insert_at = anchor_col + 1 if anchor_col >= 0 else t.columnCount()

        if existing_idx >= 0 and existing_idx == insert_at:
            return existing_idx

        if existing_idx >= 0 and existing_idx != insert_at:
            t.removeColumn(existing_idx)
            if existing_idx < insert_at:
                insert_at -= 1
            existing_idx = -1

        if existing_idx < 0:
            t.insertColumn(insert_at)
            t.setHorizontalHeaderItem(insert_at, QTableWidgetItem('Identified Orphans'))
            for r in range(t.rowCount()):
                if t.item(r, insert_at) is None:
                    t.setItem(r, insert_at, QTableWidgetItem(''))
            try:
                px = _excel_width_to_px(t, 16)
                t.setColumnWidth(insert_at, px)
            except Exception:
                pass
        return insert_at

    def _get_orphan_parent_change_value(self, part: str) -> str:
        """Look up the immediate level-1 parent of *part* in the WU table and return
        its Change value ('Obsolete' or 'Inactivate') from the OBS Parts table.
        Falls back to 'Obsolete' when no parent mapping is found."""
        pcol = self._find_col('Part')
        wucol = self._find_col('WU Level')
        if pcol < 0 or wucol < 0:
            return 'Obsolete'
        part_upper = part.strip().upper()
        obs_change_map = self._build_obs_change_map()  # {PART_UPPER: 'Obsolete'/'Inactivate'}
        row_count = self.table.rowCount()
        for r in range(row_count):
            wu_it = self.table.item(r, wucol)
            wu_val = (wu_it.text() if wu_it else '').strip()
            # Find a WU Level 0 row matching this part
            if wu_val not in ('0', '0.0'):
                continue
            pit = self.table.item(r, pcol)
            if not pit or pit.text().strip().upper() != part_upper:
                continue
            # Found the block root row for this part; scan for WU level 1 parents
            block_end = row_count
            for i in range(r + 1, row_count):
                nxt = self.table.item(i, wucol)
                if nxt and nxt.text().strip() in ('0', '0.0'):
                    block_end = i
                    break
            for i in range(r + 1, block_end):
                wu_p = self.table.item(i, wucol)
                if not wu_p or wu_p.text().strip() not in ('1', '1.0'):
                    continue
                parent_pit = self.table.item(i, pcol)
                parent_part = (parent_pit.text().lstrip() if parent_pit else '').strip().upper()
                if parent_part in obs_change_map:
                    chg = obs_change_map[parent_part].strip()
                    if chg in ('Obsolete', 'Inactivate'):
                        return chg
        return 'Obsolete'

    def _set_obs_change_value(self, row: int, value: str):
        """Ensure Change column has a combo box and set it to value.
        Allowed values are '', 'Obsolete', 'Inactivate'."""
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return
        t = self.obs_provider.table
        change_col = self._find_obs_col('Change')
        if change_col < 0:
            change_col = 2

        cw = t.cellWidget(row, change_col)
        if not isinstance(cw, QComboBox):
            # Remove any stale text item and restore dropdown behavior.
            t.setItem(row, change_col, QTableWidgetItem(''))
            cw = QComboBox()
            cw.addItems(['', 'Obsolete', 'Inactivate'])
            t.setCellWidget(row, change_col, cw)

        idx = cw.findText(value)
        if idx < 0:
            idx = 0
        cw.setCurrentIndex(idx)
        cw.setStyleSheet('')

    def append_orphans_to_obs_parts(self, remove_from_orphan_table=False, include_orphan_parent=False):
        if not self.obs_provider:
            return
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphans List')
        if pcol < 0 or ocol < 0:
            return
        orphan_map = {}
        orphan_rows = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, pcol)
            oit = self.table.item(r, ocol)
            if not (pit and oit):
                continue
            orphan_label = oit.text().strip()
            orphan_label_low = orphan_label.lower()
            if not orphan_label_low.startswith('orphan'):
                continue
            # Keep Orphan Parent identification in WU table, but do not append it
            # to OBS list unless explicitly requested.
            if (not include_orphan_parent) and orphan_label_low == 'orphan parent':
                continue
            orphan_map[pit.text().strip().upper()] = orphan_label
            orphan_rows.append(r)
        if not orphan_map:
            return
        orphan_parts = set(orphan_map.keys())
        proposed_by_part = self._build_proposed_replacement_map_for_orphans(orphan_parts)
        proposed_col = self._ensure_obs_proposed_replacement_column()
        id_orphan_col = self._ensure_obs_identified_orphans_column()
        t = self.obs_provider.table
        # Refresh column indices after possible insertions above
        change_col = self._find_obs_col('Change')
        if change_col < 0:
            change_col = 2
        existing = {t.item(r,1).text().strip().upper() for r in range(t.rowCount()) if t.item(r,1) and t.item(r,1).text().strip()}
        added_items = []
        for part, lvl in orphan_map.items():
            if part in existing:
                # Refresh 'Identified Orphans' label, Change value, and proposed replacement
                # for existing orphan rows.
                for rr in range(t.rowCount()):
                    pit = t.item(rr, 1)
                    if not pit or (pit.text() or '').strip().upper() != part:
                        continue
                    if not self._is_obs_orphan_row(rr, change_col):
                        # Also check if this row previously had the label in Change (legacy)
                        id_col_check = self._find_obs_col('Identified Orphans')
                        if id_col_check >= 0:
                            id_it = t.item(rr, id_col_check)
                            if not (id_it and (id_it.text() or '').strip().lower().startswith('orphan')):
                                continue
                        else:
                            continue
                    # Update 'Identified Orphans' column
                    if id_orphan_col >= 0:
                        id_it = t.item(rr, id_orphan_col)
                        if id_it is None:
                            id_it = QTableWidgetItem('')
                            t.setItem(rr, id_orphan_col, id_it)
                        id_it.setText(lvl)
                        _bold_f = id_it.font(); _bold_f.setBold(True); id_it.setFont(_bold_f)
                        col = get_orphan_color(lvl)
                        if col:
                            id_it.setForeground(col)
                    # Keep Change blank for Orphan Parent; otherwise set Obsolete/Inactivate.
                    if lvl.strip().lower() == 'orphan parent':
                        self._set_obs_change_value(rr, '')
                    else:
                        change_val = self._get_orphan_parent_change_value(part)
                        self._set_obs_change_value(rr, change_val)
                    # Update proposed replacement
                    if proposed_col >= 0:
                        val = proposed_by_part.get(part, '')
                        it = t.item(rr, proposed_col)
                        if it is None:
                            it = QTableWidgetItem('')
                            t.setItem(rr, proposed_col, it)
                        it.setText(val)
                continue
            target = next((r for r in range(t.rowCount()) if not t.item(r,1) or not t.item(r,1).text().strip()), None)
            if target is None:
                target = t.rowCount(); t.setRowCount(target+1); t._init_rows(target,target+1)
            t.setItem(target, 1, QTableWidgetItem(part))
            # 'Identified Orphans' column: set the orphan label with colour and bold
            if id_orphan_col >= 0:
                id_it = QTableWidgetItem(lvl)
                _bold_f = id_it.font(); _bold_f.setBold(True); id_it.setFont(_bold_f)
                col = get_orphan_color(lvl)
                if col:
                    id_it.setForeground(col)
                t.setItem(target, id_orphan_col, id_it)
            # Change column: keep blank for Orphan Parent, otherwise set Obsolete/Inactivate.
            if lvl.strip().lower() == 'orphan parent':
                self._set_obs_change_value(target, '')
            else:
                change_val = self._get_orphan_parent_change_value(part)
                self._set_obs_change_value(target, change_val)
            replacement_text = '' if self.prompt_replacement_update_for_orphans else 'OBSOLETE: NO REPLACEMENT'
            t.setItem(target, 3, QTableWidgetItem(replacement_text))
            if proposed_col >= 0:
                proposed_val = proposed_by_part.get(part, '')
                t.setItem(target, proposed_col, QTableWidgetItem(proposed_val))
            added_items.append((part, lvl))
        if remove_from_orphan_table:
            for r in reversed(orphan_rows): self.table.removeRow(r)

        # Always refresh relevant orphan rows using the latest analysis output.
        self._update_obs_proposed_replacement_rows(proposed_by_part, orphan_parts)

        if self.prompt_replacement_update_for_orphans and added_items:
            prompt_lines = [f"- {p} ({lvl})" for p, lvl in added_items[:12]]
            remaining = len(added_items) - len(prompt_lines)
            if remaining > 0:
                prompt_lines.append(f"- ... and {remaining} more item(s)")
            QMessageBox.information(
                self,
                'Action Required: Update Replacement Parts',
                'Orphan items were appended to OBS Parts with blank Replacement values.\n\n'
                'Please update the Replacement Part for these Orphan items in OBS Parts and run "Execute Orphan Item Review" again.\n'
                'Repeat this cycle until no new Orphan items are found.\n\n'
                + '\n'.join(prompt_lines)
            )

    def __init__(self, obs_provider, imp_bom_provider=None):
        super().__init__()
        self.obs_provider = obs_provider
        self.imp_bom_provider = imp_bom_provider
        self.replacement_bom_provider = None
        self.compare_with_replacement = False
        self.prompt_replacement_update_for_orphans = False
        outer = QVBoxLayout(self)

        # ── Controls rows (split to avoid overlap on narrower screens) ───────
        top_row = QHBoxLayout()
        top_row.setSpacing(8)
        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(10)
        wu_lbl = QLabel("WU Level (1-6):")
        self.wu_level_input = QLineEdit()
        self.wu_level_input.setFixedWidth(45)
        self.wu_level_input.setText("1")
        self.wu_level_input.setPlaceholderText("1-6")
        self.wu_level_input.setToolTip("Enter the maximum Where Used depth (1 to 6)")

        plant_lbl = QLabel("Plant:")
        self.plant_combo = QComboBox()
        self.plant_combo.addItems(["4020", "4055", "4060", "4070", "4080", "4090"])
        self.plant_combo.setCurrentText("4070")
        self.plant_combo.setFixedWidth(70)
        self.plant_combo.setToolTip("Plant code to filter the WU query")

        self.btn_import = QPushButton("Import WU - Rem Items")
        self.btn_analyze = QPushButton("Run Orphan Analysis")
        self.btn_append_selected = QPushButton("Append Orphan Parents to OBS List")
        self.btn_reset = QPushButton("Reset")

        # Keep button labels readable when the row is dense.
        self.btn_import.setMinimumWidth(170)
        self.btn_analyze.setMinimumWidth(165)
        self.btn_append_selected.setMinimumWidth(255)

        def _add_parent_radio_group(layout, label_text, group_attr, retain_attr, remove_attr):
            grp_box = QWidget()
            grp_layout = QHBoxLayout(grp_box)
            grp_layout.setContentsMargins(0, 0, 0, 0)
            grp_layout.setSpacing(6)

            lbl = QLabel(f"<b>{label_text}</b>")
            grp = QButtonGroup(self)
            grp.setExclusive(True)
            rb_retain = QRadioButton("Retain")
            rb_remove = QRadioButton("Remove")
            # Add to group FIRST so the group tracks state correctly,
            # then set the default checked state.
            grp.addButton(rb_retain, 0)
            grp.addButton(rb_remove, 1)
            rb_remove.setChecked(True)   # default: Remove

            grp_layout.addWidget(lbl)
            grp_layout.addWidget(rb_retain)
            grp_layout.addWidget(rb_remove)
            layout.addWidget(grp_box)

            setattr(self, group_attr, grp)
            setattr(self, retain_attr, rb_retain)
            setattr(self, remove_attr, rb_remove)

        top_row.addWidget(wu_lbl)
        top_row.addWidget(self.wu_level_input)
        top_row.addWidget(plant_lbl)
        top_row.addWidget(self.plant_combo)
        top_row.addWidget(self.btn_import)
        top_row.addWidget(self.btn_analyze)
        top_row.addWidget(self.btn_append_selected)
        top_row.addStretch(1)
        self.btn_reset.setMinimumWidth(110)
        top_row.addWidget(self.btn_reset)
        outer.addLayout(top_row)

        _add_parent_radio_group(bottom_row, "9024 Parents:", 'radio_9024_group', 'radio_9024_retain', 'radio_9024_remove')
        _add_parent_radio_group(bottom_row, "ESW Parents:", 'radio_esw_group', 'radio_esw_retain', 'radio_esw_remove')
        _add_parent_radio_group(
            bottom_row,
            "SmBOM above Config:",
            'radio_above_cfg_group',
            'radio_above_cfg_retain',
            'radio_above_cfg_remove',
        )
        bottom_row.addStretch(1)
        outer.addLayout(bottom_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self.import_from_databricks_for_removed_items)
        self.btn_analyze.clicked.connect(self.perform_orphan_analysis)
        self.btn_append_selected.clicked.connect(self.append_orphan_parents_to_obs_list)
        self.btn_reset.clicked.connect(self.reset_tab)

    def enable_with_replacement_compare(self, replacement_bom_provider):
        self.replacement_bom_provider = replacement_bom_provider
        self.compare_with_replacement = replacement_bom_provider is not None

    def configure_orphan_append_behavior(self, prompt_replacement_updates: bool = False):
        self.prompt_replacement_update_for_orphans = bool(prompt_replacement_updates)

    def append_orphan_parents_to_obs_list(self):
        if not self.obs_provider:
            return

        pcol = self._find_col('Part')
        ocol = self._find_col('Orphans List')
        if pcol < 0:
            QMessageBox.information(self, 'Append Orphan Parents', "Couldn't find a 'Part' column.")
            return

        orphan_parents = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, pcol)
            if not pit:
                continue
            part = pit.text().strip()
            if not part:
                continue

            label = ''
            if ocol >= 0:
                oit = self.table.item(r, ocol)
                label = (oit.text() if oit else '').strip()
            if label.lower() == 'orphan parent':
                orphan_parents.append((part.upper(), label))

        if not orphan_parents:
            QMessageBox.information(self, 'Append Orphan Parents', 'No identified Orphan Parent parts found to append.')
            return

        t = self.obs_provider.table
        id_orphan_col = self._ensure_obs_identified_orphans_column()
        change_col = self._find_obs_col('Change')
        if change_col < 0:
            change_col = 2

        existing = {
            t.item(r,1).text().strip().upper()
            for r in range(t.rowCount())
            if t.item(r,1) and t.item(r,1).text().strip()
        }

        added = 0
        for part, lvl in orphan_parents:
            if part in existing:
                # Refresh existing row: keep Change blank and set marker in Identified Orphans.
                for rr in range(t.rowCount()):
                    pit = t.item(rr, 1)
                    if not pit or (pit.text() or '').strip().upper() != part:
                        continue
                    if id_orphan_col >= 0:
                        id_it = t.item(rr, id_orphan_col)
                        if id_it is None:
                            id_it = QTableWidgetItem('')
                            t.setItem(rr, id_orphan_col, id_it)
                        id_it.setText(lvl)
                        _bold_f = id_it.font(); _bold_f.setBold(True); id_it.setFont(_bold_f)
                        col = get_orphan_color(lvl)
                        if col:
                            id_it.setForeground(col)
                    self._set_obs_change_value(rr, '')
                continue

            target = next((r for r in range(t.rowCount()) if not t.item(r,1) or not t.item(r,1).text().strip()), None)
            if target is None:
                target = t.rowCount(); t.setRowCount(target+1); t._init_rows(target,target+1)
            t.setItem(target,1,QTableWidgetItem(part))

            # Keep Change column blank for Orphan Parent entries.
            self._set_obs_change_value(target, '')

            if id_orphan_col >= 0:
                id_it = QTableWidgetItem(lvl)
                _bold_f = id_it.font(); _bold_f.setBold(True); id_it.setFont(_bold_f)
                col = get_orphan_color(lvl)
                if col:
                    id_it.setForeground(col)
                t.setItem(target, id_orphan_col, id_it)

            replacement_text = '' if self.prompt_replacement_update_for_orphans else 'OBSOLETE: NO REPLACEMENT'
            t.setItem(target,3,QTableWidgetItem(replacement_text))
            existing.add(part)
            added += 1

        QMessageBox.information(self, 'Append Orphan Parents', f'Appended {added} Orphan Parent part(s) to OBS List.')

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def _find_col(self, name):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text().strip().lower() == name.lower():
                return i
        return -1

    def _import_wu_by_tool_comment(self, required_comment: str, flow_name: str):
        """Generic WU import by matching Imp BOM Tool comments value."""
        from PyQt6.QtCore import QThread, pyqtSignal, QObject
        from PyQt6.QtGui import QColor
        from PyQt6.QtWidgets import QProgressDialog

        wu_text = self.wu_level_input.text().strip()
        if not wu_text:
            QMessageBox.warning(self, 'WU Level Required',
                                f'Please enter a WU Level (1 to 6) before {flow_name}.')
            return
        try:
            max_level = int(wu_text)
        except ValueError:
            QMessageBox.warning(self, 'Invalid WU Level',
                                f'WU Level must be a number between 1 and 6, got: "{wu_text}"')
            return
        if max_level < 1 or max_level > 6:
            QMessageBox.warning(self, 'Invalid WU Level',
                                f'WU Level must be between 1 and 6, got: {max_level}')
            return

        selected_plant = self.plant_combo.currentText().strip() or '4070'

        if self.imp_bom_provider is None:
            QMessageBox.warning(self, 'Imp BOM Not Available',
                                'The Imp BOM tab is not connected. Cannot read removed parts.')
            return

        imp_table = self.imp_bom_provider.table
        part_col = self.imp_bom_provider.find_column('Part')
        tc_col = self.imp_bom_provider.find_column('Tool comments')

        if part_col < 0 or tc_col < 0:
            QMessageBox.warning(self, 'Imp BOM Missing Columns',
                                'Imp BOM does not contain both a Part and a Tool comments column.\n'
                                'Please import a BOM file first.')
            return

        seen = set()
        target_parts = []
        for r in range(imp_table.rowCount()):
            tc_item = imp_table.item(r, tc_col)
            if not tc_item:
                continue
            comment = tc_item.text().strip()
            # Accept a list of valid comments (case-sensitive)
            if isinstance(required_comment, list):
                if comment not in required_comment:
                    continue
            else:
                if comment != required_comment:
                    continue

            part_item = imp_table.item(r, part_col)
            if not part_item:
                continue
            part = part_item.text().strip()
            key = part.upper()
            if part and key not in seen:
                seen.add(key)
                target_parts.append(part)

        if not target_parts:
            label = ', '.join(required_comment) if isinstance(required_comment, list) else required_comment
            QMessageBox.information(
                self,
                'No Matching Parts',
                f'No parts found in Imp BOM where Tool comments = {label}.',
            )
            return

        if self.table.rowCount() > 0 and self.table.columnCount() > 0:
            reply = QMessageBox.question(
                self, 'Existing Data',
                'This tab already contains data.\nDelete it and import fresh WU data?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        try:
            import sys as _sys
            from pathlib import Path as _Path
            _mod_dir = str(_Path(__file__).parent)
            if _mod_dir not in _sys.path:
                _sys.path.insert(0, _mod_dir)
            from where_used_query import DISPLAY_HEADERS  # type: ignore[import]
        except ImportError as exc:
            QMessageBox.warning(self, 'Module Not Found',
                                f'where_used_query.py could not be imported:\n{exc}')
            return

        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QPlainTextEdit, QApplication
        from PyQt6.QtGui import QColor
        import time as _time_ui

        _label_parts = ', '.join(required_comment) if isinstance(required_comment, list) else required_comment
        log_dialog = QDialog(self)
        log_dialog.setWindowTitle('WU Query — Databricks')
        log_dialog.setMinimumWidth(620)
        log_dialog.setMinimumHeight(300)
        _dlg_layout = QVBoxLayout(log_dialog)
        _dlg_header = QLabel(
            f'Querying {len(target_parts)} part(s)  │  Tool comments = "{_label_parts}"'
            f'  │  WU Level {max_level}  │  Plant {selected_plant}'
        )
        _dlg_header.setStyleSheet('font-weight:bold; padding:6px; color:#7A1C21;')
        _dlg_layout.addWidget(_dlg_header)
        log_text = QPlainTextEdit()
        log_text.setReadOnly(True)
        log_text.setMaximumBlockCount(500)
        log_text.setStyleSheet('font-family: Consolas, monospace; font-size: 11px;')
        _dlg_layout.addWidget(log_text)
        log_dialog.setModal(True)
        log_dialog.show()
        QApplication.processEvents()

        _t_start = _time_ui.perf_counter()

        def _log_cb(msg):
            """Updates the log dialog immediately on the main thread."""
            log_text.appendPlainText(msg)
            log_text.verticalScrollBar().setValue(log_text.verticalScrollBar().maximum())
            QApplication.processEvents()

        _log_cb(
            f'[{_time_ui.strftime("%H:%M:%S")}] Starting: {len(target_parts)} part(s)'
            f' | WU Level {max_level} | Plant {selected_plant}'
        )

        _display_headers = DISPLAY_HEADERS
        _obs_map = self._build_obs_replacement_map()

        # Run synchronously on the main thread — same approach as the working
        # "Where Used of OBS Parts" button (import_from_databricks).  The log
        # dialog stays live because _log_cb calls processEvents() on each step.
        try:
            if max_level == 1:
                from where_used_query import fetch_where_used_level1_fast as _fwu_fast
                records = _fwu_fast(
                    target_parts,
                    selected_plant,
                    retain_9024=self.radio_9024_retain.isChecked(),
                    retain_esw=self.radio_esw_retain.isChecked(),
                    retain_above_cfg=self.radio_above_cfg_retain.isChecked(),
                    log_callback=_log_cb,
                )
            else:
                from where_used_query import fetch_where_used_parents_only as _fwu
                records = _fwu(
                    target_parts,
                    max_level,
                    selected_plant,
                    retain_9024=self.radio_9024_retain.isChecked(),
                    retain_esw=self.radio_esw_retain.isChecked(),
                    retain_above_cfg=self.radio_above_cfg_retain.isChecked(),
                    log_callback=_log_cb,
                )
        except Exception as exc:
            _log_cb(f'[ERROR] {exc}')
            log_dialog.close()
            QMessageBox.warning(self, 'Databricks Query Error', str(exc))
            return

        _elapsed = _time_ui.perf_counter() - _t_start
        _log_cb(f'[{_time_ui.strftime("%H:%M:%S")}] Done. {len(records)} record(s) in {_elapsed:.1f}s.')
        log_dialog.close()

        if not records:
            QMessageBox.information(
                self, 'No Data',
                f'Databricks returned no records for {len(target_parts)} part(s) at WU level {max_level}.',
            )
            return

        _WU_COL = 1
        _PART_COL = 2
        _ORPHAN_COL = 3
        _REPL_COL = 4
        _DATA_START = 5

        wu_headers = list(_display_headers)
        wu_headers.insert(2, 'Orphans List')
        all_headers = ['Select'] + wu_headers

        _DB_COL_FUNCS = [
            lambda r: r.get('rev_ln', ''),
            lambda r: r.get('plant', ''),
            lambda r: r.get('description', ''),
            lambda r: r.get('item_status', ''),
            lambda r: r.get('base_qty', ''),
            lambda r: r.get('ext_qty', ''),
            lambda r: r.get('uom', ''),
            lambda r: r.get('eco_number', ''),
            lambda r: r.get('procurement_type', ''),
            lambda r: r.get('effectivity_date', ''),
            lambda r: r.get('user_item_type', ''),
            lambda r: r.get('item_seq', ''),
            lambda r: r.get('kit_code', ''),
            lambda r: r.get('sparable_flag', ''),
            lambda r: r.get('designator', ''),
            lambda r: r.get('option_class', ''),
            lambda r: (r.get('pace_or_dash', '') if 'pace' in r.get('pace_or_dash', '').lower() else ''),
            lambda r: r.get('mlo_class', ''),
        ]

        self.table.setUpdatesEnabled(False)
        self.table.clear()
        self.table.setColumnCount(len(all_headers))
        self.table.setHorizontalHeaderLabels(all_headers)
        self.table.setRowCount(len(records))

        for row_idx, record in enumerate(records):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            cont._chk = chk
            self.table.setCellWidget(row_idx, 0, cont)

            wu_val = str(record.get('wu_level', ''))
            self.table.setItem(row_idx, _WU_COL, QTableWidgetItem(wu_val))

            raw_part = record.get('part', '')
            try:
                level_int = int(wu_val)
            except (ValueError, TypeError):
                level_int = 0
            self.table.setItem(row_idx, _PART_COL,
                               QTableWidgetItem(('      ' * level_int) + raw_part))

            self.table.setItem(row_idx, _ORPHAN_COL, QTableWidgetItem(''))

            replacement = _obs_map.get(raw_part.strip().upper(), '')
            rep_item = QTableWidgetItem(replacement)
            rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, _REPL_COL, rep_item)

            for c_off, fn in enumerate(_DB_COL_FUNCS):
                self.table.setItem(row_idx, _DATA_START + c_off,
                                   QTableWidgetItem(fn(record)))

            if wu_val == '0':
                _blue = QColor('#C7DEFA')
                cont.setStyleSheet('background-color: #C7DEFA;')
                for _col in range(1, len(all_headers)):
                    _item = self.table.item(row_idx, _col)
                    if _item is not None:
                        _item.setBackground(_blue)

        self.table.setUpdatesEnabled(True)

        hdr = self.table.horizontalHeader()
        if len(records) > 4000:
            for i in range(len(all_headers)):
                if i == len(all_headers) - 1:
                    hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
                else:
                    hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
            if len(all_headers) > 0:
                self.table.setColumnWidth(0, 65)
            if len(all_headers) > 1:
                self.table.setColumnWidth(1, 85)
            if len(all_headers) > 2:
                self.table.setColumnWidth(2, 260)
        else:
            for i in range(len(all_headers)):
                if i == len(all_headers) - 1:
                    hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
                else:
                    hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)

        _total_elapsed = _time_ui.perf_counter() - _t_start
        _comment_label = ', '.join(required_comment) if isinstance(required_comment, list) else required_comment

        QMessageBox.information(
            self,
            'Import Complete',
            f'Imported {len(records)} row(s) from Databricks.\n'
            f'Parts queried (Tool comments = "{_comment_label}"): {len(target_parts)}\n'
            f'WU Level: {max_level} | Plant: {selected_plant}\n'
            f'Radio filters were applied during query traversal (pre-import).\n'
            f'Total time: {_total_elapsed:.1f}s',
        )

    def _normalize_bom_part(self, value: str) -> str:
        return (value or '').strip()

    def _parse_bom_level_value(self, value: str):
        txt = (value or '').strip()
        if not txt:
            return None
        try:
            return int(float(txt))
        except (ValueError, TypeError):
            return None

    def _is_provider_loaded(self, provider) -> bool:
        return bool(provider and provider.table.columnCount() > 0 and provider.table.rowCount() > 0)

    def _collect_bom_rows_by_root_and_level(self, provider):
        if provider is None:
            return {}
        part_col = provider.find_column('Part')
        bom_col = provider.find_column('BOM Level')
        if part_col < 0 or bom_col < 0:
            return {}

        groups = {}
        current_root_key = None
        current_root_text = ''
        table = provider.table
        for r in range(table.rowCount()):
            part_item = table.item(r, part_col)
            bom_item = table.item(r, bom_col)
            part = self._normalize_bom_part(part_item.text() if part_item else '')
            level = self._parse_bom_level_value(bom_item.text() if bom_item else '')
            if not part or level is None:
                continue

            if level == 0:
                current_root_key = part.upper()
                current_root_text = part
                groups.setdefault(current_root_key, {
                    'root_part': current_root_text,
                    'levels': {},
                })
                continue

            if current_root_key is None:
                continue

            root_entry = groups.setdefault(current_root_key, {
                'root_part': current_root_text or current_root_key,
                'levels': {},
            })
            level_entry = root_entry['levels'].setdefault(level, [])
            level_entry.append((r, part.upper()))

        return groups

    def _clear_obs_compare_comments(self):
        provider = self.imp_bom_provider
        if provider is None:
            return
        tc_col = provider.find_column('Tool comments')
        bom_col = provider.find_column('BOM Level')
        if tc_col < 0 or bom_col < 0:
            return

        for r in range(provider.table.rowCount()):
            bom_item = provider.table.item(r, bom_col)
            level = self._parse_bom_level_value(bom_item.text() if bom_item else '')
            if level is None or level == 0:
                continue
            provider.table.setItem(r, tc_col, QTableWidgetItem(''))

    def _mark_obs_compare_rows(self, row_indexes):
        provider = self.imp_bom_provider
        if provider is None:
            return
        tc_col = provider.find_column('Tool comments')
        if tc_col < 0:
            return

        for r in sorted(set(row_indexes)):
            it = QTableWidgetItem('Removed BOM Item')
            f = it.font()
            f.setBold(True)
            it.setFont(f)
            it.setForeground(QColor('orange'))
            provider.table.setItem(r, tc_col, it)

    def _prepare_with_replacement_compare_comments(self):
        if not self._is_provider_loaded(self.imp_bom_provider):
            QMessageBox.warning(
                self,
                'OBS BOM Required',
                'Please import Imp BOM_OBS Parts before importing WU of Removed BOM Items.'
            )
            return False

        if not self._is_provider_loaded(self.replacement_bom_provider):
            QMessageBox.warning(
                self,
                'Replacement BOM Required',
                'Please import Imp REPL Parts BOM before importing WU of Removed BOM Items.'
            )
            return False

        obs_map = self._build_obs_replacement_map()
        obs_groups = self._collect_bom_rows_by_root_and_level(self.imp_bom_provider)
        repl_groups = self._collect_bom_rows_by_root_and_level(self.replacement_bom_provider)
        if not obs_groups:
            QMessageBox.warning(
                self,
                'OBS BOM Required',
                'No OBS BOM rows were found to compare. Please import Imp BOM_OBS Parts first.'
            )
            return False

        self._clear_obs_compare_comments()

        mismatch_rows = []
        for obs_root_key, obs_group in obs_groups.items():
            replacement_part = (obs_map.get(obs_root_key) or '').strip()
            has_replacement = replacement_part and not re.search(r'no\s*replacement', replacement_part, re.IGNORECASE)
            if not has_replacement:
                for level_rows in obs_group['levels'].values():
                    mismatch_rows.extend(r for r, _ in level_rows)
                continue

            repl_group = repl_groups.get(replacement_part.upper())
            if repl_group is None:
                for level_rows in obs_group['levels'].values():
                    mismatch_rows.extend(r for r, _ in level_rows)
                continue

            for level, obs_level_rows in obs_group['levels'].items():
                repl_level_rows = repl_group['levels'].get(level, [])
                repl_parts_at_level = {part for _, part in repl_level_rows}
                if not repl_parts_at_level:
                    mismatch_rows.extend(r for r, _ in obs_level_rows)
                    continue
                for row_idx, obs_part in obs_level_rows:
                    if obs_part not in repl_parts_at_level:
                        mismatch_rows.append(row_idx)

        mismatch_rows = sorted(set(mismatch_rows))
        self._mark_obs_compare_rows(mismatch_rows)

        if not mismatch_rows:
            QMessageBox.information(
                self,
                'No Mismatched BOM Items',
                'OBS and Replacement BOM items match at each BOM Level. No WU import is needed.'
            )
            return False

        return True

    def import_from_databricks_for_removed_items(self):
        """Imports WU for Removed BOM items, with compare-first logic in with-replacement mode."""
        if self.compare_with_replacement:
            if not self._prepare_with_replacement_compare_comments():
                return
        # Only accept exact 'Removed BOM Item' (case-sensitive)
        self._import_wu_by_tool_comment(
            required_comment=['Removed BOM Item'],
            flow_name='Import WU of Removed BOM items',
        )

    def _sap_import_removed_placeholder(self):
        pass  # SAP import removed

    def _remove_parents_by_radio_selection(self):
        """Remove parent rows (WU Level != 0) based on radio buttons.
        Called after import AND before orphan analysis.
        Rows where the Replacement column is NOT blank are always kept."""
        wucol = pcol = repl_col = -1
        for ci in range(self.table.columnCount()):
            hdr = self.table.horizontalHeaderItem(ci)
            if hdr is None:
                continue
            ht = hdr.text().strip().lower()
            if ht == 'wu level':
                wucol = ci
            elif ht == 'part':
                pcol = ci
            elif ht == 'replacement':
                repl_col = ci
        if wucol < 0 or pcol < 0:
            return 0

        remove_9024 = self.radio_9024_remove.isChecked()
        remove_esw = self.radio_esw_remove.isChecked()
        remove_above_cfg = self.radio_above_cfg_remove.isChecked()
        if not remove_9024 and not remove_esw and not remove_above_cfg:
            return 0

        def _wu_level_of_row(row_idx: int):
            wu_it = self.table.item(row_idx, wucol)
            if wu_it is None:
                return None
            try:
                return int(float(wu_it.text().strip()))
            except (ValueError, TypeError):
                return None

        def _has_replacement(row_idx: int) -> bool:
            if repl_col < 0:
                return False
            repl_it = self.table.item(row_idx, repl_col)
            return bool(repl_it is not None and repl_it.text().strip())

        rows_to_delete = []

        # 9024/ESW parent filters
        for r in range(self.table.rowCount()):
            part_it = self.table.item(r, pcol)
            if part_it is None:
                continue

            lvl = _wu_level_of_row(r)
            if lvl is None or lvl == 0:
                continue

            # Keep rows where Replacement is not blank regardless of radio selections.
            if _has_replacement(r):
                continue

            # Normalize: strip ALL whitespace and uppercase
            pv = ''.join(part_it.text().split()).upper()
            if remove_9024 and pv.startswith('9024'):
                rows_to_delete.append(r)
            elif remove_esw and pv.startswith('ESW'):
                rows_to_delete.append(r)

        # SmBOM-above-config filter (same behavior as Where Used import path)
        if remove_above_cfg:
            config_prefixes = {
                '0490', '0491', '0495', '0497', '0430', '0350', '0355', '0351', '0357',
                '0390', '0395', '0397', '0335', '0391', '0431', '0435', '0437',
                '0440', '0445', '0455', '0450', '0441', '0447', '0457',
                '0460', '0465', '0461', '0467', '0410', '0415', '0417',
                '0411', '0412', '0413', '0414', '0360', '0365', '0361', '0367'
            }
            listing_prefixes = {'0243', '0288', '0289', '0290'}
            prune_anchor_prefixes = config_prefixes | listing_prefixes

            n = self.table.rowCount()
            r = 0
            while r < n:
                lvl = _wu_level_of_row(r)
                if lvl != 0:
                    r += 1
                    continue

                block_start = r
                block_end = n
                for i in range(r + 1, n):
                    lvl_i = _wu_level_of_row(i)
                    if lvl_i == 0:
                        block_end = i
                        break

                for i in range(block_start + 1, block_end):
                    lvl_i = _wu_level_of_row(i)
                    if lvl_i is None or lvl_i == 0:
                        continue

                    part_i_it = self.table.item(i, pcol)
                    part_i = (part_i_it.text() if part_i_it else '').strip()
                    if part_i[:4] in prune_anchor_prefixes and not _has_replacement(i):
                        for j in range(i + 1, block_end):
                            lvl_j = _wu_level_of_row(j)
                            if lvl_j is None:
                                continue
                            if lvl_j <= lvl_i:
                                break
                            if lvl_j != 0 and not _has_replacement(j):
                                rows_to_delete.append(j)

                r = block_end

        for r in sorted(set(rows_to_delete), reverse=True):
            self.table.removeRow(r)
        return len(rows_to_delete)

    def _build_obs_change_map(self):
        mapping = {}
        t = self.obs_provider.table
        for r in range(t.rowCount()):
            obs_it = t.item(r, 1)
            chg_w = t.cellWidget(r, 2)
            key = obs_it.text().strip().upper() if obs_it else ''
            if key:
                mapping[key] = chg_w.currentText() if chg_w else ''
        return mapping

    def _build_obs_replacement_map(self):
        mapping = {}
        t = self.obs_provider.table
        for r in range(t.rowCount()):
            obs_it = t.item(r, 1)
            rep_it = t.item(r, 3)
            key = obs_it.text().strip().upper() if obs_it else ''
            if key:
                mapping[key] = (rep_it.text() if rep_it else '').strip()
        return mapping

    def _get_orphan_label_start_level(self) -> int:
        """Return the first orphan level to assign for the current analysis run.

        In With / Without Replacement mode, continue numbering from the highest
        value already present in OBS Parts -> Identified Orphans. In all other
        modes, preserve the legacy behavior starting at Orphan1.
        """
        if not self.compare_with_replacement:
            return 1
        if not self.obs_provider or not hasattr(self.obs_provider, 'table'):
            return 1

        identified_col = self._find_obs_col('Identified Orphans')
        if identified_col < 0:
            return 1

        max_level = 0
        t = self.obs_provider.table
        for r in range(t.rowCount()):
            it = t.item(r, identified_col)
            value = (it.text() if it else '').strip()
            if not value:
                continue
            match = re.match(r'^orphan\s*(\d+)$', value, flags=re.IGNORECASE)
            if not match:
                continue
            try:
                max_level = max(max_level, int(match.group(1)))
            except ValueError:
                pass
        return max_level + 1 if max_level > 0 else 1

    def _build_orphan_blocks_from_wu(self) -> List[Dict[str, Any]]:
        """Build independent orphan blocks from WU of Removed BOM Items.

        Each block contains:
        - parents: immediate parent part(s) at WU Level 1 in the same WU-0 block
        - child: orphan part at WU Level 0
        """
        blocks: List[Dict[str, Any]] = []
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphans List')
        wucol = self._find_col('WU Level')
        rcol = self._find_col('Replacement')
        if pcol < 0 or ocol < 0 or wucol < 0:
            return blocks

        def is_zero(v):
            return str(v).strip() in ('0', '0.0')

        def is_one(v):
            return str(v).strip() in ('1', '1.0')

        # Keep aligned with the parent ignore rule used in orphan derivation.
        ignore_prefixes = ('0243', '0299', '0289', '0290')

        row_count = self.table.rowCount()
        r = 0
        while r < row_count:
            wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
            if not is_zero(wu):
                r += 1
                continue

            child_it = self.table.item(r, pcol)
            orphan_it = self.table.item(r, ocol)
            child_part = (child_it.text() if child_it else '').strip().upper()
            orphan_label = (orphan_it.text() if orphan_it else '').strip().lower()

            block_end = row_count
            for i in range(r + 1, row_count):
                nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                if is_zero(nxt):
                    block_end = i
                    break

            if child_part and orphan_label.startswith('orphan'):
                parent_list: List[str] = []
                wu_replacements: Dict[str, str] = {}

                if rcol >= 0:
                    child_rep_it = self.table.item(r, rcol)
                    child_rep = (child_rep_it.text() if child_rep_it else '').strip()
                    if child_rep:
                        wu_replacements[child_part] = child_rep

                for i in range(r + 1, block_end):
                    wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if not is_one(wu_p):
                        continue
                    pit = self.table.item(i, pcol)
                    parent_part = (pit.text().lstrip() if pit else '').strip().upper()
                    if not parent_part or parent_part.startswith(ignore_prefixes):
                        continue
                    if parent_part != child_part and parent_part not in parent_list:
                        parent_list.append(parent_part)
                    if rcol >= 0 and parent_part:
                        parent_rep_it = self.table.item(i, rcol)
                        parent_rep = (parent_rep_it.text() if parent_rep_it else '').strip()
                        if parent_rep and parent_part not in wu_replacements:
                            wu_replacements[parent_part] = parent_rep
                if parent_list:
                    blocks.append({'parents': parent_list, 'child': child_part, 'wu_replacements': wu_replacements})

            r = block_end

        return blocks

    def perform_orphan_analysis(self):
        # Applicable only for With / Without Replacement mode.
        if self.compare_with_replacement:
            self._ensure_obs_proposed_replacement_column()

        # Snapshot existing orphan assignments so we can detect true new findings.
        _existing_orphan_entries = set()
        _part_col_pre = self._find_col('Part')
        _orph_col_pre = self._find_col('Orphans List')
        if _part_col_pre >= 0 and _orph_col_pre >= 0:
            for _r in range(self.table.rowCount()):
                _pit = self.table.item(_r, _part_col_pre)
                _oit = self.table.item(_r, _orph_col_pre)
                _p = (_pit.text() if _pit else '').strip().upper()
                _o = (_oit.text() if _oit else '').strip()
                if _p and _o and _o.lower().startswith('orphan'):
                    _existing_orphan_entries.add((_p, _o.lower()))

        # Find required columns
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphans List')
        wucol = self._find_col('WU Level')

        if pcol < 0 or ocol < 0 or wucol < 0:
            QMessageBox.information(self, 'Orphan Analysis', 'Required columns not found.')
            return

        row_count = self.table.rowCount()

        def is_zero(v):
            return str(v).strip() in ('0', '0.0')

        def is_one(v):
            return str(v).strip() in ('1', '1.0')

        ignore_prefixes = ('0243', '0299', '0289', '0290')

        # Step 0: OBS → Obsolete / Inactivate mapping (existing behavior)
        obs_map = self._build_obs_change_map()
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if not it:
                continue
            key = it.text().lstrip().upper()
            if key in obs_map:
                self.table.setItem(r, ocol, QTableWidgetItem(obs_map[key]))

        start_level = self._get_orphan_label_start_level()
        first_orphan_label = f'Orphan{start_level}'

        # Step 1: first orphan level for this run
        orphan_parts = set()
        r = 0
        while r < row_count:
            wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
            if is_zero(wu):
                child_row = r
                part_it = self.table.item(child_row, pcol)
                orphan_it = self.table.item(child_row, ocol)
                part = part_it.text().strip() if part_it else ''

                if orphan_it and orphan_it.text().strip():
                    r += 1
                    continue

                block_end = row_count
                for i in range(r + 1, row_count):
                    nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if is_zero(nxt):
                        block_end = i
                        break

                parents = []
                for i in range(child_row + 1, block_end):
                    wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if not is_one(wu_p):
                        continue
                    pit = self.table.item(i, pcol)
                    pval = pit.text().lstrip() if pit else ''
                    if pval.startswith(ignore_prefixes):
                        continue
                    parents.append(i)

                mark = False
                if not parents:
                    mark = True
                else:
                    all_bad = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        status = oit.text().strip().lower() if oit else ''
                        if status not in ('obsolete', 'inactivate'):
                            all_bad = False
                            break
                    if all_bad:
                        mark = True

                if mark and part:
                    self.table.setItem(child_row, ocol, QTableWidgetItem(first_orphan_label))
                    orphan_parts.add(part.upper())

                r = block_end
            else:
                r += 1

        # Propagate first orphan level for this run
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if it and it.text().strip().upper() in orphan_parts:
                oit = self.table.item(r, ocol)
                if not oit or not oit.text().strip():
                    self.table.setItem(r, ocol, QTableWidgetItem(first_orphan_label))

        # Step 2+: continue numbering after the first orphan level for this run
        current_level = start_level + 1
        while True:
            new_found = False
            new_parts = set()
            r = 0
            while r < row_count:
                wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
                orphan_it = self.table.item(r, ocol)
                if is_zero(wu) and (not orphan_it or not orphan_it.text().strip()):
                    part_it = self.table.item(r, pcol)
                    part = part_it.text().strip() if part_it else ''

                    block_end = row_count
                    for i in range(r + 1, row_count):
                        nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if is_zero(nxt):
                            block_end = i
                            break

                    parents = []
                    for i in range(r + 1, block_end):
                        wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if not is_one(wu_p):
                            continue
                        pit = self.table.item(i, pcol)
                        pval = pit.text().lstrip() if pit else ''
                        if pval.startswith(ignore_prefixes):
                            continue
                        parents.append(i)

                    parents_filled = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        if not oit or not oit.text().strip():
                            parents_filled = False
                            break

                    if part and (not parents or parents_filled):
                        self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))
                        new_parts.add(part.upper())
                        new_found = True
                r += 1

            for r in range(row_count):
                it = self.table.item(r, pcol)
                oit = self.table.item(r, ocol)
                if it and it.text().strip().upper() in new_parts and (not oit or not oit.text().strip()):
                    self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))

            if not new_found:
                break
            current_level += 1

            self.append_orphans_to_obs_parts()

        # Ensure Orphan1-only results are also pushed to OBS parts.
        self.append_orphans_to_obs_parts()

        # ── Orphan Parent detection ──────────────────────────────────────────
        # Executes once, only when the dataset contains rows with WU Level > 1.
        # Must run after all Orphan detection logic is fully complete.
        # Does NOT overwrite any cell in the Orphans List column that already
        # has a value.
        def _as_wu_int(_val) -> int | None:
            """Parse WU level values like 2, 2.0, '2', '2.0' safely."""
            _txt = str(_val).strip()
            if not _txt:
                return None
            try:
                return int(float(_txt))
            except (ValueError, TypeError):
                return None

        _max_wu_in_data = 0
        for _r in range(row_count):
            _wu_it = self.table.item(_r, wucol)
            if _wu_it:
                _v = _as_wu_int(_wu_it.text())
                if _v is not None and _v > _max_wu_in_data:
                    _max_wu_in_data = _v

        if _max_wu_in_data > 1:
            # Part-number prefixes whose presence on the immediate parent row
            # exempts the evaluated row from being flagged as Orphan Parent:
            #   Option prefixes, Option Class prefixes, Commodity Codes 0288/0289/0243/0290
            _orphan_parent_exempt_prefixes = (
                '0288', '0289', '0243', '0290',
                '0490', '0491', '0495', '0497', '0430',
                '0350', '0355', '0351', '0357',
                '0390', '0395', '0397', '0335',
                '0391', '0431', '0435', '0437',
                '0440', '0445', '0455', '0450', '0441', '0447', '0457',
                '0460', '0465', '0461', '0467',
                '0410', '0415', '0417', '0411', '0412', '0413', '0414',
                '0360', '0365', '0361', '0367',
            )

            _r = 0
            while _r < row_count:
                _wu_it = self.table.item(_r, wucol)
                _wu_val = _wu_it.text().strip() if _wu_it else ''
                if is_zero(_wu_val):
                    # Identify block boundaries for this WU Level 0 row
                    _block_start = _r
                    _block_end = row_count
                    for _i in range(_r + 1, row_count):
                        _nxt = self.table.item(_i, wucol)
                        if _nxt and is_zero(_nxt.text()):
                            _block_end = _i
                            break

                    # Evaluate every row in the block with WU Level > 1
                    for _i in range(_block_start + 1, _block_end):
                        _wu_row_it = self.table.item(_i, wucol)
                        if not _wu_row_it:
                            continue
                        _wu_int = _as_wu_int(_wu_row_it.text())
                        if _wu_int is None:
                            continue

                        # Condition 1: WU Level must be > 1
                        if _wu_int <= 1:
                            continue

                        # Condition 2b: Skip if the row itself is an Option or
                        # Option Class part (starts with an option prefix).
                        # Options not reporting to Option Class should NOT be
                        # flagged as Orphan Parent.
                        _self_part_it = self.table.item(_i, pcol)
                        _self_part = (_self_part_it.text().lstrip()
                                      if _self_part_it else '')
                        if _self_part.strip().startswith(_orphan_parent_exempt_prefixes):
                            continue

                        # Condition 3: Orphans List column must be blank
                        _oit = self.table.item(_i, ocol)
                        if _oit and _oit.text().strip():
                            continue

                        # Find the immediate parent row: nearest preceding row in this
                        # block whose WU Level equals (_wu_int - 1)
                        _parent_wu_target = _wu_int - 1
                        _parent_row = -1
                        for _j in range(_i - 1, _block_start, -1):
                            _pj_wu_it = self.table.item(_j, wucol)
                            if not _pj_wu_it:
                                continue
                            _pj_wu = _as_wu_int(_pj_wu_it.text())
                            if _pj_wu is None:
                                continue
                            if _pj_wu == _parent_wu_target:
                                _parent_row = _j
                                break
                            elif _pj_wu < _parent_wu_target:
                                # Hierarchy went below the expected level; stop
                                break

                        # Skip if no clear immediate parent found (conservative)
                        if _parent_row < 0:
                            continue

                        # Condition 2: parent must NOT be an Option / Option Class /
                        # Commodity Code (0288, 0289, 0243, 0290)
                        _parent_part_it = self.table.item(_parent_row, pcol)
                        _parent_part = (_parent_part_it.text().lstrip()
                                        if _parent_part_it else '')
                        if _parent_part.strip().startswith(_orphan_parent_exempt_prefixes):
                            # Parent is an exempt type -> not an Orphan Parent
                            continue

                        # All conditions satisfied -> mark as Orphan Parent
                        self.table.setItem(_i, ocol, QTableWidgetItem('Orphan Parent'))

                    _r = _block_end
                else:
                    _r += 1
        # ── End Orphan Parent detection ──────────────────────────────────────

        # Build/update Orphan Hierarchy view in OBS Parts after each run.
        if self.obs_provider and hasattr(self.obs_provider, 'update_orphan_hierarchy_view'):
            _hier_blocks = self._build_orphan_blocks_from_wu()
            _wu_repl_lookup: Dict[str, str] = {}
            for _b in _hier_blocks:
                for _k, _v in (_b.get('wu_replacements', {}) or {}).items():
                    _kk = (_k or '').strip().upper()
                    _vv = (_v or '').strip()
                    if _kk and _vv and _kk not in _wu_repl_lookup:
                        _wu_repl_lookup[_kk] = _vv
            self.obs_provider.update_orphan_hierarchy_view(_hier_blocks, _wu_repl_lookup)

        from PyQt6.QtGui import QColor
        for r in range(self.table.rowCount()):
            wu = self.table.item(r, wucol)
            if wu and wu.text().strip() == '0':
                for c in range(self.table.columnCount()):
                    it = self.table.item(r, c)
                    if it:
                        it.setBackground(QColor('#87CEEB'))

        # Notify user when no newly identified orphan entries were added by this run.
        _new_orphan_entries = set()
        for _r in range(self.table.rowCount()):
            _pit = self.table.item(_r, pcol)
            _oit = self.table.item(_r, ocol)
            _p = (_pit.text() if _pit else '').strip().upper()
            _o = (_oit.text() if _oit else '').strip()
            if _p and _o and _o.lower().startswith('orphan'):
                _new_orphan_entries.add((_p, _o.lower()))
        if not (_new_orphan_entries - _existing_orphan_entries):
            QMessageBox.information(self, 'Orphan Analysis', 'Orphan Analysis completed no new Orphans Found.')


    def remove_by_prefix(self, prefix):
        pcol = self._find_col('Part')
        if pcol < 0:
            return
        rows = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, pcol)
            if it and it.text().lstrip().startswith(prefix):
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)


class  OBSAllPartsWithoutReplacementTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)

        # --- Radio buttons for Replacement Mode ---
        radio_row = QHBoxLayout()
        radio_row.setContentsMargins(0, 0, 0, 0)
        radio_row.setSpacing(16)
        self.rb_without = QRadioButton("Without Replacement*")
        self.rb_with = QRadioButton("With / Without Replacement")
        self.rb_without.setChecked(True)
        radio_group = QButtonGroup(self)
        radio_group.addButton(self.rb_without, 0)
        radio_group.addButton(self.rb_with, 1)
        radio_row.addWidget(QLabel("Mode:"))
        radio_row.addWidget(self.rb_without)
        radio_row.addWidget(self.rb_with)
        radio_row.addStretch(1)
        outer.addLayout(radio_row)

        # --- Sub-tabs for each mode ---
        self.subtabs = QTabWidget()
        self.subtabs.setStyleSheet("""
        QTabBar::tab {
            min-width: 220px;
            padding: 6px 12px;
            text-align: center;
        }
        """)

        # --- Sub-tabs for 'Without Replacement' ---
        self.tab_without_imp_bom = OrphanOBSSubTab()
        self.tab_without_wu_removed = WURemovedBOMItemsTab(self.obs_provider, imp_bom_provider=self.tab_without_imp_bom)
        self.subtabs.addTab(self.tab_without_imp_bom, "Imp BOM")
        self.subtabs.addTab(self.tab_without_wu_removed, "WU of Removed BOM Items")

        # --- Sub-tabs for 'With / Without Replacement' ---
        # 1. Imp BOM_OBS Parts (same as Imp BOM, but for this mode)
        self.tab_with_obs = OrphanOBSSubTab()
        # 2. Imp BOM_REPL Parts (duplicate of Imp BOM, but for replacement parts)
        self.tab_with_repl = OrphanOBSSubTab()
        # 3. WU of Removed BOM Items (reuse logic, but for this mode)
        self.tab_with_wu_removed = WURemovedBOMItemsTab(self.obs_provider, imp_bom_provider=self.tab_with_obs)

        # Store tab indices for switching
        self.without_tabs = [self.tab_without_imp_bom, self.tab_without_wu_removed]
        self.with_tabs = [self.tab_with_obs, self.tab_with_repl, self.tab_with_wu_removed]

        # Add only 'Without Replacement' tabs initially
        self._set_subtabs('without')
        outer.addWidget(self.subtabs)

        # Connect radio buttons to switch sub-tabs
        self.rb_without.toggled.connect(lambda checked: checked and self._set_subtabs('without'))
        self.rb_with.toggled.connect(lambda checked: checked and self._set_subtabs('with'))

    def _set_subtabs(self, mode):
        self.subtabs.clear()
        if mode == 'without':
            self.subtabs.addTab(self.tab_without_imp_bom, "Imp BOM_OBS Parts")
            self.subtabs.addTab(self.tab_without_wu_removed, "WU of Removed BOM Items")
        elif mode == 'with':
            self.subtabs.addTab(self.tab_with_obs, "Imp BOM_OBS Parts")
            self.subtabs.addTab(self.tab_with_repl, "Imp BOM_REPL Parts")
            self.subtabs.addTab(self.tab_with_wu_removed, "WU of Removed BOM Items")

class _OrphanAnalysisTab_OLD(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        outer = QVBoxLayout(self)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        self.subtabs = QTabWidget()
        self.subtabs.setStyleSheet("""
        QTabBar::tab:selected {
            background-color: #87CEEB;
            color: #0F2D46;
            font-weight: 600;
        }
        QTabBar::tab {
            background-color: #EAF6FD;
        }
        """)
        self.subtabs.addTab(
            OBSAllPartsWithoutReplacementTab(obs_provider),
            "OBS all Parts without Replacements"
        )
        self.subtabs.addTab(
            PlaceholderTab("OBS with or without Replacement"),
            "OBS with or without Replacement"
        )

        outer.addWidget(self.subtabs)




# === EC Creation Form (Embedded) ===

EC_CATEGORY_DESC = {
    "A1": "SMBoM Options as revised items and having CDW (Cost Disposition Worksheet)",
    "A2": "SMBoM Options as revised items, and No CDW",
    "B1": "No SMBoM Options as revised items and having CDW (Cost Disposition Worksheet)",
    "B2": "No SMBoM Options as revised items, No CDW, and revised item status at Eval and/or moving to Eval or adding already released parts to Proto buckets",
    "B3": "No SMBoM Option as revised items, No CDW, and revised Item status at Production and/or moving to Production",
}


def header(text):
    lbl = QLabel(text)
    lbl.setFont(QFont("Segoe UI", 12, QFont.Weight.DemiBold))
    return lbl


def highlight_label(text):
    lbl = QLabel(text)
    lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
    lbl.setStyleSheet("color:#1F4E79")
    return lbl


class LimitedTextEdit(QTextEdit):
    def __init__(self, limit=2000):
        super().__init__()
        self.limit = limit
        self.textChanged.connect(self._limit)

    def _limit(self):
        if self.limit is None:
            return
        text = self.toPlainText()
        if len(text) > self.limit:
            self.blockSignals(True)
            self.setPlainText(text[:self.limit])
            self.moveCursor(QTextCursor.MoveOperation.End)
            self.blockSignals(False)

    def _update_counter(self):
        if not self.counter_label:
            return
        count = len(self.toPlainText())
        self.counter_label.setText(f"{count} / {self.limit} characters")
        if count >= self.limit:
            self.counter_label.setStyleSheet("color:red;font-size:10px")
        else:
            self.counter_label.setStyleSheet("color:gray;font-size:10px")

class ECCreationInputsFormTab(QWidget):
    ec_category_selected = pyqtSignal(str)
    reset_form_requested = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.email_file_path = None  # Stores the path of the browsed email file
        root = QVBoxLayout(self)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        root.addWidget(scroll)
        container = QWidget(); scroll.setWidget(container)
        outer = QVBoxLayout(container)

        # ---- Section A ----
        section_a_title_row = QHBoxLayout()
        section_a_title_row.addWidget(header("Section A: EC Category Form"))
        section_a_title_row.addStretch(1)

        self.btn_reset_ec_form = QPushButton("Reset All")
        self.btn_reset_ec_form.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_reset_ec_form.setStyleSheet(
            "QPushButton {"
            "  color:#FFFFFF; padding:6px 14px; border-radius:6px; border:1px solid #0D5EA6;"
            "  background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2994FF, stop:1 #0A67C2);"
            "}"
            "QPushButton:hover {"
            "  background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2FA0FF, stop:1 #0D6ED0);"
            "}"
        )
        self.btn_reset_ec_form.clicked.connect(self._request_reset_ec_form)
        section_a_title_row.addWidget(self.btn_reset_ec_form)
        outer.addLayout(section_a_title_row)

        secA = QFrame(); a = QVBoxLayout(secA)

        scope_row = QHBoxLayout(); scope_grp = QButtonGroup(self)
        self.scope_group = scope_grp
        for scope in ["Up Revision", "Status Roll", "Production Release", "OBS / Inactivate", "Product Release"]:
            rb = QRadioButton(scope)
            scope_grp.addButton(rb); rb.toggled.connect(self.start_flow)
            scope_row.addWidget(rb)
        a.addLayout(scope_row)

        self.flow_area = QVBoxLayout(); a.addLayout(self.flow_area)
        self.ec_result_lbl = QLabel(""); self.ec_result_lbl.setWordWrap(True)
        self.ec_result_lbl.setStyleSheet("color:green;font-weight:600;font-size:14px")
        # Hidden – category is shown in the 'EC Category based on Selected Criteria' field below

        # ── AI Proposed EC Category row ───────────────────────────────────────
        ai_ec_row = QHBoxLayout()
        ai_ec_row.setSpacing(8)
        ai_ec_prop_lbl = QLabel("EC Category based on Selected Criteria:")
        ai_ec_prop_lbl.setStyleSheet("color:#1F4E79; font-weight:600;")
        ai_ec_row.addWidget(ai_ec_prop_lbl)

        self.ai_ec_category_edit = QLineEdit()
        self.ai_ec_category_edit.setReadOnly(True)
        self.ai_ec_category_edit.setPlaceholderText("")
        self.ai_ec_category_edit.setFixedHeight(26)
        self.ai_ec_category_edit.setMinimumWidth(60)
        self.ai_ec_category_edit.setMaximumWidth(120)
        self.ai_ec_category_edit.setStyleSheet(
            "background:#E8F5E9; border:1px solid #81C784; border-radius:4px; "
            "color:#1B5E20; font-weight:bold; padding:2px 6px;"
        )
        ai_ec_row.addWidget(self.ai_ec_category_edit)

        self.ai_ec_justification_btn = QPushButton("\u24d8")
        self.ai_ec_justification_btn.setFixedSize(22, 22)
        self.ai_ec_justification_btn.setStyleSheet(
            "QPushButton { background:#1F4E79; color:white; border-radius:11px; "
            "font-weight:bold; font-size:12px; border:none; }"
            "QPushButton:hover { background:#2E75B6; }"
        )
        self.ai_ec_justification_btn.setToolTip("Run Problem Summary to see EC Category justification")
        self.ai_ec_justification_btn.setCursor(Qt.CursorShape.WhatsThisCursor)
        self.ai_ec_justification_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        ai_ec_row.addWidget(self.ai_ec_justification_btn)
        ai_ec_row.addStretch(1)
        a.addLayout(ai_ec_row)
        # ───────────────────────────────────────────────────────────────────────

        self.ec_divider = QFrame(); self.ec_divider.setFrameShape(QFrame.Shape.HLine)
        self.ec_divider.setVisible(False); a.addWidget(self.ec_divider)
        outer.addWidget(secA)
        outer.addSpacing(12)

# ---- Section B ----
        self.secB_header = header("Section B: ECR Change Assessment")
        self.secB_header.setVisible(False)
        outer.addWidget(self.secB_header)

        self.secB = QFrame()
        self.secB.setVisible(False)
        outer.addWidget(self.secB)

        # === Main horizontal layout for Section B ===
        section_b_layout = QHBoxLayout(self.secB)
        section_b_left = QVBoxLayout()
        section_b_right = QVBoxLayout()

    #    setContentsMargins(left, top, right, bottom)
        BOX_WIDTH = 240          # width of the box
        BOX_HEIGHT = 300         # height of the box
        BOX_OFFSET_FROM_RIGHT = 500   # 👈 INCREASE → moves box LEFT


        section_b_layout.addLayout(section_b_left, 3)   # Left: questions
        section_b_layout.addLayout(section_b_right, 1)  # Right: PN box


        # Reference fields
        self.ref_boxes = {}
        self.ref_radios = {}
        self.ref_labels = {}

        def yes_no_with_box(key, label_text):
                row = QHBoxLayout()

                lbl = QLabel(label_text)
                row.addWidget(lbl)

                grp = QButtonGroup(self)
                rb_y = QRadioButton("Yes")
                rb_n = QRadioButton("No")
                grp.addButton(rb_y)
                grp.addButton(rb_n)

                row.addWidget(rb_y)
                row.addWidget(rb_n)

                txt = QTextEdit()
                txt.setFixedHeight(30)
                txt.setMaximumWidth(500)
                txt.setStyleSheet("background:#FFF2CC; border:none")
                txt.setVisible(False)

                row.addWidget(txt)
                row.addStretch(1)

                rb_y.toggled.connect(lambda c: txt.setVisible(c))
                rb_n.toggled.connect(lambda c: txt.setVisible(False))

                self.ref_boxes[key] = txt
                self.ref_radios[key] = (rb_y, rb_n)
                self.ref_labels[key] = lbl

                section_b_left.addLayout(row)


        yes_no_with_box("PCR_PCN", "1.  Does this Project include Product Change Request (PCR)")
        yes_no_with_box("PSN", "             Is there any Associated Product Safety Note (PSN)")
        yes_no_with_box("PCR", "2.  Associated Project created for PCR")
        yes_no_with_box("SPS", "3.  Associated Open SPSs")
        yes_no_with_box("ESW", "4.  Associated ESWs")
        yes_no_with_box("REF_ECR", "5.  Reference ECR Number(s)")


        # PCR → PSN dependency
        pcr_yes, pcr_no = self.ref_radios["PCR_PCN"]
        psn_yes, psn_no = self.ref_radios["PSN"]
        psn_txt = self.ref_boxes["PSN"]
        psn_lbl = self.ref_labels["PSN"]

        def disable_psn():
                # Temporarily disable auto-exclusive to allow both radios to be unchecked
                grp_psn = psn_yes.group()
                if grp_psn:
                    grp_psn.setExclusive(False)
                
                # Block signals and uncheck both
                psn_yes.blockSignals(True)
                psn_no.blockSignals(True)
                psn_yes.setChecked(False)
                psn_no.setChecked(False)
                psn_yes.blockSignals(False)
                psn_no.blockSignals(False)
                
                # Re-enable auto-exclusive
                if grp_psn:
                    grp_psn.setExclusive(True)
                
                # Disable and hide
                psn_yes.setEnabled(False)
                psn_no.setEnabled(False)
                psn_lbl.setEnabled(False)
                psn_txt.setPlainText("")
                psn_txt.setVisible(False)

        def enable_psn():
                psn_yes.setEnabled(True)
                psn_no.setEnabled(True)
                psn_lbl.setEnabled(True)

        pcr_no.toggled.connect(lambda c: disable_psn() if c else None)
        pcr_yes.toggled.connect(lambda c: enable_psn() if c else None)


        # Reference Email Attachments
        rowm = QHBoxLayout()
        rowm.addWidget(QLabel("6.  Reference e-mail/attachments?"))

        grp_m = QButtonGroup(self)
        rb_my = QRadioButton("Yes")
        rb_mn = QRadioButton("No")
        grp_m.addButton(rb_my)
        grp_m.addButton(rb_mn)
        self.ref_attachment_yes_radio = rb_my
        self.ref_attachment_no_radio = rb_mn

        rowm.addWidget(rb_my)
        rowm.addWidget(rb_mn)

        browse = QPushButton("Browse Attachment")
        browse.setFixedHeight(24)
        browse.setVisible(False)

        rowm.addWidget(browse)
        rowm.addStretch(1)
        section_b_left.addLayout(rowm)

        rb_my.toggled.connect(self._on_reference_attachments_toggled)
        browse.clicked.connect(self._on_browse_email_clicked)
        self._browse_email_btn = browse  # keep reference to update button label


        # Impact caused by
        row = QHBoxLayout()
        row.addWidget(QLabel("7.  Impact caused by:"))

        grp_sc = QButtonGroup(self)
        rb_sup = QRadioButton("Supplier")
        rb_cust = QRadioButton("Customer")
        grp_sc.addButton(rb_sup)
        grp_sc.addButton(rb_cust)
        self.impact_supplier_radio = rb_sup
        self.impact_customer_radio = rb_cust

        row.addWidget(rb_sup)
        row.addWidget(rb_cust)
        row.addStretch(1)
        section_b_left.addLayout(row)

        sub_row = QHBoxLayout()
        sub_row.setContentsMargins(20, 0, 0, 0)

        grp_c = QButtonGroup(self)
        rb_int = QRadioButton("Internal")
        rb_ext = QRadioButton("External")
        grp_c.addButton(rb_int)
        grp_c.addButton(rb_ext)
        self.impact_internal_radio = rb_int
        self.impact_external_radio = rb_ext

        sub_row.addWidget(rb_int)
        sub_row.addWidget(rb_ext)
        sub_row.addStretch(1)
        section_b_left.addLayout(sub_row)

        def clear_customer_scope_selection():
            # Temporarily disable exclusivity so both radios can be fully unchecked.
            grp_c.setExclusive(False)
            rb_int.setChecked(False)
            rb_ext.setChecked(False)
            grp_c.setExclusive(True)

        rb_sup.toggled.connect(
            lambda c: (
                clear_customer_scope_selection(),
                rb_int.setEnabled(False),
                rb_ext.setEnabled(False)
            ) if c else None
        )

        rb_cust.toggled.connect(
            lambda c: (
                rb_int.setEnabled(c),
                rb_ext.setEnabled(c)
            )
        )


        yes_no_with_box("QN", "8. Is there any QN")

        # Reason Code
        rc_row = QHBoxLayout()
        rc_row.setSpacing(8)
        rc_row.addWidget(QLabel("9.  ECR Reason Code:"))
        self.reason_cb = QComboBox()
        self.reason_cb.addItems([
            "Beyond Spec Request","Cap Code Management","CES","Cost Reduction",
            "Design Correction","Document Correction","Manufacturing Improvement",
            "Obsolescence","Option Reduction and Product End of Life","Order BOM Change",
            "Product Improvement","Product Release","Safety Event"
        ])
        rc_row.addWidget(self.reason_cb)

        # Proposed Reason Code – read-only label next to the dropdown
        ai_prop_lbl = QLabel("Proposed Reason Code:")
        ai_prop_lbl.setStyleSheet("color:#1F4E79; font-weight:600;")
        rc_row.addWidget(ai_prop_lbl)

        self.ai_reason_code_edit = QLineEdit()
        self.ai_reason_code_edit.setReadOnly(True)
        self.ai_reason_code_edit.setPlaceholderText("")
        self.ai_reason_code_edit.setFixedHeight(26)
        self.ai_reason_code_edit.setMinimumWidth(180)
        self.ai_reason_code_edit.setStyleSheet(
            "background:#E8F5E9; border:1px solid #81C784; border-radius:4px; "
            "color:#1B5E20; font-weight:bold; padding:2px 6px;"
        )
        rc_row.addWidget(self.ai_reason_code_edit)

        # ⓘ info button – opens Scenario & Examples dialog for the proposed Reason Code
        self.ai_justification_btn = QPushButton("\u24d8")
        self.ai_justification_btn.setFixedSize(22, 22)
        self.ai_justification_btn.setStyleSheet(
            "QPushButton { background:#1F4E79; color:white; border-radius:11px; "
            "font-weight:bold; font-size:12px; border:none; }"
            "QPushButton:hover { background:#2E75B6; }"
        )
        self.ai_justification_btn.setToolTip("Click to see Scenario and Examples for the proposed Reason Code")
        self.ai_justification_btn.setCursor(Qt.CursorShape.WhatsThisCursor)
        self.ai_justification_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.ai_justification_btn.clicked.connect(self._show_rc_info_dialog)
        rc_row.addWidget(self.ai_justification_btn)

        rc_row.addStretch(1)
        section_b_left.addLayout(rc_row)


        scope_lbl = QLabel("Initial Scope Part Numbers")
        scope_lbl.setStyleSheet("font-weight: bold;")

        self.scope_parts_txt = QTextEdit()
        self.scope_parts_txt.setPlaceholderText(
                "Enter Part Number (up to 500 lines)"
        )

        # --- Size tuning ---
        self.scope_parts_txt.setFixedWidth(235)      # ~10% narrower
        self.scope_parts_txt.setMaximumHeight(280)   # slightly taller

        # --- Scrollbars ---
        self.scope_parts_txt.setVerticalScrollBarPolicy(
                Qt.ScrollBarPolicy.ScrollBarAsNeeded
        )
        self.scope_parts_txt.setHorizontalScrollBarPolicy(
                Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )

        # --- Styling (box + scrollbar) ---
        self.scope_parts_txt.setStyleSheet("""
                QTextEdit {
                        background: #F8F8F8;
                        border: 1px solid #999;
                }
                QScrollBar:vertical {
                        background: #E6E6E6;
                        width: 10px;
                        margin: 0px;
                }
                QScrollBar::handle:vertical {
                        background: #A6A6A6;
                        min-height: 20px;
                        border-radius: 4px;
                }
                QScrollBar::handle:vertical:hover {
                        background: #8C8C8C;
                }
                QScrollBar::add-line:vertical,
                QScrollBar::sub-line:vertical {
                        height: 0px;
                }
        """)

        
        # Keep this margin small so Section B stays inside smaller displays.
        section_b_right.setContentsMargins(0, 0, 24, 0)
        section_b_right.addWidget(scope_lbl)
        section_b_right.addWidget(self.scope_parts_txt)
        section_b_right.addSpacing(10)

        include_tabs_lbl = QLabel("Include Data From Tabs")
        include_tabs_lbl.setStyleSheet("font-weight: bold;")
        section_b_right.addWidget(include_tabs_lbl)

        self.cb_where_used = QCheckBox("Where Used")
        self.cb_obs_parts = QCheckBox("OBS Parts")
        self.cb_structure_sheet = QCheckBox("Structure Sheet")

        section_b_right.addWidget(self.cb_where_used)
        section_b_right.addWidget(self.cb_obs_parts)
        section_b_right.addWidget(self.cb_structure_sheet)

        self.selected_tab_payload = {}

        self.cb_where_used.stateChanged.connect(self.refresh_selected_tab_payload)
        self.cb_obs_parts.stateChanged.connect(self.refresh_selected_tab_payload)
        self.cb_structure_sheet.stateChanged.connect(self.refresh_selected_tab_payload)

        # Keep Include-Data checkboxes aligned with source-tab data availability.
        self._include_tabs_sync_timer = QTimer(self)
        self._include_tabs_sync_timer.setInterval(1200)
        self._include_tabs_sync_timer.timeout.connect(self._update_include_tab_checkbox_enablement)
        self._include_tabs_sync_timer.start()
        QTimer.singleShot(0, self._update_include_tab_checkbox_enablement)

        self.btn_change_summary = QPushButton("Change Summary")
        self.btn_change_summary.setFixedHeight(26)
        self.btn_change_summary.setToolTip(
            "Generate row-wise change summary from Structure Sheet and optionally export"
        )
        self.btn_change_summary.clicked.connect(self.on_change_summary_clicked)
        section_b_right.addWidget(self.btn_change_summary)

        section_b_right.addStretch(1)
       
        # Short Title
        title_row = QHBoxLayout()
        title_row.addWidget(highlight_label("Short Title"))
        title_row.addStretch(1)
        section_b_left.addLayout(title_row)

        self.short_title_edit = QLineEdit()
        self.short_title_edit.setPlaceholderText("Enter short title")
        self.short_title_edit.setFixedHeight(30)
        self.short_title_edit.setMaximumWidth(900)
        section_b_left.addWidget(self.short_title_edit)

        section_b_left.addSpacing(6)

    # Problem Summary
        ps_row = QHBoxLayout()
        ps_row.addWidget(
                highlight_label("Problem Summary from PCR, PCN, SPS and ESW")
        )
        self.btn_ps = QPushButton("Generate Change Summary")
        self.btn_ps.setFixedSize(180, 26)
        ps_row.addWidget(self.btn_ps)
        self.btn_problem_regen = QPushButton("Regenerate Problem Statement")
        self.btn_problem_regen.setFixedSize(250, 26)
        self.btn_problem_regen.setToolTip(
            "Regenerates only the Problem Statement while keeping other fields unchanged."
        )
        ps_row.addWidget(self.btn_problem_regen)
        ps_row.addStretch(1)

        section_b_left.addLayout(ps_row)

        self.problem_txt = LimitedTextEdit(None)
        self.problem_txt.setMinimumHeight(260)
        self.problem_txt.setMaximumWidth(900)
        section_b_left.addWidget(self.problem_txt)

        self.btn_ps.clicked.connect(self.on_problem_summary_clicked)
        self.btn_problem_regen.clicked.connect(self.on_regenerate_problem_statement_clicked)

        section_b_left.addSpacing(6)




        # Solution
        sol_row = QHBoxLayout()
        sol_row.addWidget(highlight_label("Proposed Solution"))
        self.btn_sol = QPushButton("Regenerate Proposed Solution")
        self.btn_sol.setFixedSize(220, 26)
        self.btn_sol.setToolTip(
            "Clears the existing solution and regenerates a new solution using AI."
        )
        sol_row.addWidget(self.btn_sol)
        sol_row.addStretch(1)

        section_b_left.addLayout(sol_row)

        self.solution_txt = LimitedTextEdit(None)
        self.solution_txt.setMinimumHeight(260)
        self.solution_txt.setMaximumWidth(900)

        section_b_left.addWidget(self.solution_txt)

        self.btn_sol.clicked.connect(self.on_regenerate_solution_clicked)

        outer.addWidget(self.secB)
        outer.addStretch(1)


    def refresh_selected_tab_payload(self):
        self._update_include_tab_checkbox_enablement()
        main = self.window()
        payload = {}

        if self.cb_where_used.isChecked():
            whereused_tab = getattr(main, "whereused_tab", None)
            if whereused_tab and hasattr(whereused_tab, "table"):
                payload["Where Used"] = self._table_to_rows(whereused_tab.table)

        if self.cb_obs_parts.isChecked():
            obs_tab = getattr(main, "obs_tab", None)
            if obs_tab and hasattr(obs_tab, "table"):
                payload["OBS Parts"] = self._table_to_rows(obs_tab.table)

        if self.cb_structure_sheet.isChecked():
            structure_tab = getattr(main, "structure_tab", None)
            if structure_tab:
                structure_data = {}

                impacted_text = getattr(structure_tab, "impacted_text", None)
                if impacted_text:
                    structure_data["Impacted Options/Parts"] = [
                        line.strip()
                        for line in impacted_text.toPlainText().splitlines()
                        if line.strip()
                    ]

                structure_table = getattr(structure_tab, "table", None)
                if structure_table:
                    structure_data["Structure Sheet"] = self._table_to_rows(structure_table)

                payload["Structure Sheet"] = structure_data

        self.selected_tab_payload = payload

    def _find_table_col_index(self, table, header_names: List[str]) -> int:
        if table is None:
            return -1

        targets = {' '.join((h or '').strip().lower().split()) for h in header_names}
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            if not h:
                continue
            nh = ' '.join((h.text() or '').strip().lower().split())
            if nh in targets:
                return c
        return -1

    def _table_has_nonempty_col_value(self, table, col_index: int) -> bool:
        if table is None or col_index < 0:
            return False
        for r in range(table.rowCount()):
            it = table.item(r, col_index)
            txt = (it.text() if it else '') or ''
            if txt.strip():
                return True
        return False

    def _has_where_used_data(self) -> bool:
        main = self.window()
        whereused_tab = getattr(main, 'whereused_tab', None)
        table = getattr(whereused_tab, 'table', None) if whereused_tab else None
        part_col = self._find_table_col_index(table, ['Part'])
        return self._table_has_nonempty_col_value(table, part_col)

    def _has_obs_parts_data(self) -> bool:
        main = self.window()
        obs_tab = getattr(main, 'obs_tab', None)
        table = getattr(obs_tab, 'table', None) if obs_tab else None
        obs_col = self._find_table_col_index(table, ['OBS Parts', 'OBS Part', 'Part'])
        return self._table_has_nonempty_col_value(table, obs_col)

    def _has_structure_sheet_data(self) -> bool:
        main = self.window()
        structure_tab = getattr(main, 'structure_tab', None)
        table = getattr(structure_tab, 'table', None) if structure_tab else None
        part_col = self._find_table_col_index(table, ['Part'])
        return self._table_has_nonempty_col_value(table, part_col)

    def _set_checkbox_enabled_by_data(self, checkbox: QCheckBox, has_data: bool):
        checkbox.blockSignals(True)
        try:
            checkbox.setEnabled(has_data)
            if not has_data:
                checkbox.setChecked(False)
        finally:
            checkbox.blockSignals(False)

    def _update_include_tab_checkbox_enablement(self):
        self._set_checkbox_enabled_by_data(self.cb_where_used, self._has_where_used_data())
        self._set_checkbox_enabled_by_data(self.cb_obs_parts, self._has_obs_parts_data())
        self._set_checkbox_enabled_by_data(self.cb_structure_sheet, self._has_structure_sheet_data())

    def _collect_obs_parts_for_summary(self) -> Dict[str, List[str]]:
        """Return grouped OBS summary lines by 4 requested cases."""
        main = self.window()
        obs_tab = getattr(main, 'obs_tab', None)
        table = getattr(obs_tab, 'table', None) if obs_tab else None
        if table is None:
            return {
                'obsolete_with_repl': [],
                'obsolete_without_repl': [],
                'inactivate_with_repl': [],
                'inactivate_without_repl': [],
            }

        obs_col = self._find_table_col_index(table, ['OBS Parts', 'OBS Part', 'Part'])
        change_col = self._find_table_col_index(table, ['Change'])
        repl_col = self._find_table_col_index(table, ['Replacement'])

        grouped = {
            'obsolete_with_repl': [],
            'obsolete_without_repl': [],
            'inactivate_with_repl': [],
            'inactivate_without_repl': [],
        }

        if obs_col < 0:
            return grouped

        for r in range(table.rowCount()):
            obs_it = table.item(r, obs_col)
            obs_part = ((obs_it.text() if obs_it else '') or '').strip()
            if not obs_part:
                continue

            change_txt = 'Obsolete'
            if change_col >= 0:
                w = table.cellWidget(r, change_col)
                if isinstance(w, QComboBox):
                    change_txt = (w.currentText() or '').strip() or 'Obsolete'
                else:
                    ch_it = table.item(r, change_col)
                    change_txt = ((ch_it.text() if ch_it else '') or '').strip() or 'Obsolete'

            repl_txt = ''
            if repl_col >= 0:
                repl_it = table.item(r, repl_col)
                repl_txt = ((repl_it.text() if repl_it else '') or '').strip()

            has_repl = bool(repl_txt) and ('no replacement' not in repl_txt.lower())
            is_inactivate = change_txt.strip().lower().startswith('inactivate')

            if is_inactivate and has_repl:
                grouped['inactivate_with_repl'].append(f"{obs_part} with {repl_txt}")
            elif is_inactivate and not has_repl:
                grouped['inactivate_without_repl'].append(obs_part)
            elif (not is_inactivate) and has_repl:
                grouped['obsolete_with_repl'].append(f"{obs_part} with {repl_txt}")
            else:
                grouped['obsolete_without_repl'].append(obs_part)

        return grouped

    def _build_obs_parts_summary_lines(self) -> List[str]:
        grouped = self._collect_obs_parts_for_summary()
        total = sum(len(v) for v in grouped.values())
        if total == 0:
            return []

        lines = ['Disabled Parts List from SmBOM:', '']

        def _append_case(title: str, items: List[str]):
            if not items:
                return
            lines.append(title)
            for it in items:
                lines.append(f"     {it}")
            lines.append('')

        _append_case('Obsolete below parts with replacement:', grouped['obsolete_with_repl'])
        _append_case('Obsolete below parts without replacement:', grouped['obsolete_without_repl'])
        _append_case('Inactivate below parts with replacement:', grouped['inactivate_with_repl'])
        _append_case('Inactivate below parts without replacement:', grouped['inactivate_without_repl'])

        while lines and lines[-1] == '':
            lines.pop()
        return lines

    def get_selected_tab_payload(self):
        self.refresh_selected_tab_payload()
        return self.selected_tab_payload

    def _table_to_rows(self, table):
        if table is None or table.columnCount() == 0:
            return []

        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text().strip() if h and h.text() else f"Column {c + 1}")

        rows = []
        for r in range(table.rowCount()):
            row_data = {}
            has_value = False

            for c, header in enumerate(headers):
                value = ""
                cell_widget = table.cellWidget(r, c)

                if isinstance(cell_widget, QComboBox):
                    value = cell_widget.currentText().strip()
                elif isinstance(cell_widget, QTextEdit):
                    value = cell_widget.toPlainText().strip()
                elif isinstance(cell_widget, QWidget):
                    chk = cell_widget.findChild(QCheckBox)
                    if chk:
                        value = "Yes" if chk.isChecked() else "No"
                    else:
                        item = table.item(r, c)
                        value = item.text().strip() if item and item.text() else ""
                else:
                    item = table.item(r, c)
                    value = item.text().strip() if item and item.text() else ""

                if value:
                    has_value = True
                # Preserve values when duplicate header captions exist in the table.
                # This is important for selector columns (Yes/No) that may share a
                # caption with a later data column.
                if header in row_data:
                    existing = (row_data.get(header, '') or '').strip().lower()
                    incoming = (value or '').strip().lower()

                    # Keep checkbox state if already captured and later duplicate is blank.
                    if existing in {'yes', 'no'} and incoming == '':
                        continue

                    # Keep first non-empty value when later duplicate is blank.
                    if row_data.get(header, '') and not value:
                        continue

                    # Store duplicate under a unique key so no data is lost.
                    dup_idx = 2
                    dup_key = f"{header} ({dup_idx})"
                    while dup_key in row_data:
                        dup_idx += 1
                        dup_key = f"{header} ({dup_idx})"
                    row_data[dup_key] = value
                else:
                    row_data[header] = value

            if has_value:
                rows.append(row_data)

        return rows

    def on_change_summary_clicked(self):
        self._update_include_tab_checkbox_enablement()

        if not (
            self.cb_where_used.isChecked()
            or self.cb_obs_parts.isChecked()
            or self.cb_structure_sheet.isChecked()
        ):
            QMessageBox.information(
                self,
                'Change Summary',
                'Update the Where used, OBS Parts and Structure Sheet tabs and select thecheck boxes and re-run.',
            )
            return

        # Prompt for OBS inclusion if data exists but checkbox is not selected.
        if self._has_obs_parts_data() and not self.cb_obs_parts.isChecked() and self.cb_obs_parts.isEnabled():
            reply = QMessageBox.question(
                self,
                'Include OBS Parts',
                'OBS Parts data is available but not selected.\n\n'
                'Do you want to include OBS Parts in the Change Summary (PSS)?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.cb_obs_parts.setChecked(True)

        if generate_change_summary_sentences is None:
            QMessageBox.warning(
                self,
                'Change Summary',
                'PSS_Change_Summary.py is not available. Please ensure the file exists.',
            )
            return

        sentences = []
        if self.cb_structure_sheet.isChecked():
            main = self.window()
            structure_tab = getattr(main, 'structure_tab', None)
            structure_table = getattr(structure_tab, 'table', None) if structure_tab else None
            if structure_table is None:
                QMessageBox.information(
                    self,
                    'Change Summary',
                    'Structure Sheet tab is not available yet.',
                )
                return

            structure_rows = self._table_to_rows(structure_table)
            if not structure_rows:
                QMessageBox.information(
                    self,
                    'Change Summary',
                    'Structure Sheet has no data. Please add/import data first.',
                )
                return

            # Validate Remove Item rows before generating summary.
            def _norm_local(text: str) -> str:
                return ''.join(ch for ch in (text or '').strip().lower() if ch.isalnum())

            invalid_by_parent = {}
            current_parent = ''
            for row in structure_rows:
                action_val = ''
                repl_val = ''
                part_val = ''
                bom_val = ''
                for k, v in row.items():
                    nk = _norm_local(k)
                    if nk in {'action', 'changetype'}:
                        action_val = (str(v) if v is not None else '').strip()
                    elif nk == 'replacement':
                        repl_val = (str(v) if v is not None else '').strip()
                    elif nk == 'part':
                        part_val = (str(v) if v is not None else '').strip()
                    elif nk in {'bomlevel', 'wulevel', 'level'}:
                        bom_val = (str(v) if v is not None else '').strip()

                try:
                    bom_int = int(float(bom_val)) if bom_val != '' else -1
                except Exception:
                    bom_int = -1

                if bom_int == 0 and part_val:
                    current_parent = part_val

                if _norm_local(action_val).startswith('removeitem') and repl_val:
                    parent_key = current_parent or '(unknown parent)'
                    child_key = part_val or '(blank)'
                    invalid_by_parent.setdefault(parent_key, [])
                    if child_key not in invalid_by_parent[parent_key]:
                        invalid_by_parent[parent_key].append(child_key)

            if invalid_by_parent:
                parents = list(invalid_by_parent.keys())

                grouped_lines = []
                for p in parents[:6]:
                    items = invalid_by_parent.get(p, [])
                    item_preview = ', '.join(items[:12])
                    item_more = '' if len(items) <= 12 else f' and {len(items) - 12} more'
                    grouped_lines.append(f'Affected Parent: {p}\nAffected BOM Items: {item_preview}{item_more}')

                if len(parents) > 6:
                    grouped_lines.append(f'...and {len(parents) - 6} more parent group(s).')

                QMessageBox.warning(
                    self,
                    'Change Summary Validation',
                    'Remove replacement part numbers for items marked as "Remove Item" and re-run.'
                    + "\n\n"
                    + '\n\n'.join(grouped_lines),
                )
                return

            try:
                sentences = generate_change_summary_sentences(structure_rows)
            except Exception as e:
                QMessageBox.warning(self, 'Change Summary Error', str(e))
                return

            if not sentences:
                QMessageBox.information(
                    self,
                    'Change Summary',
                    'No change-summary lines could be generated from the current Structure Sheet rows.',
                )
                return

        # Build required output block and append it inside Proposed Solution.
        block_lines = ['Change Summary:-'] + sentences
        obs_lines = []
        if self.cb_obs_parts.isChecked():
            obs_lines = self._build_obs_parts_summary_lines()
            if obs_lines:
                block_lines += ['', ''] + obs_lines
        if not sentences and not obs_lines:
            QMessageBox.information(
                self,
                'Change Summary',
                'No change-summary content available for the selected tabs.',
            )
            return
        import html as _html

        def _line_to_html(line: str) -> str:
            raw = line or ''
            # Preserve left indentation visually in rich text.
            lead_spaces = len(raw) - len(raw.lstrip(' '))
            prefix = '&nbsp;' * lead_spaces

            body = raw.lstrip(' ')
            # Support markdown and plain-token forms for From/To in change-summary lines.
            body = re.sub(r"\*\*(from\s*:?)\*\*", r"<b>\1</b>", body, flags=re.IGNORECASE)
            body = re.sub(r"\*\*(to\s*:?)\*\*", r"<b>\1</b>", body, flags=re.IGNORECASE)
            body = re.sub(r"(?<!\w)(from\s*:)(?!\w)", r"<b>\1</b>", body, flags=re.IGNORECASE)
            body = re.sub(r"(?<!\w)(to\s*:)(?!\w)", r"<b>\1</b>", body, flags=re.IGNORECASE)

            # Escape first, then allow intended <b> tags from generator and local transforms.
            escaped = _html.escape(body)
            escaped = escaped.replace('&lt;b&gt;', '<b>').replace('&lt;/b&gt;', '</b>')
            return prefix + escaped

        block_html = '<br>'.join(_line_to_html(ln) for ln in block_lines)

        existing_text = self.solution_txt.toPlainText().strip()
        if existing_text:
            self.solution_txt.insertHtml('<br><br><br>')
        self.solution_txt.insertHtml(block_html)

        for line in sentences:
            print(line)
        return
# ...existing code...

    def _selected_radio_text(self, key: str) -> str:
        pair = self.ref_radios.get(key)
        if not pair:
            return ""
        rb_y, rb_n = pair
        if rb_y.isChecked():
            return "Yes"
        if rb_n.isChecked():
            return "No"
        return ""

# ...existing code...
    def _collect_pss_payload(self) -> Dict[str, Any]:
        self.refresh_selected_tab_payload()

        scope_parts = [
            ln.strip()
            for ln in self.scope_parts_txt.toPlainText().splitlines()
            if ln.strip()
        ]

        reference_inputs = {}
        for key, txt in self.ref_boxes.items():
            reference_inputs[key] = {
                "answer": self._selected_radio_text(key),   # "Yes" / "No" / ""
                "text": txt.toPlainText().strip(),
            }

        return {
            "short_title": self.short_title_edit.text().strip(),
            "reason_code": self.reason_cb.currentText().strip(),
            "scope_parts": scope_parts,
            "reference_inputs": reference_inputs,
            "selected_tab_payload": self.selected_tab_payload,
            "current_problem_text": self.problem_txt.toPlainText().strip(),
            "proposed_solution_text": self.solution_txt.toPlainText().strip(),
            "solution_regeneration_requested": bool(
                getattr(self, "_solution_regen_requested", False)
            ),
            "solution_regeneration_reason": (
                "Existing solution is incomplete or incorrect. "
                "Regenerate a complete, engineer-quality solution."
                if getattr(self, "_solution_regen_requested", False)
                else ""
            ),
            "include_tabs_flags": {
                "where_used": self.cb_where_used.isChecked(),
                "obs_parts": self.cb_obs_parts.isChecked(),
                "structure_sheet": self.cb_structure_sheet.isChecked(),
            },
        }
# ...existing code...

    def _load_ai_pss_module(self):
        mod_path = Path(__file__).with_name("AI_Assisted_PSS.py")
        if not mod_path.exists():
            raise FileNotFoundError(f"AI_Assisted_PSS.py not found: {mod_path}")

        spec = importlib.util.spec_from_file_location("AI_Assisted_PSS", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load AI_Assisted_PSS module spec")

        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    # ...existing code...

    def _normalize_pss_output(self, out: Any) -> Dict[str, str]:
        if isinstance(out, dict):
            title = str(out.get("title") or "").strip()
            problem = str(
                out.get("problem_statement")
                or out.get("problem_summary")
                or out.get("problem")
                or ""
            ).strip()
            solution = str(
                out.get("solution_statement")
                or out.get("solution")
                or ""
            ).strip()
            return {"title": title, "problem": problem, "solution": solution}

        text = str(out or "").strip()
        return {"title": "", "problem": text, "solution": ""}

    def _run_ai_pss_full(self, payload: Dict[str, Any]) -> Dict[str, str]:
        module = self._load_ai_pss_module()

        fn_full = getattr(module, "generate_full_pss", None)
        if callable(fn_full):
            try:
                out = fn_full(payload)
            except TypeError:
                out = fn_full()
            return self._normalize_pss_output(out)

        for fn_name in ("generate_problem_summary", "generate_pss", "run_problem_summary"):
            fn = getattr(module, fn_name, None)
            if callable(fn):
                try:
                    out = fn(payload)
                except TypeError:
                    out = fn()
                return self._normalize_pss_output(out)

        raise AttributeError(
            "No supported entry point in AI_Assisted_PSS.py. "
            "Expected generate_full_pss(payload) or generate_problem_summary(payload)."
        )

# ...existing code...

# ...existing code...

    def _run_ai_pss(self, payload: Dict[str, Any]) -> str:
        out = self._run_ai_pss_full(payload)
        return out.get("problem", "")

# ...existing code...

        for fn_name in ("generate_problem_summary", "generate_pss", "run_problem_summary"):
            fn = getattr(module, fn_name, None)
            if callable(fn):
                try:
                    out = fn(payload)
                except TypeError:
                    out = fn()

                if isinstance(out, dict):
                    out = out.get("problem_summary") or out.get("problem") or ""
                return str(out or "").strip()

        for cls_name in ("AIAssistedPSS", "ProblemSolutionAgent", "PSSAgent"):
            cls = getattr(module, cls_name, None)
            if cls is None:
                continue
            obj = cls()
            for m_name in ("generate_problem_summary", "generate_pss", "run"):
                method = getattr(obj, m_name, None)
                if callable(method):
                    try:
                        out = method(payload)
                    except TypeError:
                        out = method()

                    if isinstance(out, dict):
                        out = out.get("problem_summary") or out.get("problem") or ""
                    return str(out or "").strip()

        raise AttributeError(
            "No supported entry point in AI_Assisted_PSS.py. "
            "Expected generate_problem_summary(payload) or equivalent."
        )

# ...existing code...

    def _on_reference_attachments_toggled(self, checked: bool):
        """Show/hide browse button and clear selected attachment when toggled off."""
        if hasattr(self, "_browse_email_btn"):
            self._browse_email_btn.setVisible(checked)

        if not checked:
            self.email_file_path = None
            if hasattr(self, "_browse_email_btn"):
                self._browse_email_btn.setText("Browse Attachment")
                self._browse_email_btn.setToolTip("")

    def _on_browse_email_clicked(self):
        """Open file dialog, store selected attachment path, update button label."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Reference Attachment",
            "",
            "Supported Files (*.msg *.eml *.txt *.pdf *.ppt *.pptx *.doc *.docx)"
        )
        if file_path:
            self.email_file_path = file_path
            file_name = Path(file_path).name
            self._browse_email_btn.setText(f"📎 {file_name}")
            self._browse_email_btn.setToolTip(file_path)
            QMessageBox.information(
                self,
                "Attachment Loaded",
                f"Reference attachment loaded:\n{file_name}\n\nClick 'Problem Summary' to process it.",
            )

    def _validate_problem_summary_inputs(self) -> str:
        """Validate ECR change assessment inputs before generating Problem Summary."""
        missing_answers = []
        missing_values = []
        qn_missing_answer = None
        qn_missing_value = None
        pcr_pair = self.ref_radios.get("PCR_PCN")
        pcr_yes_selected = bool(pcr_pair and pcr_pair[0].isChecked())
        field_names = {
            "PCR_PCN": "PCR",
            "PSN": "PSN",
            "PCR": "PCR Project",
            "SPS": "SPS",
            "ESW": "ESW",
            "REF_ECR": "Reference ECR",
            "QN": "QN",
        }

        for key, (rb_yes, rb_no) in self.ref_radios.items():
            # PSN is auto-derived from PCR records when Question 1 (PCR_PCN) is Yes.
            # Do not require manual PSN Yes/No or value in that flow.
            if key == "PSN" and pcr_yes_selected:
                continue

            # Skip questions that are currently disabled by dependency logic (for example PSN).
            if not rb_yes.isEnabled() and not rb_no.isEnabled():
                continue

            if not rb_yes.isChecked() and not rb_no.isChecked():
                label = self.ref_labels.get(key)
                q_text = label.text() if label else key
                if key == "QN":
                    qn_missing_answer = q_text
                else:
                    missing_answers.append(q_text)
                continue

            if rb_yes.isChecked():
                value_widget = self.ref_boxes.get(key)
                value_text = value_widget.toPlainText().strip() if value_widget else ""
                if not value_text:
                    if key == "QN":
                        qn_missing_value = field_names.get(key, key)
                    else:
                        missing_values.append(field_names.get(key, key))

        attachment_yes = getattr(self, "ref_attachment_yes_radio", None)
        attachment_no = getattr(self, "ref_attachment_no_radio", None)
        if attachment_yes and attachment_no:
            if not attachment_yes.isChecked() and not attachment_no.isChecked():
                missing_answers.append(
                    "6. Reference e-mail/attachments"
                )
            elif attachment_yes.isChecked() and not getattr(self, "email_file_path", None):
                missing_answers.append(
                    "6. Reference e-mail/attachments: please attach relevant attachments"
                )

        impact_supplier = getattr(self, "impact_supplier_radio", None)
        impact_customer = getattr(self, "impact_customer_radio", None)
        impact_internal = getattr(self, "impact_internal_radio", None)
        impact_external = getattr(self, "impact_external_radio", None)

        if impact_supplier and impact_customer:
            if not impact_supplier.isChecked() and not impact_customer.isChecked():
                missing_answers.append("7. Impact caused by: select Supplier or Customer")
            elif impact_customer.isChecked() and impact_internal and impact_external:
                if not impact_internal.isChecked() and not impact_external.isChecked():
                    missing_answers.append(
                        "7. Impact caused by Customer: select Internal or External"
                    )

        if qn_missing_answer:
            missing_answers.append(qn_missing_answer)

        if qn_missing_value:
            missing_values.append(qn_missing_value)

        if missing_answers:
            return (
                "Please address Assessment question(s) with Yes/No selection:\n\n"
                + "\n".join(f"{q}" for q in missing_answers)
            )

        if missing_values:
            return (
                "Please provide details in the corresponding field for these 'Yes' selections:\n\n"
                + "\n".join(f"{name}: details/number is missing" for name in missing_values)
            )

        return ""

    # ------------------------------------------------------------------
    # PCR-driven Problem Summary helpers (Question 1 – Databricks path)
    # ------------------------------------------------------------------

    def _load_pcr_query_module(self):
        """Load pcr_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "pcr_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"pcr_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("pcr_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load pcr_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_project_query_module(self):
        """Load project_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "project_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"project_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("project_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load project_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_sps_query_module(self):
        """Load sps_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "sps_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"sps_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("sps_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load sps_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_esw_query_module(self):
        """Load esw_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "esw_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"esw_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("esw_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load esw_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _load_ecr_query_module(self):
        """Load ecr_details_query.py from the parent folder via importlib."""
        mod_path = Path(__file__).parent.parent / "ecr_details_query.py"
        if not mod_path.exists():
            raise FileNotFoundError(
                f"ecr_details_query.py not found at: {mod_path}\n"
                "Ensure the file exists one level above the Problem_Solution_Agent_PSS folder."
            )
        spec = importlib.util.spec_from_file_location("ecr_details_query", str(mod_path))
        if spec is None or spec.loader is None:
            raise ImportError("Unable to load ecr_details_query module spec.")
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def _fetch_q2_project_context(self) -> Dict[str, Any]:
        """Fetch Question 2 (Associated Project created for PCR) records if enabled.

        Returns empty structure when Q2 is not selected as Yes.
        """
        pair = self.ref_radios.get("PCR")
        if not pair:
            self._project_fetch_result = {"skipped": [], "not_found": []}
            return {"project_records": [], "skipped_projects": [], "not_found_projects": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._project_fetch_result = {"skipped": [], "not_found": []}
            return {"project_records": [], "skipped_projects": [], "not_found_projects": []}

        project_text = self.ref_boxes["PCR"].toPlainText().strip()
        if not project_text:
            raise ValueError(
                "Please enter at least one Project number in Question 2 "
                "(Associated Project created for PCR)."
            )

        mod = self._load_project_query_module()
        result = mod.fetch_project_records(project_text)

        project_records = result.get("valid", [])
        skipped_projects = result.get("skipped", [])
        not_found_projects = result.get("not_found", [])

        self._project_fetch_result = {
            "skipped": skipped_projects,
            "not_found": not_found_projects,
        }

        return {
            "project_records": project_records,
            "skipped_projects": skipped_projects,
            "not_found_projects": not_found_projects,
        }

    def _fetch_q3_sps_context(self) -> Dict[str, Any]:
        """Fetch Question 3 (Associated Open SPSs) records if enabled.

        Returns empty structure when Q3 is not selected as Yes.
        """
        pair = self.ref_radios.get("SPS")
        if not pair:
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            return {"sps_records": [], "skipped_sps": [], "not_found_sps": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            return {"sps_records": [], "skipped_sps": [], "not_found_sps": []}

        sps_text = self.ref_boxes["SPS"].toPlainText().strip()
        if not sps_text:
            raise ValueError(
                "Please enter at least one SPS number in Question 3 "
                "(Associated Open SPSs)."
            )

        mod = self._load_sps_query_module()
        result = mod.fetch_sps_records(sps_text)

        sps_records = result.get("valid", [])
        skipped_sps = result.get("skipped", [])
        not_found_sps = result.get("not_found", [])

        self._sps_fetch_result = {
            "skipped": skipped_sps,
            "not_found": not_found_sps,
        }

        return {
            "sps_records": sps_records,
            "skipped_sps": skipped_sps,
            "not_found_sps": not_found_sps,
        }

    def _fetch_q4_esw_context(self) -> Dict[str, Any]:
        """Fetch Question 4 (Associated ESWs) records if enabled.

        Returns empty structure when Q4 is not selected as Yes.
        """
        pair = self.ref_radios.get("ESW")
        if not pair:
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            return {"esw_records": [], "skipped_esw": [], "not_found_esw": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            return {"esw_records": [], "skipped_esw": [], "not_found_esw": []}

        esw_text = self.ref_boxes["ESW"].toPlainText().strip()
        if not esw_text:
            raise ValueError(
                "Please enter at least one ESW number in Question 4 "
                "(Associated ESWs)."
            )

        mod = self._load_esw_query_module()
        result = mod.fetch_esw_records(esw_text)

        esw_records = result.get("valid", [])
        # Keep an alias so downstream code that expects ec_number can still read ESW IDs.
        for rec in esw_records:
            if rec.get("ec_number") is None and rec.get("esw_number") is not None:
                rec["ec_number"] = rec.get("esw_number")
        skipped_esw = result.get("skipped", [])
        not_found_esw = result.get("not_found", [])

        self._esw_fetch_result = {
            "skipped": skipped_esw,
            "not_found": not_found_esw,
        }

        return {
            "esw_records": esw_records,
            "skipped_esw": skipped_esw,
            "not_found_esw": not_found_esw,
        }

    def _fetch_q5_ecr_context(self) -> Dict[str, Any]:
        """Fetch Question 5 (Reference ECR Numbers) records if enabled.

        Connects to Databricks via ODBC, reads problem and solution for each
        ECR number entered, stores results in a transient instance attribute
        (_ecr_fetch_result) and returns them for inclusion in the AI payload.

        Returns empty structure when Q5 is not selected as Yes.
        """
        pair = self.ref_radios.get("REF_ECR")
        if not pair:
            self._ecr_fetch_result = {"skipped": [], "not_found": []}
            return {"ecr_records": [], "skipped_ecr": [], "not_found_ecr": []}

        rb_yes, _rb_no = pair
        if not rb_yes.isChecked():
            self._ecr_fetch_result = {"skipped": [], "not_found": []}
            return {"ecr_records": [], "skipped_ecr": [], "not_found_ecr": []}

        ecr_text = self.ref_boxes["REF_ECR"].toPlainText().strip()
        if not ecr_text:
            raise ValueError(
                "Please enter at least one ECR number in Question 5 "
                "(Reference ECR Number(s))."
            )

        mod = self._load_ecr_query_module()
        result = mod.fetch_ecr_records(ecr_text)

        ecr_records = result.get("valid", [])
        skipped_ecr = result.get("skipped", [])
        not_found_ecr = result.get("not_found", [])

        # Store transient fetch metadata for surface warnings after generation.
        self._ecr_fetch_result = {
            "skipped": skipped_ecr,
            "not_found": not_found_ecr,
        }

        return {
            "ecr_records": ecr_records,
            "skipped_ecr": skipped_ecr,
            "not_found_ecr": not_found_ecr,
        }

    def _append_project_context(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Enrich payload with Q2 project, Q3 SPS, Q4 ESW and Q5 ECR context when applicable."""
        enriched = dict(payload or {})
        project_ctx = self._fetch_q2_project_context()
        sps_ctx = self._fetch_q3_sps_context()
        esw_ctx = self._fetch_q4_esw_context()
        ecr_ctx = self._fetch_q5_ecr_context()
        enriched.update(project_ctx)
        enriched.update(sps_ctx)
        enriched.update(esw_ctx)
        enriched.update(ecr_ctx)
        return enriched

    def _extract_numeric_ids(self, text: str) -> List[int]:
        """Extract numeric IDs from free-form text preserving first-seen order."""
        raw = re.findall(r"\d+", text or "")
        if not raw:
            return []
        return list(dict.fromkeys(int(x) for x in raw))

    def _fetch_existing_ec_numbers(self, ids: List[int]) -> List[int]:
        """Return IDs that exist in tbl_projectx_ec.ec_number."""
        if not ids:
            return []

        try:
            import pyodbc
        except Exception as exc:
            raise RuntimeError(f"pyodbc import failed: {exc}") from exc

        # Databricks Hive SQL does not support positional ? parameters; inline validated ints.
        placeholders = ", ".join(str(n) for n in ids)
        sql = f"""
        SELECT DISTINCT ec.ec_number
        FROM prd.rd_core.tbl_projectx_ec ec
        WHERE ec.ec_number IN ({placeholders})
        """

        conn = pyodbc.connect("DSN=Spark-PRD", autocommit=True)
        try:
            cursor = conn.cursor()
            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
            finally:
                cursor.close()
        finally:
            conn.close()

        found = []
        for row in rows:
            if row and row[0] is not None:
                try:
                    found.append(int(row[0]))
                except Exception:
                    continue
        return list(dict.fromkeys(found))

    def _fetch_ec_records(self, ids: List[int]) -> List[Dict[str, Any]]:
        """Return EC records with problem/solution/status for the provided IDs."""
        if not ids:
            return []

        try:
            import pyodbc
        except Exception as exc:
            raise RuntimeError(f"pyodbc import failed: {exc}") from exc

        # Databricks Hive SQL does not support positional ? parameters; inline validated ints.
        placeholders = ", ".join(str(n) for n in ids)
        sql = f"""
        SELECT
            ec.ec_number,
            ec.ec_problem,
            ec.ec_solution,
            st.ec_status
        FROM prd.rd_core.tbl_projectx_ec ec
        LEFT JOIN prd.rd_core.tbl_projectx_ec_status st
            ON ec.ec_status_id = st.ec_status_id
        WHERE ec.ec_number IN ({placeholders})
        ORDER BY ec.ec_number
        """

        conn = pyodbc.connect("DSN=Spark-PRD", autocommit=True)
        try:
            cursor = conn.cursor()
            try:
                cursor.execute(sql)
                rows = cursor.fetchall()
            finally:
                cursor.close()
        finally:
            conn.close()

        records: List[Dict[str, Any]] = []
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                ec_number = int(row[0])
            except Exception:
                continue
            records.append({
                "ec_number": ec_number,
                "problem": str(row[1]).strip() if row[1] is not None else "",
                "solution": str(row[2]).strip() if row[2] is not None else "",
                "status": str(row[3]).strip() if row[3] is not None else "",
            })
        return records

    def _fetch_selected_ec_reference_context(self) -> Dict[str, Any]:
        """Fetch detailed EC context for yes-selected reference fields."""
        mapping = {
            "SPS": "sps_records",
            "ESW": "esw_records",
            "REF_ECR": "reference_ecr_records",
            "QN": "qn_records",
        }
        context: Dict[str, Any] = {value: [] for value in mapping.values()}

        for key, payload_key in mapping.items():
            pair = self.ref_radios.get(key)
            if not pair:
                continue
            rb_yes, _rb_no = pair
            if not rb_yes.isChecked():
                continue

            text_widget = self.ref_boxes.get(key)
            field_text = text_widget.toPlainText().strip() if text_widget else ""
            ids = self._extract_numeric_ids(field_text)
            if not ids:
                continue

            context[payload_key] = self._fetch_ec_records(ids)

        return context

    def _collect_not_found_verification_items(
        self,
        pcr_not_found: List[int],
        project_not_found: List[int],
        sps_not_found: List[int],
        esw_not_found: List[int],
        ecr_not_found: List[int] = None,
    ) -> List[str]:
        """Collect all user-facing verification items for missing/unverifiable IDs."""
        issues: List[str] = []

        if pcr_not_found:
            issues.append(
                "PCR ID(s) not found: "
                + ", ".join(str(x) for x in pcr_not_found)
            )

        if project_not_found:
            issues.append(
                "Project ID(s) not found: "
                + ", ".join(str(x) for x in project_not_found)
            )

        if sps_not_found:
            issues.append(
                "SPS ID(s) not found: "
                + ", ".join(str(x) for x in sps_not_found)
            )

        if esw_not_found:
            issues.append(
                "ESW ID(s) not found: "
                + ", ".join(str(x) for x in esw_not_found)
            )

        if ecr_not_found:
            issues.append(
                "Reference ECR number(s) not found: "
                + ", ".join(str(x) for x in ecr_not_found)
            )

        return issues

    def _validate_q1_pcr(self) -> str:
        """Validate Question 1 (PCR_PCN) only – used by the PCR-driven summary path."""
        pair = self.ref_radios.get("PCR_PCN")
        if not pair:
            return "PCR question widgets not found."
        rb_yes, rb_no = pair
        if not rb_yes.isChecked() and not rb_no.isChecked():
            return (
                "Please answer Question 1: "
                "Does this Project include a Product Change Request (PCR)?"
            )
        if rb_yes.isChecked():
            pcr_text = (self.ref_boxes.get("PCR_PCN") or {}).toPlainText().strip() \
                if hasattr(self.ref_boxes.get("PCR_PCN"), "toPlainText") else ""
            if not pcr_text:
                return (
                    "Please enter at least one PCR number in the PCR field "
                    "before generating a Problem Summary."
                )
        return ""

    def _collect_pcr_payload(self) -> dict:
        """Fetch PCR/PSN records from Databricks and build an AI-ready payload.

        The PCR records and derived PSN state are stored only in runtime
        memory – no files are written.
        """
        pcr_text = self.ref_boxes["PCR_PCN"].toPlainText().strip()

        mod = self._load_pcr_query_module()
        result = mod.fetch_pcr_records(pcr_text)

        valid_records  = result.get("valid", [])
        skipped_records = result.get("skipped", [])
        not_found      = result.get("not_found", [])

        # Derive PSN Yes/No from the psnnumber field on each valid PCR row.
        psn_numbers = [
            str(rec["psnnumber"]).strip()
            for rec in valid_records
            if rec.get("psnnumber")
        ]
        derived_psn = (
            {"answer": "Yes", "numbers": psn_numbers}
            if psn_numbers
            else {"answer": "No", "numbers": []}
        )

        # Keep fetch metadata in a transient instance attribute so the
        # button handler can surface skipped/not-found IDs after generation.
        self._pcr_fetch_result = {
            "skipped":   skipped_records,
            "not_found": not_found,
        }

        scope_parts = [
            ln.strip()
            for ln in self.scope_parts_txt.toPlainText().splitlines()
            if ln.strip()
        ]

        reference_inputs = {}
        for key, txt in self.ref_boxes.items():
            reference_inputs[key] = {
                "answer": self._selected_radio_text(key),
                "text": txt.toPlainText().strip(),
            }

        return {
            "short_title":           self.short_title_edit.text().strip(),
            "reason_code":           self.reason_cb.currentText().strip(),
            "scope_parts":           scope_parts,
            "pcr_records":           valid_records,
            "derived_psn":           derived_psn,
            "skipped_pcrs":          skipped_records,
            "current_problem_text":  self.problem_txt.toPlainText().strip(),
            "proposed_solution_text": self.solution_txt.toPlainText().strip(),
            "solution_regeneration_requested": bool(
                getattr(self, "_solution_regen_requested", False)
            ),
            "solution_regeneration_reason": (
                "Existing solution is incomplete or incorrect. "
                "Regenerate a complete, engineer-quality solution."
                if getattr(self, "_solution_regen_requested", False)
                else ""
            ),
            "selected_tab_payload":  self.get_selected_tab_payload(),
            "reference_inputs": reference_inputs,
        }

    def on_regenerate_solution_clicked(self):
        """Clear existing solution and trigger AI regeneration of proposed solution."""
        try:
            # User clicked regenerate because current output is incomplete/incorrect.
            previous_title = self.short_title_edit.text()
            previous_problem_html = self.problem_txt.toHtml()
            previous_problem_text = self.problem_txt.toPlainText().strip()

            self.solution_txt.clear()
            self._solution_regen_requested = True
            self.on_problem_summary_clicked()

            # Keep this action solution-focused: preserve existing title/problem content.
            self.short_title_edit.setText(previous_title)
            if previous_problem_text:
                self.problem_txt.setHtml(previous_problem_html)
            else:
                self.problem_txt.clear()
        finally:
            self._solution_regen_requested = False

    def on_regenerate_problem_statement_clicked(self):
        """Regenerate only the problem statement and keep other generated fields unchanged."""
        previous_title = self.short_title_edit.text()
        previous_problem_html = self.problem_txt.toHtml()
        previous_problem_text = self.problem_txt.toPlainText().strip()
        previous_solution_html = self.solution_txt.toHtml()
        previous_solution_text = self.solution_txt.toPlainText().strip()
        previous_reason_code = self.ai_reason_code_edit.text() if hasattr(self, "ai_reason_code_edit") else ""
        previous_rc_scenario = getattr(self, "_current_rc_scenario", "")
        previous_rc_evidence = getattr(self, "_current_rc_evidence", None)
        previous_rc_reason_lines = getattr(self, "_current_rc_reason_lines", [])

        self.on_problem_summary_clicked()

        # Keep this action problem-focused: preserve title, solution, and reason-code fields.
        self.short_title_edit.setText(previous_title)
        if previous_solution_text:
            self.solution_txt.setHtml(previous_solution_html)
        else:
            self.solution_txt.clear()

        if hasattr(self, "ai_reason_code_edit"):
            self.ai_reason_code_edit.setText(previous_reason_code)
        self._current_rc_scenario = previous_rc_scenario
        self._current_rc_evidence = previous_rc_evidence
        self._current_rc_reason_lines = previous_rc_reason_lines

        # If regeneration fails and no new problem text is produced, restore original text.
        if not self.problem_txt.toPlainText().strip() and previous_problem_text:
            self.problem_txt.setHtml(previous_problem_html)

    # ------------------------------------------------------------------
    # Reason Code classification – called after Problem Summary is built
    # ------------------------------------------------------------------
    _REASON_CODE_DATA: Dict[str, Dict] = {
        "Cost Reduction": {
            "justification": (
                "Design change intended to reduce cost. Cost savings must be quantified "
                "in ECR PSS, showing values before and after ECO."
            ),
            "scenario": (
                "4.7\tCost Reduction \u2013 Cost reduction of part or assembly\n"
                "Scenarios\n"
                "This includes design changes in order to reduce cost of material and/or assembly cycle time.\n"
                "The EC engineer shall state quantitatively on the ECO what the existing cost was, along with the estimated savings after ECO implementation.\n"
                "Examples:\n"
                "\u2022\tA base plate is redesigned to change it to a casting in order to save overall material cost of the module.\n"
                "\u2022\tA machined part was redesigned with looser tolerance and less number of drilled holes as a part of cost reduction efforts. The drawing was released with a new PN to reflect the design changes.\n"
                "\u2022\tThe material on a part was changed from Stainless Steel to Aluminum without sacrificing design intent and functionality. The new Aluminum design is less expensive than the Stainless-Steel version."
            ),
            "keywords": [
                "cost reduction", "cost saving", "reduce cost", "dcr", "savings",
                "cost reduc", "cost improv", "price reduc",
            ],
        },
        "Product Release": {
            "justification": (
                "Used for initial release of new products, part status promotion from "
                "Concept or Eval to Prod, or first-pass FAI changes only."
            ),
            "scenario": (
                "4.1\tProduct Release \u2013 Release of New Products (Customer Release) / and Components\n"
                "Scenarios\n"
                "Used for initial release of new products/components. This includes NSR/CES releases and item status change from evaluation to production. It also includes Concept to Eval item status changes.\n"
                "\u2022\tOn changes from Concept to Evaluation\n"
                "\u2022\tAs a result, from managed builds during new product introduction\n"
                "\u2022\tFor the first pass FAI(s)\n"
                "\tChanges that result from subsequent FAIs on the same part number shall be counted as Corrective.\n"
                "\u2022\tDrawing/document corrections only while product-released part numbers shall be considered \u201cCorrective.\u201d\n"
                
            ),
            "keywords": [
                "product release", "new product", "initial release", "fai",
                "eval to prod", "concept to prod", "first release",
                "part status promotion", "new design release",
            ],
        },
        "Product Improvement": {
            "justification": (
                "Change improves functionality, quality, or reliability. Quantitative "
                "performance data must be provided in the ECR (PSS or attachments)."
            ),
            "scenario": (
                "4.2\tProduct Improvement - Improvement(Redesign or replacement New Design) to the Quality/Reliability of the Tool\n"
                "Scenarios\n"
                "This applies to design changes made to parts and assemblies as an improvement, released after July 1, 2003, that currently meet ERAMS+ and C-RAMS performance requirements in Quality (QPPM) and reliability (MTBF, MTBI). This applies to an increase in functionality or performance.\n"
                "It does apply to changes that are required to address Customer specific issues.\n"
                "It does not apply to correcting problems in the design or Drawing.\n"
                "The Design Engineering Authority (DEA) shall state quantitatively on the ECO what the product performance was, (i.e., MTBF, MTTR, MTBI, etc.) and what the product performance is after ECO implementation.\n"
                "Example:\n"
                "\u2022\tA new lift assembly is redesigned to make it more reliable. The current lift assembly already meets ERAMS+ MTBF."
            ),
            "keywords": [
                "product improvement", "improve function", "improve quality",
                "improve reliab", "performance improv", "enhance", "enhancement",
                "increased performance", "better performance",
            ],
        },
        "Manufacturing Improvement": {
            "justification": (
                "Change supports manufacturing efficiency through BOM restructuring, "
                "OMS or test procedure updates, second source additions, "
                "policy/procedure updates, or sparing improvements."
            ),
            "scenario": (
                "4.3\tManufacturing Improvement \u2013 Manufacturing Improvement/Maintenance\n"
                "Scenarios\n"
                "ECOs written to improve/maintain product producibility or manufacturing methods including:\n"
                "\u2022\tPolicy/Procedure creation/updates (all)\n"
                "\u2022\tChanging BOM structure to support work center or supplier changes (no corrections)\n"
                "\u2022\tOMS creation/updates\n"
                "\u2022\tCreating/updating test procedures, fixtures, tools etc.\n"
                "\u2022\tUpdating the OEM PDF to add second sources, update manufacturer\u2019s part numbers, etc. alternate part (no obsolete)\n"
                "\u2022\tIn the process of adding an approved second source supplier for a part released after July 1, 2003, if a SPS, SNF or SCAR is generated that requires an ECO to correct a Design or a document, as defined in the Design Correction or the Document Correction category, that ECO shall be classified as \u201cCorrective\u201d and not included in this category.\n"
                "\u2022\tAdd/updates to the critical parts lists.\n"
                "The code also applies to BOM restructuring due to outsourcing.\n"
                "Examples:\n"
                "\u2022\tThe chamber work center is reconfiguring the feeder lines and incorporating previously released options into it. The ECO restructures to configurable BOM to support this.\n"
                "\u2022\tThe spares group requests making a component part in a PDF part sparable to make the field fix simpler and less expensive. The ECO restructures the kit to accomplish this.\n"
                "\u2022\tAn ECO is written to release OMSs on the integration line.\n"
                "\u2022\tAn excess and zero demand report identifies parts in stock from two years ago that have no usage. The ECO deletes and obsoletes the parts."
            ),
            "keywords": [
                "manufacturing improvement", "manufacturing efficiency", "bom restructur",
                "oms update", "test procedure", "second source", "sparing", "spare",
                "manufacturing process", "assembly process",
            ],
        },
        "Design Correction": {
            "justification": (
                "Change corrects design issues impacting form, fit, or function, "
                "including field failures, safety compliance issues, or dimensional "
                "non-conformance."
            ),
            "scenario": (
                "4.4\tDesign Correction \u2013 Correct Design to meet Specs.\n"
                "Scenarios\n"
                "Changes made to the functionality, reliability, or quality of a product to bring it up to the specification that it was originally specified to meet including the requirements set forth by C-RAMS and ERAMS+. This includes changes to meet the products originally established safety standards.\n"
                "The change is a core engineering change that affects form, fit and function. It does not include document changes that do not affect the basic design. It does not apply to design changes made to reduce material cost. It does not apply to changes that are required to address Customer specific issues.\n"
                "Examples:\n"
                "\u2022\tA chamber was released to Production status with a specification that it could process 20 wafers / hour, however, customers reported that they could only process 10 wafers / hour. Redesign was done to enable the product to meet the original specification and an ECO was written to document the changes.\n"
                "\u2022\tAn SPS was received from a module supplier to correct a dimensional error that resulted in two mating parts not fitting together. The ECO modifies a drawing to fix the problem.\n"
                "\u2022\tAn ECO is fixing a label previously released to make the product S2 compliant.\n"
                "\u2022\tA part/assembly has poor performance in the field, high failure rate, low life one, this shall be considered as design correction not product improvement."
            ),
            "keywords": [
                "design correction", "form fit function", "fff", "field failure",
                "dimensional", "non-conformance", "nonconformance",
                "design issue", "design defect", "fit issue", "fit problem",
                "function issue", "function problem",
            ],
        },
        "Document Correction": {
            "justification": (
                "Change corrects errors in drawings, OMSs, procedures, or policy "
                "documents without impacting product form, fit, or function."
            ),
            "scenario": (
                "4.5\tDocument Correction \u2013 Correction of Document/Drawing/Model/etc.\n"
                "Scenarios\n"
                "Changes to correct document or drawing errors; this includes drawing dimension, Revision, Tolerance, Ballon Callouts, OMSs, Procedures and Policy type documents. It does not apply to drawing changes that will result in how the part will be manufactured. It does not apply to changes that result from engineering, managed and controlled builds or resulting from the initial FAI.\n"
                "Examples:\n"
                "\u2022\tAn SPS was written on volume released product to remove a feature that was double dimensioned. An ECO is written to correct and up rev the drawing.\n"
                "\u2022\tA controlled build was completed in manufacturing during a transition program. Seventy OMS/document errors were identified from the build.\n"
                "\u2022\tA reference designator was found that incorrectly identifies an option on chamber position D. An ECO is written to fix it."
            ),
            "keywords": [
                "document correction", "drawing error", "document error",
                "procedure error", "oms error", "typographical", "documentation error",
                "policy document", "drawing update", "document update",
            ],
        },
        "Obsolescence": {
            "justification": (
                "Supplier or sub-supplier part / OEM is becoming obsolete, End of Life(EOL) "
                "Includes a DEA-approved replacement. Not intended for Kit or Assembly BTP parts."
            ),
            "scenario": (
                "4.6\tObsolescence \u2013 Changes due to supplier part obsolescence or preference\n"
                "Scenarios\n"
                "Changes due to supplier or sub-supplier part obsolescence for parts/components replacement if authorized by DEA.\n"
                "Examples:\n"
                "\u2022\tA cable supplier submits an SPS because the connector is no longer available. The drawing is updated to reflect another supplier\u2019s PN.\n"
                "\u2022\tThe sensor for a cable is no longer available from the manufacturer. The engineering group selects a new sensor and releases a new PN with the new sensor and inactivates the old sensor cable.\n"
                "\u2022\tA Design Engineer wants a fastener purchased from a specific vendor. The documentation needs to be updated to reflect the preferred vendor."
            ),
            "keywords": [
                "obsolescence", "obsolete", "end of life", "eol", "discontinu",
                "oem replacement", "supplier discontinu", "obs part",
                "last time buy", "ltb", "no longer available",
            ],
        },
        "Compliance Product Specification": {
            "justification": (
                "Change addresses non-compliance with customer requirements. "
                "A unique PCR is mandatory."
            ),
            "scenario": (
                "4.9\tCompliance Product Specification\n"
                "Scenarios\n"
                "This is used when we have to update items that are not in Compliance with Customer Requirements. When selected, a unique PCR is required (CORE coded to require the PCR)."
            ),
            "keywords": [
                "compliance", "non-compliance", "noncompliance",
                "customer requirement", "customer spec", "regulatory",
                "specification violation", "does not meet spec",
            ],
        },
        "Safety Event": {
            "justification": (
                "Change impacts product safety and may require PSER reporting, "
                "Core Plus Project (PCR), or compliance-related support."
            ),
            "scenario": (
                "4.10\tSafety Event\n"
                "Scenarios\n"
                "An ECO shall be classified a Safety Event if any of the following impact criteria is met:\n"
                "\u2022\tPSER:  Mandatory capturing internal or external of corporation product Safety Event.\n"
                "\u2022\tPCR:  Mandatory unique Core Plus Project required in order to track all required changes and overall implementation of the Safety Event.\n"
                "\u2022\tSUPPORT:\n"
                "o\tSupplier or Manufacturing: Build, Testing, Shipment while adhering to corporation manufacturing practices & Compliance Product Specification\n"
                "o\tCustomer: Installation, Maintenance, Improvement while adhering to corporation manufacturing practices.  Production usage while adhering with Compliance Product Specification."
            ),
            "keywords": [
                "safety", "pser", "safety event", "safety incident",
                "hazard", "safety compliance", "safety issue", "safety concern",
                "injury", "fire", "electric shock",
            ],
        },
        "Order BOM Change \u2013 CRN Support": {
            "justification": (
                "Change requests updates to System BOMs and is directly linked "
                "to a valid CRN request."
            ),
            "scenario": (
                "4.11\tOrder BOM Change \u2013 CRN Support\n"
                "Scenarios\n"
                "This \u201cReason Code\u201d is not actually a reason code, but rather the way to support the CRN ECR Workflow rather than going through on of the ECR Workflows.  This is only to be selected for CRN ECRs requesting updates to System BOMs."
            ),
            "keywords": [
                "crn", "order bom", "system bom", "crn support",
                "configuration request number",
            ],
        },
        "Cap Code Management": {
            "justification": (
                "Change manages or modifies CAP Code options, with validation of "
                "CAP Code logic and configuration."
            ),
            "scenario": (
                "4.12\tCap Code Management\n"
                "Scenarios\n"
                "Use for ECRs managing/changing CAP Code Options"
            ),
            "keywords": [
                "cap code", "capcode", "cap code management",
                "configuration option cap",
            ],
        },
        "Option Reduction and Product End of Life": {
            "justification": (
                "Change supports structure tree trimming, movement of options to "
                "Unmaintained product structures, or obsolescence of orphaned components."
            ),
            "scenario": (
                "4.14\tOption Reduction and Product End of Life\n"
                "Scenarios\n"
                "Used for changes related to trimming of structure trees to meet corporate option optimization targets.  Options in these changes could be moved to Unmaintained product structure, removed from a specific product structure (in cases where it is still used elsewhere), or obsoleted.\n"
                "It can also be used, or the obsoleting of components orphaned by the unmaintained option process."
            ),
            "keywords": [
                "option reduction", "product end of life", "peol",
                "structure tree", "unmaintained", "orphan",
                "orphaned component", "option trim", "option elim",
            ],
        },
    }

    # ------------------------------------------------------------------
    # EC Category classification – called after Problem Summary is built
    # ------------------------------------------------------------------
    _EC_CATEGORY_DATA: Dict[str, Dict] = {
        "A1": {
            "description": EC_CATEGORY_DESC["A1"],
            "justification": (
                "Revised items include SMBoM options (e.g. kit or assembly options) "
                "and the change involves a CDW (Cost Disposition Worksheet). "
                "OBS Kit or Pieces are present."
            ),
            "keywords": [
                "smbom", "option", "cdw", "obs kit", "kit option",
                "assembly option", "configuration design", "cdw required",
            ],
        },
        "A2": {
            "description": EC_CATEGORY_DESC["A2"],
            "justification": (
                "Revised items include SMBoM options but there is no CDW involved. "
                "Typical for product release with SmBOM impact or structure changes "
                "without a configuration design work order."
            ),
            "keywords": [
                "smbom", "option", "no cdw", "product release smbom",
                "new option", "structure option", "bom option",
            ],
        },
        "B1": {
            "description": EC_CATEGORY_DESC["B1"],
            "justification": (
                "No SMBoM options among revised items but a CDW is required. "
                "Typical for OBS / Inactivate changes where parts are removed "
                "and a configuration design work order is needed."
            ),
            "keywords": [
                "obsolete", "obsolescence", "inactivate", "obs", "cdw",
                "no smbom option", "obs part", "replace obsolete", "end of life",
            ],
        },
        "B2": {
            "description": EC_CATEGORY_DESC["B2"],
            "justification": (
                "No SMBoM options, no CDW, and the revised item(s) are at Eval "
                "status or moving to Eval / being added to Proto buckets."
            ),
            "keywords": [
                "eval", "evaluation", "eval release", "proto", "prototype",
                "eval status", "moving to eval", "added to proto",
                "no cdw", "no smbom",
            ],
        },
        "B3": {
            "description": EC_CATEGORY_DESC["B3"],
            "justification": (
                "No SMBoM options, no CDW, and the revised item(s) are at Production "
                "status or moving to Production. Typical for standard production "
                "revisions, design corrections, or document corrections."
            ),
            "keywords": [
                "production", "prod release", "production release",
                "moving to production", "production status",
                "design correction", "document correction", "revision",
                "production revision",
            ],
        },
    }

    def _classify_ec_category(self, problem: str, solution: str) -> tuple:
        """Analyse problem and solution text and return (ec_category_code, justification).

        Uses keyword-scoring heuristic over EC_CATEGORY_DATA.
        Falls back to 'B3' when no keywords match (most common category).
        """
        combined = (problem + " " + solution).lower()
        best_code = ""
        best_score = -1

        for code, info in self._EC_CATEGORY_DATA.items():
            score = sum(1 for kw in info["keywords"] if kw in combined)
            if score > best_score:
                best_score = score
                best_code = code

        if best_score == 0 or not best_code:
            best_code = "B3"

        info = self._EC_CATEGORY_DATA[best_code]
        tooltip = f"{info['description']}\n\n{info['justification']}"
        return best_code, tooltip

    # ------------------------------------------------------------------

    def _extract_reason_evidence(self, problem: str, solution: str, reason_code: str, seed_texts: List[str]) -> List[str]:
        """Collect short source snippets from problem/solution for highlight display."""
        source_text = (problem or "") + "\n" + (solution or "")
        snippets: List[str] = []

        def _add_snippet(text: str):
            t = (text or "").strip()
            if not t:
                return
            if t.lower() in {s.lower() for s in snippets}:
                return
            snippets.append(t)

        # Keep LLM-provided evidence first when available.
        for s in seed_texts or []:
            _add_snippet(s)

        # Heuristic fallback: pull lines that contain rule keywords for selected code.
        for kw in self._REASON_CODE_DATA.get(reason_code, {}).get("keywords", []):
            kw = (kw or "").strip().lower()
            if not kw:
                continue
            for line in source_text.splitlines():
                line_clean = line.strip()
                if len(line_clean) < 8:
                    continue
                if kw in line_clean.lower():
                    _add_snippet(line_clean[:220])
                    break
            if len(snippets) >= 3:
                break

        return snippets[:3]

    def _highlight_evidence_html(self, text: str, evidence_snippets: List[str]) -> str:
        """Return HTML with matched evidence snippets bold-highlighted."""
        base = text or ""
        needles = [s.strip() for s in (evidence_snippets or []) if s and s.strip()]
        if not needles:
            return html.escape(base).replace("\n", "<br>")

        # Prioritize longer phrases so smaller phrases do not split larger matches.
        needles = sorted(set(needles), key=len, reverse=True)
        pattern = re.compile("|".join(re.escape(n) for n in needles), re.IGNORECASE)

        out: List[str] = []
        last = 0
        for m in pattern.finditer(base):
            out.append(html.escape(base[last:m.start()]))
            out.append(
                "<b style='background:#FFF59D; color:#111; padding:0 1px;'>"
                + html.escape(m.group(0))
                + "</b>"
            )
            last = m.end()
        out.append(html.escape(base[last:]))
        return "".join(out).replace("\n", "<br>")

    def _pick_scenario_reason_lines(
        self,
        scenario_text: str,
        reason_code: str,
        problem: str,
        solution: str,
        evidence_snippets: List[str],
    ) -> List[str]:
        """Pick existing scenario lines that best explain why the reason code was selected."""
        lines = [ln.strip() for ln in (scenario_text or "").splitlines() if ln.strip()]
        if not lines:
            return []

        combined = ((problem or "") + " " + (solution or "")).lower()
        kw_all = [str(k).strip().lower() for k in self._REASON_CODE_DATA.get(reason_code, {}).get("keywords", []) if str(k).strip()]
        kw_active = [k for k in kw_all if k in combined]
        keywords_to_use = kw_active or kw_all

        picked: List[str] = []
        picked_norm = set()

        for raw in lines:
            line = raw.strip()
            line_l = line.lower()
            if any(k in line_l for k in keywords_to_use):
                n = re.sub(r"\s+", " ", line_l)
                if n not in picked_norm:
                    picked.append(line)
                    picked_norm.add(n)

        # Optional second signal from evidence snippets.
        if len(picked) < 3:
            ev_words = set()
            for ev in evidence_snippets or []:
                for tok in re.findall(r"[A-Za-z]{4,}", ev.lower()):
                    ev_words.add(tok)
            if ev_words:
                for raw in lines:
                    if len(picked) >= 3:
                        break
                    line = raw.strip()
                    line_l = line.lower()
                    if len(line_l) < 16:
                        continue
                    if any(w in line_l for w in ev_words):
                        n = re.sub(r"\s+", " ", line_l)
                        if n not in picked_norm:
                            picked.append(line)
                            picked_norm.add(n)

        # Fallback: choose key explanatory lines already present in scenario.
        if not picked:
            for raw in lines:
                line = raw.strip()
                line_l = line.lower()
                if line_l in {"scenarios", "examples:"}:
                    continue
                if line.startswith("4.") or line.startswith("•"):
                    continue
                if len(line) >= 30:
                    picked.append(line)
                if len(picked) >= 2:
                    break

        return picked[:3]

    def _render_scenario_with_reason_lines_html(self, scenario_text: str, reason_lines: List[str]) -> str:
        """Render scenario text as HTML and bold-highlight selected existing lines."""
        chosen = {
            re.sub(r"\s+", " ", (ln or "").strip().lower())
            for ln in (reason_lines or [])
            if (ln or "").strip()
        }

        html_lines: List[str] = []
        for raw in (scenario_text or "").splitlines():
            esc = html.escape(raw)
            norm = re.sub(r"\s+", " ", raw.strip().lower())
            if norm and norm in chosen:
                html_lines.append(
                    "<b style='background:#FFF59D; color:#111; padding:0 1px;'>"
                    + esc
                    + "</b>"
                )
            else:
                html_lines.append(esc)

        return "<br>".join(html_lines)

    def _classify_reason_code(self, problem: str, solution: str) -> tuple:
        """You are a Senior Manufacturing Engineer, analyse problem and solution text and return (reason_code, scenario).

        Attempts LLM-based classification first via AI_Assisted_PSS.classify_reason_code_with_llm,
        using a comprehensive ECM expert prompt with all 14 reason codes and governing rules.
        Falls back to keyword-scoring heuristic on any LLM error.
        """
        # ── Attempt LLM-based classification ────────────────────────────────
        try:
            module = self._load_ai_pss_module()
            classify_fn = getattr(module, "classify_reason_code_with_llm", None)
            if callable(classify_fn):
                try:
                    llm_result = classify_fn(problem, solution, return_evidence=True)
                except TypeError:
                    llm_result = classify_fn(problem, solution)

                llm_code = ""
                llm_just = ""
                llm_evidence: List[str] = []
                if isinstance(llm_result, tuple):
                    if len(llm_result) >= 2:
                        llm_code, llm_just = llm_result[0], llm_result[1]
                    if len(llm_result) >= 3 and isinstance(llm_result[2], list):
                        llm_evidence = llm_result[2]

                if llm_code:
                    info = self._REASON_CODE_DATA.get(llm_code, {})
                    justification = llm_just or info.get("justification", llm_just)
                    scenario = info.get("scenario", justification)
                    evidence = self._extract_reason_evidence(problem, solution, llm_code, llm_evidence)
                    return llm_code, justification, scenario, evidence
        except Exception:
            pass

        # ── Keyword-scoring heuristic fallback ──────────────────────────────
        combined = (problem + " " + solution).lower()
        best_code = ""
        best_score = -1

        for code, info in self._REASON_CODE_DATA.items():
            score = sum(1 for kw in info["keywords"] if kw in combined)
            if score > best_score:
                best_score = score
                best_code = code

        if best_score == 0 or not best_code:
            best_code = "Product Improvement"

        justification = self._REASON_CODE_DATA[best_code]["justification"]
        scenario = self._REASON_CODE_DATA[best_code].get("scenario", justification)
        evidence = self._extract_reason_evidence(problem, solution, best_code, [])
        return best_code, justification, scenario, evidence

    # ------------------------------------------------------------------

    def _show_rc_info_dialog(self):
        """Show Scenario and Examples for the currently proposed Reason Code."""
        scenario_text = getattr(self, "_current_rc_scenario", "")
        reason_lines = getattr(self, "_current_rc_reason_lines", [])
        rc_code = self.ai_reason_code_edit.text().strip() if hasattr(self, "ai_reason_code_edit") else ""
        if not scenario_text:
            QMessageBox.information(
                self,
                "Reason Code Info",
                "Run Problem Summary first to see the Scenario and Examples for the proposed Reason Code.",
            )
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Reason Code: {rc_code} – Scenario & Examples")
        dlg.setMinimumWidth(640)
        dlg.setMinimumHeight(420)
        layout = QVBoxLayout(dlg)
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)

        scenario_html = self._render_scenario_with_reason_lines_html(scenario_text, reason_lines)
        reason_lines_html = "<br>".join(f"- {html.escape(x)}" for x in reason_lines) if reason_lines else "- No explicit scenario lines matched"

        detail_html = (
            "<div style='font-size:12px; line-height:1.45;'>"
            "<h3 style='margin:0 0 8px 0; color:#1F4E79;'>Scenario and Examples</h3>"
            "<div style='margin:0 0 8px 0; color:#444;'><i>Highlighted lines are the specific scenario lines used to justify the proposed Reason Code.</i></div>"
            f"<div>{scenario_html or 'No scenario details available.'}</div>"
            "<h3 style='margin:14px 0 8px 0; color:#1F4E79;'>Reason-Driving Scenario Lines</h3>"
            f"<div>{reason_lines_html}</div>"
            "</div>"
        )
        text_edit.setHtml(detail_html)
        text_edit.setStyleSheet(
            "font-size: 12px; background: #FAFAFA; border: 1px solid #CCCCCC; padding: 6px;"
        )
        layout.addWidget(text_edit)
        close_btn = QPushButton("Close")
        close_btn.setFixedWidth(90)
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)
        dlg.exec()

    # ------------------------------------------------------------------

    def on_problem_summary_clicked(self):
        try:
            # Reset transient fetch state per click to avoid stale cross-field notices.
            self._project_fetch_result = {"skipped": [], "not_found": []}
            self._sps_fetch_result = {"skipped": [], "not_found": []}
            self._esw_fetch_result = {"skipped": [], "not_found": []}
            self._ecr_fetch_result = {"skipped": [], "not_found": []}

            # Always validate the full assessment first so unanswered Yes/No
            # selections are prompted regardless of PCR branch.
            validation_error = self._validate_problem_summary_inputs()
            if validation_error:
                QMessageBox.warning(self, "Validation Required", validation_error)
                return

            pcr_yes = self.ref_radios.get("PCR_PCN", (None, None))[0]
            use_pcr_path = pcr_yes is not None and pcr_yes.isChecked()

            if use_pcr_path:
                # ---- Question 1 – PCR-driven Databricks path ----
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                payload = self._append_project_context(self._collect_pcr_payload())

                # Auto-fill PSN radio button and text field from derived_psn (PCR path only).
                # This must happen immediately after PCR fetch so users are not asked to enter PSN manually.
                derived_psn = payload.get("derived_psn") or {}
                psn_pair = self.ref_radios.get("PSN")
                psn_box = self.ref_boxes.get("PSN")
                if psn_pair:
                    psn_yes_rb, psn_no_rb = psn_pair
                    if derived_psn.get("answer") == "Yes":
                        psn_yes_rb.setChecked(True)
                        if psn_box:
                            psn_box.setPlainText(", ".join(derived_psn.get("numbers") or []))
                            psn_box.setVisible(True)
                    else:
                        psn_no_rb.setChecked(True)
                        if psn_box:
                            psn_box.setPlainText("")
                            psn_box.setVisible(False)

                fetch_meta = getattr(self, "_pcr_fetch_result", {})
                skipped    = fetch_meta.get("skipped", [])
                not_found  = fetch_meta.get("not_found", [])
                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])
                ecr_meta = getattr(self, "_ecr_fetch_result", {})
                skipped_ecr = ecr_meta.get("skipped", [])
                not_found_ecr = ecr_meta.get("not_found", [])

                if not payload.get("pcr_records"):
                    msgs = []
                    if skipped:
                        msgs.append(
                            "The following PCR(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  PCR {s['pcr_id']} – {s['status']}" for s in skipped
                            )
                        )
                    if not_found:
                        msgs.append(
                            "The following PCR ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found)
                        )
                    raise ValueError(
                        "No eligible PCR records found to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("PCR", (None, None))[0] and self.ref_radios["PCR"][0].isChecked() and not payload.get("project_records"):
                    msgs = []
                    if skipped_projects:
                        msgs.append(
                            "The following Project(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  Project {s['project_id']} - {s['status']}" for s in skipped_projects
                            )
                        )
                    if not_found_projects:
                        msgs.append(
                            "The following Project ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_projects)
                        )
                    raise ValueError(
                        "No eligible Project records found for Question 2 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("SPS", (None, None))[0] and self.ref_radios["SPS"][0].isChecked() and not payload.get("sps_records"):
                    msgs = []
                    if skipped_sps:
                        msgs.append(
                            "The following SPS(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                            )
                        )
                    if not_found_sps:
                        msgs.append(
                            "The following SPS ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_sps)
                        )
                    raise ValueError(
                        "No eligible SPS records found for Question 3 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("ESW", (None, None))[0] and self.ref_radios["ESW"][0].isChecked() and not payload.get("esw_records"):
                    msgs = []
                    if skipped_esw:
                        msgs.append(
                            "The following ESW(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                            )
                        )
                    if not_found_esw:
                        msgs.append(
                            "The following ESW ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_esw)
                        )
                    raise ValueError(
                        "No eligible ESW records found for Question 4 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("REF_ECR", (None, None))[0] and self.ref_radios["REF_ECR"][0].isChecked() and not payload.get("ecr_records"):
                    msgs = []
                    if skipped_ecr:
                        msgs.append(
                            "The following Reference ECR(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ECR {s['ecr_number']} - {s['status']}" for s in skipped_ecr
                            )
                        )
                    if not_found_ecr:
                        msgs.append(
                            "The following Reference ECR number(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_ecr)
                        )
                    raise ValueError(
                        "No eligible Reference ECR records found for Question 5 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                verify_items = self._collect_not_found_verification_items(
                    pcr_not_found=not_found,
                    project_not_found=not_found_projects,
                    sps_not_found=not_found_sps,
                    esw_not_found=not_found_esw,
                    ecr_not_found=not_found_ecr,
                )
                if verify_items:
                    QMessageBox.warning(
                        self,
                        "Reference Verification Required",
                        "Some entered values were not found. Please verify and correct before generating Problem Summary:\n\n"
                        + "\n".join(f"- {x}" for x in verify_items),
                    )
                    return

                out = self._run_ai_pss_full(payload)

            else:
                # ---- Existing all-questions path ----
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                payload = self._append_project_context(self._collect_pss_payload())

                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])
                ecr_meta = getattr(self, "_ecr_fetch_result", {})
                skipped_ecr = ecr_meta.get("skipped", [])
                not_found_ecr = ecr_meta.get("not_found", [])

                if self.ref_radios.get("SPS", (None, None))[0] and self.ref_radios["SPS"][0].isChecked() and not payload.get("sps_records"):
                    msgs = []
                    if skipped_sps:
                        msgs.append(
                            "The following SPS(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                            )
                        )
                    if not_found_sps:
                        msgs.append(
                            "The following SPS ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_sps)
                        )
                    raise ValueError(
                        "No eligible SPS records found for Question 3 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("ESW", (None, None))[0] and self.ref_radios["ESW"][0].isChecked() and not payload.get("esw_records"):
                    msgs = []
                    if skipped_esw:
                        msgs.append(
                            "The following ESW(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                            )
                        )
                    if not_found_esw:
                        msgs.append(
                            "The following ESW ID(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_esw)
                        )
                    raise ValueError(
                        "No eligible ESW records found for Question 4 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                if self.ref_radios.get("REF_ECR", (None, None))[0] and self.ref_radios["REF_ECR"][0].isChecked() and not payload.get("ecr_records"):
                    msgs = []
                    if skipped_ecr:
                        msgs.append(
                            "The following Reference ECR(s) have an inactive status and were skipped:\n"
                            + "\n".join(
                                f"  ECR {s['ecr_number']} - {s['status']}" for s in skipped_ecr
                            )
                        )
                    if not_found_ecr:
                        msgs.append(
                            "The following Reference ECR number(s) were not found in Databricks:\n"
                            + "\n".join(f"  {n}" for n in not_found_ecr)
                        )
                    raise ValueError(
                        "No eligible Reference ECR records found for Question 5 to generate a summary.\n\n"
                        + "\n\n".join(msgs)
                    )

                verify_items = self._collect_not_found_verification_items(
                    pcr_not_found=[],
                    project_not_found=not_found_projects,
                    sps_not_found=not_found_sps,
                    esw_not_found=not_found_esw,
                    ecr_not_found=not_found_ecr,
                )
                if verify_items:
                    QMessageBox.warning(
                        self,
                        "Reference Verification Required",
                        "Some entered values were not found. Please verify and correct before generating Problem Summary:\n\n"
                        + "\n".join(f"- {x}" for x in verify_items),
                    )
                    return

                if getattr(self, "email_file_path", None):
                    module = self._load_ai_pss_module()
                    email_content = module.read_email_file(self.email_file_path)
                    if str(email_content).startswith("Error reading"):
                        raise ValueError(f"Could not read email file:\n{email_content}")
                    user_problem = self.problem_txt.toPlainText().strip()
                    out = module.correlate_email_with_problem(
                        email_content,
                        user_problem,
                        payload
                    )
                    if out.get("error"):
                        raise ValueError(out["error"])
                else:
                    out = self._run_ai_pss_full(payload)

                fetch_meta = {}
                skipped    = []
                not_found  = []
                project_meta = getattr(self, "_project_fetch_result", {})
                skipped_projects = project_meta.get("skipped", [])
                not_found_projects = project_meta.get("not_found", [])
                sps_meta = getattr(self, "_sps_fetch_result", {})
                skipped_sps = sps_meta.get("skipped", [])
                not_found_sps = sps_meta.get("not_found", [])
                esw_meta = getattr(self, "_esw_fetch_result", {})
                skipped_esw = esw_meta.get("skipped", [])
                not_found_esw = esw_meta.get("not_found", [])

            # ---- Common output handling ----
            title    = (out.get("title") or "").strip()
            problem  = (out.get("problem") or out.get("problem_statement") or "").strip()
            solution = (out.get("solution") or out.get("solution_statement") or "").strip()

            if not any([title, problem, solution]):
                raise ValueError("AI_Assisted_PSS returned empty output.")

            if title:
                self.short_title_edit.setText(title)
            if problem:
                self.problem_txt.setHtml(_format_pss_for_html_display(problem[:2000]))
            if solution:
                formatted_sol = _format_solution_for_display(solution)[:2000]
                self.solution_txt.setHtml(_format_pss_for_html_display(formatted_sol))

            # Surface skipped / not-found PCRs as a non-blocking info message.
            if use_pcr_path and (skipped or not_found):
                info_parts = []
                if skipped:
                    info_parts.append(
                        "Skipped PCR(s) due to inactive status:\n"
                        + "\n".join(
                            f"  PCR {s['pcr_id']} – {s['status']}" for s in skipped
                        )
                    )
                if not_found:
                    info_parts.append(
                        "PCR ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found)
                    )
                QMessageBox.information(
                    self,
                    "PCR Lookup Notice",
                    "Summary generated from eligible PCR(s).\n\n"
                    + "\n\n".join(info_parts),
                )

            if (payload.get("project_records") or skipped_projects or not_found_projects):
                info_parts = []
                if skipped_projects:
                    info_parts.append(
                        "Skipped Project(s) due to inactive status:\n"
                        + "\n".join(
                            f"  Project {s['project_id']} - {s['status']}" for s in skipped_projects
                        )
                    )
                if not_found_projects:
                    info_parts.append(
                        "Project ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_projects)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "Project Lookup Notice",
                        "Summary generated from eligible Project record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            if (payload.get("sps_records") or skipped_sps or not_found_sps):
                info_parts = []
                if skipped_sps:
                    info_parts.append(
                        "Skipped SPS(s) due to inactive status:\n"
                        + "\n".join(
                            f"  SPS {s['sps_id']} - {s['status']}" for s in skipped_sps
                        )
                    )
                if not_found_sps:
                    info_parts.append(
                        "SPS ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_sps)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "SPS Lookup Notice",
                        "Summary generated from eligible SPS record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            if (payload.get("esw_records") or skipped_esw or not_found_esw):
                info_parts = []
                if skipped_esw:
                    info_parts.append(
                        "Skipped ESW(s) due to inactive status:\n"
                        + "\n".join(
                            f"  ESW {s['esw_number']} - {s['status']}" for s in skipped_esw
                        )
                    )
                if not_found_esw:
                    info_parts.append(
                        "ESW ID(s) not found in Databricks:\n"
                        + "\n".join(f"  {n}" for n in not_found_esw)
                    )
                if info_parts:
                    QMessageBox.information(
                        self,
                        "ESW Lookup Notice",
                        "Summary generated from eligible ESW record(s).\n\n"
                        + "\n\n".join(info_parts),
                    )

            # ── Propose Reason Code based on generated problem + solution ────
            _prob = self.problem_txt.toPlainText()
            _sol  = self.solution_txt.toPlainText()
            try:
                rc_code, rc_just, rc_scenario, rc_evidence = self._classify_reason_code(_prob, _sol)
                rc_reason_lines = self._pick_scenario_reason_lines(
                    rc_scenario,
                    rc_code,
                    _prob,
                    _sol,
                    rc_evidence,
                )
                self.ai_reason_code_edit.setText(rc_code)
                self._current_rc_scenario = rc_scenario
                self._current_rc_evidence = rc_evidence
                self._current_rc_reason_lines = rc_reason_lines
                self.ai_justification_btn.setToolTip(
                    "Click \u24d8 to see Scenario and highlighted reason-driving lines for the proposed Reason Code"
                )
            except Exception:
                pass
            # ─────────────────────────────────────────────────────────────────

        except Exception as e:
            QMessageBox.warning(self, "Problem Summary Error", str(e))
        finally:
            QApplication.restoreOverrideCursor()


    # ---- Flow Logic ----
    def clear_flow(self):
        while self.flow_area.count():
            w = self.flow_area.takeAt(0).widget()
            if w: w.deleteLater()
        self.ec_result_lbl.setText(""); self.ec_divider.setVisible(False)

    def add_q(self,q,opts,cb):
        row = QHBoxLayout(); row.addWidget(QLabel(q))
        grp = QButtonGroup(self)
        for o in opts:
            rb = QRadioButton(o); grp.addButton(rb)
            rb.toggled.connect(lambda c,v=o: c and cb(v))
            row.addWidget(rb)
        row.addStretch(1)
        w = QWidget(); w.setLayout(row); self.flow_area.addWidget(w)

    def start_flow(self):
        if not self.sender().isChecked(): return
        self.clear_flow(); s=self.sender().text()
        if s=="Production Release": return self.finish("B3")
        if s=="OBS / Inactivate": return self.finish("B1")
        if s=="Up Revision": return self.add_q("SmBOM Impacted?",["Yes","No"],self.up1)
        if s=="Status Roll": return self.add_q("Transition Type?",["EVAL → Production","Concept → BOM List"],self.status1)
        if s=="Product Release": return self.add_q("SmBOM Impacted?",["Yes","No"],self.prod1)

    def up1(self,v): self.trim(1); self.finish("A1" if v=="Yes" else "A2") if v in ["Yes","No"] and False else (self.add_q("OBS Kit / Pieces?",["Yes","No"],lambda x:self.finish("A1" if x=="Yes" else "A2")) if v=="Yes" else self.add_q("Part Status?",["EVAL","PROD"],lambda x:self.finish("B2" if x=="EVAL" else "B3")))
    def status1(self,v): self.trim(1); self.finish("B3") if v.startswith("EVAL") else self.add_q("OBS Kit / Pieces?",["Yes","No"],lambda x:self.finish("A1" if x=="Yes" else "A2"))
    def prod1(self,v): self.trim(1); self.finish("A2") if v=="Yes" else self.add_q("Part Status?",["EVAL","PROD"],lambda x:self.finish("B2" if x=="EVAL" else "B3"))

    def trim(self,k):
        while self.flow_area.count()>k:
            w=self.flow_area.takeAt(self.flow_area.count()-1).widget()
            if w: w.deleteLater()

    def finish(self,c):
        self.ec_result_lbl.setText(f"EC Category: {c} \u2013 {EC_CATEGORY_DESC[c]}")
        self.ec_divider.setVisible(True); self.secB_header.setVisible(True); self.secB.setVisible(True)
        # Populate the AI Proposed EC Category widgets from the user's flow selection
        try:
            self.ai_ec_category_edit.setText(c)
            info = self._EC_CATEGORY_DATA.get(c, {})
            desc = info.get('description', EC_CATEGORY_DESC.get(c, ''))
            just = info.get('justification', '')
            self.ai_ec_justification_btn.setToolTip(
                f"<b>EC Category {c}:</b><br>{desc}<br><br>{just}"
            )
        except Exception:
            pass
        self.ec_category_selected.emit(c)

    def set_selected_category(self, category_code: str):
        c = (category_code or "").strip().upper()
        if c not in EC_CATEGORY_DESC:
            return
        self.ec_result_lbl.setText(f"EC Category: {c} \u2013 {EC_CATEGORY_DESC[c]}")
        self.ec_divider.setVisible(True)
        self.secB_header.setVisible(True)
        self.secB.setVisible(True)
        self.ai_ec_category_edit.setText(c)
        info = self._EC_CATEGORY_DATA.get(c, {})
        desc = info.get('description', EC_CATEGORY_DESC.get(c, ''))
        just = info.get('justification', '')
        self.ai_ec_justification_btn.setToolTip(
            f"<b>EC Category {c}:</b><br>{desc}<br><br>{just}"
        )

    def _request_reset_ec_form(self):
        self.reset_form_requested.emit()

    def reset(self):
        self._solution_regen_requested = False
        self.email_file_path = None
        self.selected_tab_payload = {}
        self._current_rc_scenario = ""
        self._current_rc_evidence = []
        self._current_rc_reason_lines = []

        # Clear all text inputs and combo selections back to default.
        for line_edit in self.findChildren(QLineEdit):
            line_edit.clear()
        for text_edit in self.findChildren(QTextEdit):
            text_edit.clear()
        for combo in self.findChildren(QComboBox):
            if combo.count() > 0:
                combo.setCurrentIndex(0)
        for check in self.findChildren(QCheckBox):
            check.setChecked(False)

        # Fully clear radio selections by disabling exclusivity temporarily.
        groups = self.findChildren(QButtonGroup)
        for grp in groups:
            grp.setExclusive(False)
        for rb in self.findChildren(QRadioButton):
            rb.setChecked(False)
            rb.setEnabled(True)
        for grp in groups:
            grp.setExclusive(True)

        for txt in self.ref_boxes.values():
            txt.clear()
            txt.setVisible(False)
        for lbl in self.ref_labels.values():
            lbl.setEnabled(True)

        if hasattr(self, "_browse_email_btn") and self._browse_email_btn is not None:
            self._browse_email_btn.setText("Browse Attachment")
            self._browse_email_btn.setVisible(False)

        self.ai_ec_justification_btn.setToolTip("Run Problem Summary to see EC Category justification")
        self.ai_justification_btn.setToolTip("Click to see Scenario and Examples for the proposed Reason Code")

        self.clear_flow()
        self.ec_divider.setVisible(False)
        self.secB_header.setVisible(False)
        self.secB.setVisible(False)


class ECTabVisibilityPolicy:
    BASE_VISIBLE_TABS = {"READ ME", "EC Creation Form"}
    ALWAYS_VISIBLE_AFTER_CATEGORY = {
        "Revised Items & Components",
        "Reports",
        "User Notes",
    }
    CONDITIONAL_BY_CATEGORY = {
        "A1": {"OBS Parts", "Orphan Analysis", "Structure Sheet", "Where Used", "Inventory_Cost", "Watch_List"},
        "B1": {"OBS Parts", "Orphan Analysis", "Where Used", "Inventory_Cost", "Watch_List"},
        "A2": {"Structure Sheet"},
    }

    @classmethod
    def normalize_category(cls, category: str) -> str:
        return (category or "").strip().upper()

    @classmethod
    def tabs_for_category(cls, category: str) -> set[str]:
        cat = cls.normalize_category(category)
        conditional = cls.CONDITIONAL_BY_CATEGORY.get(cat, set())
        return set(cls.BASE_VISIBLE_TABS) | set(cls.ALWAYS_VISIBLE_AFTER_CATEGORY) | set(conditional)

    @classmethod
    def conditional_tabs_removed_by_change(cls, current_category: str, new_category: str) -> list[str]:
        curr = cls.CONDITIONAL_BY_CATEGORY.get(cls.normalize_category(current_category), set())
        new = cls.CONDITIONAL_BY_CATEGORY.get(cls.normalize_category(new_category), set())
        return sorted(curr - new)


class TabDirtyStateManager:
    def __init__(self, tabs_widget: QTabWidget):
        self.tabs_widget = tabs_widget
        self._tab_widgets: Dict[str, QWidget] = {}
        self._tab_titles: Dict[str, str] = {}
        self._dirty: Dict[str, bool] = {}

    def register(self, tab_name: str, tab_widget: QWidget, title: str):
        self._tab_widgets[tab_name] = tab_widget
        self._tab_titles[tab_name] = title
        self._dirty.setdefault(tab_name, False)
        self._refresh_title(tab_name)

    def mark_dirty(self, tab_name: str):
        if tab_name not in self._dirty:
            return
        if self._dirty[tab_name]:
            return
        self._dirty[tab_name] = True
        self._refresh_title(tab_name)

    def clear_dirty(self, tab_name: str):
        if tab_name not in self._dirty:
            return
        if not self._dirty[tab_name]:
            return
        self._dirty[tab_name] = False
        self._refresh_title(tab_name)

    def clear_many(self, tab_names: List[str]):
        for tab_name in tab_names:
            self.clear_dirty(tab_name)

    def clear_all(self):
        for tab_name in list(self._dirty.keys()):
            self.clear_dirty(tab_name)

    def is_dirty(self, tab_name: str) -> bool:
        return bool(self._dirty.get(tab_name, False))

    def _refresh_title(self, tab_name: str):
        tab_widget = self._tab_widgets.get(tab_name)
        if tab_widget is None:
            return
        idx = self.tabs_widget.indexOf(tab_widget)
        if idx < 0:
            return
        base = self._tab_titles.get(tab_name, tab_name)
        star = " *" if self._dirty.get(tab_name, False) else ""
        self.tabs_widget.setTabText(idx, f"{base}{star}")


class TabDirtyWatcher(QObject):
    def __init__(self, tab_name: str, root_widget: QWidget, dirty_manager: TabDirtyStateManager):
        super().__init__(root_widget)
        self.tab_name = tab_name
        self.root_widget = root_widget
        self.dirty_manager = dirty_manager
        self._wire_widget(root_widget)

    def _mark_dirty(self, *_args):
        self.dirty_manager.mark_dirty(self.tab_name)

    def _connect_once(self, widget: QWidget, attr_name: str, signal_name: str):
        if widget.property(attr_name):
            return
        signal = getattr(widget, signal_name, None)
        if signal is None:
            return
        signal.connect(self._mark_dirty)
        widget.setProperty(attr_name, True)

    def _wire_widget(self, widget: QWidget):
        if widget is None:
            return
        if widget.property('_tab_dirty_filter_installed'):
            return
        widget.setProperty('_tab_dirty_filter_installed', True)
        widget.installEventFilter(self)

        if isinstance(widget, QLineEdit):
            self._connect_once(widget, '_tab_dirty_connected_textChanged', 'textChanged')
        elif isinstance(widget, QTextEdit):
            self._connect_once(widget, '_tab_dirty_connected_textChanged', 'textChanged')
        elif isinstance(widget, QComboBox):
            self._connect_once(widget, '_tab_dirty_connected_currentTextChanged', 'currentTextChanged')
        elif isinstance(widget, QCheckBox):
            self._connect_once(widget, '_tab_dirty_connected_toggled', 'toggled')
        elif isinstance(widget, QRadioButton):
            self._connect_once(widget, '_tab_dirty_connected_toggled', 'toggled')
        elif isinstance(widget, QTableWidget):
            self._connect_once(widget, '_tab_dirty_connected_itemChanged', 'itemChanged')

        for child in widget.findChildren(QWidget):
            self._wire_widget(child)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.ChildAdded:
            child = event.child()
            if isinstance(child, QWidget):
                self._wire_widget(child)
        return super().eventFilter(obj, event)



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        screen = QGuiApplication.primaryScreen()
        if screen is not None:
            avail = screen.availableGeometry()
            w = max(640, min(1280, avail.width() - 32))
            h = max(520, min(860, avail.height() - 32))
            self.resize(w, h)
            self.move(
                avail.x() + max(0, (avail.width() - w) // 2),
                avail.y() + max(0, (avail.height() - h) // 2),
            )
        else:
            self.resize(1280, 860)
        self.setStyleSheet("""
            QScrollBar:vertical {
                background: #EAF1F8;
                width: 12px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #BDD0E3;
                min-height: 28px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover {
                background: #AFC6DD;
            }
            QScrollBar:horizontal {
                background: #EAF1F8;
                height: 12px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #BDD0E3;
                min-width: 28px;
                border-radius: 6px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #AFC6DD;
            }
            QScrollBar::add-line, QScrollBar::sub-line,
            QScrollBar::add-page, QScrollBar::sub-page {
                background: transparent;
                border: none;
            }
        """)
        self.tabs = QTabWidget()
        
        self.tabs.setStyleSheet("""
            QTabBar::tab { background: #EAF2FB; color: #1F3B57; padding: 8px 14px; border: 1px solid #D5E3F6; border-bottom: none; border-top-left-radius:6px; border-top-right-radius:6px; }
            QTabBar::tab:selected { background: #FFFFFF; color: #0F2D46; font-weight: 600; }
            QTabWidget::pane { border: 1px solid #D5E3F6; top: -1px; }
        """)

        readme_path = Path(__file__).with_name('README.txt')
        self.readme_tab = ReadmeTab(readme_path)
        self.obs_tab = OBSPartsTab()
        self.structure_tab = None  # to be created later
        self.whereused_tab = WhereUsedTabV2(obs_provider=self.obs_tab)
        self.obs_tab.where_used_tab = self.whereused_tab  # enables OBS → Where Used import

        self.ec_form_tab = ECCreationInputsFormTab()

        self.revised_items_tab = QWidget()
        revised_layout = QVBoxLayout(self.revised_items_tab)
        revised_row = QHBoxLayout()
        revised_title = QLabel("Revised Items & Components")
        revised_title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        self.btn_update_revised_items = QPushButton("Update Revised Items & Components")
        revised_row.addWidget(revised_title)
        revised_row.addStretch(1)
        revised_row.addWidget(self.btn_update_revised_items)
        revised_layout.addLayout(revised_row)
        revised_info = QTextEdit()
        revised_info.setReadOnly(True)
        revised_info.setPlaceholderText('Use this tab to review or update revised items and components.')
        revised_layout.addWidget(revised_info)
        self.btn_update_revised_items.clicked.connect(
            lambda: QMessageBox.information(
                self,
                'Update Revised Items & Components',
                'Update action is ready. Integrate revised-item source logic here.'
            )
        )

        if WatchListTab is not None:
            self.watch_list_tab = WatchListTab(self)
        else:
            self.watch_list_tab = QWidget()
            watch_layout = QVBoxLayout(self.watch_list_tab)
            watch_title = QLabel("Watch_List")
            watch_title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
            watch_layout.addWidget(watch_title)
            watch_notes = QTextEdit()
            watch_notes.setPlaceholderText('Watch_List module is not available.')
            watch_layout.addWidget(watch_notes)

        self.tabs.addTab(self.readme_tab, "READ ME")
        self.tabs.addTab(self.ec_form_tab, "EC Creation Form"); self.tabs.setDocumentMode(True); self.tabs.setMovable(True)
        self.tabs.addTab(self.revised_items_tab, "Revised Items & Components")
        self.tabs.addTab(self.obs_tab, "OBS Parts")
        self.tabs.addTab(self.whereused_tab, "Where Used")
        self.orphan_tab = OrphanAnalysisTab(self.obs_tab)
        self.tabs.addTab(self.orphan_tab, "Orphan Analysis")
        self.structure_tab = StructureSheetTab(); self.tabs.addTab(self.structure_tab, "Structure Sheet")
        self.inventory_cost_tab=InventoryCostTab(); self.tabs.addTab(self.inventory_cost_tab,'Inventory_Cost')
        self.tabs.addTab(self.watch_list_tab, 'Watch_List')
        self.report_tab = ReportTab()
        self.tabs.addTab(self.report_tab, "Reports")
        self.user_notes_tab = PlaceholderTab("User Notes")
        self.tabs.addTab(self.user_notes_tab, "User Notes")
        self.setCentralWidget(self.tabs)
        self._managed_tabs: Dict[str, QWidget] = {
            "READ ME": self.readme_tab,
            "EC Creation Form": self.ec_form_tab,
            "Revised Items & Components": self.revised_items_tab,
            "OBS Parts": self.obs_tab,
            "Where Used": self.whereused_tab,
            "Orphan Analysis": self.orphan_tab,
            "Structure Sheet": self.structure_tab,
            "Inventory_Cost": self.inventory_cost_tab,
            "Watch_List": self.watch_list_tab,
            "Reports": self.report_tab,
            "User Notes": self.user_notes_tab,
        }
        self._tab_visibility_policy = ECTabVisibilityPolicy()
        self._dirty_state = TabDirtyStateManager(self.tabs)
        self._dirty_watchers: List[TabDirtyWatcher] = []
        self._current_ec_category = None
        self._category_switch_in_progress = False
        self._initialize_ec_category_tab_management()

        self._tab_switch_guard = False
        self._prev_tab_index = self.tabs.currentIndex()
        self.tabs.currentChanged.connect(self._on_main_tab_changed)
        self._structure_attention_signals_connected = False
        self._connect_structure_tab_attention_signals()

        self._find_last_query = ''
        self._find_hits = []
        self._find_hit_index = -1
        self._find_prev_highlights = []

        tb=QToolBar("File"); self.addToolBar(tb)
        find_edit = QLineEdit(); find_edit.setPlaceholderText('Find part / text'); find_edit.setFixedWidth(220); tb.addWidget(find_edit)

        def _clear_find_highlights():
            for item, old_brush in self._find_prev_highlights:
                try:
                    item.setBackground(old_brush)
                except Exception:
                    pass
            self._find_prev_highlights = []

        def _do_find():
            text = find_edit.text().strip().lower()
            if not text:
                return

            w = self.tabs.currentWidget()
            from PyQt6.QtWidgets import QTableWidget, QTextEdit


            # Always strip and lowercase the search text for comparison
            search_text = text.strip().lower()
            if search_text != self._find_last_query:
                _clear_find_highlights()
                self._find_hits = []
                self._find_hit_index = -1
                self._find_last_query = search_text

                for t in w.findChildren(QTableWidget):
                    for r in range(t.rowCount()):
                        for c in range(t.columnCount()):
                            it = t.item(r, c)
                            if it:
                                cell_text = it.text().strip().lower()
                                if search_text and search_text in cell_text:
                                    self._find_hits.append((t, r, c, it))
                                    self._find_prev_highlights.append((it, it.background()))
                                    it.setBackground(QColor('#FFF59D'))

                if not self._find_hits:
                    found_in_text = False
                    for te in w.findChildren(QTextEdit):
                        if search_text in te.toPlainText().strip().lower():
                            te.moveCursor(QTextCursor.MoveOperation.Start)
                            te.find(search_text)
                            found_in_text = True
                            break
                    if found_in_text:
                        self.statusBar().showMessage('Found text in document.', 4000)
                        return
                    self.statusBar().showMessage('Found 0 matches.', 4000)
                    QMessageBox.information(self, 'Find', 'Not Found')
                    return

                self.statusBar().showMessage(
                    f'Found {len(self._find_hits)} match(es). Press Enter to jump to the next match.',
                    5000,
                )

            if not self._find_hits:
                QMessageBox.information(self, 'Find', 'Not Found')
                return

            self._find_hit_index = (self._find_hit_index + 1) % len(self._find_hits)
            t, r, c, it = self._find_hits[self._find_hit_index]
            t.setCurrentCell(r, c)
            t.scrollToItem(it, QTableWidget.ScrollHint.PositionAtCenter)
            self.statusBar().showMessage(
                f'Match {self._find_hit_index + 1} of {len(self._find_hits)}',
                3000,
            )
        find_edit.returnPressed.connect(_do_find)
        act_save=QAction("Save", self); act_save.triggered.connect(self.save_data); tb.addAction(act_save)
        
        spacer=QWidget(); spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred); tb.addWidget(spacer)

        reset_btn=QPushButton("Reset App")
        reset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reset_btn.setStyleSheet("""
            QPushButton { color:#FFFFFF; padding:6px 14px; border-radius:6px; border:1px solid #0D5EA6;
            background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2994FF, stop:1 #0A67C2); }
            QPushButton:hover { background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2FA0FF, stop:1 #0D6ED0); }
            QPushButton:pressed { padding-top:7px; padding-bottom:5px; }
        """)
        shadow=QGraphicsDropShadowEffect(self); shadow.setBlurRadius(12); shadow.setXOffset(0); shadow.setYOffset(2); shadow.setColor(QColor(0,0,0,80))
        reset_btn.setGraphicsEffect(shadow); reset_btn.clicked.connect(self.reset_app); tb.addWidget(reset_btn)

        self.load_data_if_exists()
        self._refresh_structure_tab_attention()

    def _connect_structure_tab_attention_signals(self):
        if getattr(self, '_structure_attention_signals_connected', False):
            return
        if self.structure_tab is None:
            return
        table = getattr(self.structure_tab, 'table', None)
        if table is None:
            return

        self._structure_attention_signals_connected = True
        table.itemChanged.connect(self._refresh_structure_tab_attention)

        model = table.model()
        if model is not None:
            model.dataChanged.connect(self._refresh_structure_tab_attention)
            model.rowsInserted.connect(self._refresh_structure_tab_attention)
            model.rowsRemoved.connect(self._refresh_structure_tab_attention)
            model.columnsInserted.connect(self._refresh_structure_tab_attention)
            model.columnsRemoved.connect(self._refresh_structure_tab_attention)

    def _structure_tab_has_incomplete_details(self) -> bool:
        tab = self.structure_tab
        if tab is None or not hasattr(tab, '_collect_incomplete_actions'):
            return False
        try:
            summary = tab._collect_incomplete_actions()
            return int(summary.get('incomplete_total', 0)) > 0
        except Exception:
            return False

    def _refresh_structure_tab_attention(self, *_args):
        tab = self.structure_tab
        if tab is None:
            return
        idx = self.tabs.indexOf(tab)
        if idx < 0:
            return

        if self._structure_tab_has_incomplete_details():
            self.tabs.tabBar().setTabTextColor(idx, QColor('#E6A23C'))
            return

        if self.tabs.currentIndex() == idx:
            self.tabs.tabBar().setTabTextColor(idx, QColor('#0F2D46'))
        else:
            self.tabs.tabBar().setTabTextColor(idx, QColor('#1F3B57'))

    def _initialize_ec_category_tab_management(self):
        for tab_name, tab_widget in self._managed_tabs.items():
            self._dirty_state.register(tab_name, tab_widget, tab_name)

        trackable_tabs = [
            "EC Creation Form",
            "Revised Items & Components",
            "OBS Parts",
            "Where Used",
            "Orphan Analysis",
            "Structure Sheet",
            "Inventory_Cost",
            "Watch_List",
            "Reports",
            "User Notes",
        ]
        for tab_name in trackable_tabs:
            tab_widget = self._managed_tabs.get(tab_name)
            if tab_widget is not None:
                self._dirty_watchers.append(TabDirtyWatcher(tab_name, tab_widget, self._dirty_state))

        self.ec_form_tab.ec_category_selected.connect(self._on_ec_category_selected)
        self.ec_form_tab.reset_form_requested.connect(self._on_reset_ec_creation_form_requested)
        self._apply_visibility_for_category(None)

    def _set_tab_visible(self, tab_name: str, visible: bool):
        tab_widget = self._managed_tabs.get(tab_name)
        if tab_widget is None:
            return
        idx = self.tabs.indexOf(tab_widget)
        if idx < 0:
            return

        if hasattr(self.tabs, 'setTabVisible'):
            self.tabs.setTabVisible(idx, visible)
        self.tabs.setTabEnabled(idx, visible)

    def _apply_visibility_for_category(self, category: str | None):
        if category:
            visible_tabs = self._tab_visibility_policy.tabs_for_category(category)
        else:
            visible_tabs = set(self._tab_visibility_policy.BASE_VISIBLE_TABS)

        for tab_name in self._managed_tabs:
            self._set_tab_visible(tab_name, tab_name in visible_tabs)

        current_widget = self.tabs.currentWidget()
        if current_widget is not None:
            current_name = None
            for name, widget in self._managed_tabs.items():
                if widget is current_widget:
                    current_name = name
                    break
            if current_name not in visible_tabs:
                self.tabs.setCurrentWidget(self.ec_form_tab)

    def _on_ec_category_selected(self, new_category: str):
        if self._category_switch_in_progress:
            return

        new_cat = self._tab_visibility_policy.normalize_category(new_category)
        if not new_cat:
            return

        if self._current_ec_category is None:
            self._current_ec_category = new_cat
            self._apply_visibility_for_category(new_cat)
            return

        if new_cat == self._current_ec_category:
            self._apply_visibility_for_category(new_cat)
            return

        tabs_to_hide = self._tab_visibility_policy.conditional_tabs_removed_by_change(
            self._current_ec_category,
            new_cat,
        )
        dirty_tabs_to_hide = [name for name in tabs_to_hide if self._dirty_state.is_dirty(name)]

        if dirty_tabs_to_hide:
            tab_names_text = ", ".join(dirty_tabs_to_hide)
            answer = QMessageBox.question(
                self,
                "Unsaved Changes",
                f"Changes made in [{tab_names_text}] will not be saved. Do you want to continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if answer != QMessageBox.StandardButton.Yes:
                self._category_switch_in_progress = True
                try:
                    self.ec_form_tab.set_selected_category(self._current_ec_category)
                finally:
                    self._category_switch_in_progress = False
                return

            self._discard_hidden_tab_changes(dirty_tabs_to_hide)
            self._dirty_state.clear_many(dirty_tabs_to_hide)

        self._current_ec_category = new_cat
        self._apply_visibility_for_category(new_cat)

    def _discard_hidden_tab_changes(self, tab_names: List[str]):
        for tab_name in tab_names:
            tab_widget = self._managed_tabs.get(tab_name)
            if tab_widget is None:
                continue
            if tab_name == "OBS Parts" and hasattr(tab_widget, 'reset'):
                tab_widget.reset()
            elif tab_name == "Structure Sheet" and hasattr(tab_widget, '_reset_struct_table'):
                tab_widget._reset_struct_table()

    def _on_main_tab_changed(self, new_index: int):
        if self._tab_switch_guard:
            return
        prev = getattr(self, '_prev_tab_index', -1)
        if prev < 0 or prev == new_index:
            self._prev_tab_index = new_index
            self._refresh_structure_tab_attention()
            return

        prev_widget = self.tabs.widget(prev)
        if isinstance(prev_widget, StructureSheetTab):
            self._refresh_structure_tab_attention()

        new_widget = self.tabs.widget(new_index)
        if isinstance(new_widget, StructureSheetTab) and hasattr(new_widget, '_reapply_action_visibility_all_rows'):
            new_widget._reapply_action_visibility_all_rows()

        self._prev_tab_index = new_index
        self._refresh_structure_tab_attention()

    def aggregate_data(self)->Dict[str,Any]:
        payload = {'obs_parts': self.obs_tab.to_dict()}
        if hasattr(self.ec_form_tab, 'to_dict'):
            payload['front_page'] = self.ec_form_tab.to_dict()
        return payload

    def apply_data(self, data: Dict[str,Any]):
        if not data: return
        if 'front_page' in data and hasattr(self.ec_form_tab, 'from_dict'):
            self.ec_form_tab.from_dict(data['front_page'])
        if 'obs_parts' in data: self.obs_tab.from_dict(data['obs_parts'])

    def save_data(self):
        data=self.aggregate_data()
        try:
            DATA_FILE.write_text(json.dumps(data, indent=2), encoding='utf-8')
            self._dirty_state.clear_all()
            QMessageBox.information(self, "Saved", f"Data saved to {DATA_FILE.name}")
        except Exception as e:
            QMessageBox.warning(self, "Save Failed", str(e))

    def load_data_if_exists(self):
        try:
            if DATA_FILE.exists():
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
        except Exception as e:
            QMessageBox.warning(self, "Load Failed", str(e))

    def load_data_dialog(self):
        if DATA_FILE.exists():
            try:
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
                QMessageBox.information(self, "Loaded", f"Data loaded from {DATA_FILE.name}")
            except Exception as e:
                QMessageBox.warning(self, "Load Failed", str(e))
        else:
            QMessageBox.information(self, "No Save Found", "No saved data file found yet. Click Save to create one.")

    def _on_reset_ec_creation_form_requested(self):
        ans = QMessageBox.question(
            self,
            "Reset All",
            "This will reset EC Creation Form and hide all tabs except READ ME and EC Creation Form. Do you want to continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if ans != QMessageBox.StandardButton.Yes:
            return

        try:
            self.ec_form_tab.reset()
            self._current_ec_category = None
            self._apply_visibility_for_category(None)
            self.tabs.setCurrentWidget(self.ec_form_tab)
            self._dirty_state.clear_all()
            QMessageBox.information(
                self,
                "Reset Complete",
                "EC Creation Form has been reset. Only READ ME and EC Creation Form are visible now.",
            )
        except Exception as e:
            QMessageBox.warning(self, "Reset Failed", str(e))

    def reset_app(self):
        ans=QMessageBox.question(self, "Reset App", "This will clear all fields, reset tables and remove any saved data. Do you want to continue?",
                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if ans==QMessageBox.StandardButton.Yes:
            try:
                if hasattr(self.ec_form_tab, 'reset'):
                    self.ec_form_tab.reset()
                self.obs_tab.reset()
                if hasattr(self.structure_tab, '_reset_struct_table'):
                    self.structure_tab._reset_struct_table()
                if hasattr(self.inventory_cost_tab, 'reset_tab'):
                    self.inventory_cost_tab.reset_tab()
                if hasattr(self.report_tab, 'reset_report'):
                    self.report_tab.reset_report()
                if DATA_FILE.exists(): DATA_FILE.unlink()
                self._current_ec_category = None
                self._apply_visibility_for_category(None)
                self._dirty_state.clear_all()
                QMessageBox.information(self, "Reset Complete", "Application data has been reset.")
            except Exception as e:
                QMessageBox.warning(self, "Reset Failed", str(e))

class OrphanAnalysisTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)

        header = QWidget(self)
        h = QHBoxLayout(header)
        h.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))

        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        self.rb_without = QRadioButton("Without Replacement*")
        self.rb_with = QRadioButton("With / Without Replacement")
        self.rb_without.setChecked(True)

        grp = QButtonGroup(self)
        grp.addButton(self.rb_without, 0)
        grp.addButton(self.rb_with, 1)

        h.addWidget(title)
        h.addSpacing(12)
        h.addWidget(self.rb_without)
        h.addWidget(self.rb_with)
        h.addStretch(1)
        self.btn_export_excel = QPushButton("Export Excel")
        self.btn_export_excel.setToolTip("Export OBS Parts and active Orphan Analysis tabs to Excel")
        h.addWidget(self.btn_export_excel)
        outer.addWidget(header)

        self.lbl_without_replacement_note = QLabel(
            "* This Orphan Analysis evaluates multiple BOM levels in a single step, "
            "assuming all identified orphan components are obsolete with no replacements."
        )
        self.lbl_without_replacement_note.setWordWrap(True)
        self.lbl_without_replacement_note.setStyleSheet("color:#7A1C21; font-size:11px;")
        outer.addWidget(self.lbl_without_replacement_note)

        self.lbl_with_replacement_note = QLabel(
            "Note: Export WU to multiple level only if Orphan Parents to be identified."
        )
        self.lbl_with_replacement_note.setWordWrap(True)
        self.lbl_with_replacement_note.setStyleSheet("color:#C1272D; font-size:11px;")
        outer.addWidget(self.lbl_with_replacement_note)


        # --- Sub-tabs for each mode ---
        self.tabs = QTabWidget()
        self.tab_without_imp_bom = OrphanOBSSubTab()
        self.tab_without_wu_removed = WURemovedBOMItemsTab(obs_provider, imp_bom_provider=self.tab_without_imp_bom)
        self.tab_with_obs = OrphanOBSSubTab()
        self.tab_with_repl = OrphanOBSSubTab()
        self.tab_with_wu_removed = WURemovedBOMItemsTab(obs_provider, imp_bom_provider=self.tab_with_obs)

        fixed_bom_tooltip = 'BOM Level is fixed to 1 for With / Without Replacement mode.'
        self.tab_with_obs.set_fixed_bom_level(1, fixed_bom_tooltip)
        self.tab_with_repl.set_fixed_bom_level(1, fixed_bom_tooltip)
        self.tab_with_obs.suppress_import_tool_comments = True
        self.tab_with_repl.suppress_import_tool_comments = True
        self.tab_with_wu_removed.enable_with_replacement_compare(self.tab_with_repl)
        self.tab_with_wu_removed.configure_orphan_append_behavior(prompt_replacement_updates=True)

        # Customize only the With/Without Replacement -> replacement BOM button behavior.
        self.tab_with_repl.btn_import.setText("Import BOM of REPL Parts")
        try:
            self.tab_with_repl.btn_import.clicked.disconnect()
        except Exception:
            pass
        self.tab_with_repl.btn_import.clicked.connect(self.tab_with_repl._import_bom_repl_parts)

        def set_subtabs(mode):
            self.tabs.clear()
            if mode == 'without':
                self.tabs.addTab(self.tab_without_imp_bom, "Imp BOM_OBS Parts")
                self.tabs.addTab(self.tab_without_wu_removed, "WU of Removed BOM Items")
            elif mode == 'with':
                self.tabs.addTab(self.tab_with_obs, "Imp BOM_OBS Parts")
                self.tabs.addTab(self.tab_with_repl, "Imp REPL Parts BOM")
                self.tabs.addTab(self.tab_with_wu_removed, "WU of Removed BOM Items")
        self._set_subtabs = set_subtabs

        outer.addWidget(self.tabs)
        self.rb_without.toggled.connect(lambda checked: checked and self._set_subtabs('without'))
        self.rb_with.toggled.connect(lambda checked: checked and self._set_subtabs('with'))
        self.rb_without.toggled.connect(lambda _checked: self._update_without_replacement_note())
        self.rb_with.toggled.connect(lambda _checked: self._update_without_replacement_note())
        self.btn_export_excel.clicked.connect(self.export_orphan_analysis_excel)
        self._set_subtabs('without')
        self._update_without_replacement_note()

        obs_table = getattr(self.obs_provider, 'table', None)
        if obs_table is not None:
            obs_table.itemChanged.connect(self._on_obs_table_changed)
            try:
                model = obs_table.model()
                model.rowsInserted.connect(self._on_obs_table_structure_changed)
                model.rowsRemoved.connect(self._on_obs_table_structure_changed)
            except Exception:
                pass
        self._auto_switch_mode_from_obs()

    def _on_obs_table_changed(self, _item):
        self._auto_switch_mode_from_obs()

    def _on_obs_table_structure_changed(self, *_args):
        self._auto_switch_mode_from_obs()

    def _update_without_replacement_note(self):
        self.lbl_without_replacement_note.setVisible(self.rb_without.isChecked())
        self.lbl_with_replacement_note.setVisible(self.rb_with.isChecked())

    def _has_any_valid_replacement(self) -> bool:
        t = getattr(self.obs_provider, 'table', None)
        if t is None:
            return False
        for r in range(t.rowCount()):
            rep_item = t.item(r, 3)
            rep = (rep_item.text() if rep_item else '').strip()
            if rep and not re.search(r'no\s*replacement', rep, re.IGNORECASE):
                return True
        return False

    def _auto_switch_mode_from_obs(self):
        has_replacement = self._has_any_valid_replacement()

        if has_replacement:
            if not self.rb_with.isChecked():
                self.rb_with.setChecked(True)
            self.rb_with.setEnabled(True)
            self.rb_without.setEnabled(False)
            self._set_subtabs('with')
        else:
            if not self.rb_without.isChecked():
                self.rb_without.setChecked(True)
            self.rb_without.setEnabled(True)
            self.rb_with.setEnabled(False)
            self._set_subtabs('without')

    def _table_to_export_dataframe(self, table: QTableWidget) -> pd.DataFrame:
        if table is None or table.columnCount() == 0:
            return pd.DataFrame()

        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text().strip() if h and h.text() else f'Column {c + 1}')

        rows = []
        for r in range(table.rowCount()):
            row_data = {}
            has_value = False
            for c, header in enumerate(headers):
                value = ''
                cell_widget = table.cellWidget(r, c)

                if isinstance(cell_widget, QComboBox):
                    value = cell_widget.currentText().strip()
                elif isinstance(cell_widget, QTextEdit):
                    value = cell_widget.toPlainText().strip()
                elif isinstance(cell_widget, QCheckBox):
                    value = 'Yes' if cell_widget.isChecked() else 'No'
                elif isinstance(cell_widget, QWidget):
                    chk = cell_widget.findChild(QCheckBox)
                    if chk is not None:
                        value = 'Yes' if chk.isChecked() else 'No'
                    else:
                        item = table.item(r, c)
                        value = item.text() if item and item.text() else ''  # Preserve indents
                else:
                    item = table.item(r, c)
                    value = item.text() if item and item.text() else ''  # Preserve indents

                if value:
                    has_value = True
                row_data[header] = value

            if has_value:
                rows.append(row_data)

        export_headers = [h for h in headers if h.strip().lower() != 'select']
        df = pd.DataFrame(rows, columns=headers)
        if not export_headers:
            return pd.DataFrame()
        return df.loc[:, export_headers]

    @staticmethod
    def _safe_excel_sheet_name(name: str) -> str:
        cleaned = re.sub(r'[\\/*?:\[\]]', '_', (name or '').strip())
        return (cleaned or 'Sheet')[:31]

    def _mode_export_sheets(self):
        if self.rb_with.isChecked():
            return [
                ('Imp BOM_OBS Parts', self.tab_with_obs.table),
                ('Imp REPL Parts BOM', self.tab_with_repl.table),
                ('WU of Removed BOM Items', self.tab_with_wu_removed.table),
            ]
        return [
            ('Imp BOM_OBS Parts', self.tab_without_imp_bom.table),
            ('WU of Removed BOM Items', self.tab_without_wu_removed.table),
        ]

    def _build_orphan_overview_chains(self, wu_df: pd.DataFrame) -> dict[int, list[list[tuple[str, str]]]]:
        if wu_df is None or wu_df.empty:
            return {}

        cols = {str(c).strip().lower(): c for c in wu_df.columns}
        part_col = cols.get('part')
        status_col = cols.get('orphans list')
        level_col = cols.get('wu level')
        if part_col is None or status_col is None or level_col is None:
            return {}

        rows = []
        for _, row in wu_df.iterrows():
            part = str(row.get(part_col, '') or '').strip()
            status = str(row.get(status_col, '') or '').strip()
            level_txt = str(row.get(level_col, '') or '').strip()
            if not part:
                continue
            try:
                level = int(float(level_txt))
            except (ValueError, TypeError):
                continue
            rows.append({'part': part, 'status': status, 'level': level})

        if not rows:
            return {}

        parent_idx = {}
        stack = []
        for idx, rec in enumerate(rows):
            lvl = rec['level']
            while stack and rows[stack[-1]]['level'] >= lvl:
                stack.pop()
            parent_idx[idx] = stack[-1] if stack else None
            stack.append(idx)

        orphan_re = re.compile(r'^orphan\d+$', re.IGNORECASE)
        obsolete_re = re.compile(r'^obsolete$', re.IGNORECASE)

        def _orphan_num(status_text: str):
            m = re.match(r'^orphan\s*(\d+)$', (status_text or '').strip(), re.IGNORECASE)
            if not m:
                return None
            try:
                return int(m.group(1))
            except (ValueError, TypeError):
                return None

        chains_by_level: dict[int, list[list[tuple[str, str]]]] = {}
        seen_by_level: dict[int, set[tuple[tuple[str, str], ...]]] = {}
        for idx, rec in enumerate(rows):
            status = rec['status'].strip()
            if not orphan_re.match(status):
                continue
            orphan_level = _orphan_num(status)
            if orphan_level is None:
                continue

            chain_idx = []
            visited = set()
            cur = idx
            while cur is not None and cur not in visited:
                visited.add(cur)
                chain_idx.append(cur)
                cur_status = rows[cur]['status'].strip()
                if obsolete_re.match(cur_status):
                    break
                cur = parent_idx.get(cur)

            chain_items = [(rows[i]['part'], rows[i]['status'].strip()) for i in chain_idx]
            if not chain_items:
                continue

            key = tuple(chain_items)
            seen_for_level = seen_by_level.setdefault(orphan_level, set())
            if key in seen_for_level:
                continue
            seen_for_level.add(key)
            chains_by_level.setdefault(orphan_level, []).append(chain_items)

        return chains_by_level

    def _write_orphans_overview_sheet(self, wb, chains_by_level: dict[int, list[list[tuple[str, str]]]]):
        if 'Orphans Overview' in wb.sheetnames:
            del wb['Orphans Overview']
        ws = wb.create_sheet('Orphans Overview')
        ws.sheet_view.showGridLines = False

        section_levels = sorted(chains_by_level.keys())
        if not section_levels:
            ws.cell(row=1, column=1, value='No orphan chains found in WU of Removed BOM Items.')
            ws.column_dimensions['A'].width = 70
            return

        section_gap = 2
        section_width = 1
        first_data_row = 3
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        for sec_idx, orphan_level in enumerate(section_levels):
            col = 1 + sec_idx * (section_width + section_gap)
            ws.cell(row=1, column=col, value=f'ORPHAN {orphan_level} CHAINS')

            row_ptr = first_data_row
            for chain in chains_by_level.get(orphan_level, []):
                for depth, (part, status) in enumerate(chain):
                    indent = ' ' * (depth * 3)
                    line = f"{indent}{part}  ({status})"
                    cell = ws.cell(row=row_ptr, column=col, value=line)
                    cell.fill = white_fill
                    row_ptr += 1
                row_ptr += 1

            max_len = 0
            for r in range(1, max(ws.max_row, row_ptr) + 1):
                v = ws.cell(row=r, column=col).value
                txt = '' if v is None else str(v)
                if len(txt) > max_len:
                    max_len = len(txt)
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 28), 120)

    def export_orphan_analysis_excel(self):
        try:
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                'Export Orphan Analysis',
                'orphan_analysis_export.xlsx',
                'Excel Files (*.xlsx)',
            )
            if not save_path:
                return

            if not save_path.lower().endswith('.xlsx'):
                save_path += '.xlsx'

            sheets_to_export = []

            obs_table = getattr(self.obs_provider, 'table', None)
            if obs_table is not None:
                sheets_to_export.append(('OBS Parts', obs_table))

            sheets_to_export.extend(self._mode_export_sheets())

            written = 0
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                for sheet_name, table in sheets_to_export:
                    if table is None:
                        continue
                    df = self._table_to_export_dataframe(table)
                    df.to_excel(
                        writer,
                        index=False,
                        sheet_name=self._safe_excel_sheet_name(sheet_name),
                    )
                    written += 1

            # Post-process workbook: add Orphans Overview, sky-blue level-0 rows, autofit.
            orphan_chains_by_level: dict[int, list[list[tuple[str, str]]]] = {}
            try:
                wu_df = pd.read_excel(save_path, sheet_name='WU of Removed BOM Items', dtype=str)
                orphan_chains_by_level = self._build_orphan_overview_chains(wu_df)
            except Exception:
                orphan_chains_by_level = {}

            wb = load_workbook(save_path)
            sky_blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

            for sheet_name, table in sheets_to_export:
                if table is None:
                    continue
                safe_name = self._safe_excel_sheet_name(sheet_name)
                if safe_name not in wb.sheetnames:
                    continue
                sheet = wb[safe_name]

                # Find hierarchy level column (BOM Level preferred, WU Level fallback).
                level_col = None
                for col_idx, cell in enumerate(sheet[1], 1):
                    header = str(cell.value).strip().lower() if cell.value is not None else ''
                    if header == 'bom level':
                        level_col = col_idx
                        break
                if level_col is None:
                    for col_idx, cell in enumerate(sheet[1], 1):
                        header = str(cell.value).strip().lower() if cell.value is not None else ''
                        if header == 'wu level':
                            level_col = col_idx
                            break

                # Apply sky blue to rows where level == 0.
                if level_col is not None:
                    for row_idx in range(2, sheet.max_row + 1):
                        level_cell = sheet.cell(row=row_idx, column=level_col)
                        if level_cell.value is None:
                            continue
                        try:
                            if float(str(level_cell.value).strip()) == 0.0:
                                for col_idx in range(1, sheet.max_column + 1):
                                    sheet.cell(row=row_idx, column=col_idx).fill = sky_blue_fill
                        except (ValueError, TypeError):
                            pass

            # Add Orphans Overview as the last sheet.
            self._write_orphans_overview_sheet(wb, orphan_chains_by_level)

            # Auto-fit all columns in all sheets.
            for sheet in wb.worksheets:
                for col_idx in range(1, sheet.max_column + 1):
                    max_len = 0
                    for row_idx in range(1, sheet.max_row + 1):
                        value = sheet.cell(row=row_idx, column=col_idx).value
                        txt = '' if value is None else str(value)
                        if len(txt) > max_len:
                            max_len = len(txt)
                    sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 120)

            wb.save(save_path)

            mode_label = 'With / Without Replacement' if self.rb_with.isChecked() else 'Without Replacement'
            QMessageBox.information(
                self,
                'Export Complete',
                f'Exported {written} sheet(s) to:\n{save_path}\nMode: {mode_label}'
            )
        except Exception as e:
            QMessageBox.warning(self, 'Export Error', str(e))



def run():
    app=QApplication(sys.argv); app.setStyle("Fusion")
    pal=QPalette(); pal.setColor(QPalette.ColorRole.Window, QColor(247,250,253)); pal.setColor(QPalette.ColorRole.Base, QColor(255,255,255))
    pal.setColor(QPalette.ColorRole.AlternateBase, QColor(241,246,252)); pal.setColor(QPalette.ColorRole.Text, QColor(28,41,56)); app.setPalette(pal)
    win=MainWindow(); win.show(); sys.exit(app.exec())

if __name__=='__main__': run()

