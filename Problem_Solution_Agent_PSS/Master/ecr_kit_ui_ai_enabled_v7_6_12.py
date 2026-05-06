# ecr_kit_ui.py (Enhanced v7.5.5 – Auto Excel conversion: silently open in Excel, SaveAs .xlsx, then import; OBS copy buttons)
import sys
import json
from pathlib import Path
from typing import Any, Dict, List
from PyQt6.QtWidgets import (
    QStyle,
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QTextEdit, QTableWidget, QTableWidgetItem, QCheckBox,
    QPushButton, QProgressBar, QSizePolicy, QScrollArea, QToolBar,
    QMessageBox, QComboBox, QFileDialog, QHeaderView, QGraphicsDropShadowEffect, QFrame,
 QGridLayout
)
from PyQt6.QtGui import QPalette, QColor, QFont, QGuiApplication, QAction
from PyQt6.QtCore import Qt
APP_TITLE = "ECR Kit Assistant"
DATA_FILE = Path(__file__).with_name('ecr_kit_data.json')
TEMPLATE_FILE = Path(__file__).with_name('Obs_parts_template.xlsx')
import traceback

def _handle_exception(exc_type, exc_value, exc_tb):
    try:
        log_path = Path(__file__).with_name('error_log.txt')
        msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log_path.write_text(msg, encoding='utf-8')
        m = QMessageBox()
        m.setIcon(QMessageBox.Icon.Critical)
        m.setWindowTitle('Unexpected Error')
        m.setText('An unexpected error occurred. A log was written to error_log.txt')
        m.setDetailedText(msg)
        m.setStandardButtons(QMessageBox.StandardButton.Ok)
        m.show()
        app = QApplication.instance()
        if app is not None:
            if not hasattr(app, '_error_dialogs'):
                app._error_dialogs = []
            app._error_dialogs.append(m)
    except Exception:
        print('Unhandled exception:', file=sys.stderr)
        traceback.print_exception(exc_type, exc_value, exc_tb)

sys.excepthook = _handle_exception

class ReadmeTab(QWidget):
    def __init__(self, readme_path: Path):
        super().__init__()
        layout = QVBoxLayout(self)
        title = QLabel("READ ME")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.DemiBold))
        layout.addWidget(title)
        text = QTextEdit(); text.setReadOnly(True)
        content = "--Wait for the Instructions on how to use--."
        try:
            if readme_path.exists():
                content = readme_path.read_text(encoding="utf-8")
        except Exception as e:
            content = f"Error opening README: {e}"
        text.setPlainText(content)
        layout.addWidget(text)

# ---------- Helpers ----------

from PyQt6.QtGui import QColor

def get_orphan_color(orphan_level: str):
    lvl = orphan_level.lower()
    if lvl == 'orphan1': return QColor('#C0392B')
    if lvl == 'orphan2': return QColor('#E67E22')
    if lvl.startswith('orphan'): return QColor('#2980B9')
    return None


def _excel_width_to_px(widget: QWidget, chars: int) -> int:
    fm = widget.fontMetrics()
    return fm.horizontalAdvance('0' * max(1, chars)) + 22

class ShiftEnterTextEdit(QTextEdit):
    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if event.modifiers() & Qt.KeyboardModifier.ShiftModifier:
                self.insertPlainText('\n'); return
        super().keyPressEvent(event)

class ECRFrontPageTab(QWidget):
    ROW_HEIGHT = 40  # compact rows to show ~5 checkpoints per view
    # Updated checklist per user request
    CHECKLIST_DEF = [
        ("Project Association", ["Does the ECR include a Project (PCR)?"]),
        ("Safety & Compliance", ["Provide PSER details if any safety incident occurred.", "Is a PCN required and approved by ROW and CE?"]),
        ("Part Release Status", ["Are all BTPs and kits EVAL released?", "If the parent part is in production, is the replacement part (BTP) production released?"]),
        ("Design Analysis", ["Is VA/VB analysis completed (for new designs/parts only)?", "PACE / DASH parts addressed?"]),
        ("Watchlist & Spares", ["Are new parts/designs MLO certified / were parent/previous parts MLO certified?", "Do we have AGS approval for OBS parts (sparable) without replacement?"]),
        ("Config Documents", ["Is CCR available with change matrix details for new options/reference designator updates?"]),
        ("ECR Strategies Identified", ["Provide reason code, effective strategy, priority, and disposition."]),
        ("Multi BU Alignment", ["If scope impacts multiple BUs/products, verify and confirm all affected BUs/products are listed in the project and approved."]),
        ("Interchangeability & Testing", [
            "Are the changes FFF compatible? Provide system details.",
            "Are test results/FDR available for all FFF impacted parts/critical parts?",
        ]),
        ("Costs and Savings", [
            "If the project relates to DCR (cost reduction), provide cost-saving details.",
        ])
    ]
    _QUESTIONS_TOTAL = sum(len(qs) for _, qs in CHECKLIST_DEF)
    assert len(CHECKLIST_DEF) == 10 and _QUESTIONS_TOTAL == 15

    def __init__(self):
        super().__init__()
        top = QVBoxLayout(self)

        scroll = QScrollArea(self); scroll.setWidgetResizable(True); top.addWidget(scroll)
        content = QWidget(); outer = QVBoxLayout(content); outer.setSpacing(10); outer.setContentsMargins(8,8,8,8); scroll.setWidget(content)

        # Header
        row1 = QWidget(); r1 = QHBoxLayout(row1); r1.setContentsMargins(12,0,0,0)
        self.ecr_no = QLineEdit(); self.ecr_no.setPlaceholderText("ECR#")
        self.eco_primer = QLineEdit(); self.eco_primer.setPlaceholderText("ECO Primer Refs#")
        self.ec_category = QLineEdit(); self.ec_category.setPlaceholderText("EC Category")
        self.bu = QLineEdit(); self.bu.setPlaceholderText("BU")
        self.tco = QLineEdit(); self.tco.setPlaceholderText("TCO")
        self.project_no = QLineEdit(); self.project_no.setPlaceholderText("Project#")
        self.product = QLineEdit(); self.product.setPlaceholderText("Product")
        for w in [self.ecr_no, self.eco_primer, self.ec_category, self.bu, self.tco, self.project_no, self.product]:
            w.setFixedHeight(28); r1.addWidget(w)
        row2 = QWidget(); r2 = QHBoxLayout(row2); r2.setContentsMargins(12,0,0,0)
        self.affected_modules = QLineEdit(); self.affected_modules.setPlaceholderText("Affected Module(s)")
        self.place = QLineEdit(); self.place.setPlaceholderText("Place")
        for w in [self.affected_modules, self.place]:
            w.setFixedHeight(28); r2.addWidget(w)
        outer.addWidget(row1); outer.addWidget(row2)

        lbl = QLabel("ECR Creation Checklist"); lbl.setFont(QFont("Segoe UI", 5, QFont.Weight.DemiBold))
        temp_lbl_row = QWidget(); tl = QHBoxLayout(temp_lbl_row); tl.setContentsMargins(12,0,0,0); tl.addWidget(lbl); outer.addWidget(temp_lbl_row)

        # Table
        table_row = QWidget(); tr = QHBoxLayout(table_row); tr.setContentsMargins(12,0,0,0)
        self.table = QTableWidget(len(self.CHECKLIST_DEF), 9)
        self.table.setHorizontalHeaderLabels(["Sl.No","Completion","Category","Check Point","Validation","Number","Action Owner","Comments","Due Date"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.table.setWordWrap(True); self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
            QHeaderView::section { background:#E1F0FF; color:#0F2D46; font-weight:600; border:1px solid #C9E2FF; padding:4px; }
            QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        for row,(cat,qlist) in enumerate(self.CHECKLIST_DEF):
            sl = QTableWidgetItem(str(row+1)); sl.setFlags(sl.flags() & ~Qt.ItemFlag.ItemIsEditable); sl.setTextAlignment(Qt.AlignmentFlag.AlignCenter); self.table.setItem(row,0,sl)
            chk = QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); cont._chk=chk; self.table.setCellWidget(row,1,cont)
            cat_item = QTableWidgetItem(cat); cat_item.setFlags(cat_item.flags() & ~Qt.ItemFlag.ItemIsEditable); self.table.setItem(row,2,cat_item)
            cp_label = QLabel("\n".join(qlist)); cp_label.setTextFormat(Qt.TextFormat.PlainText); cp_label.setWordWrap(True); self.table.setCellWidget(row,3,cp_label)
            combo = QComboBox(); combo.addItems(['','YES','NO','N/A']); self.table.setCellWidget(row,4,combo)
            def _on_change(txt, r=row):
                sel=(txt or '').strip().upper(); w=self.table.cellWidget(r,1)
                if isinstance(w, QWidget) and hasattr(w,'_chk'): w._chk.setChecked(sel in ('YES','N/A'))
            combo.currentTextChanged.connect(_on_change)
            self.table.setItem(row,5,QTableWidgetItem(""))
            self.table.setItem(row,6,QTableWidgetItem(""))
            te=ShiftEnterTextEdit(); te.setPlaceholderText("Add comments… (Shift+Enter for new line)"); te.setFrameShape(QFrame.Shape.NoFrame); te.setStyleSheet("QTextEdit { padding:4px; }"); self.table.setCellWidget(row,7,te)
        header=self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2,QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(2,_excel_width_to_px(self.table,14))
        header.setSectionResizeMode(3,QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4,QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5,QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(6,QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(7,QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(5,_excel_width_to_px(self.table,10))
        self.table.setColumnWidth(6,_excel_width_to_px(self.table,14))
        for r in range(self.table.rowCount()): self.table.setRowHeight(r,self.ROW_HEIGHT)
        tr.addWidget(self.table); outer.addWidget(table_row)

        # Progress
        progress_row = QWidget(); pr=QHBoxLayout(progress_row); pr.setContentsMargins(12,0,0,0)
        self.progress_label = QLabel(f"Checklist Progress: 0 / {len(self.CHECKLIST_DEF)}")
        self.progress_bar=QProgressBar(); self.progress_bar.setRange(0,len(self.CHECKLIST_DEF)); self.progress_bar.setFixedHeight(26); self.progress_bar.setTextVisible(True)
        pr.addWidget(self.progress_label); pr.addWidget(self.progress_bar,3); pr.addStretch(1); outer.addWidget(progress_row)

        # Inputs
        action_row = QWidget(); ar=QHBoxLayout(action_row); ar.setContentsMargins(12,0,0,0)
        self.btn_generate=QPushButton("Generate Problem Statement")
        self.title_edit=QLineEdit(); self.title_edit.setPlaceholderText("Title (max 75 chars)"); self.title_edit.setMaxLength(75)
        ar.addWidget(self.btn_generate,1); ar.addWidget(self.title_edit,3); ar.addStretch(1); outer.addWidget(action_row)

        ps_row = QWidget(); psr=QHBoxLayout(ps_row); psr.setContentsMargins(12,0,0,0)
        self.problem_edit=QTextEdit(); self.problem_edit.setPlaceholderText("Write the problem statement here (max 2000 characters)…"); self.problem_edit.setFixedHeight(140)
        psr.addWidget(self.problem_edit,3); psr.addStretch(1); outer.addWidget(ps_row)
        self.problem_edit.textChanged.connect(lambda: self._limit_text(self.problem_edit,2000))

        ss_row = QWidget(); ssr=QHBoxLayout(ss_row); ssr.setContentsMargins(12,0,0,0)
        self.solution_edit=QTextEdit(); self.solution_edit.setPlaceholderText("Write the proposed solution here (max 2000 characters)…"); self.solution_edit.setFixedHeight(140)
        ssr.addWidget(self.solution_edit,3); ssr.addStretch(1); outer.addWidget(ss_row)

        # --- AI Assisted ECR Drafting Panel (Right Side) ---
        ai_row = QWidget(); ai_layout = QHBoxLayout(ai_row); ai_layout.setContentsMargins(12,0,0,0)
        left_col = QVBoxLayout(); left_col.addWidget(self.problem_edit); left_col.addWidget(self.solution_edit)

        ai_panel = QWidget(); ai_panel.setFixedWidth(300); ai_v = QVBoxLayout(ai_panel)
        ai_title = QLabel('AI Assisted ECR Drafting'); ai_title.setFont(QFont('Segoe UI',11,QFont.Weight.DemiBold))
        btn_paste = QPushButton('-||-')
        btn_file = QPushButton('Upload PDF / Word / PPT')
        btn_email = QPushButton('Upload Email (.msg / .eml)')
        scope_lbl = QLabel('Include Data From:')
        cb_obs = QCheckBox('OBS Parts'); cb_obs.setChecked(True)
        cb_wu = QCheckBox('Where Used'); cb_wu.setChecked(True)
        cb_chk = QCheckBox('Checklist'); cb_chk.setChecked(True)
        cb_struct = QCheckBox('Structure Sheet'); cb_struct.setChecked(True)
        btn_ai = QPushButton('Generate ECR using AI'); btn_ai.setFixedHeight(34)
        info = QLabel('AI will analyze selected inputs and other tabs to generate Title, Problem and Solution.'); info.setWordWrap(True)

        for w in [ai_title, btn_paste, btn_file, btn_email, scope_lbl, cb_obs, cb_wu, cb_chk, cb_struct, btn_ai, info]: ai_v.addWidget(w)
        ai_v.addStretch(1)

        ai_layout.addLayout(left_col,3); ai_layout.addWidget(ai_panel,1); outer.addWidget(ai_row)

        self.solution_edit.textChanged.connect(lambda: self._limit_text(self.solution_edit,2000))

        self.setStyleSheet("""
            QLabel { color: #12324A; }
            QLineEdit, QTextEdit { background:#FFFFFF; border:1px solid #BBD3EA; border-radius:4px; padding:4px; }
            QLineEdit:focus, QTextEdit:focus { border-color:#639AD2; }
            QPushButton { background-color:#3BAFDA; color:#ffffff; border:1px solid #2C9CC8; border-radius:5px; padding:6px 10px; }
            QPushButton:hover { background-color:#35A0C9; }
            QProgressBar { border:1px solid #BBD3EA; border-radius:4px; background:#ECF4FF; text-align:center; color:#12324A; }
            QProgressBar::chunk { background-color:#5CC0FF; }
        """)

        def recalc():
            checked=0
            for r in range(self.table.rowCount()):
                w=self.table.cellWidget(r,1)
                if isinstance(w,QWidget) and hasattr(w,'_chk') and w._chk.isChecked(): checked+=1
            self.progress_label.setText(f"Checklist Progress: {checked} / {self.table.rowCount()}")
            self.progress_bar.setValue(checked)
        for r in range(self.table.rowCount()):
            w=self.table.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.stateChanged.connect(recalc)

    def _limit_text(self, editor: QTextEdit, max_chars: int):
        doc = editor.toPlainText()
        if len(doc)>max_chars:
            cursor=editor.textCursor(); pos=cursor.position()
            editor.blockSignals(True); editor.setPlainText(doc[:max_chars]); cursor.setPosition(min(pos,max_chars)); editor.setTextCursor(cursor); editor.blockSignals(False)

    def to_dict(self)->Dict[str,Any]:
        data={'fields':{
            'ecr_no':self.ecr_no.text(),'eco_primer':self.eco_primer.text(),'ec_category':self.ec_category.text(),
            'bu':self.bu.text(),'tco':self.tco.text(),'project_no':self.project_no.text(),'product':self.product.text(),
            'affected_modules':self.affected_modules.text(),'place':self.place.text(),'title':self.title_edit.text(),
            'problem':self.problem_edit.toPlainText(),'solution':self.solution_edit.toPlainText(),},'checklist':[]}
        t=self.table
        for r in range(t.rowCount()):
            completion=False; w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): completion=w._chk.isChecked()
            cp_widget=t.cellWidget(r,3); cp_text=cp_widget.text() if isinstance(cp_widget,QLabel) else ''
            v_widget=t.cellWidget(r,4); v_choice=v_widget.currentText() if isinstance(v_widget,QComboBox) else ''
            row={'slno':t.item(r,0).text() if t.item(r,0) else str(r+1),'completion':completion,'category':t.item(r,2).text() if t.item(r,2) else '',
                 'check_point_text':cp_text,'validation_choice':v_choice,'action_owner':t.item(r,5).text() if t.item(r,5) else '',
                 'due_date':t.item(r,6).text() if t.item(r,6) else '',
                 'comments': t.cellWidget(r,7).toPlainText() if isinstance(t.cellWidget(r,7), QTextEdit) else (t.item(r,7).text() if t.item(r,7) else '')}
            data['checklist'].append(row)
        return data

    def from_dict(self, data: Dict[str,Any]):
        f=data.get('fields',{})
        self.ecr_no.setText(f.get('ecr_no','')); self.eco_primer.setText(f.get('eco_primer','')); self.ec_category.setText(f.get('ec_category',''))
        self.bu.setText(f.get('bu','')); self.tco.setText(f.get('tco','')); self.project_no.setText(f.get('project_no','')); self.product.setText(f.get('product',''))
        self.affected_modules.setText(f.get('affected_modules','')); self.place.setText(f.get('place','')); self.title_edit.setText(f.get('title',''))
        self.problem_edit.setPlainText(f.get('problem','')); self.solution_edit.setPlainText(f.get('solution',''))
        checklist=data.get('checklist',[]); t=self.table; rows=min(len(checklist), t.rowCount())
        for r in range(rows):
            row=checklist[r]
            w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.setChecked(bool(row.get('completion',False)))
            vc=row.get('validation_choice',None); wv=t.cellWidget(r,4)
            if vc is not None and isinstance(wv,QComboBox):
                idx=wv.findText(vc); wv.setCurrentIndex(idx if idx>=0 else 0)
            def _ensure(col):
                it=t.item(r,col)
                if it is None:
                    it=QTableWidgetItem(""); t.setItem(r,col,it)
                return it
            _ensure(5).setText(row.get('action_owner','') or row.get('actioner_owner',''))
            _ensure(6).setText(row.get('due_date',''))
            w_comments=t.cellWidget(r,7)
            if isinstance(w_comments, QTextEdit): w_comments.setPlainText(row.get('comments',''))
        for r in range(t.rowCount()):
            try: t.setRowHeight(r,self.ROW_HEIGHT)
            except Exception: pass

    def reset(self):
        for w in [self.ecr_no,self.eco_primer,self.ec_category,self.bu,self.tco,self.project_no,self.product,self.affected_modules,self.place,self.title_edit]: w.clear()
        self.problem_edit.clear(); self.solution_edit.clear()
        t=self.table
        for r in range(t.rowCount()):
            w=t.cellWidget(r,1)
            if isinstance(w,QWidget) and hasattr(w,'_chk'): w._chk.setChecked(False)
        for c in (5,6):
            it=t.item(r,c)
            if it: it.setText("")
        w_comments=t.cellWidget(r,7)
        if isinstance(w_comments, QTextEdit): w_comments.clear()
        self.progress_label.setText(f"Checklist Progress: 0 / {self.table.rowCount()}"); self.progress_bar.setValue(0)

class OBSTable(QTableWidget):
    def __init__(self, parent=None, initial_rows: int = 10):
        super().__init__(initial_rows, 4, parent)
        self.setHorizontalHeaderLabels(["Select","OBS Parts","Change","Replacement"])
        self.verticalHeader().setVisible(False)
        try:
            header=self.horizontalHeader(); header.setStretchLastSection(False)
            header.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(1,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(2,QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(3,QHeaderView.ResizeMode.Interactive)
        except Exception: pass
        self._init_rows(0,self.rowCount()); self._apply_excel_widths(); self.setAlternatingRowColors(True)

    def _apply_excel_widths(self, chars:int=14):
        px=_excel_width_to_px(self,chars)
        for col in (1,2,3):
            try: self.setColumnWidth(col,px)
            except Exception: pass

    def _make_change_combo(self)->QComboBox:
        combo=QComboBox(); combo.addItems(["Obsolete","Inactivate"]); combo.setCurrentText("Obsolete"); return combo

    def _init_rows(self,start,end):
        for r in range(start,end):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); self.setCellWidget(r,0,cont)
            if not self.item(r,1): self.setItem(r,1,QTableWidgetItem(""))
            combo=self._make_change_combo(); self.setCellWidget(r,2,combo)
            if not self.item(r,3): self.setItem(r,3,QTableWidgetItem(""))

    def keyPressEvent(self,event):
        try:
            if event.matches(event.StandardKey.Copy): self._copy_selection_to_clipboard(); return
            if event.matches(event.StandardKey.Paste): self._paste_from_clipboard(); return
        except Exception: traceback.print_exc()
        super().keyPressEvent(event)

    def _copy_selection_to_clipboard(self):
        sel = self.selectedRanges()
        if not sel:
            return
        r=sel[0]; rows=[]
        for i in range(r.rowCount()):
            cols=[]
            for j in range(r.columnCount()):
                row_i=r.topRow()+i; col_j=r.leftColumn()+j
                if col_j==2:
                    w=self.cellWidget(row_i,2); txt=w.currentText() if isinstance(w,QComboBox) else ''
                else:
                    it=self.item(row_i,col_j); txt=it.text() if it else ''
                cols.append(txt)
            rows.append('\t'.join(cols))
        QGuiApplication.clipboard().setText('\n'.join(rows))

    def _ensure_rows(self,upto_row_inclusive:int):
        if upto_row_inclusive>=self.rowCount():
            old=self.rowCount(); self.setRowCount(upto_row_inclusive+1); self._init_rows(old,self.rowCount())

    def _paste_from_clipboard(self):
        text = QGuiApplication.clipboard().text()
        if not text:
            return
        start_row=self.currentRow(); start_col=self.currentColumn()
        if start_row<0:
            start_row=self._first_empty_row()
            if start_row<0: start_row=self.rowCount()
        lines=[ln for ln in text.splitlines() if ln.strip()]
        for r_offset,line in enumerate(lines):
            parts=[p.strip() for p in line.split('\t')]
            row=start_row+r_offset; self._ensure_rows(row)
            for c_offset,val in enumerate(parts):
                col=start_col+c_offset
                if col==0: continue
                if col==2:
                    w=self.cellWidget(row,2)
                    if isinstance(w,QComboBox):
                        idx=w.findText(val)
                        if idx>=0: w.setCurrentIndex(idx)
                else:
                    self.setItem(row,col,QTableWidgetItem(val))
        self._apply_excel_widths()

    def _first_empty_row(self):
        for r in range(self.rowCount()):
            it=self.item(r,1)
            if it is None or not it.text().strip(): return r
        return -1

    def delete_selected_rows(self):
        rows_to_delete = []
        for r in range(self.rowCount()):
            w = self.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and cb.isChecked():
                    rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            self.removeRow(r)
        if self.rowCount() == 0:
            self.setRowCount(10)
            self._init_rows(0, 10)

class OBSPartsTab(QWidget):
    def toggle_select_all(self):
        any_unchecked = False
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb and not cb.isChecked():
                    any_unchecked = True
                    break
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w:
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(any_unchecked)


    def select_all_rows(self):
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, 'findChild'):
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(True)

    def __init__(self):
        super().__init__()
        outer=QVBoxLayout(self)

        title_row=QHBoxLayout()
        title=QLabel("Final OBS List"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold)); title.setStyleSheet("color:#C1272D;")
        btn_template=QPushButton("Download Template")
        btn_upload=QPushButton("Upload Template")
        # New copy buttons
        btn_copy_obs=QPushButton("Copy OBS Parts")
        btn_copy_rep=QPushButton("Copy Repl Parts")
        delete_btn=QPushButton("Delete Selected")
        title_row.addWidget(title)
        title_row.addStretch(1)
        title_row.addWidget(btn_template)
        title_row.addWidget(btn_upload)
        title_row.addWidget(btn_copy_obs)
        title_row.addWidget(btn_copy_rep)
        select_all_btn = QPushButton(" Select All")
        select_all_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))
        select_all_btn.setToolTip("Toggle select / unselect all rows")
        title_row.addWidget(select_all_btn)
        title_row.addWidget(delete_btn)
        outer.addLayout(title_row)

        legend = QLabel(
            "<b>Orphan Legend:</b> "
            "<span style='color:#C0392B; font-weight:600;'>Orphan1</span> | "
            "<span style='color:#E67E22; font-weight:600;'>Orphan2</span> | "
            "<span style='color:#2980B9; font-weight:600;'>Orphan3+</span>"
        )
        legend.setStyleSheet("padding:4px;")
        outer.addWidget(legend)


        self.table=OBSTable(self, initial_rows=1); outer.addWidget(self.table)

        self.setStyleSheet("""
            QTableWidget { background:#FFFFFF; alternate-background-color:#FFF5F5; gridline-color:#F3C2C2; }
            QHeaderView::section { background:#F8D7DA; color:#7A1C21; font-weight:600; border:1px solid #E3AEB2; padding:4px; }
            QTableWidget::item:selected { background:#F5B5B8; color:#4A0E10; }
            QPushButton { background-color:#C1272D; color:#FFFFFF; border:1px solid #9F1F24; border-radius:5px; padding:6px 10px; }
            QPushButton:hover { background-color:#AD2227; }
            QComboBox { border:1px solid #E3AEB2; border-radius:4px; padding:2px 6px; }
        """)

        delete_btn.clicked.connect(self.table.delete_selected_rows)
        btn_template.clicked.connect(self.download_template)
        btn_upload.clicked.connect(self.upload_from_excel)
        btn_copy_obs.clicked.connect(self.copy_obs_parts)
        btn_copy_rep.clicked.connect(self.copy_replacement_parts)
        select_all_btn.clicked.connect(self.toggle_select_all)

    def download_template(self):
        try:
            path,_=QFileDialog.getSaveFileName(self,'Save Template',str(TEMPLATE_FILE),'Excel Files (*.xlsx)')
            if not path: return
            import pandas as pd
            df=pd.DataFrame({'OBS Parts':[''],'Change':['Obsolete'],'Replacement':['']})
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='OBS_Template')
            QMessageBox.information(self,'Template Saved',f'Template saved to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self,'Template Error',str(e))

    def upload_from_excel(self):
        try:
            path,_=QFileDialog.getOpenFileName(self,'Open OBS Parts Excel','', 'Excel Files (*.xlsx *.xls)')
            if not path: return
            import pandas as pd
            if path.lower().endswith('.xls'):
                df=pd.read_excel(path, engine='xlrd')
            else:
                df=pd.read_excel(path, engine='openpyxl')
            cols={c.strip().lower(): c for c in df.columns}
            def pick(name):
                for key in cols:
                    if key==name: return cols[key]
                return None
            c_obs=pick('obs parts') or pick('obs part') or pick('part')
            c_change=pick('change')
            c_rep=pick('replacement') or pick('replace') or pick('new part')
            if not c_obs: raise ValueError('Column "OBS Parts" is required in the Excel file.')
            rows=[]
            for _,r in df.iterrows():
                obs=str(r.get(c_obs,'')).strip()
                if not obs: continue
                change_val=str(r.get(c_change,'Obsolete')).strip() if c_change else 'Obsolete'
                if change_val not in ['Obsolete','Inactivate']: change_val='Obsolete'
                rep=str(r.get(c_rep,'')).strip() if c_rep else ''
                rows.append((obs,change_val,rep))
            if not rows:
                QMessageBox.information(self,'No Data','No valid rows found in the Excel file.'); return
            t=self.table; t.setRowCount(len(rows)); t._init_rows(0,len(rows))
            for r,(obs,change,rep) in enumerate(rows):
                t.setItem(r,1,QTableWidgetItem(obs))
                w=t.cellWidget(r,2)
                if isinstance(w,QComboBox):
                    idx=w.findText(change); w.setCurrentIndex(idx if idx>=0 else 0)
                t.setItem(r,3,QTableWidgetItem(rep))
            t._apply_excel_widths(); QMessageBox.information(self,'Upload Complete',f'Loaded {len(rows)} rows from Excel.')
        except Exception as e:
            QMessageBox.warning(self,'Upload Error',str(e))

    # NEW: Copy helpers
    def _collect_column_values(self, col_index: int) -> List[str]:
        t=self.table
        values=[]
        for r in range(t.rowCount()):
            it=t.item(r,col_index)
            if it:
                val=(it.text() or '').strip()
                if val:
                    values.append(val)
        return values

    def copy_obs_parts(self):
        values=self._collect_column_values(1)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} OBS Part number(s) to clipboard.')

    def copy_replacement_parts(self):
        values=self._collect_column_values(3)
        QGuiApplication.clipboard().setText('\n'.join(values))
        QMessageBox.information(self,'Copied', f'Copied {len(values)} Replacement part number(s) to clipboard (Image).')

    def to_dict(self)->Dict[str,Any]:
        t=self.table; rows:List[Dict[str,Any]]=[]
        for r in range(t.rowCount()):
            obs=t.item(r,1).text() if t.item(r,1) else ''
            rep=t.item(r,3).text() if t.item(r,3) else ''
            w=t.cellWidget(r,2); change=w.currentText() if isinstance(w,QComboBox) else 'Obsolete'
            if any([obs.strip(), rep.strip()]): rows.append({'obs_part':obs,'change':change,'replacement':rep})
        return {'rows': rows}

    def from_dict(self,data:Dict[str,Any]):
        rows=data.get('rows',[]); t=self.table; needed=max(10,len(rows)); t.setRowCount(needed); t._init_rows(0,needed)
        for r,row in enumerate(rows):
            t.setItem(r,1,QTableWidgetItem(row.get('obs_part','')))
            w=t.cellWidget(r,2)
            if isinstance(w,QComboBox):
                idx=w.findText(row.get('change','Obsolete'))
                if idx>=0: w.setCurrentIndex(idx)
            t.setItem(r,3,QTableWidgetItem(row.get('replacement','')))
        t._apply_excel_widths()

    def reset(self):
        t=self.table; t.setRowCount(10); t._init_rows(0,10)


class WhereUsedTab(QWidget):
    """Import 'Where Used' parents with cleanup and OBS mapping.
    - Cleans rows (backend-trim of Part) then filters per rules; display keeps original Part text
    - Preserves Excel cell background colors when available (.xlsx)
    - Adds a leading Select column (with checkboxes)
    - Inserts a Replacement column right after the Part column and auto-fills using OBS Parts tab
    - Provides utility actions: select subsets, delete, move to Structure sheet, append to OBS, export, reset
    - Does NOT modify source files on disk
    """
    def __init__(self, obs_provider=None):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)
        title_row = QHBoxLayout()
        title = QLabel("Where Used - Parents"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold))
        btn_import = QPushButton("Import 'Where Used' Parents")
        btn_import.setToolTip("Select Excel/CSV/HTML. Legacy/mismatched files are auto-converted to .xlsx using Excel, then imported. Cleanup + OBS mapping are applied to the in-app view only.")
        title_row.addWidget(title); title_row.addStretch(1); title_row.addWidget(btn_import)
        outer.addLayout(title_row)

    


        # Action buttons row
        btn_row = QHBoxLayout()
        self.btn_import_obs_multi = QPushButton("Import Where Used of OBS Parts (Multiple Level)")
        self.btn_sel_9024 = QPushButton("Select all 9024 Parents")
        self.btn_sel_options = QPushButton("Select all Options/O Class")
        self.btn_sel_esw = QPushButton("Select ESW Parents")
        self.btn_delete_sel = QPushButton("Delete the selected")
        self.btn_move_to_struct = QPushButton("Move selected Items to Structure Sheet")
        self.btn_append_obs = QPushButton("Append the selected to OBS List")
        self.btn_export = QPushButton("Export WhereUsed")
        self.btn_reset = QPushButton("Reset Where_Used")
        for b in [self.btn_import_obs_multi, self.btn_sel_9024, self.btn_sel_options, self.btn_sel_esw, self.btn_delete_sel, self.btn_move_to_struct, self.btn_append_obs, self.btn_export, self.btn_reset]:
            btn_row.addWidget(b)
        outer.addLayout(btn_row)

        self.table=QTableWidget(0,0); self.table.verticalHeader().setVisible(False); self.table.setAlternatingRowColors(True); self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        outer.addWidget(self.table)

        self.setStyleSheet("""
        QTableWidget { background:#FFFFFF; alternate-background-color:#F7FBFF; gridline-color:#D4E5F7; }
        QHeaderView::section { background:#E1F0FF; color:#0F2D46; font-weight:600; border:1px solid #C9E2FF; padding:4px; }
        QTableWidget::item:selected { background:#CDE8FF; color:#0F2D46; }
        """)
        btn_import.clicked.connect(self.import_where_used)
        self.btn_import_obs_multi.clicked.connect(self.import_where_used_of_obs_multi_level)
        self.btn_sel_9024.clicked.connect(self.select_all_9024_parents)
        self.btn_sel_options.clicked.connect(self.select_all_options)
        self.btn_sel_esw.clicked.connect(self.select_esw_parents)
        self.btn_delete_sel.clicked.connect(self.delete_selected_rows)
        self.btn_move_to_struct.clicked.connect(self.move_selected_to_structure)
        self.btn_append_obs.clicked.connect(self.append_selected_to_obs)
        self.btn_export.clicked.connect(self.export_where_used)
        self.btn_reset.clicked.connect(self.reset_where_used)

    # ---------------- Excel conversion helpers ----------------
    def _is_html_like(self, path: str)->bool:
        try:
            with open(path, 'rb') as f:
                head = f.read(2048).lstrip().lower()
                return (head.startswith(b'<!') or head.startswith(b'<html') or b'<table' in head)
        except Exception:
            return False

    def _convert_to_xlsx_via_excel(self, src_path: str)->str|None:
        try:
            from pathlib import Path as _Path
            import win32com.client
            src_abs=str(_Path(src_path).resolve())
            dst_path=str(_Path(src_abs).with_suffix(''))+'_converted.xlsx'
            excel=win32com.client.DispatchEx('Excel.Application'); excel.Visible=False; excel.DisplayAlerts=False
            wb=excel.Workbooks.Open(src_abs); wb.SaveAs(dst_path, FileFormat=51); wb.Close(SaveChanges=False); excel.Quit(); return dst_path
        except Exception:
            try:
                excel.Quit()
            except Exception:
                pass
            return None

    # ---------------- Helpers for OBS, colors, selection ----------------
    def _find_part_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            if str(c).strip().lower()=='part': return i
        return -1
    def _find_parent_col_index(self, cols:list[str])->int:
        keys=['parent','parent part','parent pn','parent number','parent part number']
        low=[str(c).strip().lower() for c in cols]
        for i,name in enumerate(low):
            for k in keys:
                if k==name or k in name: return i
        return -1
    def _find_class_col_index(self, cols:list[str])->int:
        for i,c in enumerate(cols):
            s=str(c).strip().lower()
            if any(k in s for k in ['class','type','category']): return i
        return -1
    def _build_obs_map(self)->dict:
        mapping={}
        try:
            if self.obs_provider is None: return mapping
            t=self.obs_provider.table
            for r in range(t.rowCount()):
                obs_item=t.item(r,1); rep_item=t.item(r,3)
                obs=(obs_item.text() if obs_item else '').strip(); rep=(rep_item.text() if rep_item else '')
                if obs: mapping[obs.upper()]=rep
        except Exception: pass
        return mapping
    def _read_xlsx_background_colors(self, xlsx_path:str, target_ncols:int):
        try:
            from openpyxl import load_workbook
            wb=load_workbook(xlsx_path, data_only=True); ws=wb.active
            colors=[]; first_data_row=2; ncols=target_ncols
            for r in range(first_data_row, ws.max_row+1):
                row_colors=[]
                for c in range(1, ncols+1):
                    cell=ws.cell(row=r,column=c); col=None
                    try:
                        fill=cell.fill
                        if fill and getattr(fill,'fill_type',None):
                            start=getattr(fill,'start_color',None); rgb=getattr(start,'rgb',None)
                            if isinstance(rgb,str) and len(rgb) in (6,8):
                                rgb_hex=rgb[-6:]; col=QColor('#'+rgb_hex)
                    except Exception: col=None
                    row_colors.append(col)
                colors.append(row_colors)
            return colors
        except Exception:
            return None
    def _apply_cleanup_rules(self, df, part_idx:int):
        import pandas as pd
        part_raw=df.iloc[:,part_idx].astype(str); part_trim=part_raw.str.strip(); lengths=part_trim.str.len().fillna(0)
        mask_gt=(lengths>10) & part_trim.str.upper().str.startswith('ESW')
        mask_eq=(lengths==10) & (part_trim.str[4]=='-')
        mask=mask_gt | mask_eq
        df_kept=df.loc[mask].copy(); df_kept.reset_index(drop=True, inplace=True)
        return df_kept, mask.reset_index(drop=True)
    def _find_replacement_col(self)->int:
        for i in range(self.table.columnCount()):
            h=self.table.horizontalHeaderItem(i)
            if h and str(h.text()).strip().lower()=='replacement': return i
        return -1
    def _center_replacement_column(self):
        rep_col=self._find_replacement_col()
        if rep_col>=0:
            for r in range(self.table.rowCount()):
                it=self.table.item(r,rep_col)
                if it is None:
                    it=QTableWidgetItem(''); self.table.setItem(r,rep_col,it)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def _populate_table_with_extras(self, df, orig_cols:list[str], part_idx:int, colors_2d=None):
        headers=['Select']
        for i,name in enumerate(orig_cols):
            headers.append(str(name))
            if i==part_idx: headers.append('Replacement')
        self.table.clear(); self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers); self.table.setRowCount(len(df))
        obs_map=self._build_obs_map()
        def map_col(c:int)->int: return 1 + c + (1 if c>part_idx else 0)
        for r in range(len(df)):
            chk=QCheckBox(); cont=QWidget(); h=QHBoxLayout(cont); h.setContentsMargins(0,0,0,0); h.setAlignment(Qt.AlignmentFlag.AlignCenter); h.addWidget(chk); cont._chk=chk; self.table.setCellWidget(r,0,cont)
            for c in range(len(orig_cols)):
                val=df.iloc[r,c]; txt='' if val is None else str(val)
                item=QTableWidgetItem(txt)
                if colors_2d and r<len(colors_2d) and c<len(colors_2d[r]):
                    col=colors_2d[r][c]
                    if isinstance(col,QColor): item.setBackground(col)
                self.table.setItem(r, map_col(c), item)
            part_val=str(df.iloc[r,part_idx]) if df.iloc[r,part_idx] is not None else ''
            part_key=part_val.strip().upper(); replacement=obs_map.get(part_key,'')
            rep_item=QTableWidgetItem(replacement); rep_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if colors_2d and r<len(colors_2d) and part_idx<len(colors_2d[r]):
                col=colors_2d[r][part_idx]
                if isinstance(col,QColor): rep_item.setBackground(col)
            self.table.setItem(r, 1+part_idx+1, rep_item)
        header=self.table.horizontalHeader()
        for i in range(len(headers)):
            if i==len(headers)-1: header.setSectionResizeMode(i,QHeaderView.ResizeMode.Stretch)
            else: header.setSectionResizeMode(i,QHeaderView.ResizeMode.ResizeToContents)

    # -----------------------------------------------------------
    def import_where_used(self):
        try:
            path,_=QFileDialog.getOpenFileName(self, "Import 'Where Used' Parents", '', "Excel/CSV/HTML (*.xlsx *.xls *.csv *.htm *.html);;All Files (*.*)")
            if not path: return
            import pandas as pd
            df=None; errors=[]; lower=path.lower()
            needs_excel_conversion = lower.endswith('.xls') or self._is_html_like(path)
            converted_path=None
            if needs_excel_conversion:
                converted_path=self._convert_to_xlsx_via_excel(path)
                if converted_path:
                    try: df=pd.read_excel(converted_path, engine='openpyxl')
                    except Exception as e: errors.append(f'Converted .xlsx read failed: {e}')
            if df is None:
                try:
                    if lower.endswith('.xlsx'): df=pd.read_excel(path, engine='openpyxl')
                except Exception as e: errors.append(f'XLSX read failed: {e}')
            if df is None and (lower.endswith('.htm') or lower.endswith('.html') or self._is_html_like(path)):
                try:
                    tables=pd.read_html(path)
                    if tables: df=tables[0]
                except Exception as e: errors.append(f'HTML parse failed: {e}')
            if df is None and lower.endswith('.csv'):
                try: df=pd.read_csv(path)
                except Exception as e: errors.append(f'CSV comma failed: {e}')
                if df is None:
                    try: df=pd.read_csv(path, sep='\t')
                    except Exception as e: errors.append(f'CSV tab failed: {e}')
            if df is None:
                try: df=pd.read_excel(path)
                except Exception as e: errors.append(f'Generic read_excel failed: {e}')
            if df is None:
                msg = "Failed to open the file. Tried Excel (with Excel-based conversion), HTML and CSV paths.\n" + "\n".join(errors[-6:])
                QMessageBox.warning(self,'Import Error', msg); return
            cols=[str(c) if c is not None else '' for c in df.columns]
            part_idx=self._find_part_col_index(cols)
            if part_idx<0:
                QMessageBox.warning(self,'Missing Column', "Couldn't find a 'Part' column (case-insensitive) in the selected file."); return
            # Optional OBS-only prefilter
            if hasattr(self,'_obs_only_filter') and self._obs_only_filter:
                df=self._apply_obs_only_filter(df, part_idx)
            colors_2d=None
            xlsx_source=converted_path if converted_path else (path if path.lower().endswith('.xlsx') else None)
            if xlsx_source: colors_raw=self._read_xlsx_background_colors(xlsx_source, target_ncols=len(cols))
            else: colors_raw=None
            df_kept, mask=self._apply_cleanup_rules(df, part_idx)
            if colors_raw is not None:
                kept_colors=[]; mask_list=mask.tolist()
                for ok,row_colors in zip(mask_list, colors_raw):
                    if ok: kept_colors.append(row_colors)
                colors_2d=kept_colors
            self._populate_table_with_extras(df_kept, list(df_kept.columns), part_idx, colors_2d=colors_2d)
            self._center_replacement_column()
            msg=f"Imported {len(df_kept)} cleaned row(s) from:\n{converted_path or path}"
            if converted_path: msg+="\n(The source file was auto-converted to .xlsx using Excel.)"
            QMessageBox.information(self,'Import Complete', msg)
        except Exception as e:
            QMessageBox.warning(self,'Import Error', str(e))

    # ----------------- Extra actions -----------------
    def import_where_used_of_obs_multi_level(self):
        try:
            self._obs_only_filter=True; self.import_where_used()
        finally:
            if hasattr(self,'_obs_only_filter'): delattr(self,'_obs_only_filter')
    def _apply_obs_only_filter(self, df, part_idx:int):
        obs_map=self._build_obs_map()
        if not obs_map: return df
        keys=set(obs_map.keys()); col=df.iloc[:,part_idx].astype(str).fillna('')
        mask=col.str.strip().str.upper().isin(keys)
        return df.loc[mask].reset_index(drop=True)
    def _selected_row_indices(self):
        rows=[]
        for r in range(self.table.rowCount()):
            w=self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk') and w._chk.isChecked(): rows.append(r)
        return rows
    def _headers(self)->list[str]:
        return [self.table.horizontalHeaderItem(i).text() if self.table.horizontalHeaderItem(i) else '' for i in range(self.table.columnCount())]
    def _map_original_to_table_col(self, orig_idx:int, part_idx:int)->int:
        rep_col=self._find_replacement_col()
        if part_idx>=0 and rep_col>=0 and (orig_idx>(rep_col-2)): return 1+orig_idx+1
        else: return 1+orig_idx
    def select_all_9024_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Select 9024 Parents', "Couldn't find a 'Parent' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=it.text() if it else ''
            self.table.cellWidget(r,0)._chk.setChecked(txt.strip().startswith('9024'))
    def select_all_options(self):
        headers=self._headers(); cidx=self._find_class_col_index(headers[1:])
        if cidx<0: QMessageBox.information(self,'Select Options/O Class', "Couldn't find a 'Class/Type/Category' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(cidx, part_idx)
            it=self.table.item(r, tbl_col); t=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(('OPTION' in t) or ('O CLASS' in t))
    def select_esw_parents(self):
        headers=self._headers(); pidx=self._find_parent_col_index(headers[1:])
        if pidx<0:
            pidx=self._find_part_col_index(headers[1:])
            if pidx<0: QMessageBox.information(self,'Select ESW Parents', "Couldn't find 'Parent' or 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:])
        for r in range(self.table.rowCount()):
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r, tbl_col); txt=(it.text() if it else '').strip().upper()
            self.table.cellWidget(r,0)._chk.setChecked(txt.startswith('ESW'))
    def delete_selected_rows(self):
        rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Delete Selected','No rows are selected (checkbox).'); return
        for r in reversed(rows): self.table.removeRow(r)
    def move_selected_to_structure(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'append_rows'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.'); return
            headers=self._headers(); rows_idx=self._selected_row_indices()
            if not rows_idx: QMessageBox.information(self,'Move to Structure Sheet','No rows selected.'); return
            data=[]
            for r in rows_idx:
                row=[(self.table.item(r,c).text() if self.table.item(r,c) else '') for c in range(1,self.table.columnCount())]
                data.append(row)
            target.append_rows(headers[1:], data)
            QMessageBox.information(self,'Move to Structure Sheet', f'Moved {len(data)} row(s) to Structure Sheet.')
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))
    def _append_obs_part(self, part:str):
        try:
            if not self.obs_provider: return
            t=self.obs_provider.table; key=(part or '').strip().upper()
            if not key: return
            existing=set()
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it: existing.add((it.text() or '').strip().upper())
            if key in existing: return
            target=None
            for r in range(t.rowCount()):
                it=t.item(r,1)
                if it is None or not (it.text() or '').strip(): target=r; break
            if target is None:
                target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
            t.setItem(target,1,QTableWidgetItem(part))
            w=t.cellWidget(target,2)
            if isinstance(w,QComboBox): idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
        except Exception: pass
    def append_selected_to_obs(self):
        headers=self._headers(); pidx=self._find_part_col_index(headers[1:])
        if pidx<0: QMessageBox.information(self,'Append to OBS', "Couldn't find a 'Part' column."); return
        part_idx=self._find_part_col_index(headers[1:]); rows=self._selected_row_indices()
        if not rows: QMessageBox.information(self,'Append to OBS','No rows selected.'); return
        count=0
        for r in rows:
            tbl_col=self._map_original_to_table_col(pidx, part_idx)
            it=self.table.item(r,tbl_col); part=it.text() if it else ''
            if part.strip(): self._append_obs_part(part); count+=1
        QMessageBox.information(self,'Append to OBS', f'Appended {count} part(s) to OBS List.')
    def export_where_used(self):
        try:
            path,_=QFileDialog.getSaveFileName(self,'Export WhereUsed','', 'Excel Files (*.xlsx)')
            if not path: return
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            wb=Workbook(); ws=wb.active; ws.title='WhereUsed'
            headers=self._headers(); ws.append(headers)
            for c in range(1,len(headers)+1):
                cell=ws.cell(row=1,column=c); cell.font=Font(bold=True)
                if headers[c-1].strip().lower()=='replacement': cell.alignment=Alignment(horizontal='center')
            for r in range(self.table.rowCount()):
                row_vals=[]
                for c in range(self.table.columnCount()):
                    if c==0:
                        w=self.table.cellWidget(r,0); row_vals.append('Yes' if (w and hasattr(w,'_chk') and w._chk.isChecked()) else 'No')
                    else:
                        it=self.table.item(r,c); row_vals.append(it.text() if it else '')
                ws.append(row_vals)
                for c in range(self.table.columnCount()):
                    it=self.table.item(r,c)
                    if it and it.background():
                        qcol=it.background().color()
                        if qcol.isValid():
                            rgb=f"{qcol.red():02X}{qcol.green():02X}{qcol.blue():02X}"; ws.cell(row=r+2,column=c+1).fill=PatternFill('solid', fgColor=rgb)
                    if headers[c].strip().lower()=='replacement': ws.cell(row=r+2,column=c+1).alignment=Alignment(horizontal='center')
            wb.save(path); QMessageBox.information(self,'Export Complete', f'Exported {self.table.rowCount()} row(s) to:\n{path}')
        except Exception as e:
            QMessageBox.warning(self,'Export Error', str(e))
    def reset_where_used(self):
        try:
            self.table.clear(); self.table.setRowCount(0); self.table.setColumnCount(0)
            QMessageBox.information(self,'Reset','Where Used view has been reset.')
        except Exception as e:
            QMessageBox.warning(self,'Reset Error', str(e))




class WhereUsedTabV2(WhereUsedTab):
    """Customized Where Used tab per latest requirements.
    - Remove the "Import Where Used of OBS Parts (Multiple Level)" button (hidden)
    - New two-row action panel with colored buttons
    - Selections based on Part column (with WU Level filter for 9024)
    - Select Opt & Opt Class using first 4-digit prefixes list
    - Select ESW parents using Part prefix 'ESW'
    - Move to Structure Sheet: only Part number
    - Append to OBS List: OBS Part + Replacement (if available)
    - Refresh: recompute Replacement column from current OBS Parts list
    - Delete selected, Reset Tab labels updated
    """
    def __init__(self, obs_provider=None):
        super().__init__(obs_provider=obs_provider)
        # Hide legacy buttons/row
        try:
            # Hide old buttons if they exist
            for btn_name in [
                'btn_import_obs_multi','btn_sel_9024','btn_sel_options','btn_sel_esw',
                'btn_delete_sel','btn_move_to_struct','btn_append_obs','btn_export','btn_reset'
            ]:
                btn = getattr(self, btn_name, None)
                if btn is not None:
                    btn.hide()
        except Exception:
            pass
        
        # Build new two-row action panel
        from PyQt6.QtWidgets import QWidget, QGridLayout, QLabel
        panel = QWidget(self)
        grid = QGridLayout(panel)
        grid.setContentsMargins(0,0,0,0)
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(8)

        # Buttons (new)
        self.v2_btn_sel_9024 = QPushButton("Select all 9024 Parents")
        self.v2_btn_sel_opt = QPushButton("Select all Opt & Opt Class")
        self.v2_btn_sel_esw = QPushButton("Select ESW Parents")
        self.v2_btn_sel_above_cfg = QPushButton("Select above Config")
        self.v2_btn_move = QPushButton("Move to Structure Sheet")
        self.v2_btn_append = QPushButton("Append to OBS List")
        self.v2_btn_refresh = QPushButton("Refresh")
        self.v2_btn_delete = QPushButton("Delete selected")
        self.v2_btn_reset = QPushButton("Reset Tab")
        self.v2_btn_export = QPushButton("Export WhereUsed")
        # Place Import first, then Export (no custom colors)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # Import button is the last widget; insert Export right after Import
            title_row.insertWidget(title_row.count(), self.v2_btn_export)
        except Exception:
            pass
        # Move Export button to title row (right side)
        try:
            outer = self.layout()
            title_row = outer.itemAt(0).layout()
            # export will be inserted after import
        except Exception:
            pass

        # Colors (smooth, subtle gradients)
        def btn_style(bg1, bg2, border, text='#FFFFFF'):
            return f"""
            QPushButton {{
                color:{text}; padding:4px 8px; font-size:11px; border-radius:5px;
                border:1px solid {border};
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg2});
            }}
            QPushButton:hover {{
                background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {bg1}, stop:1 {bg1});
            }}
            """
        
        self.v2_btn_sel_9024.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))    # cyan
        self.v2_btn_sel_opt.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_sel_esw.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_move.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_append.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_delete.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_export.setStyleSheet(btn_style('#26C6DA','#14A7BE','#0F93AA'))        # cyan
        self.v2_btn_reset.setStyleSheet(btn_style('#90A4AE','#7C919B','#6A7E87'))       # gray
        self.v2_btn_refresh.setStyleSheet(btn_style('#66BB6A','#4EA85A','#3F8F4A'))     # fresh green

        # Arrange in 2 rows
        row1 = [self.v2_btn_sel_9024, self.v2_btn_sel_opt, self.v2_btn_sel_esw, self.v2_btn_sel_above_cfg, self.v2_btn_refresh]
        row2 = [self.v2_btn_move, self.v2_btn_append, self.v2_btn_delete, self.v2_btn_reset]
        for c,btn in enumerate(row1):
            grid.addWidget(btn, 0, c)
        for c,btn in enumerate(row2):
            grid.addWidget(btn, 1, c)

        # Insert our panel right after the title row (index 1)
        try:
            outer: QVBoxLayout = self.layout()
            # self.layout() returns the QWidget layout of WhereUsedTab (outer QVBoxLayout)
            outer.insertWidget(1, panel)
        except Exception:
            # If insertion fails, just add at end
            self.layout().addWidget(panel)

        # Wire up actions
        self.v2_btn_sel_9024.clicked.connect(self._v2_select_9024_by_part)
        self.v2_btn_sel_opt.clicked.connect(self._v2_select_opt_optclass_by_prefix)
        self.v2_btn_sel_esw.clicked.connect(self._v2_select_esw_by_part)
        self.v2_btn_sel_above_cfg.clicked.connect(self._v2_select_above_config_block)
        self.v2_btn_move.clicked.connect(self._v2_move_selected_part_only)
        self.v2_btn_append.clicked.connect(self._v2_append_selected_with_replacement)
        self.v2_btn_refresh.clicked.connect(self._v2_refresh_replacements)
        self.v2_btn_delete.clicked.connect(self.delete_selected_rows)
        self.v2_btn_reset.clicked.connect(self.reset_where_used)
        self.v2_btn_export.clicked.connect(self.export_where_used)

        # Unified single-color style for Where Used buttons
        for b in (self.v2_btn_sel_9024, self.v2_btn_sel_opt, self.v2_btn_sel_esw,
                  self.v2_btn_sel_above_cfg, self.v2_btn_move, self.v2_btn_append,
                  self.v2_btn_refresh, self.v2_btn_delete, self.v2_btn_export, self.v2_btn_reset):
            b.setStyleSheet("QPushButton{background:#2F80ED;color:white;padding:6px 12px;border-radius:6px;}QPushButton:hover{background:#2567C7;}")

        # Also hide the legacy "Import Where Used of OBS Parts (Multiple Level)" if present in title row
        try:
            if hasattr(self, 'btn_import_obs_multi'):
                self.btn_import_obs_multi.hide()
        except Exception:
            pass

    # ---------- Helpers to find current table columns ----------
    def _v2_find_table_col_index(self, header_exact: str) -> int:
        name = (header_exact or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and (h.text() or '').strip().lower() == name:
                return i
        return -1

    def _v2_find_table_col_contains(self, keyword: str) -> int:
        key = (keyword or '').strip().lower()
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and key in (h.text() or '').strip().lower():
                return i
        return -1

    # ---------- New selection logics ----------
    def _v2_select_9024_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select 9024 Parents', "Couldn't find a 'Part' column.")
            return
        wucol = self._v2_find_table_col_contains('wu level')
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
            ok_prefix = part.startswith('9024')
            ok_wu = True
            if wucol >= 0:
                val = (self.table.item(r, wucol).text() if self.table.item(r, wucol) else '').strip()
                try:
                    ok_wu = int(val) != 0
                except Exception:
                    ok_wu = (val != '0')
            chk = ok_prefix and ok_wu
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    def _v2_select_opt_optclass_by_prefix(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select Opt & Opt Class', "Couldn't find a 'Part' column.")
            return
        prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335',
            '0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465',
            '0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'
        }
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
            first4 = part[:4]
            chk = first4 in prefixes
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    def _v2_select_esw_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select ESW Parents', "Couldn't find a 'Part' column.")
            return
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip().upper()
            chk = part.startswith('ESW')
            w = self.table.cellWidget(r,0)
            if w and hasattr(w,'_chk'):
                w._chk.setChecked(chk)

    # ---------- Move / Append / Refresh ----------
    
    def _v2_toggle_rows(self, rows):
        checks = []
        for r in rows:
            w = self.table.cellWidget(r, 0)
            if w and hasattr(w, '_chk'):
                checks.append(w._chk)
        if not checks:
            return
        select = any(not c.isChecked() for c in checks)
        for c in checks:
            c.setChecked(select)

    def _v2_select_9024_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select 9024 Parents',"Couldn't find a 'Part' column.")
            return
        wucol = self._v2_find_table_col_contains('wu level')
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip()
            ok = part.startswith('9024')
            if ok and wucol >= 0:
                wu = (self.table.item(r,wucol).text() if self.table.item(r,wucol) else '').strip()
                ok = wu != '0'
            if ok:
                rows.append(r)
        self._v2_toggle_rows(rows)

    def _v2_select_opt_optclass_by_prefix(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select Opt & Opt Class',"Couldn't find a 'Part' column.")
            return
        prefixes = {'0490','0491','0495','0497','0430','0350','0355','0351','0357','0390','0395','0397','0335','0391','0431','0435','0437','0440','0445','0455','0450','0441','0447','0457','0460','0465','0461','0467','0410','0415','0417','0411','0412','0413','0414','0360','0365','0361','0367'}
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip()
            if part[:4] in prefixes:
                rows.append(r)
        self._v2_toggle_rows(rows)

    def _v2_select_esw_by_part(self):
        pcol = self._v2_find_table_col_index('Part')
        if pcol < 0:
            QMessageBox.information(self,'Select ESW Parents',"Couldn't find a 'Part' column.")
            return
        rows = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r,pcol).text() if self.table.item(r,pcol) else '').strip().upper()
            if part.startswith('ESW'):
                rows.append(r)
        self._v2_toggle_rows(rows)
    def _v2_move_selected_part_only(self):
        try:
            main=self.window(); target=getattr(main,'structure_tab',None)
            if not target or not hasattr(target,'append_rows'):
                QMessageBox.information(self,'Move to Structure Sheet','Structure Sheet is not available yet.')
                return
            rows_idx = self._selected_row_indices()
            if not rows_idx:
                QMessageBox.information(self,'Move to Structure Sheet','No rows selected.')
                return
            pcol = self._v2_find_table_col_index('Part')
            if pcol < 0:
                QMessageBox.information(self,'Move to Structure Sheet',"Couldn't find a 'Part' column.")
                return
            data = []
            for r in rows_idx:
                val = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '')
                data.append([val])
            target.append_rows(['Part'], data)
            QMessageBox.information(self,'Move to Structure Sheet', f'Moved {len(data)} row(s) to Structure Sheet (Part only).')
        except Exception as e:
            QMessageBox.warning(self,'Move Error', str(e))

    def _v2_append_selected_with_replacement(self):
        try:
            if not self.obs_provider:
                QMessageBox.information(self,'Append to OBS', 'OBS Parts tab is not available.')
                return
            t = self.obs_provider.table
            rows = self._selected_row_indices()
            if not rows:
                QMessageBox.information(self,'Append to OBS','No rows selected.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            count = 0
            for r in rows:
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '').strip()
                rep  = (self.table.item(r, rcol).text() if (rcol>=0 and self.table.item(r, rcol)) else '').strip()
                if not part:
                    continue
                # Check existing
                existing=set()
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it: existing.add((it.text() or '').strip().upper())
                if part.upper() in existing:
                    # If exists, update replacement if empty
                    if rep:
                        for rr in range(t.rowCount()):
                            it=t.item(rr,1)
                            if it and (it.text() or '').strip().upper()==part.upper():
                                t.setItem(rr,3,QTableWidgetItem(rep))
                                break
                    continue
                target=None
                for rr in range(t.rowCount()):
                    it=t.item(rr,1)
                    if it is None or not (it.text() or '').strip():
                        target=rr; break
                if target is None:
                    target=t.rowCount(); t.setRowCount(target+1); t._init_rows(target, target+1)
                t.setItem(target,1,QTableWidgetItem(part))
                w=t.cellWidget(target,2)
                if isinstance(w,QComboBox):
                    idx=w.findText('Obsolete'); w.setCurrentIndex(idx if idx>=0 else 0)
                if rep:
                    t.setItem(target,3,QTableWidgetItem(rep))
                count += 1
            QMessageBox.information(self,'Append to OBS', f'Appended/Updated {count} part(s) to OBS List.')
        except Exception as e:
            QMessageBox.warning(self,'Append Error', str(e))

    def _v2_refresh_replacements(self):
        try:
            obs_map = self._build_obs_map()
            if not obs_map:
                QMessageBox.information(self, 'Refresh', 'No OBS parts found to refresh from.')
                return
            pcol = self._v2_find_table_col_index('Part')
            rcol = self._find_replacement_col()
            if pcol < 0 or rcol < 0:
                QMessageBox.information(self,'Refresh','Missing Part or Replacement column.')
                return
            updated = 0
            for r in range(self.table.rowCount()):
                part = (self.table.item(r, pcol).text() if self.table.item(r, pcol) else '')
                key = (part or '').strip().upper()
                rep = obs_map.get(key, '')
                it = self.table.item(r, rcol)
                if it is None:
                    it = QTableWidgetItem('')
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(r, rcol, it)
                old = it.text()
                if rep != old:
                    it.setText(rep)
                    updated += 1
            QMessageBox.information(self,'Refresh', f'Refreshed Replacement column using OBS Parts list. Updated {updated} row(s).')
        except Exception as e:
            QMessageBox.warning(self,'Refresh Error', str(e))



    def _v2_select_above_config_block(self):
        part_col = self._v2_find_table_col_index('Part')
        wu_col = self._v2_find_table_col_contains('wu level')

        if part_col < 0 or wu_col < 0:
            return

        config_prefixes = {
            '0490','0491','0495','0497','0430','0350','0355','0351','0357',
            '0390','0395','0397','0335','0391','0431','0435','0437',
            '0440','0445','0455','0450','0441','0447','0457',
            '0460','0465','0461','0467','0410','0415','0417',
            '0411','0412','0413','0414','0360','0365','0361','0367'
        }

        rows = self.table.rowCount()
        r = 0

        while r < rows:
            wu_val = (self.table.item(r, wu_col).text()
                      if self.table.item(r, wu_col) else '').strip()

            if wu_val == '0':
                block_start = r
                block_end = rows

                for i in range(r + 1, rows):
                    nxt = (self.table.item(i, wu_col).text()
                           if self.table.item(i, wu_col) else '').strip()
                    if nxt == '0':
                        block_end = i
                        break

                config_row = None
                config_indent = None

                # Step 1: find CONFIG row by part prefix (do NOT select it)
                for i in range(block_start + 1, block_end):
                    it = self.table.item(i, part_col)
                    if not it:
                        continue
                    raw = it.text()
                    stripped = raw.lstrip()
                    if not stripped:
                        continue
                    part_prefix = stripped[:4]
                    indent = len(raw) - len(stripped)

                    if part_prefix in config_prefixes:
                        config_row = i
                        config_indent = indent
                        break

                # Step 2: select only rows deeper than config indentation
                if config_row is not None:
                    for i in range(config_row + 1, block_end):
                        it = self.table.item(i, part_col)
                        if not it:
                            continue
                        raw = it.text()
                        stripped = raw.lstrip()
                        indent = len(raw) - len(stripped)

                        if indent > config_indent:
                            w = self.table.cellWidget(i, 0)
                            if w and hasattr(w, '_chk'):
                                w._chk.setChecked(True)
                        else:
                            break

                r = block_end
            else:
                r += 1


class StructureSheetTab(QWidget):
    def __init__(self):
        super().__init__()
        outer=QVBoxLayout(self)
        title=QLabel("Structure Sheet"); title.setFont(QFont("Segoe UI",14,QFont.Weight.DemiBold))
        outer.addWidget(title)
        self.table=QTableWidget(0,0); self.table.verticalHeader().setVisible(False); self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)
    def append_rows(self, headers:list[str], rows:list[list[str]]):
        if self.table.columnCount()==0:
            self.table.setColumnCount(len(headers)); self.table.setHorizontalHeaderLabels(headers)
        if self.table.columnCount()!=len(headers):
            norm=[]
            for r in rows:
                r2=(r+['']*self.table.columnCount())[:self.table.columnCount()]; norm.append(r2)
            rows=norm
        start=self.table.rowCount(); self.table.setRowCount(start+len(rows))
        for i,row in enumerate(rows):
            for j,val in enumerate(row): self.table.setItem(start+i,j,QTableWidgetItem(val))

class PlaceholderTab(QWidget):
    def __init__(self, title: str):
        super().__init__()
        l = QVBoxLayout(self)
        l.addWidget(QLabel(f"{title} – UI under development"))



class OrphanOBSSubTab(QWidget):
    def __init__(self):
        super().__init__()
        outer = QVBoxLayout(self)

        btn_row = QHBoxLayout()
        self.btn_import = QPushButton("Import Imp BOM of OBS Parts (.xlsx)")
        self.btn_select_oem = QPushButton("Select OEM's")
        self.btn_delete = QPushButton("Delete")
        self.btn_reset = QPushButton("Reset")
        self.btn_copy_removed = QPushButton("Copy Removed Child Parts")

        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(self.btn_select_oem)
        btn_row.addWidget(self.btn_delete)
        btn_row.addWidget(self.btn_reset)
        btn_row.addWidget(self.btn_copy_removed)
        btn_row.addStretch(1)
        outer.addLayout(btn_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self.import_excel)
        self.btn_select_oem.clicked.connect(self.select_oems)
        self.btn_delete.clicked.connect(self.delete_selected)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.btn_copy_removed.clicked.connect(self.copy_removed_child_parts)

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import Imp BOM", '', 'Excel Files (*.xlsx)')
        if not path:
            return
        import pandas as pd
        from PyQt6.QtGui import QColor

        df = pd.read_excel(path, engine='openpyxl')
        cols = list(df.columns)
        if 'Part' not in cols or 'BOM Level' not in cols:
            QMessageBox.warning(self, 'Missing Column', 'Required columns not found')
            return

        part_idx = cols.index('Part')
        cols2 = cols[:part_idx+1] + ['Tool comments'] + cols[part_idx+1:]

        headers = ['Select'] + cols2
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(df))

        bom_df_idx = cols.index('BOM Level')
        tc_tbl_idx = headers.index('Tool comments')

        for r in range(len(df)):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            self.table.setCellWidget(r, 0, cont)

            for c, col in enumerate(cols):
                val = '' if pd.isna(df.iloc[r, c]) else str(df.iloc[r, c])
                tbl_c = 1 + c if c <= part_idx else 1 + c + 1
                self.table.setItem(r, tbl_c, QTableWidgetItem(val))

            bom_val = str(df.iloc[r, bom_df_idx]).strip()
            if bom_val != '0':
                it = QTableWidgetItem('Removed child Part')
                f = it.font(); f.setBold(True); it.setFont(f)
                it.setForeground(QColor('orange'))
                self.table.setItem(r, tc_tbl_idx, it)
            else:
                # BOM Level 0 -> BLUE row
                for c in range(self.table.columnCount()):
                    cell = self.table.item(r, c)
                    if cell:
                        cell.setBackground(QColor('#87CEEB'))  # Sky Blue

        self.table.resizeColumnsToContents()

    def find_column(self, header):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text() == header:
                return i
        return -1

    def select_oems(self):
        part_col = self.find_column('Part')
        bom_col = self.find_column('BOM Level')
        if part_col < 0 or bom_col < 0:
            return

        eligible = []
        for r in range(self.table.rowCount()):
            part = (self.table.item(r, part_col).text() if self.table.item(r, part_col) else '').lstrip()
            bom = (self.table.item(r, bom_col).text() if self.table.item(r, bom_col) else '').strip()
            prefix = part[:4]
            if bom != '0' and prefix.isdigit() and int(prefix) >= 500:
                eligible.append(r)

        should_select = any(
            not self.table.cellWidget(r, 0).findChild(QCheckBox).isChecked()
            for r in eligible
        )

        for r in eligible:
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            chk.setChecked(should_select)

    def delete_selected(self):
        rows = []
        for r in range(self.table.rowCount()):
            chk = self.table.cellWidget(r, 0).findChild(QCheckBox)
            if chk.isChecked():
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)

    def copy_removed_child_parts(self):
        part_col = self.find_column('Part')
        tc_col = self.find_column('Tool comments')

        if part_col < 0 or tc_col < 0:
            QMessageBox.warning(self, 'Column Missing',
                                'Required columns (Part / Tool comments) not found.')
            return

        unique_parts = set()

        for r in range(self.table.rowCount()):
            tc_item = self.table.item(r, tc_col)
            if not tc_item or not tc_item.text().strip():
                continue

            part_item = self.table.item(r, part_col)
            if not part_item:
                continue

            part = part_item.text().replace(' ', '').strip()
            if part:
                unique_parts.add(part)

        if not unique_parts:
            QMessageBox.information(self, 'No Data',
                                    'No removed child parts found to copy.')
            return

        from PyQt6.QtGui import QGuiApplication
        result = '\n'.join(sorted(unique_parts))
        QGuiApplication.clipboard().setText(result)

        QMessageBox.information(self, 'Copied',
                                f'Copied {len(unique_parts)} unique removed child part(s) to clipboard.')


class WURemovedBOMItemsTab(QWidget):

    def append_orphans_to_obs_parts(self, remove_from_orphan_table=False):
        if not self.obs_provider:
            return
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphan Child')
        if pcol < 0 or ocol < 0:
            return
        orphan_map = {}
        orphan_rows = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, pcol)
            oit = self.table.item(r, ocol)
            if pit and oit and oit.text().lower().startswith('orphan'):
                orphan_map[pit.text().strip().upper()] = oit.text().strip()
                orphan_rows.append(r)
        if not orphan_map:
            return
        t = self.obs_provider.table
        existing = {t.item(r,1).text().strip().upper() for r in range(t.rowCount()) if t.item(r,1) and t.item(r,1).text().strip()}
        for part, lvl in orphan_map.items():
            if part in existing:
                continue
            target = next((r for r in range(t.rowCount()) if not t.item(r,1) or not t.item(r,1).text().strip()), None)
            if target is None:
                target = t.rowCount(); t.setRowCount(target+1); t._init_rows(target,target+1)
            t.setItem(target,1,QTableWidgetItem(part))
            cb = QComboBox(); cb.addItem(lvl); cb.setCurrentText(lvl)
            col = get_orphan_color(lvl)
            if col: cb.setStyleSheet(f"QComboBox {{ color:{col.name()}; font-weight:bold; }}")
            t.setCellWidget(target,2,cb)
            t.setItem(target,3,QTableWidgetItem('No Replacement'))
        if remove_from_orphan_table:
            for r in reversed(orphan_rows): self.table.removeRow(r)

    def __init__(self, obs_provider):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)

        btn_row = QHBoxLayout()
        self.btn_import = QPushButton("Import WU of Removed BOM items")
        self.btn_analyze = QPushButton("Perform Orphan Analysis")
        self.btn_reset = QPushButton("Reset")
        self.btn_remove_esw = QPushButton("Remove ESW Parents")
        self.btn_remove_9024 = QPushButton("Remove 9024")

        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(self.btn_analyze)
        btn_row.addWidget(self.btn_reset)
        btn_row.addWidget(self.btn_remove_esw)
        btn_row.addWidget(self.btn_remove_9024)
        btn_row.addStretch(1)
        outer.addLayout(btn_row)

        self.table = QTableWidget(0, 0)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        outer.addWidget(self.table)

        self.btn_import.clicked.connect(self.import_excel)
        self.btn_analyze.clicked.connect(self.perform_orphan_analysis)
        self.btn_reset.clicked.connect(self.reset_tab)
        self.btn_remove_esw.clicked.connect(lambda: self.remove_by_prefix('ESW'))
        self.btn_remove_9024.clicked.connect(lambda: self.remove_by_prefix('9024'))

    def reset_tab(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def _find_col(self, name):
        for i in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(i)
            if h and h.text().strip().lower() == name.lower():
                return i
        return -1

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import WU of Removed BOM items", '', "Excel Files (*.xlsx)")
        if not path:
            return

        import pandas as pd
        from PyQt6.QtGui import QColor

        df = pd.read_excel(path, engine='openpyxl')
        if 'Part' not in df.columns:
            QMessageBox.warning(self, 'Missing Column', 'Part column not found')
            return

        cleaned = []
        for _, row in df.iterrows():
            original_part = '' if pd.isna(row['Part']) else str(row['Part'])
            p = original_part.lstrip()
            lp = len(p)

            keep = False
            if lp > 10 and p.upper().startswith('ESW'):
                keep = True
            elif lp == 10 and p[4] == '-':
                keep = True

            if keep:
                cleaned.append(row)

        if not cleaned:
            QMessageBox.information(self, 'No Valid Data', 'All rows removed by cleanup rules.')
            return

        clean_df = pd.DataFrame(cleaned)
        clean_df.insert(2, 'Orphan Child', '')

        headers = ['Select'] + list(clean_df.columns)
        self.table.setRowCount(len(clean_df))
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        for r in range(len(clean_df)):
            chk = QCheckBox()
            cont = QWidget()
            h = QHBoxLayout(cont)
            h.setContentsMargins(0, 0, 0, 0)
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.addWidget(chk)
            self.table.setCellWidget(r, 0, cont)

            wu_zero = False
            if 'WU Level' in clean_df.columns:
                wu_zero = str(clean_df.iloc[r]['WU Level']).strip() == '0'

            for c, col in enumerate(clean_df.columns):
                val = '' if pd.isna(clean_df.iloc[r, c]) else str(clean_df.iloc[r, c])
                item = QTableWidgetItem(val)
                if wu_zero:
                    item.setBackground(QColor('#87CEEB'))
                self.table.setItem(r, c + 1, item)

        self.table.resizeColumnsToContents()

    def _build_obs_change_map(self):
        mapping = {}
        t = self.obs_provider.table
        for r in range(t.rowCount()):
            obs_it = t.item(r, 1)
            chg_w = t.cellWidget(r, 2)
            key = obs_it.text().strip().upper() if obs_it else ''
            if key:
                mapping[key] = chg_w.currentText() if chg_w else ''
        return mapping

    def perform_orphan_analysis(self):
        # Find required columns
        pcol = self._find_col('Part')
        ocol = self._find_col('Orphan Child')
        wucol = self._find_col('WU Level')

        if pcol < 0 or ocol < 0 or wucol < 0:
            QMessageBox.information(self, 'Orphan Analysis', 'Required columns not found.')
            return

        row_count = self.table.rowCount()

        def is_zero(v):
            return str(v).strip() in ('0', '0.0')

        def is_one(v):
            return str(v).strip() in ('1', '1.0')

        ignore_prefixes = ('0243', '0299', '0289', '0290')

        # Step 0: OBS → Obsolete / Inactivate mapping (existing behavior)
        obs_map = self._build_obs_change_map()
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if not it:
                continue
            key = it.text().lstrip().upper()
            if key in obs_map:
                self.table.setItem(r, ocol, QTableWidgetItem(obs_map[key]))

        # Step 1: Orphan1
        orphan_parts = set()
        r = 0
        while r < row_count:
            wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
            if is_zero(wu):
                child_row = r
                part_it = self.table.item(child_row, pcol)
                orphan_it = self.table.item(child_row, ocol)
                part = part_it.text().strip() if part_it else ''

                if orphan_it and orphan_it.text().strip():
                    r += 1
                    continue

                block_end = row_count
                for i in range(r + 1, row_count):
                    nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if is_zero(nxt):
                        block_end = i
                        break

                parents = []
                for i in range(child_row + 1, block_end):
                    wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                    if not is_one(wu_p):
                        continue
                    pit = self.table.item(i, pcol)
                    pval = pit.text().lstrip() if pit else ''
                    if pval.startswith(ignore_prefixes):
                        continue
                    parents.append(i)

                mark = False
                if not parents:
                    mark = True
                else:
                    all_bad = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        status = oit.text().strip().lower() if oit else ''
                        if status not in ('obsolete', 'inactivate'):
                            all_bad = False
                            break
                    if all_bad:
                        mark = True

                if mark and part:
                    self.table.setItem(child_row, ocol, QTableWidgetItem('Orphan1'))
                    orphan_parts.add(part.upper())

                r = block_end
            else:
                r += 1

        # Propagate Orphan1
        for r in range(row_count):
            it = self.table.item(r, pcol)
            if it and it.text().strip().upper() in orphan_parts:
                oit = self.table.item(r, ocol)
                if not oit or not oit.text().strip():
                    self.table.setItem(r, ocol, QTableWidgetItem('Orphan1'))

        # Step 2+: Orphan2, Orphan3...
        current_level = 2
        while True:
            new_found = False
            new_parts = set()
            r = 0
            while r < row_count:
                wu = self.table.item(r, wucol).text() if self.table.item(r, wucol) else ''
                orphan_it = self.table.item(r, ocol)
                if is_zero(wu) and (not orphan_it or not orphan_it.text().strip()):
                    part_it = self.table.item(r, pcol)
                    part = part_it.text().strip() if part_it else ''

                    block_end = row_count
                    for i in range(r + 1, row_count):
                        nxt = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if is_zero(nxt):
                            block_end = i
                            break

                    parents = []
                    for i in range(r + 1, block_end):
                        wu_p = self.table.item(i, wucol).text() if self.table.item(i, wucol) else ''
                        if not is_one(wu_p):
                            continue
                        pit = self.table.item(i, pcol)
                        pval = pit.text().lstrip() if pit else ''
                        if pval.startswith(ignore_prefixes):
                            continue
                        parents.append(i)

                    parents_filled = True
                    for pr in parents:
                        oit = self.table.item(pr, ocol)
                        if not oit or not oit.text().strip():
                            parents_filled = False
                            break

                    if part and (not parents or parents_filled):
                        self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))
                        new_parts.add(part.upper())
                        new_found = True
                r += 1

            for r in range(row_count):
                it = self.table.item(r, pcol)
                oit = self.table.item(r, ocol)
                if it and it.text().strip().upper() in new_parts and (not oit or not oit.text().strip()):
                    self.table.setItem(r, ocol, QTableWidgetItem(f'Orphan{current_level}'))

            if not new_found:
                break
            current_level += 1

            self.append_orphans_to_obs_parts()
        from PyQt6.QtGui import QColor
        for r in range(self.table.rowCount()):
            wu = self.table.item(r, wucol)
            if wu and wu.text().strip() == '0':
                for c in range(self.table.columnCount()):
                    it = self.table.item(r, c)
                    if it:
                        it.setBackground(QColor('#87CEEB'))


    def remove_by_prefix(self, prefix):
        pcol = self._find_col('Part')
        if pcol < 0:
            return
        rows = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, pcol)
            if it and it.text().lstrip().startswith(prefix):
                rows.append(r)
        for r in reversed(rows):
            self.table.removeRow(r)


class  OBSAllPartsWithoutReplacementTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        self.obs_provider = obs_provider
        outer = QVBoxLayout(self)

        subtabs = QTabWidget()
        # Equal width secondary tabs
        subtabs.setStyleSheet("""
        QTabBar::tab {
            min-width: 220px;
            padding: 6px 12px;
            text-align: center;
        }
        """)

        subtabs.addTab(OrphanOBSSubTab(), "Imp BOM")
        subtabs.addTab(WURemovedBOMItemsTab(self.obs_provider), "WU of Removed BOM Items")

        outer.addWidget(subtabs)

class _OrphanAnalysisTab_OLD(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        outer = QVBoxLayout(self)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))
        outer.addWidget(title)

        self.subtabs = QTabWidget()
        self.subtabs.setStyleSheet("""
        QTabBar::tab:selected {
            background-color: #87CEEB;
            color: #0F2D46;
            font-weight: 600;
        }
        QTabBar::tab {
            background-color: #EAF6FD;
        }
        """)
        self.subtabs.addTab(
            OBSAllPartsWithoutReplacementTab(obs_provider),
            "OBS all Parts without Replacements"
        )
        self.subtabs.addTab(
            PlaceholderTab("OBS with or without Replacement"),
            "OBS with or without Replacement"
        )

        outer.addWidget(self.subtabs)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1280, 860)
        self.tabs = QTabWidget(); self.tabs.setDocumentMode(True); self.tabs.setMovable(True)
        self.tabs.setStyleSheet("""
            QTabBar::tab { background: #EAF2FB; color: #1F3B57; padding: 8px 14px; border: 1px solid #D5E3F6; border-bottom: none; border-top-left-radius:6px; border-top-right-radius:6px; }
            QTabBar::tab:selected { background: #FFFFFF; color: #0F2D46; font-weight: 600; }
            QTabWidget::pane { border: 1px solid #D5E3F6; top: -1px; }
        """)

        readme_path = Path(__file__).with_name('README.txt')
        self.readme_tab = ReadmeTab(readme_path)
        self.front_tab = ECRFrontPageTab()
        self.obs_tab = OBSPartsTab()
        self.structure_tab = None  # to be created later
        self.whereused_tab = WhereUsedTabV2(obs_provider=self.obs_tab)

        self.tabs.addTab(self.readme_tab, "READ ME")
        self.tabs.addTab(self.front_tab, "ECR Front Page")
        self.tabs.addTab(self.obs_tab, "OBS Parts")
        self.tabs.addTab(self.whereused_tab, "Where Used")
        self.tabs.addTab(OrphanAnalysisTab(self.obs_tab), "Orphan Analysis")
        self.structure_tab = StructureSheetTab(); self.tabs.addTab(self.structure_tab, "Structure sheet")
        self.tabs.addTab(PlaceholderTab("Inventory Cost Analysis"), "Inventory_Cost")
        self.tabs.addTab(PlaceholderTab("Report"), "Report")
        self.tabs.addTab(PlaceholderTab("User Notes"), "User Notes")
        self.setCentralWidget(self.tabs)
        def update_tab_colors():
            for i in range(self.tabs.count()):
                txt = self.tabs.tabText(i)
                # heuristic: Front Page checklist
                if 'ECR Front Page' in txt:
                    done = self.front_tab.progress_bar.value()==self.front_tab.progress_bar.maximum()
                    if done:
                        self.tabs.setTabText(i, '✓ ' + txt.strip('✓ '))
                        self.tabs.tabBar().setTabBackgroundColor(i, QColor('#E8F5E9'))
                    elif self.front_tab.progress_bar.value()>0:
                        self.tabs.setTabText(i, '✓ ' + txt.strip('✓ '))
  #                      self.tabs.tabBar().setTabBackgroundColor(i, QColor('#E3F2FD'))
            
        try:
            self.front_tab.progress_bar.valueChanged.connect(update_tab_colors)
        except Exception: pass


        tb=QToolBar("File"); self.addToolBar(tb)
        find_edit = QLineEdit(); find_edit.setPlaceholderText('Find part / text'); find_edit.setFixedWidth(220); tb.addWidget(find_edit)
        def _do_find():
            text = find_edit.text().strip().lower()
            if not text:
                return
            w = self.tabs.currentWidget()
            from PyQt6.QtWidgets import QTableWidget, QTextEdit
            found = False
            for t in w.findChildren(QTableWidget):
                for r in range(t.rowCount()):
                    for c in range(t.columnCount()):
                        it = t.item(r,c)
                        if it and text in it.text().lower():
                            it.setBackground(QColor('#FFF59D'))
                            t.setCurrentCell(r,c)
                            found = True
                            break
                    if found:
                        break
                if found:
                    break
            if not found:
                for te in w.findChildren(QTextEdit):
                    if text in te.toPlainText().lower():
                        te.find(text)
                        found = True
                        break
            if not found:
                QMessageBox.information(self, 'Find', 'Not Found')
        find_edit.returnPressed.connect(_do_find)
        act_save=QAction("Save", self); act_save.triggered.connect(self.save_data); tb.addAction(act_save)
        
        spacer=QWidget(); spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred); tb.addWidget(spacer)

        reset_btn=QPushButton("Reset App")
        reset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reset_btn.setStyleSheet("""
            QPushButton { color:#FFFFFF; padding:6px 14px; border-radius:6px; border:1px solid #0D5EA6;
            background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2994FF, stop:1 #0A67C2); }
            QPushButton:hover { background-color:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #2FA0FF, stop:1 #0D6ED0); }
            QPushButton:pressed { padding-top:7px; padding-bottom:5px; }
        """)
        shadow=QGraphicsDropShadowEffect(self); shadow.setBlurRadius(12); shadow.setXOffset(0); shadow.setYOffset(2); shadow.setColor(QColor(0,0,0,80))
        reset_btn.setGraphicsEffect(shadow); reset_btn.clicked.connect(self.reset_app); tb.addWidget(reset_btn)

        self.load_data_if_exists()

    def aggregate_data(self)->Dict[str,Any]:
        return {'front_page': self.front_tab.to_dict(), 'obs_parts': self.obs_tab.to_dict()}

    def apply_data(self, data: Dict[str,Any]):
        if not data: return
        if 'front_page' in data: self.front_tab.from_dict(data['front_page'])
        if 'obs_parts' in data: self.obs_tab.from_dict(data['obs_parts'])

    def save_data(self):
        data=self.aggregate_data()
        try:
            DATA_FILE.write_text(json.dumps(data, indent=2), encoding='utf-8')
            QMessageBox.information(self, "Saved", f"Data saved to {DATA_FILE.name}")
        except Exception as e:
            QMessageBox.warning(self, "Save Failed", str(e))

    def load_data_if_exists(self):
        try:
            if DATA_FILE.exists():
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
        except Exception as e:
            QMessageBox.warning(self, "Load Failed", str(e))

    def load_data_dialog(self):
        if DATA_FILE.exists():
            try:
                data=json.loads(DATA_FILE.read_text(encoding='utf-8'))
                self.apply_data(data)
                QMessageBox.information(self, "Loaded", f"Data loaded from {DATA_FILE.name}")
            except Exception as e:
                QMessageBox.warning(self, "Load Failed", str(e))
        else:
            QMessageBox.information(self, "No Save Found", "No saved data file found yet. Click Save to create one.")

    def reset_app(self):
        ans=QMessageBox.question(self, "Reset App", "This will clear all fields, reset tables and remove any saved data. Do you want to continue?",
                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if ans==QMessageBox.StandardButton.Yes:
            try:
                self.front_tab.reset(); self.obs_tab.reset()
                if DATA_FILE.exists(): DATA_FILE.unlink()
                QMessageBox.information(self, "Reset Complete", "Application data has been reset.")
            except Exception as e:
                QMessageBox.warning(self, "Reset Failed", str(e))

class OrphanAnalysisTab(QWidget):
    def __init__(self, obs_provider):
        super().__init__()
        outer = QVBoxLayout(self)

        header = QWidget(self)
        h = QHBoxLayout(header)
        h.setContentsMargins(0, 0, 0, 0)

        title = QLabel("Orphan Analysis")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.DemiBold))

        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        rb_without = QRadioButton("Without Replacement")
        rb_with = QRadioButton("With / Without Replacement")
        rb_without.setChecked(True)

        grp = QButtonGroup(self)
        grp.addButton(rb_without, 0)
        grp.addButton(rb_with, 1)

        h.addWidget(title)
        h.addSpacing(12)
        h.addWidget(rb_without)
        h.addWidget(rb_with)
        h.addStretch(1)
        outer.addWidget(header)

        self.tabs = QTabWidget()
        tab1 = OBSAllPartsWithoutReplacementTab(obs_provider)
        tab2 = PlaceholderTab("OBS with or without Replacement")

        self.tabs.addTab(tab1, "OBS all Parts without Replacements")
        self.tabs.addTab(tab2, "OBS with or without Replacement")
        outer.addWidget(self.tabs)

        def apply(idx):
            self.tabs.setCurrentIndex(idx)
            self.tabs.setTabEnabled(0, idx == 0)
            self.tabs.setTabEnabled(1, idx == 1)

        rb_without.toggled.connect(lambda v: v and apply(0))
        rb_with.toggled.connect(lambda v: v and apply(1))
        apply(0)



def run():
    app=QApplication(sys.argv); app.setStyle("Fusion")
    pal=QPalette(); pal.setColor(QPalette.ColorRole.Window, QColor(247,250,253)); pal.setColor(QPalette.ColorRole.Base, QColor(255,255,255))
    pal.setColor(QPalette.ColorRole.AlternateBase, QColor(241,246,252)); pal.setColor(QPalette.ColorRole.Text, QColor(28,41,56)); app.setPalette(pal)
    win=MainWindow(); win.show(); sys.exit(app.exec())

if __name__=='__main__': run()
